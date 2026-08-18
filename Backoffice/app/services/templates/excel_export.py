"""Template Excel export paths."""

# ========== Template Excel Import/Export Service ==========
from app.utils.datetime_helpers import utcnow
"""
Service for exporting and importing form templates to/from Excel.
"""
from flask import current_app
from flask_login import current_user
from app import db
from app.models import (
    FormTemplate, FormPage, FormSection, FormItem, FormTemplateVersion,
    FormData,
    RepeatGroupInstance, RepeatGroupData, DynamicIndicatorData, DynamicSectionContext,
)
from app.models.documents import SubmittedDocument
from app.models.indicator_bank import IndicatorBank
from contextlib import suppress
from app.services.monitoring.memory import memory_tracker
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import io
import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from app.utils.stable_key import generate_stable_key, is_valid_stable_key, normalize_stable_key

from app.services.templates.excel_base import TemplateExcelBase


class TemplateExcelExportMixin(TemplateExcelBase):
    """Mixin for TemplateExcelService."""

    @classmethod
    @memory_tracker("Template Excel Export", log_top_allocations=True)
    def export_template(cls, template_id: int, version_id: Optional[int] = None) -> io.BytesIO:
        """
        Export template structure to Excel.

        Args:
            template_id: Template ID to export
            version_id: Optional version ID (defaults to published or latest)

        Returns:
            BytesIO object containing Excel file
        """
        template = FormTemplate.query.get_or_404(template_id)

        # Determine version to export
        if version_id:
            version = FormTemplateVersion.query.filter_by(
                id=version_id, template_id=template.id
            ).first()
        else:
            if template.published_version_id:
                version = FormTemplateVersion.query.get(template.published_version_id)
            else:
                version = FormTemplateVersion.query.filter_by(
                    template_id=template.id
                ).order_by(FormTemplateVersion.created_at.desc()).first()

        if not version:
            raise ValueError(f"No version found for template {template_id}")

        # Create workbook
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        # Export Instructions sheet first (so it appears first)
        cls._export_instructions_sheet(workbook)

        # Export hidden metadata sheet
        cls._export_metadata_sheet(workbook, template, version)

        # Export Template sheet
        cls._export_template_sheet(workbook, template, version)

        # Export Pages sheet
        cls._export_pages_sheet(workbook, template, version)

        # Export Sections sheet
        cls._export_sections_sheet(workbook, template, version)

        # Export Items sheet
        cls._export_items_sheet(workbook, template, version)

        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    @classmethod
    def _export_template_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export template metadata to Template sheet (field/value rows)."""
        sheet = workbook.create_sheet("Template")

        headers = cls.TEMPLATE_SHEET_ROW_HEADERS
        required_cols = cls.REQUIRED_COLUMNS.get('Template', [])
        field_names = cls.get_template_columns()
        field_values = cls._build_template_field_values(template, version)

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cls._style_header_cell(cell, is_required=(header == 'field'))

        boolean_fields = {
            'add_to_self_report', 'display_order_visible', 'is_paginated',
            'enable_export_pdf', 'enable_export_excel', 'enable_import_excel', 'enable_ai_validation',
        }

        for row_idx, field_name in enumerate(field_names, start=2):
            field_cell = sheet.cell(row=row_idx, column=1, value=field_name)
            cls._style_header_cell(field_cell, is_required=(field_name in required_cols))
            sheet.cell(row=row_idx, column=2, value=field_values.get(field_name))

            if field_name in boolean_fields and field_name in cls.DROPDOWN_OPTIONS:
                cls._add_dropdown_validation(
                    sheet, 2, cls.DROPDOWN_OPTIONS[field_name],
                    start_row=row_idx, end_row=row_idx,
                )

        cls._auto_size_columns(sheet, 2)
        cls._create_excel_table(sheet, "TemplateTable", 2, len(field_names) + 1)

    @classmethod
    def _export_pages_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export pages to Pages sheet with sequential IDs."""
        sheet = workbook.create_sheet("Pages")

        # Get pages for this version (include archived)
        pages = FormPage.query.filter_by(
            template_id=template.id, version_id=version.id
        ).order_by(FormPage.order).all()

        # Write headers with required/optional styling
        headers = cls.get_page_columns()
        required_cols = cls.REQUIRED_COLUMNS.get('Pages', [])
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            is_required = header in required_cols
            cls._style_header_cell(cell, is_required=is_required)

        # Create mapping: sequential export ID -> database ID
        page_export_id_map = {}  # export_id -> db_id

        # Write data rows with sequential IDs
        for row_idx, page in enumerate(pages, start=2):
            export_id = row_idx - 1  # Sequential ID starting from 1
            page_export_id_map[export_id] = page.id  # Store mapping for reference

            row_data = cls._build_export_row(
                cls.PAGE_BASE_COLUMNS,
                cls.PAGE_TRANSLATABLE_FIELDS,
                {
                    'id': export_id,
                    'name': page.name,
                    'order': page.order,
                },
                {'name': page.name_translations},
            )
            cls._write_data_row(sheet, row_idx, headers, row_data)

        # Store mapping in sheet for reference (not visible, but can be used if needed)
        sheet._page_export_id_map = page_export_id_map

        # Auto-size columns
        cls._auto_size_columns(sheet, len(headers))

        # Create Excel table (always create, even with 0 rows)
        num_data_rows = len(pages)
        cls._create_excel_table(sheet, "PagesTable", len(headers), num_data_rows + 1)  # +1 for header

        # Add conditional formatting for duplicate IDs (only if there are data rows)
        if num_data_rows > 0:
            cls._add_duplicate_highlighting(sheet, 'id', headers, num_data_rows + 1)

    @classmethod
    def _export_sections_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export sections to Sections sheet with sequential IDs."""
        sheet = workbook.create_sheet("Sections")

        # Get sections for this version (include archived)
        sections = FormSection.query.filter_by(
            template_id=template.id, version_id=version.id
        ).order_by(FormSection.order).all()

        # Get page export ID mapping from Pages sheet
        pages_sheet = workbook['Pages']
        page_db_to_export = {}  # db_id -> export_id
        if hasattr(pages_sheet, '_page_export_id_map'):
            # Reverse the map: export_id -> db_id becomes db_id -> export_id
            for exp_id, db_id in pages_sheet._page_export_id_map.items():
                page_db_to_export[db_id] = exp_id

        # Build section database ID to sequential export ID mapping
        section_db_to_export = {}  # db_id -> export_id

        # Build item DB -> export mapping for rewriting relevance/validation rules.
        # This must match the Items sheet export IDs (which are sequential as well).
        item_db_to_export = cls._build_item_db_to_export_map(template, version)

        # First pass: build mapping
        for row_idx, section in enumerate(sections, start=2):
            export_id = row_idx - 1  # Sequential ID starting from 1
            section_db_to_export[section.id] = export_id

        # Write headers with required/optional styling
        headers = cls.get_section_columns()
        required_cols = cls.REQUIRED_COLUMNS.get('Sections', [])
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            is_required = header in required_cols
            cls._style_header_cell(cell, is_required=is_required)

        # Second pass: write data rows with sequential IDs and mapped references
        for row_idx, section in enumerate(sections, start=2):
            export_id = section_db_to_export[section.id]

            # Map page_id to export ID
            page_export_id = None
            if section.page_id and section.page_id in page_db_to_export:
                page_export_id = page_db_to_export[section.page_id]

            # Map parent_section_id to export ID
            parent_export_id = None
            if section.parent_section_id and section.parent_section_id in section_db_to_export:
                parent_export_id = section_db_to_export[section.parent_section_id]

            row_data = cls._build_export_row(
                cls.SECTION_BASE_COLUMNS,
                cls.SECTION_TRANSLATABLE_FIELDS,
                {
                    'id': export_id,
                    'name': section.name,
                    'order': section.order,
                    'parent_section_id': parent_export_id,
                    'page_id': page_export_id,
                    'section_type': section.section_type,
                    'max_dynamic_indicators': section.max_dynamic_indicators,
                    'allowed_sectors': cls._format_json_for_excel(section.allowed_sectors),
                    'indicator_filters': cls._format_json_for_excel(section.indicator_filters),
                    'allow_data_not_available': section.allow_data_not_available,
                    'allow_not_applicable': section.allow_not_applicable,
                    'allowed_disaggregation_options': cls._format_json_for_excel(section.allowed_disaggregation_options),
                    'data_entry_display_filters': cls._format_json_for_excel(section.data_entry_display_filters),
                    'add_indicator_note': section.add_indicator_note,
                    'relevance_condition': cls._rewrite_rule_json_item_ids(section.relevance_condition, item_db_to_export),
                    'archived': section.archived,
                },
                {'name': section.name_translations},
            )
            cls._write_data_row(sheet, row_idx, headers, row_data)

        # Store mapping for items sheet (export_id -> db_id)
        sheet._section_export_id_map = {exp_id: db_id for db_id, exp_id in section_db_to_export.items()}

        # Auto-size columns
        cls._auto_size_columns(sheet, len(headers))

        # Add data validation dropdowns
        section_type_col = headers.index('section_type') + 1  # +1 because Excel is 1-indexed
        cls._add_dropdown_validation(sheet, section_type_col, cls.DROPDOWN_OPTIONS['section_type'],
                                    start_row=2, end_row=len(sections) + 1)

        # Add boolean dropdowns
        archived_col = headers.index('archived') + 1
        cls._add_dropdown_validation(sheet, archived_col, cls.DROPDOWN_OPTIONS['archived'],
                                    start_row=2, end_row=len(sections) + 1)

        allow_data_not_available_col = headers.index('allow_data_not_available') + 1
        cls._add_dropdown_validation(sheet, allow_data_not_available_col, cls.DROPDOWN_OPTIONS['allow_data_not_available'],
                                    start_row=2, end_row=len(sections) + 1)

        allow_not_applicable_col = headers.index('allow_not_applicable') + 1
        cls._add_dropdown_validation(sheet, allow_not_applicable_col, cls.DROPDOWN_OPTIONS['allow_not_applicable'],
                                    start_row=2, end_row=len(sections) + 1)

        # Add dropdown for page_id that references Pages sheet (if Pages sheet exists)
        if 'Pages' in workbook.sheetnames:
            page_id_col = headers.index('page_id') + 1
            pages_sheet = workbook['Pages']
            cls._add_sheet_reference_dropdown(sheet, page_id_col, pages_sheet, 'id',
                                             start_row=2, end_row=len(sections) + 1)

        # Create Excel table (always create, even with 0 rows)
        num_data_rows = len(sections)
        cls._create_excel_table(sheet, "SectionsTable", len(headers), num_data_rows + 1)  # +1 for header

        # Add conditional formatting for duplicate IDs (only if there are data rows)
        if num_data_rows > 0:
            cls._add_duplicate_highlighting(sheet, 'id', headers, num_data_rows + 1)

    @classmethod
    def _export_items_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export items to Items sheet with sequential IDs."""
        sheet = workbook.create_sheet("Items")

        # Get items for this version (include archived)
        items = cls._get_items_for_version(template, version)

        # Build DB -> export mapping for rewriting rule JSON inside exported strings
        item_db_to_export = {item.id: idx + 1 for idx, item in enumerate(items)}

        # Get section export ID mapping from Sections sheet
        sections_sheet = workbook['Sections']
        section_db_to_export = {}  # db_id -> export_id
        if hasattr(sections_sheet, '_section_export_id_map'):
            # Reverse the map: export_id -> db_id becomes db_id -> export_id
            for exp_id, db_id in sections_sheet._section_export_id_map.items():
                section_db_to_export[db_id] = exp_id

        # Write headers with required/optional styling
        headers = cls.get_item_columns()
        required_cols = cls.REQUIRED_COLUMNS.get('Items', [])
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            is_required = header in required_cols
            cls._style_header_cell(cell, is_required=is_required)

        # Write data rows with sequential IDs
        for row_idx, item in enumerate(items, start=2):
            export_id = row_idx - 1  # Sequential ID starting from 1

            # Map section_id to export ID
            section_export_id = None
            if item.section_id and item.section_id in section_db_to_export:
                section_export_id = section_db_to_export[item.section_id]

            row_data = cls._build_export_row(
                cls.ITEM_BASE_COLUMNS,
                cls.ITEM_TRANSLATABLE_FIELDS,
                {
                    'id': export_id,
                    'section_id': section_export_id,
                    'item_type': item.item_type,
                    'label': item.label,
                    'order': item.order,
                    'relevance_condition': cls._rewrite_rule_json_item_ids(item.relevance_condition, item_db_to_export),
                    'archived': item.archived,
                    'config': cls._format_json_for_excel(
                        cls._normalize_matrix_item_config(item.item_type, item.config)
                    ),
                    'indicator_bank_id': item.indicator_bank_id,
                    'type': item.type,
                    'unit': item.unit,
                    'validation_condition': cls._rewrite_rule_json_item_ids(item.validation_condition, item_db_to_export),
                    'validation_message': item.validation_message,
                    'definition': item.definition,
                    'options_json': cls._format_json_for_excel(item.options_json),
                    'lookup_list_id': item.lookup_list_id,
                    'list_display_column': item.list_display_column,
                    'list_filters_json': cls._format_json_for_excel(item.list_filters_json),
                    'options_translations': cls._format_json_for_excel(item.options_translations),
                    'description': item.description,
                },
                {
                    'label': item.label_translations,
                    'definition': item.definition_translations,
                    'description': item.description_translations,
                    'validation_message': item.validation_message_translations,
                },
            )
            cls._write_data_row(sheet, row_idx, headers, row_data)

        # Auto-size columns
        cls._auto_size_columns(sheet, len(headers))

        # Add data validation dropdowns
        item_type_col = headers.index('item_type') + 1
        item_type_options = cls._get_item_type_options()
        cls._add_dropdown_validation(sheet, item_type_col, item_type_options,
                                    start_row=2, end_row=len(items) + 1)

        archived_col = headers.index('archived') + 1
        cls._add_dropdown_validation(sheet, archived_col, cls.DROPDOWN_OPTIONS['archived'],
                                    start_row=2, end_row=len(items) + 1)

        # Add dropdown for type column with dynamic values from database
        if 'type' in headers:
            type_col = headers.index('type') + 1
            type_options = cls._get_type_options_from_database()
            if type_options:
                cls._add_dropdown_validation(sheet, type_col, type_options,
                                            start_row=2, end_row=len(items) + 1)

        # Add dropdown for section_id that references Sections sheet
        if 'Sections' in workbook.sheetnames:
            section_id_col = headers.index('section_id') + 1
            sections_sheet = workbook['Sections']
            cls._add_sheet_reference_dropdown(sheet, section_id_col, sections_sheet, 'id',
                                             start_row=2, end_row=len(items) + 1)

        # Create Excel table (always create, even with 0 rows)
        num_data_rows = len(items)
        cls._create_excel_table(sheet, "ItemsTable", len(headers), num_data_rows + 1)  # +1 for header

        # Add conditional formatting for duplicate IDs (only if there are data rows)
        if num_data_rows > 0:
            cls._add_duplicate_highlighting(sheet, 'id', headers, num_data_rows + 1)

    @classmethod
    def _export_metadata_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export very hidden metadata sheet with system information."""
        sheet = workbook.create_sheet("_Metadata")
        sheet.sheet_state = 'veryHidden'  # Very hidden - cannot be unhidden via Excel UI

        # Metadata information
        metadata = {
            'Excel Export Version': cls.EXCEL_EXPORT_VERSION,
            'Export Timestamp': utcnow().isoformat(),
            'Template ID': template.id,
            'Template Name': template.name,
            'Version ID': version.id,
            'Version Number': version.version_number,
            'Version Status': version.status,
            'Exported By': current_user.email if current_user else 'System',
        }

        # Write metadata as key-value pairs
        row = 1
        for key, value in metadata.items():
            sheet.cell(row=row, column=1, value=key)
            sheet.cell(row=row, column=2, value=value)
            # Style the key column
            key_cell = sheet.cell(row=row, column=1)
            key_cell.font = Font(bold=True)
            row += 1

        # Auto-size columns
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 50

    @classmethod
    def _export_instructions_sheet(cls, workbook):
        """Export instructions sheet with formatting guide."""
        sheet = workbook.create_sheet("Instructions", 0)  # Insert at position 0 (first sheet)

        # Title
        title_cell = sheet.cell(row=1, column=1, value="Template Excel Import/Export Instructions")
        title_cell.font = Font(bold=True, size=16, color=cls.IFRC_COLORS['WHITE'])
        title_cell.fill = PatternFill(start_color=cls.IFRC_COLORS['DARK_RED'],
                                     end_color=cls.IFRC_COLORS['DARK_RED'],
                                     fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells('A1:D1')
        sheet.row_dimensions[1].height = 35

        # Section: Overview
        row = 3
        section_rows = []  # Track section header rows for border styling
        overview_cell = sheet.cell(row=row, column=1, value="📋 Overview")
        overview_cell.font = Font(bold=True, size=14, color=cls.IFRC_COLORS['DARK_RED'])
        overview_cell.fill = PatternFill(start_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                        end_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                        fill_type='solid')
        sheet.merge_cells(f'A{row}:D{row}')
        section_rows.append(row)
        row += 1

        instructions = [
            "This Excel file contains the structure of a form template that can be imported back into the system.",
            "",
            "The file contains the following sheets:",
            "  • Instructions (this sheet) - Ignored during import",
            "  • Template - Template metadata and configuration (field/value rows)",
            "  • Pages - Page definitions (if template is paginated)",
            "  • Sections - Section definitions",
            "  • Items - Form items (indicators, questions, document fields, etc.)",
            "",
            "Any sheets other than the recognized ones (Template, Pages, Sections, Items) will be ignored during import.",
        ]

        for instruction in instructions:
            cell = sheet.cell(row=row, column=1, value=instruction)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if instruction.startswith("  •"):
                cell.font = Font(size=10)
            row += 1

        # Section: Column Headers
        row += 2
        headers_cell = sheet.cell(row=row, column=1, value="🎨 Column Headers")
        headers_cell.font = Font(bold=True, size=14, color=cls.IFRC_COLORS['DARK_RED'])
        headers_cell.fill = PatternFill(start_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                       end_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                       fill_type='solid')
        sheet.merge_cells(f'A{row}:D{row}')
        section_rows.append(row)
        row += 1

        header_instructions = [
            "Column headers in Pages, Sections, and Items are color-coded to indicate whether they are required or optional:",
            "",
            "The Template sheet uses a field/value layout: column A lists field names and column B holds values (one field per row).",
            "Required Template fields (e.g. name) are shown in red in column A.",
            "",
            "  🔴 RED HEADERS (Required): These columns must have values. They are essential for the template to function.",
            "     Missing values in required columns will cause the import to fail.",
            "",
            "  🔵 BLUE HEADERS (Optional): These columns can be left empty. They provide additional configuration or metadata.",
            "     Empty optional columns will use default values or be set to NULL.",
            "",
            "Required columns for each sheet:",
        ]

        for instruction in header_instructions:
            cell = sheet.cell(row=row, column=1, value=instruction)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            row += 1

        # List required columns for each sheet
        for sheet_name in ['Template', 'Pages', 'Sections', 'Items']:
            required = cls.REQUIRED_COLUMNS.get(sheet_name, [])
            if required:
                cell = sheet.cell(row=row, column=1, value=f"  {sheet_name}: {', '.join(required)}")
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                row += 1

        # Section: Important Notes
        row += 2
        notes_cell = sheet.cell(row=row, column=1, value="⚠️ Important Notes")
        notes_cell.font = Font(bold=True, size=14, color=cls.IFRC_COLORS['DARK_RED'])
        notes_cell.fill = PatternFill(start_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                     end_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                     fill_type='solid')
        sheet.merge_cells(f'A{row}:D{row}')
        section_rows.append(row)
        row += 1

        notes = [
            "1. ID Fields: The 'id' columns in Pages, Sections, and Items are sequential export IDs (1, 2, 3...).",
            "   They are NOT database IDs. The system will create new IDs when importing.",
            "",
            "2. References: When referencing other records (e.g., section_id in Items), use the export IDs from the same file.",
            "",
            "3. JSON Fields: Fields such as config, options_json, and options_translations remain JSON. Exports are pretty-printed with line breaks for readability; re-import accepts both compact and formatted JSON.",
            "",
            "4. Translation Fields: English/base text lives in columns like name, label, definition, and description. Other languages use separate columns named {field}_{lang} (e.g. label_ar, label_fr, name_es). Empty translation cells are ignored on import.",
            "",
            "5. Boolean Fields: Use TRUE/FALSE or 1/0 for boolean values.",
            "",
            "6. Order Fields: Use numeric values (integers or decimals like 1, 1.1, 1.2) to control display order.",
            "",
            "7. Import Behavior:",
            "   • If importing into a published version, a new draft version will be created automatically.",
            "   • Existing items/sections with matching order+name will be updated, others will be created.",
            "   • The import will fail if required columns are missing or invalid.",
            "",
            "8. Do NOT modify the Instructions sheet or add unrecognized sheets if you plan to re-import the file.",
        ]

        for note in notes:
            cell = sheet.cell(row=row, column=1, value=note)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if note and note[0].isdigit():
                cell.font = Font(bold=True, size=10)
            row += 1

        # Auto-size columns
        sheet.column_dimensions['A'].width = 100
        sheet.column_dimensions['B'].width = 5
        sheet.column_dimensions['C'].width = 5
        sheet.column_dimensions['D'].width = 5

        # Add borders to section headers for better visual separation
        for section_row in section_rows:
            for col in ['A', 'B', 'C', 'D']:
                cell = sheet[f'{col}{section_row}']
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='medium'),
                    bottom=Side(style='thin')
                )

        # Freeze first row
        sheet.freeze_panes = 'A2'

        # Protect the instructions sheet (read-only, but allow formatting)
        sheet.protection.sheet = True
        sheet.protection.formatCells = False
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False

    @classmethod
    def _create_excel_table(cls, sheet, table_name: str, num_columns: int, num_rows: int):
        """Create an Excel table from the data range.

        Args:
            sheet: The worksheet to add the table to
            table_name: Name for the table
            num_columns: Number of columns
            num_rows: Number of rows (including header)
            Note: Excel requires at least 2 rows (header + 1 data row), so if num_rows=1, we create with 2 rows
        """
        if num_rows < 1:  # Need at least header row
            return

        # Excel requires at least 2 rows for a table (header + at least one data row)
        # If we only have a header, create table with header + 1 empty data row
        if num_rows == 1:
            num_rows = 2  # Add one empty data row

        # Calculate range (A1 to last column, last row)
        start_col = get_column_letter(1)
        end_col = get_column_letter(num_columns)
        table_range = f"{start_col}1:{end_col}{num_rows}"

        # Create table
        table = Table(displayName=table_name, ref=table_range)

        # Style the table
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style

        # Add table to sheet
        sheet.add_table(table)

    @classmethod
    def _add_dropdown_validation(cls, sheet, column: int, options: List[str],
                                 start_row: int = 2, end_row: int = 1000):
        """Add data validation dropdown to a column.

        Args:
            sheet: The worksheet
            column: Column number (1-indexed)
            options: List of options for the dropdown
            start_row: Starting row (default 2, after header)
            end_row: Ending row
        """
        if not options:
            return

        # Create comma-separated list of options
        # Excel list validation formula format: "option1,option2,option3"
        options_str = ','.join(options)
        formula = f'"{options_str}"'

        # Create data validation
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = 'Invalid value'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = 'Please select from the dropdown'
        dv.promptTitle = 'Select Value'

        # Apply to column range
        col_letter = get_column_letter(column)
        dv_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
        dv.add(dv_range)

        # Add validation to sheet
        sheet.add_data_validation(dv)

    @classmethod
    def _add_sheet_reference_dropdown(cls, sheet, column: int, source_sheet,
                                      source_column_name: str,
                                      start_row: int = 2, end_row: int = 1000):
        """Add data validation dropdown that references values from another sheet using table column reference.

        Uses structured table references (e.g., PagesTable[id]) which automatically expand when new rows are added.

        Args:
            sheet: The worksheet to add validation to
            column: Column number (1-indexed) in the target sheet
            source_sheet: The source worksheet to reference
            source_column_name: Name of the column in source sheet to reference
            start_row: Starting row (default 2, after header)
            end_row: Ending row
        """
        try:
            # Find the source column index in the source sheet
            source_headers = [cell.value for cell in source_sheet[1]]
            if source_column_name not in source_headers:
                current_app.logger.warning(f"Source column '{source_column_name}' not found in source sheet")
                return

            source_col_idx = source_headers.index(source_column_name) + 1
            source_col_letter = get_column_letter(source_col_idx)

            # Prepare sheet reference
            source_sheet_name = source_sheet.title
            # Escape sheet name if it contains spaces or special characters
            if ' ' in source_sheet_name or '-' in source_sheet_name:
                source_sheet_ref = f"'{source_sheet_name}'"
            else:
                source_sheet_ref = source_sheet_name

            # Use OFFSET with COUNTA to create a dynamic range that only includes non-empty cells
            # OFFSET(start_cell, rows, cols, height, width)
            # COUNTA counts all non-empty cells in the column, subtract 1 to exclude header row
            # This creates a range that automatically expands/contracts based on actual data
            # Format: OFFSET(Pages!$A$2,0,0,COUNTA(Pages!$A:$A)-1,1)
            source_range = f"OFFSET({source_sheet_ref}!${source_col_letter}$2,0,0,COUNTA({source_sheet_ref}!${source_col_letter}:${source_col_letter})-1,1)"

            current_app.logger.debug(f"Creating dropdown with OFFSET formula: {source_range}")

            # Create data validation with reference formula
            dv = DataValidation(type="list", formula1=source_range, allow_blank=True)
            # Generic error messages since this is used for both page_id and section_id
            dv.error = f'Invalid {source_column_name}. Please select from the dropdown.'
            dv.errorTitle = 'Invalid Entry'
            dv.prompt = f'Please select a {source_column_name} from the {source_sheet.title} sheet'
            dv.promptTitle = f'Select {source_column_name.replace("_", " ").title()}'

            # Apply to column range
            col_letter = get_column_letter(column)
            dv_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
            dv.add(dv_range)

            # Add validation to sheet
            sheet.add_data_validation(dv)

            current_app.logger.info(f"Successfully added sheet reference dropdown: {source_range} for column {col_letter}, range {dv_range}")

        except Exception as e:
            current_app.logger.error(f"Could not add sheet reference dropdown: {e}", exc_info=True)

    @classmethod
    def _get_item_type_options(cls) -> List[str]:
        """Get all supported item types including plugin types.

        Returns:
            List of all item types: standard types (indicator, question, document_field, matrix)
            plus all active plugin field types (prefixed with 'plugin_')

        Note: These match the hardcoded values in form_builder.html dropdown:
        - indicator (Indicator from Bank)
        - question (New Question)
        - document_field (Document Field - stored as document_field in DB, shown as 'document' in UI)
        - matrix (Matrix Table)
        - plugin_* (dynamically added from plugin system)
        """
        item_types = []

        # Add standard item types - these are the hardcoded fallback values currently in the system
        # Matching the form_builder.html dropdown options (lines 1624-1627)
        # Note: UI shows 'document' but DB stores 'document_field', so we use 'document_field' here
        standard_types = ['indicator', 'question', 'document_field', 'matrix']
        item_types.extend(standard_types)

        # Add plugin field types if plugin manager is available
        try:
            if hasattr(current_app, 'plugin_manager'):
                plugin_manager = current_app.plugin_manager
                active_field_types = plugin_manager.list_active_field_types()

                # Prefix each plugin field type with 'plugin_'
                for field_type in active_field_types:
                    plugin_item_type = f'plugin_{field_type}'
                    if plugin_item_type not in item_types:
                        item_types.append(plugin_item_type)
        except Exception as e:
            current_app.logger.warning(f"Could not fetch plugin field types: {e}")

        # Sort for consistent ordering
        return sorted(item_types)

    @classmethod
    def _get_type_options_from_database(cls) -> List[str]:
        """Get distinct type values from FormItem table dynamically.

        Returns:
            List of unique type values found in the database, sorted alphabetically
        """
        try:
            # Query distinct non-null type values from FormItem
            distinct_types = db.session.query(FormItem.type).filter(
                FormItem.type.isnot(None),
                FormItem.type != ''
            ).distinct().all()

            # Extract values and sort
            type_options = sorted([t[0] for t in distinct_types if t[0]])

            # If no types found in database, return common defaults
            if not type_options:
                type_options = ['Number', 'Percentage', 'Text', 'Boolean', 'Date', 'Choice']

            return type_options

        except Exception as e:
            current_app.logger.warning(f"Could not fetch type options from database: {e}")
            # Return common defaults as fallback
            return ['Number', 'Percentage', 'Text', 'Boolean', 'Date', 'Choice']

    @classmethod
    def _add_duplicate_highlighting(cls, sheet, column_name: str, headers: List[str],
                                     num_rows: int):
        """Add conditional formatting to highlight duplicate values in a column.

        Args:
            sheet: The worksheet
            column_name: Name of the column to check for duplicates
            headers: List of header names (to find column index)
            num_rows: Total number of rows (including header)
        """
        if num_rows < 3:  # Need at least header + 2 data rows to have duplicates
            return

        try:
            # Find column index
            if column_name not in headers:
                return

            col_idx = headers.index(column_name) + 1  # +1 because Excel is 1-indexed
            col_letter = get_column_letter(col_idx)

            # Create range for data rows (skip header row)
            data_range = f"{col_letter}2:{col_letter}{num_rows}"

            # Create duplicate values rule with yellow background
            # Use COUNTIF formula to detect duplicates: COUNTIF($A$2:$A$100, A2)>1
            # This formula checks if the current cell value appears more than once in the range
            duplicate_fill = PatternFill(start_color=cls.IFRC_COLORS['YELLOW'],
                                        end_color=cls.IFRC_COLORS['YELLOW'],
                                        fill_type='solid')

            # Formula: COUNTIF(absolute_range, relative_cell) > 1
            # $col_letter$2:$col_letter$num_rows is the absolute range
            # col_letter2 is the relative cell reference (will adjust per row)
            formula = f'COUNTIF(${col_letter}$2:${col_letter}${num_rows},{col_letter}2)>1'

            duplicate_rule = FormulaRule(formula=[formula], fill=duplicate_fill)

            # Add conditional formatting rule
            sheet.conditional_formatting.add(data_range, duplicate_rule)

        except Exception as e:
            # Log but don't fail if conditional formatting can't be added
            current_app.logger.warning(f"Could not add duplicate highlighting for column {column_name}: {e}")

    @classmethod
    def _style_header_cell(cls, cell, is_required=True):
        """Apply IFRC styling to header cell.

        Args:
            cell: The cell to style
            is_required: If True, style as required (red), if False, style as optional (blue)
        """
        cell.font = Font(bold=True, color=cls.IFRC_COLORS['WHITE'])
        # Use red for required columns, blue for optional columns
        fill_color = cls.IFRC_COLORS['DARK_RED'] if is_required else cls.IFRC_COLORS['DARK_BLUE']
        cell.fill = PatternFill(start_color=fill_color,
                               end_color=fill_color,
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    @classmethod
    def _auto_size_columns(cls, sheet, num_columns):
        """Auto-size columns for better readability."""
        for col_idx in range(1, num_columns + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in sheet[column_letter]:
                with suppress(Exception):
                    if row.value:
                        lines = str(row.value).splitlines() or [str(row.value)]
                        max_length = max(max_length, max(len(line) for line in lines))
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            sheet.column_dimensions[column_letter].width = adjusted_width

