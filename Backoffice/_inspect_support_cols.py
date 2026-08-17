import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter

path = r"app/static/templates/unified_country_plan.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)
ws = wb["Planned Bilateral Support"]

print("--- rows 4-12 cols A-K ---")
for row in range(4, 13):
    vals = []
    for col in range(1, 12):
        c = ws.cell(row, col)
        v = c.value
        if isinstance(v, str) and len(v) > 40:
            v = v[:40] + "..."
        vals.append(f"{get_column_letter(col)}={v!r}")
    print(f"row{row}: " + " | ".join(vals))

wb2 = openpyxl.load_workbook(path, data_only=True)
ws2 = wb2["Planned Bilateral Support"]
print("\n--- data_only row5-8 col C and K ---")
for row in range(5, 9):
    print(row, "C=", ws2.cell(row, 3).value, "K=", ws2.cell(row, 11).value)

wb.close()
wb2.close()
