"""One-off maintenance: set footnote translation placeholders for dynamic NS counts."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

DATA_SOURCE = {
    "EN": (
        "*{year} data is based on reports received from {upr_ns} NSs through the unified "
        "reporting process and {fdrs_ns} NSs through FDRS. There is no complete overlap "
        "between the two data collection processes (in NSs and in indicators)."
    ),
    "FR": (
        "* Les données de {year} reposent sur les rapports de {upr_ns} SN via le processus "
        "de rapport unifié et de {fdrs_ns} SN via le FDRS. Il n'y a pas de chevauchement "
        "complet entre les deux processus de collecte (ni en termes de SN ni d'indicateurs)."
    ),
    "SP": (
        "* Los datos de {year} se basan en los informes de {upr_ns} Sociedades Nacionales "
        "a través del proceso unificado de presentación de informes y de {fdrs_ns} "
        "Sociedades Nacionales a través del banco de datos y sistema de información general "
        "interno (FDRS). No existe una coincidencia completa entre los dos procesos de "
        "recopilación (ni en las Sociedades Nacionales ni en los indicadores)."
    ),
    "AR": (
        "*تستند بيانات عام {year} إلى تقارير {upr_ns} جمعيات وطنية من خلال عملية التقارير "
        "الموحدة و{fdrs_ns} جمعية وطنية من خلال نظام قاعدة البيانات ونظام الإفادة في "
        "الاتحاد الدولي. لا يوجد تطابق كامل بين عمليتي جمع البيانات (لا في الجمعيات "
        "الوطنية ولا في المؤشرات)."
    ),
}

SP1_PREFIX = {
    "EN": "See also indicator for Disaster Risk Reduction under SP2.\n",
    "FR": "Voir également l'indicateur pour la réduction des risques de catastrophe dans le cadre du PS2.\n",
    "SP": "Véase también el indicador para la reducción del riesgo de desastres en el marco de la prioridad estratégica 2.\n",
    "AR": "انظر أيضًا مؤشر الحد من مخاطر الكوارث ضمن SP2.\n",
}

FOOTNOTE_VALUES = {
    "footnote.default": dict(DATA_SOURCE),
    "footnote.ef4": dict(DATA_SOURCE),
    "footnote.sp1": {lang: SP1_PREFIX[lang] + DATA_SOURCE[lang] for lang in DATA_SOURCE},
    "footnote.sp2": dict(DATA_SOURCE),
}


def patch_workbook(path: Path) -> int:
    wb = load_workbook(path)
    ws = wb["Translations"]
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    changed = 0
    for row in ws.iter_rows(min_row=2):
        footnote_id = str(row[headers["id"] - 1].value or "").strip()
        if footnote_id not in FOOTNOTE_VALUES:
            continue
        for lang, value in FOOTNOTE_VALUES[footnote_id].items():
            cell = row[headers[lang] - 1]
            if cell.value != value:
                cell.value = value
                changed += 1
    wb.save(path)
    return changed


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    paths = [
        root / "SG Report.xlsx",
        root.parent / "instance" / "uploads" / "pb_progress" / "source" / "SG_Report.xlsx",
    ]
    for path in paths:
        if not path.exists():
            print(f"skip missing {path}")
            continue
        count = patch_workbook(path)
        print(f"updated {path} ({count} cells)")


if __name__ == "__main__":
    main()
