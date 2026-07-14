from app.utils.transactions import request_transaction_rollback
from contextlib import suppress
from app.utils.datetime_helpers import utcnow
from app.utils.advanced_validation import validate_upload_extension_and_mime
from app.utils.file_parsing import EXCEL_EXTENSIONS
from app.utils.api_helpers import GENERIC_ERROR_MESSAGE, get_json_safe
from app.utils.api_responses import (
    json_bad_request, json_forbidden, json_ok, json_ok_result,
    json_server_error, json_form_errors, require_json_data,
)
from flask import (
    render_template, request, flash, redirect, url_for,
    current_app, make_response, send_file,
)
from flask_babel import _
from flask_login import current_user
from app import db
from config import Config
from collections import defaultdict

from app.models import (
    Sector,
    SubSector,
    IndicatorBank,
    IndicatorBankHistory,
    IndicatorSuggestion,
    CommonWord,
    IndicatorBankSpef,
    IndicatorBankType,
    IndicatorBankUnit,
)
from app.models.enums import IndicatorSuggestionStatusValue
from sqlalchemy.orm import joinedload
from app.forms.system import IndicatorBankForm, CommonWordForm
from app.routes.admin.shared import permission_required, user_has_permission
from app.utils.request_utils import get_json_or_form, is_json_request, get_request_data
from sqlalchemy import func, or_, cast, String
from app.services.indicator_neural_map import build_embedding_scatter, probe_query_embedding
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import json
import os
import io
import pandas as pd

from app.routes.admin.system_admin import bp
from app.routes.admin.system_admin.helpers import track_indicator_changes, indicator_bank_history_snapshot


def _indicator_bank_dynamic_list_values():
    """Read monitoring_questions / related_programs from POST (JSON or form)."""
    if request.method != "POST":
        return {"monitoring_questions": [], "related_programs": []}
    data = get_request_data()
    return {
        "monitoring_questions": [
            str(v).strip()
            for v in data.getlist("monitoring_questions")
            if v and str(v).strip()
        ],
        "related_programs": [
            str(v).strip()
            for v in data.getlist("related_programs")
            if v and str(v).strip()
        ],
    }


_INDICATOR_BANK_TABS = frozenset({'indicators', 'sectors', 'common_words', 'types', 'units', 'spef'})


def _resolve_indicator_bank_tab() -> str:
    tab = (request.args.get('tab') or 'indicators').strip().lower()
    if tab == 'measurement':
        return 'types'
    return tab if tab in _INDICATOR_BANK_TABS else 'indicators'


# === Indicator Bank Management Routes ===
@bp.route("/indicator_bank", methods=["GET"])
@permission_required('admin.indicator_bank.view')
def manage_indicator_bank():
    active_tab = _resolve_indicator_bank_tab()
    search = request.args.get('search', '')
    sector_filter = request.args.get('sector', '')
    type_filter = request.args.get('type', '')

    if active_tab == 'sectors' and not user_has_permission('admin.organization.manage'):
        active_tab = 'indicators'
    if active_tab in ('common_words', 'types', 'units', 'spef') and not user_has_permission('admin.indicator_bank.edit'):
        active_tab = 'indicators'

    query = IndicatorBank.query.options(
        joinedload(IndicatorBank.measurement_type),
        joinedload(IndicatorBank.measurement_unit),
        joinedload(IndicatorBank.spef_area),
    )

    if search:
        query = query.filter(
            or_(
                IndicatorBank.name.contains(search),
                IndicatorBank.definition.contains(search),
                IndicatorBank.type.contains(search)
            )
        )

    if sector_filter:
        sector_ids = [
            row[0]
            for row in db.session.query(Sector.id)
            .filter(Sector.name.ilike(f"%{sector_filter}%"))
            .all()
        ]
        if sector_ids:
            query = query.filter(
                or_(*[
                    cast(IndicatorBank.sector['primary'], String) == str(sid)
                    for sid in sector_ids
                ])
            )
        elif sector_filter.isdigit():
            sid = int(sector_filter)
            query = query.filter(cast(IndicatorBank.sector['primary'], String) == str(sid))
        else:
            query = query.filter(db.false())

    if type_filter:
        query = query.filter(IndicatorBank.type == type_filter)

    indicators = query.order_by(IndicatorBank.name).all()
    total_count = len(indicators)

    sector_ids = set()
    subsector_ids = set()

    for indicator in indicators:
        if indicator.sector:
            for level in ['primary', 'secondary', 'tertiary']:
                if level in indicator.sector:
                    sector_ids.add(indicator.sector[level])
        if indicator.sub_sector:
            for level in ['primary', 'secondary', 'tertiary']:
                if level in indicator.sub_sector:
                    subsector_ids.add(indicator.sub_sector[level])

    sectors_dict = {}
    if sector_ids:
        sectors = Sector.query.filter(Sector.id.in_(sector_ids)).all()
        sectors_dict = {sector.id: sector for sector in sectors}

    subsectors_dict = {}
    if subsector_ids:
        subsectors = SubSector.query.filter(SubSector.id.in_(subsector_ids)).all()
        subsectors_dict = {subsector.id: subsector for subsector in subsectors}

    from app.services.indicator_bank_service import attach_indicator_usage_cache
    attach_indicator_usage_cache(indicators)

    for indicator in indicators:
        indicator._cached_sectors = {}
        indicator._cached_subsectors = {}

        if indicator.sector:
            for level in ['primary', 'secondary', 'tertiary']:
                if level in indicator.sector:
                    sector_id = indicator.sector[level]
                    if sector_id in sectors_dict:
                        indicator._cached_sectors[level] = sectors_dict[sector_id]

        if indicator.sub_sector:
            for level in ['primary', 'secondary', 'tertiary']:
                if level in indicator.sub_sector:
                    subsector_id = indicator.sub_sector[level]
                    if subsector_id in subsectors_dict:
                        indicator._cached_subsectors[level] = subsectors_dict[subsector_id]

    if is_json_request():
        indicators_data = []
        for indicator in indicators:
            sector_name = None
            subsector_name = None
            if indicator.sector:
                for level in ['primary', 'secondary', 'tertiary']:
                    if level in indicator.sector and indicator.sector[level]:
                        sector_id = indicator.sector[level]
                        if sector_id in sectors_dict:
                            sector_name = sectors_dict[sector_id].name
                            break
            if indicator.sub_sector:
                for level in ['primary', 'secondary', 'tertiary']:
                    if level in indicator.sub_sector and indicator.sub_sector[level]:
                        subsector_id = indicator.sub_sector[level]
                        if subsector_id in subsectors_dict:
                            subsector_name = subsectors_dict[subsector_id].name
                            break

            indicators_data.append({
                'id': indicator.id,
                'name': indicator.name or '',
                'definition': indicator.definition if hasattr(indicator, 'definition') else None,
                'type': indicator.type if hasattr(indicator, 'type') else None,
                'sector': sector_name,
                'sub_sector': subsector_name,
                'unit': indicator.unit if hasattr(indicator, 'unit') else None,
                'fdrs_kpi_code': getattr(indicator, 'fdrs_kpi_code', None),
                'area': getattr(indicator, 'area', None),
                'area_label': getattr(indicator, 'area_label', None),
                'spef_label': getattr(indicator, 'area_label', None),
                'template_count': indicator.template_count,
                'data_value_count': indicator.data_value_count,
                'usage_count': indicator.template_count,
            })
        return json_ok(indicators=indicators_data, count=len(indicators_data), total_count=total_count)

    sector_names = db.session.query(Sector.name).distinct().order_by(Sector.name).all()
    types = db.session.query(IndicatorBank.type).distinct().filter(IndicatorBank.type.isnot(None)).order_by(IndicatorBank.type).all()

    pending_suggestions_count = IndicatorSuggestion.query.filter(
        IndicatorSuggestion.status == IndicatorSuggestionStatusValue.pending
    ).count()

    template_ctx = {
        'indicators': indicators,
        'sectors': [s[0] for s in sector_names],
        'types': [t[0] for t in types if t[0]],
        'search': search,
        'sector_filter': sector_filter,
        'type_filter': type_filter,
        'title': 'Manage Indicator Bank',
        'total_count': total_count,
        'pending_suggestions_count': pending_suggestions_count,
        'active_tab': active_tab,
        'can_edit_ml': user_has_permission('admin.indicator_bank.edit'),
    }

    if user_has_permission('admin.organization.manage'):
        all_sectors = Sector.query.order_by(Sector.name).all()
        all_subsectors = SubSector.query.order_by(SubSector.name).all()
        subsectors_by_sector = defaultdict(list)
        for sub in all_subsectors:
            subsectors_by_sector[sub.sector_id].append(sub)
        from app.routes.admin.system_admin.indicator_lookups import (
            batch_sector_indicator_counts,
            batch_subsector_indicator_counts,
        )
        template_ctx.update(
            sectors=all_sectors,
            subsectors=all_subsectors,
            subsectors_by_sector=dict(subsectors_by_sector),
            sector_indicator_counts=batch_sector_indicator_counts([s.id for s in all_sectors]),
            subsector_indicator_counts=batch_subsector_indicator_counts([s.id for s in all_subsectors]),
        )

    if user_has_permission('admin.indicator_bank.edit'):
        cw_query = CommonWord.query
        cw_search = request.args.get('search', '')
        if cw_search:
            cw_query = cw_query.filter(
                or_(CommonWord.term.contains(cw_search), CommonWord.meaning.contains(cw_search))
            )
        template_ctx['common_words'] = cw_query.order_by(CommonWord.term).all()
        template_ctx['search'] = cw_search

        from app.routes.admin.system_admin.indicator_lookups import (
            _ensure_national_society_unit_row,
            batch_spef_usage_counts,
            batch_type_usage_counts,
            batch_unit_usage_counts,
        )
        _ensure_national_society_unit_row()
        ml_types = IndicatorBankType.query.order_by(
            IndicatorBankType.sort_order, IndicatorBankType.name
        ).all()
        ml_units = IndicatorBankUnit.query.order_by(
            IndicatorBankUnit.sort_order, IndicatorBankUnit.name
        ).all()
        spef_rows = IndicatorBankSpef.query.order_by(
            IndicatorBankSpef.sort_order, IndicatorBankSpef.code
        ).all()
        template_ctx.update(
            types=ml_types,
            units=ml_units,
            type_usage=batch_type_usage_counts([t.id for t in ml_types]),
            unit_usage=batch_unit_usage_counts([u.id for u in ml_units]),
            spef_rows=spef_rows,
            spef_usage=batch_spef_usage_counts([r.id for r in spef_rows]),
        )

    return render_template("admin/indicator_bank/indicator_bank.html", **template_ctx)


@bp.route("/indicator_bank/neural_map", methods=["GET"])
@permission_required('admin.indicator_bank.view')
def indicator_bank_neural_map():
    """Interactive neural map visualization of indicator bank vector embeddings."""
    return render_template("admin/indicator_bank/neural_map.html", title="Indicators Neural Map")


@bp.route("/indicator_bank/neural_map/data", methods=["GET"])
@permission_required('admin.indicator_bank.view')
def indicator_bank_neural_map_data():
    """JSON: 2D/3D scatter data for the indicators neural map."""
    max_nodes = request.args.get("max_nodes", type=int) or 5000
    n_neighbors = request.args.get("neighbors", type=int) or 8
    method = request.args.get("method", "pca")
    dimensions = request.args.get("dimensions", 3, type=int)
    if method not in ("pca", "tsne"):
        method = "pca"
    if dimensions not in (2, 3):
        dimensions = 3
    try:
        result = build_embedding_scatter(
            max_nodes=max_nodes,
            n_neighbors=n_neighbors,
            method=method,
            dimensions=dimensions,
            exclude_archived=True,
        )
        return json_ok_result(result)
    except Exception as e:
        current_app.logger.exception("Neural map data failed")
        return json_server_error(GENERIC_ERROR_MESSAGE, nodes=[], groups=[], count=0)


@bp.route("/indicator_bank/neural_map/probe", methods=["POST"])
@permission_required('admin.indicator_bank.view')
def indicator_bank_neural_map_probe():
    """Embed a free-text query and return nearest indicator neighbours."""
    data = get_json_safe()
    query = (data.get("query") or "").strip()
    if not query:
        return json_bad_request("query is required")
    try:
        result = probe_query_embedding(query, top_k=10)
        return json_ok_result(result)
    except Exception as e:
        current_app.logger.exception("Neural map probe failed")
        return json_server_error(GENERIC_ERROR_MESSAGE)


@bp.route("/indicator_bank/sync_remote", methods=["POST"])
@permission_required("admin.indicator_bank.edit")
def sync_indicator_bank_remote():
    """Trigger background sync from the external IFRC Indicator Bank platform."""
    from app.services.indicatorbank_remote_sync_service import start_remote_sync, get_remote_sync_state

    api_key = os.getenv("IFRC_INDICATORBANK_API_KEY", "").strip()
    api_url = os.getenv("IFRC_INDICATORBANK_API_URL", "https://ifrc-indicatorbank.azurewebsites.net/api/indicator").strip()
    data = get_json_or_form()
    limit_raw = data.get("limit") or request.form.get("limit")
    try:
        limit = int(limit_raw) if limit_raw not in (None, "", "null") else None
    except Exception as e:
        current_app.logger.debug("indicator bank sync limit parse failed: %s", e)
        limit = None

    ok, msg = start_remote_sync(current_app._get_current_object(), api_url=api_url, api_key=api_key, limit=limit)
    return (json_ok(success=bool(ok), message=msg, state=get_remote_sync_state())
            if ok else json_bad_request(msg, success=False, state=get_remote_sync_state()))


@bp.route("/indicator_bank/sync_remote/status", methods=["GET"])
@permission_required("admin.indicator_bank.edit")
def sync_indicator_bank_remote_status():
    """Fetch background sync status/result."""
    from app.services.indicatorbank_remote_sync_service import get_remote_sync_state
    return json_ok(state=get_remote_sync_state())

@bp.route("/indicator_bank/view/<int:id>", methods=["GET"])
@permission_required('admin.indicator_bank.view')
def view_indicator_bank(id):
    indicator = IndicatorBank.query.get_or_404(id)
    return render_template("admin/indicator_bank/view_indicator_bank.html",
                         indicator=indicator,
                         title=f"View Indicator: {indicator.name}")

@bp.route("/indicator_bank/add", methods=["GET", "POST"])
@permission_required('admin.indicator_bank.create')
def add_indicator_bank():
    try:
        from app.services.authorization_service import AuthorizationService
        can_archive = (
            AuthorizationService.is_system_manager(current_user)
            or AuthorizationService.has_rbac_permission(current_user, 'admin.indicator_bank.archive')
        )
    except Exception as e:
        current_app.logger.debug("can_archive check failed (add_indicator_bank): %s", e)
        can_archive = False

    if request.method == 'POST':
        form = IndicatorBankForm(formdata=get_request_data())
    else:
        form = IndicatorBankForm()

    form._populate_choices()

    if form.validate_on_submit():
        try:
            new_indicator = IndicatorBank(
                name=form.name.data,
                definition=form.definition.data,
                # Placeholder; populate_indicator_bank sets FKs and syncs canonical type/unit strings
                type="number",
                unit=None,
                fdrs_kpi_code=(form.fdrs_kpi_code.data or '').strip() or None,
                emergency=form.emergency.data,
                archived=(form.archived.data if can_archive else False)
            )

            form.populate_indicator_bank(new_indicator)

            if not can_archive:
                new_indicator.archived = False

            db.session.add(new_indicator)
            db.session.flush()

            history = IndicatorBankHistory(
                indicator_bank_id=new_indicator.id,
                user_id=current_user.id,
                change_type='CREATED',
                change_description=f'Indicator "{new_indicator.name}" created by {current_user.name or current_user.email}',
                **indicator_bank_history_snapshot(new_indicator),
            )
            db.session.add(history)

            db.session.flush()

            flash(f"Indicator '{new_indicator.name}' added successfully.", "success")
            return redirect(url_for("system_admin.manage_indicator_bank"))

        except Exception as e:
            request_transaction_rollback()
            flash("Error adding indicator.", "danger")
            current_app.logger.error(f"Error adding indicator: {e}", exc_info=True)

    return render_template("admin/indicator_bank/add_indicator_bank.html",
                         form=form,
                         title="Add New Indicator",
                         form_lists=_indicator_bank_dynamic_list_values())

@bp.route("/indicator_bank/edit/<int:id>", methods=["GET", "POST"])
@permission_required('admin.indicator_bank.edit')
def edit_indicator_bank(id):
    indicator = IndicatorBank.query.get_or_404(id)

    try:
        from app.services.authorization_service import AuthorizationService
        can_archive = (
            AuthorizationService.is_system_manager(current_user)
            or AuthorizationService.has_rbac_permission(current_user, 'admin.indicator_bank.archive')
        )
    except Exception as e:
        current_app.logger.debug("can_archive check failed (edit_indicator_bank): %s", e)
        can_archive = False

    if request.method == 'POST':
        is_ajax = is_json_request()
        ajax_data = None

        # WAF-bypass form submits arrive as JSON with a base64 "payload" key.
        # They are full form submissions and must go through the normal form
        # handler (which calls get_request_data() and decodes the payload).
        if is_ajax:
            _raw_json = get_json_safe()
            if _raw_json.get("payload") or _raw_json.get("payload_b64"):
                _payload_data = get_request_data().to_dict()
                if _payload_data.pop("_indicator_bank_partial_update", None):
                    ajax_data = _payload_data
                else:
                    is_ajax = False

        if is_ajax:
            try:
                data = ajax_data if ajax_data is not None else get_json_safe()
                err = require_json_data(data)
                if err:
                    return err

                if 'archived' in data and not can_archive:
                    return json_forbidden('Archive permission required.')

                updated_fields = []
                for field_name, value in data.items():

                    if value is not None:
                        if field_name == 'archived' and not can_archive:
                            continue
                        if field_name.startswith('name_') and field_name != 'name':
                            language = field_name.replace('name_', '')
                            indicator.set_name_translation(language, value)
                            updated_fields.append(field_name)
                        elif field_name.startswith('definition_') and field_name != 'definition':
                            language = field_name.replace('definition_', '')
                            indicator.set_definition_translation(language, value)
                            updated_fields.append(field_name)
                        elif field_name.startswith('aggregated_label_') and field_name != 'aggregated_label':
                            language = field_name.replace('aggregated_label_', '')
                            indicator.set_aggregated_label_translation(language, value)
                            updated_fields.append(field_name)
                        elif hasattr(indicator, field_name):
                            setattr(indicator, field_name, value)
                            updated_fields.append(field_name)
                        else:
                            continue

                if updated_fields:
                    from sqlalchemy.orm.attributes import flag_modified

                    if any(field.startswith('name_') for field in updated_fields):
                        flag_modified(indicator, 'name_translations')
                    if any(field.startswith('definition_') for field in updated_fields):
                        flag_modified(indicator, 'definition_translations')
                    if any(field.startswith('aggregated_label_') for field in updated_fields):
                        flag_modified(indicator, 'aggregated_label_translations')

                    db.session.add(indicator)

                if updated_fields:
                    history = IndicatorBankHistory(
                        indicator_bank_id=indicator.id,
                        user_id=current_user.id,
                        change_type='UPDATED',
                        change_description=f"Auto-translate update: {', '.join(updated_fields)} updated by {current_user.name or current_user.email}",
                        **indicator_bank_history_snapshot(indicator),
                    )
                    db.session.add(history)
                    db.session.flush()

                    return json_ok(
                        message=f'Indicator updated successfully. Updated fields: {", ".join(updated_fields)}'
                    )
                else:
                    return json_bad_request('No valid fields to update')

            except Exception as e:
                request_transaction_rollback()
                current_app.logger.error(f"Error updating indicator {id} via AJAX: {e}", exc_info=True)
                return json_server_error(GENERIC_ERROR_MESSAGE)

        form = IndicatorBankForm(formdata=get_request_data())

    else:
        form = IndicatorBankForm(obj=indicator)

    form._populate_choices()

    if request.method == 'GET':
        form.populate_from_indicator_bank(indicator)

    if form.validate_on_submit():
        try:
            try:
                _langs = current_app.config.get("TRANSLATABLE_LANGUAGES") or []
            except Exception as e:
                current_app.logger.debug("TRANSLATABLE_LANGUAGES config failed: %s", e)
                _langs = []

            name_translations = {"en": form.name.data}
            for _lang in _langs:
                _lang = str(_lang or "").strip().lower()
                if not _lang or _lang == "en":
                    continue
                _field = getattr(form, f"name_{_lang}", None)
                if _field is not None:
                    name_translations[_lang] = _field.data

            prev_archived = indicator.archived
            if not can_archive:
                try:
                    form.archived.data = prev_archived
                except Exception as e:
                    current_app.logger.debug("form.archived reset failed: %s", e)

            form_data = {
                'name': form.name.data,
                'type': form.type.data,
                'unit': form.unit.data,
                'fdrs_kpi_code': (form.fdrs_kpi_code.data or '').strip() or None,
                'definition': form.definition.data,
                'aggregated_label': form.aggregated_label.data,
                'name_translations': name_translations,
                'comments': form.comments.data,
                'related_programs': IndicatorBankForm.related_programs_from_request(),
                'area': form.area.data,
                'data_source': form.data_source.data,
                'disaggregation_guidance': form.disaggregation_guidance.data,
                'monitoring_questions': IndicatorBankForm.monitoring_questions_from_request(),
                'tags': [t.strip() for t in (form.tags.data or '').split(',') if t.strip()] or None,
                'emergency': form.emergency.data,
                'archived': prev_archived if not can_archive else form.archived.data,
                'sector_primary': form.sector_primary.data,
                'sector_secondary': form.sector_secondary.data,
                'sector_tertiary': form.sector_tertiary.data,
                'sub_sector_primary': form.sub_sector_primary.data,
                'sub_sector_secondary': form.sub_sector_secondary.data,
                'sub_sector_tertiary': form.sub_sector_tertiary.data
            }

            changes = track_indicator_changes(indicator, form_data, current_user)

            if changes:
                change_description = "; ".join(changes)
            else:
                change_description = f"Indicator updated by {current_user.name or current_user.email} (no specific changes detected)"

            form.populate_indicator_bank(indicator)

            if not can_archive:
                indicator.archived = prev_archived

            history = IndicatorBankHistory(
                indicator_bank_id=indicator.id,
                user_id=current_user.id,
                change_type='UPDATED',
                change_description=change_description,
                **indicator_bank_history_snapshot(indicator),
            )
            db.session.add(history)

            db.session.flush()
            success_message = f"Indicator '{indicator.name}' updated successfully."
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return json_ok(message=success_message, indicator_id=indicator.id, name=indicator.name)
            flash(success_message, "success")
            return redirect(url_for("system_admin.manage_indicator_bank"))

        except Exception as e:
            request_transaction_rollback()
            current_app.logger.error(f"Error updating indicator {id}: {e}", exc_info=True)
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return json_server_error(GENERIC_ERROR_MESSAGE)
            flash("An error occurred. Please try again.", "danger")

    if request.method == "POST" and request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return json_form_errors(form, _("Please correct the errors in the form."))

    return render_template("admin/indicator_bank/edit_indicator_bank.html",
                         form=form,
                         indicator=indicator,
                         title=f"Edit Indicator: {indicator.name}",
                         form_lists=_indicator_bank_dynamic_list_values())

@bp.route("/indicator_bank/delete/<int:id>", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def delete_indicator_bank(id):
    indicator = IndicatorBank.query.get_or_404(id)

    try:
        history = IndicatorBankHistory(
            indicator_bank_id=indicator.id,
            user_id=current_user.id,
            change_type='DELETED',
            change_description=f'Indicator "{indicator.name}" deleted by {current_user.name or current_user.email}',
            **indicator_bank_history_snapshot(indicator),
        )
        db.session.add(history)

        db.session.delete(indicator)
        db.session.flush()

        flash(f"Indicator '{indicator.name}' deleted successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("An error occurred. Please try again.", "danger")
        current_app.logger.error(f"Error deleting indicator {id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.manage_indicator_bank"))

@bp.route("/indicator_bank/archive/<int:id>", methods=["POST"])
@permission_required('admin.indicator_bank.archive')
def archive_indicator_bank(id):
    indicator = IndicatorBank.query.get_or_404(id)

    try:
        new_archived_status = not indicator.archived

        history = IndicatorBankHistory(
            indicator_bank_id=indicator.id,
            user_id=current_user.id,
            change_type='ARCHIVED' if new_archived_status else 'UNARCHIVED',
            change_description=f'Indicator "{indicator.name}" {"archived" if new_archived_status else "unarchived"} by {current_user.name or current_user.email}',
            **indicator_bank_history_snapshot(indicator),
        )
        db.session.add(history)

        indicator.archived = new_archived_status
        db.session.flush()

        action = "archived" if new_archived_status else "unarchived"
        flash(f"Indicator '{indicator.name}' {action} successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("Error archiving indicator.", "danger")
        current_app.logger.error(f"Error archiving indicator {id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.manage_indicator_bank"))

@bp.route("/indicator_bank/translations/<int:id>", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def update_indicator_translations(id):
    indicator = IndicatorBank.query.get_or_404(id)

    try:
        translation_changes = []
        languages = current_app.config.get("TRANSLATABLE_LANGUAGES", None) or getattr(Config, "TRANSLATABLE_LANGUAGES", []) or []

        for lang in languages:
            name_key = f'name_{lang}'
            definition_key = f'definition_{lang}'

            if name_key in request.form:
                new_name = request.form[name_key].strip()
                old_name = indicator.get_name_translation(lang)
                if old_name != new_name:
                    if old_name is None and new_name:
                        translation_changes.append(f"{lang.upper()} Name: Added '{new_name}'")
                    elif old_name and new_name != old_name:
                        translation_changes.append(f"{lang.upper()} Name: Changed from '{old_name}' to '{new_name}'")
                    indicator.set_name_translation(lang, new_name)

            if definition_key in request.form:
                new_definition = request.form[definition_key].strip()
                old_definition = indicator.get_definition_translation(lang)
                if old_definition != new_definition:
                    if old_definition is None and new_definition:
                        translation_changes.append(f"{lang.upper()} Definition: Added '{new_definition[:50]}{'...' if len(new_definition) > 50 else ''}'")
                    elif old_definition and new_definition != old_definition:
                        translation_changes.append(f"{lang.upper()} Definition: Changed from '{old_definition[:50]}{'...' if len(old_definition) > 50 else ''}' to '{new_definition[:50]}{'...' if len(new_definition) > 50 else ''}'")
                    indicator.set_definition_translation(lang, new_definition)

        if translation_changes:
            change_description = "; ".join(translation_changes)
        else:
            change_description = f"Translations updated by {current_user.name or current_user.email} (no specific changes detected)"

        history = IndicatorBankHistory(
            indicator_bank_id=indicator.id,
            user_id=current_user.id,
            change_type='UPDATED',
            change_description=change_description,
            **indicator_bank_history_snapshot(indicator),
        )
        db.session.add(history)

        db.session.flush()
        flash("Translations updated successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("An error occurred. Please try again.", "danger")
        current_app.logger.error(f"Error updating indicator translations {id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.view_indicator_bank", id=id))


# === Session Management Routes ===
@bp.route("/sessions/cleanup", methods=["POST"])
@permission_required('admin.analytics.view')
def cleanup_sessions():
    """Cleanup inactive sessions"""
    try:
        from app.services.user_analytics_service import cleanup_inactive_sessions
        count = cleanup_inactive_sessions()
        flash(f"Successfully cleaned up {count} inactive sessions.", "success")
    except Exception as e:
        request_transaction_rollback()
        flash("An error occurred during session cleanup.", "danger")
        current_app.logger.error(f"Error during session cleanup: {e}", exc_info=True)

    return redirect(url_for("analytics.session_logs"))

@bp.route("/sessions/status", methods=["GET"])
@permission_required('admin.analytics.view')
def session_status():
    """Legacy endpoint. Redirect to analytics session logs."""
    return redirect(url_for("analytics.session_logs"))


# === Export/Import Routes ===
@bp.route("/indicator_bank/export", methods=["GET", "POST"])
@permission_required('admin.indicator_bank.view')
def export_indicators():
    """Export indicators to a fully human-readable, importable Excel workbook.

    Sheets:
      - Indicators   – main data (names, definitions, sectors, etc.)
      - Types        – measurement type catalog
      - Units        – measurement unit catalog
      - Sectors      – sector catalog
      - Sub-Sectors  – sub-sector catalog (with parent sector name)
      - Common Words – common words / glossary

    Supports:
    - GET: export all indicators
    - POST: export only selected IDs passed as comma-separated 'selected_ids'
    """
    try:
        selected_ids = None
        if request.method == "POST":
            ids_str = request.form.get('selected_ids')
            if ids_str:
                try:
                    selected_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
                except Exception as e:
                    current_app.logger.debug("selected_ids parse failed: %s", e)
                    selected_ids = None

        query = IndicatorBank.query
        if selected_ids:
            query = query.filter(IndicatorBank.id.in_(selected_ids))
        indicators = query.order_by(IndicatorBank.name).all()

        wb = Workbook()

        # ── Styling helpers ──────────────────────────────────────────────────
        def _hdr_fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        _FILL_BLUE   = _hdr_fill("1F4E79")   # Indicators
        _FILL_TEAL   = _hdr_fill("1A6B5A")   # Types / Units
        _FILL_PURPLE = _hdr_fill("5B3A8E")   # Sectors
        _FILL_GREEN  = _hdr_fill("2D6A4F")   # Sub-Sectors
        _FILL_BROWN  = _hdr_fill("7B3F00")   # Common Words
        _FONT_WHITE  = Font(bold=True, color="FFFFFF")
        _ALIGN_CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def _write_headers(ws, headers, fill, row=1):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = _FONT_WHITE
                c.fill = fill
                c.alignment = _ALIGN_CTR
            ws.row_dimensions[row].height = 22

        def _autofit(ws, max_width=60):
            for col in ws.columns:
                mx = max(
                    (len(str(cell.value or "")) for cell in col),
                    default=0
                )
                ws.column_dimensions[col[0].column_letter].width = min(mx + 2, max_width)

        def _mq_to_str(val):
            """Convert monitoring_questions list/json to semicolon-separated string."""
            if not val:
                return ""
            if isinstance(val, list):
                return "; ".join(str(v) for v in val if v)
            try:
                parsed = json.loads(str(val))
                if isinstance(parsed, list):
                    return "; ".join(str(v) for v in parsed if v)
            except Exception:
                pass
            return str(val)

        languages = (
            current_app.config.get("SUPPORTED_LANGUAGES")
            or getattr(Config, "LANGUAGES", None)
            or ["en"]
        )

        # ── Sheet 1: Indicators ──────────────────────────────────────────────
        ws = wb.active
        ws.title = "Indicators"

        name_lang_headers = [f"Name ({code})" for code in languages]
        def_lang_headers  = [f"Definition ({code})" for code in languages]
        agg_lang_headers  = [f"Aggregated Label ({code})" for code in languages if code != "en"]
        ind_headers = (
            ['ID', 'Name', 'Definition', 'Aggregated Label', 'Area', 'Type', 'Unit',
             'FDRS KPI Code', 'Data Source', 'Disaggregation Guidance',
             'Monitoring Questions', 'Tags', 'Emergency', 'Related Programs',
             'Archived', 'Comments', 'Created At']
            + name_lang_headers
            + def_lang_headers
            + agg_lang_headers
            + ['Sector Primary', 'Sector Secondary', 'Sector Tertiary',
               'SubSector Primary', 'SubSector Secondary', 'SubSector Tertiary']
        )
        _write_headers(ws, ind_headers, _FILL_BLUE)
        ws.freeze_panes = "B2"

        # Resolve sector/subsector id → name
        sector_ids, subsector_ids = set(), set()
        for ind in indicators:
            for level in ("primary", "secondary", "tertiary"):
                if ind.sector and ind.sector.get(level):
                    sector_ids.add(ind.sector[level])
                if ind.sub_sector and ind.sub_sector.get(level):
                    subsector_ids.add(ind.sub_sector[level])

        sector_names = (
            {s.id: s.name for s in Sector.query.filter(Sector.id.in_(sector_ids)).all()}
            if sector_ids else {}
        )
        subsector_names = (
            {ss.id: ss.name for ss in SubSector.query.filter(SubSector.id.in_(subsector_ids)).all()}
            if subsector_ids else {}
        )

        for row, ind in enumerate(indicators, 2):
            nt  = ind.name_translations or {}
            dt  = ind.definition_translations or {}
            alt = ind.aggregated_label_translations or {}
            sec = ind.sector or {}
            sub = ind.sub_sector or {}
            col = 1
            ws.cell(row=row, column=col, value=ind.id);                                                     col += 1
            ws.cell(row=row, column=col, value=ind.name);                                                   col += 1
            ws.cell(row=row, column=col, value=ind.definition);                                             col += 1
            ws.cell(row=row, column=col, value=getattr(ind, 'aggregated_label', None) or '');               col += 1
            ws.cell(row=row, column=col, value=getattr(ind, 'area', None) or '');                           col += 1
            ws.cell(row=row, column=col, value=ind.type);                                                   col += 1
            ws.cell(row=row, column=col, value=ind.unit);                                                   col += 1
            ws.cell(row=row, column=col, value=getattr(ind, 'fdrs_kpi_code', None) or '');                  col += 1
            ws.cell(row=row, column=col, value=getattr(ind, 'data_source', None) or '');                    col += 1
            ws.cell(row=row, column=col, value=getattr(ind, 'disaggregation_guidance', None) or '');        col += 1
            ws.cell(row=row, column=col, value=_mq_to_str(ind.monitoring_questions_list));                  col += 1
            ws.cell(row=row, column=col, value=", ".join(ind.tags_list));                                   col += 1
            ws.cell(row=row, column=col, value=ind.emergency);                                              col += 1
            ws.cell(row=row, column=col, value=", ".join(ind.related_programs_list));                     col += 1
            ws.cell(row=row, column=col, value=ind.archived);                                               col += 1
            ws.cell(row=row, column=col, value=getattr(ind, 'comments', None) or '');                       col += 1
            ws.cell(row=row, column=col,
                    value=ind.created_at.strftime('%Y-%m-%d') if ind.created_at else '');                   col += 1
            for code in languages:
                ws.cell(row=row, column=col,
                        value=nt.get(code) if code != "en" else (nt.get("en") or ind.name));               col += 1
            for code in languages:
                ws.cell(row=row, column=col,
                        value=dt.get(code) if code != "en" else (dt.get("en") or ind.definition));         col += 1
            for code in languages:
                if code == "en":
                    continue
                ws.cell(row=row, column=col, value=alt.get(code) or '');                                   col += 1
            for level in ("primary", "secondary", "tertiary"):
                sid = sec.get(level)
                ws.cell(row=row, column=col, value=sector_names.get(sid, '') if sid else '');               col += 1
            for level in ("primary", "secondary", "tertiary"):
                ssid = sub.get(level)
                ws.cell(row=row, column=col, value=subsector_names.get(ssid, '') if ssid else '');          col += 1

        _autofit(ws)

        # ── Sheet 2: Types ───────────────────────────────────────────────────
        ws_types = wb.create_sheet(title="Types")
        types_headers = (
            ['ID', 'Code', 'Name', 'Active', 'Sort Order']
            + [f"Name ({code})" for code in languages if code != "en"]
        )
        _write_headers(ws_types, types_headers, _FILL_TEAL)
        for row, mt in enumerate(
            IndicatorBankType.query.order_by(IndicatorBankType.sort_order, IndicatorBankType.name).all(), 2
        ):
            nt = mt.name_translations or {}
            col = 1
            ws_types.cell(row=row, column=col, value=mt.id);          col += 1
            ws_types.cell(row=row, column=col, value=mt.code);         col += 1
            ws_types.cell(row=row, column=col, value=mt.name);         col += 1
            ws_types.cell(row=row, column=col, value=mt.is_active);    col += 1
            ws_types.cell(row=row, column=col, value=mt.sort_order);   col += 1
            for code in languages:
                if code == "en":
                    continue
                ws_types.cell(row=row, column=col, value=nt.get(code, '')); col += 1
        _autofit(ws_types)

        # ── Sheet 3: Units ───────────────────────────────────────────────────
        ws_units = wb.create_sheet(title="Units")
        units_headers = (
            ['ID', 'Code', 'Name', 'Active', 'Sort Order', 'Allows Disaggregation']
            + [f"Name ({code})" for code in languages if code != "en"]
        )
        _write_headers(ws_units, units_headers, _FILL_TEAL)
        for row, mu in enumerate(
            IndicatorBankUnit.query.order_by(IndicatorBankUnit.sort_order, IndicatorBankUnit.name).all(), 2
        ):
            nt = mu.name_translations or {}
            col = 1
            ws_units.cell(row=row, column=col, value=mu.id);                    col += 1
            ws_units.cell(row=row, column=col, value=mu.code);                   col += 1
            ws_units.cell(row=row, column=col, value=mu.name);                   col += 1
            ws_units.cell(row=row, column=col, value=mu.is_active);              col += 1
            ws_units.cell(row=row, column=col, value=mu.sort_order);             col += 1
            ws_units.cell(row=row, column=col, value=mu.allows_disaggregation);  col += 1
            for code in languages:
                if code == "en":
                    continue
                ws_units.cell(row=row, column=col, value=nt.get(code, '')); col += 1
        _autofit(ws_units)

        # ── Sheet 4: Sectors ─────────────────────────────────────────────────
        ws_sectors = wb.create_sheet(title="Sectors")
        sectors_headers = (
            ['ID', 'Name', 'Description', 'Active', 'Display Order', 'Icon Class']
            + [f"Name ({code})" for code in languages if code != "en"]
        )
        _write_headers(ws_sectors, sectors_headers, _FILL_PURPLE)
        for row, s in enumerate(
            Sector.query.order_by(Sector.display_order, Sector.name).all(), 2
        ):
            nt = s.name_translations or {}
            col = 1
            ws_sectors.cell(row=row, column=col, value=s.id);             col += 1
            ws_sectors.cell(row=row, column=col, value=s.name);           col += 1
            ws_sectors.cell(row=row, column=col, value=s.description);    col += 1
            ws_sectors.cell(row=row, column=col, value=s.is_active);      col += 1
            ws_sectors.cell(row=row, column=col, value=s.display_order);  col += 1
            ws_sectors.cell(row=row, column=col, value=s.icon_class);     col += 1
            for code in languages:
                if code == "en":
                    continue
                ws_sectors.cell(row=row, column=col, value=nt.get(code, '')); col += 1
        _autofit(ws_sectors)

        # ── Sheet 5: Sub-Sectors ─────────────────────────────────────────────
        ws_subsectors = wb.create_sheet(title="Sub-Sectors")
        subsectors_headers = (
            ['ID', 'Sector', 'Name', 'Description', 'Active', 'Display Order', 'Icon Class']
            + [f"Name ({code})" for code in languages if code != "en"]
        )
        _write_headers(ws_subsectors, subsectors_headers, _FILL_GREEN)
        all_sectors_map = {s.id: s.name for s in Sector.query.all()}
        for row, ss in enumerate(
            SubSector.query.order_by(SubSector.display_order, SubSector.name).all(), 2
        ):
            nt = ss.name_translations or {}
            col = 1
            ws_subsectors.cell(row=row, column=col, value=ss.id);          col += 1
            ws_subsectors.cell(row=row, column=col,
                               value=all_sectors_map.get(ss.sector_id, '') if ss.sector_id else ''); col += 1
            ws_subsectors.cell(row=row, column=col, value=ss.name);        col += 1
            ws_subsectors.cell(row=row, column=col, value=ss.description); col += 1
            ws_subsectors.cell(row=row, column=col, value=ss.is_active);   col += 1
            ws_subsectors.cell(row=row, column=col, value=ss.display_order); col += 1
            ws_subsectors.cell(row=row, column=col, value=ss.icon_class);  col += 1
            for code in languages:
                if code == "en":
                    continue
                ws_subsectors.cell(row=row, column=col, value=nt.get(code, '')); col += 1
        _autofit(ws_subsectors)

        # ── Sheet 6: Common Words ────────────────────────────────────────────
        ws_cw = wb.create_sheet(title="Common Words")
        cw_headers = (
            ['ID', 'Term', 'Meaning', 'Active']
            + [f"Meaning ({code})" for code in languages if code != "en"]
        )
        _write_headers(ws_cw, cw_headers, _FILL_BROWN)
        for row, cw in enumerate(CommonWord.query.order_by(CommonWord.term).all(), 2):
            mt = cw.meaning_translations or {}
            col = 1
            ws_cw.cell(row=row, column=col, value=cw.id);       col += 1
            ws_cw.cell(row=row, column=col, value=cw.term);     col += 1
            ws_cw.cell(row=row, column=col, value=cw.meaning);  col += 1
            ws_cw.cell(row=row, column=col, value=cw.is_active); col += 1
            for code in languages:
                if code == "en":
                    continue
                ws_cw.cell(row=row, column=col, value=mt.get(code, '')); col += 1
        _autofit(ws_cw)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=indicators_export.xlsx'

        return response

    except Exception as e:
        current_app.logger.error(f"Error exporting indicators: {e}", exc_info=True)
        flash("Error exporting indicators.", "danger")
        return redirect(url_for("system_admin.manage_indicator_bank"))


# === API Endpoints ===

@bp.route("/api/indicator-count", methods=["POST"])
@permission_required('admin.indicator_bank.view')
def get_filtered_indicator_count():
    """API endpoint to get count of indicators matching filters"""
    try:
        data = request.json
        filters = data.get('filters', [])
        section_id = data.get('section_id')
        include_indicators = data.get('include_indicators', False)

        query = db.session.query(IndicatorBank)

        if section_id:
            from app.models import FormSection
            section = FormSection.query.get(section_id)
            if section and section.indicator_filters_list:
                filters = section.indicator_filters_list

        for filter_obj in filters:
            field = filter_obj.get('field')
            values = filter_obj.get('values', [])

            if not field or not values:
                continue

            try:
                if field == 'type':
                    query = query.filter(IndicatorBank.type.in_(values))
                elif field == 'unit':
                    query = query.filter(IndicatorBank.unit.in_(values))
                elif field == 'emergency':
                    bool_values = [v.lower() == 'true' for v in values]
                    query = query.filter(IndicatorBank.emergency.in_(bool_values))
                elif field == 'archived':
                    bool_values = [v.lower() == 'true' for v in values]
                    query = query.filter(IndicatorBank.archived.in_(bool_values))
                elif field == 'related_programs':
                    conditions = []
                    for value in values:
                        conditions.append(
                            cast(IndicatorBank._related_programs_list, String).ilike(f'%{value}%')
                        )
                    if conditions:
                        query = query.filter(db.or_(*conditions))
                elif field == 'sector':
                    conditions = []
                    primary_only = filter_obj.get('primary_only', False)

                    for value in values:
                        if primary_only:
                            conditions.append(IndicatorBank.sector.cast(db.Text).like(f'%"primary": "{value}"%'))
                        else:
                            conditions.append(IndicatorBank.sector.cast(db.Text).like(f'%"primary": "{value}"%'))
                            conditions.append(IndicatorBank.sector.cast(db.Text).like(f'%"secondary": "{value}"%'))
                            conditions.append(IndicatorBank.sector.cast(db.Text).like(f'%"tertiary": "{value}"%'))
                    if conditions:
                        query = query.filter(db.or_(*conditions))
                elif field == 'subsector':
                    conditions = []
                    primary_only = filter_obj.get('primary_only', False)

                    for value in values:
                        if primary_only:
                            conditions.append(IndicatorBank.sub_sector.cast(db.Text).like(f'%"primary": "{value}"%'))
                        else:
                            conditions.append(IndicatorBank.sub_sector.cast(db.Text).like(f'%"primary": "{value}"%'))
                            conditions.append(IndicatorBank.sub_sector.cast(db.Text).like(f'%"secondary": "{value}"%'))
                            conditions.append(IndicatorBank.sub_sector.cast(db.Text).like(f'%"tertiary": "{value}"%'))
                    if conditions:
                        query = query.filter(db.or_(*conditions))
            except Exception as filter_error:
                current_app.logger.error(f"Error applying filter {field}: {str(filter_error)}")
                continue

        query = query.order_by(IndicatorBank.name)

        try:
            count = query.count()
            result = {
                'count': count,
                'success': True
            }
        except Exception as count_error:
            current_app.logger.error(f"Error executing count query: {str(count_error)}")
            result = {
                'count': 0,
                'success': True
            }

        if include_indicators:
            try:
                indicators = query.all()
                result['indicators'] = [{
                    'id': indicator.id,
                    'name': indicator.name,
                    'type': indicator.type,
                    'unit': indicator.unit,
                    'fdrs_kpi_code': getattr(indicator, 'fdrs_kpi_code', None),
                    'definition': indicator.definition,
                    'related_programs': indicator.related_programs_list,
                    'emergency': indicator.emergency,
                    'archived': indicator.archived,
                    'sector': indicator.sector,
                    'sub_sector': indicator.sub_sector
                } for indicator in indicators]
            except Exception as indicators_error:
                current_app.logger.error(f"Error retrieving indicators: {str(indicators_error)}")
                result['indicators'] = []

        return json_ok(**result)

    except Exception as e:
        with suppress(Exception):
            request_transaction_rollback()

        current_app.logger.error(f"Error in get_filtered_indicator_count: {str(e)}", exc_info=True)
        return json_server_error(GENERIC_ERROR_MESSAGE)


# === Common Words Management Routes ===
@bp.route("/common_words", methods=["GET"])
@permission_required('admin.indicator_bank.edit')
def manage_common_words():
    """Manage common words used in indicators."""
    return redirect(url_for('system_admin.manage_indicator_bank', tab='common_words', **request.args))

@bp.route("/common_words/add", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def add_common_word_modal():
    """Add a new common word via modal."""
    form = CommonWordForm(request.form)

    if form.validate_on_submit():
        try:
            new_common_word = CommonWord(
                created_by_user_id=current_user.id
            )

            form.populate_common_word(new_common_word)

            db.session.add(new_common_word)
            db.session.flush()

            return json_ok(
                success=True,
                message=f"Common word '{new_common_word.term}' added successfully.",
                common_word={
                    'id': new_common_word.id,
                    'term': new_common_word.term,
                    'meaning': new_common_word.meaning,
                    'is_active': new_common_word.is_active,
                    'created_at': new_common_word.created_at.strftime('%Y-%m-%d') if new_common_word.created_at else None
                }
            )

        except Exception as e:
            request_transaction_rollback()
            error_msg = GENERIC_ERROR_MESSAGE
            current_app.logger.error(f"Error adding common word: {e}", exc_info=True)

            return json_bad_request(error_msg, success=False)
    else:
        return json_form_errors(form, "Please correct the errors in the form.")

@bp.route("/common_words/edit/<int:id>", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def edit_common_word_modal(id):
    """Edit an existing common word via modal."""
    common_word = CommonWord.query.get_or_404(id)
    form = CommonWordForm(request.form)

    if form.validate_on_submit():
        try:
            form.populate_common_word(common_word)
            db.session.flush()

            return json_ok(
                success=True,
                message=f"Common word '{common_word.term}' updated successfully.",
                common_word={
                    'id': common_word.id,
                    'term': common_word.term,
                    'meaning': common_word.meaning,
                    'is_active': common_word.is_active,
                    'updated_at': common_word.updated_at.strftime('%Y-%m-%d') if common_word.updated_at else None
                }
            )

        except Exception as e:
            request_transaction_rollback()
            current_app.logger.error(f"Error updating common word {id}: {e}", exc_info=True)
            return json_bad_request("Error updating common word.", success=False)
    else:
        return json_form_errors(form, "Please correct the errors in the form.")


@bp.route("/common_words/delete/<int:id>", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def delete_common_word(id):
    """Delete a common word."""
    common_word = CommonWord.query.get_or_404(id)

    try:
        term_name = common_word.term
        db.session.delete(common_word)
        db.session.flush()

        flash(f"Common word '{term_name}' deleted successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("An error occurred. Please try again.", "danger")
        current_app.logger.error(f"Error deleting common word {id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.manage_common_words"))

@bp.route("/common_words/export", methods=["GET"])
@permission_required('admin.indicator_bank.view')
def export_common_words():
    """Export common words to Excel file."""
    try:
        common_words = CommonWord.query.filter_by(is_active=True).all()

        translatable = current_app.config.get("TRANSLATABLE_LANGUAGES") or []
        display_names = getattr(Config, "ALL_LANGUAGES_DISPLAY_NAMES", {}) or {}

        data = []
        for word in common_words:
            row = {
                'Term': word.term,
                'Meaning': word.meaning,
            }
            for code in translatable:
                header = display_names.get(code, code.upper())
                row[header] = word.get_meaning_translation(code) if word.meaning_translations else ''
            row.update({
                'Active': 'TRUE' if word.is_active else 'FALSE',
                'Created': word.created_at.strftime('%Y-%m-%d') if word.created_at else '',
                'Updated': word.updated_at.strftime('%Y-%m-%d') if word.updated_at else ''
            })
            data.append(row)

        df = pd.DataFrame(data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Common Words', index=False)

            worksheet = writer.sheets['Common Words']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    with suppress(Exception):
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)

        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'common_words_export_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        current_app.logger.exception("Error exporting common words: %s", e)
        flash('Error exporting common words.', 'danger')
        return redirect(url_for('system_admin.manage_common_words'))

@bp.route("/common_words/import", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def import_common_words():
    """Import common words from Excel file."""
    try:
        if 'excel_file' not in request.files:
            flash('No file selected', 'danger')
            return redirect(url_for('system_admin.manage_common_words'))

        file = request.files['excel_file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(url_for('system_admin.manage_common_words'))

        valid, error_msg, ext = validate_upload_extension_and_mime(file, EXCEL_EXTENSIONS)
        if not valid:
            flash(error_msg or 'Invalid file format. Please upload an Excel file (.xlsx or .xls).', 'danger')
            return redirect(url_for('system_admin.manage_common_words'))

        df = pd.read_excel(file, engine='openpyxl')

        required_columns = ['Term', 'Meaning']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            flash(f'Missing required columns: {", ".join(missing_columns)}', 'danger')
            return redirect(url_for('system_admin.manage_common_words'))

        overwrite = request.form.get('overwrite_existing') == 'on'
        imported_count = 0
        updated_count = 0
        errors = []

        for index, row in df.iterrows():
            try:
                term = str(row['Term']).strip()
                meaning = str(row['Meaning']).strip()

                if not term or not meaning:
                    continue

                existing_word = CommonWord.query.filter_by(term=term).first()

                if existing_word and not overwrite:
                    errors.append(f'Term "{term}" already exists (row {index + 2})')
                    continue

                translations = {}
                translatable = current_app.config.get("TRANSLATABLE_LANGUAGES") or []
                display_names = getattr(Config, "ALL_LANGUAGES_DISPLAY_NAMES", {}) or {}
                candidates = {}
                for code in translatable:
                    candidates[code] = {
                        code,
                        code.upper(),
                        display_names.get(code, '').strip(),
                    }
                for lang_code, possible_cols in candidates.items():
                    for col in possible_cols:
                        if not col:
                            continue
                        if col in df.columns and pd.notna(row[col]):
                            translation = str(row[col]).strip()
                            if translation:
                                translations[lang_code] = translation
                            break

                is_active = True
                if 'Active' in df.columns and pd.notna(row['Active']):
                    active_value = str(row['Active']).strip().upper()
                    is_active = active_value in ['TRUE', 'YES', '1', 'ACTIVE']

                if existing_word and overwrite:
                    existing_word.meaning = meaning
                    existing_word.meaning_translations = translations
                    existing_word.is_active = is_active
                    existing_word.updated_at = utcnow()
                    updated_count += 1
                else:
                    new_word = CommonWord(
                        term=term,
                        meaning=meaning,
                        meaning_translations=translations,
                        is_active=is_active,
                        created_by_user_id=current_user.id
                    )
                    db.session.add(new_word)
                    imported_count += 1

            except Exception as e:
                current_app.logger.warning("Error processing common words row %d: %s", index + 2, e, exc_info=True)
                errors.append(f'Error processing row {index + 2}.')
                continue

        db.session.flush()

        if imported_count > 0 or updated_count > 0:
            success_msg = f'Successfully imported {imported_count} new common words'
            if updated_count > 0:
                success_msg += f' and updated {updated_count} existing ones'
            flash(success_msg, 'success')

        if errors:
            error_msg = f'Import completed with {len(errors)} errors: ' + '; '.join(errors[:5])
            if len(errors) > 5:
                error_msg += f'... and {len(errors) - 5} more'
            flash(error_msg, 'warning')

        return redirect(url_for('system_admin.manage_common_words'))

    except Exception as e:
        current_app.logger.exception("Error importing common words: %s", e)
        flash('Error importing common words.', 'danger')
        return redirect(url_for('system_admin.manage_common_words'))

@bp.route("/common_words/template", methods=["GET"])
@permission_required('admin.indicator_bank.view')
def download_common_words_template():
    """Download Excel template for common words import."""
    try:
        translatable = current_app.config.get("TRANSLATABLE_LANGUAGES") or []
        display_names = getattr(Config, "ALL_LANGUAGES_DISPLAY_NAMES", {}) or {}

        base_rows = [
            {
                'Term': 'Emergency',
                'Meaning': 'A serious, unexpected, and often dangerous situation requiring immediate action.',
                'Active': 'TRUE'
            },
            {
                'Term': 'Response',
                'Meaning': 'An answer or reply, especially in the context of humanitarian aid or disaster relief.',
                'Active': 'TRUE'
            }
        ]
        sample_data = []
        for r in base_rows:
            row = dict(r)
            for code in translatable:
                header = display_names.get(code, code.upper())
                row.setdefault(header, '')
            sample_data.append(row)

        df = pd.DataFrame(sample_data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Common Words Template', index=False)

            worksheet = writer.sheets['Common Words Template']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    with suppress(Exception):
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='common_words_template.xlsx'
        )

    except Exception as e:
        current_app.logger.exception("Error downloading template: %s", e)
        flash("An error occurred. Please try again.", "danger")
        return redirect(url_for('system_admin.manage_common_words'))
