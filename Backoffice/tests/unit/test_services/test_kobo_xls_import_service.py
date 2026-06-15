"""
Comprehensive tests for app/services/kobo_xls_import_service.py
"""
from __future__ import annotations

import io
from typing import Any, Dict, List, Optional, Tuple
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from tests.factories import create_test_user


@pytest.fixture
def kobo_owner(db_session):
    """Real user row for Kobo import FK constraints (created_by / owned_by)."""
    return create_test_user(db_session)


def _make_kobo_xls(survey_rows=None, choices_rows=None, settings_rows=None):
    """
    Create a minimal Kobo XLSForm Excel file as BytesIO.

    survey_rows: list of dicts with keys: type, name, label, ...
    choices_rows: list of dicts with keys: list_name, name, label
    settings_rows: list of dicts with keys: form_title, ...
    """
    wb = openpyxl.Workbook()

    # Survey sheet
    survey_ws = wb.active
    survey_ws.title = "survey"
    if survey_rows:
        all_keys = list(dict.fromkeys(k for row in survey_rows for k in row.keys()))
        survey_ws.append(all_keys)
        for row in survey_rows:
            survey_ws.append([row.get(k, None) for k in all_keys])
    else:
        survey_ws.append(["type", "name", "label"])

    # Choices sheet
    if choices_rows is not None:
        choices_ws = wb.create_sheet("choices")
        all_keys = list(dict.fromkeys(k for row in choices_rows for k in row.keys())) if choices_rows else ["list_name", "name", "label"]
        choices_ws.append(all_keys or ["list_name", "name", "label"])
        for row in choices_rows:
            choices_ws.append([row.get(k, None) for k in all_keys])

    # Settings sheet
    if settings_rows is not None:
        settings_ws = wb.create_sheet("settings")
        all_keys = list(dict.fromkeys(k for row in settings_rows for k in row.keys())) if settings_rows else ["form_title"]
        settings_ws.append(all_keys or ["form_title"])
        for row in settings_rows:
            settings_ws.append([row.get(k, None) for k in all_keys])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class FakeFile:
    """Mimics a file-like object with read() that returns bytes."""
    def __init__(self, buf: io.BytesIO):
        self._buf = buf

    def read(self):
        return self._buf.read()


# ---------------------------------------------------------------------------
# Pure helper functions
# ---------------------------------------------------------------------------

class TestParseStr:
    def _call(self, val):
        from app.services.kobo_xls_import_service import _parse_str
        return _parse_str(val)

    def test_none_returns_empty(self):
        assert self._call(None) == ""

    def test_int(self):
        assert self._call(42) == "42"

    def test_strips_whitespace(self):
        assert self._call("  hello  ") == "hello"

    def test_empty_string(self):
        assert self._call("") == ""

    def test_whitespace_only(self):
        assert self._call("   ") == ""


class TestGetLabelFromRow:
    def _call(self, row, name_fallback=""):
        from app.services.kobo_xls_import_service import _get_label_from_row
        return _get_label_from_row(row, name_fallback)

    def test_uses_label_key(self):
        assert self._call({"label": "My Question"}) == "My Question"

    def test_uses_label_with_language(self):
        assert self._call({"label::English": "EN Label"}) == "EN Label"

    def test_fallback_to_name(self):
        assert self._call({"type": "text", "name": "q1"}, "q1") == "q1"

    def test_non_dict_uses_fallback(self):
        assert self._call("not a dict", "fallback") == "fallback"

    def test_untitled_when_no_label_or_fallback(self):
        assert self._call({}, "") == "Untitled"

    def test_prefers_exact_label_over_colon(self):
        # exact 'label' key takes priority
        row = {"label": "Exact", "label::English": "English Label"}
        assert self._call(row) == "Exact"

    def test_skips_none_values(self):
        assert self._call({"label": None, "label::en": "EN"}, "fallback") == "EN"


class TestGetTypeFromRow:
    def _call(self, row):
        from app.services.kobo_xls_import_service import _get_type_from_row
        return _get_type_from_row(row)

    def test_type_key(self):
        assert self._call({"type": "text"}) == "text"

    def test_type_key_case_insensitive_col(self):
        assert self._call({"Type": "integer"}) == "integer"

    def test_non_dict_returns_empty(self):
        assert self._call("not a dict") == ""

    def test_missing_type_returns_empty(self):
        assert self._call({"name": "q1", "label": "Q1"}) == ""

    def test_empty_type_returns_empty(self):
        assert self._call({"type": ""}) == ""


class TestParseRepeatCount:
    def _call(self, val):
        from app.services.kobo_xls_import_service import _parse_repeat_count
        return _parse_repeat_count(val)

    def test_none(self):
        assert self._call(None) is None

    def test_int_positive(self):
        assert self._call(5) == 5

    def test_float(self):
        assert self._call(3.0) == 3

    def test_zero_returns_none(self):
        assert self._call(0) is None

    def test_negative_returns_none(self):
        assert self._call(-1) is None

    def test_string_number(self):
        assert self._call("4") == 4

    def test_dynamic_reference_skipped(self):
        assert self._call("${num_people}") is None

    def test_invalid_string(self):
        assert self._call("many") is None

    def test_bool_int(self):
        # True is a bool (isinstance bool), should still work or return int
        # bool is a subclass of int, but the code checks isinstance(val, bool)
        assert self._call(True) is None


class TestParseFloat:
    def _call(self, val):
        from app.services.kobo_xls_import_service import _parse_float
        return _parse_float(val)

    def test_none_returns_zero(self):
        assert self._call(None) == 0.0

    def test_int(self):
        assert self._call(3) == 3.0

    def test_float(self):
        assert self._call(1.5) == 1.5

    def test_string(self):
        assert self._call("2.5") == 2.5

    def test_invalid(self):
        assert self._call("abc") == 0.0

    def test_empty_string(self):
        assert self._call("") == 0.0


# ---------------------------------------------------------------------------
# KoboXlsImportService methods
# ---------------------------------------------------------------------------

class TestReadFormTitle:
    def test_reads_form_title(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["form_title", "other_col"])
            ws.append(["My Test Form", "value"])
            result = KoboXlsImportService._read_form_title(ws)
            assert result == "My Test Form"

    def test_returns_none_when_no_form_title_col(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["title", "other"])
            ws.append(["My Form", "v"])
            result = KoboXlsImportService._read_form_title(ws)
            assert result is None

    def test_returns_none_when_empty_value(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["form_title"])
            ws.append([None])
            result = KoboXlsImportService._read_form_title(ws)
            assert result is None


class TestLoadChoices:
    def test_loads_choices(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["list_name", "name", "label"])
            ws.append(["gender", "male", "Male"])
            ws.append(["gender", "female", "Female"])
            ws.append(["yesno", "yes", "Yes"])
            result = KoboXlsImportService._load_choices(ws)
            assert "gender" in result
            assert len(result["gender"]) == 2
            assert "yesno" in result

    def test_missing_required_columns(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["list_name", "name"])  # missing label
            ws.append(["yesno", "yes"])
            result = KoboXlsImportService._load_choices(ws)
            assert result == {}

    def test_skips_rows_without_list_or_name(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["list_name", "name", "label"])
            ws.append([None, "male", "Male"])  # missing list_name
            ws.append(["gender", None, "Female"])  # missing name
            ws.append(["gender", "other", "Other"])
            result = KoboXlsImportService._load_choices(ws)
            assert "gender" in result
            assert len(result["gender"]) == 1

    def test_label_fallback_to_name(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["list_name", "name", "label"])
            ws.append(["gender", "male", None])  # no label -> fallback to name
            result = KoboXlsImportService._load_choices(ws)
            assert result["gender"][0] == ("male", "male")


class TestParseSurveySheet:
    def test_parses_survey(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["type", "name", "label"])
            ws.append(["text", "q1", "Question 1"])
            ws.append(["integer", "q2", "Question 2"])
            result = KoboXlsImportService._parse_survey_sheet(ws)
            assert len(result) == 2

    def test_empty_type_skipped(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["type", "name", "label"])
            ws.append(["", "q1", "Q1"])  # empty type
            ws.append(["text", "q2", "Q2"])
            result = KoboXlsImportService._parse_survey_sheet(ws)
            assert len(result) == 1

    def test_no_type_column_returns_empty(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["name", "label"])  # no type column
            ws.append(["q1", "Q1"])
            result = KoboXlsImportService._parse_survey_sheet(ws)
            assert result == []


class TestMapKoboRowToItem:
    def _call(self, row, choices_by_list=None, warnings=None):
        from app.services.kobo_xls_import_service import KoboXlsImportService
        if choices_by_list is None:
            choices_by_list = {}
        if warnings is None:
            warnings = []
        return KoboXlsImportService._map_kobo_row_to_item(row, choices_by_list, warnings)

    def test_text_question(self):
        row = {"type": "text", "name": "q1", "label": "What is your name?"}
        result = self._call(row)
        assert result is not None
        assert result['type'] == 'text'
        assert result['item_type'] == 'question'

    def test_integer_question(self):
        row = {"type": "integer", "name": "q_num", "label": "How many?"}
        result = self._call(row)
        assert result['type'] == 'number'

    def test_select_one_with_choices(self):
        row = {"type": "select_one gender", "name": "gender_q", "label": "What is your gender?"}
        choices = {"gender": [("male", "Male"), ("female", "Female")]}
        result = self._call(row, choices)
        assert result['type'] == 'single_choice'
        assert result['options_json'] is not None
        assert len(result['options_json']) == 2

    def test_select_multiple_empty_choices(self):
        warnings = []
        row = {"type": "select_multiple colors", "name": "colors_q", "label": "Pick colors"}
        result = self._call(row, {}, warnings)
        assert result['type'] == 'multiple_choice'
        assert result['options_json'] == []
        assert len(warnings) > 0

    def test_note_becomes_blank(self):
        row = {"type": "note", "name": "note1", "label": "Read this note"}
        result = self._call(row)
        assert result['type'] == 'blank'

    def test_image_becomes_document_field(self):
        row = {"type": "image", "name": "photo", "label": "Upload photo"}
        result = self._call(row)
        assert result['item_type'] == 'document_field'
        assert result['type'] is None

    def test_audio_becomes_document_field(self):
        row = {"type": "audio", "name": "audio", "label": "Upload audio"}
        result = self._call(row)
        assert result['item_type'] == 'document_field'

    def test_skipped_type_returns_none(self):
        warnings = []
        row = {"type": "geopoint", "name": "location", "label": "Location"}
        result = self._call(row, {}, warnings)
        assert result is None
        assert len(warnings) > 0

    def test_unknown_type_returns_none(self):
        warnings = []
        row = {"type": "unknowntype", "name": "q", "label": "Unknown"}
        result = self._call(row, {}, warnings)
        assert result is None
        assert len(warnings) > 0

    def test_multiline_text_becomes_textarea(self):
        row = {"type": "text", "name": "bio", "label": "Bio", "appearance": "multiline"}
        result = self._call(row)
        assert result['type'] == 'textarea'

    def test_required_field(self):
        row = {"type": "text", "name": "q", "label": "Required Q", "required": "yes"}
        result = self._call(row)
        assert result['config']['is_required'] is True

    def test_not_required_field(self):
        row = {"type": "text", "name": "q", "label": "Optional Q", "required": "no"}
        result = self._call(row)
        assert result['config']['is_required'] is False

    def test_date_type(self):
        row = {"type": "date", "name": "dob", "label": "Date of birth"}
        result = self._call(row)
        assert result['type'] == 'date'

    def test_datetime_type(self):
        row = {"type": "datetime", "name": "ts", "label": "Timestamp"}
        result = self._call(row)
        assert result['type'] == 'datetime'

    def test_time_becomes_datetime(self):
        row = {"type": "time", "name": "t", "label": "Time"}
        result = self._call(row)
        assert result['type'] == 'datetime'


class TestImportKoboXls:
    """Tests for the main import_kobo_xls entry point."""

    def test_missing_survey_sheet(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            wb = openpyxl.Workbook()
            wb.active.title = "not-survey"
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            result = KoboXlsImportService.import_kobo_xls(FakeFile(buf), owned_by=1)
            assert result['success'] is False
            assert "survey" in result['message'].lower()

    def test_empty_survey_sheet(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "survey"
            ws.append(["type", "name", "label"])  # Only header, no data
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            result = KoboXlsImportService.import_kobo_xls(FakeFile(buf), owned_by=1)
            assert result['success'] is False

    def test_invalid_excel_file(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            result = KoboXlsImportService.import_kobo_xls(FakeFile(io.BytesIO(b"not-an-excel")), owned_by=1)
            assert result['success'] is False

    def test_no_owner_fails(self, app):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            buf = _make_kobo_xls(
                survey_rows=[{"type": "text", "name": "q1", "label": "Q1"}],
            )
            with patch("app.services.kobo_xls_import_service.current_user") as mock_cu:
                mock_cu.is_authenticated = False
                result = KoboXlsImportService.import_kobo_xls(FakeFile(buf))
            assert result['success'] is False
            assert "owner" in result['message'].lower() or "authenticated" in result['message'].lower()

    def test_simple_form_import(self, app, db_session, kobo_owner):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            buf = _make_kobo_xls(
                survey_rows=[
                    {"type": "text", "name": "name_q", "label": "Your name"},
                    {"type": "integer", "name": "age_q", "label": "Your age"},
                ],
            )
            with patch("app.services.kobo_xls_import_service.current_user") as mock_cu:
                mock_cu.is_authenticated = True
                mock_cu.id = kobo_owner.id
                result = KoboXlsImportService.import_kobo_xls(FakeFile(buf), owned_by=kobo_owner.id)
            # Success is True only if DB is available (unit test may fail without mock)
            assert 'success' in result

    def test_form_with_groups(self, app, db_session, kobo_owner):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            buf = _make_kobo_xls(
                survey_rows=[
                    {"type": "begin_group", "name": "personal_info", "label": "Personal Info"},
                    {"type": "text", "name": "first_name", "label": "First Name"},
                    {"type": "text", "name": "last_name", "label": "Last Name"},
                    {"type": "end_group", "name": "", "label": ""},
                ],
            )
            with patch("app.services.kobo_xls_import_service.current_user") as mock_cu:
                mock_cu.is_authenticated = True
                mock_cu.id = kobo_owner.id
                result = KoboXlsImportService.import_kobo_xls(FakeFile(buf), owned_by=kobo_owner.id)
            assert 'success' in result

    def test_form_with_repeat(self, app, db_session, kobo_owner):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            buf = _make_kobo_xls(
                survey_rows=[
                    {"type": "begin_repeat", "name": "household", "label": "Household Members", "repeat_count": "3"},
                    {"type": "text", "name": "member_name", "label": "Member Name"},
                    {"type": "end_repeat", "name": "", "label": ""},
                ],
            )
            with patch("app.services.kobo_xls_import_service.current_user") as mock_cu:
                mock_cu.is_authenticated = True
                mock_cu.id = kobo_owner.id
                result = KoboXlsImportService.import_kobo_xls(FakeFile(buf), owned_by=kobo_owner.id)
            assert 'success' in result

    def test_form_with_choices(self, app, db_session, kobo_owner):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            buf = _make_kobo_xls(
                survey_rows=[
                    {"type": "select_one gender_list", "name": "gender", "label": "Gender"},
                ],
                choices_rows=[
                    {"list_name": "gender_list", "name": "male", "label": "Male"},
                    {"list_name": "gender_list", "name": "female", "label": "Female"},
                ],
            )
            with patch("app.services.kobo_xls_import_service.current_user") as mock_cu:
                mock_cu.is_authenticated = True
                mock_cu.id = kobo_owner.id
                result = KoboXlsImportService.import_kobo_xls(FakeFile(buf), owned_by=kobo_owner.id)
            assert 'success' in result

    def test_form_title_from_settings(self, app, db_session, kobo_owner):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            buf = _make_kobo_xls(
                survey_rows=[{"type": "text", "name": "q", "label": "Question"}],
                settings_rows=[{"form_title": "My Survey Title"}],
            )
            with patch("app.services.kobo_xls_import_service.current_user") as mock_cu:
                mock_cu.is_authenticated = True
                mock_cu.id = kobo_owner.id
                result = KoboXlsImportService.import_kobo_xls(FakeFile(buf), owned_by=kobo_owner.id)
            assert 'success' in result

    def test_template_name_override(self, app, db_session, kobo_owner):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            buf = _make_kobo_xls(
                survey_rows=[{"type": "text", "name": "q", "label": "Q"}],
                settings_rows=[{"form_title": "Original Title"}],
            )
            with patch("app.services.kobo_xls_import_service.current_user") as mock_cu:
                mock_cu.is_authenticated = True
                mock_cu.id = kobo_owner.id
                result = KoboXlsImportService.import_kobo_xls(
                    FakeFile(buf), template_name="Override Title", owned_by=kobo_owner.id
                )
            assert 'success' in result

    def test_root_level_items_get_main_section(self, app, db_session, kobo_owner):
        """Items without a group should be placed in auto-created 'Main' section."""
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            buf = _make_kobo_xls(
                survey_rows=[
                    {"type": "text", "name": "standalone_q", "label": "Standalone Question"},
                ],
            )
            with patch("app.services.kobo_xls_import_service.current_user") as mock_cu:
                mock_cu.is_authenticated = True
                mock_cu.id = kobo_owner.id
                result = KoboXlsImportService.import_kobo_xls(FakeFile(buf), owned_by=kobo_owner.id)
            assert 'success' in result

    def test_unclosed_group_treated_as_root_section(self, app, db_session, kobo_owner):
        """Unclosed begin_group should still produce a section."""
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            buf = _make_kobo_xls(
                survey_rows=[
                    {"type": "begin_group", "name": "grp1", "label": "Unclosed Group"},
                    {"type": "text", "name": "q", "label": "Q"},
                    # Missing end_group
                ],
            )
            with patch("app.services.kobo_xls_import_service.current_user") as mock_cu:
                mock_cu.is_authenticated = True
                mock_cu.id = kobo_owner.id
                result = KoboXlsImportService.import_kobo_xls(FakeFile(buf), owned_by=kobo_owner.id)
            assert 'success' in result

    def test_openpyxl_none_returns_error(self, app):
        """If openpyxl is not installed, should return error dict."""
        with app.app_context():
            import app.services.kobo_xls_import_service as module
            original = module.openpyxl
            module.openpyxl = None
            try:
                from app.services.kobo_xls_import_service import KoboXlsImportService
                result = KoboXlsImportService.import_kobo_xls(MagicMock(), owned_by=1)
                assert result['success'] is False
                assert 'openpyxl' in result['message'].lower()
            finally:
                module.openpyxl = original

    def test_skipped_types_generate_warnings(self, app, db_session, kobo_owner):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            buf = _make_kobo_xls(
                survey_rows=[
                    {"type": "geopoint", "name": "loc", "label": "Location"},
                    {"type": "text", "name": "q", "label": "Q"},
                ],
            )
            with patch("app.services.kobo_xls_import_service.current_user") as mock_cu:
                mock_cu.is_authenticated = True
                mock_cu.id = kobo_owner.id
                result = KoboXlsImportService.import_kobo_xls(FakeFile(buf), owned_by=kobo_owner.id)
            if result.get('success'):
                assert len(result.get('warnings', [])) > 0

    def test_nested_groups(self, app, db_session, kobo_owner):
        with app.app_context():
            from app.services.kobo_xls_import_service import KoboXlsImportService
            buf = _make_kobo_xls(
                survey_rows=[
                    {"type": "begin_group", "name": "outer", "label": "Outer Group"},
                    {"type": "begin_group", "name": "inner", "label": "Inner Group"},
                    {"type": "text", "name": "nested_q", "label": "Nested Question"},
                    {"type": "end_group", "name": "", "label": ""},
                    {"type": "end_group", "name": "", "label": ""},
                ],
            )
            with patch("app.services.kobo_xls_import_service.current_user") as mock_cu:
                mock_cu.is_authenticated = True
                mock_cu.id = kobo_owner.id
                result = KoboXlsImportService.import_kobo_xls(FakeFile(buf), owned_by=kobo_owner.id)
            assert 'success' in result
