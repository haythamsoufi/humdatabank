#!/usr/bin/env python3
"""
Per-country Unified Country Plan Excel template round-trip for Template 24 (planning).

Uses the structured IFRC planning workbook (named cells + Excel tables) for export/import
from a single country assignment (aes_id). Workbook: app/static/templates/unified_country_plan.xlsx
"""

from __future__ import annotations

import json
import os
import re
import sys
import zipfile
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

script_dir = os.path.dirname(os.path.abspath(__file__))
backoffice_dir = os.path.dirname(os.path.dirname(script_dir))
if backoffice_dir not in sys.path:
    sys.path.insert(0, backoffice_dir)
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

if "FLASK_CONFIG" not in os.environ:
    os.environ["FLASK_CONFIG"] = "development"

from import_fdrs_form_data import upsert_form_data_rows  # noqa: E402
from import_upr_excel_data import (  # noqa: E402
    EMERGENCY_APPEALS_COLUMN,
    FUNDING_MATRIX_BY_YEAR_OFFSET,
    ITEM_BILATERAL_SUPPORT,
    ITEM_COMMENTS,
    ITEM_EMERGENCY_APPEALS,
    ITEM_LONGER_TERM_PROGRAMMES,
    PLANNING_EA_FUNDING_AREAS,
    build_import_context,
    parse_value_num,
    round_to_period,
    _ensure_funding_ea_col_header,
    _matrix_row,
    _resolve_emergency_matrix_cells,
    _resolve_ns_row_id,
    _scalar_row,
    _year_offset,
)
from upr_country_reporting_excel_template import (  # noqa: E402
    GENERIC_FORM_EXPORT_SHEETS,
    _bilateral_ns_name_for_row,
    _bilateral_table_row_info,
    _load_form_data_map,
    _matrix_cell_scalar,
    _matrix_cells,
    _normalize_matrix_cells,
    _normalize_workbook_header,
    _quiet_openpyxl_io,
    _scalar_value,
    _table_data_row_capacity,
    _write_bilateral_ns_source_cell,
    dedupe_upr_import_warnings,
    serialize_upr_import_warnings,
    read_named_cell,
    read_named_table,
    read_table_cell,
    write_named_cell,
    write_table_cell,
)

_PLANNING_SUPPORT_CELL_KEY_RE = re.compile(r"^(\d+)_(SP\d|EFs)$")

PLANNING_COUNTRY_TEMPLATE_ID = 24  # Unified Country Plan

PEOPLE_SHEET = "People to be reached"
PEOPLE_TABLE = "Data_People"
SUPPORT_SHEET = "Planned Bilateral Support"
SUPPORT_TABLE = "Data_Support"
SUPPORT_NS_REGION_COL = 2  # Column B — must be set before column C (NS dropdown depends on region).
NS_REGION_TABLE = "Table9"
FUNDING_SHEET = "Funding requirements"
FUNDING_TABLE = "Data_FR"
COMMENTS_SHEET = "Comments"
COMMENT_NAMED_CELL = "Comment"
FUNDING_NETWORK_MEMBER_COL = 2  # "National Society name" / IFRC network member (column B).
FUNDING_PNS_FIRST_ROW = 10
FUNDING_PNS_LAST_ROW = 34
FUNDING_HNS_ROW = FUNDING_PNS_FIRST_ROW - 2
FUNDING_IFRC_ROW = FUNDING_PNS_FIRST_ROW - 1
FUNDING_PNS_ROW_PLACEHOLDER_AREA = "SP2"
FUNDING_PNS_ARRAY_FORMULA_TEXT = (
    "IFERROR(_xlfn.UNIQUE(_xlfn._xlws.FILTER('Planned Bilateral Support'!$C$5:$C$34,"
    "'Planned Bilateral Support'!$C$5:$C$34<>\"\")),\"\")"
)

START_SHEET = "Start"
START_REGION_CELL = "C12"
COUNTRY_REGION_TABLE_SHEET = "TemplateData"
COUNTRY_REGION_TABLE = "Table7"

# Workbook Table8 labels differ from platform canonical names (e.g. MENA).
WORKBOOK_REGION_LABELS = {
    "MENA": "Middle East and North Africa",
}

SP_AREAS: Tuple[str, ...] = ("SP2", "SP1", "SP3", "SP4", "SP5")
FUNDING_AREAS_PER_YEAR: Tuple[str, ...] = ("EA1", "EA2", "EA3", *SP_AREAS, "EFs")

NS_DATA_BANK_IDS: Dict[str, int] = {
    "volunteers": 724,
    "staff": 727,
    "local_units": 723,
    "branches": 1117,
}

NS_DATA_NAMED_CELLS: Dict[str, str] = {
    "volunteers": "Data_vol",
    "staff": "Data_staff",
    "local_units": "Data_units",
    "branches": "Data_branches",
}

UNIFIED_COUNTRY_PLAN_REQUIRED_NAMED_RANGES: Tuple[str, ...] = (
    "Version",
    "Data_Country",
    COMMENT_NAMED_CELL,
    *NS_DATA_NAMED_CELLS.values(),
)

UNIFIED_COUNTRY_PLAN_REQUIRED_TABLES: Tuple[Tuple[str, str], ...] = (
    (PEOPLE_SHEET, PEOPLE_TABLE),
    (SUPPORT_SHEET, SUPPORT_TABLE),
    (FUNDING_SHEET, FUNDING_TABLE),
)

UNIFIED_COUNTRY_PLAN_COMPATIBLE_ROUND_PREFIXES: Tuple[str, ...] = ("P",)

EA_SLOT_NAMED_CELLS: Tuple[Tuple[str, str, str, str], ...] = (
    ("Reach_EA1", "Data_MDR1", "Data_EA1", "EA1"),
    ("Reach_EA2", "Data_MDR2", "Data_EA2", "EA2"),
    ("Reach_EA3", "Data_MDR3", "Data_EA3", "EA3"),
)

_FUNDING_HEADER_RE = re.compile(r"^(.+)_(\d{4})$")


def _require_openpyxl():
    import openpyxl  # noqa: F401


def period_to_workbook_version(period_name: str) -> str:
    """Map assignment period name to planning workbook Version cell (e.g. 2027 -> P27.V1.0)."""
    period = (period_name or "").strip()
    match = re.match(r"^(\d{4})$", period)
    if match:
        year = int(match.group(1))
        return f"P{year - 2000}.V1.0"
    return "P27.V1.0"


def planning_base_year(period_name: str) -> Optional[int]:
    period = (period_name or "").strip()
    match = re.match(r"^(\d{4})$", period)
    if match:
        return int(match.group(1))
    return None


def planning_year_triplet(period_name: str) -> Tuple[int, int, int]:
    base = planning_base_year(period_name) or 2027
    return base, base + 1, base + 2


def funding_column_header(area: str, year: int) -> str:
    return f"{area}_{year}"


def parse_funding_column_header(header: str) -> Optional[Tuple[str, int]]:
    text = str(header or "").strip()
    if not text or text.upper() == "NS":
        return None
    match = _FUNDING_HEADER_RE.match(text)
    if not match:
        return None
    area, year_text = match.group(1), match.group(2)
    return area, int(year_text)


def parse_version(wb) -> Tuple[str, str]:
    raw = read_named_cell(wb, "Version")
    text = str(raw or "").strip()
    if not text:
        return "", ""
    round_code = text.split(".")[0].upper()
    period = round_to_period(round_code) or ""
    return round_code, period


def _workbook_table_exists(wb, sheet_name: str, table_name: str) -> bool:
    if sheet_name not in wb.sheetnames:
        return False
    return table_name in wb[sheet_name].tables


def _rewrite_table_header(wb, sheet_name: str, table_name: str, header: str, new_value: str) -> None:
    from openpyxl.utils import range_boundaries

    ws = wb[sheet_name]
    tbl = ws.tables[table_name]
    min_col, min_row, max_col, _max_row = range_boundaries(tbl.ref)
    target = _normalize_workbook_header(header)
    for col in range(min_col, max_col + 1):
        raw = ws.cell(min_row, col).value
        h = str(raw).strip() if raw is not None else ""
        if h == header or _normalize_workbook_header(h) == target:
            ws.cell(min_row, col).value = new_value
            # Keep Excel table metadata in sync with the header row (prevents table*.xml repair errors).
            col_idx = col - min_col
            if tbl.tableColumns and 0 <= col_idx < len(tbl.tableColumns):
                tbl.tableColumns[col_idx].name = new_value
            return


def rewrite_planning_year_headers(wb, period_name: str) -> None:
    """Align Data_People year rows and Data_FR column headers with the assignment period."""
    y0, y1, y2 = planning_year_triplet(period_name)

    _, people_rows = read_named_table(wb, PEOPLE_SHEET, PEOPLE_TABLE)
    for offset, year in enumerate((y0, y1, y2)):
        if offset < len(people_rows):
            write_table_cell(wb, PEOPLE_SHEET, PEOPLE_TABLE, offset, "Year", year)

    headers, _ = read_named_table(wb, FUNDING_SHEET, FUNDING_TABLE)
    funding_headers = [h for h in headers if parse_funding_column_header(h)]
    # The workbook uses 9 EA+SP columns for year 1 but only 6 SP columns for years 2–3,
    # so idx // len(FUNDING_AREAS_PER_YEAR) must not be used to infer the target year.
    template_years = sorted({parse_funding_column_header(h)[1] for h in funding_headers})
    target_years = (y0, y1, y2)
    year_map = {
        template_years[i]: target_years[i]
        for i in range(min(len(template_years), len(target_years)))
    }
    for header in funding_headers:
        parsed = parse_funding_column_header(header)
        if not parsed:
            continue
        area, old_year = parsed
        new_year = year_map.get(old_year, old_year)
        new_header = funding_column_header(area, new_year)
        if new_header != header:
            _rewrite_table_header(wb, FUNDING_SHEET, FUNDING_TABLE, header, new_header)


def _workbook_region_label(region_name: str) -> str:
    """Map a platform/canonical region name to the Start-sheet dropdown label."""
    from app.services.organization.secretariat_regional_office_service import normalize_region_label

    canonical = normalize_region_label(region_name) or str(region_name or "").strip()
    if not canonical:
        return ""
    return WORKBOOK_REGION_LABELS.get(canonical, canonical)


def _workbook_region_for_country(wb, country_name: str, *, fallback_region: str = "") -> str:
    """Resolve the Start-sheet region dropdown value for a country."""
    name = str(country_name or "").strip()
    if name and _workbook_table_exists(wb, COUNTRY_REGION_TABLE_SHEET, COUNTRY_REGION_TABLE):
        try:
            _, rows = read_named_table(wb, COUNTRY_REGION_TABLE_SHEET, COUNTRY_REGION_TABLE)
        except ValueError:
            rows = []
        lower = name.lower()
        for row in rows:
            row_country = str(row.get("Country") or "").strip()
            if row_country.lower() == lower:
                region = str(row.get("Region") or "").strip()
                if region:
                    return region
    return _workbook_region_label(fallback_region)


def _write_start_sheet_selection(wb, country_name: str, region_name: str) -> None:
    """Pre-fill Start sheet region (C12) and country (Data_Country / K12)."""
    if START_SHEET not in wb.sheetnames:
        return
    ws = wb[START_SHEET]
    region = str(region_name or "").strip()
    country = str(country_name or "").strip()
    if region:
        ws[START_REGION_CELL].value = region
    if country:
        write_named_cell(wb, "Data_Country", country)


def _load_assignment_meta(aes_id: int):
    from app.models.assignments import AssignmentEntityStatus
    from app.utils.api_serialization import _country_for_aes

    aes = AssignmentEntityStatus.query.get(aes_id)
    if not aes:
        raise ValueError(f"Assignment {aes_id} not found")
    country = _country_for_aes(aes)
    country_name = (country.name if country else "") or ""
    iso3 = (country.iso3 if country else "") or ""
    region = (country.region if country else "") or ""
    period = (aes.assigned_form.period_name if aes.assigned_form else "") or ""
    template_id = int(getattr(aes.assigned_form, "template_id", 0) or 0)
    if template_id != PLANNING_COUNTRY_TEMPLATE_ID:
        raise ValueError(
            f"Assignment {aes_id} is template {template_id}, not T{PLANNING_COUNTRY_TEMPLATE_ID}"
        )
    return aes, country_name.strip(), iso3.strip().upper(), period.strip(), region.strip()


def validate_unified_country_plan_import_file(
    wb,
    *,
    expected_country: str = "",
    expected_period: str = "",
) -> Dict[str, Any]:
    """Validate workbook structure and assignment match before import."""
    errors: List[str] = []
    warnings: List[str] = []
    preview: Dict[str, Any] = {
        "country": None,
        "period": None,
        "round_code": None,
    }

    generic_hits = [name for name in GENERIC_FORM_EXPORT_SHEETS if name in wb.sheetnames]
    if generic_hits:
        errors.append(
            "This file looks like a generic form Excel export (sheets: "
            f"{', '.join(generic_hits)}), not a Unified Country Plan workbook. "
            "Use Export Unified Country Plan from this assignment."
        )

    for named_range in UNIFIED_COUNTRY_PLAN_REQUIRED_NAMED_RANGES:
        if named_range not in wb.defined_names:
            errors.append(
                f"Missing named range {named_range!r}. "
                "Download a current Unified Country Plan template from this assignment."
            )

    for sheet_name, table_name in UNIFIED_COUNTRY_PLAN_REQUIRED_TABLES:
        if not _workbook_table_exists(wb, sheet_name, table_name):
            if sheet_name not in wb.sheetnames:
                errors.append(f"Missing sheet {sheet_name!r}.")
            else:
                errors.append(
                    f"Missing table {table_name!r} on sheet {sheet_name!r}. "
                    "Older IFRC templates are not compatible with this import."
                )

    round_code, period_name = parse_version(wb)
    preview["round_code"] = round_code or None
    preview["period"] = period_name or None
    country_name = str(read_named_cell(wb, "Data_Country") or "").strip()
    preview["country"] = country_name or None

    if not round_code:
        errors.append(
            "Missing or unreadable Version cell. "
            "Use a Unified Country Plan template exported from this assignment."
        )
    elif not any(round_code.startswith(prefix) for prefix in UNIFIED_COUNTRY_PLAN_COMPATIBLE_ROUND_PREFIXES):
        errors.append(
            f"Unsupported workbook version {round_code!r}. "
            f"Expected a planning round code starting with "
            f"{' or '.join(UNIFIED_COUNTRY_PLAN_COMPATIBLE_ROUND_PREFIXES)}."
        )

    if expected_period and period_name and period_name != expected_period:
        warnings.append(
            f"Workbook period {period_name!r} differs from this assignment ({expected_period!r}). "
            "Values will be loaded into the current assignment."
        )

    if expected_country and country_name and country_name.lower() != expected_country.lower():
        errors.append(
            f"Workbook country {country_name!r} does not match this assignment ({expected_country!r})."
        )

    return {
        "valid": not errors,
        "message": errors[0] if errors else "Workbook is valid.",
        "errors": errors,
        "warnings": warnings,
        "preview": preview,
    }


def _cell_is_tick(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value) == 1.0
    text = str(value).strip().lower()
    return text in ("1", "x", "yes", "true")


def _parse_funding_row_entity(ns_label: Any) -> Tuple[str, str]:
    text = str(ns_label or "").replace("\r\n", "\n").strip()
    lower = text.lower()
    if lower == "ifrc secretariat":
        return "ifrc", "IFRC Secretariat"
    if "host national society" in lower:
        return "hns", "HNS"
    if "\n" in text:
        text = text.split("\n", 1)[-1].strip()
    return "pns", text


def _funding_network_member_label(wb, row_idx: int) -> str:
    from openpyxl.worksheet.formula import ArrayFormula

    val = wb[FUNDING_SHEET].cell(row_idx, FUNDING_NETWORK_MEMBER_COL).value
    if val is None or isinstance(val, ArrayFormula):
        return ""
    if isinstance(val, str):
        return val.replace("\r\n", "\n").strip()
    return str(val).strip()


def _funding_row_entity_from_workbook(wb, row: Dict[str, Any]) -> Tuple[str, str]:
    """Resolve HNS/IFRC/PNS for a funding row."""
    row_idx = int(row.get("_row") or 0)
    if row_idx >= FUNDING_PNS_FIRST_ROW:
        member_label = _funding_network_member_label(wb, row_idx)
        if member_label:
            return "pns", member_label
        names = _collect_bilateral_support_ns_names(wb)
        pns_idx = row_idx - FUNDING_PNS_FIRST_ROW
        if 0 <= pns_idx < len(names):
            return "pns", names[pns_idx]
    return _parse_funding_row_entity(row.get("NS"))


def _collect_bilateral_support_ns_names(wb) -> List[str]:
    names: List[str] = []
    capacity = _table_data_row_capacity(wb, SUPPORT_SHEET, SUPPORT_TABLE)
    for offset in range(capacity):
        ns_name = _bilateral_ns_name_for_row(wb, SUPPORT_SHEET, SUPPORT_TABLE, offset)
        if ns_name:
            names.append(ns_name)
    return names


def _funding_table_headers(wb) -> List[str]:
    from openpyxl.utils import range_boundaries

    ws = wb[FUNDING_SHEET]
    tbl = ws.tables[FUNDING_TABLE]
    ref = tbl.ref if hasattr(tbl, "ref") else tbl
    min_col, min_row, max_col, _max_row = range_boundaries(ref)
    headers: List[str] = []
    for col in range(min_col, max_col + 1):
        raw = ws.cell(min_row, col).value
        headers.append(str(raw).strip() if raw is not None else "")
    return headers


def _read_funding_table_row(wb, excel_row: int, headers: List[str]) -> Dict[str, Any]:
    """Read one funding table row by Excel row number (works with data_only workbooks)."""
    from openpyxl.utils import range_boundaries

    ws = wb[FUNDING_SHEET]
    tbl = ws.tables[FUNDING_TABLE]
    ref = tbl.ref if hasattr(tbl, "ref") else tbl
    min_col, min_row, max_col, _max_row = range_boundaries(ref)
    header_to_col: Dict[str, int] = {}
    for col in range(min_col, max_col + 1):
        raw = ws.cell(min_row, col).value
        if raw is not None and str(raw).strip():
            header_to_col[str(raw).strip()] = col
    record: Dict[str, Any] = {
        "_row": excel_row,
        "_sheet": FUNDING_SHEET,
        "_table": FUNDING_TABLE,
    }
    for header in headers:
        if not header:
            continue
        col = header_to_col.get(header)
        record[header] = ws.cell(excel_row, col).value if col else None
    return record


def _import_funding_cells_for_row(
    matrix_cells: Dict[Tuple[int, int], Dict[str, Any]],
    ctx,
    *,
    aes_id: int,
    iso3: str,
    period: str,
    rnd: str,
    row: Dict[str, Any],
    row_key: str,
    funding_headers: List[str],
    reach_ea_codes: Dict[Tuple[str, str, str], str],
    reach_ea_names: Dict[Tuple[str, str, str], str],
    warnings: List[str],
) -> None:
    for header in funding_headers:
        parsed = parse_funding_column_header(header)
        if not parsed:
            continue
        area, year = parsed
        offset = _year_offset(period, year)
        if offset is None:
            continue
        amount = parse_value_num(row.get(header))
        if amount is None:
            continue
        item_id = FUNDING_MATRIX_BY_YEAR_OFFSET.get(offset)
        if not item_id:
            continue
        if area in PLANNING_EA_FUNDING_AREAS:
            if not _ensure_funding_ea_col_header(
                matrix_cells,
                ctx,
                aes_id=aes_id,
                funding_item_id=item_id,
                iso3=iso3,
                rnd=rnd,
                area=area,
                ea_code_raw=reach_ea_codes.get((iso3, rnd, area)),
                reach_ea_codes=reach_ea_codes,
                excel_name_raw=reach_ea_names.get((iso3, rnd, area)),
                reach_ea_names=reach_ea_names,
            ):
                warnings.append(f"Could not resolve EA column header for {area} ({year}).")
                continue
        matrix_cells[(aes_id, item_id)][f"{row_key}_{area}"] = amount


def _ensure_funding_pns_rows_in_matrices(
    matrix_cells: Dict[Tuple[int, int], Dict[str, Any]],
    *,
    aes_id: int,
    ns_id: int,
) -> None:
    """Ensure hybrid funding matrices include a PNS row (mirrors bilateral-driven Excel rows)."""
    row_prefix = f"{ns_id}_"
    for item_id in FUNDING_MATRIX_BY_YEAR_OFFSET.values():
        cells = matrix_cells[(aes_id, item_id)]
        if any(k.startswith(row_prefix) for k in cells if not k.startswith("col_header|")):
            continue
        cells[f"{ns_id}_{FUNDING_PNS_ROW_PLACEHOLDER_AREA}"] = ""


def _funding_pns_array_formula_text(wb) -> str:
    from openpyxl.worksheet.formula import ArrayFormula

    val = wb[FUNDING_SHEET].cell(FUNDING_PNS_FIRST_ROW, FUNDING_NETWORK_MEMBER_COL).value
    if isinstance(val, ArrayFormula):
        return val.text
    return FUNDING_PNS_ARRAY_FORMULA_TEXT


def _refresh_funding_pns_array_formula(wb) -> None:
    """Re-apply the bilateral PNS listing formula and clear the spill area below it."""
    from openpyxl.worksheet.formula import ArrayFormula

    ws = wb[FUNDING_SHEET]
    ws.cell(FUNDING_PNS_FIRST_ROW, FUNDING_NETWORK_MEMBER_COL).value = ArrayFormula(
        text=_funding_pns_array_formula_text(wb),
        ref=f"B{FUNDING_PNS_FIRST_ROW}",
    )
    for row_idx in range(FUNDING_PNS_FIRST_ROW + 1, FUNDING_PNS_LAST_ROW + 1):
        ws.cell(row_idx, FUNDING_NETWORK_MEMBER_COL).value = None


def _dynamic_array_cell_flags_from_sheet_xml(xml: str) -> Dict[str, Dict[str, str]]:
    flags_by_cell: Dict[str, Dict[str, str]] = {}
    for match in re.finditer(r'<c r="([^"]+)"([^>]*)>', xml):
        ref, attrs = match.group(1), match.group(2)
        if 'cm="1"' not in attrs:
            continue
        flags: Dict[str, str] = {"cm": "1"}
        type_match = re.search(r'\bt="([^"]+)"', attrs)
        if type_match:
            flags["t"] = type_match.group(1)
        flags_by_cell[ref] = flags
    return flags_by_cell


def _patch_sheet_xml_dynamic_array_flags(xml: str, flags_by_cell: Dict[str, Dict[str, str]]) -> str:
    if not flags_by_cell:
        return xml

    def _patch_cell(match: re.Match[str]) -> str:
        ref, attrs = match.group(1), match.group(2)
        flags = flags_by_cell.get(ref)
        if not flags:
            return match.group(0)
        if flags.get("cm") and 'cm="1"' not in attrs:
            attrs += ' cm="1"'
        if flags.get("t") and not re.search(r'\bt="', attrs):
            attrs += f' t="{flags["t"]}"'
        return f'<c r="{ref}"{attrs}>'

    return re.sub(r'<c r="([^"]+)"([^>]*)>', _patch_cell, xml)


def restore_workbook_dynamic_array_metadata(template_path: str, output_path: str) -> None:
    """
    openpyxl drops Excel 365 dynamic-array metadata (xl/metadata.xml, cm=\"1\" flags).
    Copy them back from the canonical template so FILTER/UNIQUE formulas spill on open.
    """
    metadata_part = "xl/metadata.xml"
    rels_part = "xl/_rels/workbook.xml.rels"
    content_types_part = "[Content_Types].xml"
    metadata_rel_type = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sheetMetadata"
    )
    metadata_content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheetMetadata+xml"
    )

    with zipfile.ZipFile(template_path, "r") as template_zip:
        if metadata_part not in template_zip.namelist():
            return
        template_metadata = template_zip.read(metadata_part)
        sheet_flags = {
            name: _dynamic_array_cell_flags_from_sheet_xml(template_zip.read(name).decode("utf-8"))
            for name in template_zip.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        }

    patched_path = f"{output_path}.dynamic-array-patch"
    try:
        with zipfile.ZipFile(output_path, "r") as src, zipfile.ZipFile(
            patched_path, "w", zipfile.ZIP_DEFLATED
        ) as dst:
            wrote_metadata = False
            for item in src.infolist():
                data = src.read(item.filename)
                if item.filename == rels_part:
                    text = data.decode("utf-8")
                    if "sheetMetadata" not in text:
                        rel_ids = [int(value) for value in re.findall(r'Id="rId(\d+)"', text)]
                        next_id = max(rel_ids, default=0) + 1
                        insert = (
                            f'<Relationship Id="rId{next_id}" '
                            f'Type="{metadata_rel_type}" Target="metadata.xml"/>'
                        )
                        text = text.replace("</Relationships>", insert + "</Relationships>")
                    data = text.encode("utf-8")
                elif item.filename == content_types_part:
                    text = data.decode("utf-8")
                    if 'PartName="/xl/metadata.xml"' not in text:
                        override = (
                            f'<Override PartName="/xl/metadata.xml" '
                            f'ContentType="{metadata_content_type}"/>'
                        )
                        text = text.replace("</Types>", override + "</Types>")
                    data = text.encode("utf-8")
                elif item.filename == metadata_part:
                    data = template_metadata
                    wrote_metadata = True
                elif item.filename in sheet_flags and sheet_flags[item.filename]:
                    data = _patch_sheet_xml_dynamic_array_flags(
                        data.decode("utf-8"),
                        sheet_flags[item.filename],
                    ).encode("utf-8")
                dst.writestr(item, data)
            if not wrote_metadata:
                dst.writestr(metadata_part, template_metadata)
        os.replace(patched_path, output_path)
    finally:
        if os.path.isfile(patched_path):
            try:
                os.unlink(patched_path)
            except OSError:
                pass


def _ensure_workbook_recalculates_on_open(wb) -> None:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcOnSave = True


def _workbook_reach_ea_codes(wb, *, iso3: str, rnd: str) -> Dict[Tuple[str, str, str], str]:
    codes: Dict[Tuple[str, str, str], str] = {}
    for _reach, mdr_name, _ea_name, area in EA_SLOT_NAMED_CELLS:
        code = str(read_named_cell(wb, mdr_name) or "").strip().upper()
        if code:
            codes[(iso3, rnd, area)] = code
    return codes


def _workbook_reach_ea_names(wb, *, iso3: str, rnd: str) -> Dict[Tuple[str, str, str], str]:
    names: Dict[Tuple[str, str, str], str] = {}
    for _reach, _mdr_name, ea_name, area in EA_SLOT_NAMED_CELLS:
        name = str(read_named_cell(wb, ea_name) or "").strip()
        if name:
            names[(iso3, rnd, area)] = name
    return names


def _import_reach_matrices(
    wb,
    ctx,
    *,
    aes_id: int,
    iso3: str,
    period: str,
    rnd: str,
    warnings: List[str],
) -> Dict[int, Dict[str, Any]]:
    matrices: Dict[int, Dict[str, Any]] = defaultdict(dict)
    reach_ea_codes = _workbook_reach_ea_codes(wb, iso3=iso3, rnd=rnd)
    reach_ea_names = _workbook_reach_ea_names(wb, iso3=iso3, rnd=rnd)

    _, people_rows = read_named_table(wb, PEOPLE_SHEET, PEOPLE_TABLE)
    for row in people_rows:
        year_val = row.get("Year")
        if year_val in (None, ""):
            continue
        try:
            row_year = str(int(float(str(year_val))))
        except (ValueError, TypeError):
            continue
        if _year_offset(period, row_year) is None:
            warnings.append(f"Reach row year {row_year!r} is outside the planning window for {period}.")
            continue
        for area in SP_AREAS:
            amount = parse_value_num(row.get(area))
            if amount is None:
                continue
            matrices[ITEM_LONGER_TERM_PROGRAMMES][f"{row_year}_{area}"] = amount

    for (reach_name, mdr_name, ea_name, area) in EA_SLOT_NAMED_CELLS:
        amount_raw = read_named_cell(wb, reach_name)
        amount = parse_value_num(amount_raw)
        ea_code = read_named_cell(wb, mdr_name)
        excel_name = read_named_cell(wb, ea_name)
        if amount is None:
            continue
        ea_cells = _resolve_emergency_matrix_cells(
            ctx,
            iso3=iso3,
            area=area,
            ea_code=ea_code,
            excel_name=excel_name,
            amount=amount,
        )
        if not ea_cells:
            warnings.append(f"Could not resolve emergency appeal row for {area} on Reach sheet.")
            continue
        matrices[ITEM_EMERGENCY_APPEALS].update(ea_cells)
        code_text = str(ea_code or "").strip().upper()
        if code_text:
            reach_ea_codes[(iso3, rnd, area)] = code_text

    return matrices


def _parse_planning_support_ticks(cells: Dict[str, Any]) -> Dict[int, Dict[str, bool]]:
    """Parse item 955 matrix keys into ns_id -> {area: True}."""
    ticks: Dict[int, Dict[str, bool]] = {}
    for cell_key, raw in cells.items():
        if not _cell_is_tick(raw):
            continue
        match = _PLANNING_SUPPORT_CELL_KEY_RE.match(str(cell_key).strip())
        if not match:
            continue
        ns_id = int(match.group(1))
        area = match.group(2)
        ticks.setdefault(ns_id, {})[area] = True
    return ticks


def _write_support_region_cell(wb, row_offset: int, region_name: str) -> None:
    _min_col, min_row = _bilateral_table_row_info(wb, SUPPORT_SHEET, SUPPORT_TABLE)
    wb[SUPPORT_SHEET].cell(min_row + 1 + row_offset, SUPPORT_NS_REGION_COL).value = str(region_name or "").strip()


def _build_workbook_ns_region_index(wb) -> Dict[str, str]:
    """NS name (lower) -> workbook region label from TemplateData/Table9."""
    out: Dict[str, str] = {}
    if not _workbook_table_exists(wb, COUNTRY_REGION_TABLE_SHEET, NS_REGION_TABLE):
        return out
    try:
        _, rows = read_named_table(wb, COUNTRY_REGION_TABLE_SHEET, NS_REGION_TABLE)
    except ValueError:
        return out
    for row in rows:
        ns_name = str(row.get("NS name") or row.get("Value") or "").strip()
        region = str(row.get("Region") or "").strip()
        if ns_name and region:
            out[ns_name.lower()] = region
    return out


def _workbook_region_for_ns_name(
    wb,
    ns_name: str,
    *,
    table_index: Optional[Dict[str, str]] = None,
    fallback_region: str = "",
) -> str:
    name = str(ns_name or "").strip()
    if name:
        index = table_index if table_index is not None else _build_workbook_ns_region_index(wb)
        region = index.get(name.lower())
        if region:
            return region
    return _workbook_region_label(fallback_region)


def _planning_support_ns_regions(
    wb,
    ns_ids: List[int],
    id_to_name: Dict[int, str],
) -> Dict[int, str]:
    table_index = _build_workbook_ns_region_index(wb)
    regions: Dict[int, str] = {}
    db_regions: Dict[int, str] = {}
    try:
        from app.models.organization import NationalSociety

        for ns in NationalSociety.query.filter(NationalSociety.id.in_(list(ns_ids))).all():
            country_region = (ns.country.region if ns.country else "") or ""
            db_regions[int(ns.id)] = _workbook_region_label(country_region)
    except Exception:
        db_regions = {}

    for ns_id in ns_ids:
        ns_name = id_to_name.get(int(ns_id), "")
        region = _workbook_region_for_ns_name(
            wb,
            ns_name,
            table_index=table_index,
            fallback_region=db_regions.get(int(ns_id), ""),
        )
        if region:
            regions[int(ns_id)] = region
    return regions


def _planning_support_ns_display_names(ctx, ns_ids: List[int]) -> Dict[int, str]:
    id_to_name: Dict[int, str] = {}
    try:
        from app.models.organization import NationalSociety

        for ns in NationalSociety.query.filter(NationalSociety.id.in_(list(ns_ids))).all():
            id_to_name[int(ns.id)] = (ns.name or "").strip()
    except Exception:
        id_to_name = {}

    for ns_id in ns_ids:
        if int(ns_id) in id_to_name:
            continue
        for db_name, db_id in ctx.ns_name_to_id.items():
            if db_id == int(ns_id):
                id_to_name[int(ns_id)] = db_name
                break
    return id_to_name


def _clear_planning_support_table(wb) -> None:
    capacity = _table_data_row_capacity(wb, SUPPORT_SHEET, SUPPORT_TABLE)
    for offset in range(capacity):
        _write_support_region_cell(wb, offset, "")
        _write_bilateral_ns_source_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, offset, "")
        for area in (*SP_AREAS, "EFs"):
            write_table_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, offset, area, None)


def _import_support_matrix(wb, ctx, *, aes_id: int, warnings: List[str]) -> Dict[str, Any]:
    cells: Dict[str, Any] = {}
    capacity = _table_data_row_capacity(wb, SUPPORT_SHEET, SUPPORT_TABLE)
    for offset in range(capacity):
        ns_name = _bilateral_ns_name_for_row(wb, SUPPORT_SHEET, SUPPORT_TABLE, offset)
        if not ns_name:
            continue
        ns_id = _resolve_ns_row_id(ctx, ns_name)
        if ns_id is None:
            continue
        for area in (*SP_AREAS, "EFs"):
            val = read_table_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, offset, area)
            if _cell_is_tick(val):
                cells[f"{ns_id}_{area}"] = 1
    return cells


def _import_funding_matrices(
    wb,
    ctx,
    *,
    aes_id: int,
    iso3: str,
    period: str,
    rnd: str,
    warnings: List[str],
) -> Dict[int, Dict[str, Any]]:
    matrices: Dict[int, Dict[str, Any]] = defaultdict(dict)
    reach_ea_codes = _workbook_reach_ea_codes(wb, iso3=iso3, rnd=rnd)
    reach_ea_names = _workbook_reach_ea_names(wb, iso3=iso3, rnd=rnd)
    headers = _funding_table_headers(wb)
    funding_headers = [h for h in headers if parse_funding_column_header(h)]

    matrix_cells: Dict[Tuple[int, int], Dict[str, Any]] = defaultdict(dict)
    bilateral_names = _collect_bilateral_support_ns_names(wb)

    for funding_item_id in FUNDING_MATRIX_BY_YEAR_OFFSET.values():
        for area in sorted(PLANNING_EA_FUNDING_AREAS):
            _ensure_funding_ea_col_header(
                matrix_cells,
                ctx,
                aes_id=aes_id,
                funding_item_id=funding_item_id,
                iso3=iso3,
                rnd=rnd,
                area=area,
                ea_code_raw=reach_ea_codes.get((iso3, rnd, area)),
                reach_ea_codes=reach_ea_codes,
                excel_name_raw=reach_ea_names.get((iso3, rnd, area)),
                reach_ea_names=reach_ea_names,
            )

    for excel_row, row_key in (
        (FUNDING_HNS_ROW, "HNS"),
        (FUNDING_IFRC_ROW, "IFRC Secretariat"),
    ):
        row = _read_funding_table_row(wb, excel_row, headers)
        _import_funding_cells_for_row(
            matrix_cells,
            ctx,
            aes_id=aes_id,
            iso3=iso3,
            period=period,
            rnd=rnd,
            row=row,
            row_key=row_key,
            funding_headers=funding_headers,
            reach_ea_codes=reach_ea_codes,
            reach_ea_names=reach_ea_names,
            warnings=warnings,
        )

    for pns_idx, ns_name in enumerate(bilateral_names):
        excel_row = FUNDING_PNS_FIRST_ROW + pns_idx
        if excel_row > FUNDING_PNS_LAST_ROW:
            break
        ns_id = _resolve_ns_row_id(ctx, ns_name)
        if ns_id is None:
            continue
        row = _read_funding_table_row(wb, excel_row, headers)
        _import_funding_cells_for_row(
            matrix_cells,
            ctx,
            aes_id=aes_id,
            iso3=iso3,
            period=period,
            rnd=rnd,
            row=row,
            row_key=str(ns_id),
            funding_headers=funding_headers,
            reach_ea_codes=reach_ea_codes,
            reach_ea_names=reach_ea_names,
            warnings=warnings,
        )
        _ensure_funding_pns_rows_in_matrices(matrix_cells, aes_id=aes_id, ns_id=ns_id)

    for (aid, item_id), cells in matrix_cells.items():
        if aid == aes_id and cells:
            matrices[item_id].update(cells)
    return matrices


def _resolve_comments_item_id(ctx) -> Optional[int]:
    labels = ctx.item_ids_by_label.get(PLANNING_COUNTRY_TEMPLATE_ID, {})
    for key, item_id in labels.items():
        if "comment" in str(key).lower():
            return int(item_id)
    return ITEM_COMMENTS


def _resolve_comment_named_cell(wb) -> Optional[str]:
    if COMMENT_NAMED_CELL in wb.defined_names:
        return COMMENT_NAMED_CELL
    for name in wb.defined_names:
        if str(name).lower() == COMMENT_NAMED_CELL.lower():
            return str(name)
    return None


def parse_comment(wb) -> str:
    """Read the single Comments sheet cell aligned with form item 956."""
    named_cell = _resolve_comment_named_cell(wb)
    if not named_cell:
        return ""
    return str(read_named_cell(wb, named_cell) or "").strip()


def _import_comment_field_from_workbook(wb, ctx) -> Dict[str, Dict[str, Any]]:
    text = parse_comment(wb)
    if not text:
        return {}
    item_id = _resolve_comments_item_id(ctx)
    if not item_id:
        return {}
    return {str(item_id): {"value": text}}


def _export_comment_to_workbook(wb, entry) -> None:
    named_cell = _resolve_comment_named_cell(wb)
    if not named_cell:
        return
    text = str(_scalar_value(entry) or "").strip()
    write_named_cell(wb, named_cell, text or None)


def _ns_fields_from_workbook(wb, ctx) -> Dict[str, Dict[str, Any]]:
    fields: Dict[str, Dict[str, Any]] = {}
    bank_map = ctx.items_by_bank_id.get(PLANNING_COUNTRY_TEMPLATE_ID, {})
    for _key, bank_id in NS_DATA_BANK_IDS.items():
        item_id = bank_map.get(bank_id)
        if not item_id:
            continue
        cell_name = NS_DATA_NAMED_CELLS[_key]
        raw = read_named_cell(wb, cell_name)
        if raw is None or str(raw).strip() == "":
            continue
        fields[str(item_id)] = {"value": raw}
    return fields


def _workbook_matrices_from_payload(
    wb,
    ctx,
    *,
    aes_id: int,
    iso3: str,
    period: str,
    warnings: List[str],
) -> Dict[int, Dict[str, Any]]:
    rnd, _ = parse_version(wb)
    reach = _import_reach_matrices(
        wb, ctx, aes_id=aes_id, iso3=iso3, period=period, rnd=rnd, warnings=warnings
    )
    support = _import_support_matrix(wb, ctx, aes_id=aes_id, warnings=warnings)
    funding = _import_funding_matrices(
        wb, ctx, aes_id=aes_id, iso3=iso3, period=period, rnd=rnd, warnings=warnings
    )
    matrices: Dict[int, Dict[str, Any]] = {}
    for item_id, cells in reach.items():
        if cells:
            matrices[item_id] = dict(cells)
    if support:
        matrices[ITEM_BILATERAL_SUPPORT] = support
    for item_id, cells in funding.items():
        if cells:
            matrices[item_id] = dict(cells)
    return matrices


def build_unified_country_plan_client_payload(
    aes_id: int,
    wb,
    ctx,
    *,
    iso3: str,
    period: str,
    warnings: Optional[List[str]] = None,
) -> Dict[str, Any]:
    warn = warnings if warnings is not None else []
    matrices = _workbook_matrices_from_payload(
        wb,
        ctx,
        aes_id=aes_id,
        iso3=iso3,
        period=period,
        warnings=warn,
    )
    return {
        "fields": {
            **_ns_fields_from_workbook(wb, ctx),
            **_import_comment_field_from_workbook(wb, ctx),
        },
        "matrices": {str(k): v for k, v in matrices.items()},
        "dynamic_indicators": [],
        "repeat_slots": [],
        "meta": {"iso3": iso3, "period": period},
    }


def _parse_emergency_row_id(row_id: str) -> Tuple[str, str]:
    text = (row_id or "").strip()
    if text.endswith(")"):
        open_idx = text.rfind("(")
        if open_idx > 0:
            return text[:open_idx].strip(), text[open_idx + 1 : -1].strip()
    return text, ""


def _export_reach_to_workbook(wb, entry_954, entry_960) -> None:
    cells_954 = _normalize_matrix_cells(_matrix_cells(entry_954))
    _, people_rows = read_named_table(wb, PEOPLE_SHEET, PEOPLE_TABLE)
    for offset, row in enumerate(people_rows):
        year_val = row.get("Year")
        if year_val in (None, ""):
            continue
        try:
            row_year = str(int(float(str(year_val))))
        except (ValueError, TypeError):
            continue
        for area in SP_AREAS:
            val = _matrix_cell_scalar(cells_954.get(f"{row_year}_{area}"))
            if val is not None:
                write_table_cell(wb, PEOPLE_SHEET, PEOPLE_TABLE, offset, area, val)

    cells_960 = _normalize_matrix_cells(_matrix_cells(entry_960))
    suffix = f"_{EMERGENCY_APPEALS_COLUMN}"
    ea_entries: List[Tuple[str, Any]] = []
    for key, raw in cells_960.items():
        if not str(key).endswith(suffix):
            continue
        row_id = str(key)[: -len(suffix)]
        val = _matrix_cell_scalar(raw)
        if val is None:
            continue
        ea_entries.append((row_id, val))
    ea_entries.sort(key=lambda item: item[0])
    for idx, (reach_name, mdr_name, ea_name, _area) in enumerate(EA_SLOT_NAMED_CELLS):
        if idx >= len(ea_entries):
            break
        row_id, val = ea_entries[idx]
        name, code = _parse_emergency_row_id(row_id)
        write_named_cell(wb, reach_name, val)
        if code:
            write_named_cell(wb, mdr_name, code)
        if name:
            write_named_cell(wb, ea_name, name)


def _export_support_to_workbook(wb, entry_955, ctx) -> None:
    cells = _normalize_matrix_cells(_matrix_cells(entry_955))
    ticks_by_ns = _parse_planning_support_ticks(cells)
    if not ticks_by_ns:
        return

    ns_ids = sorted(ticks_by_ns.keys())
    id_to_name = _planning_support_ns_display_names(ctx, ns_ids)
    id_to_region = _planning_support_ns_regions(wb, ns_ids, id_to_name)
    _clear_planning_support_table(wb)

    for offset, ns_id in enumerate(ns_ids):
        ns_name = id_to_name.get(int(ns_id), "")
        if not ns_name:
            continue
        region = id_to_region.get(int(ns_id), "")
        if region:
            _write_support_region_cell(wb, offset, region)
        _write_bilateral_ns_source_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, offset, ns_name)
        for area, ticked in ticks_by_ns[int(ns_id)].items():
            if ticked:
                write_table_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, offset, area, "X")


def _export_funding_to_workbook(wb, entries, period: str, ctx) -> None:
    headers, rows = read_named_table(wb, FUNDING_SHEET, FUNDING_TABLE)
    funding_headers = [h for h in headers if parse_funding_column_header(h)]
    for offset, item_id in FUNDING_MATRIX_BY_YEAR_OFFSET.items():
        cells = _normalize_matrix_cells(_matrix_cells(entries.get(item_id)))
        if not cells:
            continue
        for row_offset, row in enumerate(rows):
            entity, ns_name = _funding_row_entity_from_workbook(wb, row)
            for header in funding_headers:
                parsed = parse_funding_column_header(header)
                if not parsed:
                    continue
                area, year = parsed
                if _year_offset(period, year) != offset:
                    continue
                if entity == "hns":
                    row_key = "HNS"
                elif entity == "ifrc":
                    row_key = "IFRC Secretariat"
                else:
                    ns_id = _resolve_ns_row_id(ctx, ns_name)
                    if ns_id is None:
                        continue
                    row_key = str(ns_id)
                val = _matrix_cell_scalar(cells.get(f"{row_key}_{area}"))
                if val is not None:
                    write_table_cell(wb, FUNDING_SHEET, FUNDING_TABLE, row_offset, header, val)


def build_unified_country_plan_export(aes_id: int, template_path: str, output_path: str) -> Dict[str, Any]:
    """Fill Unified Country Plan template with assignment data and save to output_path."""
    _require_openpyxl()
    import openpyxl

    _aes, country_name, iso3, period, region = _load_assignment_meta(aes_id)
    if not os.path.isfile(template_path):
        raise FileNotFoundError(f"Unified Country Plan template not found: {template_path}")

    with _quiet_openpyxl_io():
        wb = openpyxl.load_workbook(template_path)
    rewrite_planning_year_headers(wb, period)
    write_named_cell(wb, "Version", period_to_workbook_version(period))
    _write_start_sheet_selection(
        wb,
        country_name,
        _workbook_region_for_country(wb, country_name, fallback_region=region),
    )

    ctx = build_import_context([PLANNING_COUNTRY_TEMPLATE_ID])
    entries = _load_form_data_map(aes_id)
    for key, bank_id in NS_DATA_BANK_IDS.items():
        item_id = ctx.items_by_bank_id.get(PLANNING_COUNTRY_TEMPLATE_ID, {}).get(bank_id)
        if not item_id:
            continue
        value = _scalar_value(entries.get(item_id))
        if value is not None:
            write_named_cell(wb, NS_DATA_NAMED_CELLS[key], value)

    _export_reach_to_workbook(
        wb,
        entries.get(ITEM_LONGER_TERM_PROGRAMMES),
        entries.get(ITEM_EMERGENCY_APPEALS),
    )
    _export_support_to_workbook(wb, entries.get(ITEM_BILATERAL_SUPPORT), ctx)
    _refresh_funding_pns_array_formula(wb)
    _export_funding_to_workbook(wb, entries, period, ctx)
    comments_item_id = _resolve_comments_item_id(ctx)
    _export_comment_to_workbook(wb, entries.get(comments_item_id))

    _ensure_workbook_recalculates_on_open(wb)
    with _quiet_openpyxl_io():
        wb.save(output_path)
    wb.close()
    restore_workbook_dynamic_array_metadata(template_path, output_path)

    safe_country = re.sub(r"[^\w\-]+", "_", country_name or iso3 or "country").strip("_") or "country"
    safe_period = re.sub(r"[^\w\-]+", "_", period or "period").strip("_") or "period"
    filename = f"Unified_Country_Plan_{safe_country}_{safe_period}.xlsx"
    return {"filename": filename, "aes_id": aes_id}


def _payload_to_import_rows(
    aes_id: int,
    payload: Dict[str, Any],
    *,
    iso3: str,
    period: str,
) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for item_id, field_data in (payload.get("fields") or {}).items():
        built = _scalar_row(
            aes_id=aes_id,
            item_id=int(item_id),
            value=field_data.get("value"),
            iso3=iso3,
            period=period,
            debug_kpi=f"field_{item_id}",
        )
        if built:
            rows.append(built)
    for item_id, cells in (payload.get("matrices") or {}).items():
        if not cells:
            continue
        rows.append(
            _matrix_row(
                aes_id=aes_id,
                item_id=int(item_id),
                cells=cells,
                iso3=iso3,
                period=period,
                debug_kpi=f"matrix_{item_id}",
            )
        )
    return rows


def run_unified_country_plan_import(
    aes_id: int,
    workbook_path: str,
    *,
    dry_run: bool = False,
    persist: bool = True,
) -> Dict[str, Any]:
    """Import a filled Unified Country Plan workbook into the assignment."""
    _require_openpyxl()
    import openpyxl

    from app.extensions import db
    from app.models.form_items import FormItem

    _aes, country_name, iso3, period, _region = _load_assignment_meta(aes_id)
    ctx = build_import_context([PLANNING_COUNTRY_TEMPLATE_ID])

    with _quiet_openpyxl_io():
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
    try:
        validation = validate_unified_country_plan_import_file(
            wb,
            expected_country=country_name,
            expected_period=period,
        )
        if not validation.get("valid"):
            return {
                "success": False,
                "errors": len(validation.get("errors") or []),
                "message": validation.get("message") or "Validation failed.",
                "warnings": validation.get("warnings") or [],
            }

        warnings: List[str] = list(validation.get("warnings") or [])
        payload = build_unified_country_plan_client_payload(
            aes_id,
            wb,
            ctx,
            iso3=iso3,
            period=period,
            warnings=warnings,
        )
        warnings.extend(ctx.warnings)
        warnings = dedupe_upr_import_warnings(warnings)
        warning_texts, warning_items = serialize_upr_import_warnings(warnings)

        field_count = len(payload.get("fields") or {})
        matrix_count = sum(len(v or {}) for v in (payload.get("matrices") or {}).values())
        updated_count = field_count + matrix_count

        if not persist:
            return {
                "success": True,
                "stage_only": True,
                "payload": payload,
                "warnings": warning_texts,
                "warning_items": warning_items,
                "updated_count": updated_count,
            }

        import_rows = _payload_to_import_rows(aes_id, payload, iso3=iso3, period=period)
        valid_item_ids = {
            int(fid)
            for (fid,) in db.session.query(FormItem.id)
            .filter(FormItem.template_id == PLANNING_COUNTRY_TEMPLATE_ID)
            .all()
        }
        stats = upsert_form_data_rows(
            import_rows,
            dry_run=dry_run,
            valid_form_item_ids=valid_item_ids,
        )
        stats["success"] = True
        stats["warnings"] = warning_texts
        stats["warning_items"] = warning_items
        stats["updated_count"] = int(stats.get("inserted", 0) or 0) + int(stats.get("updated", 0) or 0)
        return stats
    finally:
        wb.close()


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Unified Country Plan Excel round-trip (T24)")
    parser.add_argument("--aes-id", type=int, required=True)
    parser.add_argument(
        "--template",
        default=os.path.join(backoffice_dir, "app", "static", "templates", "unified_country_plan.xlsx"),
    )
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    print(json.dumps(build_unified_country_plan_export(args.aes_id, args.template, args.output), indent=2))
