#!/usr/bin/env python3
"""Compare UPR Master ``PNS Data`` sheet vs ``UPR Data`` for PNS planning import (P26)."""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, Optional, Set, Tuple

script_dir = Path(__file__).resolve().parent
backoffice_dir = script_dir.parent
if str(backoffice_dir) not in sys.path:
    sys.path.insert(0, str(backoffice_dir))
if str(script_dir) not in sys.path:
    sys.path.insert(0, str(script_dir))

from import_upr_excel_data import (  # noqa: E402
    HEADER_ROW_INDEX,
    T22_BREAKDOWN_AREAS,
    _t22_pns_import_cell_value,
    _year_offset,
    is_planning_funding_requirement_row,
    load_upr_data_sheet,
    parse_pns_reported_yes,
    parse_value_num,
    round_to_period,
)

PNS_DATA_SHEET = "PNS Data"
PNS_BREAKDOWN_COLS = ("SP1", "SP2", "SP3", "SP4", "SP5", "EFs")
PNS_TOTAL_COL = "Funding Requirement"
PNS_INFO_COLS = ("Confirmed Funding",)


def _norm_ns(name: Any) -> str:
    return str(name or "").strip().lower()


def _norm_country(name: Any) -> str:
    return str(name or "").strip().lower()


def _num(raw: Any) -> Optional[float]:
    if raw is None or raw == "":
        return None
    try:
        val = float(raw)
        return val if val != 0 else None
    except (TypeError, ValueError):
        return None


def load_pns_data_sheet(path: str) -> Tuple[list[str], list[Dict[str, Any]]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if PNS_DATA_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet {PNS_DATA_SHEET!r} not found")
    ws = wb[PNS_DATA_SHEET]
    headers: list[str] = []
    rows: list[Dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h is not None else "" for h in row]
            continue
        if not any(c is not None and str(c).strip() for c in row):
            continue
        record = {}
        for j, h in enumerate(headers):
            if not h:
                continue
            record[h] = row[j] if j < len(row) else None
        rows.append(record)
    wb.close()
    return headers, rows


def _planning_year_matches(row: Dict[str, Any], rnd: str) -> bool:
    period = round_to_period(rnd)
    if not period:
        return False
    year_val = row.get("Year")
    if year_val in (None, ""):
        return False
    return _year_offset(period, year_val) == 0


def pns_data_p26_records(rows: Iterable[Dict[str, Any]], rnd: str = "P26") -> Dict[Tuple[str, str], Dict[str, Any]]:
    """Wide PNS Data rows keyed by (ns_lower, country_lower) for the planning year only."""
    out: Dict[Tuple[str, str], Dict[str, Any]] = {}
    rnd_u = rnd.strip().upper()
    for row in rows:
        if str(row.get("Round") or "").strip().upper() != rnd_u:
            continue
        if not _planning_year_matches(row, rnd):
            continue
        key = (_norm_ns(row.get("NS")), _norm_country(row.get("Country")))
        if not key[0] or not key[1]:
            continue
        metrics: Dict[str, Optional[float]] = {}
        for col in PNS_BREAKDOWN_COLS:
            metrics[col] = _num(row.get(col))
        metrics[PNS_TOTAL_COL] = _num(row.get(PNS_TOTAL_COL))
        for col in PNS_INFO_COLS:
            metrics[col] = _num(row.get(col))
        metrics["Funding"] = _num(row.get("Funding"))
        metrics["Expenditure"] = _num(row.get("Expenditure"))
        metrics["Transferred"] = _num(row.get("Transferred"))
        out[key] = {
            "ns": row.get("NS"),
            "country": row.get("Country"),
            "date_submitted": row.get("Date Submitted"),
            "metrics": metrics,
            "staff": {
                "intl_delegates_hns": _num(row.get("Delegates integrated with the HNS")),
                "national_staff_hns_hns": _num(
                    row.get(
                        "National staff hired through the HNS (PNS operating under HNS legal umbrella)"
                    )
                ),
                "intl_delegates_ifrc": _num(row.get("Delegates integrated with the IFRC")),
                "national_staff_ifrc_ifrc": _num(
                    row.get(
                        "National staff hired through the IFRC (PNS operating under IFRC legal umbrella)"
                    )
                ),
                "national_staff_hns_ifrc": _num(
                    row.get(
                        "National staff hired through the HNS (PNS operating under IFRC legal umbrella)"
                    )
                ),
            },
        }
    return out


def upr_data_p26_pns_funding(
    rows: Iterable[Dict[str, Any]], rnd: str = "P26"
) -> Dict[Tuple[str, str], Dict[str, Any]]:
    """Aggregate UPR Data PNS funding rows the import would use for T22 (reported Yes only)."""
    rnd_u = rnd.strip().upper()
    staging: Dict[Tuple[str, str, str], Tuple[Optional[float], Optional[float]]] = defaultdict(
        lambda: (None, None)
    )
    reported_pairs: Set[Tuple[str, str]] = set()
    all_pns_pairs: Set[Tuple[str, str]] = set()

    for row in rows:
        if str(row.get("Round") or "").strip().upper() != rnd_u:
            continue
        if str(row.get("Section") or "").strip() != "Funding":
            continue
        if str(row.get("Entity") or "").strip().upper() != "PNS":
            continue
        if not is_planning_funding_requirement_row(row):
            continue
        if not _planning_year_matches(row, rnd):
            continue
        ns = _norm_ns(row.get("NS"))
        country = _norm_country(row.get("Country"))
        if not ns or not country or ns == "country":
            continue
        pair = (ns, country)
        all_pns_pairs.add(pair)
        if parse_pns_reported_yes(row):
            reported_pairs.add(pair)

        if not parse_pns_reported_yes(row):
            continue

        area = str(row.get("Area") or "").strip()
        if not area:
            continue
        cv = parse_value_num(row.get("Country Value"))
        pv = parse_value_num(row.get("PNS Value"))
        if cv is None and pv is None:
            continue
        staging[(ns, country, area)] = (cv, pv)

    out: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for (ns, country, area), (cv, pv) in staging.items():
        key = (ns, country)
        bucket = out.setdefault(
            key,
            {
                "areas": {},
                "reported": key in reported_pairs,
            },
        )
        submitted = _t22_pns_import_cell_value(cv, pv)
        if isinstance(submitted, dict):
            submitted = "CLEARED"
        bucket["areas"][area] = {
            "country_value": cv,
            "pns_value": pv,
            "submitted": submitted,
        }

    for pair in reported_pairs:
        out.setdefault(pair, {"areas": {}, "reported": True})
    return out


def compare_round(path: str, rnd: str = "P26") -> Dict[str, Any]:
    _, pns_rows = load_pns_data_sheet(path)
    _, upr_rows = load_upr_data_sheet(path)

    pns = pns_data_p26_records(pns_rows, rnd=rnd)
    upr = upr_data_p26_pns_funding(upr_rows, rnd=rnd)

    pns_keys = set(pns.keys())
    upr_reported_keys = {k for k, v in upr.items() if v.get("reported") or v.get("areas")}
    upr_all_keys = set(upr.keys())

    only_pns_sheet = sorted(pns_keys - upr_reported_keys)
    only_upr_reported = sorted(upr_reported_keys - pns_keys)
    common = sorted(pns_keys & upr_reported_keys)

    funding_mismatches = []
    breakdown_mismatches = []
    for key in common:
        p = pns[key]["metrics"]
        u_areas = upr[key]["areas"]

        # Compare total: PNS sheet Funding Requirement vs UPR Area=Total (Funding Requirement only)
        pns_total = p.get(PNS_TOTAL_COL)
        upr_total = None
        if "Total" in u_areas:
            upr_total = u_areas["Total"].get("submitted")
        if pns_total != upr_total and not (pns_total is None and upr_total is None):
            funding_mismatches.append(
                {
                    "ns": pns[key]["ns"],
                    "country": pns[key]["country"],
                    "pns_sheet_total": pns_total,
                    "upr_total_submitted": upr_total,
                }
            )

        # Compare SP/EF breakdown columns vs UPR area rows (submitted scalar logic)
        for area in PNS_BREAKDOWN_COLS:
            pns_val = p.get(area)
            upr_val = u_areas.get(area, {}).get("submitted") if area in u_areas else None
            if pns_val != upr_val and not (pns_val is None and upr_val is None):
                breakdown_mismatches.append(
                    {
                        "ns": pns[key]["ns"],
                        "country": pns[key]["country"],
                        "area": area,
                        "pns_sheet": pns_val,
                        "upr_import": upr_val,
                    }
                )

    # PNS sheet rows with any funding data but missing from UPR reported
    pns_with_funding_not_reported = []
    for key in only_pns_sheet:
        metrics = pns[key]["metrics"]
        if any(metrics.get(c) for c in (*PNS_BREAKDOWN_COLS, PNS_TOTAL_COL)):
            pns_with_funding_not_reported.append(
                {
                    "ns": pns[key]["ns"],
                    "country": pns[key]["country"],
                    "funding_requirement": metrics.get("Funding Requirement"),
                    "confirmed_funding": metrics.get("Confirmed Funding"),
                }
            )

    return {
        "round": rnd,
        "pns_sheet_rows": len(pns_keys),
        "upr_pns_funding_pairs_any": len(upr_all_keys),
        "upr_pns_reported_pairs": len(upr_reported_keys),
        "common_pairs": len(common),
        "only_on_pns_sheet": len(only_pns_sheet),
        "only_on_upr_reported": len(only_upr_reported),
        "total_value_mismatches": len(funding_mismatches),
        "breakdown_value_mismatches": len(breakdown_mismatches),
        "sample_only_pns_sheet": [
            {"ns": pns[k]["ns"], "country": pns[k]["country"]} for k in only_pns_sheet[:10]
        ],
        "sample_only_upr_reported": [
            {"ns": upr[k].get("ns", k[0]), "country": k[1]} for k in only_upr_reported[:10]
        ],
        "sample_total_mismatches": funding_mismatches[:15],
        "sample_breakdown_mismatches": breakdown_mismatches[:20],
        "pns_sheet_funding_not_in_upr_reported_count": len(pns_with_funding_not_reported),
        "sample_pns_funding_not_reported": pns_with_funding_not_reported[:15],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("path", help="Path to UPR Master.xlsx")
    parser.add_argument("--round", default="P26", help="Round code (default: P26)")
    args = parser.parse_args()

    result = compare_round(args.path, rnd=args.round)
    print(f"=== PNS Data vs UPR Data comparison ({result['round']}) ===")
    print(f"PNS Data sheet rows:              {result['pns_sheet_rows']}")
    print(f"UPR Data PNS funding pairs (any): {result['upr_pns_funding_pairs_any']}")
    print(f"UPR Data PNS reported=Yes pairs:  {result['upr_pns_reported_pairs']}")
    print(f"Common (NS, Country) pairs:       {result['common_pairs']}")
    print(f"Only on PNS Data sheet:           {result['only_on_pns_sheet']}")
    print(f"Only on UPR Data (reported Yes):  {result['only_on_upr_reported']}")
    print(f"Total funding mismatches:         {result['total_value_mismatches']}")
    print(f"SP/EF breakdown mismatches:       {result['breakdown_value_mismatches']}")
    print(
        "PNS sheet funding rows not in UPR reported:",
        result["pns_sheet_funding_not_in_upr_reported_count"],
    )

    if result["sample_only_pns_sheet"]:
        print("\nSample only on PNS Data sheet:")
        for row in result["sample_only_pns_sheet"]:
            print(f"  - {row['ns']} × {row['country']}")

    if result["sample_only_upr_reported"]:
        print("\nSample only on UPR Data (PNS reported=Yes):")
        for row in result["sample_only_upr_reported"]:
            print(f"  - {row['ns']} × {row['country']}")

    if result["sample_total_mismatches"]:
        print("\nSample total funding mismatches:")
        for row in result["sample_total_mismatches"]:
            print(
                f"  - {row['ns']} × {row['country']}: "
                f"PNS sheet={row['pns_sheet_total']} vs UPR import={row['upr_total_submitted']}"
            )

    if result["sample_breakdown_mismatches"]:
        print("\nSample SP/EF breakdown mismatches:")
        for row in result["sample_breakdown_mismatches"]:
            print(
                f"  - {row['ns']} × {row['country']} {row['area']}: "
                f"PNS sheet={row['pns_sheet']} vs UPR import={row['upr_import']}"
            )

    if result["sample_pns_funding_not_reported"]:
        print("\nSample PNS sheet funding rows missing UPR reported=Yes:")
        for row in result["sample_pns_funding_not_reported"]:
            print(
                f"  - {row['ns']} × {row['country']}: "
                f"req={row['funding_requirement']} confirmed={row['confirmed_funding']}"
            )


if __name__ == "__main__":
    main()
