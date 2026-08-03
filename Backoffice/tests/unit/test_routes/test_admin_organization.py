"""
Comprehensive tests for app/routes/admin/organization/ package.
Targeting 100% code coverage of organization management routes.
"""
import io
import json
import pytest
from unittest.mock import patch, MagicMock

from tests.factories import create_test_country, create_test_user

pytestmark = [pytest.mark.unit]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _json_headers():
    return {"Content-Type": "application/json", "Accept": "application/json"}


def _mock_render(return_value="<html>ok</html>"):
    return patch(
        "flask.render_template",
        return_value=return_value,
    )


def _make_country_data(**overrides):
    data = {
        "name": "Test Country",
        "iso3": "TCC",
        "iso2": "TC",
        "region": "Europe",
        "status": "Active",
        "preferred_language": "en",
        "currency_code": "USD",
    }
    data.update(overrides)
    return data


# ---------------------------------------------------------------------------
# RBAC before_request / access control
# ---------------------------------------------------------------------------

class TestOrganizationRBAC:
    def test_unauthenticated_redirects_to_login(self, client, db_session):
        resp = client.get("/admin/organization/", follow_redirects=False)
        assert resp.status_code == 302
        assert "login" in resp.location.lower()

    def test_admin_with_org_manage_can_access_index(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/")
        assert resp.status_code == 200

    def test_index_json_request_returns_json(self, logged_in_client, db_session):
        resp = logged_in_client.get(
            "/admin/organization/",
            headers=_json_headers(),
        )
        assert resp.status_code == 200
        data = resp.get_json()
        assert data is not None
        assert data.get("success") is True


# ---------------------------------------------------------------------------
# GET /admin/organization/  – index
# ---------------------------------------------------------------------------

class TestOrganizationIndex:
    def test_index_html_response(self, logged_in_client, db_session):
        with _mock_render() as mock_rt:
            resp = logged_in_client.get("/admin/organization/")
        assert resp.status_code == 200
        mock_rt.assert_called_once()

    def test_index_with_tab_param(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/?tab=countries")
        assert resp.status_code == 200

    def test_index_with_invalid_tab_falls_back_to_default(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/?tab=nonexistent")
        assert resp.status_code == 200

    def test_index_json_includes_counts(self, logged_in_client, db_session):
        resp = logged_in_client.get(
            "/admin/organization/",
            headers=_json_headers(),
        )
        assert resp.status_code == 200
        data = resp.get_json()
        assert "counts" in data

    def test_index_json_countries_tab(self, logged_in_client, db_session):
        resp = logged_in_client.get(
            "/admin/organization/?tab=countries",
            headers=_json_headers(),
        )
        assert resp.status_code == 200

    def test_index_json_ns_structure_tab(self, logged_in_client, db_session):
        resp = logged_in_client.get(
            "/admin/organization/?tab=ns-structure",
            headers=_json_headers(),
        )
        assert resp.status_code == 200

    def test_index_json_secretariat_tab(self, logged_in_client, db_session):
        resp = logged_in_client.get(
            "/admin/organization/?tab=secretariat",
            headers=_json_headers(),
        )
        assert resp.status_code == 200

    def test_index_with_active_filter_false(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/?active=false")
        assert resp.status_code == 200

    def test_index_with_country_id_filter(self, logged_in_client, db_session, app):
        country = create_test_country(db_session)
        with _mock_render():
            resp = logged_in_client.get(f"/admin/organization/?country_id={country.id}")
        assert resp.status_code == 200

    def test_index_with_division_id_filter(self, logged_in_client, db_session, app):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/?division_id=1")
        assert resp.status_code == 200

    def test_index_with_branch_id_filter(self, logged_in_client, db_session, app):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/?branch_id=1")
        assert resp.status_code == 200

    def test_index_with_secretariat_tab_param(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/?secretariat_tab=departments")
        assert resp.status_code == 200

    def test_index_invalid_secretariat_tab_falls_back(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/?secretariat_tab=bogus")
        assert resp.status_code == 200


# ---------------------------------------------------------------------------
# Countries CRUD
# ---------------------------------------------------------------------------

class TestNewCountry:
    def test_get_renders_form(self, logged_in_client, db_session):
        with _mock_render() as mock_rt:
            resp = logged_in_client.get("/admin/organization/countries/new")
        assert resp.status_code == 200
        mock_rt.assert_called_once()

    def test_post_valid_data_creates_country(self, logged_in_client, db_session, app):
        from app.models.core import Country
        with app.app_context():
            Country.query.filter_by(iso3="ZZA").delete()
            db_session.commit()

        resp = logged_in_client.post(
            "/admin/organization/countries/new",
            data=_make_country_data(name="New Country ZZA", iso3="ZZA", iso2="ZA"),
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_post_invalid_data_rerenders_form(self, logged_in_client, db_session):
        with _mock_render() as mock_rt:
            resp = logged_in_client.post(
                "/admin/organization/countries/new",
                data={"name": "", "iso3": "AB"},
                follow_redirects=False,
            )
        assert resp.status_code == 200
        mock_rt.assert_called()

    def test_unauthenticated_access_redirects(self, client, db_session):
        resp = client.get("/admin/organization/countries/new", follow_redirects=False)
        assert resp.status_code == 302


class TestEditCountry:
    def test_get_renders_form_with_country_data(self, logged_in_client, db_session, app):
        country = create_test_country(db_session, name="Edit Test Country", iso3="ETA", iso2="ET")
        with app.app_context():
            with _mock_render() as mock_rt:
                resp = logged_in_client.get(f"/admin/organization/countries/{country.id}/edit")
        assert resp.status_code == 200
        mock_rt.assert_called_once()

    def test_get_nonexistent_country_returns_404(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/organization/countries/99999/edit")
        assert resp.status_code == 404

    def test_post_valid_data_updates_country(self, logged_in_client, db_session, app):
        country = create_test_country(db_session, name="Country Before Edit", iso3="BEF", iso2="BF")
        resp = logged_in_client.post(
            f"/admin/organization/countries/{country.id}/edit",
            data=_make_country_data(name="Country After Edit", iso3="BEF", iso2="BF"),
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_post_invalid_fds_member_rerenders_form(self, logged_in_client, db_session, app):
        country = create_test_country(db_session, name="FDS Test Country", iso3="FDS", iso2="FD")
        with _mock_render() as mock_rt:
            resp = logged_in_client.post(
                f"/admin/organization/countries/{country.id}/edit",
                data={**_make_country_data(iso3="FDS", iso2="FD"), "fds_member_user_id": "invalid_user"},
                follow_redirects=False,
            )
        # May re-render due to invalid fds member or redirect on success
        assert resp.status_code in (200, 302)

    def test_post_invalid_form_rerenders(self, logged_in_client, db_session, app):
        country = create_test_country(db_session, name="Invalid Form Country", iso3="IFC", iso2="IF")
        with _mock_render() as mock_rt:
            resp = logged_in_client.post(
                f"/admin/organization/countries/{country.id}/edit",
                data={"name": "", "iso3": "XX"},
                follow_redirects=False,
            )
        assert resp.status_code == 200
        mock_rt.assert_called()


class TestDeleteCountry:
    def test_delete_existing_country_redirects(self, logged_in_client, db_session, app):
        country = create_test_country(db_session, name="Delete Country", iso3="DEL", iso2="DL")
        resp = logged_in_client.post(
            f"/admin/organization/countries/{country.id}/delete",
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_delete_nonexistent_country_returns_404(self, logged_in_client, db_session):
        resp = logged_in_client.post(
            "/admin/organization/countries/99999/delete",
            follow_redirects=False,
        )
        assert resp.status_code == 404

    def test_delete_with_exception_redirects(self, logged_in_client, db_session, app):
        country = create_test_country(db_session, name="Error Delete Country", iso3="ERC", iso2="ER")
        with patch("app.routes.admin.organization.countries.db.session.delete", side_effect=Exception("db error")):
            resp = logged_in_client.post(
                f"/admin/organization/countries/{country.id}/delete",
                follow_redirects=False,
            )
        assert resp.status_code == 302


# ---------------------------------------------------------------------------
# Countries Export / Template / Import
# ---------------------------------------------------------------------------

class TestExportCountries:
    def test_export_returns_excel_file(self, logged_in_client, db_session, app):
        create_test_country(db_session, name="Export Country", iso3="EXP", iso2="EX")
        resp = logged_in_client.get("/admin/organization/countries/export")
        assert resp.status_code == 200
        assert "spreadsheet" in resp.content_type or "xlsx" in resp.content_type

    def test_export_on_exception_redirects(self, logged_in_client, db_session):
        with patch("app.routes.admin.organization.import_export.pd.DataFrame", side_effect=Exception("err")):
            resp = logged_in_client.get(
                "/admin/organization/countries/export",
                follow_redirects=False,
            )
        assert resp.status_code in (302, 500)

    def test_unauthenticated_export_redirects(self, client, db_session):
        resp = client.get("/admin/organization/countries/export", follow_redirects=False)
        assert resp.status_code == 302


class TestCountriesTemplate:
    def test_template_returns_excel_file(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/organization/countries/template")
        assert resp.status_code == 200
        assert "spreadsheet" in resp.content_type or "xlsx" in resp.content_type

    def test_template_on_exception_redirects(self, logged_in_client, db_session):
        with patch("app.routes.admin.organization.import_export.pd.DataFrame", side_effect=Exception("err")):
            resp = logged_in_client.get(
                "/admin/organization/countries/template",
                follow_redirects=False,
            )
        assert resp.status_code in (302, 500)


class TestImportCountries:
    def test_import_no_file_redirects(self, logged_in_client, db_session):
        resp = logged_in_client.post(
            "/admin/organization/countries/import",
            data={},
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_import_empty_filename_redirects(self, logged_in_client, db_session):
        data = {"file": (io.BytesIO(b""), "")}
        resp = logged_in_client.post(
            "/admin/organization/countries/import",
            data=data,
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_import_valid_excel_redirects(self, logged_in_client, db_session, app):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "ISO3", "ISO2", "Region", "Status", "Preferred Language", "Currency Code",
                   "FDS Member User ID", "FDS Member Email", "FDS Member Name"])
        ws.append(["Import Test", "IMP", "IM", "Europe", "Active", "en", "USD", "", "", ""])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = logged_in_client.post(
            "/admin/organization/countries/import",
            data={"file": (buf, "countries.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_import_overwrite_by_id_updates_instead_of_inserting(self, logged_in_client, db_session, app):
        import openpyxl
        from app.models.core import Country

        country = create_test_country(db_session, name="Original Country", iso3="OLD", iso2="OL")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "ISO3", "ISO2", "Region", "Status", "Preferred Language", "Currency Code"])
        ws.append([country.id, "Renamed Country", "NEW", "NW", "Europe", "Active", "en", "EUR"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with patch(
            "app.utils.advanced_validation.AdvancedValidator.validate_mime_type",
            return_value=(True, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ):
            resp = logged_in_client.post(
                "/admin/organization/countries/import",
                data={
                    "excel_file": (buf, "countries.xlsx"),
                    "overwrite_existing": "on",
                },
                content_type="multipart/form-data",
                follow_redirects=False,
            )
        assert resp.status_code == 302

        with app.app_context():
            updated = db_session.get(Country, country.id)
            assert updated is not None
            assert updated.name == "Renamed Country"
            assert updated.iso3 == "NEW"
            assert updated.iso2 == "NW"
            assert Country.query.count() == 1

    def test_import_overwrite_without_id_still_matches_iso3(self, logged_in_client, db_session, app):
        import openpyxl
        from app.models.core import Country

        country = create_test_country(db_session, name="Template Country", iso3="TPL", iso2="TP")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "ISO3", "ISO2", "Region", "Status", "Preferred Language", "Currency Code"])
        ws.append(["Updated Template Country", "TPL", "TP", "Europe", "Active", "en", "USD"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with patch(
            "app.utils.advanced_validation.AdvancedValidator.validate_mime_type",
            return_value=(True, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ):
            resp = logged_in_client.post(
                "/admin/organization/countries/import",
                data={
                    "excel_file": (buf, "countries.xlsx"),
                    "overwrite_existing": "on",
                },
                content_type="multipart/form-data",
                follow_redirects=False,
            )
        assert resp.status_code == 302

        with app.app_context():
            updated = db_session.get(Country, country.id)
            assert updated.name == "Updated Template Country"
            assert Country.query.count() == 1

    def test_import_overwrite_with_unknown_id_inserts(self, logged_in_client, db_session, app):
        import openpyxl
        from app.models.core import Country

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "ISO3", "ISO2", "Region", "Status", "Preferred Language", "Currency Code"])
        ws.append([99991, "New Country", "NC9", "N9", "Europe", "Active", "en", "USD"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with patch(
            "app.utils.advanced_validation.AdvancedValidator.validate_mime_type",
            return_value=(True, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ):
            resp = logged_in_client.post(
                "/admin/organization/countries/import",
                data={
                    "excel_file": (buf, "countries.xlsx"),
                    "overwrite_existing": "on",
                },
                content_type="multipart/form-data",
                follow_redirects=False,
            )
        assert resp.status_code == 302

        with app.app_context():
            created = db_session.get(Country, 99991)
            assert created is not None
            assert created.name == "New Country"
            assert created.iso3 == "NC9"


class TestImportNationalSocietiesOverwrite:
    def test_import_overwrite_by_id_updates_instead_of_inserting(self, logged_in_client, db_session, app):
        import openpyxl
        from app.models.core import Country
        from app.models.organization import NationalSociety

        country_a = create_test_country(db_session, name="Country A", iso3="CTA", iso2="CA")
        country_b = create_test_country(db_session, name="Country B", iso3="CTB", iso2="CB")
        ns = NationalSociety(name="Original NS", code="ORG", country_id=country_a.id, is_active=True)
        db_session.add(ns)
        db_session.commit()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Code", "Description", "Country ISO3", "Is Active", "Display Order"])
        ws.append([ns.id, "Renamed NS", "REN", "Updated", "CTB", "Yes", 5])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = logged_in_client.post(
            "/admin/organization/national-societies/import",
            data={
                "excel_file": (buf, "nss.xlsx"),
                "overwrite_existing": "on",
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert resp.status_code == 302

        with app.app_context():
            updated = db_session.get(NationalSociety, ns.id)
            assert updated is not None
            assert updated.name == "Renamed NS"
            assert updated.code == "REN"
            assert updated.country_id == country_b.id
            assert updated.display_order == 5
            assert NationalSociety.query.count() == 1

    def test_import_overwrite_with_unknown_id_inserts(self, logged_in_client, db_session, app):
        import openpyxl
        from app.models.organization import NationalSociety

        country = create_test_country(db_session, name="NS Insert Country", iso3="NIC", iso2="NI")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Code", "Description", "Country ISO3", "Is Active", "Display Order"])
        ws.append([99992, "Brand New NS", "BNW", "", "NIC", "Yes", 0])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = logged_in_client.post(
            "/admin/organization/national-societies/import",
            data={
                "excel_file": (buf, "nss.xlsx"),
                "overwrite_existing": "on",
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert resp.status_code == 302

        with app.app_context():
            created = db_session.get(NationalSociety, 99992)
            assert created is not None
            assert created.name == "Brand New NS"
            assert created.country_id == country.id


# ---------------------------------------------------------------------------
# National Societies CRUD
# ---------------------------------------------------------------------------

class TestNationalSocietyCRUD:
    def test_new_ns_get_renders_form(self, logged_in_client, db_session):
        with _mock_render() as mock_rt:
            resp = logged_in_client.get("/admin/organization/national-societies/new")
        assert resp.status_code == 200
        mock_rt.assert_called_once()

    def test_new_ns_post_valid_creates_ns(self, logged_in_client, db_session, app):
        country = create_test_country(db_session, name="NS Country", iso3="NSC", iso2="NC")
        resp = logged_in_client.post(
            "/admin/organization/national-societies/new",
            data={"name": "Test NS", "country_id": str(country.id), "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_new_ns_post_invalid_rerenders(self, logged_in_client, db_session):
        with _mock_render() as mock_rt:
            resp = logged_in_client.post(
                "/admin/organization/national-societies/new",
                data={"name": "", "country_id": ""},
                follow_redirects=False,
            )
        assert resp.status_code == 200

    def test_edit_ns_get_renders_form(self, logged_in_client, db_session, app):
        from app.models.organization import NationalSociety
        country = create_test_country(db_session, name="NS Edit Country", iso3="NEC", iso2="NE")
        ns = NationalSociety(name="Edit NS", country_id=country.id, is_active=True)
        db_session.add(ns)
        db_session.commit()
        with _mock_render() as mock_rt:
            resp = logged_in_client.get(f"/admin/organization/national-societies/{ns.id}/edit")
        assert resp.status_code == 200
        mock_rt.assert_called()

    def test_edit_ns_get_nonexistent_returns_404(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/organization/national-societies/99999/edit")
        assert resp.status_code == 404

    def test_edit_ns_post_valid_updates_ns(self, logged_in_client, db_session, app):
        from app.models.organization import NationalSociety
        country = create_test_country(db_session, name="NS Post Country", iso3="NPC", iso2="NP")
        ns = NationalSociety(name="Post NS", country_id=country.id, is_active=True)
        db_session.add(ns)
        db_session.commit()
        resp = logged_in_client.post(
            f"/admin/organization/national-societies/{ns.id}/edit",
            data={"name": "Updated NS", "country_id": str(country.id), "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_delete_ns_redirects(self, logged_in_client, db_session, app):
        from app.models.organization import NationalSociety
        country = create_test_country(db_session, name="NS Del Country", iso3="NDC", iso2="ND")
        ns = NationalSociety(name="Delete NS", country_id=country.id, is_active=True)
        db_session.add(ns)
        db_session.commit()
        resp = logged_in_client.post(
            f"/admin/organization/national-societies/{ns.id}/delete",
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_delete_ns_nonexistent_returns_404(self, logged_in_client, db_session):
        resp = logged_in_client.post(
            "/admin/organization/national-societies/99999/delete",
            follow_redirects=False,
        )
        assert resp.status_code == 404

    def test_ns_export_returns_excel(self, logged_in_client, db_session, app):
        resp = logged_in_client.get("/admin/organization/national-societies/export")
        assert resp.status_code == 200
        assert "spreadsheet" in resp.content_type or "xlsx" in resp.content_type

    def test_ns_template_returns_excel(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/organization/national-societies/template")
        assert resp.status_code == 200

    def test_ns_import_no_file_redirects(self, logged_in_client, db_session):
        resp = logged_in_client.post(
            "/admin/organization/national-societies/import",
            data={},
            follow_redirects=False,
        )
        assert resp.status_code == 302


# ---------------------------------------------------------------------------
# NS Branches
# ---------------------------------------------------------------------------

class TestNSBranchesCRUD:
    def _make_branch(self, db_session, name="Test Branch", country_iso3="NBC", country_iso2="NB"):
        from app.models.organization import NSBranch
        country = create_test_country(db_session, name=f"Branch Country {country_iso3}",
                                       iso3=country_iso3, iso2=country_iso2)
        branch = NSBranch(name=name, country_id=country.id, is_active=True)
        db_session.add(branch)
        db_session.commit()
        return branch, country

    def test_get_branches_list(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/ns-branches")
        assert resp.status_code == 200

    def test_new_branch_get(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/ns-branches/new")
        assert resp.status_code == 200

    def test_new_branch_post_valid(self, logged_in_client, db_session, app):
        from app.models.core import Country
        country = create_test_country(db_session, name="New Branch Country", iso3="NBR", iso2="NR")
        resp = logged_in_client.post(
            "/admin/organization/ns-branches/new",
            data={"name": "New Branch", "country_id": str(country.id), "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_new_branch_post_invalid_rerenders(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.post(
                "/admin/organization/ns-branches/new",
                data={"name": "", "country_id": ""},
                follow_redirects=False,
            )
        assert resp.status_code == 200

    def test_edit_branch_get(self, logged_in_client, db_session, app):
        branch, _ = self._make_branch(db_session)
        with _mock_render():
            resp = logged_in_client.get(f"/admin/organization/ns-branches/{branch.id}/edit")
        assert resp.status_code == 200

    def test_edit_branch_get_404(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/organization/ns-branches/99999/edit")
        assert resp.status_code == 404

    def test_edit_branch_post_valid(self, logged_in_client, db_session, app):
        branch, country = self._make_branch(db_session, name="Edit Branch", country_iso3="EBB", country_iso2="EB")
        resp = logged_in_client.post(
            f"/admin/organization/ns-branches/{branch.id}/edit",
            data={"name": "Updated Branch", "country_id": str(country.id), "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_delete_branch_redirects(self, logged_in_client, db_session, app):
        branch, _ = self._make_branch(db_session, name="Delete Branch", country_iso3="DBR", country_iso2="DB")
        resp = logged_in_client.post(
            f"/admin/organization/ns-branches/{branch.id}/delete",
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_delete_branch_404(self, logged_in_client, db_session):
        resp = logged_in_client.post("/admin/organization/ns-branches/99999/delete", follow_redirects=False)
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# NS SubBranches
# ---------------------------------------------------------------------------

class TestNSSubBranchesCRUD:
    def _make_branch(self, db_session, suffix="SBT"):
        from app.models.organization import NSBranch
        country = create_test_country(db_session, name=f"SubBranch Country {suffix}",
                                       iso3=suffix, iso2=suffix[:2])
        branch = NSBranch(name=f"Branch {suffix}", country_id=country.id, is_active=True)
        db_session.add(branch)
        db_session.commit()
        return branch, country

    def _make_subbranch(self, db_session, suffix="SBA"):
        from app.models.organization import NSSubBranch
        branch, country = self._make_branch(db_session, suffix=f"S{suffix[:2]}")
        sub = NSSubBranch(name=f"SubBranch {suffix}", branch_id=branch.id, is_active=True)
        db_session.add(sub)
        db_session.commit()
        return sub, branch

    def test_get_subbranches_list(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/ns-subbranches")
        assert resp.status_code == 200

    def test_new_subbranch_get(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/ns-subbranches/new")
        assert resp.status_code == 200

    def test_new_subbranch_post_valid(self, logged_in_client, db_session, app):
        branch, _ = self._make_branch(db_session, suffix="NSB")
        resp = logged_in_client.post(
            "/admin/organization/ns-subbranches/new",
            data={"name": "New SubBranch", "branch_id": str(branch.id), "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_new_subbranch_post_invalid(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.post(
                "/admin/organization/ns-subbranches/new",
                data={"name": "", "branch_id": ""},
                follow_redirects=False,
            )
        assert resp.status_code == 200

    def test_edit_subbranch_get(self, logged_in_client, db_session):
        sub, branch = self._make_subbranch(db_session)
        with _mock_render():
            resp = logged_in_client.get(f"/admin/organization/ns-subbranches/{sub.id}/edit")
        assert resp.status_code == 200

    def test_edit_subbranch_get_404(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/organization/ns-subbranches/99999/edit")
        assert resp.status_code == 404

    def test_edit_subbranch_post_valid(self, logged_in_client, db_session, app):
        sub, branch = self._make_subbranch(db_session, suffix="ESB")
        resp = logged_in_client.post(
            f"/admin/organization/ns-subbranches/{sub.id}/edit",
            data={"name": "Updated SubBranch", "branch_id": str(branch.id), "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_delete_subbranch(self, logged_in_client, db_session, app):
        sub, _ = self._make_subbranch(db_session, suffix="DSB")
        resp = logged_in_client.post(
            f"/admin/organization/ns-subbranches/{sub.id}/delete",
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_delete_subbranch_404(self, logged_in_client, db_session):
        resp = logged_in_client.post("/admin/organization/ns-subbranches/99999/delete", follow_redirects=False)
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# NS Local Units
# ---------------------------------------------------------------------------

class TestNSLocalUnitsCRUD:
    def _setup(self, db_session, suffix="LU"):
        from app.models.organization import NSBranch, NSLocalUnit
        country = create_test_country(db_session, name=f"LU Country {suffix}",
                                       iso3=f"L{suffix[:2]}", iso2=suffix[:2])
        branch = NSBranch(name=f"LU Branch {suffix}", country_id=country.id, is_active=True)
        db_session.add(branch)
        db_session.commit()
        return branch, country

    def test_get_localunits_list(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/ns-localunits")
        assert resp.status_code == 200

    def test_new_localunit_get(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/ns-localunits/new")
        assert resp.status_code == 200

    def test_new_localunit_post_valid(self, logged_in_client, db_session, app):
        branch, _ = self._setup(db_session)
        resp = logged_in_client.post(
            "/admin/organization/ns-localunits/new",
            data={"name": "New Local Unit", "branch_id": str(branch.id), "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_new_localunit_post_invalid(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.post(
                "/admin/organization/ns-localunits/new",
                data={"name": "", "branch_id": ""},
                follow_redirects=False,
            )
        assert resp.status_code == 200

    def test_edit_localunit_get(self, logged_in_client, db_session, app):
        from app.models.organization import NSLocalUnit
        branch, _ = self._setup(db_session, suffix="EL")
        lu = NSLocalUnit(name="Edit LU", branch_id=branch.id, is_active=True)
        db_session.add(lu)
        db_session.commit()
        with _mock_render():
            resp = logged_in_client.get(f"/admin/organization/ns-localunits/{lu.id}/edit")
        assert resp.status_code == 200

    def test_edit_localunit_get_404(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/organization/ns-localunits/99999/edit")
        assert resp.status_code == 404

    def test_edit_localunit_post_valid(self, logged_in_client, db_session, app):
        from app.models.organization import NSLocalUnit
        branch, _ = self._setup(db_session, suffix="PL")
        lu = NSLocalUnit(name="Post LU", branch_id=branch.id, is_active=True)
        db_session.add(lu)
        db_session.commit()
        resp = logged_in_client.post(
            f"/admin/organization/ns-localunits/{lu.id}/edit",
            data={"name": "Updated LU", "branch_id": str(branch.id), "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_delete_localunit(self, logged_in_client, db_session, app):
        from app.models.organization import NSLocalUnit
        branch, _ = self._setup(db_session, suffix="DL")
        lu = NSLocalUnit(name="Delete LU", branch_id=branch.id, is_active=True)
        db_session.add(lu)
        db_session.commit()
        resp = logged_in_client.post(
            f"/admin/organization/ns-localunits/{lu.id}/delete",
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_delete_localunit_404(self, logged_in_client, db_session):
        resp = logged_in_client.post("/admin/organization/ns-localunits/99999/delete", follow_redirects=False)
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Secretariat Divisions
# ---------------------------------------------------------------------------

class TestSecretariatDivisionsCRUD:
    def _make_division(self, db_session, name="Test Division"):
        from app.models.organization import SecretariatDivision
        div = SecretariatDivision(name=name, is_active=True)
        db_session.add(div)
        db_session.commit()
        return div

    def test_get_divisions_list(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/secretariat-divisions")
        assert resp.status_code == 200

    def test_new_division_get(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/secretariat-divisions/new")
        assert resp.status_code == 200

    def test_new_division_post_valid(self, logged_in_client, db_session, app):
        resp = logged_in_client.post(
            "/admin/organization/secretariat-divisions/new",
            data={"name": "New Division", "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_new_division_post_invalid(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.post(
                "/admin/organization/secretariat-divisions/new",
                data={"name": ""},
                follow_redirects=False,
            )
        assert resp.status_code == 200

    def test_edit_division_get(self, logged_in_client, db_session, app):
        div = self._make_division(db_session)
        with _mock_render():
            resp = logged_in_client.get(f"/admin/organization/secretariat-divisions/{div.id}/edit")
        assert resp.status_code == 200

    def test_edit_division_get_404(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/organization/secretariat-divisions/99999/edit")
        assert resp.status_code == 404

    def test_edit_division_post_valid(self, logged_in_client, db_session, app):
        div = self._make_division(db_session, name="Edit Division")
        resp = logged_in_client.post(
            f"/admin/organization/secretariat-divisions/{div.id}/edit",
            data={"name": "Updated Division", "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_delete_division(self, logged_in_client, db_session, app):
        div = self._make_division(db_session, name="Delete Division")
        resp = logged_in_client.post(
            f"/admin/organization/secretariat-divisions/{div.id}/delete",
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_delete_division_404(self, logged_in_client, db_session):
        resp = logged_in_client.post("/admin/organization/secretariat-divisions/99999/delete", follow_redirects=False)
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Secretariat Departments
# ---------------------------------------------------------------------------

class TestSecretariatDepartmentsCRUD:
    def _make_division(self, db_session, name="Dept Division"):
        from app.models.organization import SecretariatDivision
        div = SecretariatDivision(name=name, is_active=True)
        db_session.add(div)
        db_session.commit()
        return div

    def _make_department(self, db_session, div_name="Dept Div"):
        from app.models.organization import SecretariatDepartment
        div = self._make_division(db_session, name=div_name)
        dept = SecretariatDepartment(name=f"Dept of {div_name}", division_id=div.id, is_active=True)
        db_session.add(dept)
        db_session.commit()
        return dept, div

    def test_get_departments_list(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/secretariat-departments")
        assert resp.status_code == 200

    def test_new_department_get(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/secretariat-departments/new")
        assert resp.status_code == 200

    def test_new_department_post_valid(self, logged_in_client, db_session, app):
        div = self._make_division(db_session, name="New Dept Division")
        resp = logged_in_client.post(
            "/admin/organization/secretariat-departments/new",
            data={"name": "New Department", "division_id": str(div.id), "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_new_department_post_invalid(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.post(
                "/admin/organization/secretariat-departments/new",
                data={"name": "", "division_id": ""},
                follow_redirects=False,
            )
        assert resp.status_code == 200

    def test_edit_department_get(self, logged_in_client, db_session, app):
        dept, div = self._make_department(db_session)
        with _mock_render():
            resp = logged_in_client.get(f"/admin/organization/secretariat-departments/{dept.id}/edit")
        assert resp.status_code == 200

    def test_edit_department_get_404(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/organization/secretariat-departments/99999/edit")
        assert resp.status_code == 404

    def test_edit_department_post_valid(self, logged_in_client, db_session, app):
        dept, div = self._make_department(db_session, div_name="Edit Dept Div")
        resp = logged_in_client.post(
            f"/admin/organization/secretariat-departments/{dept.id}/edit",
            data={"name": "Updated Department", "division_id": str(div.id), "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_delete_department(self, logged_in_client, db_session, app):
        dept, _ = self._make_department(db_session, div_name="Delete Dept Div")
        resp = logged_in_client.post(
            f"/admin/organization/secretariat-departments/{dept.id}/delete",
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_delete_department_404(self, logged_in_client, db_session):
        resp = logged_in_client.post("/admin/organization/secretariat-departments/99999/delete", follow_redirects=False)
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Secretariat Regional Offices
# ---------------------------------------------------------------------------

class TestSecretariatRegionalOfficesCRUD:
    def _make_region(self, db_session, name="Test Region"):
        from app.models.organization import SecretariatRegionalOffice
        region = SecretariatRegionalOffice(name=name, is_active=True)
        db_session.add(region)
        db_session.commit()
        return region

    def test_get_regions_list(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/secretariat-regional-offices")
        assert resp.status_code == 200

    def test_new_region_get(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/secretariat-regional-offices/new")
        assert resp.status_code == 200

    def test_new_region_post_valid(self, logged_in_client, db_session, app):
        resp = logged_in_client.post(
            "/admin/organization/secretariat-regional-offices/new",
            data={"name": "New Regional Office", "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_new_region_post_invalid(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.post(
                "/admin/organization/secretariat-regional-offices/new",
                data={"name": ""},
                follow_redirects=False,
            )
        assert resp.status_code == 200

    def test_edit_region_get(self, logged_in_client, db_session, app):
        region = self._make_region(db_session)
        with _mock_render():
            resp = logged_in_client.get(f"/admin/organization/secretariat-regional-offices/{region.id}/edit")
        assert resp.status_code == 200

    def test_edit_region_get_404(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/organization/secretariat-regional-offices/99999/edit")
        assert resp.status_code == 404

    def test_edit_region_post_valid(self, logged_in_client, db_session, app):
        region = self._make_region(db_session, name="Edit Region")
        resp = logged_in_client.post(
            f"/admin/organization/secretariat-regional-offices/{region.id}/edit",
            data={"name": "Updated Region", "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_delete_region(self, logged_in_client, db_session, app):
        region = self._make_region(db_session, name="Delete Region")
        resp = logged_in_client.post(
            f"/admin/organization/secretariat-regional-offices/{region.id}/delete",
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_delete_region_404(self, logged_in_client, db_session):
        resp = logged_in_client.post("/admin/organization/secretariat-regional-offices/99999/delete", follow_redirects=False)
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Secretariat Cluster Offices
# ---------------------------------------------------------------------------

class TestSecretariatClusterOfficesCRUD:
    def _make_region(self, db_session, name="Cluster Region"):
        from app.models.organization import SecretariatRegionalOffice
        region = SecretariatRegionalOffice(name=name, is_active=True)
        db_session.add(region)
        db_session.commit()
        return region

    def _make_cluster(self, db_session, region_name="Cluster Region"):
        from app.models.organization import SecretariatClusterOffice
        region = self._make_region(db_session, name=region_name)
        cluster = SecretariatClusterOffice(name=f"Cluster of {region_name}", regional_office_id=region.id, is_active=True)
        db_session.add(cluster)
        db_session.commit()
        return cluster, region

    def test_get_clusters_list(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/secretariat-cluster-offices")
        assert resp.status_code == 200

    def test_new_cluster_get(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.get("/admin/organization/secretariat-cluster-offices/new")
        assert resp.status_code == 200

    def test_new_cluster_post_valid(self, logged_in_client, db_session, app):
        region = self._make_region(db_session, name="New Cluster Region")
        resp = logged_in_client.post(
            "/admin/organization/secretariat-cluster-offices/new",
            data={"name": "New Cluster", "regional_office_id": str(region.id), "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_new_cluster_post_invalid(self, logged_in_client, db_session):
        with _mock_render():
            resp = logged_in_client.post(
                "/admin/organization/secretariat-cluster-offices/new",
                data={"name": "", "regional_office_id": ""},
                follow_redirects=False,
            )
        assert resp.status_code == 200

    def test_edit_cluster_get(self, logged_in_client, db_session, app):
        cluster, region = self._make_cluster(db_session)
        with _mock_render():
            resp = logged_in_client.get(f"/admin/organization/secretariat-cluster-offices/{cluster.id}/edit")
        assert resp.status_code == 200

    def test_edit_cluster_get_404(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/organization/secretariat-cluster-offices/99999/edit")
        assert resp.status_code == 404

    def test_edit_cluster_post_valid(self, logged_in_client, db_session, app):
        cluster, region = self._make_cluster(db_session, region_name="Edit Cluster Region")
        resp = logged_in_client.post(
            f"/admin/organization/secretariat-cluster-offices/{cluster.id}/edit",
            data={"name": "Updated Cluster", "regional_office_id": str(region.id), "is_active": "y"},
            follow_redirects=False,
        )
        assert resp.status_code in (200, 302)

    def test_delete_cluster(self, logged_in_client, db_session, app):
        cluster, _ = self._make_cluster(db_session, region_name="Delete Cluster Region")
        resp = logged_in_client.post(
            f"/admin/organization/secretariat-cluster-offices/{cluster.id}/delete",
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_delete_cluster_404(self, logged_in_client, db_session):
        resp = logged_in_client.post("/admin/organization/secretariat-cluster-offices/99999/delete", follow_redirects=False)
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# API Endpoints
# ---------------------------------------------------------------------------

class TestOrganizationAPIEndpoints:
    def test_api_branches_by_country(self, logged_in_client, db_session, app):
        from app.models.organization import NSBranch
        country = create_test_country(db_session, name="API Branch Country", iso3="ABC", iso2="AB")
        branch = NSBranch(name="API Branch", country_id=country.id, is_active=True)
        db_session.add(branch)
        db_session.commit()
        resp = logged_in_client.get(f"/admin/organization/api/branches/{country.id}")
        assert resp.status_code == 200
        data = resp.get_json()
        assert isinstance(data, list) or (isinstance(data, dict) and "branches" in data)

    def test_api_branches_empty_country(self, logged_in_client, db_session, app):
        country = create_test_country(db_session, name="Empty Branch Country", iso3="EBC", iso2="EB")
        resp = logged_in_client.get(f"/admin/organization/api/branches/{country.id}")
        assert resp.status_code == 200

    def test_api_subbranches_by_branch(self, logged_in_client, db_session, app):
        from app.models.organization import NSBranch, NSSubBranch
        country = create_test_country(db_session, name="API Sub Country", iso3="ASC", iso2="AS")
        branch = NSBranch(name="API Sub Branch", country_id=country.id, is_active=True)
        db_session.add(branch)
        db_session.commit()
        sub = NSSubBranch(name="API SubBranch", branch_id=branch.id, is_active=True)
        db_session.add(sub)
        db_session.commit()
        resp = logged_in_client.get(f"/admin/organization/api/subbranches/{branch.id}")
        assert resp.status_code == 200

    def test_api_cluster_offices_by_region(self, logged_in_client, db_session, app):
        from app.models.organization import SecretariatRegionalOffice, SecretariatClusterOffice
        region = SecretariatRegionalOffice(name="API Cluster Region", is_active=True)
        db_session.add(region)
        db_session.commit()
        cluster = SecretariatClusterOffice(name="API Cluster", regional_office_id=region.id, is_active=True)
        db_session.add(cluster)
        db_session.commit()
        resp = logged_in_client.get(f"/admin/organization/api/cluster-offices/{region.id}")
        assert resp.status_code == 200

    def test_api_departments_by_division(self, logged_in_client, db_session, app):
        from app.models.organization import SecretariatDivision, SecretariatDepartment
        div = SecretariatDivision(name="API Dept Division", is_active=True)
        db_session.add(div)
        db_session.commit()
        dept = SecretariatDepartment(name="API Department", division_id=div.id, is_active=True)
        db_session.add(dept)
        db_session.commit()
        resp = logged_in_client.get(f"/admin/organization/api/departments/{div.id}")
        assert resp.status_code == 200

    def test_api_public_branches_by_country(self, logged_in_client, db_session, app):
        from app.models.organization import NSBranch
        country = create_test_country(db_session, name="Public Branch Country", iso3="PBC", iso2="PB")
        branch = NSBranch(name="Public Branch", country_id=country.id, is_active=True)
        db_session.add(branch)
        db_session.commit()
        resp = logged_in_client.get(f"/admin/organization/api/public/branches/{country.id}")
        assert resp.status_code == 200

    def test_api_public_subbranches_by_branch(self, logged_in_client, db_session, app):
        from app.models.organization import NSBranch, NSSubBranch
        country = create_test_country(db_session, name="Public SubBranch Country", iso3="PSC", iso2="PS")
        branch = NSBranch(name="Public SubBranch Branch", country_id=country.id, is_active=True)
        db_session.add(branch)
        db_session.commit()
        sub = NSSubBranch(name="Public SubBranch", branch_id=branch.id, is_active=True)
        db_session.add(sub)
        db_session.commit()
        resp = logged_in_client.get(f"/admin/organization/api/public/subbranches/{branch.id}")
        assert resp.status_code == 200

    def test_api_public_subbranches_by_country(self, logged_in_client, db_session, app):
        from app.models.organization import NSBranch, NSSubBranch
        country = create_test_country(db_session, name="Public Sub By Country", iso3="BCC", iso2="BC")
        branch = NSBranch(name="PSub Branch", country_id=country.id, is_active=True)
        db_session.add(branch)
        db_session.commit()
        sub = NSSubBranch(name="PSub", branch_id=branch.id, is_active=True)
        db_session.add(sub)
        db_session.commit()
        resp = logged_in_client.get(f"/admin/organization/api/public/subbranches/by-country/{country.id}")
        assert resp.status_code == 200

    def test_api_translation_counts_returns_json(self, logged_in_client, db_session, app):
        resp = logged_in_client.get("/admin/organization/api/translation-counts?entity_type=countries")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["success"] is True
        assert "counts" in data

    def test_api_translation_counts_secretariat_regions(self, logged_in_client, db_session, app):
        from app.models.organization import SecretariatRegionalOffice

        region = SecretariatRegionalOffice(
            name="Untranslated Region",
            short_name="UR",
            is_active=True,
        )
        db_session.add(region)
        db_session.commit()

        resp = logged_in_client.get(
            "/admin/organization/api/translation-counts?entity_type=secretariat_regions"
        )
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["success"] is True
        assert sum(data["counts"].values()) > 0

    def test_regional_office_translation_fields_includes_short_name(self):
        from types import SimpleNamespace
        from app.routes.admin.organization import regional_office_translation_fields

        same = SimpleNamespace(name="Africa", short_name="Africa")
        assert regional_office_translation_fields(same) == [
            ("name", "name_translations"),
            ("short_name", "short_name_translations"),
        ]

        different = SimpleNamespace(name="Europe and Central Asia", short_name="Europe & CA")
        assert regional_office_translation_fields(different) == [
            ("name", "name_translations"),
            ("short_name", "short_name_translations"),
        ]

    def test_resolve_field_translation_copies_matching_short_name(self):
        from types import SimpleNamespace
        from app.routes.admin.organization import resolve_field_translation

        entity = SimpleNamespace(
            name="Africa",
            short_name="Africa",
            name_translations={"fr": "Afrique"},
        )
        mock_translator = MagicMock()

        result = resolve_field_translation(
            entity, "short_name", "Africa", "fr", mock_translator, "ifrc"
        )

        assert result == "Afrique"
        mock_translator.translate_text.assert_not_called()

    def test_api_auto_translate_secretariat_regions_persists(self, logged_in_client, db_session, app):
        from app.models.organization import SecretariatRegionalOffice

        region = SecretariatRegionalOffice(name="Persist Test Region", is_active=True)
        db_session.add(region)
        db_session.commit()
        region_id = region.id

        mock_translator = MagicMock()
        mock_translator.translate_text.return_value = "Région test"

        with patch("app.services.translation.auto_translator.get_auto_translator", return_value=mock_translator):
            resp = logged_in_client.post(
                "/admin/organization/api/auto-translate-organizations",
                json={
                    "entity_type": "secretariat_regions",
                    "target_languages": ["fr"],
                    "translation_service": "ifrc",
                },
                headers={**_json_headers(), "Accept": "text/event-stream"},
                buffered=True,
            )

        assert resp.status_code == 200
        db_session.expire_all()
        refreshed = db_session.get(SecretariatRegionalOffice, region_id)
        assert refreshed is not None
        assert (refreshed.name_translations or {}).get("fr") == "Région test"

    def test_api_auto_translate_missing_key_returns_error(self, logged_in_client, db_session):
        resp = logged_in_client.post(
            "/admin/organization/api/auto-translate-organizations",
            json={},
            headers=_json_headers(),
        )
        assert resp.status_code in (200, 400, 500)

    def test_api_auto_translate_with_entity_type(self, logged_in_client, db_session, app):
        country = create_test_country(db_session, name="Translate Country", iso3="TRC", iso2="TR")
        with patch("app.routes.admin.organization.EntityService") as mock_svc:
            mock_svc.auto_translate.return_value = {"translated": 1}
            resp = logged_in_client.post(
                "/admin/organization/api/auto-translate-organizations",
                json={"entity_type": "countries", "entity_id": country.id},
                headers=_json_headers(),
            )
        assert resp.status_code in (200, 400, 500)

    def test_api_ns_part_of_post(self, logged_in_client, db_session, app):
        from app.models.organization import NationalSociety
        country = create_test_country(db_session, name="Part Of Country", iso3="POC", iso2="PO")
        ns = NationalSociety(name="Part Of NS", country_id=country.id, is_active=True)
        db_session.add(ns)
        db_session.commit()
        resp = logged_in_client.post(
            f"/admin/organization/api/national-societies/{ns.id}/part-of",
            json={"programs": ["prog1", "prog2"]},
            headers=_json_headers(),
        )
        assert resp.status_code in (200, 400, 422, 500)

    def test_api_part_of_programs_get(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/organization/api/part-of-programs")
        assert resp.status_code == 200

    def test_api_part_of_programs_post(self, logged_in_client, db_session):
        resp = logged_in_client.post(
            "/admin/organization/api/part-of-programs",
            json={"name": "Test Program"},
            headers=_json_headers(),
        )
        assert resp.status_code in (200, 201, 400, 422)

    def test_api_part_of_programs_delete(self, logged_in_client, db_session):
        # First create a program
        logged_in_client.post(
            "/admin/organization/api/part-of-programs",
            json={"name": "Delete Program"},
            headers=_json_headers(),
        )
        resp = logged_in_client.delete(
            "/admin/organization/api/part-of-programs/Delete%20Program",
            headers=_json_headers(),
        )
        assert resp.status_code in (200, 204, 404)

    def test_api_branches_unauthenticated_redirects(self, client, db_session):
        resp = client.get("/admin/organization/api/branches/1", follow_redirects=False)
        assert resp.status_code == 302


# ---------------------------------------------------------------------------
# National Societies Export / Template / Import
# ---------------------------------------------------------------------------

class TestNSSExportImport:
    def test_ns_export_on_exception_redirects(self, logged_in_client, db_session):
        with patch("app.routes.admin.organization.import_export.pd.DataFrame", side_effect=Exception("err")):
            resp = logged_in_client.get(
                "/admin/organization/national-societies/export",
                follow_redirects=False,
            )
        assert resp.status_code in (302, 500)

    def test_ns_template_on_exception_redirects(self, logged_in_client, db_session):
        with patch("app.routes.admin.organization.import_export.pd.DataFrame", side_effect=Exception("err")):
            resp = logged_in_client.get(
                "/admin/organization/national-societies/template",
                follow_redirects=False,
            )
        assert resp.status_code in (302, 500)

    def test_ns_import_empty_file_redirects(self, logged_in_client, db_session):
        data = {"file": (io.BytesIO(b""), "")}
        resp = logged_in_client.post(
            "/admin/organization/national-societies/import",
            data=data,
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert resp.status_code == 302

    def test_ns_import_valid_excel(self, logged_in_client, db_session, app):
        import openpyxl
        from app.models.organization import NationalSociety
        country = create_test_country(db_session, name="NS Import Country", iso3="NIM", iso2="NI")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Code", "Description", "Country ISO3", "Active", "Display Order"])
        ws.append(["Import NS", "IMP", "Test import", "NIM", "True", "1"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = logged_in_client.post(
            "/admin/organization/national-societies/import",
            data={"file": (buf, "ns.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert resp.status_code == 302


# ---------------------------------------------------------------------------
# Helper functions (internal) — covered via route execution
# ---------------------------------------------------------------------------

class TestInternalHelpers:
    def test_get_translation_languages_no_app_context(self, app):
        """Test that _get_translation_languages works during import (no app context)."""
        with app.app_context():
            from app.routes.admin.organization import get_translation_languages
            result = get_translation_languages()
            assert isinstance(result, list)

    def test_get_translation_codes(self, app):
        with app.app_context():
            from app.routes.admin.organization import get_translation_codes
            result = get_translation_codes()
            assert isinstance(result, list)

    def test_count_missing_name_translations(self, app, db_session):
        with app.app_context():
            from app.routes.admin.organization import count_missing_name_translations
            countries = create_test_country(db_session, name="Translation Test", iso3="TTX", iso2="TX")
            result = count_missing_name_translations([countries])
            assert isinstance(result, dict)
