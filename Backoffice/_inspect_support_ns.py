import openpyxl
from openpyxl.utils import range_boundaries

path = r"app/static/templates/unified_country_plan.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)
ws = wb["Planned Bilateral Support"]
tbl = ws.tables["Data_Support"]
ref = tbl.ref if hasattr(tbl, "ref") else tbl
min_col, min_row, max_col, max_row = range_boundaries(ref)
headers = [ws.cell(min_row, c).value for c in range(min_col, max_col + 1)]
print("headers:", headers)
ns_col = None
for i, h in enumerate(headers):
    if h == "NS":
        ns_col = min_col + i
        break
print("NS col:", ns_col)
for row in range(min_row + 1, min_row + 8):
    cell = ws.cell(row, ns_col)
    print(f"row{row} NS value={cell.value!r} formula={getattr(cell, '_value', None)}")

# what read_named_table returns
import os, sys
IMPORTS = os.path.join(os.path.dirname(__file__), "scripts", "imports")
sys.path.insert(0, IMPORTS)
from unified_country_plan_excel_template import read_named_table, SUPPORT_SHEET, SUPPORT_TABLE
_, rows = read_named_table(wb, SUPPORT_SHEET, SUPPORT_TABLE)
print("read_named_table row count:", len(rows))
if rows:
    print("first row:", rows[0])
    print("row with NS:", next((r for r in rows if r.get("NS")), None))
wb.close()
