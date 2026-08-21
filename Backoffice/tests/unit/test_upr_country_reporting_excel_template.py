"""Unit tests for UPR Country Reporting Excel template parsing helpers."""

from __future__ import annotations

import os
import sys

import pytest

BACKOFFICE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
IMPORTS_DIR = os.path.join(BACKOFFICE_DIR, "scripts", "imports")
if IMPORTS_DIR not in sys.path:
    sys.path.insert(0, IMPORTS_DIR)

from upr_country_reporting_excel_template import (  # noqa: E402
    INDICATOR_APPLICABLE_VALUE,
    INDICATOR_DNA_VALUE,
    INDICATOR_DNA_HEADER,
    INDICATOR_ID_HEADER,
    INDICATOR_MATCH_THRESHOLD,
    SP_EF_HEADER,
    INDICATOR_HEADER,
    _build_bank_id_row_locations,
    _build_indicator_row_index,
    _build_kpi_display_map,
    _build_upr_country_reporting_disagg_header_maps,
    _disagg_payload_to_workbook_cells,
    _find_row_for_form_item,
    _find_row_in_indicator_table,
    _indicator_similarity,
    _bank_id_row_integrity_warning,
    _merge_non_binary_into_unknown_breakdown,
    _parse_support_matrix_ticks,
    _matrix_cell_is_set,
    _normalize_matrix_cells,
    _export_funding_breakdown,
    _entity_ids_from_matrix_disagg,
    _parse_support_matrix_ns_ids,
    _matrix_cell_scalar,
    _parse_emergency_selection_from_entry,
    _format_emergency_operation_display,
    _resolve_workbook_emergency_slot_metadata,
    _upsert_emergency_repeat_choice,
    _resolve_workbook_indicator_bank_id,
    _workbook_yes_no_value,
    _resolve_indicator_import_value,
    _is_percentage_indicator_type,
    _iter_numeric_leaves,
    _percentage_range_warning,
    _yes_no_value_is_applicable,
    _write_indicator_entry,
    read_table_cell,
    _collect_workbook_indicator_bank_ids,
    _reporting_funding_matrix_column,
    _entry_is_yes_no,
    _is_yes_no_indicator_type,
    _parse_workbook_row_disagg,
    _disagg_consistency_warning,
    _extract_disagg_components,
    import_rows_to_client_payload,
    dedupe_upr_import_warnings,
    build_kpi_lookup,
    parse_comments,
    parse_emergency_slot_metadata,
    parse_funding,
    parse_indicators,
    parse_ns_key_data,
    parse_version,
    period_to_workbook_version,
    _apply_reporting_assignment_label,
    _assignment_display_label,
    _quiet_openpyxl_io,
    ASSIGNMENT_LABEL_NAMED_CELL,
    read_named_cell,
    read_named_table,
    write_table_cell,
    validate_upr_country_reporting_import_file,
    _table_data_row_capacity,
    DATA_OTHER_SHEET,
    DATA_OTHER_TABLE,
)


TEMPLATE_PATH = os.path.join(
    BACKOFFICE_DIR,
    "app",
    "static",
    "templates",
    "unified_country_report.xlsx",
)


@pytest.fixture(scope="module")
def upr_country_reporting_workbook():
    if not os.path.isfile(TEMPLATE_PATH):
        pytest.skip("UPR Country Reporting template file not present")
    import openpyxl

    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    yield wb
    wb.close()


def test_period_to_workbook_version():
    assert period_to_workbook_version("Jan-Jun 2026") == "MYR26.V1.0"
    assert period_to_workbook_version("2025") == "AR25.V1.0"


def test_parse_version_from_template(upr_country_reporting_workbook):
    round_code, period = parse_version(upr_country_reporting_workbook)
    assert round_code == "MYR26"
    assert period == "Jan-Jun 2026"


def test_build_kpi_lookup_has_entries(upr_country_reporting_workbook):
    lookup = build_kpi_lookup(upr_country_reporting_workbook)
    assert len(lookup) > 100
    sample = next(iter(lookup.values()))
    assert isinstance(sample, int)


def test_read_named_table_data_core(upr_country_reporting_workbook):
    headers, rows = read_named_table(upr_country_reporting_workbook, "Overall action Indicators", "Data_core")
    assert INDICATOR_ID_HEADER in headers
    assert "Indicator" in headers
    assert len(rows) > 5
    assert rows[0].get(INDICATOR_ID_HEADER) is not None


def test_parse_indicators_returns_rows(upr_country_reporting_workbook):
    rows = parse_indicators(upr_country_reporting_workbook)
    assert rows
    assert all("indicator" in row and "sp_ef" in row for row in rows)
    assert all(row.get("bank_id") is not None for row in rows[:5])


def test_find_row_for_form_items_by_bank_id(upr_country_reporting_workbook):
    import openpyxl

    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    bank_locations = _build_bank_id_row_locations(wb)
    core_index = _build_indicator_row_index(wb, tables=(("Overall action Indicators", "Data_core"),))
    kpi = build_kpi_lookup(wb)
    loc = _find_row_for_form_item(
        section_name="Cross Cutting",
        label="Number of people reached with emergency response and early recovery programmes.",
        bank_id=619,
        indicator_row_index=core_index,
        kpi_lookup=kpi,
        bank_id_locations=bank_locations,
    )
    assert loc is not None
    assert loc[1] == "Data_core"
    wb.close()


def test_parse_funding_structure(upr_country_reporting_workbook):
    funding = parse_funding(upr_country_reporting_workbook)
    assert "sources" in funding
    assert "breakdown" in funding
    assert "funding_column" in funding


def test_parse_comments_empty_by_default(upr_country_reporting_workbook):
    assert parse_comments(upr_country_reporting_workbook) == ""


def test_export_import_comments_single_cell():
    import openpyxl

    from upr_country_reporting_excel_template import COMMENTS_NAMED_CELL, write_named_cell

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    write_named_cell(wb, COMMENTS_NAMED_CELL, "Notes from the NS focal point.")
    assert parse_comments(wb) == "Notes from the NS focal point."
    wb.close()


def test_parse_ns_key_data_reads_named_cells(upr_country_reporting_workbook):
    ns = parse_ns_key_data(upr_country_reporting_workbook)
    assert set(ns.keys()) == {"volunteers", "staff", "local_units", "branches"}


def test_read_named_cell_version(upr_country_reporting_workbook):
    version = read_named_cell(upr_country_reporting_workbook, "Version")
    assert str(version).startswith("MYR")


def test_data_other_table_has_overflow_capacity(upr_country_reporting_workbook):
    capacity = _table_data_row_capacity(upr_country_reporting_workbook, DATA_OTHER_SHEET, DATA_OTHER_TABLE)
    assert capacity >= 4


def test_find_row_for_form_items_in_data_core(upr_country_reporting_workbook):
    """Core T33 indicators with an explicit ID cell in the workbook should
    resolve to their Data_core row via the direct bank-id lookup (the primary,
    intended path — every currently-published T33 indicator item has an
    explicit ID cell in the real template; fuzzy text matching is only a
    fallback for corrupted/edited files, see test below).
    """
    import openpyxl

    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    kpi = build_kpi_lookup(wb)
    index = _build_indicator_row_index(wb)
    bank_locations = _build_bank_id_row_locations(wb)
    samples = [
        ("Cross Cutting", "Number of people reached with emergency response and early recovery programmes.", 619),
    ]
    for section, label, bank_id in samples:
        loc = _find_row_for_form_item(
            section_name=section,
            label=label,
            bank_id=bank_id,
            indicator_row_index=index,
            kpi_lookup=kpi,
            bank_id_locations=bank_locations,
        )
        assert loc is not None, f"No row for {section!r} bank={bank_id}"
    wb.close()


def test_find_row_for_form_item_falls_back_to_fuzzy_text_match():
    """When a form item's bank id has no explicit ID cell anywhere in the
    workbook (e.g. the row was pasted without its ID, or the item is newer
    than the last template regeneration), _find_row_for_form_item should
    still locate the right row by matching its (section, label) text against
    kpi_lookup and the workbook's own indicator_row_index — as long as the
    text is a close, high-confidence match, not merely a superficially
    similar one (see the "conditional" bank_id in the low-confidence sample).
    """
    kpi_lookup = {
        ("health", "number of people vaccinated against measles"): 501,
    }
    indicator_row_index = {
        ("health", "number of people vaccinated against measles"): ("Sheet1", "Data_core", 3),
    }
    loc = _find_row_for_form_item(
        section_name="Health",
        label="Number of people vaccinated against measles",
        bank_id=501,
        indicator_row_index=indicator_row_index,
        kpi_lookup=kpi_lookup,
        bank_id_locations={},
    )
    assert loc == ("Sheet1", "Data_core", 3)

    # Two genuinely different indicators that only share a long generic
    # opening phrase must NOT match, even though a naive character-level
    # ratio alone would score them ~0.72 (see test_indicator_similarity_*
    # below) — this is the real-world pair ("long-term services" vs
    # "emergency response") that motivated tightening _indicator_similarity.
    other_index = {
        (
            "cross cutting",
            "number of people reached with long-term services and programmes",
        ): ("Sheet1", "Data_core", 0),
    }
    loc_wrong = _find_row_for_form_item(
        section_name="Cross Cutting",
        label="Number of people reached with emergency response and early recovery programmes",
        bank_id=619,
        indicator_row_index=other_index,
        kpi_lookup={
            (
                "cross cutting",
                "number of people reached with emergency response and early recovery programmes",
            ): 619,
        },
        bank_id_locations={},
    )
    assert loc_wrong is None


def test_indicator_match_threshold_allows_truncated_excel_labels():
    assert INDICATOR_MATCH_THRESHOLD <= 0.65


def test_indicator_similarity_exact_and_truncated_match():
    assert _indicator_similarity("Number of shelters built", "Number of shelters built") == 1.0
    assert _indicator_similarity("Number of shelters built.", "Number of shelters built") == 1.0
    # Genuine truncation/abbreviation: the whole shorter label is a prefix of the longer one.
    assert _indicator_similarity(
        "Number of shelters built",
        "Number of shelters built in the reporting period",
    ) == 0.8


def test_indicator_similarity_rejects_short_generic_substring_fragment():
    """A short/generic leftover fragment (a stray word left behind by a
    partially cleared or mis-pasted cell) must NOT win the prefix/substring
    shortcut just because it happens to be a literal substring of some
    unrelated, much longer KPI sentence — that would score an automatic 0.8
    and silently win a blank-ID row's fuzzy match against the wrong
    indicator. Genuine truncations (see test above) are still full phrases
    with real content, not a single word, so they clear the minimums fine.
    """
    score = _indicator_similarity(
        "Support",
        "Number of national society branches receiving direct technical support",
    )
    assert score < INDICATOR_MATCH_THRESHOLD


def test_indicator_similarity_rejects_shared_boilerplate_phrase():
    """Real regression: two completely different KPIs that only share a long
    generic opening ("Number of people reached with ...") must NOT be
    considered similar just because a naive character-ratio is fooled by the
    shared boilerplate — the distinctive words carry no overlap at all here.
    """
    score = _indicator_similarity(
        "Number of people reached with long-term services and programmes",
        "Number of people reached with emergency response and early recovery programmes",
    )
    assert score < INDICATOR_MATCH_THRESHOLD


def test_indicator_similarity_high_word_overlap_still_matches():
    """Sanity check: genuinely close rewordings (same distinctive words, minor
    phrasing differences) must still clear the threshold — the fix must not
    over-correct into rejecting everything that isn't a verbatim match.
    """
    score = _indicator_similarity(
        "Number of people reached with mental health and psychosocial services",
        "Number of people reached with psychosocial and mental health services",
    )
    assert score >= INDICATOR_MATCH_THRESHOLD


def test_bank_id_row_integrity_warning_flags_genuine_mismatch():
    kpi_display = {
        619: ("Cross-cutting", "Number of people reached with emergency response and early recovery programmes."),
    }
    row = {
        "sp_ef": "Cross-cutting",
        "indicator": "Number of people reached with long-term services and programmes",
    }
    warning = _bank_id_row_integrity_warning(row, 619, kpi_display)
    assert warning is not None
    assert "619" in warning


def test_bank_id_row_integrity_warning_ignores_section_only_difference():
    """The same indicator can legitimately be grouped under a different SP/EF
    in the reporting tables than its single canonical entry in the master
    list (real observed case: bank id 711 files under 'Accountability and
    agility' in Data_core but 'Respect - Values, power and inclusion' in the
    master KPI list) — only the indicator NAME should be checked, not section.
    """
    kpi_display = {
        711: ("Respect - Values, power and inclusion", "Shared indicator text"),
    }
    row = {"sp_ef": "Accountability and agility", "indicator": "Shared indicator text"}
    assert _bank_id_row_integrity_warning(row, 711, kpi_display) is None


def test_bank_id_row_integrity_warning_no_reference_data_is_silent():
    """If the id isn't in the reference map at all (e.g. deleted from the DB,
    or DB lookup failed), don't guess — stay silent rather than risk a false
    positive against data we can't actually verify.
    """
    assert _bank_id_row_integrity_warning({"sp_ef": "X", "indicator": "Y"}, 999999, {}) is None
    assert _bank_id_row_integrity_warning({"sp_ef": "X", "indicator": "Y"}, 999999, None) is None


def test_parse_workbook_row_disagg_total_only_uses_total_key():
    row = {"Total Direct": 423.0}
    disagg = _parse_workbook_row_disagg(row)
    assert disagg == {"mode": "total", "values": {"total": 423.0}}


def test_parse_workbook_row_disagg_total_with_indirect_uses_direct_key():
    row = {"Total Direct": 100.0, "Indirectly reached": 25.0}
    disagg = _parse_workbook_row_disagg(row)
    assert disagg == {"mode": "total", "values": {"direct": 100.0, "indirect": 25.0}}


def test_parse_workbook_row_disagg_sex_age(upr_country_reporting_workbook):
    _, rows = read_named_table(upr_country_reporting_workbook, "Overall action Indicators", "Data_core")
    sample = dict(rows[0])
    sample["Male <5"] = 10
    sample["Male 5-17"] = 20
    sample["Female <5"] = 5
    sample["Indirectly reached"] = 3
    disagg = _parse_workbook_row_disagg(sample)
    assert disagg is not None
    assert disagg["mode"] == "sex_age"
    # "<5" slugifies to "_5" (the "<" becomes "_"), so the real sex-age key is
    # "male__5" (double underscore) — matching processing_service.slugify_age_group
    # and the entry-form field name "indicator_{id}_sexage_male__5".
    assert disagg["values"]["direct"]["male__5"] == 10
    assert disagg["values"]["direct"]["male_5_17"] == 20
    assert disagg["values"]["direct"]["female__5"] == 5
    assert disagg["values"]["indirect"] == 3


def test_parse_workbook_row_disagg_other_unknown_goes_to_form_unknown():
    disagg = _parse_workbook_row_disagg(
        {"Total Male": 10, "Total Female": 20, "Other/Unknown": 7}
    )
    assert disagg == {"mode": "sex", "values": {"male": 10, "female": 20, "unknown": 7}}


def test_parse_workbook_row_disagg_sex_age_keeps_other_unknown_on_unknown_row():
    disagg = _parse_workbook_row_disagg(
        {"Male <5": 10, "Female <5": 5, "Other/Unknown": 7, "Indirectly reached": 3}
    )
    assert disagg is not None
    assert disagg["mode"] == "sex_age"
    assert disagg["values"]["direct"]["male__5"] == 10
    assert disagg["values"]["direct"]["female__5"] == 5
    assert disagg["values"]["direct"]["unknown_unknown"] == 7
    assert disagg["values"]["indirect"] == 3


def test_disagg_consistency_warning_none_when_only_sex_age_present():
    row = {"Male <5": 10, "Female <5": 5}
    assert _disagg_consistency_warning(row, indicator_label="X") is None


def test_disagg_consistency_warning_none_when_only_total_present():
    row = {"Total Direct": 423.0}
    assert _disagg_consistency_warning(row, indicator_label="X") is None


def test_disagg_consistency_warning_none_when_sex_age_matches_sex_totals():
    row = {"Male <5": 10, "Male 5-17": 20, "Total Male": 30, "Total Female": 0}
    assert _disagg_consistency_warning(row, indicator_label="X") is None


def test_disagg_consistency_warning_none_when_other_unknown_complements_sex_age():
    row = {
        "Male <5": 10,
        "Female <5": 5,
        "Other/Unknown": 7,
        "Total Male": 10,
        "Total Female": 5,
        "Total Direct": 22,
    }
    assert _disagg_consistency_warning(row, indicator_label="X") is None


def test_disagg_consistency_warning_fires_when_sex_age_disagrees_with_sex_totals():
    """A row that has BOTH a full sex+age breakdown AND separate Total Male/Female
    values that don't match is a strong signal of a moved/stale/duplicated cell —
    _parse_workbook_row_disagg() would silently import the sex_age breakdown and
    drop the mismatched totals, so this must be surfaced as a warning instead.
    """
    row = {"Male <5": 10, "Male 5-17": 20, "Total Male": 999, "Total Female": 0}
    warning = _disagg_consistency_warning(row, indicator_label="Some indicator", context_label="PAK")
    assert warning is not None
    assert "Some indicator" in warning
    assert "PAK" in warning
    assert "sex+age" in warning
    assert "Total Male/Female" in warning


def test_disagg_consistency_warning_fires_when_sex_age_disagrees_with_direct_total():
    row = {"Male <5": 10, "Male 5-17": 20, "Total Direct": 999}
    warning = _disagg_consistency_warning(row, indicator_label="Some indicator")
    assert warning is not None
    assert "Total Direct" in warning


def test_disagg_consistency_warning_fires_when_sex_disagrees_with_direct_total():
    row = {"Total Male": 345, "Total Female": 534, "Total Direct": 999}
    warning = _disagg_consistency_warning(row, indicator_label="Some indicator")
    assert warning is not None
    assert "Male/Female" in warning
    assert "Total Direct" in warning


def test_disagg_consistency_warning_none_when_sex_matches_direct_total():
    row = {"Total Male": 345, "Total Female": 534, "Total Direct": 879}
    assert _disagg_consistency_warning(row, indicator_label="X") is None


def test_disagg_consistency_warning_tolerates_rounding_noise():
    row = {"Total Male": 345, "Total Female": 534.2, "Total Direct": 879}
    assert _disagg_consistency_warning(row, indicator_label="X") is None


def test_extract_disagg_components_ignores_id_and_underscore_keys():
    row = {"ID": 42, "_internal": 1, "Total Direct": 10}
    sex_age, sex, direct_total, indirect = _extract_disagg_components(row)
    assert sex_age == {}
    assert sex == {}
    assert direct_total == 10
    assert indirect is None


def test_disagg_payload_roundtrip_to_excel_headers(upr_country_reporting_workbook):
    key_to_header, _ = _build_upr_country_reporting_disagg_header_maps(upr_country_reporting_workbook)
    payload = {
        "mode": "total",
        "values": {"direct": 100, "indirect": 25},
    }
    cells = _disagg_payload_to_workbook_cells(payload, key_to_header)
    assert key_to_header["direct"] in cells
    assert cells[key_to_header["direct"]] == 100
    assert key_to_header["indirect"] in cells
    assert cells[key_to_header["indirect"]] == 25
    combined_header = key_to_header.get("combined")
    assert combined_header not in cells


def test_disagg_total_only_writes_total_direct(upr_country_reporting_workbook):
    key_to_header, _ = _build_upr_country_reporting_disagg_header_maps(upr_country_reporting_workbook)
    cells = _disagg_payload_to_workbook_cells({"mode": "total", "values": {"total": 42}}, key_to_header)
    assert cells[key_to_header["direct"]] == 42
    assert key_to_header.get("combined") not in cells


def test_disagg_by_sex_sums_total_direct(upr_country_reporting_workbook):
    key_to_header, _ = _build_upr_country_reporting_disagg_header_maps(upr_country_reporting_workbook)
    cells = _disagg_payload_to_workbook_cells(
        {"mode": "sex", "values": {"male": 345, "female": 534, "unknown": 555}},
        key_to_header,
    )
    assert cells[key_to_header["direct"]] == 1434
    assert cells[key_to_header["male"]] == 345
    assert cells[key_to_header["female"]] == 534
    assert cells[key_to_header["unknown"]] == 555


def test_disagg_by_sex_merges_non_binary_into_unknown(upr_country_reporting_workbook):
    key_to_header, _ = _build_upr_country_reporting_disagg_header_maps(upr_country_reporting_workbook)
    cells = _disagg_payload_to_workbook_cells(
        {
            "mode": "sex",
            "values": {"male": 345, "female": 534, "non_binary": 12, "unknown": 555},
        },
        key_to_header,
    )
    assert cells[key_to_header["unknown"]] == 567
    assert cells[key_to_header["male"]] == 345
    assert cells[key_to_header["female"]] == 534
    assert cells[key_to_header["direct"]] == 1446


def test_merge_non_binary_into_unknown_breakdown():
    merged = _merge_non_binary_into_unknown_breakdown(
        {
            "male": 1,
            "non_binary": 2,
            "unknown": 3,
            "non_binary_5_17": 4,
            "unknown_5_17": 5,
            "male_unknown": 6,
        }
    )
    assert merged["male"] == 1
    assert merged["male_unknown"] == 6
    assert merged["unknown"] == 14


def test_disagg_by_sex_age_merges_non_binary_and_unknown_rows(upr_country_reporting_workbook):
    key_to_header, _ = _build_upr_country_reporting_disagg_header_maps(upr_country_reporting_workbook)
    cells = _disagg_payload_to_workbook_cells(
        {
            "mode": "sex_age",
            "values": {
                "direct": {
                    "male__5": 10,
                    "female__5": 5,
                    "non_binary__5": 2,
                    "non_binary_5_17": 3,
                    "unknown__5": 4,
                    "unknown_5_17": 6,
                }
            },
        },
        key_to_header,
    )
    assert cells[key_to_header["unknown"]] == 15
    assert cells[key_to_header["male__5"]] == 10
    assert cells[key_to_header["female__5"]] == 5
    assert cells[key_to_header["direct"]] == 30


def test_disagg_by_sex_age_sums_total_direct(upr_country_reporting_workbook):
    key_to_header, _ = _build_upr_country_reporting_disagg_header_maps(upr_country_reporting_workbook)
    cells = _disagg_payload_to_workbook_cells(
        {
            "mode": "sex_age",
            "values": {"direct": {"male__5": 10, "female__5": 5, "male_5_17": 20}, "indirect": 3},
        },
        key_to_header,
    )
    assert cells[key_to_header["direct"]] == 35
    assert cells[key_to_header["indirect"]] == 3


def test_disagg_key_to_header_uses_double_underscore_for_under5(upr_country_reporting_workbook):
    """Regression test for a real data-loss bug: the "<5" age-group label slugifies
    to "_5" (processing_service.slugify_age_group replaces "<" with "_"), so the
    entry form's actual field name / disagg JSON key is "male__5" (double
    underscore), not "male_5". A single-underscore key here would silently orphan
    every "<5" sex-age value on both import and export.
    """
    key_to_header, _ = _build_upr_country_reporting_disagg_header_maps(upr_country_reporting_workbook)
    assert key_to_header["male__5"] == "Male <5"
    assert key_to_header["female__5"] == "Female <5"
    assert "male_5" not in key_to_header
    assert "female_5" not in key_to_header

    cells = _disagg_payload_to_workbook_cells(
        {"mode": "sex_age", "values": {"direct": {"male__5": 10, "female__5": 5}}},
        key_to_header,
    )
    assert cells[key_to_header["direct"]] == 15


def test_applicable_status_constants():
    assert INDICATOR_APPLICABLE_VALUE == "Applicable"
    assert INDICATOR_DNA_VALUE == "Data not available"


def test_emergency_row_lookup_scoped_to_target_table(upr_country_reporting_workbook):
    kpi_lookup = build_kpi_lookup(upr_country_reporting_workbook)
    bank_id_locations = _build_bank_id_row_locations(upr_country_reporting_workbook)
    indicator_row_index = _build_indicator_row_index(upr_country_reporting_workbook)

    # Bank 754 exists in both Data_core and emergency tables — lookup must stay on EA1.
    loc = _find_row_in_indicator_table(
        sheet_name="Emergency Appeal 1",
        table_name="Data_emergency1",
        bank_id=754,
        bank_id_locations=bank_id_locations,
        indicator_row_index=indicator_row_index,
        kpi_lookup=kpi_lookup,
    )
    assert loc is not None
    assert loc[0] == "Emergency Appeal 1"
    assert loc[1] == "Data_emergency1"

    # Bank 611 is not in the emergency template — must not fall back to Data_core.
    loc_other = _find_row_in_indicator_table(
        sheet_name="Emergency Appeal 2",
        table_name="Data_emergency2",
        bank_id=611,
        bank_id_locations=bank_id_locations,
        indicator_row_index=indicator_row_index,
        kpi_lookup=kpi_lookup,
    )
    assert loc_other is None


def test_kpi_display_map_has_labels(upr_country_reporting_workbook):
    display = _build_kpi_display_map(upr_country_reporting_workbook)
    kpi_lookup = build_kpi_lookup(upr_country_reporting_workbook)
    bank_id = next(iter(kpi_lookup.values()))
    assert bank_id in display
    sp_ef, indicator = display[bank_id]
    assert sp_ef
    assert indicator


def test_parse_emergency_selection_from_entry_disagg():
    entry = type("Entry", (), {"disagg_data": {"name": "Afghanistan - Earthquake", "code": "MDRAF019"}, "value": None})()
    meta = _parse_emergency_selection_from_entry(entry)
    assert meta == {
        "code": "MDRAF019",
        "name": "Afghanistan - Earthquake",
        "label": "Afghanistan - Earthquake (MDRAF019)",
    }


def test_parse_emergency_selection_from_entry_display_value():
    entry = type("Entry", (), {"disagg_data": None, "value": "Appeal Name (MDR001)"})()
    meta = _parse_emergency_selection_from_entry(entry)
    assert meta["code"] == "MDR001"
    assert meta["name"] == "Appeal Name"


def test_format_emergency_operation_display():
    assert _format_emergency_operation_display("Afghanistan - Earthquake", "MDRAF019") == (
        "Afghanistan - Earthquake (MDRAF019)"
    )
    assert _format_emergency_operation_display("Appeal", "") == "Appeal"
    assert _format_emergency_operation_display("", "MDR001") == "MDR001"


def test_resolve_workbook_emergency_slot_metadata_uses_go_api_when_code_matches():
    """A retyped-but-correct MDR code should resolve to the canonical GO API name,
    self-healing minor Excel typos in the appeal name column."""
    import openpyxl

    from import_upr_excel_data import UprImportContext
    from upr_country_reporting_excel_template import write_named_cell

    if not os.path.isfile(TEMPLATE_PATH):
        pytest.skip("UPR Country Reporting template file not present")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    try:
        write_named_cell(wb, "Data_MDR1", "MDRNG041")
        write_named_cell(wb, "Data_EO1", "Nigeria Floods EA (typo'd Excel name)")

        ctx = UprImportContext(template_ids=[33])
        # Pre-seeded so _ensure_emergency_ops short-circuits before any live GO API call.
        ctx.emergency_ops_by_iso["NGA"] = {
            "MDRNG041": {"name": "Nigeria - Floods", "code": "MDRNG041"},
        }
        ctx.emergency_ops_ordered_by_iso["NGA"] = [ctx.emergency_ops_by_iso["NGA"]["MDRNG041"]]

        resolved = _resolve_workbook_emergency_slot_metadata(ctx, wb, "NGA")

        assert resolved[1]["appeal_name"] == "Nigeria - Floods"
        assert resolved[1]["mdr_code"] == "MDRNG041"
        assert resolved[1]["display_value"] == "Nigeria - Floods (MDRNG041)"
        assert ctx.warnings == []
    finally:
        wb.close()


def test_resolve_workbook_emergency_slot_metadata_warns_when_code_unknown():
    """A retyped/incorrect MDR code that matches no real GO appeal for the country must
    not be silently trusted -- the slot's indicator values would otherwise be attributed
    to a fictitious operation with zero visibility. Excel's name/code are kept so the
    slot's data still imports (never blocks), but a warning must surface the mismatch."""
    import openpyxl

    from import_upr_excel_data import UprImportContext
    from upr_country_reporting_excel_template import write_named_cell

    if not os.path.isfile(TEMPLATE_PATH):
        pytest.skip("UPR Country Reporting template file not present")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    try:
        write_named_cell(wb, "Data_MDR1", "MDRNG999")
        write_named_cell(wb, "Data_EO1", "Nigeria Floods EA")

        ctx = UprImportContext(template_ids=[33])
        ctx.emergency_ops_by_iso["NGA"] = {
            "MDRNG041": {"name": "Nigeria - Floods", "code": "MDRNG041"},
        }
        ctx.emergency_ops_ordered_by_iso["NGA"] = [ctx.emergency_ops_by_iso["NGA"]["MDRNG041"]]

        resolved = _resolve_workbook_emergency_slot_metadata(ctx, wb, "NGA")

        # Excel values are preserved as a fallback -- the slot's data is never dropped.
        assert resolved[1]["appeal_name"] == "Nigeria Floods EA"
        assert resolved[1]["mdr_code"] == "MDRNG999"
        from upr_import_warnings import warning_text

        assert any(
            "MDRNG999" in warning_text(w) and "is not listed for this country in GO" in warning_text(w)
            for w in ctx.warnings
        ), f"Expected a GO API mismatch warning; got: {ctx.warnings}"
    finally:
        wb.close()


def test_resolve_workbook_emergency_slot_metadata_matches_raw_parse_when_no_ctx_data():
    """Sanity check: with no GO API data seeded for the country at all (e.g. a country
    the GO API has no operations for), resolution degrades to the raw Excel values --
    same as calling parse_emergency_slot_metadata directly -- rather than erroring."""
    import openpyxl

    from import_upr_excel_data import UprImportContext
    from upr_country_reporting_excel_template import write_named_cell

    if not os.path.isfile(TEMPLATE_PATH):
        pytest.skip("UPR Country Reporting template file not present")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    try:
        write_named_cell(wb, "Data_MDR2", "MDRZZ001")
        write_named_cell(wb, "Data_EO2", "Some Appeal")

        ctx = UprImportContext(template_ids=[33])
        ctx.emergency_ops_by_iso["ZZZ"] = {}
        ctx.emergency_ops_ordered_by_iso["ZZZ"] = []

        raw = parse_emergency_slot_metadata(wb)
        resolved = _resolve_workbook_emergency_slot_metadata(ctx, wb, "ZZZ")

        assert resolved[2]["appeal_name"] == raw[2]["appeal_name"] == "Some Appeal"
        assert resolved[2]["mdr_code"] == raw[2]["mdr_code"] == "MDRZZ001"
    finally:
        wb.close()


def test_upsert_emergency_repeat_choice(app):
    from unittest.mock import MagicMock, patch

    inst = MagicMock()
    inst.id = 501

    with app.app_context():
        with patch("app.models.forms.RepeatGroupData") as mock_rgd_cls, patch(
            "app.extensions.db.session.add"
        ):
            mock_entry = MagicMock()
            mock_rgd_cls.query.filter_by.return_value.first.return_value = None
            mock_rgd_cls.return_value = mock_entry

            _upsert_emergency_repeat_choice(
                repeat_instance=inst,
                choice_item_id=1374,
                appeal_name="Afghanistan - Earthquake",
                mdr_code="MDRAF019",
            )

            mock_rgd_cls.assert_called_once_with(repeat_instance_id=501, form_item_id=1374)
            assert mock_entry.value == "Afghanistan - Earthquake (MDRAF019)"
            assert mock_entry.disagg_type == "emergency_operation"
            assert mock_entry.disagg_data == {"name": "Afghanistan - Earthquake", "code": "MDRAF019"}


def test_yes_no_indicator_detected_from_bank():
    entry = type(
        "Entry",
        (),
        {"value": "yes", "indicator_bank": type("Bank", (), {"type": "YesNo"})()},
    )()
    assert _entry_is_yes_no(entry)


def test_is_yes_no_indicator_type():
    assert _is_yes_no_indicator_type("YesNo")
    assert _is_yes_no_indicator_type("yes/no")
    assert _is_yes_no_indicator_type("Yes No")
    assert _is_yes_no_indicator_type("Yes-No")
    assert _is_yes_no_indicator_type("Boolean")
    assert not _is_yes_no_indicator_type("Number")


def test_parse_support_matrix_ticks_scalar_and_nested():
    cells = {
        "166_EFs Supported": 1,
        "42_SP1 Supported": {"original": 0, "modified": 1, "isModified": True},
        "99_SP2 Supported": {"original": 0, "modified": 0},
        "not_a_key": 1,
    }
    ticks = _parse_support_matrix_ticks(cells)
    assert ticks == {166: {"EFs": True}, 42: {"SP1": True}}


def test_matrix_cell_is_set():
    assert _matrix_cell_is_set(1)
    assert _matrix_cell_is_set("X")
    assert not _matrix_cell_is_set(0)
    assert not _matrix_cell_is_set({"original": 0, "modified": 0})
    assert _matrix_cell_is_set({"original": 0, "modified": 1})


def test_normalize_matrix_cells():
    raw = {"a": {"original": 0, "modified": 1}, "b": 2}
    assert _normalize_matrix_cells(raw) == {"a": 1, "b": 2}


def test_bilateral_row_index_reads_column_c(upr_country_reporting_workbook):
    from upr_country_reporting_excel_template import _bilateral_ns_name_for_row

    name = _bilateral_ns_name_for_row(upr_country_reporting_workbook, "Bilateral Support", "Data_act", 1)
    assert name


def test_entity_ids_from_matrix_disagg_with_tick_filter():
    disagg = {"166_SP1": 1, "166_SP2": 0, "42_SP1": 1, "42_SP3": 1, "_table": "national_society"}
    assert _entity_ids_from_matrix_disagg(disagg, tick_column_names=["SP1"], require_tick=True) == {166, 42}
    assert _entity_ids_from_matrix_disagg(disagg, require_tick=False) == {166, 42}


def test_export_bilateral_support_splits_planned_and_manual():
    import openpyxl
    from unittest.mock import patch
    from upr_country_reporting_excel_template import (
        BILATERAL_MANUAL_TABLE,
        BILATERAL_PLANNED_TABLE,
        _export_bilateral_support,
        _bilateral_ns_name_for_row,
        read_table_cell,
    )

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ctx = type("Ctx", (), {"ns_name_to_id": {"swiss red cross": 166, "norwegian red cross": 42}})()
    support_cells = {
        "166_EFs Supported": 1,
        "99_SP1 Supported": 1,
    }
    id_to_name = {166: "Swiss Red Cross", 99: "Manual NS Example"}

    with patch("upr_country_reporting_excel_template._resolve_autoloaded_bilateral_ns_ids", return_value={166}), patch(
        "app.models.organization.NationalSociety"
    ) as mock_ns:
        mock_ns.query.filter.return_value.all.return_value = [
            type("NS", (), {"id": 166, "name": "Swiss Red Cross"})(),
            type("NS", (), {"id": 99, "name": "Manual NS Example"})(),
        ]
        _export_bilateral_support(wb, ctx, support_cells, aes=type("AES", (), {})())

    planned_sheet, planned_table = BILATERAL_PLANNED_TABLE
    manual_sheet, manual_table = BILATERAL_MANUAL_TABLE
    assert _bilateral_ns_name_for_row(wb, planned_sheet, planned_table, 0) == "Swiss Red Cross"
    assert read_table_cell(wb, planned_sheet, planned_table, 0, "EFs_Supported") == "X"
    assert _bilateral_ns_name_for_row(wb, manual_sheet, manual_table, 0) == "Manual NS Example"
    assert read_table_cell(wb, manual_sheet, manual_table, 0, "SP1_Supported") == "X"
    wb.close()


def test_export_funding_breakdown_writes_excel_columns():
    import openpyxl

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    cells = {
        "Resilience - Climate and environment_Funding (CHF)": 1000,
        "Resilience - Climate and environment_Expenditure (CHF)": 800,
        "Response - Disasters and crises_Funding (CHF)": {"original": 0, "modified": 500},
    }
    _export_funding_breakdown(wb, cells)
    _, rows = read_named_table(wb, "Funding", "Data_funding2")
    sp1 = next(r for r in rows if r.get("Attribute") == "SP1")
    sp2 = next(r for r in rows if r.get("Attribute") == "SP2")
    assert sp1["Funding"] == 1000
    assert sp1["Expenditure"] == 800
    assert sp2["Funding"] == 500
    wb.close()


def test_parse_funding_breakdown_reads_funding_column():
    import openpyxl

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    write_table_cell(wb, "Funding", "Data_funding2", 1, "Funding", 12345)
    write_table_cell(wb, "Funding", "Data_funding2", 1, "Expenditure", 6789)
    parsed = parse_funding(wb)
    row = parsed["breakdown"]["Resilience - Climate and environment"]
    assert row["Funding (CHF)"] == 12345
    assert row["Expenditure (CHF)"] == 6789
    wb.close()


def test_assignment_display_label_prefers_custom_name():
    template = type("Template", (), {"name": "Reporting – Country"})()
    assigned_form = type(
        "AssignedForm",
        (),
        {"custom_name": "Afghanistan MYR 2026", "period_name": "Jan-Jun 2026", "template": template},
    )()
    assigned_form.display_name = "Afghanistan MYR 2026"
    aes = type("AES", (), {"assigned_form": assigned_form})()
    assert _assignment_display_label(aes) == "Afghanistan MYR 2026"


def test_assignment_display_label_falls_back_to_template_and_period():
    tpl = type("Template", (), {"name": "Reporting – Country"})()

    class FakeAssignedForm:
        custom_name = None
        period_name = "Jan-Jun 2026"
        template = tpl

        @property
        def display_name(self):
            return f"{self.template.name} – {self.period_name}"

    aes = type("AES", (), {"assigned_form": FakeAssignedForm()})()
    assert _assignment_display_label(aes) == "Reporting – Country – Jan-Jun 2026"


def test_apply_reporting_assignment_label_updates_start_and_headers(upr_country_reporting_workbook):
    import openpyxl

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    _apply_reporting_assignment_label(wb, "Custom Assignment Label")
    assert read_named_cell(wb, ASSIGNMENT_LABEL_NAMED_CELL) == "Custom Assignment Label"
    assert wb["Start"]["C2"].value == "Custom Assignment Label - Start here"
    assert wb["Start"]["C2"].font.size == 22.0
    assert "Data_AssignmentLabel" in str(wb["Funding"]["B1"].value)
    assert "Midyear Reporting" not in str(wb["Funding"]["B1"].value)
    wb.close()


def test_workbook_yes_no_value_mapping():
    assert _workbook_yes_no_value("Applicable") == "yes"
    assert _workbook_yes_no_value(" applicable ") == "yes"
    assert _workbook_yes_no_value("Data not available") == "no"
    assert _workbook_yes_no_value("") == "no"


def test_yes_no_value_is_applicable():
    assert _yes_no_value_is_applicable("yes")
    assert _yes_no_value_is_applicable("YES")
    assert not _yes_no_value_is_applicable("no")
    assert not _yes_no_value_is_applicable("")
    assert not _yes_no_value_is_applicable(None)


def test_export_yes_no_indicator_writes_applicable_only_for_yes(upr_country_reporting_workbook):
    import openpyxl
    from upr_country_reporting_excel_template import INDICATOR_DNA_HEADER

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    yes_entry = type(
        "Entry",
        (),
        {
            "value": "yes",
            "is_data_not_available": False,
            "indicator_bank": type("Bank", (), {"type": "YesNo"})(),
            "disagg_data": None,
            "form_item": None,
        },
    )()
    no_entry = type(
        "Entry",
        (),
        {
            "value": "no",
            "is_data_not_available": False,
            "indicator_bank": type("Bank", (), {"type": "YesNo"})(),
            "disagg_data": None,
            "form_item": None,
        },
    )()
    dna_entry = type(
        "Entry",
        (),
        {
            "value": "no",
            "is_data_not_available": True,
            "indicator_bank": type("Bank", (), {"type": "YesNo"})(),
            "disagg_data": None,
            "form_item": None,
        },
    )()

    _write_indicator_entry(wb, "Overall action Indicators", "Data_core", 0, yes_entry, {}, yes_no=True)
    assert read_table_cell(wb, "Overall action Indicators", "Data_core", 0, INDICATOR_DNA_HEADER) == INDICATOR_APPLICABLE_VALUE

    _write_indicator_entry(wb, "Overall action Indicators", "Data_core", 1, no_entry, {}, yes_no=True)
    assert read_table_cell(wb, "Overall action Indicators", "Data_core", 1, INDICATOR_DNA_HEADER) in (None, "")

    _write_indicator_entry(
        wb,
        "Overall action Indicators",
        "Data_core",
        2,
        dna_entry,
        {},
        data_not_available=True,
        yes_no=True,
    )
    assert read_table_cell(wb, "Overall action Indicators", "Data_core", 2, INDICATOR_DNA_HEADER) == INDICATOR_DNA_VALUE
    wb.close()


def test_collect_workbook_indicator_bank_ids_includes_table_id_column(
    upr_country_reporting_workbook,
):
    kpi = build_kpi_lookup(upr_country_reporting_workbook)
    table_ids = _collect_workbook_indicator_bank_ids(upr_country_reporting_workbook, kpi)
    assert table_ids - set(kpi.values()), "fixture should include table IDs outside Final KPI list"


def test_parse_indicators_yes_no_applicable_not_skipped(upr_country_reporting_workbook):
    """Applicable-only rows must not be dropped when the bank id is known Yes/No."""
    kpi = build_kpi_lookup(upr_country_reporting_workbook)
    baseline = parse_indicators(upr_country_reporting_workbook, yes_no_bank_ids=set(), kpi_lookup=kpi)
    sample = next(r for r in baseline if r.get("bank_id"))
    bank_id = int(sample["bank_id"])
    rows = parse_indicators(
        upr_country_reporting_workbook,
        yes_no_bank_ids={bank_id},
        kpi_lookup=kpi,
    )
    target = next(r for r in rows if int(r["bank_id"]) == bank_id)
    assert target["value"] in ("yes", "no")


def test_parse_indicators_skips_applicable_only_for_numeric_banks(upr_country_reporting_workbook):
    kpi = build_kpi_lookup(upr_country_reporting_workbook)
    rows = parse_indicators(
        upr_country_reporting_workbook,
        yes_no_bank_ids=set(),
        kpi_lookup=kpi,
    )
    applicable_only = [
        r
        for r in rows
        if "applicable" in str(r.get("applicable_text") or "")
        and r.get("value") is None
        and r.get("disagg") is None
        and not r.get("data_not_available")
    ]
    assert not applicable_only


def test_parse_indicators_yes_no_dna_preserves_data_not_available():
    """Regression test: a Yes/No indicator marked "Data not available" in Excel
    must come out of parse_indicators() with data_not_available=True and
    value=None — never silently coerced to a "no" answer. "No" and "we don't
    have this data" are materially different claims and conflating them was a
    real bug (the code used to force is_dna=False for every Yes/No row).

    Uses a fresh workbook load (not the shared module-scoped
    upr_country_reporting_workbook fixture) since this test writes into a
    table row, and other tests in this module rely on that fixture's rows
    being in their original state.
    """
    import openpyxl

    if not os.path.isfile(TEMPLATE_PATH):
        pytest.skip("UPR Country Reporting template file not present")

    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    try:
        sheet_name, table_name = "Overall action Indicators", "Data_core"
        bank_id = 555555
        write_table_cell(wb, sheet_name, table_name, 0, SP_EF_HEADER, "Health")
        write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_HEADER, "A Yes/No indicator")
        write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_ID_HEADER, bank_id)
        write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_DNA_HEADER, INDICATOR_DNA_VALUE)

        rows = parse_indicators(wb, yes_no_bank_ids={bank_id}, kpi_lookup={})
        target = next(r for r in rows if r.get("bank_id") == bank_id)

        assert target["data_not_available"] is True
        assert target["value"] is None
    finally:
        wb.close()


def test_parse_indicators_blank_applicable_is_not_applicable():
    """A blank Applicable cell imports as Not Applicable only when the caller
    opts in (i.e. the destination FormItem has allow_not_applicable=True)."""
    import openpyxl

    if not os.path.isfile(TEMPLATE_PATH):
        pytest.skip("UPR Country Reporting template file not present")

    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    try:
        sheet_name, table_name = "Overall action Indicators", "Data_core"
        bank_id = 555556
        write_table_cell(wb, sheet_name, table_name, 0, SP_EF_HEADER, "Health")
        write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_HEADER, "A core indicator")
        write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_ID_HEADER, bank_id)
        write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_DNA_HEADER, None)

        rows = parse_indicators(wb, yes_no_bank_ids={bank_id}, kpi_lookup={})
        target = next(r for r in rows if r.get("bank_id") == bank_id)
        value, is_dna, _disagg, should_import, is_na = _resolve_indicator_import_value(
            target, bank_id, {bank_id}, blank_is_not_applicable=True
        )

        assert target["applicable_text"] == ""
        assert value is None
        assert is_dna is False
        assert is_na is True
        assert should_import is True
    finally:
        wb.close()


def test_resolve_indicator_import_value_yes_no():
    yes_row = {
        "data_not_available": False,
        "applicable_text": "applicable",
        "value": None,
        "disagg": None,
    }
    value, is_dna, disagg, should_import, is_na = _resolve_indicator_import_value(yes_row, 999, {999})
    assert value == "yes"
    assert not is_dna
    assert not is_na
    assert disagg is None
    assert should_import

    blank_row = {
        "data_not_available": False,
        "applicable_text": "",
        "value": None,
        "disagg": None,
    }

    # Default (blank_is_not_applicable=False): destination has no Not
    # Applicable state to record, so a blank cell keeps its historical
    # meaning of "no" rather than being skipped or mis-mapped. This is what
    # every dynamic Other-indicator / Emergency Appeal row uses, since most
    # blank rows there just mean "not selected for this operation".
    value, is_dna, disagg, should_import, is_na = _resolve_indicator_import_value(blank_row, 999, {999})
    assert value == "no"
    assert not is_dna
    assert not is_na
    assert disagg is None
    assert should_import

    # Opt-in (blank_is_not_applicable=True): only passed by callers for a
    # destination FormItem that actually has allow_not_applicable=True.
    value, is_dna, disagg, should_import, is_na = _resolve_indicator_import_value(
        blank_row, 999, {999}, blank_is_not_applicable=True
    )
    assert value is None
    assert not is_dna
    assert is_na
    assert should_import

    # Regression test: a Yes/No indicator marked "Data not available" in Excel
    # must import as DNA (is_data_not_available=True, no value), never silently
    # coerced to a "No" answer — those are materially different claims ("we
    # don't know" vs "confirmed no").
    dna_row = {
        "data_not_available": True,
        "applicable_text": "data not available",
        "value": None,
        "disagg": None,
    }
    value, is_dna, disagg, should_import, is_na = _resolve_indicator_import_value(
        dna_row, 999, {999}, blank_is_not_applicable=True
    )
    assert value is None
    assert is_dna
    assert not is_na
    assert disagg is None
    assert should_import

    numeric_row = {
        "data_not_available": False,
        "applicable_text": "applicable",
        "value": None,
        "disagg": None,
    }
    value, _, _, should_import, is_na = _resolve_indicator_import_value(numeric_row, 123, {999})
    assert value is None
    assert not should_import
    assert not is_na

    # Regression test: a real value present must never be discarded in favor
    # of Not Applicable just because the Applicable/DNA cell is blank — that
    # column is unrelated to the numeric answer cell for non-Yes/No indicators.
    numeric_blank_with_value = {
        "data_not_available": False,
        "applicable_text": "  ",
        "value": 12,
        "disagg": None,
    }
    value, is_dna, disagg, should_import, is_na = _resolve_indicator_import_value(
        numeric_blank_with_value, 123, {999}, blank_is_not_applicable=True
    )
    assert value == 12
    assert not is_dna
    assert not is_na
    assert should_import

    # A genuinely empty numeric row (no value, no disagg) with the opt-in set
    # imports as Not Applicable instead of being silently skipped.
    numeric_blank_empty = {
        "data_not_available": False,
        "applicable_text": "  ",
        "value": None,
        "disagg": None,
    }
    value, is_dna, disagg, should_import, is_na = _resolve_indicator_import_value(
        numeric_blank_empty, 123, {999}, blank_is_not_applicable=True
    )
    assert value is None
    assert not is_dna
    assert is_na
    assert should_import

    # Same empty numeric row without the opt-in: skipped, exactly like before
    # Not Applicable existed — this is what dynamic Other-indicator / Emergency
    # Appeal rows (and static items without allow_not_applicable) still get.
    value, is_dna, disagg, should_import, is_na = _resolve_indicator_import_value(
        numeric_blank_empty, 123, {999}
    )
    assert value is None
    assert not should_import
    assert not is_na


def test_reporting_funding_matrix_column_matches_form_item(app):
    with app.app_context():
        from import_upr_excel_data import (
            ITEM_REPORTING_COUNTRY_FUNDING,
            _matrix_column_name_from_form_item,
        )
        from app.models import FormItem

        item = FormItem.query.get(ITEM_REPORTING_COUNTRY_FUNDING)
        expected = _matrix_column_name_from_form_item(item) or "ns_fun"
        assert _reporting_funding_matrix_column() == expected


def test_validate_upr_country_reporting_import_file_accepts_current_template(
    upr_country_reporting_workbook,
):
    result = validate_upr_country_reporting_import_file(
        upr_country_reporting_workbook,
        expected_country="Afghanistan",
        expected_period="Jan-Jun 2026",
    )
    assert result["valid"]
    assert result["preview"]["kpi_count"] > 0
    assert result["preview"]["core_indicator_rows"] > 0


def test_validate_upr_country_reporting_import_file_rejects_generic_export():
    import openpyxl

    wb = openpyxl.Workbook()
    wb.create_sheet("Template")
    wb.create_sheet("Pages")
    wb.create_sheet("Sections")
    wb.create_sheet("Items")
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    result = validate_upr_country_reporting_import_file(wb)
    assert not result["valid"]
    assert any("generic form Excel export" in err for err in result["errors"])
    wb.close()


def test_validate_upr_country_reporting_import_file_rejects_country_mismatch(
    upr_country_reporting_workbook,
):
    result = validate_upr_country_reporting_import_file(
        upr_country_reporting_workbook,
        expected_country="Netherlands",
        expected_period="Jan-Jun 2026",
    )
    assert not result["valid"]
    assert any("country" in err.lower() for err in result["errors"])


def test_validate_upr_country_reporting_import_file_allows_period_mismatch(
    upr_country_reporting_workbook,
):
    result = validate_upr_country_reporting_import_file(
        upr_country_reporting_workbook,
        expected_country="Afghanistan",
        expected_period="2025",
    )
    assert result["valid"]
    assert any("period" in w.lower() for w in result["warnings"])
    assert not any("period" in err.lower() for err in result["errors"])


def test_quiet_openpyxl_io_suppresses_pil_debug(capsys):
    import logging

    pil_logger = logging.getLogger("PIL.PngImagePlugin")
    previous = pil_logger.level
    pil_logger.setLevel(logging.DEBUG)
    try:
        with _quiet_openpyxl_io():
            pil_logger.debug("STREAM b'IHDR' 16 13")
        captured = capsys.readouterr()
        assert "IHDR" not in captured.out
        assert "IHDR" not in captured.err
    finally:
        pil_logger.setLevel(previous)


def test_import_rows_to_client_payload_splits_fields_and_matrices():
    from import_fdrs_form_data import COL_DATA_NA, COL_DISAGG, COL_ITEM, COL_NA, COL_VALUE

    rows = [
        {
            COL_ITEM: "100",
            COL_VALUE: "42",
            COL_DISAGG: "",
            COL_DATA_NA: "",
        },
        {
            COL_ITEM: "200",
            COL_VALUE: "",
            COL_DISAGG: '{"mode":"total","values":{"total":99}}',
            COL_DATA_NA: "",
        },
        {
            COL_ITEM: "300",
            COL_VALUE: "",
            COL_DISAGG: '{"IFRC_tot_fn":1500}',
            COL_DATA_NA: "",
        },
        {
            COL_ITEM: "400",
            COL_VALUE: "",
            COL_DISAGG: "",
            COL_DATA_NA: "1",
        },
        {
            COL_ITEM: "500",
            COL_VALUE: "",
            COL_DISAGG: "",
            COL_DATA_NA: "",
            COL_NA: "1",
        },
    ]
    fields, matrices = import_rows_to_client_payload(rows)
    assert fields["100"]["value"] == "42"
    assert fields["200"]["disagg_data"]["mode"] == "total"
    assert matrices["300"]["IFRC_tot_fn"] == 1500
    assert fields["400"]["data_not_available"] is True
    assert fields["500"]["not_applicable"] is True
    assert "400" not in matrices


def test_dedupe_upr_import_warnings_collapses_period_mismatch():
    warnings = dedupe_upr_import_warnings([
        "Workbook Version period '2025' does not match assignment period 'Jan-Jun 2026'",
        "Workbook period '2025' differs from this assignment ('Jan-Jun 2026'). Values will be loaded into the current assignment.",
        "No T33 form item for indicator 'Example' in 'Section' (AFG)",
        "No T33 form item for indicator 'Example' in 'Section' (AFG)",
    ])
    assert len(warnings) == 2
    assert sum("period" in w.lower() for w in warnings) == 1
    assert any("Example" in w for w in warnings)


def test_resolve_workbook_indicator_bank_id_prefers_explicit_id():
    """An explicit ID cell always wins, even if the indicator text would
    fuzzy-match a completely different bank id in kpi_lookup."""
    row = {"bank_id": 555, "sp_ef": "Health", "indicator": "Totally unrelated text"}
    kpi_lookup = {("Health", "Totally unrelated text"): 999}
    assert _resolve_workbook_indicator_bank_id(row, kpi_lookup) == 555


def test_resolve_workbook_indicator_bank_id_picks_best_fuzzy_match_not_first():
    """Regression test: when a row's ID cell is blank (e.g. cleared or a row
    was pasted without it), the fallback must pick the BEST-scoring candidate
    across the section, not merely the first one that clears the 0.60
    threshold. Humanitarian indicator labels are often near-duplicates
    ("...reached with X" vs "...reached with X and Y"), so picking the first
    match found in kpi_lookup's (arbitrary) iteration order can silently
    attribute a value to the wrong indicator.
    """
    kpi_lookup = {
        # Inserted first — a strict-prefix substring of the row's text below,
        # so it also clears the threshold (short-in-long => 0.8) but is NOT
        # the row's actual indicator.
        ("Health", "Number of people reached with health services"): 101,
        # The row's exact indicator — inserted second, would be skipped by a
        # naive first-match loop even though it's a perfect (1.0) match.
        ("Health", "Number of people reached with health services and referrals"): 102,
    }
    row = {
        "bank_id": None,
        "sp_ef": "Health",
        "indicator": "Number of people reached with health services and referrals",
    }
    assert _resolve_workbook_indicator_bank_id(row, kpi_lookup) == 102


def test_resolve_workbook_indicator_bank_id_returns_none_below_threshold():
    row = {"bank_id": None, "sp_ef": "Health", "indicator": "Completely different indicator"}
    kpi_lookup = {("Health", "Number of shelters distributed"): 42}
    assert _resolve_workbook_indicator_bank_id(row, kpi_lookup) is None


def test_resolve_workbook_indicator_bank_id_rejects_short_generic_fragment():
    """Same real-world scenario as test_indicator_similarity_rejects_short_
    generic_substring_fragment, exercised through the actual blank-ID
    fallback used during import: a stray short fragment left in the
    indicator cell must not fuzzy-match an unrelated KPI merely because it
    is a literal substring of that KPI's (much longer) label.
    """
    row = {"bank_id": None, "sp_ef": "Health", "indicator": "Support"}
    kpi_lookup = {
        ("Health", "Number of national society branches receiving direct technical support"): 42,
    }
    assert _resolve_workbook_indicator_bank_id(row, kpi_lookup) is None


def test_resolve_workbook_indicator_bank_id_respects_section(upr_country_reporting_workbook):
    """A blank-ID row must not fuzzy-match an indicator in a different SP/EF section."""
    kpi_lookup = build_kpi_lookup(upr_country_reporting_workbook)
    (sp_ef, indicator_label), bank_id = next(iter(kpi_lookup.items()))
    row = {"bank_id": None, "sp_ef": f"NOT-{sp_ef}", "indicator": indicator_label}
    assert _resolve_workbook_indicator_bank_id(row, kpi_lookup) is None


def test_transform_warns_when_row_has_no_id_and_matched_by_name(app):
    """End-to-end regression test for two related fixes:
    (1) _resolve_workbook_indicator_bank_id resolving a blank-ID row via the
        best fuzzy match, and
    (2) run_upr_country_reporting_import no longer silently dropping
        dynamic/emergency-indicator warnings that are appended to
        ctx.warnings AFTER the initial snapshot was taken.
    This test exercises the first fix directly and the underlying warning
    plumbing that the second fix depends on (transform_... appending to
    ctx.warnings, which the caller must read after all steps complete).
    """
    import openpyxl

    if not os.path.isfile(TEMPLATE_PATH):
        pytest.skip("UPR Country Reporting template file not present")

    from import_upr_excel_data import build_import_context, REPORTING_COUNTRY_TEMPLATE_ID
    from upr_country_reporting_excel_template import transform_upr_country_reporting_to_import_rows

    with app.app_context():
        wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
        try:
            sheet_name, table_name = "Overall action Indicators", "Data_core"
            kpi_lookup = build_kpi_lookup(wb)
            (sp_ef, indicator_label), _bank_id = next(iter(kpi_lookup.items()))

            # Simulate a corrupted row: ID cell cleared, but Applicable + a
            # value entered — exactly the "moved/cleared cell" scenario users
            # can accidentally trigger in Excel. Write the SP/EF + indicator
            # text verbatim from kpi_lookup's own key so this test exercises
            # the fuzzy-match fallback deterministically, independent of
            # whether Data_core row 0 happens to line up 1:1 with the master
            # KPI list's exact text/section-label formatting.
            write_table_cell(wb, sheet_name, table_name, 0, SP_EF_HEADER, sp_ef)
            write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_HEADER, indicator_label)
            write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_ID_HEADER, None)
            write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_DNA_HEADER, INDICATOR_APPLICABLE_VALUE)
            write_table_cell(wb, sheet_name, table_name, 0, "Total\nDirect", 123)

            ctx = build_import_context([REPORTING_COUNTRY_TEMPLATE_ID])
            transform_upr_country_reporting_to_import_rows(
                999999999, wb, ctx, iso3="ZZZ", period="Jan-Jun 2026"
            )

            assert any(
                "had no ID" in w and str(indicator_label) in w for w in ctx.warnings
            ), f"Expected a 'had no ID' warning mentioning {indicator_label!r}; got: {ctx.warnings}"
        finally:
            wb.close()


# ---------------------------------------------------------------------------
# Percentage out-of-range detection (e.g. 500 entered instead of 50).
# ---------------------------------------------------------------------------


def test_is_percentage_indicator_type_matches_common_spellings():
    assert _is_percentage_indicator_type("percentage") is True
    assert _is_percentage_indicator_type("Percentage") is True
    assert _is_percentage_indicator_type("PERCENT") is True
    assert _is_percentage_indicator_type("pct") is True
    assert _is_percentage_indicator_type("number") is False
    assert _is_percentage_indicator_type("yesno") is False
    assert _is_percentage_indicator_type(None) is False


def test_iter_numeric_leaves_scalar_values():
    assert list(_iter_numeric_leaves(45)) == [45.0]
    assert list(_iter_numeric_leaves(45.5)) == [45.5]
    assert list(_iter_numeric_leaves("45")) == [45.0]
    assert list(_iter_numeric_leaves(None)) == []
    assert list(_iter_numeric_leaves("")) == []
    assert list(_iter_numeric_leaves("not a number")) == []
    # Booleans are a bool/int subtype in Python — must not be misread as 0/1.
    assert list(_iter_numeric_leaves(True)) == []


def test_iter_numeric_leaves_walks_nested_disagg_structure():
    disagg = {
        "mode": "sex_age",
        "values": {
            "direct": {"male_5_17": 60, "female_5_17": "45"},
            "indirect": 10,
        },
    }
    assert sorted(_iter_numeric_leaves(disagg)) == [10.0, 45.0, 60.0]


def test_percentage_range_warning_ignores_non_percentage_indicator():
    row = {"indicator": "Some indicator", "sp_ef": "Health"}
    warning = _percentage_range_warning(row, 42, 500, None, percentage_bank_ids=set(), allow_over_100_bank_ids=set())
    assert warning is None


def test_percentage_range_warning_flags_scalar_over_100():
    row = {"indicator": "Percentage of X reached", "sp_ef": "Health"}
    warning = _percentage_range_warning(
        row, 42, 500, None, percentage_bank_ids={42}, allow_over_100_bank_ids=set()
    )
    assert warning is not None
    assert "500" in warning
    assert "Percentage of X reached" in warning
    assert "0-100%" in warning


def test_percentage_range_warning_flags_negative_value():
    row = {"indicator": "Percentage of X reached", "sp_ef": "Health"}
    warning = _percentage_range_warning(
        row, 42, -5, None, percentage_bank_ids={42}, allow_over_100_bank_ids=set()
    )
    assert warning is not None
    assert "-5" in warning


def test_percentage_range_warning_allows_in_range_value():
    row = {"indicator": "Percentage of X reached", "sp_ef": "Health"}
    warning = _percentage_range_warning(
        row, 42, 45, None, percentage_bank_ids={42}, allow_over_100_bank_ids=set()
    )
    assert warning is None


def test_percentage_range_warning_flags_disagg_sub_value_out_of_range():
    """The scalar 'value' can be None while the real number lives inside a
    sex/age disaggregation breakdown — the range check must look there too."""
    row = {"indicator": "Percentage of X reached", "sp_ef": "Health"}
    disagg = {"mode": "sex_age", "values": {"direct": {"male_5_17": 250, "female_5_17": 10}}}
    warning = _percentage_range_warning(
        row, 42, None, disagg, percentage_bank_ids={42}, allow_over_100_bank_ids=set()
    )
    assert warning is not None
    assert "250" in warning


def test_percentage_range_warning_respects_allow_over_100_override():
    """Indicators explicitly configured with allow_over_100 (cumulative/ratio
    KPIs) must not be flagged for exceeding 100 — but a negative value is
    still nonsensical for a percentage and must still be flagged."""
    row = {"indicator": "Cumulative coverage ratio", "sp_ef": "Health"}
    over_100_warning = _percentage_range_warning(
        row, 42, 250, None, percentage_bank_ids={42}, allow_over_100_bank_ids={42}
    )
    assert over_100_warning is None

    negative_warning = _percentage_range_warning(
        row, 42, -5, None, percentage_bank_ids={42}, allow_over_100_bank_ids={42}
    )
    assert negative_warning is not None


def test_transform_warns_on_out_of_range_percentage_value(app):
    """End-to-end regression test: a percentage-type indicator given a wildly
    out-of-range value (the classic '500 instead of 50' data-entry mistake in
    Excel) must produce a warning instead of silently importing as-is.

    Creates its own temporary percentage-type IndicatorBank row (cleaned up
    afterwards) so this test is deterministic regardless of what the test
    database happens to have seeded.
    """
    import openpyxl
    from app.extensions import db
    from app.models import IndicatorBank

    if not os.path.isfile(TEMPLATE_PATH):
        pytest.skip("UPR Country Reporting template file not present")

    from import_upr_excel_data import build_import_context, REPORTING_COUNTRY_TEMPLATE_ID
    from upr_country_reporting_excel_template import transform_upr_country_reporting_to_import_rows

    indicator_name = "__test_percentage_range_indicator__"
    with app.app_context():
        temp_indicator = IndicatorBank(name=indicator_name, type="percentage")
        db.session.add(temp_indicator)
        db.session.flush()
        bank_id = int(temp_indicator.id)
        wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
        try:
            sheet_name, table_name = "Overall action Indicators", "Data_core"
            write_table_cell(wb, sheet_name, table_name, 0, SP_EF_HEADER, "Health")
            write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_HEADER, indicator_name)
            write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_ID_HEADER, bank_id)
            write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_DNA_HEADER, None)
            write_table_cell(wb, sheet_name, table_name, 0, "Total\nDirect", 450)

            ctx = build_import_context([REPORTING_COUNTRY_TEMPLATE_ID])
            transform_upr_country_reporting_to_import_rows(
                999999999, wb, ctx, iso3="ZZZ", period="Jan-Jun 2026"
            )

            assert any(
                "outside the valid 0-100%" in w and "450" in w for w in ctx.warnings
            ), f"Expected an out-of-range percentage warning; got: {ctx.warnings}"
        finally:
            wb.close()
            db.session.delete(temp_indicator)
            db.session.commit()


def test_transform_warns_on_id_present_but_text_mismatched(app):
    """End-to-end regression test for the OTHER half of the ID-integrity check:
    the workbook's ID cell is filled in (so this is NOT the blank-ID/fuzzy-match
    path), but the indicator TEXT next to it belongs to a different KPI — e.g.
    a row was inserted/deleted elsewhere in Excel and the ID column no longer
    lines up with the text column it was pasted next to.

    Per _bank_id_row_integrity_warning's contract the workbook's stated ID must
    still win the import (never silently switched), but a warning must reach
    ctx.warnings so a reviewer can catch the swap before saving — this exercises
    that full chain (parse_indicators -> transform_..._to_import_rows ->
    ctx.warnings), not just the isolated helper function.
    """
    import openpyxl

    if not os.path.isfile(TEMPLATE_PATH):
        pytest.skip("UPR Country Reporting template file not present")

    from import_upr_excel_data import build_import_context, REPORTING_COUNTRY_TEMPLATE_ID
    from upr_country_reporting_excel_template import transform_upr_country_reporting_to_import_rows

    with app.app_context():
        wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
        try:
            sheet_name, table_name = "Overall action Indicators", "Data_core"
            kpi_lookup = build_kpi_lookup(wb)
            (sp_ef, _indicator_label), bank_id = next(iter(kpi_lookup.items()))

            # Deliberately unrelated text sharing no meaningful words with any
            # real indicator label, so the fuzzy-similarity check reliably
            # falls below the match threshold regardless of which KPI ends up
            # first in kpi_lookup.
            mismatched_text = "ZZZZZ_MISMATCH_MARKER_424242 unrelated placeholder text"
            write_table_cell(wb, sheet_name, table_name, 0, SP_EF_HEADER, sp_ef)
            write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_HEADER, mismatched_text)
            write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_ID_HEADER, bank_id)
            write_table_cell(wb, sheet_name, table_name, 0, INDICATOR_DNA_HEADER, INDICATOR_APPLICABLE_VALUE)
            write_table_cell(wb, sheet_name, table_name, 0, "Total\nDirect", 123)

            ctx = build_import_context([REPORTING_COUNTRY_TEMPLATE_ID])
            transform_upr_country_reporting_to_import_rows(
                999999999, wb, ctx, iso3="ZZZ", period="Jan-Jun 2026"
            )

            assert any(
                f"Row ID {bank_id}" in w and "moved, copied, or swapped" in w for w in ctx.warnings
            ), f"Expected an ID/text mismatch integrity warning for bank_id {bank_id}; got: {ctx.warnings}"
        finally:
            wb.close()
