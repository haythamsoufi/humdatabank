"""
Comprehensive tests for TemplateExcelService.

Targets maximum code coverage for app/services/template_excel_service.py.

Strategy:
- Pure-function methods (_rewrite_rule_json_item_ids, _parse_json, helpers) tested without DB.
- Export / Import round-trip tests use real DB fixtures + a request context with a
  logged-in user so current_user is available.
"""
import io
import json
import pytest
import openpyxl
from unittest.mock import patch, MagicMock, call
from flask_login import login_user

from app import db
from app.models import FormTemplate, FormTemplateVersion, FormPage, FormSection, FormItem
from app.services.templates.excel_service import TemplateExcelService
from tests.factories import (
    create_test_admin,
    create_test_template,
    create_test_section,
    create_test_item,
    create_test_draft_version,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _excel_columns(app=None):
    if app is not None:
        with app.app_context():
            return {
                'template': TemplateExcelService.get_template_columns(),
                'pages': TemplateExcelService.get_page_columns(),
                'sections': TemplateExcelService.get_section_columns(),
                'items': TemplateExcelService.get_item_columns(),
            }
    return {
        'template': TemplateExcelService.get_template_columns(),
        'pages': TemplateExcelService.get_page_columns(),
        'sections': TemplateExcelService.get_section_columns(),
        'items': TemplateExcelService.get_item_columns(),
    }


def _blank_row(columns):
    return [None] * len(columns)


def _append_template_sheet_row_layout(ws, template_name='Test', extra_values=None):
    """Write Template sheet using field/value row layout (matches export format)."""
    fields = TemplateExcelService.get_template_columns()
    ws.append(TemplateExcelService.TEMPLATE_SHEET_ROW_HEADERS)
    for field in fields:
        value = None
        if extra_values and field in extra_values:
            value = extra_values[field]
        elif field == 'name':
            value = template_name
        ws.append([field, value])


def _append_template_sheet_legacy_columns(ws, template_name='Test', extra_values=None):
    """Write Template sheet using legacy column layout (header row + data row)."""
    fields = TemplateExcelService.get_template_columns()
    ws.append(fields)
    row = _blank_row(fields)
    row[fields.index('name')] = template_name
    if extra_values:
        for key, val in extra_values.items():
            if key in fields:
                row[fields.index(key)] = val
    ws.append(row)


def _make_export_workbook_bytes(
    template_name="Test",
    pages=None,
    sections=None,
    items=None,
    extra_sheets=None,
    app=None,
):
    """Create a minimal compliant export workbook as bytes."""
    cols = _excel_columns(app)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_t = wb.create_sheet("Template")
    _append_template_sheet_row_layout(ws_t, template_name=template_name)

    ws_p = wb.create_sheet("Pages")
    ws_p.append(cols['pages'])
    for page in (pages or []):
        ws_p.append(page)

    ws_s = wb.create_sheet("Sections")
    ws_s.append(cols['sections'])
    for section in (sections or []):
        ws_s.append(section)

    ws_i = wb.create_sheet("Items")
    ws_i.append(cols['items'])
    for item in (items or []):
        ws_i.append(item)

    for sheet_name in (extra_sheets or []):
        wb.create_sheet(sheet_name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# _rewrite_rule_json_item_ids
# ---------------------------------------------------------------------------

class TestRewriteRuleJsonItemIds:
    """Pure-function tests – no DB or app context needed."""

    def test_none_returns_none(self):
        assert TemplateExcelService._rewrite_rule_json_item_ids(None, {}) is None

    def test_empty_string_returns_unchanged(self):
        assert TemplateExcelService._rewrite_rule_json_item_ids("  ", {}) == "  "

    def test_dict_input_already_parsed(self):
        rule = {"item_id": "5"}
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {5: 99})
        assert json.loads(result)["item_id"] == "99"

    def test_list_input_rewritten(self):
        rule = [{"item_id": "3"}, {"item_id": "4"}]
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {3: 30, 4: 40})
        parsed = json.loads(result)
        assert parsed[0]["item_id"] == "30"
        assert parsed[1]["item_id"] == "40"

    def test_numeric_string_id(self):
        rule = json.dumps({"item_id": "7"})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {7: 77})
        assert json.loads(result)["item_id"] == "77"

    def test_integer_id(self):
        rule = json.dumps({"item_id": 8})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {8: 88})
        assert json.loads(result)["item_id"] == "88"

    def test_float_id(self):
        rule = json.dumps({"item_id": 9.0})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {9: 99})
        assert json.loads(result)["item_id"] == "99"

    def test_plugin_prefix_two_parts(self):
        rule = json.dumps({"item_id": "plugin_10"})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {10: 100})
        assert json.loads(result)["item_id"] == "plugin_100"

    def test_plugin_prefix_with_measure(self):
        rule = json.dumps({"item_id": "plugin_11_measure_id"})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {11: 111})
        parsed = json.loads(result)
        assert parsed["item_id"] == "plugin_111_measure_id"

    def test_plugin_prefix_no_digit_part_unchanged(self):
        rule = json.dumps({"item_id": "plugin_abc"})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {})
        assert json.loads(result)["item_id"] == "plugin_abc"

    def test_question_prefix(self):
        rule = json.dumps({"item_id": "question_12"})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {12: 120})
        assert json.loads(result)["item_id"] == "question_120"

    def test_indicator_prefix(self):
        rule = json.dumps({"item_id": "indicator_13"})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {13: 130})
        assert json.loads(result)["item_id"] == "indicator_130"

    def test_document_field_prefix(self):
        rule = json.dumps({"item_id": "document_field_14"})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {14: 140})
        assert json.loads(result)["item_id"] == "document_field_140"

    def test_form_item_prefix(self):
        rule = json.dumps({"item_id": "form_item_15"})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {15: 150})
        assert json.loads(result)["item_id"] == "form_item_150"

    def test_id_not_in_map_returns_original(self):
        rule = json.dumps({"item_id": "99"})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {})
        assert json.loads(result)["item_id"] == "99"

    def test_question_prefix_id_not_in_map(self):
        rule = json.dumps({"item_id": "question_55"})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {})
        assert json.loads(result)["item_id"] == "question_55"

    def test_plugin_prefix_id_not_in_map(self):
        rule = json.dumps({"item_id": "plugin_77"})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {})
        assert json.loads(result)["item_id"] == "plugin_77"

    def test_non_item_id_keys_are_walked(self):
        rule = json.dumps({"nested": {"item_id": "20"}})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {20: 200})
        assert json.loads(result)["nested"]["item_id"] == "200"

    def test_item_id_none_value_unchanged(self):
        rule = json.dumps({"item_id": None})
        result = TemplateExcelService._rewrite_rule_json_item_ids(rule, {})
        assert json.loads(result)["item_id"] is None

    def test_double_encoded_json(self):
        inner = json.dumps({"item_id": "6"})
        outer = json.dumps(inner)  # double-encoded
        result = TemplateExcelService._rewrite_rule_json_item_ids(outer, {6: 60})
        # After rewrite result should be a JSON-encoded string
        parsed = json.loads(result)
        assert "60" in str(parsed)

    def test_invalid_json_returns_original(self):
        bad = "not_json_{}"
        result = TemplateExcelService._rewrite_rule_json_item_ids(bad, {})
        # Should return original since it can't be parsed
        assert result == bad


# ---------------------------------------------------------------------------
# _parse_json
# ---------------------------------------------------------------------------

class TestParseJson:
    def test_none_returns_none(self):
        assert TemplateExcelService._parse_json(None) is None

    def test_empty_string_returns_none(self):
        assert TemplateExcelService._parse_json("") is None

    def test_none_string_returns_none(self):
        assert TemplateExcelService._parse_json("None") is None

    def test_dict_input_passthrough(self):
        d = {"key": "value"}
        assert TemplateExcelService._parse_json(d) is d

    def test_list_input_passthrough(self):
        lst = [1, 2, 3]
        assert TemplateExcelService._parse_json(lst) is lst

    def test_valid_json_string(self):
        assert TemplateExcelService._parse_json('{"a": 1}') == {"a": 1}

    def test_valid_json_list_string(self):
        assert TemplateExcelService._parse_json('[1, 2]') == [1, 2]

    def test_invalid_json_returns_none(self):
        assert TemplateExcelService._parse_json("{bad json}") is None

    def test_empty_json_object_returns_none(self):
        # json.loads('{}') == {} which is falsy; returns None
        assert TemplateExcelService._parse_json("{}") is None

    def test_empty_json_array_returns_none(self):
        assert TemplateExcelService._parse_json("[]") is None

    def test_non_empty_json_object_returned(self):
        result = TemplateExcelService._parse_json('{"x": 1}')
        assert result == {"x": 1}


# ---------------------------------------------------------------------------
# _format_json_for_excel
# ---------------------------------------------------------------------------

class TestFormatJsonForExcel:
    def test_pretty_prints_dict_with_line_breaks(self):
        value = {"is_required": False, "layout_column_width": "12", "matrix_config": {"type": "matrix"}}
        formatted = TemplateExcelService._format_json_for_excel(value)
        assert formatted is not None
        assert ",\n" in formatted
        assert '"is_required": false' in formatted

    def test_none_returns_none(self):
        assert TemplateExcelService._format_json_for_excel(None) is None

    def test_compact_json_string_is_reformatted(self):
        compact = '{"a": 1, "b": 2}'
        formatted = TemplateExcelService._format_json_for_excel(compact)
        assert formatted is not None
        assert ",\n" in formatted
        assert TemplateExcelService._parse_json(formatted) == {"a": 1, "b": 2}

    def test_non_json_string_passthrough(self):
        assert TemplateExcelService._format_json_for_excel("plain text") == "plain text"

    def test_compact_json_with_unicode_escapes_renders_readable_text(self):
        compact = (
            '{"ar": "\\u062a\\u0645\\u0648\\u064a\\u0644 NS 2025 \\u0627\\u0644\\u0625\\u062c\\u0645\\u0627\\u0644\\u064a", '
            '"es": "NS 2025 Financiaci\\u00f3n Total", "fr": "Financement total NS 2025"}'
        )
        formatted = TemplateExcelService._format_json_for_excel(compact)
        assert formatted is not None
        assert '\\u062a' not in formatted
        assert '\\u00f3' not in formatted
        assert 'تم' in formatted
        assert 'Financiación' in formatted

    def test_dict_with_literal_unicode_escapes_in_values(self):
        value = {
            "ar": "\\u062a\\u0644\\u0642\\u064a\\u062a \\u0627\\u0644\\u062f\\u0639\\u0645",
            "es": "Apoyo recibido",
            "fr": "Soutien re\\u00e7u",
        }
        formatted = TemplateExcelService._format_json_for_excel(value)
        assert formatted is not None
        assert 'تلقيت الدعم' in formatted
        assert 'Soutien reçu' in formatted
        assert '\\u062a' not in formatted
        assert '\\u00e7' not in formatted

    def test_double_encoded_json_string_is_unwrapped(self):
        inner = json.dumps(
            {"ar": "تم", "fr": "Soutien reçu"},
            ensure_ascii=True,
        )
        outer = json.dumps(inner)
        formatted = TemplateExcelService._format_json_for_excel(outer)
        assert formatted is not None
        assert 'تم' in formatted
        assert 'Soutien reçu' in formatted
        assert '\\u062a' not in formatted


# ---------------------------------------------------------------------------
# Helper method unit tests (no DB)
# ---------------------------------------------------------------------------

class TestHelperMethods:
    """Tests for helper/style methods that operate on openpyxl objects."""

    def _make_cell(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        return ws.cell(row=1, column=1)

    def test_style_header_cell_required(self):
        cell = self._make_cell()
        TemplateExcelService._style_header_cell(cell, is_required=True)
        # Red fill for required
        assert cell.fill.fgColor.rgb == TemplateExcelService.IFRC_COLORS['DARK_RED']
        assert cell.font.bold is True

    def test_style_header_cell_optional(self):
        cell = self._make_cell()
        TemplateExcelService._style_header_cell(cell, is_required=False)
        # Blue fill for optional
        assert cell.fill.fgColor.rgb == TemplateExcelService.IFRC_COLORS['DARK_BLUE']

    def test_auto_size_columns_runs_without_error(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Hello World")
        ws.cell(row=2, column=1, value="Short")
        # Should not raise
        TemplateExcelService._auto_size_columns(ws, 1)
        assert ws.column_dimensions['A'].width > 0

    def test_create_excel_table_normal(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="id")
        ws.cell(row=2, column=1, value=1)
        # Should create table without error
        TemplateExcelService._create_excel_table(ws, "TestTable", 1, 2)

    def test_create_excel_table_header_only_adds_empty_row(self):
        """num_rows=1 (header only) should use num_rows=2 (add empty data row)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="id")
        TemplateExcelService._create_excel_table(ws, "HeaderOnly", 1, 1)
        # Table should have been created
        assert len(ws.tables) == 1

    def test_create_excel_table_zero_rows_skipped(self):
        """num_rows < 1 => no table created."""
        wb = openpyxl.Workbook()
        ws = wb.active
        TemplateExcelService._create_excel_table(ws, "ZeroTable", 1, 0)
        assert len(ws.tables) == 0

    def test_add_dropdown_validation_empty_options(self):
        """Empty options list should skip adding validation."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Should not raise
        TemplateExcelService._add_dropdown_validation(ws, 1, [], start_row=2, end_row=10)

    def test_add_dropdown_validation_with_options(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="section_type")
        TemplateExcelService._add_dropdown_validation(
            ws, 1,
            TemplateExcelService.DROPDOWN_OPTIONS['section_type'],
            start_row=2, end_row=5
        )
        # Validation was added
        assert len(ws.data_validations.dataValidation) > 0

    def test_add_duplicate_highlighting_less_than_3_rows(self):
        """Less than 3 rows => no conditional formatting added."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="id")
        ws.cell(row=2, column=1, value=1)
        TemplateExcelService._add_duplicate_highlighting(ws, 'id', ['id'], 2)
        # No rule should be added
        assert not ws.conditional_formatting._cf_rules

    def test_add_duplicate_highlighting_column_not_in_headers(self, app):
        """Column name not in headers => silently skipped."""
        with app.app_context():
            wb = openpyxl.Workbook()
            ws = wb.active
            for i in range(1, 6):
                ws.cell(row=i, column=1, value=i)
            TemplateExcelService._add_duplicate_highlighting(ws, 'nonexistent', ['id'], 5)

    def test_add_duplicate_highlighting_normal(self, app):
        with app.app_context():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="id")
            for i in range(2, 6):
                ws.cell(row=i, column=1, value=i - 1)
            TemplateExcelService._add_duplicate_highlighting(ws, 'id', ['id'], 5)

    def test_get_item_type_options_includes_standard_types(self, app):
        with app.app_context():
            opts = TemplateExcelService._get_item_type_options()
            assert 'indicator' in opts
            assert 'question' in opts
            assert 'document_field' in opts
            assert 'matrix' in opts

    def test_get_item_type_options_with_plugin_manager(self, app):
        with app.app_context():
            plugin_mgr = MagicMock()
            plugin_mgr.list_active_field_types.return_value = ['custom_type']
            with patch.object(app, 'plugin_manager', plugin_mgr, create=True):
                opts = TemplateExcelService._get_item_type_options()
                assert 'plugin_custom_type' in opts

    def test_get_item_type_options_plugin_exception_handled(self, app):
        """Plugin manager raising exception is handled gracefully."""
        with app.app_context():
            plugin_mgr = MagicMock()
            plugin_mgr.list_active_field_types.side_effect = Exception("Plugin error")
            with patch.object(app, 'plugin_manager', plugin_mgr, create=True):
                opts = TemplateExcelService._get_item_type_options()
                assert 'indicator' in opts  # Standard types still there

    def test_get_type_options_from_database_with_existing_types(self, db_session, app):
        with app.app_context():
            opts = TemplateExcelService._get_type_options_from_database()
            assert isinstance(opts, list)
            assert len(opts) > 0

    def test_get_type_options_from_database_exception_returns_defaults(self, app):
        with app.app_context():
            with patch("app.services.templates.excel_export.db") as mock_db:
                mock_db.session.query.side_effect = Exception("DB error")
                opts = TemplateExcelService._get_type_options_from_database()
                assert 'Number' in opts

    def test_add_sheet_reference_dropdown_column_not_found(self, app):
        with app.app_context():
            wb = openpyxl.Workbook()
            source = wb.create_sheet("Pages")
            source.cell(row=1, column=1, value="id")
            target = wb.create_sheet("Sections")
            # Column 'page_id' not in source headers
            TemplateExcelService._add_sheet_reference_dropdown(
                target, 1, source, 'nonexistent_col', start_row=2, end_row=10
            )

    def test_add_sheet_reference_dropdown_normal(self, app):
        with app.app_context():
            wb = openpyxl.Workbook()
            source = wb.create_sheet("Pages")
            source.cell(row=1, column=1, value="id")
            source.cell(row=2, column=1, value=1)
            target = wb.create_sheet("Sections")
            # Should add a data validation referencing Pages
            TemplateExcelService._add_sheet_reference_dropdown(
                target, 1, source, 'id', start_row=2, end_row=10
            )

    def test_add_sheet_reference_dropdown_sheet_name_with_space(self, app):
        with app.app_context():
            wb = openpyxl.Workbook()
            source = wb.create_sheet("My Pages")
            source.cell(row=1, column=1, value="id")
            target = wb.create_sheet("Sections")
            TemplateExcelService._add_sheet_reference_dropdown(
                target, 1, source, 'id', start_row=2, end_row=10
            )


# ---------------------------------------------------------------------------
# _build_item_db_to_export_map
# ---------------------------------------------------------------------------

class TestBuildItemDbToExportMap:
    def test_map_sequential_starting_from_one(self, db_session, app):
        with app.app_context():
            template = create_test_template(db_session, name="BuildMap Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            section = create_test_section(db_session, template, version=version, order=1)
            item1 = create_test_item(db_session, section, template, version=version, order=1)
            item2 = create_test_item(db_session, section, template, version=version, order=2)

            mapping = TemplateExcelService._build_item_db_to_export_map(template, version)
            assert item1.id in mapping
            assert item2.id in mapping
            # Sequential from 1
            assert sorted(mapping.values()) == list(range(1, len(mapping) + 1))

    def test_empty_template_returns_empty_map(self, db_session, app):
        with app.app_context():
            template = create_test_template(db_session, name="EmptyMap Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            mapping = TemplateExcelService._build_item_db_to_export_map(template, version)
            assert mapping == {}

    def test_items_ordered_by_section_then_item_order(self, db_session, app):
        """Export order must follow section order, then item order within each section."""
        with app.app_context():
            template = create_test_template(db_session, name="SectionOrder Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()

            section_first = create_test_section(
                db_session, template, version=version, name="First Section", order=1
            )
            section_second = create_test_section(
                db_session, template, version=version, name="Second Section", order=2
            )

            # Lower item.order in the later section must not appear before first section items.
            item_second_section = create_test_item(
                db_session, section_second, template, version=version,
                order=1, label="Second section item"
            )
            item_first_section = create_test_item(
                db_session, section_first, template, version=version,
                order=100, label="First section item"
            )

            items = TemplateExcelService._get_items_for_version(template, version)
            assert [item.id for item in items] == [item_first_section.id, item_second_section.id]


# ---------------------------------------------------------------------------
# export_template integration tests
# ---------------------------------------------------------------------------

class TestExportTemplate:
    """Integration tests for export_template – exercises all _export_* methods."""

    def _create_full_template(self, db_session):
        """Create a template with page, section, and item."""
        template = create_test_template(db_session, name="Full Export Template")
        version = db_session.query(FormTemplateVersion).filter_by(
            id=template.published_version_id
        ).first()
        # Update version with various options to exercise all export paths
        version.is_paginated = True
        version.add_to_self_report = True
        version.display_order_visible = True
        version.enable_export_pdf = True
        version.enable_export_excel = True
        version.enable_import_excel = True
        version.enable_ai_validation = True
        version.name_translations = {"fr": "Modèle de test"}
        version.variables = {"var1": "value1"}
        db_session.add(version)
        db_session.commit()
        db_session.refresh(version)

        page = FormPage(
            template_id=template.id,
            version_id=version.id,
            name="Test Page",
            order=1,
            name_translations={"fr": "Page Test"},
        )
        db_session.add(page)
        db_session.flush()

        section = create_test_section(
            db_session, template, version=version,
            name="Export Section", order=1,
            section_type="standard"
        )
        section.page_id = page.id
        section.name_translations = {"fr": "Section Test"}
        section.relevance_condition = json.dumps({"item_id": "1"})
        db_session.add(section)

        section2 = create_test_section(
            db_session, template, version=version,
            name="Sub Section", order=2,
            section_type="standard",
            parent_section_id=section.id
        )
        db_session.add(section2)

        item = create_test_item(
            db_session, section, template, version=version,
            item_type="indicator", order=1,
            label="Test Indicator"
        )
        item.relevance_condition = json.dumps({"item_id": "1"})
        item.validation_condition = json.dumps({"item_id": "1"})
        item.config = {"key": "val"}
        item.label_translations = {"fr": "Indicateur Test"}
        db_session.add(item)

        db_session.commit()
        db_session.refresh(template)
        db_session.refresh(version)
        return template, version, page, section, item

    def test_export_returns_bytes_io(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template, version, page, section, item = self._create_full_template(db_session)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.export_template(template.id)

            assert isinstance(result, io.BytesIO)
            # Verify it's a valid Excel file
            result.seek(0)
            wb = openpyxl.load_workbook(result)
            assert "Template" in wb.sheetnames
            assert "Pages" in wb.sheetnames
            assert "Sections" in wb.sheetnames
            assert "Items" in wb.sheetnames
            assert "Instructions" in wb.sheetnames

    def test_export_with_specific_version_id(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template, version, page, section, item = self._create_full_template(db_session)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.export_template(template.id, version_id=version.id)

            assert isinstance(result, io.BytesIO)

    def test_export_with_nonexistent_version_falls_back(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template, version, page, section, item = self._create_full_template(db_session)

            with app.test_request_context():
                login_user(admin)
                # version_id=999999 should fall back to published version
                result = TemplateExcelService.export_template(template.id, version_id=999999)

            assert isinstance(result, io.BytesIO)

    def test_export_no_version_raises_for_template_with_no_version(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            # Create template without a version
            template = FormTemplate()
            db_session.add(template)
            db_session.commit()

            with app.test_request_context():
                login_user(admin)
                with pytest.raises(ValueError, match="No version found"):
                    TemplateExcelService.export_template(template.id)

    def test_export_template_without_pages(self, db_session, app):
        """Template with no pages should still export successfully."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="No Pages Export Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            section = create_test_section(db_session, template, version=version)
            item = create_test_item(db_session, section, template, version=version)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.export_template(template.id)

            result.seek(0)
            wb = openpyxl.load_workbook(result)
            assert "Pages" in wb.sheetnames

    def test_export_draft_version_when_no_published(self, db_session, app):
        """Template with only a draft version should use latest."""
        with app.app_context():
            admin = create_test_admin(db_session)
            # Create a template with draft version only
            template = FormTemplate()
            db_session.add(template)
            db_session.flush()
            version = FormTemplateVersion(
                template_id=template.id,
                version_number=1,
                status='draft',
                name='Draft Only',
            )
            db_session.add(version)
            db_session.commit()

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.export_template(template.id)

            assert isinstance(result, io.BytesIO)

    def test_export_produces_metadata_sheet(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template, version, _, _, _ = self._create_full_template(db_session)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.export_template(template.id)

            result.seek(0)
            wb = openpyxl.load_workbook(result)
            # _Metadata is veryHidden but should exist
            assert "_Metadata" in wb.sheetnames

    def test_export_sections_with_many_items(self, db_session, app):
        """Exercises the items export with multiple items for conditional formatting."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Many Items Export")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            section = create_test_section(db_session, template, version=version)
            for i in range(3):
                create_test_item(
                    db_session, section, template, version=version,
                    order=i + 1, label=f"Item {i}"
                )

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.export_template(template.id)

            result.seek(0)
            wb = openpyxl.load_workbook(result)
            ws_items = wb["Items"]
            # Header row + 3 data rows
            assert ws_items.max_row >= 4


# ---------------------------------------------------------------------------
# import_template integration tests
# ---------------------------------------------------------------------------

class TestImportTemplate:
    """Integration tests for import_template."""

    def test_import_missing_required_sheets_fails(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import MissingSheets Template")

            # Workbook missing Template sheet
            wb = openpyxl.Workbook()
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(template.id, buf)

            assert result['success'] is False
            assert "Missing required sheets" in result['message']

    def test_import_with_no_version_fails(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            # Template with no version
            template = FormTemplate()
            db_session.add(template)
            db_session.commit()

            buf = _make_export_workbook_bytes()

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(template.id, buf)

            assert result['success'] is False
            assert "No version" in result['message']

    def test_import_basic_success(self, db_session, app):
        """Import a minimal template with one section and one item."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import Basic Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()

            # Build a fresh draft version for import
            draft = create_test_draft_version(db_session, template, name="Draft for Import")

            buf = _make_export_workbook_bytes(
                template_name="Imported Template",
                pages=[],
                sections=[
                    [1, "Section A", 1.0, None, None, "standard",
                     None, None, None, False, False, None, None,
                     None, None, None, False]
                ],
                items=[
                    [1, 1, "indicator", "Item A", 1.0,
                     None, False, None, None, "Number", None,
                     None, None, None, None, None, None, None,
                     None, None, None, None, None]
                ],
            )

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is True
            assert result['created_count']['sections'] == 1

    def test_import_creates_draft_from_published(self, db_session, app):
        """When target version is published, a new draft is created automatically."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import AutoDraft Template")

            buf = _make_export_workbook_bytes(
                template_name="AutoDraft Import",
                sections=[
                    [1, "Auto Section", 1.0, None, None, "standard",
                     None, None, None, False, False, None, None,
                     None, None, None, False]
                ],
            )

            with app.test_request_context():
                login_user(admin)
                # Use the published version_id => should auto-create a draft
                result = TemplateExcelService.import_template(template.id, buf)

            assert result['success'] is True

    def test_parse_import_version_mode(self, db_session, app):
        with app.app_context():
            template = create_test_template(db_session, name="Parse Mode Template")
            published = template.published_version
            draft = create_test_draft_version(db_session, template, name="Parse Draft")

            assert TemplateExcelService._parse_import_version_mode('create_draft', published) is True
            assert TemplateExcelService._parse_import_version_mode('current_version', published) is False
            assert TemplateExcelService._parse_import_version_mode(None, published) is True
            assert TemplateExcelService._parse_import_version_mode(None, draft) is False

    def test_get_or_create_draft_for_import_reuses_existing(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Reuse Draft Import Template")
            published = template.published_version
            existing = create_test_draft_version(db_session, template, name="Existing Draft")

            with app.test_request_context():
                login_user(admin)
                draft = TemplateExcelService._get_or_create_draft_for_import(template, published)

            assert draft.id == existing.id
            assert FormTemplateVersion.query.filter_by(
                template_id=template.id, status='draft'
            ).count() == 1

    def test_get_or_create_draft_for_import_creates_new(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Create Draft Import Template")
            published = template.published_version

            with app.test_request_context():
                login_user(admin)
                draft = TemplateExcelService._get_or_create_draft_for_import(template, published)

            assert draft.status == 'draft'
            assert draft.based_on_version_id == published.id

    def test_resolve_import_target_version_id_uses_existing_draft(self, db_session, app):
        with app.app_context():
            template = create_test_template(db_session, name="Resolve Draft Target Template")
            published = template.published_version
            draft = create_test_draft_version(db_session, template, name="Existing Draft")

            resolved = TemplateExcelService.resolve_import_target_version_id(
                template.id,
                published.id,
                import_version_mode='create_draft',
            )
            assert resolved == draft.id

    def test_resolve_import_target_version_id_keeps_published_for_current_mode(self, db_session, app):
        with app.app_context():
            template = create_test_template(db_session, name="Resolve Published Target Template")
            published = template.published_version

            resolved = TemplateExcelService.resolve_import_target_version_id(
                template.id,
                published.id,
                import_version_mode='current_version',
            )
            assert resolved == published.id

    def test_import_with_pages(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import Pages Template")
            draft = create_test_draft_version(db_session, template, name="Draft Pages")

            buf = _make_export_workbook_bytes(
                pages=[
                    [1, "Page 1", 1, None]
                ],
                sections=[
                    [1, "Sec with Page", 1.0, None, 1, "standard",
                     None, None, None, False, False, None, None,
                     None, None, None, False]
                ],
                items=[
                    [1, 1, "question", "Q1", 1.0,
                     None, False, None, None, "text", None,
                     None, None, None, None, None, None, None,
                     None, None, None, None, None]
                ],
            )

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is True
            assert result['created_count']['pages'] >= 1

    def test_import_with_unrecognized_sheets_ignored(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import ExtraSheets Template")
            draft = create_test_draft_version(db_session, template, name="Draft Extra")

            buf = _make_export_workbook_bytes(
                sections=[
                    [1, "Sec", 1.0, None, None, "standard",
                     None, None, None, False, False, None, None,
                     None, None, None, False]
                ],
                extra_sheets=["CustomData", "SomeOtherSheet"],
            )

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is True

    def test_import_bad_pages_headers_returns_error(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import BadPageHdr Template")
            draft = create_test_draft_version(db_session, template, name="Draft BadHdr")

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws_t = wb.create_sheet("Template")
            _append_template_sheet_row_layout(ws_t, "Name")
            ws_p = wb.create_sheet("Pages")
            ws_p.append(["wrong", "headers"])  # Bad headers
            ws_s = wb.create_sheet("Sections")
            ws_s.append(TemplateExcelService.get_section_columns())
            ws_i = wb.create_sheet("Items")
            ws_i.append(TemplateExcelService.get_item_columns())
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is False

    def test_import_bad_sections_headers_returns_error(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import BadSecHdr Template")
            draft = create_test_draft_version(db_session, template, name="Draft BadSecHdr")

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws_t = wb.create_sheet("Template")
            _append_template_sheet_row_layout(ws_t, "Name")
            ws_p = wb.create_sheet("Pages")
            ws_p.append(TemplateExcelService.get_page_columns())
            ws_s = wb.create_sheet("Sections")
            ws_s.append(["wrong", "headers"])  # Bad headers
            ws_i = wb.create_sheet("Items")
            ws_i.append(TemplateExcelService.get_item_columns())
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is False

    def test_import_bad_items_headers_returns_error(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import BadItemHdr Template")
            draft = create_test_draft_version(db_session, template, name="Draft BadItemHdr")

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws_t = wb.create_sheet("Template")
            _append_template_sheet_row_layout(ws_t, "Name")
            ws_p = wb.create_sheet("Pages")
            ws_p.append(TemplateExcelService.get_page_columns())
            ws_s = wb.create_sheet("Sections")
            ws_s.append(TemplateExcelService.get_section_columns())
            ws_i = wb.create_sheet("Items")
            ws_i.append(["bad", "columns"])  # Bad headers
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is False

    def test_import_template_metadata_missing_required_column(self, db_session, app):
        """Template sheet missing the 'name' required column."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import MissingName Template")
            draft = create_test_draft_version(db_session, template, name="Draft MissingName")

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws_t = wb.create_sheet("Template")
            ws_t.append(["description"])  # Missing 'name'
            ws_t.append(["some desc"])
            ws_p = wb.create_sheet("Pages")
            ws_p.append(TemplateExcelService.get_page_columns())
            ws_s = wb.create_sheet("Sections")
            ws_s.append(TemplateExcelService.get_section_columns())
            ws_i = wb.create_sheet("Items")
            ws_i.append(TemplateExcelService.get_item_columns())
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is False

    def test_import_page_missing_id_records_error(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import PageNoID Template")
            draft = create_test_draft_version(db_session, template, name="Draft PageNoID")

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws_t = wb.create_sheet("Template")
            _append_template_sheet_row_layout(ws_t, "Name")
            ws_p = wb.create_sheet("Pages")
            ws_p.append(TemplateExcelService.get_page_columns())
            # Page row with None id
            ws_p.append([None, "Page 1", 1, None])
            ws_s = wb.create_sheet("Sections")
            ws_s.append(TemplateExcelService.get_section_columns())
            ws_i = wb.create_sheet("Items")
            ws_i.append(TemplateExcelService.get_item_columns())
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is False
            assert any("Missing export ID" in e for e in result['errors'])

    def test_import_section_missing_id_records_error(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import SecNoID Template")
            draft = create_test_draft_version(db_session, template, name="Draft SecNoID")

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws_t = wb.create_sheet("Template")
            _append_template_sheet_row_layout(ws_t, "Name")
            ws_p = wb.create_sheet("Pages")
            ws_p.append(TemplateExcelService.get_page_columns())
            ws_s = wb.create_sheet("Sections")
            ws_s.append(TemplateExcelService.get_section_columns())
            # Section with None id
            ws_s.append([None, "Sec", 1.0, None, None, "standard",
                         None, None, None, False, False, None, None,
                         None, None, None, False])
            ws_i = wb.create_sheet("Items")
            ws_i.append(TemplateExcelService.get_item_columns())
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is False

    def test_import_item_unresolvable_section_id_records_error(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import ItemBadSec Template")
            draft = create_test_draft_version(db_session, template, name="Draft ItemBadSec")

            buf = _make_export_workbook_bytes(
                items=[
                    [1, 999, "indicator", "Orphan Item", 1.0,
                     None, False, None, None, "Number", None,
                     None, None, None, None, None, None, None,
                     None, None, None, None, None]
                ],
            )

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is False
            assert any("Could not resolve section_id" in e for e in result['errors'])

    def test_import_updates_existing_section(self, db_session, app):
        """When a section with matching order+name already exists, it should be updated."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import UpdateSec Template")
            draft = create_test_draft_version(db_session, template, name="Draft UpdateSec")
            # Pre-create the section in the draft version
            existing_section = create_test_section(
                db_session, template, version=draft,
                name="Existing Section", order=1,
                section_type="standard"
            )

            buf = _make_export_workbook_bytes(
                sections=[
                    [1, "Existing Section", 1.0, None, None, "standard",
                     None, None, None, False, False, None, None,
                     None, None, None, False]
                ],
                items=[
                    [1, 1, "indicator", "Indicator Item", 1.0,
                     None, False, None, None, "Number", None,
                     None, None, None, None, None, None, None,
                     None, None, None, None, None]
                ],
            )

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is True

    def test_import_replaces_removed_subsections(self, db_session, app):
        """Excel import replaces structure; removed subsections and stale items must not remain."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import Replace Struct Template")
            draft = create_test_draft_version(db_session, template, name="Draft Replace")

            parent = create_test_section(
                db_session, template, version=draft,
                name="Main Section", order=1, section_type="standard",
            )
            create_test_section(
                db_session, template, version=draft,
                name="Old Subsection", order=1, section_type="standard",
                parent_section_id=parent.id,
            )

            section_cols = _excel_columns(app)['sections']
            item_cols = _excel_columns(app)['items']
            section_row = _blank_row(section_cols)
            section_row[section_cols.index('id')] = 1
            section_row[section_cols.index('order')] = 1.0
            section_row[section_cols.index('name')] = "Main Section"
            section_row[section_cols.index('section_type')] = "dynamic"
            item_row = _blank_row(item_cols)
            item_row[item_cols.index('id')] = 1
            item_row[item_cols.index('section_id')] = 1
            item_row[item_cols.index('item_type')] = "indicator"
            item_row[item_cols.index('order')] = 1.0
            item_row[item_cols.index('label')] = "New Item"
            item_row[item_cols.index('type')] = "Number"

            buf = _make_export_workbook_bytes(
                sections=[section_row],
                items=[item_row],
                app=app,
            )

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is True
            sections = FormSection.query.filter_by(
                template_id=template.id, version_id=draft.id
            ).all()
            assert {s.name for s in sections} == {"Main Section"}
            main = next(s for s in sections if s.name == "Main Section")
            assert main.section_type == "dynamic"
            items = FormItem.query.filter_by(
                template_id=template.id, version_id=draft.id
            ).all()
            assert len(items) == 1
            assert items[0].label == "New Item"

    def test_import_with_parent_section_references(self, db_session, app):
        """Parent section references are resolved correctly in third pass."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import ParentSec Template")
            draft = create_test_draft_version(db_session, template, name="Draft ParentSec")

            buf = _make_export_workbook_bytes(
                sections=[
                    [1, "Parent", 1.0, None, None, "standard",
                     None, None, None, False, False, None, None,
                     None, None, None, False],
                    [2, "Child", 2.0, 1, None, "standard",
                     None, None, None, False, False, None, None,
                     None, None, None, False],
                ],
                items=[
                    [1, 1, "indicator", "Parent Item", 1.0,
                     None, False, None, None, "Number", None,
                     None, None, None, None, None, None, None,
                     None, None, None, None, None],
                    [2, 2, "question", "Child Item", 1.0,
                     None, False, None, None, "text", None,
                     None, None, None, None, None, None, None,
                     None, None, None, None, None],
                ],
            )

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is True
            assert result['created_count']['sections'] >= 2

    def test_import_template_row_with_none_values(self, db_session, app):
        """Template sheet with an empty data row is gracefully skipped."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import EmptyRow Template")
            draft = create_test_draft_version(db_session, template, name="Draft EmptyRow")

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws_t = wb.create_sheet("Template")
            fields = TemplateExcelService.get_template_columns()
            ws_t.append(fields)
            ws_t.append([None] * len(fields))
            ws_p = wb.create_sheet("Pages")
            ws_p.append(TemplateExcelService.get_page_columns())
            ws_s = wb.create_sheet("Sections")
            ws_s.append(TemplateExcelService.get_section_columns())
            ws_i = wb.create_sheet("Items")
            ws_i.append(TemplateExcelService.get_item_columns())
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is True

    def test_import_exception_returns_failure(self, db_session, app):
        """A corrupted file triggers exception handler and returns failure."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import Exception Template")

            # Pass garbage bytes
            buf = io.BytesIO(b"this is not an excel file")

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(template.id, buf)

            assert result['success'] is False

    def test_import_with_name_translations_json(self, db_session, app):
        """Template sheet with name_translations JSON is applied to version."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import NameTrans Template")
            draft = create_test_draft_version(db_session, template, name="Draft NameTrans")

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws_t = wb.create_sheet("Template")
            _append_template_sheet_row_layout(
                ws_t,
                extra_values={
                    'name': 'Translated Template',
                    'name_fr': 'Modèle',
                    'variables': json.dumps({"v": "1"}),
                },
            )
            ws_p = wb.create_sheet("Pages")
            ws_p.append(TemplateExcelService.get_page_columns())
            ws_s = wb.create_sheet("Sections")
            ws_s.append(TemplateExcelService.get_section_columns())
            ws_i = wb.create_sheet("Items")
            ws_i.append(TemplateExcelService.get_item_columns())
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is True


# ---------------------------------------------------------------------------
# _clone_template_structure
# ---------------------------------------------------------------------------

class TestCloneTemplateStructure:
    def test_clone_creates_copies_of_pages_sections_items(self, db_session, app):
        with app.app_context():
            template = create_test_template(db_session, name="Clone Structure Template")
            src_version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            tgt_version = create_test_draft_version(db_session, template, name="Target Draft")

            page = FormPage(
                template_id=template.id,
                version_id=src_version.id,
                name="Src Page",
                order=1,
            )
            db_session.add(page)
            db_session.flush()

            section = create_test_section(
                db_session, template, version=src_version, name="Src Section", order=1
            )
            section.page_id = page.id
            db_session.add(section)
            db_session.flush()

            child = create_test_section(
                db_session, template, version=src_version,
                name="Src Child", order=2, parent_section_id=section.id
            )
            item = create_test_item(
                db_session, section, template, version=src_version,
                order=1, label="Src Item"
            )
            db_session.commit()

            TemplateExcelService._clone_template_structure(
                template.id, src_version.id, tgt_version.id
            )
            db_session.commit()

            # New sections exist for target version
            new_sections = FormSection.query.filter_by(
                template_id=template.id, version_id=tgt_version.id
            ).all()
            assert len(new_sections) >= 2

            # New pages exist
            new_pages = FormPage.query.filter_by(
                template_id=template.id, version_id=tgt_version.id
            ).all()
            assert len(new_pages) >= 1

            # New items exist
            new_items = FormItem.query.filter_by(
                template_id=template.id, version_id=tgt_version.id
            ).all()
            assert len(new_items) >= 1

    def test_clone_with_item_config_json_roundtrip_failure(self, db_session, app):
        """Item config that can't be JSON-serialized still clones via deepcopy fallback."""
        with app.app_context():
            template = create_test_template(db_session, name="Clone Config Fail Template")
            src_version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            tgt_version = create_test_draft_version(db_session, template, name="Clone Tgt Draft")

            section = create_test_section(
                db_session, template, version=src_version, name="Config Section", order=1
            )
            item = create_test_item(
                db_session, section, template, version=src_version,
                order=1, label="Config Item"
            )
            # Config is already a dict (stored in DB)
            item.config = {"nested": {"key": "value"}}
            db_session.add(item)
            db_session.commit()

            TemplateExcelService._clone_template_structure(
                template.id, src_version.id, tgt_version.id
            )
            db_session.commit()

            new_items = FormItem.query.filter_by(
                template_id=template.id, version_id=tgt_version.id
            ).all()
            assert len(new_items) >= 1


# ---------------------------------------------------------------------------
# _get_existing_items_lookup
# ---------------------------------------------------------------------------

class TestGetExistingItemsLookup:
    def test_returns_dict_with_items(self, db_session, app):
        with app.app_context():
            template = create_test_template(db_session, name="ExistLookup Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            section = create_test_section(db_session, template, version=version)
            item = create_test_item(
                db_session, section, template, version=version,
                item_type="indicator", order=1, label="Lookup Item"
            )

            lookup = TemplateExcelService._get_existing_items_lookup(template.id, version.id)
            key = (item.section_id, item.order, item.item_type, item.label or '')
            assert key in lookup
            assert lookup[key].id == item.id

    def test_returns_empty_for_no_items(self, db_session, app):
        with app.app_context():
            template = create_test_template(db_session, name="ExistLookupEmpty Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            lookup = TemplateExcelService._get_existing_items_lookup(template.id, version.id)
            assert lookup == {}


# ---------------------------------------------------------------------------
# Round-trip export -> import
# ---------------------------------------------------------------------------

class TestExportImportRoundTrip:
    """Full round-trip: export a template, then import it into a new draft."""

    def test_round_trip_preserves_section_count(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="RoundTrip Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()

            s1 = create_test_section(db_session, template, version=version, name="Sec 1", order=1)
            s2 = create_test_section(db_session, template, version=version, name="Sec 2", order=2)
            create_test_item(
                db_session, s1, template, version=version, order=1, label="Item 1"
            )

            with app.test_request_context():
                login_user(admin)
                export_buf = TemplateExcelService.export_template(template.id)

            # Now import into a new draft (auto-created since published version will be used)
            export_buf.seek(0)

            # Create a new import target (draft version) so we don't mutate the original
            draft = create_test_draft_version(db_session, template, name="Import Draft RT")

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, export_buf, version_id=draft.id
                )

            assert result['success'] is True

    def test_round_trip_preserves_matrix_column_groups(self, db_session, app):
        """Matrix column groups and per-column group keys survive export then import."""
        matrix_config = {
            'is_required': False,
            'layout_column_width': '12',
            'matrix_config': {
                'type': 'matrix',
                'row_mode': 'manual',
                'columns': [
                    {'name': 'SP1', 'type': 'number', 'group': 'Funding'},
                    {'name': 'SP2', 'type': 'number', 'group': 'Funding'},
                    {'name': 'Other', 'type': 'number'},
                ],
                'column_groups': {'Funding': {'fr': 'Financement'}},
                'rows': [{'text': 'Row 1'}],
            },
        }
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Matrix Groups RT")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            section = create_test_section(
                db_session, template, version=version, name="Matrix Sec", order=1
            )
            create_test_item(
                db_session,
                section,
                template,
                version=version,
                item_type='matrix',
                order=1,
                label='Funding Matrix',
                config=matrix_config,
            )

            with app.test_request_context():
                login_user(admin)
                export_buf = TemplateExcelService.export_template(template.id)

            draft = create_test_draft_version(db_session, template, name="Matrix Import Draft")
            export_buf.seek(0)
            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, export_buf, version_id=draft.id
                )

            assert result['success'] is True
            imported_item = (
                db_session.query(FormItem)
                .filter_by(template_id=template.id, version_id=draft.id, item_type='matrix')
                .first()
            )
            assert imported_item is not None
            mc = imported_item.config.get('matrix_config') or {}
            assert mc.get('column_groups', {}).get('Funding', {}).get('fr') == 'Financement'
            grouped = [c for c in mc.get('columns', []) if c.get('group') == 'Funding']
            assert len(grouped) == 2
            assert {c['name'] for c in grouped} == {'SP1', 'SP2'}
        """Rules with item_id references are rewritten correctly."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="RoundTrip Rules Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()

            section = create_test_section(
                db_session, template, version=version, name="Rules Sec", order=1
            )
            item1 = create_test_item(
                db_session, section, template, version=version,
                order=1, label="Rule Source Item"
            )
            item2 = create_test_item(
                db_session, section, template, version=version,
                order=2, label="Rule Target Item"
            )
            # Add a relevance condition referencing item1
            item2.relevance_condition = json.dumps({"item_id": str(item1.id)})
            db_session.add(item2)
            db_session.commit()

            with app.test_request_context():
                login_user(admin)
                export_buf = TemplateExcelService.export_template(template.id)

            draft = create_test_draft_version(db_session, template, name="Import Draft Rules")

            export_buf.seek(0)
            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, export_buf, version_id=draft.id
                )

            assert result['success'] is True


# ---------------------------------------------------------------------------
# Import indicator_bank_id validation
# ---------------------------------------------------------------------------

class TestImportIndicatorBankValidation:
    """Tests for indicator_bank_id validation logic during item import."""

    def test_import_item_with_null_indicator_bank_id_flagged(self, db_session, app):
        """indicator type with null indicator_bank_id should flag import issue in config."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import IB Null Template")
            draft = create_test_draft_version(db_session, template, name="Draft IB Null")

            buf = _make_export_workbook_bytes(
                sections=[
                    [1, "IB Section", 1.0, None, None, "standard",
                     None, None, None, False, False, None, None,
                     None, None, None, False]
                ],
                items=[
                    [1, 1, "indicator", "Null IB Item", 1.0,
                     None, False, None, None, "Number", None,
                     None, None, None, None, None, None, None,
                     None, None, None, None, None]
                    # indicator_bank_id at index 8 = None
                ],
            )

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is True
            # Item should be created with _import_issues in config
            new_items = FormItem.query.filter_by(
                template_id=template.id, version_id=draft.id
            ).all()
            assert len(new_items) >= 1

    def test_import_item_with_missing_indicator_bank_reference(self, db_session, app):
        """indicator type referencing a non-existent IndicatorBank ID."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import IB Missing Template")
            draft = create_test_draft_version(db_session, template, name="Draft IB Missing")

            buf = _make_export_workbook_bytes(
                sections=[
                    [1, "IB Section 2", 1.0, None, None, "standard",
                     None, None, None, False, False, None, None,
                     None, None, None, False]
                ],
                items=[
                    [1, 1, "indicator", "Missing IB Item", 1.0,
                     None, False, None, 9999999, "Number", None,
                     None, None, None, None, None, None, None,
                     None, None, None, None, None]
                    # indicator_bank_id=9999999 (non-existent)
                ],
            )

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is True

    def test_import_non_indicator_item_clears_indicator_issues(self, db_session, app):
        """Non-indicator item with any indicator_bank_id doesn't trigger validation."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Import NonInd Template")
            draft = create_test_draft_version(db_session, template, name="Draft NonInd")

            buf = _make_export_workbook_bytes(
                sections=[
                    [1, "NonInd Section", 1.0, None, None, "standard",
                     None, None, None, False, False, None, None,
                     None, None, None, False]
                ],
                items=[
                    [1, 1, "question", "Question Item", 1.0,
                     None, False, None, None, "text", None,
                     None, None, None, None, None, None, None,
                     None, None, None, None, None]
                ],
            )

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is True


# ---------------------------------------------------------------------------
# Export Instructions sheet content
# ---------------------------------------------------------------------------

class TestExportInstructionsSheet:
    def test_instructions_sheet_has_title(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Instructions Sheet Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            section = create_test_section(db_session, template, version=version)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.export_template(template.id)

            result.seek(0)
            wb = openpyxl.load_workbook(result)
            ws = wb["Instructions"]
            # First cell should be the title
            title = ws.cell(row=1, column=1).value
            assert "Instructions" in (title or "")

    def test_instructions_sheet_has_required_columns_info(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Instructions Content Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            section = create_test_section(db_session, template, version=version)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.export_template(template.id)

            result.seek(0)
            wb = openpyxl.load_workbook(result)
            ws = wb["Instructions"]
            # Check that some content exists beyond row 1
            has_content = any(
                ws.cell(row=r, column=1).value
                for r in range(2, ws.max_row + 1)
            )
            assert has_content


# ---------------------------------------------------------------------------
# Flat translation columns
# ---------------------------------------------------------------------------

class TestFlatTranslationColumns:
    def test_export_flattens_label_translations_into_columns(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Flat Translations Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            section = create_test_section(db_session, template, version=version)
            item = create_test_item(
                db_session, section, template, version=version,
                label="Support received", order=1,
            )
            item.label_translations = {
                'ar': 'تلقيت الدعم',
                'fr': 'Soutien reçu',
                'es': 'Apoyo recibido',
            }
            db_session.add(item)
            db_session.commit()

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.export_template(template.id)

            result.seek(0)
            wb = openpyxl.load_workbook(result)
            items_sheet = wb['Items']
            headers = [cell.value for cell in items_sheet[1]]
            assert 'label_ar' in headers
            assert 'label_fr' in headers
            row = next(items_sheet.iter_rows(min_row=2, values_only=True))
            row_data = dict(zip(headers, row))
            assert row_data['label'] == 'Support received'
            assert row_data['label_ar'] == 'تلقيت الدعم'
            assert row_data['label_fr'] == 'Soutien reçu'
            assert 'label_translations' not in headers


class TestResolveSheetHeaders:
    def test_accepts_reordered_translation_columns(self, app):
        with app.app_context():
            pages_cols = TemplateExcelService.get_page_columns()
            reordered = ['id', 'order', 'name', 'name_fr', 'name_es', 'name_ar']
            assert reordered != pages_cols or set(reordered) == set(pages_cols)
            _, legacy, err = TemplateExcelService._resolve_sheet_headers('Pages', reordered)
            assert err is None
            assert legacy is False

            items_cols = TemplateExcelService.get_item_columns()
            got_items = list(items_cols)
            if 'label_fr' in got_items and 'label_es' in got_items:
                fr_idx = got_items.index('label_fr')
                es_idx = got_items.index('label_es')
                got_items[fr_idx], got_items[es_idx] = got_items[es_idx], got_items[fr_idx]
            _, legacy, err = TemplateExcelService._resolve_sheet_headers('Items', got_items)
            assert err is None
            assert legacy is False

    def test_reports_missing_required_columns(self, app):
        with app.app_context():
            _, _, err = TemplateExcelService._resolve_sheet_headers('Pages', ['id', 'order'])
            assert err is not None
            assert 'missing required column' in err.lower()
            assert 'name' in err.lower()

    def test_reports_unrecognized_columns(self, app):
        with app.app_context():
            cols = TemplateExcelService.get_page_columns() + ['unexpected_column']
            _, _, err = TemplateExcelService._resolve_sheet_headers('Pages', cols)
            assert err is not None
            assert 'unrecognized column' in err.lower()
            assert 'unexpected_column' in err

    def test_accepts_v2_export_without_stable_key_column(self, app):
        with app.app_context():
            section_headers = [c for c in TemplateExcelService.get_section_columns() if c != 'stable_key']
            _, legacy, err = TemplateExcelService._resolve_sheet_headers('Sections', section_headers)
            assert err is None
            assert legacy is False

            item_headers = [c for c in TemplateExcelService.get_item_columns() if c != 'stable_key']
            _, legacy, err = TemplateExcelService._resolve_sheet_headers('Items', item_headers)
            assert err is None
            assert legacy is False


class TestNormalizeMatrixItemConfig:
    def test_wraps_flat_matrix_config(self):
        flat = {
            'type': 'matrix',
            'columns': [
                {'name': 'SP1', 'type': 'number', 'group': 'Funding'},
                {'name': 'SP2', 'type': 'number', 'group': 'Funding'},
            ],
            'column_groups': {'Funding': {'fr': 'Financement'}},
            'rows': [{'text': 'Row 1'}],
        }
        normalized = TemplateExcelService._normalize_matrix_item_config('matrix', flat)
        assert 'matrix_config' in normalized
        mc = normalized['matrix_config']
        assert mc['column_groups']['Funding']['fr'] == 'Financement'
        assert mc['columns'][0]['group'] == 'Funding'

    def test_hoists_root_column_groups_into_matrix_config(self):
        config = {
            'matrix_config': {
                'type': 'matrix',
                'columns': [{'name': 'A', 'type': 'number', 'group': 'G1'}],
            },
            'column_groups': {'G1': {'ar': 'مجموعة'}},
        }
        normalized = TemplateExcelService._normalize_matrix_item_config('matrix', config)
        assert normalized['matrix_config']['column_groups']['G1']['ar'] == 'مجموعة'

    def test_non_matrix_items_unchanged(self):
        config = {'plugin_config': {'x': 1}}
        assert TemplateExcelService._normalize_matrix_item_config('indicator', config) == config

    def test_repair_column_groups_from_column_name_lists(self):
        mc = {
            'type': 'matrix',
            'columns': [
                {'name': 'SP1', 'type': 'number'},
                {'name': 'SP2', 'type': 'number'},
            ],
            'column_groups': {'Funding': ['SP1', 'SP2']},
        }
        repaired = TemplateExcelService._repair_matrix_column_groups(mc)
        assert repaired['columns'][0]['group'] == 'Funding'
        assert repaired['columns'][1]['group'] == 'Funding'
        assert repaired['column_groups']['Funding'] == {}

    def test_repair_column_groups_from_name_prefix(self):
        mc = {
            'type': 'matrix',
            'columns': [
                {'name': 'SP1 Planned', 'type': 'tick'},
                {'name': 'SP1 Supported', 'type': 'tick'},
            ],
            'column_groups': {'SP1': {'fr': 'PS1'}},
        }
        repaired = TemplateExcelService._repair_matrix_column_groups(mc)
        assert repaired['columns'][0]['group'] == 'SP1'
        assert repaired['columns'][1]['group'] == 'SP1'

    def test_parse_json_preserves_empty_object(self):
        assert TemplateExcelService._parse_json('{}') == {}


class TestMatrixImportMatching:
    def test_label_match_preferred_over_export_id(self):
        section_id = 10
        matrix_item = FormItem(
            section_id=section_id,
            template_id=1,
            version_id=1,
            item_type='matrix',
            label='Received Support',
            order=5.0,
        )
        matrix_item.id = 100

        wrong_export_item = FormItem(
            section_id=section_id,
            template_id=1,
            version_id=1,
            item_type='indicator',
            label='Some Indicator',
            order=1.0,
        )
        wrong_export_item.id = 200

        export_id_to_existing = {40: wrong_export_item}
        existing_items_lookup = {
            (section_id, 5.0, 'matrix', 'Received Support'): matrix_item,
        }

        found, method = TemplateExcelService._find_existing_item_for_import(
            section_id=section_id,
            item_type='matrix',
            item_label='Received Support',
            item_order=5.0,
            export_id=40,
            export_id_to_existing=export_id_to_existing,
            existing_items_lookup=existing_items_lookup,
        )
        assert found is matrix_item
        assert 'label' in method

    def test_clean_imported_matrix_config_strips_root_duplicates(self):
        config = {
            'is_required': False,
            'matrix_config': {
                'type': 'matrix',
                'columns': [{'name': 'A', 'group': 'G1'}],
                'column_groups': {'G1': {}},
            },
            'columns': [{'name': 'stale'}],
            'column_groups': {'stale': {}},
        }
        cleaned = TemplateExcelService._clean_imported_matrix_config(config)
        assert 'columns' not in cleaned
        assert 'column_groups' not in cleaned
        assert cleaned['matrix_config']['columns'][0]['group'] == 'G1'


# ---------------------------------------------------------------------------
# Template sheet row layout
# ---------------------------------------------------------------------------

class TestTemplateSheetRowLayout:
    def test_export_template_sheet_uses_field_value_rows(self, db_session, app):
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Row Layout Template")
            version = db_session.query(FormTemplateVersion).filter_by(
                id=template.published_version_id
            ).first()
            version.name_translations = {'fr': 'Modèle'}
            version.variables = {'key': 'val'}
            db_session.add(version)
            db_session.commit()

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.export_template(template.id)

            result.seek(0)
            wb = openpyxl.load_workbook(result)
            ws = wb['Template']
            headers = [cell.value for cell in ws[1]]
            assert headers == ['field', 'value']

            row_data = {
                str(row[0]): row[1]
                for row in ws.iter_rows(min_row=2, values_only=True)
                if row and row[0] is not None
            }
            assert row_data['name'] == template.name or row_data['name'] == version.name
            assert row_data.get('name_fr') == 'Modèle'
            assert row_data.get('variables') is not None

    def test_import_legacy_column_layout_template_sheet(self, db_session, app):
        """Older exports with column headers + single data row still import."""
        with app.app_context():
            admin = create_test_admin(db_session)
            template = create_test_template(db_session, name="Legacy Column Template")
            draft = create_test_draft_version(db_session, template, name="Draft Legacy")

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws_t = wb.create_sheet("Template")
            _append_template_sheet_legacy_columns(
                ws_t,
                template_name='Legacy Import Name',
                extra_values={'name_fr': 'Nom'},
            )
            ws_p = wb.create_sheet("Pages")
            ws_p.append(TemplateExcelService.get_page_columns())
            ws_s = wb.create_sheet("Sections")
            ws_s.append(TemplateExcelService.get_section_columns())
            ws_i = wb.create_sheet("Items")
            ws_i.append(TemplateExcelService.get_item_columns())
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            with app.test_request_context():
                login_user(admin)
                result = TemplateExcelService.import_template(
                    template.id, buf, version_id=draft.id
                )

            assert result['success'] is True
            db_session.refresh(draft)
            assert draft.name == 'Legacy Import Name'
            assert draft.name_translations.get('fr') == 'Nom'

    def test_validate_import_file_accepts_row_layout(self, app):
        buf = _make_export_workbook_bytes(template_name='Valid Row Layout')
        with app.app_context():
            result = TemplateExcelService.validate_import_file(buf)
        assert result['valid'] is True
        assert result['preview']['name'] == 'Valid Row Layout'


# ---------------------------------------------------------------------------
# Split-module layout (god-file refactor)
# ---------------------------------------------------------------------------

class TestModuleSplit:
    """Ensure TemplateExcelService public API survives the module split."""

    def test_orchestrator_inherits_export_and_import_mixins(self):
        from app.services.templates.excel_base import TemplateExcelBase
        from app.services.templates.excel_export import TemplateExcelExportMixin
        from app.services.templates.excel_import import TemplateExcelImportMixin
        from app.services.templates.matrix_import import TemplateExcelMatrixMixin

        assert issubclass(TemplateExcelService, TemplateExcelImportMixin)
        assert issubclass(TemplateExcelService, TemplateExcelExportMixin)
        assert issubclass(TemplateExcelImportMixin, TemplateExcelMatrixMixin)
        assert issubclass(TemplateExcelMatrixMixin, TemplateExcelBase)
        assert issubclass(TemplateExcelExportMixin, TemplateExcelBase)

    def test_public_methods_resolved_on_orchestrator(self):
        assert getattr(TemplateExcelService.export_template, '__func__', TemplateExcelService.export_template)
        assert TemplateExcelService.import_template
        assert TemplateExcelService.matrix_import_entry_log
        assert TemplateExcelService.validate_import_file

    def test_constants_available_on_orchestrator(self):
        assert TemplateExcelService.EXCEL_EXPORT_VERSION == 'V2'
        assert 'matrix' not in TemplateExcelService.REQUIRED_COLUMNS  # sanity: unrelated key absent


# ---------------------------------------------------------------------------
# Class-level constants / configuration
# ---------------------------------------------------------------------------

class TestClassConstants:
    def test_ifrc_colors_defined(self):
        assert 'RED' in TemplateExcelService.IFRC_COLORS
        assert 'WHITE' in TemplateExcelService.IFRC_COLORS
        assert 'DARK_RED' in TemplateExcelService.IFRC_COLORS

    def test_required_columns_defined(self):
        assert 'name' in TemplateExcelService.REQUIRED_COLUMNS['Template']
        assert 'id' in TemplateExcelService.REQUIRED_COLUMNS['Pages']

    def test_dropdown_options_defined(self):
        assert 'TRUE' in TemplateExcelService.DROPDOWN_OPTIONS['archived']
        assert 'FALSE' in TemplateExcelService.DROPDOWN_OPTIONS['archived']

    def test_column_definitions_defined(self):
        assert 'name' in TemplateExcelService.get_template_columns()
        assert 'id' in TemplateExcelService.get_page_columns()
        assert 'id' in TemplateExcelService.get_section_columns()
        assert 'id' in TemplateExcelService.get_item_columns()
        assert 'label_ar' in TemplateExcelService.get_item_columns()

    def test_item_columns_follow_logical_order(self):
        cols = TemplateExcelService.get_item_columns()
        langs = TemplateExcelService._get_translatable_languages()

        # Identifiers and placement before display text
        assert cols.index('id') < cols.index('section_id') < cols.index('item_type')
        assert cols.index('item_type') < cols.index('label')

        # Each translation block sits immediately after its English/base column
        for field in ('label', 'definition', 'description'):
            base_idx = cols.index(field)
            for lang in langs:
                lang_col = f'{field}_{lang}'
                assert lang_col in cols
                assert cols.index(lang_col) > base_idx
            next_fields = {
                'label': 'definition',
                'definition': 'description',
                'description': 'options_json',
            }
            assert cols.index(next_fields[field]) > base_idx + len(langs)

        # Question/list fields after text, indicator fields next, rules, config last
        assert cols.index('description') < cols.index('options_json')
        assert cols.index('options_json') < cols.index('indicator_bank_id')
        assert cols.index('indicator_bank_id') < cols.index('relevance_condition')
        assert cols.index('relevance_condition') < cols.index('config')
        assert cols[-1] == 'config'

    def test_page_columns_place_order_before_name(self):
        cols = TemplateExcelService.get_page_columns()
        assert cols.index('id') < cols.index('order') < cols.index('name')

    def test_excel_export_version(self):
        assert TemplateExcelService.EXCEL_EXPORT_VERSION == 'V2'

    def test_item_columns_exclude_stable_key(self):
        cols = TemplateExcelService.get_item_columns()
        assert 'stable_key' not in cols

    def test_section_columns_exclude_stable_key(self):
        cols = TemplateExcelService.get_section_columns()
        assert 'stable_key' not in cols


class TestStableKeyExcelRoundTrip:
    def test_export_omits_stable_key(self, app, db_session, admin_user):
        template = create_test_template(db_session, owner_id=admin_user.id)
        version = template.published_version
        section = create_test_section(db_session, template, version=version)
        item = create_test_item(db_session, section, template, version=version, item_type='question')
        item.stable_key = 'aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee'
        db_session.commit()

        with app.test_request_context():
            login_user(admin_user)
            buf = TemplateExcelService.export_template(template.id, version.id)
        wb = openpyxl.load_workbook(buf)
        assert 'stable_key' not in [cell.value for cell in wb['Sections'][1]]
        assert 'stable_key' not in [cell.value for cell in wb['Items'][1]]

    def test_import_inherits_published_stable_keys_without_excel_column(
        self, app, db_session, admin_user
    ):
        section_key = 'cccccccc-dddd-eeee-ffff-000000000001'
        item_key = 'dddddddd-eeee-ffff-0000-111111111111'
        template = create_test_template(db_session, owner_id=admin_user.id)
        version = template.published_version
        section = create_test_section(db_session, template, version=version)
        section.stable_key = section_key
        item = create_test_item(db_session, section, template, version=version, item_type='question')
        item.stable_key = item_key
        db_session.commit()

        with app.test_request_context():
            login_user(admin_user)
            export_buf = TemplateExcelService.export_template(template.id, version.id)

        draft = create_test_draft_version(db_session, template, name='Stable Key RT Draft')
        export_buf.seek(0)
        with app.test_request_context():
            login_user(admin_user)
            result = TemplateExcelService.import_template(
                template.id, export_buf, version_id=draft.id
            )
        assert result['success'] is True

        imported_section = db_session.query(FormSection).filter_by(
            template_id=template.id, version_id=draft.id
        ).one()
        imported_item = db_session.query(FormItem).filter_by(
            template_id=template.id, version_id=draft.id
        ).one()
        assert imported_section.stable_key == section_key
        assert imported_item.stable_key == item_key

    def test_duplicate_stable_key_in_file_rejected(self, app, db_session, admin_user):
        template = create_test_template(db_session, owner_id=admin_user.id)
        draft = create_test_draft_version(db_session, template)
        dup_key = '11111111-2222-3333-4444-555555555555'
        cols = _excel_columns(app)

        section_row = _blank_row(cols['sections'])
        section_row[cols['sections'].index('id')] = 1
        section_row[cols['sections'].index('order')] = 1.0
        section_row[cols['sections'].index('name')] = 'Section A'
        section_row[cols['sections'].index('section_type')] = 'standard'

        def _item_row(export_id, label):
            row = _blank_row(cols['items']) + [dup_key]
            row[cols['items'].index('id')] = export_id
            row[cols['items'].index('section_id')] = 1
            row[cols['items'].index('item_type')] = 'question'
            row[cols['items'].index('order')] = export_id
            row[cols['items'].index('label')] = label
            return row

        buf = io.BytesIO()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws_t = wb.create_sheet('Template')
        _append_template_sheet_row_layout(ws_t)
        for name, headers, rows in (
            ('Pages', cols['pages'], []),
            ('Sections', cols['sections'], [section_row]),
            ('Items', cols['items'] + ['stable_key'], [_item_row(1, 'Q1'), _item_row(2, 'Q2')]),
        ):
            ws = wb.create_sheet(name)
            ws.append(headers)
            for row in rows:
                ws.append(row)
        wb.save(buf)
        buf.seek(0)

        with app.test_request_context():
            login_user(admin_user)
            result = TemplateExcelService.import_template(template.id, buf, version_id=draft.id)

        assert result['success'] is False
        assert any('Duplicate stable_key' in err for err in result.get('errors', []))

    def test_published_import_blocked_when_submission_data_exists(self, app, db_session, admin_user):
        from app.models import FormData
        from tests.factories import create_test_assignment_entity_status

        template = create_test_template(db_session, owner_id=admin_user.id)
        published = template.published_version
        section = create_test_section(db_session, template, version=published)
        item = create_test_item(db_session, section, template, version=published, item_type='question')
        aes = create_test_assignment_entity_status(db_session, template=template)
        db_session.add(FormData(assignment_entity_status_id=aes.id, form_item_id=item.id, value='1'))
        db_session.commit()

        cols = _excel_columns(app)
        section_row = _blank_row(cols['sections'])
        section_row[cols['sections'].index('id')] = 1
        section_row[cols['sections'].index('order')] = 1.0
        section_row[cols['sections'].index('name')] = 'Section A'
        section_row[cols['sections'].index('section_type')] = 'standard'
        item_row = _blank_row(cols['items'])
        item_row[cols['items'].index('id')] = 1
        item_row[cols['items'].index('section_id')] = 1
        item_row[cols['items'].index('item_type')] = 'question'
        item_row[cols['items'].index('order')] = 1.0
        item_row[cols['items'].index('label')] = 'Q1'
        buf = _make_export_workbook_bytes(sections=[section_row], items=[item_row], app=app)

        with app.test_request_context():
            login_user(admin_user)
            result = TemplateExcelService.import_template(
                template.id,
                buf,
                version_id=published.id,
                import_version_mode='current_version',
            )

        assert result['success'] is False
        assert any('submission data' in err.lower() for err in result.get('errors', []))

    def test_new_item_auto_generates_stable_key(self, db_session, app):
        template = create_test_template(db_session)
        version = template.published_version
        section = create_test_section(db_session, template, version=version)
        item = create_test_item(db_session, section, template, version=version, item_type='question')
        assert item.stable_key is not None
        from app.utils.stable_key import is_valid_stable_key
        assert is_valid_stable_key(item.stable_key)
