#!/usr/bin/env python3
"""
Per-country UPR Country Reporting Excel template round-trip for Template 33 (Reporting – Country).

Uses the structured IFRC reporting workbook (named cells + Excel tables) for export/import
from a single country assignment (aes_id), distinct from the bulk UPR Master import.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import tempfile
from contextlib import contextmanager
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

logger = logging.getLogger(__name__)

script_dir = os.path.dirname(os.path.abspath(__file__))
backoffice_dir = os.path.dirname(os.path.dirname(script_dir))
if backoffice_dir not in sys.path:
    sys.path.insert(0, backoffice_dir)
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

if "FLASK_CONFIG" not in os.environ:
    os.environ["FLASK_CONFIG"] = "development"

from import_fdrs_form_data import (  # noqa: E402
    COL_ASSIGNMENT,
    COL_DATA_NA,
    COL_DISAGG,
    COL_IMPUTED,
    COL_ITEM,
    COL_NA,
    COL_PREFILLED,
    COL_PUBLIC,
    COL_SUBMITTED,
    COL_VALUE,
    upsert_form_data_rows,
)
from import_upr_excel_data import (  # noqa: E402
    ITEM_REPORTING_COUNTRY_EXPENDITURE,
    ITEM_REPORTING_COUNTRY_FUNDING,
    ITEM_REPORTING_COUNTRY_SP_BREAKDOWN,
    ITEM_REPORTING_COUNTRY_SUPPORT,
    REPORTING_COUNTRY_TEMPLATE_ID,
    REPORTING_FUNDING_ROW_IFRC,
    REPORTING_FUNDING_ROW_OTHER,
    REPORTING_FUNDING_ROW_PNS,
    REPORTING_FUNDING_MATRIX_COLUMN,
    REPORTING_SP_BREAKDOWN_AREA_TO_ROW,
    REPORTING_SP_BREAKDOWN_COLUMNS,
    UprImportContext,
    _data_na_row,
    _matrix_row,
    _resolve_item_by_bank_and_area,
    _resolve_ns_row_id,
    _scalar_row,
    build_import_context,
    reporting_funding_matrix_column,
    reporting_special_item,
    round_to_period,
    upsert_upr_discussion_comments,
    upsert_dynamic_indicator_entries,
    _matrix_column_name_from_item_id,
    _queue_other_dynamic_indicator,
)

# NS Data indicator bank IDs (same as UPR T33 / T24 NS Data)
NS_DATA_BANK_IDS: Dict[str, int] = {
    "volunteers": 724,
    "staff": 727,
    "local_units": 723,
    "branches": 1117,
}

NS_DATA_NAMED_CELLS: Dict[str, str] = {
    "volunteers": "Data_vol",
    "staff": "Data_staff",
    "local_units": "Data_units",
    "branches": "Data_branches",
}

FUNDING_SOURCE_ATTR_TO_ROW: Dict[str, str] = {
    "IFRC": REPORTING_FUNDING_ROW_IFRC,
    "PNS": REPORTING_FUNDING_ROW_PNS,
    "HNS": REPORTING_FUNDING_ROW_OTHER,
}

SP_BREAKDOWN_ROW_TO_AREA: Dict[str, str] = {v: k for k, v in REPORTING_SP_BREAKDOWN_AREA_TO_ROW.items()}
# Form matrix column labels differ from workbook table headers (Funding / Expenditure).
SP_BREAKDOWN_MATRIX_COL_TO_EXCEL: Dict[str, str] = {
    "Funding (CHF)": "Funding",
    "Expenditure (CHF)": "Expenditure",
}

SUPPORT_COLUMN_TO_AREA: Dict[str, str] = {
    "SP1_Supported": "SP1",
    "SP2_Supported": "SP2",
    "SP3_Supported": "SP3",
    "SP4_Supported": "SP4",
    "SP5_Supported": "SP5",
    "EFs_Supported": "EFs",
}

COMMENTS_NAMED_CELL = "Commetns_overall"
ASSIGNMENT_LABEL_NAMED_CELL = "Data_AssignmentLabel"
ASSIGNMENT_LABEL_SHEET = "Start"
ASSIGNMENT_LABEL_CELL = "K11"
START_HERE_CELL = "C2"
# Sheet cells that show "{assignment label}: {country}" (legacy templates used a fixed year + "Midyear Reporting").
REPORTING_SHEET_TITLE_CELLS: Tuple[Tuple[str, str], ...] = (
    ("Overall action Indicators", "C1"),
    ("Emergency Appeal 1", "C1"),
    ("Emergency Appeal 2", "C1"),
    ("Emergency Appeal 3", "C1"),
    ("Funding", "B1"),
    ("Bilateral Support", "B1"),
)
_LEGACY_REPORTING_TITLE_MARKERS = ("Midyear Reporting", "Mid-Year Reporting", "Mid Year Reporting")

INDICATOR_TABLES: Tuple[Tuple[str, str], ...] = (
    ("Overall action Indicators", "Data_core"),
    ("Overall action Indicators", "Data_other"),
)

BILATERAL_TABLES: Tuple[Tuple[str, str], ...] = (
    ("Bilateral Support", "Data_act"),
    ("Bilateral Support", "Data_act2"),
)
BILATERAL_PLANNED_TABLE = BILATERAL_TABLES[0]
BILATERAL_MANUAL_TABLE = BILATERAL_TABLES[1]
BILATERAL_NS_SOURCE_COL = 3  # Column C holds the partner NS name (export writes here; NS column mirrors it).
AREA_TO_SUPPORT_COLUMN: Dict[str, str] = {area: col for col, area in SUPPORT_COLUMN_TO_AREA.items()}
_SUPPORT_CELL_KEY_RE = re.compile(r"^(\d+)_(SP\d|EFs) Supported$")

EMERGENCY_SLOTS: Tuple[Tuple[str, str, int, str, str], ...] = (
    ("Emergency Appeal 1", "Data_emergency1", 1, "Data_MDR1", "Data_EO1"),
    ("Emergency Appeal 2", "Data_emergency2", 2, "Data_MDR2", "Data_EO2"),
    ("Emergency Appeal 3", "Data_emergency3", 3, "Data_MDR3", "Data_EO3"),
)

INDICATOR_TOTAL_HEADER = "Total\nDirect + Indirect"
INDICATOR_DNA_HEADER = "Applicable/\nData not available"
INDICATOR_APPLICABLE_VALUE = "Applicable"
INDICATOR_DNA_VALUE = "Data not available"
INDICATOR_ID_HEADER = "ID"
SP_EF_HEADER = "Strategic Priority / Enabling Function"
INDICATOR_HEADER = "Indicator"
DATA_OTHER_SHEET = "Overall action Indicators"
DATA_OTHER_TABLE = "Data_other"

# Workbooks produced by the generic assignment Excel export (not UPR Country Reporting).
GENERIC_FORM_EXPORT_SHEETS: Tuple[str, ...] = ("Template", "Pages", "Sections", "Items")

# Minimum structure required by the T33 UPR Country Reporting import script.
UPR_COUNTRY_REPORTING_REQUIRED_NAMED_RANGES: Tuple[str, ...] = (
    "Version",
    "Data_Country",
)
UPR_COUNTRY_REPORTING_REQUIRED_TABLES: Tuple[Tuple[str, str, Tuple[str, ...]], ...] = (
    ("Indicators list", "Final", ("KPI ID", "SP/EF", "Indicator Name")),
    (
        "Overall action Indicators",
        "Data_core",
        (INDICATOR_ID_HEADER, SP_EF_HEADER, INDICATOR_HEADER, INDICATOR_DNA_HEADER),
    ),
    ("Funding", "Data_funding1", ()),
)
UPR_COUNTRY_REPORTING_COMPATIBLE_ROUND_PREFIXES: Tuple[str, ...] = ("MYR", "AR")

# Fuzzy label matching thresholds (Excel truncates long indicator names).
INDICATOR_MATCH_THRESHOLD = 0.60
INDICATOR_CROSS_SECTION_THRESHOLD = 0.95

# Normalized workbook indicator-table column header -> internal disagg key.
WORKBOOK_NORM_HEADER_TO_KEY: Dict[str, str] = {
    "total direct": "direct",
    "total male": "male",
    "total female": "female",
    "other/ unknown": "unknown",
    "male <5": "male_5",
    "male 5-17": "male_5_17",
    "male 18-49": "male_18_49",
    "male 50+": "male_50_",
    "female <5": "female_5",
    "female 5-17": "female_5_17",
    "female 18-49": "female_18_49",
    "female 50+": "female_50_",
    "indirectly reached": "indirect",
    "total direct + indirect": "combined",
}

# Excel SP/EF labels that differ from published T33 section names.
SECTION_ALIASES: Dict[str, str] = {
    "cross-cutting": "cross cutting",
    "cross cutting": "cross cutting",
}


@dataclass
class UprCountryReportingWorkbookContext:
    """Parsed UPR Country Reporting workbook metadata for validation."""

    round_code: str = ""
    period_name: str = ""
    country_name: str = ""
    warnings: List[str] = field(default_factory=list)


def _normalize_workbook_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ").strip()).lower()


def _coerce_disagg_number(raw: Any) -> Optional[float]:
    val = _parse_value_num(raw)
    if val is None:
        return None
    if val == int(val):
        return float(int(val))
    return val


def _sum_disagg_breakdown(breakdown: Dict[str, Any]) -> Optional[float]:
    """Sum numeric breakdown cells into Total Direct (excludes indirect/total meta keys)."""
    if not isinstance(breakdown, dict):
        return None
    total = 0.0
    found = False
    for key, val in breakdown.items():
        if str(key).lower() in ("indirect", "total", "combined", "direct"):
            continue
        num = _coerce_disagg_number(val)
        if num is not None:
            total += num
            found = True
    return total if found else None


def _disagg_breakdown_dict(mode: str, values: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if mode == "sex_age":
        direct = values.get("direct")
        return direct if isinstance(direct, dict) else None
    if mode in ("sex", "age"):
        direct = values.get("direct", values)
        return direct if isinstance(direct, dict) else None
    return None


def _is_non_binary_disagg_key(key: str) -> bool:
    norm = str(key or "").strip().lower().replace("-", "_")
    return norm == "non_binary" or norm.startswith("non_binary_")


def _merge_non_binary_into_unknown_breakdown(breakdown: Dict[str, Any]) -> Dict[str, float]:
    """Map form non_binary counts into workbook Other/Unknown (merged with unknown)."""
    merged: Dict[str, float] = {}
    other_unknown = 0.0
    for key, val in breakdown.items():
        num = _coerce_disagg_number(val)
        if num is None:
            continue
        if _is_non_binary_disagg_key(str(key)):
            other_unknown += num
            continue
        norm = str(key).strip().lower().replace("-", "_")
        if norm == "unknown":
            other_unknown += num
            continue
        merged[str(key)] = merged.get(str(key), 0.0) + num
    if other_unknown:
        merged["unknown"] = merged.get("unknown", 0.0) + other_unknown
    return merged


def _parse_row_bank_id(row: Dict[str, Any]) -> Optional[int]:
    raw = row.get(INDICATOR_ID_HEADER)
    if raw is None or raw == "":
        return None
    try:
        return int(float(str(raw).strip()))
    except (ValueError, TypeError):
        return None


def _build_upr_country_reporting_disagg_header_maps(wb) -> Tuple[Dict[str, str], Dict[str, str]]:
    """Return (key -> exact header, normalized header -> exact header) for indicator tables."""
    key_to_header: Dict[str, str] = {}
    norm_to_header: Dict[str, str] = {}
    sample_tables = (("Overall action Indicators", "Data_core"),)
    for sheet_name, table_name in sample_tables:
        if sheet_name not in wb.sheetnames or table_name not in wb[sheet_name].tables:
            continue
        headers, _ = read_named_table(wb, sheet_name, table_name)
        for header in headers:
            if not header:
                continue
            norm = _normalize_workbook_header(header)
            norm_to_header[norm] = header
            logical_key = WORKBOOK_NORM_HEADER_TO_KEY.get(norm)
            if logical_key:
                key_to_header[logical_key] = header
    return key_to_header, norm_to_header


def _parse_workbook_row_disagg(row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Build entry-form disagg_data JSON from a workbook indicator table row."""
    sex_age: Dict[str, float] = {}
    sex: Dict[str, float] = {}
    direct_total: Optional[float] = None
    indirect: Optional[float] = None

    for header, raw in row.items():
        if not header or str(header).startswith("_"):
            continue
        norm = _normalize_workbook_header(header)
        if norm == "id":
            continue
        logical_key = WORKBOOK_NORM_HEADER_TO_KEY.get(norm)
        if not logical_key or logical_key == "combined":
            continue
        val = _coerce_disagg_number(raw)
        if val is None:
            continue
        if logical_key in ("male_5", "male_5_17", "male_18_49", "male_50_", "female_5", "female_5_17", "female_18_49", "female_50_"):
            sex_age[logical_key] = val
        elif logical_key in ("male", "female", "unknown"):
            sex[logical_key] = val
        elif logical_key == "direct":
            direct_total = val
        elif logical_key == "indirect":
            indirect = val

    if sex_age:
        values: Dict[str, Any] = {"direct": sex_age}
        if indirect is not None:
            values["indirect"] = indirect
        return {"mode": "sex_age", "values": values}
    if sex:
        if indirect is not None:
            return {"mode": "sex", "values": {"direct": sex, "indirect": indirect}}
        return {"mode": "sex", "values": sex}
    if direct_total is not None or indirect is not None:
        values: Dict[str, Any] = {}
        if direct_total is not None:
            # Entry-form total mode without indirect reach stores a scalar under ``total``,
            # not ``direct`` (see FormItemProcessor._process_numeric_indicator).
            if indirect is not None:
                values["direct"] = direct_total
            else:
                values["total"] = direct_total
        if indirect is not None:
            values["indirect"] = indirect
        return {"mode": "total", "values": values}
    return None


def _entry_disagg_payload(entry) -> Optional[Dict[str, Any]]:
    if not entry:
        return None
    raw = entry.get_display_disagg_data() if hasattr(entry, "get_display_disagg_data") else None
    if raw is None:
        raw = getattr(entry, "disagg_data", None)
    if isinstance(raw, str) and raw.strip():
        try:
            raw = json.loads(raw)
        except json.JSONDecodeError:
            return None
    if isinstance(raw, dict) and raw.get("mode"):
        return raw
    return None


def _disagg_payload_to_workbook_cells(
    disagg: Dict[str, Any],
    key_to_header: Dict[str, str],
) -> Dict[str, Any]:
    """Map disagg_data JSON to exact workbook Excel column headers."""
    mode = disagg.get("mode") or "total"
    values = disagg.get("values") or {}
    out: Dict[str, Any] = {}

    def _set(logical_key: str, val: Any) -> None:
        header = key_to_header.get(logical_key)
        if header and val is not None:
            out[header] = val

    if mode == "sex_age":
        direct = values.get("direct")
        if isinstance(direct, dict):
            for key, val in _merge_non_binary_into_unknown_breakdown(direct).items():
                _set(str(key), val)
    elif mode == "sex":
        direct = values.get("direct", values)
        if isinstance(direct, dict):
            for key, val in _merge_non_binary_into_unknown_breakdown(direct).items():
                _set(str(key), val)
    elif mode == "age":
        direct = values.get("direct", values)
        if isinstance(direct, dict):
            for key, val in _merge_non_binary_into_unknown_breakdown(direct).items():
                _set(str(key), val)
    elif mode == "total":
        if "direct" in values:
            _set("direct", values.get("direct"))
        elif "total" in values:
            _set("direct", values.get("total"))
        if values.get("indirect") is not None:
            _set("indirect", values.get("indirect"))

    indirect_val = values.get("indirect")
    if indirect_val is not None and mode != "total":
        _set("indirect", indirect_val)

    if mode in ("sex", "age", "sex_age"):
        breakdown = _disagg_breakdown_dict(mode, values)
        direct_sum = _sum_disagg_breakdown(breakdown or {})
        if direct_sum is not None:
            _set("direct", direct_sum)
    return out


def _indicator_disagg_row(
    *,
    aes_id: int,
    item_id: int,
    disagg_payload: Dict[str, Any],
    iso3: str,
    period: str,
    debug_kpi: str,
) -> Dict[str, str]:
    return {
        "_debug_iso3": iso3,
        "_debug_year": period,
        "_debug_kpi_code": debug_kpi,
        COL_ASSIGNMENT: str(aes_id),
        COL_PUBLIC: "",
        COL_ITEM: str(item_id),
        COL_VALUE: "",
        COL_DISAGG: json.dumps(disagg_payload),
        COL_DATA_NA: "",
        COL_NA: "",
        COL_PREFILLED: "",
        COL_IMPUTED: "",
        COL_SUBMITTED: "",
    }


def _is_yes_no_indicator_type(type_value: Any) -> bool:
    normalized = str(type_value or "").strip().lower().replace("/", "").replace("-", "")
    return normalized in ("yesno", "boolean", "bool")


def _entry_is_yes_no(entry, *, form_item: Any = None, indicator_bank: Any = None) -> bool:
    """True when the submitted row is a Yes/No indicator (Applicable only in the workbook)."""
    if form_item is not None and _is_yes_no_indicator_type(getattr(form_item, "type", None)):
        return True
    if indicator_bank is not None and _is_yes_no_indicator_type(getattr(indicator_bank, "type", None)):
        return True
    if entry is None:
        return False
    linked_item = getattr(entry, "form_item", None)
    if linked_item is not None and _is_yes_no_indicator_type(getattr(linked_item, "type", None)):
        return True
    linked_bank = getattr(entry, "indicator_bank", None)
    if linked_bank is not None and _is_yes_no_indicator_type(getattr(linked_bank, "type", None)):
        return True
    return False


def _load_yes_no_bank_ids(bank_ids: Iterable[int]) -> Set[int]:
    """Return indicator bank ids whose type is Yes/No."""
    ids = {int(bid) for bid in bank_ids if bid}
    if not ids:
        return set()
    from app.models import IndicatorBank

    rows = IndicatorBank.query.filter(IndicatorBank.id.in_(ids)).all()
    return {int(row.id) for row in rows if _is_yes_no_indicator_type(row.type)}


def _workbook_indicator_table_names(wb) -> Tuple[Tuple[str, str], ...]:
    """All indicator tables present in a workbook (core, overflow, emergency slots)."""
    tables: List[Tuple[str, str]] = []
    for sheet_name, table_name in INDICATOR_TABLES:
        if sheet_name in wb.sheetnames and table_name in wb[sheet_name].tables:
            tables.append((sheet_name, table_name))
    for sheet_name, table_name, *_rest in EMERGENCY_SLOTS:
        if sheet_name in wb.sheetnames and table_name in wb[sheet_name].tables:
            tables.append((sheet_name, table_name))
    return tuple(tables)


def _collect_workbook_indicator_bank_ids(
    wb,
    kpi_lookup: Optional[Dict[Tuple[str, str], int]] = None,
) -> Set[int]:
    """Union of Final KPI ids and every ID column value from indicator tables."""
    bank_ids: Set[int] = set()
    if kpi_lookup:
        bank_ids.update(int(bid) for bid in kpi_lookup.values() if bid)
    for sheet_name, table_name in _workbook_indicator_table_names(wb):
        _, rows = read_named_table(wb, sheet_name, table_name)
        for row in rows:
            bank_id = _parse_row_bank_id(row)
            if bank_id:
                bank_ids.add(int(bank_id))
    return bank_ids


def _load_workbook_yes_no_bank_ids(
    wb,
    kpi_lookup: Optional[Dict[Tuple[str, str], int]] = None,
) -> Set[int]:
    """Yes/No bank ids referenced by this workbook (ID column + Final KPI list)."""
    return _load_yes_no_bank_ids(_collect_workbook_indicator_bank_ids(wb, kpi_lookup))


def _resolve_workbook_indicator_bank_id(
    row: Dict[str, Any],
    kpi_lookup: Optional[Dict[Tuple[str, str], int]] = None,
) -> Optional[int]:
    bank_id = row.get("bank_id")
    if bank_id:
        return int(bank_id)
    if not kpi_lookup:
        return None
    for (sp, ind), bid in kpi_lookup.items():
        if not _section_names_match(row["sp_ef"], sp):
            continue
        if _indicator_similarity(row["indicator"], ind) >= INDICATOR_MATCH_THRESHOLD:
            return int(bid)
    return None


def _workbook_yes_no_value(applicable_text: str) -> str:
    """Map workbook Applicable column to entry-form yes/no storage."""
    text = str(applicable_text or "").strip().lower()
    if "data not available" in text:
        return "no"
    if "applicable" in text:
        return "yes"
    return "no"


def _yes_no_value_is_applicable(value: Any) -> bool:
    """True when a Yes/No form answer should export as Applicable in the workbook."""
    return str(value or "").strip().lower() in ("yes", "y", "1", "true")


def _resolve_indicator_import_value(
    row: Dict[str, Any],
    bank_id: Optional[int],
    yes_no_bank_ids: Set[int],
) -> Tuple[Optional[Any], bool, Optional[Dict[str, Any]], bool]:
    """Return (value, data_not_available, disagg, should_import)."""
    if bank_id and bank_id in yes_no_bank_ids:
        return _workbook_yes_no_value(row.get("applicable_text", "")), False, None, True
    if row.get("data_not_available"):
        return None, True, None, True
    disagg = row.get("disagg")
    value = row.get("value")
    return value, False, disagg, bool(disagg or value is not None)


def _write_indicator_applicable_cell(
    wb,
    sheet_name: str,
    table_name: str,
    row_offset: int,
    *,
    data_not_available: bool,
    applicable: bool,
) -> None:
    if data_not_available:
        write_table_cell(
            wb,
            sheet_name,
            table_name,
            row_offset,
            INDICATOR_DNA_HEADER,
            INDICATOR_DNA_VALUE,
        )
    elif applicable:
        write_table_cell(
            wb,
            sheet_name,
            table_name,
            row_offset,
            INDICATOR_DNA_HEADER,
            INDICATOR_APPLICABLE_VALUE,
        )


def _write_indicator_entry(
    wb,
    sheet_name: str,
    table_name: str,
    row_offset: int,
    entry,
    key_to_header: Dict[str, str],
    *,
    fallback_value: Any = None,
    data_not_available: bool = False,
    yes_no: Optional[bool] = None,
) -> None:
    if data_not_available:
        _write_indicator_applicable_cell(
            wb,
            sheet_name,
            table_name,
            row_offset,
            data_not_available=True,
            applicable=False,
        )
        return

    is_yes_no = yes_no if yes_no is not None else _entry_is_yes_no(entry)
    val = fallback_value if fallback_value is not None else _scalar_value(entry)

    if is_yes_no:
        if _yes_no_value_is_applicable(val):
            _write_indicator_applicable_cell(
                wb,
                sheet_name,
                table_name,
                row_offset,
                data_not_available=False,
                applicable=True,
            )
        return

    disagg = _entry_disagg_payload(entry)
    disagg_cells: Dict[str, Any] = {}
    if disagg:
        disagg_cells = _disagg_payload_to_workbook_cells(disagg, key_to_header)

    if not disagg_cells and val is None:
        return

    _write_indicator_applicable_cell(
        wb,
        sheet_name,
        table_name,
        row_offset,
        data_not_available=False,
        applicable=True,
    )
    for header, cell_val in disagg_cells.items():
        write_table_cell(wb, sheet_name, table_name, row_offset, header, cell_val)
    if not disagg_cells and val is not None:
        direct_header = key_to_header.get("direct")
        if direct_header:
            write_table_cell(wb, sheet_name, table_name, row_offset, direct_header, val)


def _normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _parse_value_num(raw: Any) -> Optional[float]:
    if raw is None or raw == "":
        return None
    try:
        return float(raw)
    except (ValueError, TypeError):
        return None


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise RuntimeError("Excel support requires openpyxl: pip install openpyxl") from exc


@contextmanager
def _quiet_openpyxl_io():
    """Silence Pillow PNG chunk DEBUG lines while openpyxl reads embedded template images."""
    names = ("PIL", "PIL.PngImagePlugin", "PIL.Image")
    previous_levels: Dict[str, int] = {}
    for name in names:
        pil_logger = logging.getLogger(name)
        previous_levels[name] = pil_logger.level
        pil_logger.setLevel(logging.WARNING)
    try:
        yield
    finally:
        for name, level in previous_levels.items():
            logging.getLogger(name).setLevel(level)


def read_named_cell(wb, name: str) -> Any:
    """Read a scalar value from a workbook named range."""
    if name not in wb.defined_names:
        return None
    defn = wb.defined_names[name]
    for title, coord in defn.destinations:
        ws = wb[title]
        if ":" in coord:
            coord = coord.split(":")[0]
        return ws[coord].value
    return None


def write_named_cell(wb, name: str, value: Any) -> None:
    """Write a scalar value to a workbook named range."""
    if name not in wb.defined_names:
        return
    defn = wb.defined_names[name]
    for title, coord in defn.destinations:
        ws = wb[title]
        if ":" in coord:
            coord = coord.split(":")[0]
        ws[coord].value = value


def _ensure_assignment_label_named_range(wb) -> None:
    """Ensure Data_AssignmentLabel points at Start!K11 (hidden metadata cell for title formulas)."""
    if ASSIGNMENT_LABEL_NAMED_CELL not in wb.defined_names:
        from openpyxl.workbook.defined_name import DefinedName

        wb.defined_names.add(
            DefinedName(
                ASSIGNMENT_LABEL_NAMED_CELL,
                attr_text=f"{ASSIGNMENT_LABEL_SHEET}!${ASSIGNMENT_LABEL_CELL}",
            )
        )


def _assignment_display_label(aes) -> str:
    """Custom assignment name, or '<template> – <period>' (matches AssignedForm.display_name)."""
    assigned_form = getattr(aes, "assigned_form", None)
    if not assigned_form:
        return "UPR Country Reporting"
    display_name = getattr(assigned_form, "display_name", None)
    if callable(display_name):
        label = display_name()
    else:
        label = display_name
    text = str(label or "").strip()
    if text:
        return text
    template_name = assigned_form.template.name if getattr(assigned_form, "template", None) else "UPR Country Reporting"
    period = (assigned_form.period_name or "").strip()
    return f"{template_name} – {period}" if period else template_name


def _reporting_sheet_title_formula() -> str:
    return (
        '=IF(Data_Country="","Please select your country at the Start page",'
        f"{ASSIGNMENT_LABEL_NAMED_CELL}&\": \"&Data_Country)"
    )


def _start_here_title_text(assignment_label: str) -> str:
    label = str(assignment_label or "").strip()
    if not label:
        return "Reporting - Start here"
    return f"{label} - Start here"


def _write_cell_value_preserving_style(ws, cell_ref: str, value: Any) -> None:
    """Replace cell contents without clearing existing font/fill/alignment."""
    ws[cell_ref].value = value


def _cell_uses_legacy_reporting_title(value: Any) -> bool:
    if value is None:
        return False
    text = str(value)
    return any(marker in text for marker in _LEGACY_REPORTING_TITLE_MARKERS)


def _apply_reporting_assignment_label(wb, assignment_label: str) -> None:
    """Write assignment label and retarget sheet titles away from fixed-year template text."""
    _ensure_assignment_label_named_range(wb)
    write_named_cell(wb, ASSIGNMENT_LABEL_NAMED_CELL, assignment_label)

    title_formula = _reporting_sheet_title_formula()
    for sheet_name, cell_ref in REPORTING_SHEET_TITLE_CELLS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if _cell_uses_legacy_reporting_title(ws[cell_ref].value) or (
            isinstance(ws[cell_ref].value, str) and ws[cell_ref].value.startswith("=IF")
        ):
            ws[cell_ref].value = title_formula

    start_cell = wb[ASSIGNMENT_LABEL_SHEET][START_HERE_CELL]
    if _cell_uses_legacy_reporting_title(start_cell.value) or (
        isinstance(start_cell.value, str)
        and (
            start_cell.value.startswith("=")
            or "Start here" in start_cell.value
        )
    ):
        _write_cell_value_preserving_style(
            wb[ASSIGNMENT_LABEL_SHEET],
            START_HERE_CELL,
            _start_here_title_text(assignment_label),
        )


def read_named_table(wb, sheet_name: str, table_name: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Read an Excel table's header row and data rows as dicts."""
    from openpyxl.utils import range_boundaries

    ws = wb[sheet_name]
    if table_name not in ws.tables:
        raise ValueError(f"Table {table_name!r} not found on sheet {sheet_name!r}")
    tbl = ws.tables[table_name]
    ref = tbl.ref if hasattr(tbl, "ref") else tbl
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    headers: List[str] = []
    for col in range(min_col, max_col + 1):
        raw = ws.cell(min_row, col).value
        headers.append(str(raw).strip() if raw is not None else "")
    rows: List[Dict[str, Any]] = []
    for row_idx in range(min_row + 1, max_row + 1):
        record: Dict[str, Any] = {}
        has_value = False
        for offset, header in enumerate(headers):
            if not header:
                continue
            val = ws.cell(row_idx, min_col + offset).value
            if val is not None and str(val).strip() != "":
                has_value = True
            record[header] = val
        if has_value:
            record["_row"] = row_idx
            record["_sheet"] = sheet_name
            record["_table"] = table_name
            rows.append(record)
    return headers, rows


def write_table_cell(wb, sheet_name: str, table_name: str, row_offset: int, header: str, value: Any) -> None:
    """Write a value into a table cell identified by header name and 0-based data row offset."""
    from openpyxl.utils import range_boundaries

    ws = wb[sheet_name]
    tbl = ws.tables[table_name]
    ref = tbl.ref if hasattr(tbl, "ref") else tbl
    min_col, min_row, max_col, _max_row = range_boundaries(ref)
    col_index: Optional[int] = None
    header_norm = _normalize_workbook_header(header)
    for col in range(min_col, max_col + 1):
        raw = ws.cell(min_row, col).value
        h = str(raw).strip() if raw is not None else ""
        if h == header or _normalize_workbook_header(h) == header_norm:
            col_index = col
            break
    if col_index is None:
        return
    ws.cell(min_row + 1 + row_offset, col_index).value = value


def read_table_cell(wb, sheet_name: str, table_name: str, row_offset: int, header: str) -> Any:
    """Read a table cell by header name and 0-based data row offset."""
    from openpyxl.utils import range_boundaries

    ws = wb[sheet_name]
    tbl = ws.tables[table_name]
    ref = tbl.ref if hasattr(tbl, "ref") else tbl
    min_col, min_row, max_col, _max_row = range_boundaries(ref)
    header_norm = _normalize_workbook_header(header)
    for col in range(min_col, max_col + 1):
        raw = ws.cell(min_row, col).value
        h = str(raw).strip() if raw is not None else ""
        if h == header or _normalize_workbook_header(h) == header_norm:
            return ws.cell(min_row + 1 + row_offset, col).value
    return None


def parse_version(wb) -> Tuple[str, str]:
    """Extract round code and period name from the Version named cell."""
    raw = read_named_cell(wb, "Version")
    text = str(raw or "").strip()
    if not text:
        return "", ""
    round_code = text.split(".")[0].upper()
    period = round_to_period(round_code) or ""
    return round_code, period


def period_to_workbook_version(period_name: str) -> str:
    """Map assignment period name to the workbook Version cell value (IFRC round codes)."""
    period = (period_name or "").strip()
    match = re.match(r"Jan-Jun\s+(\d{4})", period, re.IGNORECASE)
    if match:
        year = int(match.group(1))
        return f"MYR{year - 2000}.V1.0"
    match = re.match(r"^(\d{4})$", period)
    if match:
        year = int(match.group(1))
        return f"AR{year - 2000}.V1.0"
    return "MYR26.V1.0"


def build_kpi_lookup(wb) -> Dict[Tuple[str, str], int]:
    """Build (normalized_sp_ef, normalized_indicator_name) -> bank_id from Final table."""
    _, rows = read_named_table(wb, "Indicators list", "Final")
    lookup: Dict[Tuple[str, str], int] = {}
    for row in rows:
        bank_raw = row.get("KPI ID")
        if bank_raw is None:
            continue
        try:
            bank_id = int(bank_raw)
        except (ValueError, TypeError):
            continue
        sp_ef = _normalize_text(row.get("SP/EF"))
        indicator = _normalize_text(row.get("Indicator Name"))
        if sp_ef and indicator:
            lookup[(sp_ef, indicator)] = bank_id
    return lookup


def _build_kpi_display_map(wb) -> Dict[int, Tuple[str, str]]:
    """Map indicator bank id -> (SP/EF display label, indicator display label)."""
    _, rows = read_named_table(wb, "Indicators list", "Final")
    out: Dict[int, Tuple[str, str]] = {}
    for row in rows:
        bank_raw = row.get("KPI ID")
        if bank_raw is None:
            continue
        try:
            bank_id = int(bank_raw)
        except (ValueError, TypeError):
            continue
        sp_ef = str(row.get("SP/EF") or "").strip()
        indicator = str(row.get("Indicator Name") or "").strip()
        if sp_ef and indicator:
            out[bank_id] = (sp_ef, indicator)
    return out


def _load_items_by_section_label(template_id: int, version_id: Optional[int] = None) -> Dict[Tuple[str, str], int]:
    """Map (normalized section name, normalized item label) -> form_item_id."""
    from app.models.form_items import FormItem
    from app.models.forms import FormSection, FormTemplateVersion

    out: Dict[Tuple[str, str], int] = {}
    query = (
        FormItem.query.join(FormTemplateVersion, FormItem.version_id == FormTemplateVersion.id)
        .outerjoin(FormSection, FormItem.section_id == FormSection.id)
        .filter(
            FormItem.template_id == template_id,
            FormItem.archived == False,
        )
    )
    if version_id:
        query = query.filter(FormItem.version_id == int(version_id))
    else:
        query = query.filter(FormTemplateVersion.status == "published")
    items = query.all()
    for item in items:
        section_name = (item.form_section.name if item.form_section else "") or ""
        label = (item.label or "").strip()
        if not label:
            continue
        out[(_normalize_text(section_name), _normalize_text(label))] = int(item.id)
    return out


def _normalize_section_name(value: Any) -> str:
    text = _normalize_text(value).replace("-", " ")
    return SECTION_ALIASES.get(text, text)


def _normalize_indicator_label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower()).rstrip(".")


def _section_names_match(left: str, right: str) -> bool:
    a = _normalize_section_name(left)
    b = _normalize_section_name(right)
    return a == b or a.startswith(b) or b.startswith(a)


def _indicator_similarity(excel_indicator: str, form_label: str) -> float:
    a = _normalize_indicator_label(excel_indicator)
    b = _normalize_indicator_label(form_label)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    short, long = (a, b) if len(a) <= len(b) else (b, a)
    if len(short) >= 25 and long.startswith(short[:25]):
        return 0.9
    if short in long or long in short:
        return 0.8
    return SequenceMatcher(None, a, b).ratio()


def _indicator_names_match(excel_indicator: str, form_label: str) -> bool:
    return _indicator_similarity(excel_indicator, form_label) >= INDICATOR_MATCH_THRESHOLD


def _resolve_comments_item_id(ctx: UprImportContext) -> Optional[int]:
    labels = ctx.item_ids_by_label.get(REPORTING_COUNTRY_TEMPLATE_ID, {})
    for key, item_id in labels.items():
        if "comment" in key:
            return item_id
    return None


def _load_upr_country_reporting_section_ids() -> Dict[str, int]:
    from app.models.forms import FormSection, FormTemplateVersion

    out: Dict[str, int] = {}
    sections = (
        FormSection.query.join(FormTemplateVersion, FormSection.version_id == FormTemplateVersion.id)
        .filter(
            FormSection.template_id == REPORTING_COUNTRY_TEMPLATE_ID,
            FormTemplateVersion.status == "published",
        )
        .all()
    )
    ea_repeat_id: Optional[int] = None
    for section in sections:
        name = (section.name or "").strip().lower()
        if name == "emergency appeals indicators" and section.section_type == "repeat":
            ea_repeat_id = int(section.id)
            out["ea_repeat"] = ea_repeat_id
        elif name == "other indicators" and section.section_type == "dynamic_indicators":
            out["other_dynamic"] = int(section.id)
    if ea_repeat_id:
        for section in sections:
            if (
                section.parent_section_id == ea_repeat_id
                and (section.section_type or "").lower() == "dynamic_indicators"
            ):
                out["ea_dynamic"] = int(section.id)
                break
    if "ea_dynamic" not in out:
        for section in sections:
            if (section.name or "").strip().lower() == "emergency appeal indicators":
                out["ea_dynamic"] = int(section.id)
                break
    return out


def _load_upr_country_reporting_emergency_choice_item_id(ea_repeat_section_id: Optional[int]) -> Optional[int]:
    """Form item id for the per-repeat Emergency Operations single-choice field."""
    if not ea_repeat_section_id:
        return None
    from app.models.form_items import FormItem
    from app.models.forms import FormTemplateVersion

    item = (
        FormItem.query.join(FormTemplateVersion, FormItem.version_id == FormTemplateVersion.id)
        .filter(
            FormItem.template_id == REPORTING_COUNTRY_TEMPLATE_ID,
            FormItem.section_id == ea_repeat_section_id,
            FormItem.archived == False,
            FormItem.lookup_list_id == "emergency_operations",
            FormTemplateVersion.status == "published",
        )
        .first()
    )
    return int(item.id) if item else None


def _parse_emergency_selection_from_entry(entry) -> Optional[Dict[str, str]]:
    """Extract appeal code/name from a repeat-group emergency_operations choice entry."""
    if not entry:
        return None
    disagg = getattr(entry, "disagg_data", None)
    if isinstance(disagg, str) and disagg.strip():
        try:
            disagg = json.loads(disagg)
        except json.JSONDecodeError:
            disagg = None
    if isinstance(disagg, dict):
        code = str(disagg.get("code") or "").strip()
        name = str(disagg.get("name") or "").strip()
        if code or name:
            label = f"{name} ({code})" if code else name
            return {"code": code, "name": name, "label": label}
    text = str(getattr(entry, "value", None) or "").strip()
    if not text:
        return None
    match = re.match(r"^(.*)\s+\(([^)]+)\)\s*$", text)
    if match:
        name = match.group(1).strip()
        code = match.group(2).strip()
        return {"code": code, "name": name, "label": text}
    return {"code": "", "name": text, "label": text}


def _resolve_emergency_slots_for_export(aes) -> List[Optional[Dict[str, str]]]:
    """Resolve EO slots from repeat-instance emergency appeal selections (item 1374 pattern).

    Each repeat instance in "Emergency Appeals Indicators" carries one emergency_operations
    choice. Excel slot N maps to repeat instance_number N. Unused slots stay empty instead of
    auto-filling from the GO API (which resolve_slot_map would do).
    """
    slots: List[Optional[Dict[str, str]]] = [None, None, None]
    section_ids = _load_upr_country_reporting_section_ids()
    ea_repeat = section_ids.get("ea_repeat")
    choice_item_id = _load_upr_country_reporting_emergency_choice_item_id(ea_repeat)
    if not ea_repeat or not choice_item_id:
        return slots

    from app.models.forms import RepeatGroupData, RepeatGroupInstance

    instances = (
        RepeatGroupInstance.query.filter_by(
            assignment_entity_status_id=aes.id,
            section_id=ea_repeat,
        )
        .order_by(RepeatGroupInstance.instance_number.asc())
        .all()
    )
    for inst in instances:
        slot_num = int(inst.instance_number or 0)
        if not (1 <= slot_num <= len(slots)):
            continue
        entry = RepeatGroupData.query.filter_by(
            repeat_instance_id=inst.id,
            form_item_id=choice_item_id,
        ).first()
        meta = _parse_emergency_selection_from_entry(entry)
        if meta and (meta.get("code") or meta.get("name")):
            slots[slot_num - 1] = meta
    return slots


def _default_user_id_for_import() -> int:
    try:
        from flask_login import current_user

        if getattr(current_user, "is_authenticated", False) and getattr(current_user, "id", None):
            return int(current_user.id)
    except Exception:
        pass
    from app.models.core import User

    user = User.query.order_by(User.id.asc()).first()
    return int(user.id) if user else 1


def _build_indicator_row_index(
    wb,
    *,
    tables: Optional[Tuple[Tuple[str, str], ...]] = None,
) -> Dict[Tuple[str, str], Tuple[str, str, int]]:
    """Map (section, indicator label) -> (sheet, table, 0-based row offset in table).

    When the same indicator appears in multiple tables, the first table wins (Data_core
    before emergency slots) so core export targets the Overall Action table.
    """
    index: Dict[Tuple[str, str], Tuple[str, str, int]] = {}
    table_list = tables or (
        INDICATOR_TABLES + tuple((sheet, table) for sheet, table, *_rest in EMERGENCY_SLOTS)
    )
    for sheet_name, table_name in table_list:
        if sheet_name not in wb.sheetnames:
            continue
        if table_name not in wb[sheet_name].tables:
            continue
        _, rows = read_named_table(wb, sheet_name, table_name)
        for idx, row in enumerate(rows):
            sp_ef = str(row.get("Strategic Priority / Enabling Function") or "").strip()
            indicator = str(row.get("Indicator") or "").strip()
            if not sp_ef or not indicator:
                continue
            key = (_normalize_text(sp_ef), _normalize_indicator_label(indicator))
            if key in index:
                continue
            index[key] = (sheet_name, table_name, idx)
    return index


def _build_core_indicator_row_index(wb) -> Dict[Tuple[str, str], Tuple[str, str, int]]:
    """Row index for Overall Action tables only (Data_core + Data_other)."""
    return _build_indicator_row_index(wb, tables=INDICATOR_TABLES)


def _build_bank_id_row_locations(
    wb,
    *,
    tables: Optional[Tuple[Tuple[str, str], ...]] = None,
) -> Dict[Tuple[str, str, int], Tuple[str, str, int]]:
    """Map (sheet, table, indicator_bank_id) -> (sheet, table, 0-based row offset)."""
    locations: Dict[Tuple[str, str, int], Tuple[str, str, int]] = {}
    table_list = tables or (
        INDICATOR_TABLES + tuple((sheet, table) for sheet, table, *_rest in EMERGENCY_SLOTS)
    )
    for sheet_name, table_name in table_list:
        if sheet_name not in wb.sheetnames:
            continue
        if table_name not in wb[sheet_name].tables:
            continue
        _, rows = read_named_table(wb, sheet_name, table_name)
        for idx, row in enumerate(rows):
            bank_id = _parse_row_bank_id(row)
            if bank_id is None:
                continue
            key = (sheet_name, table_name, bank_id)
            if key not in locations:
                locations[key] = (sheet_name, table_name, idx)
    return locations


def _find_row_by_bank_id(
    bank_id: Optional[int],
    bank_id_locations: Dict[Tuple[str, str, int], Tuple[str, str, int]],
    *,
    sheet_name: Optional[str] = None,
    table_name: Optional[str] = None,
    table_filter: Optional[Tuple[Tuple[str, str], ...]] = None,
) -> Optional[Tuple[str, str, int]]:
    if bank_id is None:
        return None
    if sheet_name and table_name:
        return bank_id_locations.get((sheet_name, table_name, bank_id))
    tables = table_filter or INDICATOR_TABLES
    for sheet, table in tables:
        loc = bank_id_locations.get((sheet, table, bank_id))
        if loc:
            return loc
    return None


def _table_data_row_capacity(wb, sheet_name: str, table_name: str) -> int:
    from openpyxl.utils import range_boundaries

    ws = wb[sheet_name]
    tbl = ws.tables[table_name]
    ref = tbl.ref if hasattr(tbl, "ref") else tbl
    _min_col, min_row, _max_col, max_row = range_boundaries(ref)
    return max(0, max_row - min_row)


def _table_header_column(wb, sheet_name: str, table_name: str, header: str) -> Optional[int]:
    from openpyxl.utils import range_boundaries

    ws = wb[sheet_name]
    tbl = ws.tables[table_name]
    ref = tbl.ref if hasattr(tbl, "ref") else tbl
    min_col, min_row, max_col, _max_row = range_boundaries(ref)
    for col in range(min_col, max_col + 1):
        raw = ws.cell(min_row, col).value
        if str(raw or "").strip() == header:
            return col
    return None


def _is_table_row_empty(wb, sheet_name: str, table_name: str, row_offset: int, header: str = INDICATOR_HEADER) -> bool:
    col = _table_header_column(wb, sheet_name, table_name, header)
    if col is None:
        return True
    from openpyxl.utils import range_boundaries

    ws = wb[sheet_name]
    tbl = ws.tables[table_name]
    ref = tbl.ref if hasattr(tbl, "ref") else tbl
    _min_col, min_row, _max_col, _max_row = range_boundaries(ref)
    val = ws.cell(min_row + 1 + row_offset, col).value
    return val is None or str(val).strip() == ""


def _write_full_indicator_row(
    wb,
    sheet_name: str,
    table_name: str,
    row_offset: int,
    *,
    section_name: str,
    indicator_label: str,
    entry=None,
    key_to_header: Optional[Dict[str, str]] = None,
    value: Any = None,
    data_not_available: bool = False,
    bank_id: Optional[int] = None,
    yes_no: Optional[bool] = None,
) -> None:
    if bank_id is not None:
        write_table_cell(wb, sheet_name, table_name, row_offset, INDICATOR_ID_HEADER, bank_id)
    write_table_cell(wb, sheet_name, table_name, row_offset, SP_EF_HEADER, section_name)
    write_table_cell(wb, sheet_name, table_name, row_offset, INDICATOR_HEADER, indicator_label)
    headers = key_to_header or {}
    _write_indicator_entry(
        wb,
        sheet_name,
        table_name,
        row_offset,
        entry,
        headers,
        fallback_value=value,
        data_not_available=data_not_available,
        yes_no=yes_no,
    )


def _match_excel_row_to_bank_id(
    sp_ef: str,
    indicator: str,
    kpi_lookup: Dict[Tuple[str, str], int],
) -> Optional[int]:
    """Resolve indicator bank id from Excel row text via fuzzy KPI lookup."""
    best_bid: Optional[int] = None
    best_score = 0.0
    for (sp, ind), bid in kpi_lookup.items():
        if not _section_names_match(sp_ef, sp):
            continue
        score = _indicator_similarity(indicator, ind)
        if score > best_score:
            best_score = score
            best_bid = bid
    if best_bid and best_score >= INDICATOR_MATCH_THRESHOLD:
        return best_bid

    cross_bid: Optional[int] = None
    cross_score = 0.0
    for (_sp, ind), bid in kpi_lookup.items():
        score = _indicator_similarity(indicator, ind)
        if score > cross_score:
            cross_score = score
            cross_bid = bid
    if cross_bid and cross_score >= INDICATOR_CROSS_SECTION_THRESHOLD:
        return cross_bid
    return None


def _find_indicator_row(
    indicator_row_index: Dict[Tuple[str, str], Tuple[str, str, int]],
    section_name: str,
    label: str,
    *,
    bank_id: Optional[int] = None,
    kpi_lookup: Optional[Dict[Tuple[str, str], int]] = None,
) -> Optional[Tuple[str, str, int]]:
    direct = indicator_row_index.get(
        (_normalize_text(section_name), _normalize_indicator_label(label))
    )
    if direct:
        return direct
    best_loc: Optional[Tuple[str, str, int]] = None
    best_score = 0.0
    for (sp_norm, ind_norm), loc in indicator_row_index.items():
        if not _section_names_match(section_name, sp_norm):
            continue
        score = _indicator_similarity(ind_norm, label)
        if score > best_score:
            best_score = score
            best_loc = loc
    if best_loc and best_score >= INDICATOR_MATCH_THRESHOLD:
        return best_loc

    if bank_id and kpi_lookup:
        for (sp, ind), bid in kpi_lookup.items():
            if bid != bank_id:
                continue
            score = _indicator_similarity(label, ind)
            if score < INDICATOR_CROSS_SECTION_THRESHOLD:
                continue
            for (sp_norm, ind_norm), loc in indicator_row_index.items():
                if _indicator_similarity(ind_norm, ind) >= INDICATOR_MATCH_THRESHOLD:
                    return loc
            for (sp_norm, ind_norm), loc in indicator_row_index.items():
                row_score = _indicator_similarity(ind_norm, label)
                if row_score >= INDICATOR_CROSS_SECTION_THRESHOLD:
                    return loc
    return None


def _find_row_for_form_item(
    *,
    section_name: str,
    label: str,
    bank_id: Optional[int],
    indicator_row_index: Dict[Tuple[str, str], Tuple[str, str, int]],
    kpi_lookup: Dict[Tuple[str, str], int],
    bank_id_locations: Optional[Dict[Tuple[str, str, int], Tuple[str, str, int]]] = None,
) -> Optional[Tuple[str, str, int]]:
    """Locate the workbook row for a published T33 indicator form item."""
    if bank_id_locations and bank_id is not None:
        loc = _find_row_by_bank_id(
            bank_id,
            bank_id_locations,
            table_filter=INDICATOR_TABLES,
        )
        if loc:
            return loc

    loc = _find_indicator_row(
        indicator_row_index,
        section_name,
        label,
        bank_id=bank_id,
        kpi_lookup=kpi_lookup,
    )
    if loc:
        return loc
    if bank_id:
        for (_sp, ind), bid in kpi_lookup.items():
            if bid != bank_id:
                continue
            if _indicator_similarity(label, ind) < INDICATOR_CROSS_SECTION_THRESHOLD:
                continue
            for (sp_norm, ind_norm), candidate in indicator_row_index.items():
                if _indicator_similarity(ind_norm, ind) >= INDICATOR_MATCH_THRESHOLD:
                    return candidate
                if _indicator_similarity(ind_norm, label) >= INDICATOR_CROSS_SECTION_THRESHOLD:
                    return candidate
    return None


def _find_row_in_indicator_table(
    *,
    sheet_name: str,
    table_name: str,
    bank_id: int,
    bank_id_locations: Dict[Tuple[str, str, int], Tuple[str, str, int]],
    indicator_row_index: Dict[Tuple[str, str], Tuple[str, str, int]],
    kpi_lookup: Dict[Tuple[str, str], int],
) -> Optional[Tuple[str, str, int]]:
    """Locate a row within one indicator table (never cross-sheet/cross-table)."""
    loc = bank_id_locations.get((sheet_name, table_name, bank_id))
    if loc:
        return loc
    table_index = {
        key: value
        for key, value in indicator_row_index.items()
        if value[0] == sheet_name and value[1] == table_name
    }
    if not table_index:
        return None
    for (sp, ind), bid in kpi_lookup.items():
        if bid != bank_id:
            continue
        loc = _find_indicator_row(
            table_index,
            sp,
            ind,
            bank_id=bank_id,
            kpi_lookup=kpi_lookup,
        )
        if loc:
            return loc
    return None


def _allocate_table_overflow_row(
    wb,
    sheet_name: str,
    table_name: str,
    used_offsets: Set[int],
) -> Optional[int]:
    capacity = _table_data_row_capacity(wb, sheet_name, table_name)
    for offset in range(capacity):
        if offset in used_offsets:
            continue
        if _is_table_row_empty(wb, sheet_name, table_name, offset):
            used_offsets.add(offset)
            return offset
    return None


def _allocate_data_other_row(
    wb,
    *,
    section_name: str,
    label: str,
    used_offsets: Set[int],
) -> Optional[int]:
    return _allocate_table_overflow_row(wb, DATA_OTHER_SHEET, DATA_OTHER_TABLE, used_offsets)


def _resolve_item_by_section_and_indicator(
    ctx: UprImportContext,
    section_label_map: Dict[Tuple[str, str], int],
    sp_ef: str,
    indicator: str,
    kpi_lookup: Dict[Tuple[str, str], int],
) -> Optional[int]:
    """Resolve a T33 item from workbook indicator row text."""
    direct_key = (_normalize_text(sp_ef), _normalize_text(indicator))
    if direct_key in section_label_map:
        return section_label_map[direct_key]

    for (section_name, label), item_id in section_label_map.items():
        if _section_names_match(sp_ef, section_name) and _indicator_names_match(indicator, label):
            return item_id

    best_id: Optional[int] = None
    best_score = 0.0
    for (section_name, label), item_id in section_label_map.items():
        if not _section_names_match(sp_ef, section_name):
            continue
        score = _indicator_similarity(indicator, label)
        if score > best_score:
            best_score = score
            best_id = item_id
    if best_id and best_score >= INDICATOR_MATCH_THRESHOLD:
        return best_id

    bank_id = _match_excel_row_to_bank_id(sp_ef, indicator, kpi_lookup)
    if bank_id:
        return _resolve_item_for_workbook_indicator(ctx, bank_id, sp_ef)
    return None


def _resolve_item_for_workbook_indicator(
    ctx: UprImportContext,
    bank_id: int,
    sp_ef: str,
) -> Optional[int]:
    """Resolve T33 published-version form item for an indicator bank id and SP/EF section name."""
    section_map = ctx.items_by_bank_section.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get(bank_id)
    if section_map:
        sp_norm = _normalize_text(sp_ef)
        for section_name, item_id in section_map.items():
            if _normalize_text(section_name) == sp_norm:
                return item_id
            if sp_norm.replace("-", " ") == _normalize_text(section_name).replace("-", " "):
                return item_id
        if len(section_map) == 1:
            return next(iter(section_map.values()))
    from import_upr_excel_data import REPORTING_EXCEL_AREA_TO_SECTION

    for area_code, section in REPORTING_EXCEL_AREA_TO_SECTION.items():
        if _normalize_text(section) == _normalize_text(sp_ef):
            return _resolve_item_by_bank_and_area(ctx, REPORTING_COUNTRY_TEMPLATE_ID, bank_id, area_code)
    return ctx.items_by_bank_id.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get(bank_id)


def _funding_column_name(wb) -> str:
    headers, _ = read_named_table(wb, "Funding", "Data_funding1")
    for header in headers:
        if header and "funding" in header.lower():
            return header
    return "NS 2026 Total\nFunding "


def _reporting_funding_matrix_column() -> str:
    """Return the matrix column ``name`` for item 1403 (cell keys: ``{row}_{name}``)."""
    return _matrix_column_name_from_item_id(ITEM_REPORTING_COUNTRY_FUNDING)


def parse_ns_key_data(wb) -> Dict[str, Optional[float]]:
    out: Dict[str, Optional[float]] = {}
    for key, named in NS_DATA_NAMED_CELLS.items():
        out[key] = _parse_value_num(read_named_cell(wb, named))
    return out


def parse_indicators(
    wb,
    *,
    tables: Optional[Tuple[Tuple[str, str], ...]] = None,
    yes_no_bank_ids: Optional[Set[int]] = None,
    kpi_lookup: Optional[Dict[Tuple[str, str], int]] = None,
) -> List[Dict[str, Any]]:
    rows_out: List[Dict[str, Any]] = []
    table_list = tables or INDICATOR_TABLES
    for sheet_name, table_name in table_list:
        if sheet_name not in wb.sheetnames or table_name not in wb[sheet_name].tables:
            continue
        _, rows = read_named_table(wb, sheet_name, table_name)
        for row in rows:
            sp_ef = str(row.get("Strategic Priority / Enabling Function") or "").strip()
            indicator = str(row.get("Indicator") or "").strip()
            if not sp_ef or not indicator:
                continue
            applicable = str(row.get(INDICATOR_DNA_HEADER) or "").strip().lower()
            is_dna = "data not available" in applicable
            disagg = _parse_workbook_row_disagg(row)
            direct_val: Optional[float] = None
            for header, raw in row.items():
                if _normalize_workbook_header(header) == "total direct":
                    direct_val = _coerce_disagg_number(raw)
                    break
            combined_val = _parse_value_num(row.get(INDICATOR_TOTAL_HEADER))
            if disagg:
                value = None
            else:
                value = direct_val if direct_val is not None else combined_val
            bank_id = _parse_row_bank_id(row)
            resolved_bank_id = bank_id or _resolve_workbook_indicator_bank_id(
                {"bank_id": bank_id, "sp_ef": sp_ef, "indicator": indicator},
                kpi_lookup,
            )
            is_yes_no = bool(
                resolved_bank_id and yes_no_bank_ids and resolved_bank_id in yes_no_bank_ids
            )
            if is_yes_no:
                value = "yes" if "applicable" in applicable and not is_dna else "no"
                disagg = None
                is_dna = False
            elif not is_dna and "applicable" in applicable and disagg is None and value is None:
                # Numeric placeholders: Applicable with no values — skip on import.
                continue
            rows_out.append(
                {
                    "bank_id": bank_id,
                    "sp_ef": sp_ef,
                    "indicator": indicator,
                    "data_not_available": is_dna,
                    "applicable_text": applicable,
                    "value": value,
                    "disagg": disagg,
                    "sheet_name": sheet_name,
                    "table_name": table_name,
                }
            )
    return rows_out


def parse_emergency_slot_metadata(wb) -> Dict[int, Dict[str, str]]:
    """Read MDR code and appeal name for each emergency slot from Start sheet named cells."""
    slots: Dict[int, Dict[str, str]] = {}
    for _sheet, _table, slot_num, mdr_name, eo_name in EMERGENCY_SLOTS:
        mdr = str(read_named_cell(wb, mdr_name) or "").strip()
        name = str(read_named_cell(wb, eo_name) or "").strip()
        if mdr or name:
            slots[slot_num] = {"mdr_code": mdr, "appeal_name": name}
    return slots


def parse_funding(wb) -> Dict[str, Any]:
    funding_col = _funding_column_name(wb)
    exp_header = "NS 2026 Total\nExpenditure"
    _, rows1 = read_named_table(wb, "Funding", "Data_funding1")
    sources: Dict[str, Optional[float]] = {}
    total_expenditure: Optional[float] = None
    for row in rows1:
        attr = str(row.get("Attribute") or "").strip().upper()
        funding_val = _parse_value_num(row.get(funding_col))
        exp_val = _parse_value_num(row.get(exp_header))
        if attr == "TOTAL":
            total_expenditure = exp_val
        elif attr in FUNDING_SOURCE_ATTR_TO_ROW and funding_val is not None:
            sources[FUNDING_SOURCE_ATTR_TO_ROW[attr]] = funding_val

    breakdown: Dict[str, Dict[str, Optional[float]]] = {}
    _, rows2 = read_named_table(wb, "Funding", "Data_funding2")
    for row in rows2:
        area = str(row.get("Attribute") or "").strip()
        if not area or area.lower() == "total":
            continue
        row_name = REPORTING_SP_BREAKDOWN_AREA_TO_ROW.get(area)
        if not row_name:
            continue
        breakdown[row_name] = {
            "Funding (CHF)": _parse_value_num(row.get("Funding")),
            "Expenditure (CHF)": _parse_value_num(row.get("Expenditure")),
        }
    return {
        "sources": sources,
        "total_expenditure": total_expenditure,
        "breakdown": breakdown,
        "funding_column": funding_col.strip(),
    }


def _matrix_cell_scalar(raw: Any) -> Any:
    """Return the effective scalar from a matrix cell (handles original/modified payloads)."""
    if isinstance(raw, dict) and {"original", "modified"} & set(raw.keys()):
        chosen = raw.get("modified")
        if chosen in (None, ""):
            chosen = raw.get("original")
        return _matrix_cell_scalar(chosen)
    return raw


def _export_funding_breakdown(wb, breakdown_cells: Dict[str, Any]) -> None:
    """Write item 1405 SP/EF funding breakdown into Data_funding2."""
    cells = _normalize_matrix_cells(breakdown_cells)
    if not cells:
        return

    _, breakdown_rows = read_named_table(wb, "Funding", "Data_funding2")
    for idx, row in enumerate(breakdown_rows):
        area = str(row.get("Attribute") or "").strip()
        if not area or area.lower() == "total":
            continue
        row_name = REPORTING_SP_BREAKDOWN_AREA_TO_ROW.get(area)
        if not row_name:
            continue
        for matrix_col, excel_col in SP_BREAKDOWN_MATRIX_COL_TO_EXCEL.items():
            amount = _parse_value_num(_matrix_cell_scalar(cells.get(f"{row_name}_{matrix_col}")))
            if amount is not None:
                write_table_cell(wb, "Funding", "Data_funding2", idx, excel_col, amount)


def _matrix_cell_is_set(raw: Any) -> bool:
    """True when a matrix/checkbox cell is marked (handles nested original/modified payloads)."""
    if raw is None or raw == "":
        return False
    if isinstance(raw, dict):
        val = raw.get("modified", raw.get("original", raw.get("value")))
        if val is None or str(val).strip() in ("", "0", "0.0"):
            return False
        return _matrix_cell_is_set(val)
    if isinstance(raw, bool):
        return raw
    text = str(raw).strip().upper()
    if text in ("X", "1", "YES", "TRUE"):
        return True
    num = _parse_value_num(raw)
    return num is not None and num != 0


def _normalize_matrix_cells(raw: Dict[str, Any]) -> Dict[str, Any]:
    """Flatten matrix payloads so keys map to scalar/set values."""
    out: Dict[str, Any] = {}
    for key, val in raw.items():
        if isinstance(val, dict) and {"original", "modified"} & set(val.keys()):
            chosen = val.get("modified")
            if chosen in (None, ""):
                chosen = val.get("original")
            out[key] = chosen
        else:
            out[key] = val
    return out


def _matrix_item_config(form_item) -> Dict[str, Any]:
    if not form_item or not form_item.config:
        return {}
    cfg = form_item.config
    if isinstance(cfg, str):
        try:
            cfg = json.loads(cfg)
        except json.JSONDecodeError:
            return {}
    mc = cfg.get("matrix_config") or cfg
    return mc if isinstance(mc, dict) else {}


def _entity_ids_from_matrix_disagg(
    disagg_data: Any,
    *,
    tick_column_names: Optional[List[str]] = None,
    require_tick: bool = False,
) -> Set[int]:
    """Extract row entity ids from matrix disagg_data (same rules as matrix auto-load API)."""
    from app.services.forms.variable_resolution_service import VariableResolutionService

    if not isinstance(disagg_data, dict):
        return set()
    tick_names = set(tick_column_names or [])
    entity_info: Dict[int, bool] = {}
    for key, raw in disagg_data.items():
        if key == "_table" or "_" not in str(key):
            continue
        prefix, column_name = str(key).split("_", 1)
        try:
            entity_id = int(prefix)
        except (ValueError, TypeError):
            continue
        if entity_id not in entity_info:
            entity_info[entity_id] = False
        if not require_tick or not tick_names:
            entity_info[entity_id] = True
            continue
        if column_name in tick_names:
            val = VariableResolutionService._effective_matrix_cell_value(raw)
            if val == 1 or val == "1" or val is True:
                entity_info[entity_id] = True
    return {eid for eid, ok in entity_info.items() if ok}


def _autoload_tick_column_names(matrix_config: Dict[str, Any], template_variables: Dict[str, Any]) -> List[str]:
    """Tick column names in the *source* plan matrix (matrix_column_name on linked variables)."""
    names: List[str] = []
    for col in matrix_config.get("columns") or []:
        if not isinstance(col, dict):
            continue
        col_type = str(col.get("type") or "number").lower()
        if col_type != "tick":
            continue
        var_name = col.get("variable") or col.get("variable_name")
        if var_name and isinstance(template_variables.get(var_name), dict):
            matrix_col = template_variables[var_name].get("matrix_column_name")
            if matrix_col:
                names.append(str(matrix_col))
                continue
        col_name = col.get("name")
        if col_name:
            names.append(str(col_name))
    return names


def _resolve_autoloaded_bilateral_ns_ids(aes, form_item_id: int) -> Set[int]:
    """
    NS ids that the reporting bilateral matrix auto-loads from the country plan assignment.
    Mirrors /api/v1/matrix/auto-load-entities without tick filtering when no tick columns exist.
    """
    from app.models.assignments import AssignedForm
    from app.models.form_items import FormItem
    from app.models.forms import FormData, FormTemplateVersion
    from app.services.forms.variable_resolution_service import VariableResolutionService

    item = FormItem.query.get(form_item_id)
    matrix_config = _matrix_item_config(item)
    if not matrix_config.get("auto_load_entities"):
        return set()

    assigned_form = getattr(aes, "assigned_form", None)
    template = getattr(assigned_form, "template", None) if assigned_form else None
    version_id = getattr(template, "published_version_id", None) if template else None
    version = FormTemplateVersion.query.get(version_id) if version_id else None
    template_variables = (version.variables if version and version.variables else {}) or {}

    variable_configs: List[Dict[str, Any]] = []
    for col in matrix_config.get("columns") or []:
        if not isinstance(col, dict) or not col.get("is_variable"):
            continue
        var_name = col.get("variable") or col.get("variable_name")
        if not var_name:
            continue
        var_cfg = template_variables.get(var_name)
        if isinstance(var_cfg, dict):
            variable_configs.append(var_cfg)

    if not variable_configs:
        return set()

    tick_column_names = _autoload_tick_column_names(matrix_config, template_variables)
    require_tick = bool(tick_column_names)
    entity_ids: Set[int] = set()

    for var_cfg in variable_configs:
        source_template_id = var_cfg.get("source_template_id")
        source_form_item_id = var_cfg.get("source_form_item_id")
        source_assignment_period = var_cfg.get("source_assignment_period")
        if not all([source_template_id, source_form_item_id, source_assignment_period]):
            continue

        effective_period = VariableResolutionService._resolve_effective_period(
            source_assignment_period,
            int(source_template_id),
            aes,
        )
        if not effective_period:
            continue

        source_assigned_form = AssignedForm.query.filter_by(
            template_id=int(source_template_id),
            period_name=effective_period,
        ).first()
        if not source_assigned_form:
            continue

        matching_statuses = [
            status
            for status in source_assigned_form.entity_statuses
            if status.entity_id == aes.entity_id and status.entity_type == aes.entity_type
        ]
        if not matching_statuses:
            continue

        entries = FormData.query.filter(
            FormData.assignment_entity_status_id.in_([s.id for s in matching_statuses]),
            FormData.form_item_id == int(source_form_item_id),
        ).all()

        for entry in entries:
            entity_ids.update(
                _entity_ids_from_matrix_disagg(
                    entry.disagg_data,
                    tick_column_names=tick_column_names,
                    require_tick=require_tick,
                )
            )

    return entity_ids


def _parse_support_matrix_ns_ids(cells: Dict[str, Any]) -> Set[int]:
    """All partner NS ids present in the reporting bilateral support matrix."""
    ids: Set[int] = set()
    for cell_key in _normalize_matrix_cells(cells):
        match = _SUPPORT_CELL_KEY_RE.match(str(cell_key).strip())
        if match:
            ids.add(int(match.group(1)))
    return ids


def _clear_bilateral_table(wb, sheet_name: str, table_name: str) -> None:
    """Clear partner NS names and Supported ticks so export replaces template formulas."""
    capacity = _table_data_row_capacity(wb, sheet_name, table_name)
    for offset in range(capacity):
        _write_bilateral_ns_source_cell(wb, sheet_name, table_name, offset, "")
        for col in SUPPORT_COLUMN_TO_AREA:
            write_table_cell(wb, sheet_name, table_name, offset, col, None)


def _write_bilateral_ns_rows(
    wb,
    sheet_name: str,
    table_name: str,
    ns_ids: List[int],
    id_to_name: Dict[int, str],
) -> Dict[int, int]:
    """Write NS names into column C; return ns_id -> 0-based row offset."""
    offsets: Dict[int, int] = {}
    for offset, ns_id in enumerate(ns_ids):
        ns_name = id_to_name.get(int(ns_id), "")
        if not ns_name:
            continue
        _write_bilateral_ns_source_cell(wb, sheet_name, table_name, offset, ns_name)
        offsets[int(ns_id)] = offset
    return offsets


def _bilateral_table_row_info(wb, sheet_name: str, table_name: str) -> Tuple[int, int]:
    from openpyxl.utils import range_boundaries

    ws = wb[sheet_name]
    tbl = ws.tables[table_name]
    ref = tbl.ref if hasattr(tbl, "ref") else tbl
    min_col, min_row, _max_col, _max_row = range_boundaries(ref)
    return min_col, min_row


def _bilateral_ns_name_for_row(wb, sheet_name: str, table_name: str, row_offset: int) -> str:
    """Partner NS name lives in column C; the table NS column is usually a formula."""
    _min_col, min_row = _bilateral_table_row_info(wb, sheet_name, table_name)
    ws = wb[sheet_name]
    return str(ws.cell(min_row + 1 + row_offset, BILATERAL_NS_SOURCE_COL).value or "").strip()


def _write_bilateral_ns_source_cell(
    wb,
    sheet_name: str,
    table_name: str,
    row_offset: int,
    ns_name: str,
) -> None:
    _min_col, min_row = _bilateral_table_row_info(wb, sheet_name, table_name)
    wb[sheet_name].cell(min_row + 1 + row_offset, BILATERAL_NS_SOURCE_COL).value = ns_name


def _parse_support_matrix_ticks(cells: Dict[str, Any]) -> Dict[int, Dict[str, bool]]:
    """Parse item 1407 matrix keys into ns_id -> {area: True}."""
    ticks: Dict[int, Dict[str, bool]] = {}
    normalized = _normalize_matrix_cells(cells)
    for cell_key, raw in normalized.items():
        if not _matrix_cell_is_set(raw):
            continue
        match = _SUPPORT_CELL_KEY_RE.match(str(cell_key).strip())
        if not match:
            continue
        ns_id = int(match.group(1))
        area = match.group(2)
        ticks.setdefault(ns_id, {})[area] = True
    return ticks


def _export_bilateral_support(wb, ctx: UprImportContext, support_cells: Dict[str, Any], aes) -> None:
    """
    Write bilateral support: planned/autoloaded PNS in Data_act, manually added PNS in Data_act2.
    Ignores template PlannedActivities formulas (year-specific); NS names come from assignment data.
    """
    ticks_by_ns = _parse_support_matrix_ticks(support_cells)
    autoloaded_ids = _resolve_autoloaded_bilateral_ns_ids(aes, ITEM_REPORTING_COUNTRY_SUPPORT)
    reported_ids = _parse_support_matrix_ns_ids(support_cells)
    all_ns_ids = autoloaded_ids | reported_ids | set(ticks_by_ns.keys())
    if not all_ns_ids:
        return

    planned_ids = sorted(autoloaded_ids)
    manual_ids = sorted(all_ns_ids - autoloaded_ids)

    id_to_display_name: Dict[int, str] = {}
    try:
        from app.models.organization import NationalSociety

        for ns in NationalSociety.query.filter(NationalSociety.id.in_(list(all_ns_ids))).all():
            id_to_display_name[int(ns.id)] = (ns.name or "").strip()
    except Exception:
        id_to_display_name = {}

    for ns_id in all_ns_ids:
        if int(ns_id) in id_to_display_name:
            continue
        for db_name, db_id in ctx.ns_name_to_id.items():
            if db_id == int(ns_id):
                id_to_display_name[int(ns_id)] = db_name
                break

    planned_sheet, planned_table = BILATERAL_PLANNED_TABLE
    manual_sheet, manual_table = BILATERAL_MANUAL_TABLE
    _clear_bilateral_table(wb, planned_sheet, planned_table)
    _clear_bilateral_table(wb, manual_sheet, manual_table)

    planned_offsets = _write_bilateral_ns_rows(wb, planned_sheet, planned_table, planned_ids, id_to_display_name)
    manual_offsets = _write_bilateral_ns_rows(wb, manual_sheet, manual_table, manual_ids, id_to_display_name)

    for ns_id, areas in ticks_by_ns.items():
        offset = planned_offsets.get(int(ns_id))
        sheet_name, table_name = planned_sheet, planned_table
        if offset is None:
            offset = manual_offsets.get(int(ns_id))
            sheet_name, table_name = manual_sheet, manual_table
        if offset is None:
            continue
        for area in areas:
            col = AREA_TO_SUPPORT_COLUMN.get(area)
            if not col:
                continue
            write_table_cell(wb, sheet_name, table_name, offset, col, "X")


def parse_bilateral_support(wb) -> List[Dict[str, Any]]:
    rows_out: List[Dict[str, Any]] = []
    for sheet_name, table_name in BILATERAL_TABLES:
        capacity = _table_data_row_capacity(wb, sheet_name, table_name)
        for offset in range(capacity):
            ns_name = _bilateral_ns_name_for_row(wb, sheet_name, table_name, offset)
            if not ns_name or ns_name.upper() == "NS":
                continue
            supported_areas: List[str] = []
            for col, area in SUPPORT_COLUMN_TO_AREA.items():
                val = read_table_cell(wb, sheet_name, table_name, offset, col)
                if val is not None and str(val).strip().upper() in ("X", "1", "YES", "TRUE"):
                    supported_areas.append(area)
            if supported_areas:
                rows_out.append({"ns_name": ns_name, "areas": supported_areas})
    return rows_out


def parse_comments(wb) -> str:
    """Read the single comments cell (aligned with the T33 comments textarea)."""
    return str(read_named_cell(wb, COMMENTS_NAMED_CELL) or "").strip()


def _format_emergency_operation_display(name: str, code: str) -> str:
    name = (name or "").strip()
    code = (code or "").strip()
    if name and code:
        return f"{name} ({code})"
    return name or code


def _upsert_emergency_repeat_choice(
    *,
    repeat_instance,
    choice_item_id: int,
    appeal_name: str,
    mdr_code: str,
) -> None:
    """Persist the per-repeat emergency_operations single choice (item 1374 pattern)."""
    from app.models.forms import RepeatGroupData

    name = (appeal_name or "").strip()
    code = (mdr_code or "").strip()
    if not name and not code:
        return

    display = _format_emergency_operation_display(name, code)
    entry = RepeatGroupData.query.filter_by(
        repeat_instance_id=repeat_instance.id,
        form_item_id=choice_item_id,
    ).first()
    if not entry:
        entry = RepeatGroupData(
            repeat_instance_id=repeat_instance.id,
            form_item_id=choice_item_id,
        )
        from app.extensions import db

        db.session.add(entry)

    entry.value = display
    entry.disagg_data = {"name": name, "code": code}
    entry.disagg_type = "emergency_operation"
    entry.data_not_available = False
    entry.not_applicable = False


def _ensure_repeat_instance(
    aes_id: int,
    repeat_section_id: int,
    instance_number: int,
    label: Optional[str] = None,
    *,
    user_id: Optional[int] = None,
):
    from app.models.forms import RepeatGroupInstance

    inst = RepeatGroupInstance.query.filter_by(
        assignment_entity_status_id=aes_id,
        section_id=repeat_section_id,
        instance_number=instance_number,
    ).first()
    if not inst:
        effective_user_id = int(user_id) if user_id is not None else _default_user_id_for_import()
        inst = RepeatGroupInstance(
            assignment_entity_status_id=aes_id,
            section_id=repeat_section_id,
            instance_number=instance_number,
            instance_label=label or f"Emergency Appeal {instance_number}",
            created_by_user_id=effective_user_id,
        )
        from app.extensions import db

        db.session.add(inst)
    elif label:
        inst.instance_label = label
    return inst


def _upsert_dynamic_indicator(
    *,
    aes_id: int,
    section_id: int,
    indicator_bank_id: int,
    repeat_instance_number: Optional[int],
    value: Any,
    data_not_available: bool,
    user_id: int,
    order: float,
    disagg_payload: Optional[Dict[str, Any]] = None,
) -> str:
    from app.extensions import db
    from app.models.forms import DynamicIndicatorData

    row = DynamicIndicatorData.query.filter_by(
        assignment_entity_status_id=aes_id,
        section_id=section_id,
        indicator_bank_id=indicator_bank_id,
        repeat_instance_number=repeat_instance_number,
    ).first()
    action = "updated"
    if not row:
        row = DynamicIndicatorData(
            assignment_entity_status_id=aes_id,
            section_id=section_id,
            indicator_bank_id=indicator_bank_id,
            repeat_instance_number=repeat_instance_number,
            added_by_user_id=user_id,
            order=order,
        )
        db.session.add(row)
        action = "inserted"
    if data_not_available:
        row.set_data_availability(data_not_available=True)
    elif disagg_payload:
        row.set_disaggregated_data(disagg_payload["mode"], disagg_payload["values"])
    elif value is not None:
        if isinstance(value, float) and value.is_integer():
            row.set_simple_value(str(int(value)))
        else:
            row.set_simple_value(str(value))
    return action


def _lookup_existing_dynamic_assignment_id(
    aes_id: int,
    section_id: int,
    indicator_bank_id: int,
    repeat_instance_number: Optional[int],
) -> Optional[int]:
    from app.models.forms import DynamicIndicatorData

    row = DynamicIndicatorData.query.filter_by(
        assignment_entity_status_id=aes_id,
        section_id=section_id,
        indicator_bank_id=indicator_bank_id,
        repeat_instance_number=repeat_instance_number,
    ).first()
    return int(row.id) if row else None


def _collect_workbook_dynamic_indicator_entries(
    aes_id: int,
    wb,
    ctx: UprImportContext,
    *,
    section_label_map: Dict[Tuple[str, str], int],
    kpi_lookup: Dict[Tuple[str, str], int],
) -> List[Dict[str, Any]]:
    """Parse dynamic indicator rows from a workbook without writing to the database."""
    section_ids = _load_upr_country_reporting_section_ids()
    ea_dynamic = section_ids.get("ea_dynamic")
    other_dynamic = section_ids.get("other_dynamic")
    yes_no_bank_ids = _load_workbook_yes_no_bank_ids(wb, kpi_lookup)
    entries: List[Dict[str, Any]] = []
    order = 0.0

    emergency_tables = tuple((sheet, table) for sheet, table, *_rest in EMERGENCY_SLOTS)
    ea_rows = parse_indicators(
        wb,
        tables=emergency_tables,
        yes_no_bank_ids=yes_no_bank_ids,
        kpi_lookup=kpi_lookup,
    )
    for row in ea_rows:
        if not ea_dynamic:
            continue
        bank_id = _resolve_workbook_indicator_bank_id(row, kpi_lookup)
        if not bank_id:
            ctx.warnings.append(
                f"No indicator bank id for emergency indicator {row['indicator']!r} ({row['sp_ef']!r})"
            )
            continue
        slot_num = next(
            (num for sheet, table, num, *_rest in EMERGENCY_SLOTS if sheet == row["sheet_name"]),
            None,
        )
        if slot_num is None:
            continue
        value, is_dna, disagg, should_import = _resolve_indicator_import_value(
            row, bank_id, yes_no_bank_ids
        )
        if not should_import:
            continue
        order += 1.0
        entries.append(
            {
                "section_id": ea_dynamic,
                "indicator_bank_id": bank_id,
                "repeat_instance_number": slot_num,
                "value": value,
                "data_not_available": is_dna,
                "disagg_data": disagg,
                "order": order,
                "existing_assignment_id": _lookup_existing_dynamic_assignment_id(
                    aes_id, ea_dynamic, bank_id, slot_num
                ),
            }
        )

    if other_dynamic and "Overall action Indicators" in wb.sheetnames:
        if "Data_other" in wb["Overall action Indicators"].tables:
            other_rows = parse_indicators(
                wb,
                tables=(("Overall action Indicators", "Data_other"),),
                yes_no_bank_ids=yes_no_bank_ids,
                kpi_lookup=kpi_lookup,
            )
            for row in other_rows:
                item_id = _resolve_item_by_section_and_indicator(
                    ctx,
                    section_label_map,
                    row["sp_ef"],
                    row["indicator"],
                    kpi_lookup,
                )
                if item_id:
                    continue
                bank_id = _resolve_workbook_indicator_bank_id(row, kpi_lookup)
                if not bank_id:
                    continue
                value, is_dna, disagg, should_import = _resolve_indicator_import_value(
                    row, bank_id, yes_no_bank_ids
                )
                if not should_import:
                    continue
                order += 1.0
                entries.append(
                    {
                        "section_id": other_dynamic,
                        "indicator_bank_id": bank_id,
                        "repeat_instance_number": None,
                        "value": value,
                        "data_not_available": is_dna,
                        "disagg_data": disagg,
                        "order": order,
                        "existing_assignment_id": _lookup_existing_dynamic_assignment_id(
                            aes_id, other_dynamic, bank_id, None
                        ),
                    }
                )
    return entries


def _collect_workbook_repeat_slot_entries(
    aes_id: int,
    wb,
    dynamic_entries: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Build repeat-slot metadata for client-side repeat entry creation."""
    section_ids = _load_upr_country_reporting_section_ids()
    ea_repeat = section_ids.get("ea_repeat")
    if not ea_repeat:
        return []

    choice_item_id = _load_upr_country_reporting_emergency_choice_item_id(ea_repeat)
    slot_meta = parse_emergency_slot_metadata(wb)
    required_slots: Set[int] = set(slot_meta.keys())
    for entry in dynamic_entries:
        repeat_num = entry.get("repeat_instance_number")
        if repeat_num is not None:
            required_slots.add(int(repeat_num))

    out: List[Dict[str, Any]] = []
    for slot_num in sorted(required_slots):
        meta = slot_meta.get(slot_num) or {}
        appeal_name = meta.get("appeal_name") or ""
        mdr_code = meta.get("mdr_code") or ""
        out.append(
            {
                "repeat_section_id": ea_repeat,
                "slot_num": slot_num,
                "mdr_code": mdr_code,
                "appeal_name": appeal_name,
                "choice_item_id": choice_item_id,
                "display_value": _format_emergency_operation_display(appeal_name, mdr_code),
            }
        )
    return out


def import_rows_to_client_payload(
    import_rows: List[Dict[str, str]],
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Split form_data import rows into static field values and matrix cell maps."""
    fields: Dict[str, Any] = {}
    matrices: Dict[str, Any] = {}
    for row in import_rows:
        item_id = str(row.get(COL_ITEM) or "").strip()
        if not item_id:
            continue
        data_na = str(row.get(COL_DATA_NA) or "").strip() == "1"
        disagg_raw = str(row.get(COL_DISAGG) or "").strip()
        value = row.get(COL_VALUE)

        if disagg_raw:
            try:
                disagg = json.loads(disagg_raw)
            except json.JSONDecodeError:
                continue
            if isinstance(disagg, dict) and ("mode" in disagg or "values" in disagg):
                fields[item_id] = {"disagg_data": disagg, "data_not_available": data_na}
            elif isinstance(disagg, dict):
                matrices[item_id] = disagg
            continue
        if data_na:
            fields[item_id] = {"data_not_available": True}
        elif value is not None and str(value).strip() != "":
            fields[item_id] = {"value": str(value)}
    return fields, matrices


def dedupe_upr_import_warnings(warnings: Iterable[str]) -> List[str]:
    """Return unique import warnings, collapsing redundant period-mismatch messages."""
    seen: Set[str] = set()
    out: List[str] = []
    period_noted = False
    for raw in warnings:
        text = str(raw or "").strip()
        if not text or text in seen:
            continue
        lower = text.lower()
        if "period" in lower and ("does not match" in lower or "differs from" in lower):
            if period_noted:
                continue
            period_noted = True
        seen.add(text)
        out.append(text)
    return out


def build_upr_country_reporting_client_payload(
    aes_id: int,
    wb,
    ctx: UprImportContext,
    import_rows: List[Dict[str, str]],
    *,
    iso3: str,
    period: str,
) -> Dict[str, Any]:
    """Build a JSON-serializable payload for staging an Excel import in the browser."""
    fields, matrices = import_rows_to_client_payload(import_rows)
    section_label_map = _load_items_by_section_label(REPORTING_COUNTRY_TEMPLATE_ID)
    kpi_lookup = build_kpi_lookup(wb)
    dynamic_entries = _collect_workbook_dynamic_indicator_entries(
        aes_id,
        wb,
        ctx,
        section_label_map=section_label_map,
        kpi_lookup=kpi_lookup,
    )
    repeat_slots = _collect_workbook_repeat_slot_entries(aes_id, wb, dynamic_entries)
    return {
        "fields": fields,
        "matrices": matrices,
        "dynamic_indicators": dynamic_entries,
        "repeat_slots": repeat_slots,
        "meta": {"iso3": iso3, "period": period},
    }


def _import_dynamic_indicators_from_workbook(
    aes_id: int,
    wb,
    ctx: UprImportContext,
    *,
    section_label_map: Dict[Tuple[str, str], int],
    kpi_lookup: Dict[Tuple[str, str], int],
    dry_run: bool = False,
) -> Dict[str, int]:
    from app.extensions import db

    section_ids = _load_upr_country_reporting_section_ids()
    ea_repeat = section_ids.get("ea_repeat")
    user_id = _default_user_id_for_import()
    stats = {"inserted": 0, "updated": 0}

    if dry_run:
        return stats

    slot_meta = parse_emergency_slot_metadata(wb)
    choice_item_id = _load_upr_country_reporting_emergency_choice_item_id(ea_repeat)
    for slot_num, meta in slot_meta.items():
        if ea_repeat:
            inst = _ensure_repeat_instance(
                aes_id,
                ea_repeat,
                slot_num,
                meta.get("appeal_name") or meta.get("mdr_code"),
                user_id=user_id,
            )
            if inst and choice_item_id:
                _upsert_emergency_repeat_choice(
                    repeat_instance=inst,
                    choice_item_id=choice_item_id,
                    appeal_name=meta.get("appeal_name") or "",
                    mdr_code=meta.get("mdr_code") or "",
                )

    for entry in _collect_workbook_dynamic_indicator_entries(
        aes_id,
        wb,
        ctx,
        section_label_map=section_label_map,
        kpi_lookup=kpi_lookup,
    ):
        action = _upsert_dynamic_indicator(
            aes_id=aes_id,
            section_id=entry["section_id"],
            indicator_bank_id=entry["indicator_bank_id"],
            repeat_instance_number=entry.get("repeat_instance_number"),
            value=entry.get("value"),
            data_not_available=bool(entry.get("data_not_available")),
            user_id=user_id,
            order=float(entry.get("order") or 0),
            disagg_payload=entry.get("disagg_data"),
        )
        stats[action] += 1

    db.session.commit()
    return stats


def _load_dynamic_indicator_map(aes_id: int) -> Dict[Tuple[int, Optional[int]], List[Any]]:
    from app.models.forms import DynamicIndicatorData
    from sqlalchemy.orm import joinedload

    grouped: Dict[Tuple[int, Optional[int]], List[Any]] = {}
    rows = (
        DynamicIndicatorData.query.filter_by(assignment_entity_status_id=aes_id)
        .options(joinedload(DynamicIndicatorData.indicator_bank))
        .order_by(DynamicIndicatorData.order.asc())
        .all()
    )
    for row in rows:
        key = (int(row.section_id), row.repeat_instance_number)
        grouped.setdefault(key, []).append(row)
    return grouped


def _export_emergency_slots(
    wb,
    aes,
    *,
    slots: Optional[List[Optional[Dict[str, str]]]] = None,
) -> None:
    resolved = slots if slots is not None else _resolve_emergency_slots_for_export(aes)
    for _sheet, _table, slot_num, mdr_name, eo_name in EMERGENCY_SLOTS:
        slot = resolved[slot_num - 1] if slot_num - 1 < len(resolved) else None
        if not slot:
            continue
        write_named_cell(wb, mdr_name, slot.get("code") or "")
        write_named_cell(wb, eo_name, slot.get("name") or slot.get("label") or "")


def _export_dynamic_indicators(
    wb,
    *,
    dynamic_rows: List[Any],
    key_to_header: Dict[str, str],
    bank_id_locations: Dict[Tuple[str, str, int], Tuple[str, str, int]],
    sheet_name: str,
    table_name: str,
    indicator_row_index: Optional[Dict[Tuple[str, str], Tuple[str, str, int]]] = None,
    kpi_lookup: Optional[Dict[Tuple[str, str], int]] = None,
    kpi_display: Optional[Dict[int, Tuple[str, str]]] = None,
    overflow_used: Optional[Dict[Tuple[str, str], Set[int]]] = None,
) -> None:
    index = indicator_row_index or {}
    lookup = kpi_lookup or {}
    display_map = kpi_display or {}
    used_offsets = (overflow_used or {}).setdefault((sheet_name, table_name), set())

    for dyn in dynamic_rows:
        bank_id = int(dyn.indicator_bank_id)
        is_dna = bool(getattr(dyn, "is_data_not_available", False))
        loc = _find_row_in_indicator_table(
            sheet_name=sheet_name,
            table_name=table_name,
            bank_id=bank_id,
            bank_id_locations=bank_id_locations,
            indicator_row_index=index,
            kpi_lookup=lookup,
        )
        if loc:
            loc_sheet, loc_table, idx = loc
            _write_indicator_entry(
                wb,
                loc_sheet,
                loc_table,
                idx,
                dyn,
                key_to_header,
                data_not_available=is_dna,
            )
            continue

        if is_dna and _entry_disagg_payload(dyn) is None and _scalar_value(dyn) is None:
            continue

        sp_ef, indicator_label = display_map.get(bank_id, ("", ""))
        if not indicator_label:
            ib = getattr(dyn, "indicator_bank", None)
            if ib is not None:
                indicator_label = str(getattr(ib, "name", None) or "").strip()
        if not sp_ef or not indicator_label:
            for (sp, ind), bid in lookup.items():
                if bid != bank_id:
                    continue
                sp_ef = sp_ef or sp
                indicator_label = indicator_label or ind
                break
        if not indicator_label:
            continue

        offset = _allocate_table_overflow_row(wb, sheet_name, table_name, used_offsets)
        if offset is None:
            continue
        _write_full_indicator_row(
            wb,
            sheet_name,
            table_name,
            offset,
            section_name=sp_ef,
            indicator_label=indicator_label,
            entry=dyn,
            key_to_header=key_to_header,
            data_not_available=is_dna,
            bank_id=bank_id,
        )


def _load_assignment_meta(aes_id: int) -> Tuple[Any, str, str, str]:
    from app.models.assignments import AssignmentEntityStatus
    from app.utils.api_serialization import _country_for_aes

    aes = AssignmentEntityStatus.query.get(aes_id)
    if not aes:
        raise ValueError(f"Assignment {aes_id} not found")
    country = _country_for_aes(aes)
    country_name = (country.name if country else "") or ""
    iso3 = (country.iso3 if country else "") or ""
    period = (aes.assigned_form.period_name if aes.assigned_form else "") or ""
    template_id = int(getattr(aes.assigned_form, "template_id", 0) or 0)
    if template_id != REPORTING_COUNTRY_TEMPLATE_ID:
        raise ValueError(f"Assignment {aes_id} is template {template_id}, not T33")
    return aes, country_name.strip(), iso3.strip().upper(), period.strip()


def _load_form_data_map(aes_id: int) -> Dict[int, Any]:
    from app.models.forms import FormData

    entries = FormData.query.filter_by(assignment_entity_status_id=aes_id).all()
    return {int(entry.form_item_id): entry for entry in entries if entry.form_item_id}


def _matrix_cells(entry) -> Dict[str, Any]:
    if not entry:
        return {}
    raw = entry.disagg_data
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str) and raw.strip():
        try:
            parsed = json.loads(raw)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}
    return {}


def _scalar_value(entry) -> Any:
    if not entry:
        return None
    if entry.is_data_not_available:
        return None
    return entry.value


def _workbook_table_headers(wb, sheet_name: str, table_name: str) -> Optional[List[str]]:
    """Return Excel table header row values, or None when the sheet/table is missing."""
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    if table_name not in ws.tables:
        return None
    from openpyxl.utils import range_boundaries

    tbl = ws.tables[table_name]
    ref = tbl.ref if hasattr(tbl, "ref") else tbl
    min_col, min_row, max_col, _max_row = range_boundaries(ref)
    headers: List[str] = []
    for col in range(min_col, max_col + 1):
        raw = ws.cell(min_row, col).value
        headers.append(str(raw).strip() if raw is not None else "")
    return headers


def _header_set_includes(headers: Optional[List[str]], required: Iterable[str]) -> bool:
    if not headers:
        return False
    normalized = {_normalize_workbook_header(h) for h in headers if h}
    for header in required:
        if _normalize_workbook_header(header) not in normalized:
            return False
    return True


def validate_upr_country_reporting_import_file(
    wb,
    *,
    expected_country: str = "",
    expected_period: str = "",
) -> Dict[str, Any]:
    """Validate a workbook before UPR Country Reporting import (structure + assignment match).

    Returns dict with keys: valid, message, errors, warnings, preview.
    """
    errors: List[str] = []
    warnings: List[str] = []
    preview: Dict[str, Any] = {
        "country": None,
        "period": None,
        "round_code": None,
        "kpi_count": 0,
        "core_indicator_rows": 0,
    }

    generic_hits = [name for name in GENERIC_FORM_EXPORT_SHEETS if name in wb.sheetnames]
    if generic_hits:
        errors.append(
            "This file looks like a generic form Excel export (sheets: "
            f"{', '.join(generic_hits)}), not a UPR Country Reporting workbook. "
            "Use Export UPR Country Reporting from this assignment and fill that template."
        )

    for named_range in UPR_COUNTRY_REPORTING_REQUIRED_NAMED_RANGES:
        if named_range not in wb.defined_names:
            errors.append(
                f"Missing named range {named_range!r}. "
                "Download a current UPR Country Reporting template from this assignment."
            )

    for sheet_name, table_name, required_headers in UPR_COUNTRY_REPORTING_REQUIRED_TABLES:
        headers = _workbook_table_headers(wb, sheet_name, table_name)
        if headers is None:
            if sheet_name not in wb.sheetnames:
                errors.append(f"Missing sheet {sheet_name!r}.")
            else:
                errors.append(
                    f"Missing table {table_name!r} on sheet {sheet_name!r}. "
                    "Older IFRC templates are not compatible with this import."
                )
            continue
        if required_headers and not _header_set_includes(headers, required_headers):
            normalized = {_normalize_workbook_header(h) for h in headers if h}
            missing = [
                header
                for header in required_headers
                if _normalize_workbook_header(header) not in normalized
            ]
            if INDICATOR_ID_HEADER in missing:
                errors.append(
                    "The Overall action Indicators table is missing the ID column. "
                    "Export a fresh UPR Country Reporting template from this assignment "
                    "(older Midyear templates without bank IDs cannot be imported)."
                )
            else:
                errors.append(
                    f"Table {table_name!r} on {sheet_name!r} is missing required columns: "
                    f"{', '.join(missing)}."
                )

    round_code, period_name = parse_version(wb)
    preview["round_code"] = round_code or None
    preview["period"] = period_name or None
    country_name = str(read_named_cell(wb, "Data_Country") or "").strip()
    preview["country"] = country_name or None

    if not round_code:
        errors.append(
            "Missing or unreadable Version cell. "
            "Use a UPR Country Reporting template exported from this assignment."
        )
    elif not any(round_code.startswith(prefix) for prefix in UPR_COUNTRY_REPORTING_COMPATIBLE_ROUND_PREFIXES):
        errors.append(
            f"Unsupported workbook version {round_code!r}. "
            f"Expected a reporting round code starting with "
            f"{' or '.join(UPR_COUNTRY_REPORTING_COMPATIBLE_ROUND_PREFIXES)}."
        )

    if not errors:
        try:
            kpi_lookup = build_kpi_lookup(wb)
        except Exception:
            kpi_lookup = {}
            errors.append(
                "Could not read the Indicators list (Final) table. "
                "Ensure the workbook is an unmodified UPR Country Reporting template."
            )
        else:
            preview["kpi_count"] = len(kpi_lookup)
            if not kpi_lookup:
                errors.append(
                    "Indicators list (Final) contains no KPI mappings. "
                    "Download a current template from this assignment."
                )

    if not errors and "Overall action Indicators" in wb.sheetnames:
        try:
            _, core_rows = read_named_table(wb, "Overall action Indicators", "Data_core")
            preview["core_indicator_rows"] = len(core_rows)
            if not any(_parse_row_bank_id(row) for row in core_rows):
                errors.append(
                    "Overall action Indicators rows have no bank IDs. "
                    "Export a fresh UPR Country Reporting template from this assignment."
                )
        except Exception:
            errors.append("Could not read Overall action Indicators (Data_core).")

    if expected_period and period_name and period_name != expected_period:
        warnings.append(
            f"Workbook period {period_name!r} differs from this assignment ({expected_period!r}). "
            "Values will be loaded into the current assignment."
        )
    if expected_country and country_name:
        if _normalize_text(expected_country) != _normalize_text(country_name):
            errors.append(
                f"Workbook country {country_name!r} does not match this assignment ({expected_country!r}). "
                "Use the file exported for this country assignment."
            )
    elif expected_country and not country_name:
        warnings.append("Workbook country is blank on the Start sheet.")

    valid = len(errors) == 0
    if valid:
        message = (
            f"Compatible UPR Country Reporting workbook"
            f"{f' for {country_name}' if country_name else ''}"
            f"{f' ({period_name})' if period_name else ''}."
        )
    else:
        message = errors[0]
    return {
        "valid": valid,
        "message": message,
        "errors": errors,
        "warnings": warnings,
        "preview": preview,
    }


def validate_upr_country_reporting_workbook(
    wb,
    *,
    expected_country: str,
    expected_period: str,
    warnings: Optional[List[str]] = None,
) -> UprCountryReportingWorkbookContext:
    ctx = UprCountryReportingWorkbookContext()
    ctx.warnings = warnings if warnings is not None else []
    ctx.round_code, ctx.period_name = parse_version(wb)
    ctx.country_name = str(read_named_cell(wb, "Data_Country") or "").strip()

    if expected_period and ctx.period_name and ctx.period_name != expected_period:
        pass  # Period mismatch is reported once via validate_upr_country_reporting_import_file.
    if expected_country and ctx.country_name:
        if _normalize_text(expected_country) != _normalize_text(ctx.country_name):
            ctx.warnings.append(
                f"Workbook country {ctx.country_name!r} does not match assignment country {expected_country!r}"
            )
    return ctx


def transform_upr_country_reporting_to_import_rows(
    aes_id: int,
    wb,
    ctx: UprImportContext,
    *,
    iso3: str,
    period: str,
) -> List[Dict[str, str]]:
    """Transform UPR Country Reporting workbook content into form_data import rows for one assignment."""
    import_rows: List[Dict[str, str]] = []
    matrix_cells: Dict[Tuple[int, int], Dict[str, Any]] = {}
    dynamic_order: Dict[int, float] = {}
    items_by_bank = ctx.items_by_bank_id.get(REPORTING_COUNTRY_TEMPLATE_ID, {})
    funding_item = reporting_special_item(ctx, "funding") or ITEM_REPORTING_COUNTRY_FUNDING
    sp_breakdown_item = reporting_special_item(ctx, "sp_breakdown") or ITEM_REPORTING_COUNTRY_SP_BREAKDOWN
    support_item = reporting_special_item(ctx, "support") or ITEM_REPORTING_COUNTRY_SUPPORT
    kpi_lookup = build_kpi_lookup(wb)
    yes_no_bank_ids = _load_workbook_yes_no_bank_ids(wb, kpi_lookup)
    section_label_map = _load_items_by_section_label(REPORTING_COUNTRY_TEMPLATE_ID)

    # NS Data scalars
    ns_data = parse_ns_key_data(wb)
    bank_key_map = {
        "volunteers": NS_DATA_BANK_IDS["volunteers"],
        "staff": NS_DATA_BANK_IDS["staff"],
        "local_units": NS_DATA_BANK_IDS["local_units"],
        "branches": NS_DATA_BANK_IDS["branches"],
    }
    for key, bank_id in bank_key_map.items():
        value = ns_data.get(key)
        item_id = items_by_bank.get(bank_id)
        if item_id and value is not None:
            built = _scalar_row(
                aes_id=aes_id,
                item_id=item_id,
                value=value,
                iso3=iso3,
                period=period,
                debug_kpi=f"ns_{key}",
            )
            if built:
                import_rows.append(built)

    # Indicators (Overall Action tables only — emergency tables use dynamic import).
    for row in parse_indicators(wb, yes_no_bank_ids=yes_no_bank_ids, kpi_lookup=kpi_lookup):
        bank_id = _resolve_workbook_indicator_bank_id(row, kpi_lookup)
        if bank_id:
            item_id = _resolve_item_for_workbook_indicator(ctx, bank_id, row["sp_ef"])
        else:
            item_id = _resolve_item_by_section_and_indicator(
                ctx,
                section_label_map,
                row["sp_ef"],
                row["indicator"],
                kpi_lookup,
            )
        value, is_dna, disagg, should_import = _resolve_indicator_import_value(
            row, bank_id, yes_no_bank_ids
        )
        if not should_import:
            continue
        if not item_id:
            if bank_id and (is_dna or value is not None or disagg):
                _queue_other_dynamic_indicator(
                    ctx,
                    aes_id=aes_id,
                    indicator_bank_id=bank_id,
                    value=value,
                    data_not_available=is_dna,
                    order_counters=dynamic_order,
                    disagg_data=disagg,
                )
            elif not bank_id:
                ctx.warnings.append(
                    f"No T33 form item for indicator {row['indicator']!r} in {row['sp_ef']!r} ({iso3})"
                )
            continue
        if is_dna:
            import_rows.append(
                _data_na_row(
                    aes_id=aes_id,
                    item_id=item_id,
                    iso3=iso3,
                    period=period,
                    debug_kpi=f"indicator_{row['indicator'][:40]}",
                )
            )
        elif disagg:
            import_rows.append(
                _indicator_disagg_row(
                    aes_id=aes_id,
                    item_id=item_id,
                    disagg_payload=disagg,
                    iso3=iso3,
                    period=period,
                    debug_kpi=row["indicator"],
                )
            )
        elif value is not None:
            built = _scalar_row(
                aes_id=aes_id,
                item_id=item_id,
                value=value,
                iso3=iso3,
                period=period,
                debug_kpi=row["indicator"],
            )
            if built:
                import_rows.append(built)

    # Funding (IFRC / PNS / HNS rows → item 1403 matrix column from published FormItem config)
    funding = parse_funding(wb)
    funding_matrix_col = reporting_funding_matrix_column(ctx)
    for row_name, amount in funding.get("sources", {}).items():
        if amount:
            cell_key = f"{row_name}_{funding_matrix_col}"
            matrix_cells.setdefault((aes_id, funding_item), {})[cell_key] = amount

    total_exp = funding.get("total_expenditure")
    if total_exp is not None:
        exp_item = items_by_bank.get(734) or reporting_special_item(ctx, "expenditure")
        item_id = exp_item or ITEM_REPORTING_COUNTRY_EXPENDITURE
        built = _scalar_row(
            aes_id=aes_id,
            item_id=item_id,
            value=total_exp,
            iso3=iso3,
            period=period,
            debug_kpi="Reporting_Expenditure",
        )
        if built:
            import_rows.append(built)

    for row_name, cols in funding.get("breakdown", {}).items():
        for col_name, amount in cols.items():
            if amount is not None:
                cell_key = f"{row_name}_{col_name}"
                matrix_cells.setdefault((aes_id, sp_breakdown_item), {})[cell_key] = amount

    # Bilateral support
    for row in parse_bilateral_support(wb):
        ns_id = _resolve_ns_row_id(ctx, row["ns_name"])
        if ns_id is None:
            continue
        for area in row["areas"]:
            cell_key = f"{ns_id}_{area} Supported"
            matrix_cells.setdefault((aes_id, support_item), {})[cell_key] = 1

    # Comments (T33) → discussion panel (historical import, no author)
    comments = parse_comments(wb)
    if comments:
        ctx.discussion_comment_entries.append(
            {
                "aes_id": aes_id,
                "body": comments,
                "source": "upr_excel_import",
            }
        )

    for (aid, item_id), cells in matrix_cells.items():
        if cells:
            import_rows.append(
                _matrix_row(
                    aes_id=aid,
                    item_id=item_id,
                    cells=cells,
                    iso3=iso3,
                    period=period,
                    debug_kpi=f"matrix_{item_id}",
                )
            )
    return import_rows


def run_upr_country_reporting_import(
    aes_id: int,
    input_path: str,
    *,
    dry_run: bool = False,
    persist: bool = True,
    batch_size: int = 500,
) -> Dict[str, Any]:
    """Import a UPR Country Reporting workbook into a single T33 assignment.

    When ``persist`` is False, parse the workbook and return a client payload for
    staging in the browser without writing to the database.
    """
    from contextlib import nullcontext

    from app.extensions import db
    from app.models.form_items import FormItem

    _require_openpyxl()
    import openpyxl

    try:
        from flask import current_app as _cur

        _cur._get_current_object()
        _ctx = nullcontext()
    except RuntimeError:
        from app import create_app

        _ctx = create_app().app_context()

    stats: Dict[str, Any] = {
        "success": False,
        "loaded": 0,
        "inserted": 0,
        "updated": 0,
        "errors": 0,
        "warnings": [],
        "transformed": 0,
        "updated_count": 0,
    }

    with _ctx:
        _, country_name, iso3, period = _load_assignment_meta(aes_id)
        with _quiet_openpyxl_io():
            wb = openpyxl.load_workbook(input_path, data_only=True)
        meta = validate_upr_country_reporting_workbook(
            wb,
            expected_country=country_name,
            expected_period=period,
        )
        stats["warnings"] = list(meta.warnings)

        compat = validate_upr_country_reporting_import_file(
            wb,
            expected_country=country_name,
            expected_period=period,
        )
        if not compat.get("valid"):
            stats["success"] = False
            stats["errors"] = 1
            stats["message"] = compat.get("message") or "Workbook is not compatible with UPR Country Reporting import."
            stats["warnings"].extend(compat.get("warnings") or [])
            for err in compat.get("errors") or []:
                if err not in stats["warnings"]:
                    stats["warnings"].append(err)
            wb.close()
            return stats
        stats["warnings"].extend(compat.get("warnings") or [])

        ctx = build_import_context([REPORTING_COUNTRY_TEMPLATE_ID])
        import_rows = transform_upr_country_reporting_to_import_rows(aes_id, wb, ctx, iso3=iso3, period=period)
        stats["warnings"].extend(ctx.warnings)
        stats["warnings"] = dedupe_upr_import_warnings(stats["warnings"])
        stats["transformed"] = len(import_rows)

        if not persist:
            payload = build_upr_country_reporting_client_payload(
                aes_id,
                wb,
                ctx,
                import_rows,
                iso3=iso3,
                period=period,
            )
            field_count = len(payload.get("fields") or {})
            matrix_count = len(payload.get("matrices") or {})
            dynamic_count = len(payload.get("dynamic_indicators") or [])
            repeat_count = len(payload.get("repeat_slots") or [])
            updated_count = field_count + matrix_count + dynamic_count + repeat_count
            wb.close()
            return {
                "success": True,
                "stage_only": True,
                "payload": payload,
                "warnings": stats["warnings"],
                "transformed": len(import_rows),
                "updated_count": updated_count,
            }

        valid_item_ids: Set[int] = set(
            fid for (fid,) in db.session.query(FormItem.id).filter(FormItem.template_id == REPORTING_COUNTRY_TEMPLATE_ID).all()
        )
        upsert_stats = upsert_form_data_rows(
            import_rows,
            dry_run=dry_run,
            batch_size=batch_size,
            valid_form_item_ids=valid_item_ids,
            stats=stats,
        )
        discussion_stats = upsert_upr_discussion_comments(
            ctx.discussion_comment_entries,
            dry_run=dry_run,
        )
        upsert_stats.update(discussion_stats)
        if not dry_run:
            dyn_stats = _import_dynamic_indicators_from_workbook(
                aes_id,
                wb,
                ctx,
                section_label_map=_load_items_by_section_label(REPORTING_COUNTRY_TEMPLATE_ID),
                kpi_lookup=build_kpi_lookup(wb),
                dry_run=False,
            )
            displaced_stats = upsert_dynamic_indicator_entries(
                ctx.dynamic_indicator_entries,
                dry_run=False,
            )
            upsert_stats["dynamic_inserted"] = dyn_stats.get("inserted", 0) + displaced_stats.get(
                "dynamic_inserted", 0
            )
            upsert_stats["dynamic_updated"] = dyn_stats.get("updated", 0) + displaced_stats.get(
                "dynamic_updated", 0
            )
        upsert_stats["success"] = upsert_stats.get("errors", 0) == 0
        upsert_stats["updated_count"] = upsert_stats.get("inserted", 0) + upsert_stats.get("updated", 0)
        upsert_stats["warnings"] = stats["warnings"]
        upsert_stats["transformed"] = len(import_rows)
        wb.close()
        return upsert_stats


def build_upr_country_reporting_export(aes_id: int, template_path: str, output_path: str) -> Dict[str, Any]:
    """Fill UPR Country Reporting template with assignment data and save to output_path."""
    _require_openpyxl()
    import openpyxl

    aes, country_name, iso3, period = _load_assignment_meta(aes_id)
    if not os.path.isfile(template_path):
        raise FileNotFoundError(f"UPR Country Reporting template not found: {template_path}")

    with _quiet_openpyxl_io():
        wb = openpyxl.load_workbook(template_path)
    write_named_cell(wb, "Version", period_to_workbook_version(period))
    write_named_cell(wb, "Data_Country", country_name)
    _apply_reporting_assignment_label(wb, _assignment_display_label(aes))

    entries = _load_form_data_map(aes_id)
    ctx = build_import_context([REPORTING_COUNTRY_TEMPLATE_ID])
    kpi_lookup = build_kpi_lookup(wb)
    kpi_display = _build_kpi_display_map(wb)
    indicator_row_index = _build_indicator_row_index(wb)
    bank_id_locations = _build_bank_id_row_locations(wb)
    disagg_key_to_header, _disagg_norm_to_header = _build_upr_country_reporting_disagg_header_maps(wb)
    section_ids = _load_upr_country_reporting_section_ids()
    dynamic_map = _load_dynamic_indicator_map(aes_id)
    comments_item_id = _resolve_comments_item_id(ctx)
    emergency_overflow_used: Dict[Tuple[str, str], Set[int]] = {}

    # NS Data
    for key, bank_id in NS_DATA_BANK_IDS.items():
        item_id = ctx.items_by_bank_id.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get(bank_id)
        if not item_id:
            continue
        value = _scalar_value(entries.get(item_id))
        if value is not None:
            write_named_cell(wb, NS_DATA_NAMED_CELLS[key], value)

    # Standard Overall Action indicators (form_data scalars; overflow -> Data_other)
    from app.models.form_items import FormItem
    from app.models.forms import FormTemplateVersion

    skip_item_ids = {
        ITEM_REPORTING_COUNTRY_FUNDING,
        ITEM_REPORTING_COUNTRY_EXPENDITURE,
        ITEM_REPORTING_COUNTRY_SP_BREAKDOWN,
        ITEM_REPORTING_COUNTRY_SUPPORT,
    }
    # Key Data scalars are written to named cells above — skip only those four items.
    for bank_id in NS_DATA_BANK_IDS.values():
        ns_item_id = ctx.items_by_bank_id.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get(bank_id)
        if ns_item_id:
            skip_item_ids.add(ns_item_id)
    ea_dynamic = section_ids.get("ea_dynamic")
    other_dynamic = section_ids.get("other_dynamic")
    data_other_used: Set[int] = set()
    core_indicator_row_index = _build_core_indicator_row_index(wb)

    indicator_items = (
        FormItem.query.join(FormTemplateVersion, FormItem.version_id == FormTemplateVersion.id)
        .filter(
            FormItem.template_id == REPORTING_COUNTRY_TEMPLATE_ID,
            FormItem.archived == False,
            FormItem.item_type == "indicator",
            FormTemplateVersion.status == "published",
        )
        .all()
    )
    for item in indicator_items:
        if item.id in skip_item_ids:
            continue
        if item.section_id in (ea_dynamic, other_dynamic):
            continue
        entry = entries.get(item.id)
        if not entry:
            continue
        section_name = (item.form_section.name if item.form_section else "") or ""
        label = (item.label or "").strip()
        if not label:
            continue
        loc = _find_row_for_form_item(
            section_name=section_name,
            label=label,
            bank_id=item.indicator_bank_id,
            indicator_row_index=core_indicator_row_index,
            kpi_lookup=kpi_lookup,
            bank_id_locations=bank_id_locations,
        )
        is_dna = bool(entry.is_data_not_available)
        if loc:
            sheet_name, table_name, idx = loc
            _write_indicator_entry(
                wb,
                sheet_name,
                table_name,
                idx,
                entry,
                disagg_key_to_header,
                data_not_available=is_dna,
                yes_no=_entry_is_yes_no(entry, form_item=item),
            )
            continue
        if is_dna or (_entry_disagg_payload(entry) is None and _scalar_value(entry) is None):
            continue
        offset = _allocate_data_other_row(
            wb,
            section_name=section_name,
            label=label,
            used_offsets=data_other_used,
        )
        if offset is None:
            continue
        _write_full_indicator_row(
            wb,
            DATA_OTHER_SHEET,
            DATA_OTHER_TABLE,
            offset,
            section_name=section_name,
            indicator_label=label,
            entry=entry,
            key_to_header=disagg_key_to_header,
            data_not_available=is_dna,
            bank_id=item.indicator_bank_id,
            yes_no=_entry_is_yes_no(entry, form_item=item),
        )

    # Emergency appeal slots (Start sheet MDR/name + per-slot indicator tables)
    emergency_slots = _resolve_emergency_slots_for_export(aes)
    _export_emergency_slots(wb, aes, slots=emergency_slots)
    ea_dynamic = section_ids.get("ea_dynamic")
    if ea_dynamic:
        for _sheet, _table, slot_num, _mdr, _eo in EMERGENCY_SLOTS:
            if slot_num - 1 >= len(emergency_slots) or not emergency_slots[slot_num - 1]:
                continue
            dyn_rows = dynamic_map.get((ea_dynamic, slot_num), [])
            if dyn_rows:
                _export_dynamic_indicators(
                    wb,
                    dynamic_rows=dyn_rows,
                    key_to_header=disagg_key_to_header,
                    bank_id_locations=bank_id_locations,
                    sheet_name=_sheet,
                    table_name=_table,
                    indicator_row_index=indicator_row_index,
                    kpi_lookup=kpi_lookup,
                    kpi_display=kpi_display,
                    overflow_used=emergency_overflow_used,
                )

    # Other dynamic indicators -> Data_other rows when present in template
    other_dynamic = section_ids.get("other_dynamic")
    if other_dynamic:
        other_rows = dynamic_map.get((other_dynamic, None), [])
        if other_rows:
            _export_dynamic_indicators(
                wb,
                dynamic_rows=other_rows,
                key_to_header=disagg_key_to_header,
                bank_id_locations=bank_id_locations,
                sheet_name=DATA_OTHER_SHEET,
                table_name=DATA_OTHER_TABLE,
                indicator_row_index=indicator_row_index,
                kpi_lookup=kpi_lookup,
                kpi_display=kpi_display,
                overflow_used=emergency_overflow_used,
            )

    # Funding
    funding_col = _funding_column_name(wb).strip()
    exp_header = "NS 2026 Total\nExpenditure"
    funding_entry = entries.get(ITEM_REPORTING_COUNTRY_FUNDING)
    funding_cells = _matrix_cells(funding_entry)

    def _funding_amount_for_row(row_name: str) -> Optional[float]:
        for key, val in funding_cells.items():
            if key.startswith(f"{row_name}_") and val not in (None, ""):
                try:
                    return float(val)
                except (TypeError, ValueError):
                    continue
        return None

    _, funding_rows = read_named_table(wb, "Funding", "Data_funding1")
    funding_header = _funding_column_name(wb)
    total_funding = 0.0
    has_funding_total = False
    for idx, row in enumerate(funding_rows):
        attr = str(row.get("Attribute") or "").strip().upper()
        row_name = FUNDING_SOURCE_ATTR_TO_ROW.get(attr)
        if row_name:
            amount = _funding_amount_for_row(row_name)
            if amount is not None:
                write_table_cell(wb, "Funding", "Data_funding1", idx, funding_header, amount)
                total_funding += amount
                has_funding_total = True

    if has_funding_total:
        for idx, row in enumerate(funding_rows):
            if str(row.get("Attribute") or "").strip().upper() == "TOTAL":
                write_table_cell(wb, "Funding", "Data_funding1", idx, funding_header, total_funding)
                break

    exp_entry = entries.get(ITEM_REPORTING_COUNTRY_EXPENDITURE)
    exp_val = _scalar_value(exp_entry)
    if exp_val is not None:
        _, funding_rows = read_named_table(wb, "Funding", "Data_funding1")
        for idx, row in enumerate(funding_rows):
            if str(row.get("Attribute") or "").strip().upper() == "TOTAL":
                write_table_cell(wb, "Funding", "Data_funding1", idx, exp_header, exp_val)
                break

    breakdown_entry = entries.get(ITEM_REPORTING_COUNTRY_SP_BREAKDOWN)
    _export_funding_breakdown(wb, _matrix_cells(breakdown_entry))

    # Bilateral support
    support_entry = entries.get(ITEM_REPORTING_COUNTRY_SUPPORT)
    support_cells = _matrix_cells(support_entry)
    _export_bilateral_support(wb, ctx, support_cells, aes)

    # Comments — single cell aligned with the form comments field
    comments_entry = entries.get(comments_item_id) if comments_item_id else None
    comment_text = str(_scalar_value(comments_entry) or "").strip()
    if comment_text:
        write_named_cell(wb, COMMENTS_NAMED_CELL, comment_text)

    with _quiet_openpyxl_io():
        wb.save(output_path)
    wb.close()
    safe_country = re.sub(r"[^\w\-]+", "_", country_name) or iso3
    filename = f"{safe_country}_UPR_Country_Reporting.xlsx"
    return {"success": True, "filename": filename, "aes_id": aes_id, "iso3": iso3, "period": period}


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    parser = argparse.ArgumentParser(description="UPR Country Reporting Excel import/export for T33 assignments")
    sub = parser.add_subparsers(dest="command", required=True)

    export_p = sub.add_parser("export", help="Export assignment data into UPR Country Reporting template")
    export_p.add_argument("--aes-id", type=int, required=True)
    export_p.add_argument("--template", required=True, help="Path to UPR Country Reporting template .xlsx")
    export_p.add_argument("--output", required=True, help="Output .xlsx path")

    import_p = sub.add_parser("import", help="Import UPR Country Reporting template into assignment")
    import_p.add_argument("--aes-id", type=int, required=True)
    import_p.add_argument("--input", required=True, help="Filled UPR Country Reporting .xlsx path")
    import_p.add_argument("--dry-run", action="store_true")

    args = parser.parse_args()

    from app import create_app

    app = create_app()
    with app.app_context():
        if args.command == "export":
            result = build_upr_country_reporting_export(args.aes_id, args.template, args.output)
            print(json.dumps(result, indent=2))
            return 0 if result.get("success") else 1
        result = run_upr_country_reporting_import(args.aes_id, args.input, dry_run=args.dry_run)
        print(json.dumps({k: result.get(k) for k in ("success", "transformed", "inserted", "updated", "errors", "warnings")}, indent=2))
        return 0 if result.get("success") else 1


if __name__ == "__main__":
    raise SystemExit(main())
