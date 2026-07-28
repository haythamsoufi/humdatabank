# ========== Template Excel Import/Export Service ==========
from app.utils.datetime_helpers import utcnow
"""
Service for exporting and importing form templates to/from Excel.

This service handles:
- Exporting template structure (pages, sections, items) to Excel with exact DB column mapping
- Importing template structure from Excel with validation and ID mapping
- Preserving all configurations, translations, and skip logic rules
"""

from flask import current_app
from flask_login import current_user
from app import db
from app.models import (
    FormTemplate, FormPage, FormSection, FormItem, FormTemplateVersion,
    FormData,
    RepeatGroupInstance, RepeatGroupData, DynamicIndicatorData, DynamicSectionContext,
)
from app.models.documents import SubmittedDocument
from app.models.indicator_bank import IndicatorBank
from contextlib import suppress
from app.services.monitoring.memory import memory_tracker
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import io
import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from app.utils.stable_key import generate_stable_key, is_valid_stable_key, normalize_stable_key


class TemplateExcelService:
    """Service for template Excel export/import operations."""

    # IFRC Color Scheme (matching ExcelService)
    IFRC_COLORS = {
        'RED': "FFED1B2E",
        'DARK_RED': "FFAF0E1B",
        'LIGHT_GRAY': "FFF5F5F5",
        'MEDIUM_GRAY': "FFE0E0E0",
        'DARK_GRAY': "FF666666",
        'WHITE': "FFFFFFFF",
        'YELLOW': "FFFFF9E6",
        'BLUE': "FF0066CC",
        'DARK_BLUE': "FF004499",
    }

    # Required vs Optional columns for each sheet
    REQUIRED_COLUMNS = {
        'Template': ['name'],
        'Pages': ['id', 'name', 'order'],
        'Sections': ['id', 'name', 'order'],
        'Items': ['id', 'section_id', 'item_type', 'label', 'order']
    }

    # Import-only columns omitted from export (e.g. stable_key). May be absent or present on import.
    OPTIONAL_HEADER_COLUMNS = {
        'Sections': {'stable_key'},
        'Items': {'stable_key'},
    }

    # Excel export version (V2: per-language translation columns)
    EXCEL_EXPORT_VERSION = 'V2'

    FALLBACK_TRANSLATABLE_LANGUAGES = ['fr', 'es', 'ar', 'ru', 'zh']

    TRANSLATION_DB_FIELDS = {
        'name': 'name_translations',
        'label': 'label_translations',
        'definition': 'definition_translations',
        'description': 'description_translations',
    }

    TEMPLATE_SHEET_ROW_HEADERS = ['field', 'value']

    TEMPLATE_TRANSLATABLE_FIELDS = ['name']
    PAGE_TRANSLATABLE_FIELDS = ['name']
    SECTION_TRANSLATABLE_FIELDS = ['name']
    ITEM_TRANSLATABLE_FIELDS = ['label', 'definition', 'description']

    # Base column definitions (English/base fields only; translation JSON columns removed).
    # Order: identifiers → display text → type-specific fields → rules → large JSON blobs.
    TEMPLATE_BASE_COLUMNS = [
        'name',
        'description',
        'is_paginated',
        'add_to_self_report',
        'display_order_visible',
        'enable_export_pdf',
        'enable_export_excel',
        'enable_import_excel',
        'enable_ai_validation',
        'variables',
    ]

    PAGE_BASE_COLUMNS = ['id', 'order', 'name']

    SECTION_BASE_COLUMNS = [
        'id',
        'order',
        'name',
        'page_id',
        'parent_section_id',
        'section_type',
        'archived',
        'max_dynamic_indicators',
        'allowed_sectors',
        'indicator_filters',
        'allow_data_not_available',
        'allow_not_applicable',
        'allowed_disaggregation_options',
        'data_entry_display_filters',
        'add_indicator_note',
        'relevance_condition',
    ]

    ITEM_BASE_COLUMNS = [
        'id',
        'section_id',
        'item_type',
        'order',
        'archived',
        'label',
        'definition',
        'description',
        'options_json',
        'options_translations',
        'lookup_list_id',
        'list_display_column',
        'list_filters_json',
        'indicator_bank_id',
        'type',
        'unit',
        'relevance_condition',
        'validation_condition',
        'validation_message',
        'config',
    ]

    # Legacy V1 export columns (JSON translation blobs) — accepted on import for backward compatibility
    TEMPLATE_LEGACY_COLUMNS = [
        'name', 'description',
        'add_to_self_report', 'display_order_visible',
        'is_paginated', 'enable_export_pdf', 'enable_export_excel',
        'enable_import_excel', 'enable_ai_validation', 'name_translations', 'variables',
    ]

    PAGE_LEGACY_COLUMNS = ['id', 'name', 'order', 'name_translations']

    SECTION_LEGACY_COLUMNS = [
        'id', 'name', 'order', 'parent_section_id', 'page_id',
        'section_type', 'max_dynamic_indicators', 'allowed_sectors',
        'indicator_filters', 'allow_data_not_available', 'allow_not_applicable',
        'allowed_disaggregation_options', 'data_entry_display_filters',
        'add_indicator_note', 'name_translations', 'relevance_condition', 'archived',
    ]

    ITEM_LEGACY_COLUMNS = [
        'id', 'section_id', 'item_type', 'label', 'order',
        'relevance_condition', 'archived', 'config', 'indicator_bank_id',
        'type', 'unit', 'validation_condition', 'validation_message',
        'definition', 'options_json', 'lookup_list_id', 'list_display_column',
        'list_filters_json', 'label_translations', 'definition_translations',
        'options_translations', 'description_translations', 'description',
    ]

    JSON_EXPORT_COLUMNS = frozenset({
        'variables',
        'allowed_sectors',
        'indicator_filters',
        'allowed_disaggregation_options',
        'data_entry_display_filters',
        'relevance_condition',
        'validation_condition',
        'config',
        'options_json',
        'list_filters_json',
        'options_translations',
        # Legacy JSON translation columns (import-only)
        'name_translations',
        'label_translations',
        'definition_translations',
        'description_translations',
    })

    @classmethod
    def _get_translatable_languages(cls) -> List[str]:
        try:
            langs = current_app.config.get('TRANSLATABLE_LANGUAGES')
            if langs:
                return list(langs)
        except RuntimeError:
            pass
        return list(cls.FALLBACK_TRANSLATABLE_LANGUAGES)

    @classmethod
    def _expand_columns_with_translations(
        cls, base_columns: List[str], translatable_fields: List[str]
    ) -> List[str]:
        """Insert {field}_{lang} columns immediately after each translatable base field."""
        translatable = set(translatable_fields)
        expanded: List[str] = []
        for col in base_columns:
            expanded.append(col)
            if col in translatable:
                for lang in cls._get_translatable_languages():
                    expanded.append(f"{col}_{lang}")
        return expanded

    @classmethod
    def get_template_columns(cls) -> List[str]:
        """Ordered Template sheet field names (one row per field in export)."""
        return cls._expand_columns_with_translations(
            cls.TEMPLATE_BASE_COLUMNS, cls.TEMPLATE_TRANSLATABLE_FIELDS
        )

    @classmethod
    def _template_sheet_uses_row_layout(cls, headers: List[Any]) -> bool:
        if len(headers) < 2:
            return False
        first = str(headers[0]).strip().lower() if headers[0] is not None else ''
        second = str(headers[1]).strip().lower() if headers[1] is not None else ''
        return first == 'field' and second == 'value'

    @classmethod
    def _build_template_field_values(
        cls, template: FormTemplate, version: FormTemplateVersion
    ) -> Dict[str, Any]:
        version_name = version.name if version.name else template.name
        version_name_translations = (
            version.name_translations if version.name_translations else template.name_translations
        )
        version_description = version.description
        if version_description is None:
            version_description = getattr(template, 'description', None)

        row_values = cls._build_export_row(
            cls.TEMPLATE_BASE_COLUMNS,
            cls.TEMPLATE_TRANSLATABLE_FIELDS,
            {
                'name': version_name,
                'description': version_description,
                'add_to_self_report': (
                    version.add_to_self_report
                    if version.add_to_self_report is not None
                    else getattr(template, 'add_to_self_report', False)
                ),
                'display_order_visible': (
                    version.display_order_visible
                    if version.display_order_visible is not None
                    else getattr(template, 'display_order_visible', False)
                ),
                'is_paginated': (
                    version.is_paginated
                    if version.is_paginated is not None
                    else getattr(template, 'is_paginated', False)
                ),
                'enable_export_pdf': (
                    version.enable_export_pdf
                    if version.enable_export_pdf is not None
                    else getattr(template, 'enable_export_pdf', False)
                ),
                'enable_export_excel': (
                    version.enable_export_excel
                    if version.enable_export_excel is not None
                    else getattr(template, 'enable_export_excel', False)
                ),
                'enable_import_excel': (
                    version.enable_import_excel
                    if version.enable_import_excel is not None
                    else getattr(template, 'enable_import_excel', False)
                ),
                'enable_ai_validation': (
                    version.enable_ai_validation
                    if getattr(version, 'enable_ai_validation', None) is not None
                    else getattr(template, 'enable_ai_validation', False)
                ),
                'variables': cls._format_json_for_excel(
                    version.variables if version.variables else None
                ),
            },
            {'name': version_name_translations},
        )
        return dict(zip(cls.get_template_columns(), row_values))

    @classmethod
    def _parse_template_sheet(
        cls, sheet,
    ) -> Tuple[Dict[str, Any], bool, List[str]]:
        """Parse Template sheet. Returns (field_values, legacy_column_layout, errors)."""
        headers = [cell.value for cell in sheet[1]]
        errors: List[str] = []

        if cls._template_sheet_uses_row_layout(headers):
            row_data: Dict[str, Any] = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None for cell in row):
                    continue
                field_name = row[0]
                if field_name is None or str(field_name).strip() == '':
                    continue
                value = row[1] if len(row) > 1 else None
                row_data[str(field_name).strip()] = value
            if 'name' not in row_data or str(row_data.get('name') or '').strip() == '':
                errors.append('Template sheet is missing required field: name')
            return row_data, False, errors

        # Legacy column layout (headers in row 1, values in row 2)
        header_list = [h for h in headers if h is not None]
        required_headers = cls.REQUIRED_COLUMNS.get('Template', ['name'])
        header_set = {h for h in header_list if isinstance(h, str) and h}
        missing = set(required_headers) - header_set
        if missing:
            errors.append(
                f"Template sheet headers missing required columns. "
                f"Required: {required_headers}, Missing: {list(missing)}, Got: {headers}"
            )
            return {}, True, errors

        row = next(sheet.iter_rows(min_row=2, values_only=True), None)
        if not row or all(cell is None for cell in row):
            return {}, True, []

        row_data = cls._row_dict_from_sheet(headers, row)
        return row_data, True, errors

    @classmethod
    def get_page_columns(cls) -> List[str]:
        return cls._expand_columns_with_translations(
            cls.PAGE_BASE_COLUMNS, cls.PAGE_TRANSLATABLE_FIELDS
        )

    @classmethod
    def get_section_columns(cls) -> List[str]:
        return cls._expand_columns_with_translations(
            cls.SECTION_BASE_COLUMNS, cls.SECTION_TRANSLATABLE_FIELDS
        )

    @classmethod
    def get_item_columns(cls) -> List[str]:
        return cls._expand_columns_with_translations(
            cls.ITEM_BASE_COLUMNS, cls.ITEM_TRANSLATABLE_FIELDS
        )

    @classmethod
    def _format_sheet_header_error(
        cls,
        sheet_name: str,
        *,
        summary: str,
        got: Optional[List[str]] = None,
        missing: Optional[List[str]] = None,
        unrecognized: Optional[List[str]] = None,
    ) -> str:
        parts = [f"{sheet_name} sheet: {summary}"]
        if missing:
            parts.append(f"Missing column(s): {', '.join(missing)}")
        if unrecognized:
            parts.append(f"Unrecognized column(s): {', '.join(unrecognized)}")
        if got:
            parts.append(f"Found ({len(got)}): {', '.join(got)}")
        return ' '.join(parts)

    @classmethod
    def _sheet_uses_legacy_translation_columns(cls, sheet_name: str, header_set: set) -> bool:
        legacy_json_cols = {
            'Template': ['name_translations'],
            'Pages': ['name_translations'],
            'Sections': ['name_translations'],
            'Items': ['label_translations', 'definition_translations', 'description_translations'],
        }.get(sheet_name, [])
        return any(col in header_set for col in legacy_json_cols)

    @classmethod
    def _resolve_sheet_headers(cls, sheet_name: str, headers: List[Any]) -> Tuple[List[str], bool, Optional[str]]:
        """Return (expected_headers, legacy_format, error_message).

        Column order is ignored; only required columns, recognized names, and format matter.
        """
        normalized = [str(h).strip() for h in headers if h is not None and str(h).strip()]
        if not normalized:
            return [], False, cls._format_sheet_header_error(
                sheet_name, summary='no column headers found'
            )

        header_set = set(normalized)
        layouts = {
            'Template': (cls.get_template_columns(), cls.TEMPLATE_LEGACY_COLUMNS),
            'Pages': (cls.get_page_columns(), cls.PAGE_LEGACY_COLUMNS),
            'Sections': (cls.get_section_columns(), cls.SECTION_LEGACY_COLUMNS),
            'Items': (cls.get_item_columns(), cls.ITEM_LEGACY_COLUMNS),
        }
        current, legacy = layouts[sheet_name]
        current_set = set(current)
        legacy_set = set(legacy)
        optional_import = cls.OPTIONAL_HEADER_COLUMNS.get(sheet_name, set())
        required = set(cls.REQUIRED_COLUMNS.get(sheet_name, []))

        missing_required = sorted(required - header_set)
        if missing_required:
            return current, False, cls._format_sheet_header_error(
                sheet_name,
                summary='missing required column(s)',
                missing=missing_required,
                got=normalized,
            )

        known_set = current_set | legacy_set | optional_import
        unrecognized = sorted(header_set - known_set)
        if unrecognized:
            return current, False, cls._format_sheet_header_error(
                sheet_name,
                summary='unrecognized column(s)',
                unrecognized=unrecognized,
                got=normalized,
            )

        legacy_format = cls._sheet_uses_legacy_translation_columns(sheet_name, header_set)
        if legacy_format:
            if header_set <= legacy_set:
                return legacy, True, None
            missing_legacy = sorted(legacy_set - header_set)
            return legacy, True, cls._format_sheet_header_error(
                sheet_name,
                summary='legacy export format but missing expected column(s)',
                missing=missing_legacy,
                got=normalized,
            )

        if header_set <= current_set:
            return current, False, None

        missing_from_current = current_set - header_set
        if missing_from_current <= optional_import:
            return current, False, None

        return current, False, cls._format_sheet_header_error(
            sheet_name,
            summary='missing expected column(s) for the current export format',
            missing=sorted(missing_from_current - optional_import),
            got=normalized,
        )

    @classmethod
    def _translation_export_values(cls, translations: Any) -> List[Any]:
        data = cls._normalize_json_for_export(translations)
        if not isinstance(data, dict):
            data = {}
        values: List[Any] = []
        for lang in cls._get_translatable_languages():
            val = data.get(lang)
            if val is None or str(val).strip() == '':
                values.append(None)
            elif isinstance(val, str):
                values.append(cls._decode_unicode_escapes_in_str(val))
            else:
                values.append(val)
        return values

    @classmethod
    def _build_export_row(
        cls,
        base_columns: List[str],
        translatable_fields: List[str],
        scalar_values: Dict[str, Any],
        translation_values: Optional[Dict[str, Any]] = None,
    ) -> List[Any]:
        translatable = set(translatable_fields)
        translation_values = translation_values or {}
        row: List[Any] = []
        for col in base_columns:
            row.append(scalar_values.get(col))
            if col in translatable:
                row.extend(cls._translation_export_values(translation_values.get(col)))
        return row

    @classmethod
    def _collect_translations_from_row(
        cls,
        row_data: Dict[str, Any],
        base_field: str,
        *,
        legacy: bool = False,
    ) -> Optional[Dict[str, str]]:
        translations: Dict[str, str] = {}
        if not legacy:
            for lang in cls._get_translatable_languages():
                col = f"{base_field}_{lang}"
                val = row_data.get(col)
                if val is not None and str(val).strip() != '':
                    translations[lang] = cls._decode_unicode_escapes_in_str(str(val).strip())
        legacy_col = cls.TRANSLATION_DB_FIELDS.get(base_field)
        if legacy_col:
            parsed = cls._parse_json(row_data.get(legacy_col))
            if isinstance(parsed, dict):
                for lang, val in parsed.items():
                    if val is not None and str(val).strip() and lang not in translations:
                        translations[str(lang)] = cls._decode_unicode_escapes_in_str(str(val).strip())
        return translations or None

    @classmethod
    def _row_dict_from_sheet(cls, headers: List[Any], row: Tuple[Any, ...]) -> Dict[str, Any]:
        row_data: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if header is None:
                continue
            row_data[str(header)] = row[idx] if idx < len(row) else None
        return row_data

    @classmethod
    def _apply_item_translations_from_row(
        cls,
        item: FormItem,
        row_data: Dict[str, Any],
        *,
        legacy: bool = False,
    ) -> None:
        for base_field in cls.ITEM_TRANSLATABLE_FIELDS:
            db_field = cls.TRANSLATION_DB_FIELDS[base_field]
            setattr(
                item,
                db_field,
                cls._collect_translations_from_row(row_data, base_field, legacy=legacy),
            )

    # Dropdown options for data validation (static options)
    DROPDOWN_OPTIONS = {
        'section_type': ['standard', 'repeat', 'dynamic_indicators'],
        # Note: item_type is now dynamic - see _get_item_type_options()
        'archived': ['TRUE', 'FALSE'],
        'add_to_self_report': ['TRUE', 'FALSE'],
        'display_order_visible': ['TRUE', 'FALSE'],
        'is_paginated': ['TRUE', 'FALSE'],
        'enable_export_pdf': ['TRUE', 'FALSE'],
        'enable_export_excel': ['TRUE', 'FALSE'],
        'enable_import_excel': ['TRUE', 'FALSE'],
        'enable_ai_validation': ['TRUE', 'FALSE'],
        'allow_data_not_available': ['TRUE', 'FALSE'],
        'allow_not_applicable': ['TRUE', 'FALSE'],
    }

    _UNICODE_ESCAPE_RE = re.compile(r'\\u([0-9a-fA-F]{4})')

    @classmethod
    def _decode_unicode_escapes_in_str(cls, value: str) -> str:
        """Convert literal \\uXXXX sequences to real Unicode characters."""
        if not isinstance(value, str) or '\\u' not in value:
            return value

        def _repl(match: re.Match[str]) -> str:
            try:
                return chr(int(match.group(1), 16))
            except ValueError:
                return match.group(0)

        return cls._UNICODE_ESCAPE_RE.sub(_repl, value)

    @classmethod
    def _normalize_json_for_export(cls, value: Any) -> Any:
        """Normalize JSON payloads so exports show readable Unicode text."""
        if value is None or value == '' or value == 'None':
            return None
        if isinstance(value, str):
            parsed = cls._parse_json(value)
            if isinstance(parsed, (dict, list)):
                return cls._normalize_json_for_export(parsed)
            if isinstance(parsed, str) and parsed != value:
                return cls._normalize_json_for_export(parsed)
            return cls._decode_unicode_escapes_in_str(value)
        if isinstance(value, dict):
            return {key: cls._normalize_json_for_export(val) for key, val in value.items()}
        if isinstance(value, list):
            return [cls._normalize_json_for_export(item) for item in value]
        return value

    @classmethod
    def _format_json_for_excel(cls, value: Any) -> Optional[str]:
        """Serialize JSON values with line breaks for readable Excel cells."""
        normalized = cls._normalize_json_for_export(value)
        if normalized is None:
            return None
        if isinstance(normalized, str):
            return normalized
        try:
            return json.dumps(normalized, ensure_ascii=False, indent=2)
        except (TypeError, ValueError):
            try:
                return json.dumps(normalized, ensure_ascii=False, default=str, indent=2)
            except (TypeError, ValueError):
                return str(normalized)

    _MATRIX_CONFIG_KEYS = frozenset({
        'type', 'columns', 'column_groups', 'rows', 'row_mode', 'show_row_totals',
        'show_column_totals', 'auto_load_entities', 'highlight_manual_rows',
        'legend_text', 'legend_text_translations', 'legend_hide', 'lookup_list_id',
        'list_display_column', 'list_filters', 'group_by_column', 'group_dropdown_enabled',
        'group_table_enabled', 'search_placeholder', 'search_placeholder_translations',
        'plugin_config',
    })

    _MATRIX_IMPORT_LOG_PREFIX = '[excel-import:matrix]'
    _matrix_import_logger: Optional[logging.Logger] = None

    @classmethod
    def _get_matrix_import_logger(cls) -> logging.Logger:
        """Dedicated logger: always writes to stdout and instance/logs/excel_matrix_import.log."""
        if cls._matrix_import_logger is not None:
            return cls._matrix_import_logger

        logger = logging.getLogger('excel_matrix_import')
        logger.setLevel(logging.DEBUG)
        logger.propagate = False

        if not logger.handlers:
            formatter = logging.Formatter(
                '[%(asctime)s] %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S',
            )
            stdout_handler = logging.StreamHandler(sys.stdout)
            stdout_handler.setFormatter(formatter)
            logger.addHandler(stdout_handler)

            try:
                log_dir = Path(__file__).resolve().parents[2] / 'instance' / 'logs'
                log_dir.mkdir(parents=True, exist_ok=True)
                file_handler = logging.FileHandler(
                    log_dir / 'excel_matrix_import.log',
                    encoding='utf-8',
                )
                file_handler.setFormatter(formatter)
                logger.addHandler(file_handler)
            except OSError as exc:
                logger.warning(
                    '%s Could not open excel_matrix_import.log: %s',
                    cls._MATRIX_IMPORT_LOG_PREFIX,
                    exc,
                )

        cls._matrix_import_logger = logger
        return logger

    @classmethod
    def matrix_import_entry_log(cls, message: str) -> None:
        """High-visibility entry point log (routes, import start)."""
        cls._matrix_import_log(message)

    @classmethod
    def _matrix_import_log(cls, message: str, *, level: str = 'info') -> None:
        """Log matrix Excel import diagnostics to dedicated logger, app logger, and log file."""
        text = f"{cls._MATRIX_IMPORT_LOG_PREFIX} {message}"
        logger = cls._get_matrix_import_logger()
        log_fn = getattr(logger, level, logger.info)
        log_fn(text)
        try:
            app_log_fn = getattr(current_app.logger, level, current_app.logger.info)
            app_log_fn(text)
        except RuntimeError:
            pass

    @classmethod
    def _scan_workbook_matrix_items(cls, workbook, *, stage: str) -> int:
        """Scan Items sheet for matrix rows and log config/group hints (validate or import)."""
        if 'Items' not in workbook.sheetnames:
            cls._matrix_import_log(f"{stage}: Items sheet missing in workbook", level='warning')
            return 0

        sheet = workbook['Items']
        headers = [cell.value for cell in sheet[1]]
        if 'config' not in headers or 'item_type' not in headers:
            cls._matrix_import_log(
                f"{stage}: Items sheet missing config or item_type column (headers={headers})",
                level='warning',
            )
            return 0

        matrix_count = 0
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue
            row_data = cls._row_dict_from_sheet(headers, row)
            item_type = str(row_data.get('item_type') or '').strip().lower()
            if item_type != 'matrix':
                continue
            matrix_count += 1
            parsed = cls._parse_json(row_data.get('config'))
            normalized = cls._normalize_matrix_item_config('matrix', parsed) if parsed else None
            cls._matrix_import_log(
                f"{stage} Items row {row_idx} export_id={row_data.get('id')!r} "
                f"label={row_data.get('label')!r}: "
                f"raw={cls._summarize_raw_config_cell(row_data.get('config'))}; "
                f"normalized={cls._summarize_matrix_config_for_log(normalized if isinstance(normalized, dict) else {})}"
            )

        cls._matrix_import_log(f"{stage}: found {matrix_count} matrix row(s) in workbook")
        return matrix_count

    @classmethod
    def _summarize_matrix_config_for_log(cls, config: Any) -> str:
        if not isinstance(config, dict):
            return f"type={type(config).__name__}, value={config!r}"

        mc = config.get('matrix_config')
        if not isinstance(mc, dict):
            if config.get('type') == 'matrix' or isinstance(config.get('columns'), list):
                mc = config
            else:
                return "no matrix_config key"

        columns = mc.get('columns') if isinstance(mc.get('columns'), list) else []
        column_groups = mc.get('column_groups') if isinstance(mc.get('column_groups'), dict) else {}
        grouped = [
            f"{col.get('name')}→{col.get('group')}"
            for col in columns
            if isinstance(col, dict) and col.get('group')
        ]
        ungrouped = [
            str(col.get('name', col))
            for col in columns
            if isinstance(col, dict) and not col.get('group')
        ]
        grouped_preview = grouped[:10]
        if len(grouped) > 10:
            grouped_preview.append(f"...+{len(grouped) - 10} more")
        return (
            f"columns={len(columns)}, "
            f"column_group_keys={list(column_groups.keys())}, "
            f"grouped={grouped_preview or 'none'}, "
            f"ungrouped={ungrouped[:5]}{'...' if len(ungrouped) > 5 else ''}"
        )

    @classmethod
    def _summarize_raw_config_cell(cls, raw: Any) -> str:
        if raw is None or raw == '' or raw == 'None':
            return 'empty'
        if isinstance(raw, dict):
            return cls._summarize_matrix_config_for_log(raw)
        text = str(raw).strip()
        has_column_groups = 'column_groups' in text
        has_group_key = '"group"' in text or "'group'" in text
        preview = text[:160].replace('\n', '\\n')
        if len(text) > 160:
            preview += '...'
        return (
            f"text_len={len(text)}, has_column_groups={has_column_groups}, "
            f"has_group_key={has_group_key}, preview={preview!r}"
        )

    @classmethod
    def _log_matrix_items_in_version(cls, template_id: int, version_id: int, stage: str) -> None:
        """Log all matrix items currently in the target version."""
        matrix_items = (
            FormItem.query.join(FormSection, FormItem.section_id == FormSection.id)
            .filter(
                FormItem.template_id == template_id,
                FormSection.version_id == version_id,
                FormItem.item_type == 'matrix',
            )
            .order_by(FormItem.order, FormItem.id)
            .all()
        )
        cls._matrix_import_log(
            f"{stage}: found {len(matrix_items)} matrix item(s) in version_id={version_id}"
        )
        for item in matrix_items:
            cls._matrix_import_log(
                f"{stage}: item db_id={item.id} label={item.label!r} "
                f"section_id={item.section_id} order={item.order} "
                f"config={cls._summarize_matrix_config_for_log(item.config)}"
            )

    @classmethod
    def _deep_copy_json(cls, value: Any) -> Any:
        try:
            return json.loads(json.dumps(value))
        except (TypeError, ValueError):
            import copy
            return copy.deepcopy(value)

    @classmethod
    def _clean_imported_matrix_config(cls, config: Dict[str, Any]) -> Dict[str, Any]:
        """Remove legacy matrix keys duplicated at config root; keep matrix_config canonical."""
        out = cls._deep_copy_json(config)
        if not isinstance(out, dict):
            return out
        if isinstance(out.get('matrix_config'), dict):
            for key in cls._MATRIX_CONFIG_KEYS:
                out.pop(key, None)
            out.pop('column_groups', None)
        return out

    @classmethod
    def _find_existing_item_for_import(
        cls,
        *,
        section_id: int,
        item_type: str,
        item_label: str,
        item_order: float,
        export_id: Optional[int],
        export_id_to_existing: Dict[int, FormItem],
        existing_items_lookup: Dict[Tuple, FormItem],
    ) -> Tuple[Optional[FormItem], Optional[str]]:
        """Resolve an existing FormItem for an imported row (matrix: label match wins over export id)."""
        existing_item: Optional[FormItem] = None
        match_method: Optional[str] = None

        if item_label:
            label_key = (section_id, item_type, item_label)
            existing_item = existing_items_lookup.get(label_key)
            if existing_item:
                return existing_item, f'label_key={label_key!r}'

            for key, candidate in existing_items_lookup.items():
                if key[0] == section_id and key[2] == item_type and key[3] == item_label:
                    return candidate, f'label={item_label!r} section_id={section_id}'

        if export_id and export_id in export_id_to_existing:
            candidate = export_id_to_existing[export_id]
            if candidate.section_id == section_id:
                candidate_label = str(candidate.label or '').strip()
                if not item_label or not candidate_label or candidate_label == item_label:
                    return candidate, f'export_id={export_id}'
                cls._matrix_import_log(
                    f"export_id={export_id} label mismatch: excel={item_label!r} "
                    f"db_id={candidate.id} db_label={candidate_label!r}; skipping export_id match",
                    level='warning',
                )

        item_key = (section_id, item_order, item_type, item_label)
        existing_item = existing_items_lookup.get(item_key)
        if existing_item:
            return existing_item, f'order_key={item_key!r}'

        return None, None

    @classmethod
    def _normalize_matrix_item_config(cls, item_type: str, config: Any) -> Any:
        """Ensure matrix items store grouped columns under config.matrix_config."""
        if item_type != 'matrix' or config is None:
            return config
        if isinstance(config, str):
            config = cls._parse_json(config)
        if not isinstance(config, dict):
            cls._matrix_import_log(
                f"normalize skipped: config is {type(config).__name__}, not dict",
                level='warning',
            )
            return config

        cls._matrix_import_log(
            f"normalize input: {cls._summarize_matrix_config_for_log(config)}"
        )

        out = dict(config)
        matrix_config = out.get('matrix_config')
        if isinstance(matrix_config, str):
            matrix_config = cls._parse_json(matrix_config)

        root_column_groups = out.pop('column_groups', None)

        if isinstance(matrix_config, dict):
            mc = dict(matrix_config)
        elif out.get('type') == 'matrix' or isinstance(out.get('columns'), list):
            mc = {k: out.pop(k) for k in list(out.keys()) if k in cls._MATRIX_CONFIG_KEYS}
        else:
            cls._matrix_import_log(
                "normalize: no matrix_config/columns found in config; leaving unchanged",
                level='warning',
            )
            return out

        if root_column_groups is not None and 'column_groups' not in mc:
            cls._matrix_import_log(
                f"normalize: hoisted root column_groups keys={list(root_column_groups.keys()) if isinstance(root_column_groups, dict) else root_column_groups!r}"
            )
            mc['column_groups'] = root_column_groups

        columns = mc.get('columns')
        if isinstance(columns, list):
            normalized_columns = []
            for col in columns:
                if isinstance(col, dict):
                    normalized_columns.append(dict(col))
                elif isinstance(col, str) and col.strip():
                    normalized_columns.append({'name': col.strip(), 'type': 'number'})
            mc['columns'] = normalized_columns

        mc = cls._repair_matrix_column_groups(mc)
        out['matrix_config'] = mc
        cls._matrix_import_log(
            f"normalize output: {cls._summarize_matrix_config_for_log(out)}"
        )
        return out

    @classmethod
    def _repair_matrix_column_groups(cls, matrix_config: Any) -> Dict[str, Any]:
        """Keep column_groups translations and per-column group keys in sync."""
        if not isinstance(matrix_config, dict):
            return matrix_config

        before = cls._summarize_matrix_config_for_log({'matrix_config': matrix_config})

        mc = dict(matrix_config)
        columns = mc.get('columns')
        if not isinstance(columns, list):
            return mc

        columns = [dict(col) if isinstance(col, dict) else col for col in columns]
        raw_groups = mc.get('column_groups')
        column_groups: Dict[str, Any] = raw_groups if isinstance(raw_groups, dict) else {}

        repaired_groups: Dict[str, Any] = {}
        for group_label, group_value in column_groups.items():
            group_key = str(group_label).strip()
            if not group_key:
                continue
            if isinstance(group_value, list):
                for col_name in group_value:
                    target_name = str(col_name).strip()
                    for col in columns:
                        if isinstance(col, dict) and str(col.get('name', '')).strip() == target_name:
                            col['group'] = group_key
                repaired_groups[group_key] = {}
            elif isinstance(group_value, dict):
                repaired_groups[group_key] = group_value
            else:
                repaired_groups[group_key] = {}

        for col in columns:
            if not isinstance(col, dict):
                continue
            if 'group' not in col and col.get('group_label'):
                col['group'] = str(col.pop('group_label')).strip()
            elif 'group' not in col and col.get('groupName'):
                col['group'] = str(col.pop('groupName')).strip()
            group_name = col.get('group')
            if group_name and str(group_name).strip() and str(group_name).strip() not in repaired_groups:
                repaired_groups[str(group_name).strip()] = {}

        # Infer groups from column_groups keys when columns share a name prefix (e.g. "SP1 Planned" -> "SP1")
        for group_key in list(repaired_groups.keys()):
            prefix = f"{group_key} "
            for col in columns:
                if not isinstance(col, dict) or col.get('group'):
                    continue
                name = str(col.get('name', '')).strip()
                if name == group_key or name.startswith(prefix):
                    col['group'] = group_key

        mc['columns'] = columns
        mc['column_groups'] = repaired_groups
        after = cls._summarize_matrix_config_for_log({'matrix_config': mc})
        if before != after:
            cls._matrix_import_log(f"repair changed config: before={before} after={after}")
        else:
            cls._matrix_import_log(f"repair unchanged: {after}")
        return mc

    @classmethod
    def _apply_item_config_from_import(
        cls,
        item: FormItem,
        item_type: str,
        raw_config: Any,
        *,
        keep_existing_on_failure: bool = True,
    ) -> None:
        """Parse, normalize, and persist imported item config without wiping on parse failure."""
        from sqlalchemy.orm.attributes import flag_modified

        item_id = getattr(item, 'id', None)
        item_label = getattr(item, 'label', None)
        log_matrix = item_type == 'matrix'

        if raw_config is None or raw_config == '' or raw_config == 'None':
            if log_matrix:
                cls._matrix_import_log(
                    f"item id={item_id} label={item_label!r}: config cell empty; keeping existing config "
                    f"({cls._summarize_matrix_config_for_log(item.config if isinstance(item.config, dict) else {})})"
                )
            return

        if log_matrix:
            cls._matrix_import_log(
                f"item id={item_id} label={item_label!r}: raw config cell "
                f"{cls._summarize_raw_config_cell(raw_config)}"
            )

        parsed = cls._parse_json(raw_config)
        if parsed is None:
            if keep_existing_on_failure and log_matrix:
                cls._matrix_import_log(
                    f"item id={item_id} label={item_label!r}: JSON parse FAILED; keeping existing config "
                    f"({cls._summarize_matrix_config_for_log(item.config if isinstance(item.config, dict) else {})})",
                    level='warning',
                )
            return

        before_summary = cls._summarize_matrix_config_for_log(
            item.config if isinstance(item.config, dict) else {}
        ) if log_matrix else ''

        normalized = cls._normalize_matrix_item_config(item_type, parsed)
        if normalized is None:
            if log_matrix:
                cls._matrix_import_log(
                    f"item id={item_id} label={item_label!r}: normalize returned None; config not updated",
                    level='warning',
                )
            return

        if item_type == 'matrix' and isinstance(normalized, dict):
            item.config = cls._clean_imported_matrix_config(normalized)
            cls._sync_matrix_item_fields_from_config(item)
        else:
            item.config = cls._deep_copy_json(normalized) if isinstance(normalized, dict) else normalized

        flag_modified(item, 'config')
        if log_matrix:
            cls._matrix_import_log(
                f"item id={item_id} label={item_label!r}: config applied "
                f"before={before_summary} after={cls._summarize_matrix_config_for_log(item.config)}"
            )

    @classmethod
    def _sync_matrix_item_fields_from_config(cls, item: FormItem) -> None:
        """Mirror matrix_config list-library fields onto FormItem columns."""
        if item.item_type != 'matrix' or not isinstance(item.config, dict):
            return
        matrix_config = item.config.get('matrix_config')
        if not isinstance(matrix_config, dict):
            return
        if matrix_config.get('row_mode') == 'list_library':
            if 'lookup_list_id' in matrix_config:
                item.lookup_list_id = matrix_config['lookup_list_id']
            if matrix_config.get('list_display_column'):
                item.list_display_column = matrix_config['list_display_column']
            if matrix_config.get('list_filters') is not None:
                item.list_filters_json = matrix_config['list_filters']

    @classmethod
    def _write_data_row(cls, sheet, row_idx: int, headers: List[str], row_data: List[Any]) -> None:
        """Write a data row."""
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    @classmethod
    def _get_items_for_version(cls, template: FormTemplate, version: FormTemplateVersion) -> List[FormItem]:
        """Get items for a template version in a deterministic order.

        Rows are grouped by section display order, then item order within each section.
        Stable secondary sorts on section/item IDs keep export IDs deterministic.
        """
        return FormItem.query.join(FormSection).filter(
            FormItem.template_id == template.id,
            FormSection.version_id == version.id
        ).order_by(
            FormSection.order,
            FormSection.id,
            FormItem.order,
            FormItem.id,
        ).all()

    @classmethod
    def _build_item_db_to_export_map(cls, template: FormTemplate, version: FormTemplateVersion) -> Dict[int, int]:
        """Build mapping from DB item IDs -> sequential export IDs (1, 2, 3...)."""
        items = cls._get_items_for_version(template, version)
        return {item.id: idx + 1 for idx, item in enumerate(items)}

    @classmethod
    def _rewrite_rule_json_item_ids(cls, rule_json: Any, id_map: Dict[int, int]) -> Any:
        """Rewrite item references inside a relevance/validation rule JSON.

        The rule builder stores references under the key 'item_id'. Values can be:
        - numeric strings (e.g., "66") for regular items
        - prefixed strings (e.g., "plugin_123" or "plugin_123_measure_id")
        - legacy prefixed strings (e.g., "question_66")

        This method is used in two directions depending on the provided id_map:
        - Export: db_id -> export_id
        - Import: export_id -> new_db_id
        """
        if rule_json is None:
            return None
        if isinstance(rule_json, str) and rule_json.strip() == '':
            return rule_json

        # Parse (handle occasional double-encoded JSON)
        parsed = None
        if isinstance(rule_json, (dict, list)):
            parsed = rule_json
        else:
            raw = str(rule_json)
            with suppress(Exception):
                parsed = json.loads(raw)
                if isinstance(parsed, str):
                    parsed2 = json.loads(parsed)
                    parsed = parsed2
        if parsed is None:
            return rule_json

        def _rewrite_item_id_value(val: Any) -> Any:
            if val is None:
                return val
            # Numeric id stored as int/float/string
            if isinstance(val, (int, float)) or (isinstance(val, str) and val.isdigit()):
                old_id = int(val)
                new_id = id_map.get(old_id)
                return str(new_id) if new_id is not None else val

            if isinstance(val, str):
                # plugin_123 or plugin_123_measure_id
                if val.startswith('plugin_'):
                    parts = val.split('_')
                    if len(parts) >= 2 and parts[1].isdigit():
                        old_id = int(parts[1])
                        new_id = id_map.get(old_id)
                        if new_id is not None:
                            parts[1] = str(new_id)
                            return '_'.join(parts)
                    return val

                # Legacy prefixed IDs (question_66, indicator_12, document_field_7, form_item_99)
                m = re.match(r'^(question|indicator|document_field|form_item)_(\d+)$', val)
                if m:
                    old_id = int(m.group(2))
                    new_id = id_map.get(old_id)
                    if new_id is not None:
                        return f"{m.group(1)}_{new_id}"
                    return val

            return val

        def _walk(obj: Any) -> Any:
            if isinstance(obj, list):
                for i in range(len(obj)):
                    obj[i] = _walk(obj[i])
                return obj
            if isinstance(obj, dict):
                for k, v in list(obj.items()):
                    if k == 'item_id':
                        obj[k] = _rewrite_item_id_value(v)
                    else:
                        obj[k] = _walk(v)
                return obj
            return obj

        parsed = _walk(parsed)
        try:
            return cls._format_json_for_excel(parsed)
        except Exception as e:
            current_app.logger.debug("Rule JSON serialization failed, using original: %s", e)
            # Fall back to the original if serialization fails
            return rule_json

    @classmethod
    @memory_tracker("Template Excel Export", log_top_allocations=True)
    def export_template(cls, template_id: int, version_id: Optional[int] = None) -> io.BytesIO:
        """
        Export template structure to Excel.

        Args:
            template_id: Template ID to export
            version_id: Optional version ID (defaults to published or latest)

        Returns:
            BytesIO object containing Excel file
        """
        template = FormTemplate.query.get_or_404(template_id)

        # Determine version to export
        if version_id:
            version = FormTemplateVersion.query.filter_by(
                id=version_id, template_id=template.id
            ).first()
        else:
            if template.published_version_id:
                version = FormTemplateVersion.query.get(template.published_version_id)
            else:
                version = FormTemplateVersion.query.filter_by(
                    template_id=template.id
                ).order_by(FormTemplateVersion.created_at.desc()).first()

        if not version:
            raise ValueError(f"No version found for template {template_id}")

        # Create workbook
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        # Export Instructions sheet first (so it appears first)
        cls._export_instructions_sheet(workbook)

        # Export hidden metadata sheet
        cls._export_metadata_sheet(workbook, template, version)

        # Export Template sheet
        cls._export_template_sheet(workbook, template, version)

        # Export Pages sheet
        cls._export_pages_sheet(workbook, template, version)

        # Export Sections sheet
        cls._export_sections_sheet(workbook, template, version)

        # Export Items sheet
        cls._export_items_sheet(workbook, template, version)

        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    @classmethod
    def _export_template_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export template metadata to Template sheet (field/value rows)."""
        sheet = workbook.create_sheet("Template")

        headers = cls.TEMPLATE_SHEET_ROW_HEADERS
        required_cols = cls.REQUIRED_COLUMNS.get('Template', [])
        field_names = cls.get_template_columns()
        field_values = cls._build_template_field_values(template, version)

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cls._style_header_cell(cell, is_required=(header == 'field'))

        boolean_fields = {
            'add_to_self_report', 'display_order_visible', 'is_paginated',
            'enable_export_pdf', 'enable_export_excel', 'enable_import_excel', 'enable_ai_validation',
        }

        for row_idx, field_name in enumerate(field_names, start=2):
            field_cell = sheet.cell(row=row_idx, column=1, value=field_name)
            cls._style_header_cell(field_cell, is_required=(field_name in required_cols))
            sheet.cell(row=row_idx, column=2, value=field_values.get(field_name))

            if field_name in boolean_fields and field_name in cls.DROPDOWN_OPTIONS:
                cls._add_dropdown_validation(
                    sheet, 2, cls.DROPDOWN_OPTIONS[field_name],
                    start_row=row_idx, end_row=row_idx,
                )

        cls._auto_size_columns(sheet, 2)
        cls._create_excel_table(sheet, "TemplateTable", 2, len(field_names) + 1)

    @classmethod
    def _export_pages_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export pages to Pages sheet with sequential IDs."""
        sheet = workbook.create_sheet("Pages")

        # Get pages for this version (include archived)
        pages = FormPage.query.filter_by(
            template_id=template.id, version_id=version.id
        ).order_by(FormPage.order).all()

        # Write headers with required/optional styling
        headers = cls.get_page_columns()
        required_cols = cls.REQUIRED_COLUMNS.get('Pages', [])
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            is_required = header in required_cols
            cls._style_header_cell(cell, is_required=is_required)

        # Create mapping: sequential export ID -> database ID
        page_export_id_map = {}  # export_id -> db_id

        # Write data rows with sequential IDs
        for row_idx, page in enumerate(pages, start=2):
            export_id = row_idx - 1  # Sequential ID starting from 1
            page_export_id_map[export_id] = page.id  # Store mapping for reference

            row_data = cls._build_export_row(
                cls.PAGE_BASE_COLUMNS,
                cls.PAGE_TRANSLATABLE_FIELDS,
                {
                    'id': export_id,
                    'name': page.name,
                    'order': page.order,
                },
                {'name': page.name_translations},
            )
            cls._write_data_row(sheet, row_idx, headers, row_data)

        # Store mapping in sheet for reference (not visible, but can be used if needed)
        sheet._page_export_id_map = page_export_id_map

        # Auto-size columns
        cls._auto_size_columns(sheet, len(headers))

        # Create Excel table (always create, even with 0 rows)
        num_data_rows = len(pages)
        cls._create_excel_table(sheet, "PagesTable", len(headers), num_data_rows + 1)  # +1 for header

        # Add conditional formatting for duplicate IDs (only if there are data rows)
        if num_data_rows > 0:
            cls._add_duplicate_highlighting(sheet, 'id', headers, num_data_rows + 1)

    @classmethod
    def _export_sections_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export sections to Sections sheet with sequential IDs."""
        sheet = workbook.create_sheet("Sections")

        # Get sections for this version (include archived)
        sections = FormSection.query.filter_by(
            template_id=template.id, version_id=version.id
        ).order_by(FormSection.order).all()

        # Get page export ID mapping from Pages sheet
        pages_sheet = workbook['Pages']
        page_db_to_export = {}  # db_id -> export_id
        if hasattr(pages_sheet, '_page_export_id_map'):
            # Reverse the map: export_id -> db_id becomes db_id -> export_id
            for exp_id, db_id in pages_sheet._page_export_id_map.items():
                page_db_to_export[db_id] = exp_id

        # Build section database ID to sequential export ID mapping
        section_db_to_export = {}  # db_id -> export_id

        # Build item DB -> export mapping for rewriting relevance/validation rules.
        # This must match the Items sheet export IDs (which are sequential as well).
        item_db_to_export = cls._build_item_db_to_export_map(template, version)

        # First pass: build mapping
        for row_idx, section in enumerate(sections, start=2):
            export_id = row_idx - 1  # Sequential ID starting from 1
            section_db_to_export[section.id] = export_id

        # Write headers with required/optional styling
        headers = cls.get_section_columns()
        required_cols = cls.REQUIRED_COLUMNS.get('Sections', [])
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            is_required = header in required_cols
            cls._style_header_cell(cell, is_required=is_required)

        # Second pass: write data rows with sequential IDs and mapped references
        for row_idx, section in enumerate(sections, start=2):
            export_id = section_db_to_export[section.id]

            # Map page_id to export ID
            page_export_id = None
            if section.page_id and section.page_id in page_db_to_export:
                page_export_id = page_db_to_export[section.page_id]

            # Map parent_section_id to export ID
            parent_export_id = None
            if section.parent_section_id and section.parent_section_id in section_db_to_export:
                parent_export_id = section_db_to_export[section.parent_section_id]

            row_data = cls._build_export_row(
                cls.SECTION_BASE_COLUMNS,
                cls.SECTION_TRANSLATABLE_FIELDS,
                {
                    'id': export_id,
                    'name': section.name,
                    'order': section.order,
                    'parent_section_id': parent_export_id,
                    'page_id': page_export_id,
                    'section_type': section.section_type,
                    'max_dynamic_indicators': section.max_dynamic_indicators,
                    'allowed_sectors': cls._format_json_for_excel(section.allowed_sectors),
                    'indicator_filters': cls._format_json_for_excel(section.indicator_filters),
                    'allow_data_not_available': section.allow_data_not_available,
                    'allow_not_applicable': section.allow_not_applicable,
                    'allowed_disaggregation_options': cls._format_json_for_excel(section.allowed_disaggregation_options),
                    'data_entry_display_filters': cls._format_json_for_excel(section.data_entry_display_filters),
                    'add_indicator_note': section.add_indicator_note,
                    'relevance_condition': cls._rewrite_rule_json_item_ids(section.relevance_condition, item_db_to_export),
                    'archived': section.archived,
                },
                {'name': section.name_translations},
            )
            cls._write_data_row(sheet, row_idx, headers, row_data)

        # Store mapping for items sheet (export_id -> db_id)
        sheet._section_export_id_map = {exp_id: db_id for db_id, exp_id in section_db_to_export.items()}

        # Auto-size columns
        cls._auto_size_columns(sheet, len(headers))

        # Add data validation dropdowns
        section_type_col = headers.index('section_type') + 1  # +1 because Excel is 1-indexed
        cls._add_dropdown_validation(sheet, section_type_col, cls.DROPDOWN_OPTIONS['section_type'],
                                    start_row=2, end_row=len(sections) + 1)

        # Add boolean dropdowns
        archived_col = headers.index('archived') + 1
        cls._add_dropdown_validation(sheet, archived_col, cls.DROPDOWN_OPTIONS['archived'],
                                    start_row=2, end_row=len(sections) + 1)

        allow_data_not_available_col = headers.index('allow_data_not_available') + 1
        cls._add_dropdown_validation(sheet, allow_data_not_available_col, cls.DROPDOWN_OPTIONS['allow_data_not_available'],
                                    start_row=2, end_row=len(sections) + 1)

        allow_not_applicable_col = headers.index('allow_not_applicable') + 1
        cls._add_dropdown_validation(sheet, allow_not_applicable_col, cls.DROPDOWN_OPTIONS['allow_not_applicable'],
                                    start_row=2, end_row=len(sections) + 1)

        # Add dropdown for page_id that references Pages sheet (if Pages sheet exists)
        if 'Pages' in workbook.sheetnames:
            page_id_col = headers.index('page_id') + 1
            pages_sheet = workbook['Pages']
            cls._add_sheet_reference_dropdown(sheet, page_id_col, pages_sheet, 'id',
                                             start_row=2, end_row=len(sections) + 1)

        # Create Excel table (always create, even with 0 rows)
        num_data_rows = len(sections)
        cls._create_excel_table(sheet, "SectionsTable", len(headers), num_data_rows + 1)  # +1 for header

        # Add conditional formatting for duplicate IDs (only if there are data rows)
        if num_data_rows > 0:
            cls._add_duplicate_highlighting(sheet, 'id', headers, num_data_rows + 1)

    @classmethod
    def _export_items_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export items to Items sheet with sequential IDs."""
        sheet = workbook.create_sheet("Items")

        # Get items for this version (include archived)
        items = cls._get_items_for_version(template, version)

        # Build DB -> export mapping for rewriting rule JSON inside exported strings
        item_db_to_export = {item.id: idx + 1 for idx, item in enumerate(items)}

        # Get section export ID mapping from Sections sheet
        sections_sheet = workbook['Sections']
        section_db_to_export = {}  # db_id -> export_id
        if hasattr(sections_sheet, '_section_export_id_map'):
            # Reverse the map: export_id -> db_id becomes db_id -> export_id
            for exp_id, db_id in sections_sheet._section_export_id_map.items():
                section_db_to_export[db_id] = exp_id

        # Write headers with required/optional styling
        headers = cls.get_item_columns()
        required_cols = cls.REQUIRED_COLUMNS.get('Items', [])
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            is_required = header in required_cols
            cls._style_header_cell(cell, is_required=is_required)

        # Write data rows with sequential IDs
        for row_idx, item in enumerate(items, start=2):
            export_id = row_idx - 1  # Sequential ID starting from 1

            # Map section_id to export ID
            section_export_id = None
            if item.section_id and item.section_id in section_db_to_export:
                section_export_id = section_db_to_export[item.section_id]

            row_data = cls._build_export_row(
                cls.ITEM_BASE_COLUMNS,
                cls.ITEM_TRANSLATABLE_FIELDS,
                {
                    'id': export_id,
                    'section_id': section_export_id,
                    'item_type': item.item_type,
                    'label': item.label,
                    'order': item.order,
                    'relevance_condition': cls._rewrite_rule_json_item_ids(item.relevance_condition, item_db_to_export),
                    'archived': item.archived,
                    'config': cls._format_json_for_excel(
                        cls._normalize_matrix_item_config(item.item_type, item.config)
                    ),
                    'indicator_bank_id': item.indicator_bank_id,
                    'type': item.type,
                    'unit': item.unit,
                    'validation_condition': cls._rewrite_rule_json_item_ids(item.validation_condition, item_db_to_export),
                    'validation_message': item.validation_message,
                    'definition': item.definition,
                    'options_json': cls._format_json_for_excel(item.options_json),
                    'lookup_list_id': item.lookup_list_id,
                    'list_display_column': item.list_display_column,
                    'list_filters_json': cls._format_json_for_excel(item.list_filters_json),
                    'options_translations': cls._format_json_for_excel(item.options_translations),
                    'description': item.description,
                },
                {
                    'label': item.label_translations,
                    'definition': item.definition_translations,
                    'description': item.description_translations,
                },
            )
            cls._write_data_row(sheet, row_idx, headers, row_data)

        # Auto-size columns
        cls._auto_size_columns(sheet, len(headers))

        # Add data validation dropdowns
        item_type_col = headers.index('item_type') + 1
        item_type_options = cls._get_item_type_options()
        cls._add_dropdown_validation(sheet, item_type_col, item_type_options,
                                    start_row=2, end_row=len(items) + 1)

        archived_col = headers.index('archived') + 1
        cls._add_dropdown_validation(sheet, archived_col, cls.DROPDOWN_OPTIONS['archived'],
                                    start_row=2, end_row=len(items) + 1)

        # Add dropdown for type column with dynamic values from database
        if 'type' in headers:
            type_col = headers.index('type') + 1
            type_options = cls._get_type_options_from_database()
            if type_options:
                cls._add_dropdown_validation(sheet, type_col, type_options,
                                            start_row=2, end_row=len(items) + 1)

        # Add dropdown for section_id that references Sections sheet
        if 'Sections' in workbook.sheetnames:
            section_id_col = headers.index('section_id') + 1
            sections_sheet = workbook['Sections']
            cls._add_sheet_reference_dropdown(sheet, section_id_col, sections_sheet, 'id',
                                             start_row=2, end_row=len(items) + 1)

        # Create Excel table (always create, even with 0 rows)
        num_data_rows = len(items)
        cls._create_excel_table(sheet, "ItemsTable", len(headers), num_data_rows + 1)  # +1 for header

        # Add conditional formatting for duplicate IDs (only if there are data rows)
        if num_data_rows > 0:
            cls._add_duplicate_highlighting(sheet, 'id', headers, num_data_rows + 1)

    @classmethod
    def _export_metadata_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export very hidden metadata sheet with system information."""
        sheet = workbook.create_sheet("_Metadata")
        sheet.sheet_state = 'veryHidden'  # Very hidden - cannot be unhidden via Excel UI

        # Metadata information
        metadata = {
            'Excel Export Version': cls.EXCEL_EXPORT_VERSION,
            'Export Timestamp': utcnow().isoformat(),
            'Template ID': template.id,
            'Template Name': template.name,
            'Version ID': version.id,
            'Version Number': version.version_number,
            'Version Status': version.status,
            'Exported By': current_user.email if current_user else 'System',
        }

        # Write metadata as key-value pairs
        row = 1
        for key, value in metadata.items():
            sheet.cell(row=row, column=1, value=key)
            sheet.cell(row=row, column=2, value=value)
            # Style the key column
            key_cell = sheet.cell(row=row, column=1)
            key_cell.font = Font(bold=True)
            row += 1

        # Auto-size columns
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 50

    @classmethod
    def _export_instructions_sheet(cls, workbook):
        """Export instructions sheet with formatting guide."""
        sheet = workbook.create_sheet("Instructions", 0)  # Insert at position 0 (first sheet)

        # Title
        title_cell = sheet.cell(row=1, column=1, value="Template Excel Import/Export Instructions")
        title_cell.font = Font(bold=True, size=16, color=cls.IFRC_COLORS['WHITE'])
        title_cell.fill = PatternFill(start_color=cls.IFRC_COLORS['DARK_RED'],
                                     end_color=cls.IFRC_COLORS['DARK_RED'],
                                     fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells('A1:D1')
        sheet.row_dimensions[1].height = 35

        # Section: Overview
        row = 3
        section_rows = []  # Track section header rows for border styling
        overview_cell = sheet.cell(row=row, column=1, value="📋 Overview")
        overview_cell.font = Font(bold=True, size=14, color=cls.IFRC_COLORS['DARK_RED'])
        overview_cell.fill = PatternFill(start_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                        end_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                        fill_type='solid')
        sheet.merge_cells(f'A{row}:D{row}')
        section_rows.append(row)
        row += 1

        instructions = [
            "This Excel file contains the structure of a form template that can be imported back into the system.",
            "",
            "The file contains the following sheets:",
            "  • Instructions (this sheet) - Ignored during import",
            "  • Template - Template metadata and configuration (field/value rows)",
            "  • Pages - Page definitions (if template is paginated)",
            "  • Sections - Section definitions",
            "  • Items - Form items (indicators, questions, document fields, etc.)",
            "",
            "Any sheets other than the recognized ones (Template, Pages, Sections, Items) will be ignored during import.",
        ]

        for instruction in instructions:
            cell = sheet.cell(row=row, column=1, value=instruction)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if instruction.startswith("  •"):
                cell.font = Font(size=10)
            row += 1

        # Section: Column Headers
        row += 2
        headers_cell = sheet.cell(row=row, column=1, value="🎨 Column Headers")
        headers_cell.font = Font(bold=True, size=14, color=cls.IFRC_COLORS['DARK_RED'])
        headers_cell.fill = PatternFill(start_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                       end_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                       fill_type='solid')
        sheet.merge_cells(f'A{row}:D{row}')
        section_rows.append(row)
        row += 1

        header_instructions = [
            "Column headers in Pages, Sections, and Items are color-coded to indicate whether they are required or optional:",
            "",
            "The Template sheet uses a field/value layout: column A lists field names and column B holds values (one field per row).",
            "Required Template fields (e.g. name) are shown in red in column A.",
            "",
            "  🔴 RED HEADERS (Required): These columns must have values. They are essential for the template to function.",
            "     Missing values in required columns will cause the import to fail.",
            "",
            "  🔵 BLUE HEADERS (Optional): These columns can be left empty. They provide additional configuration or metadata.",
            "     Empty optional columns will use default values or be set to NULL.",
            "",
            "Required columns for each sheet:",
        ]

        for instruction in header_instructions:
            cell = sheet.cell(row=row, column=1, value=instruction)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            row += 1

        # List required columns for each sheet
        for sheet_name in ['Template', 'Pages', 'Sections', 'Items']:
            required = cls.REQUIRED_COLUMNS.get(sheet_name, [])
            if required:
                cell = sheet.cell(row=row, column=1, value=f"  {sheet_name}: {', '.join(required)}")
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                row += 1

        # Section: Important Notes
        row += 2
        notes_cell = sheet.cell(row=row, column=1, value="⚠️ Important Notes")
        notes_cell.font = Font(bold=True, size=14, color=cls.IFRC_COLORS['DARK_RED'])
        notes_cell.fill = PatternFill(start_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                     end_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                     fill_type='solid')
        sheet.merge_cells(f'A{row}:D{row}')
        section_rows.append(row)
        row += 1

        notes = [
            "1. ID Fields: The 'id' columns in Pages, Sections, and Items are sequential export IDs (1, 2, 3...).",
            "   They are NOT database IDs. The system will create new IDs when importing.",
            "",
            "2. References: When referencing other records (e.g., section_id in Items), use the export IDs from the same file.",
            "",
            "3. JSON Fields: Fields such as config, options_json, and options_translations remain JSON. Exports are pretty-printed with line breaks for readability; re-import accepts both compact and formatted JSON.",
            "",
            "4. Translation Fields: English/base text lives in columns like name, label, definition, and description. Other languages use separate columns named {field}_{lang} (e.g. label_ar, label_fr, name_es). Empty translation cells are ignored on import.",
            "",
            "5. Boolean Fields: Use TRUE/FALSE or 1/0 for boolean values.",
            "",
            "6. Order Fields: Use numeric values (integers or decimals like 1, 1.1, 1.2) to control display order.",
            "",
            "7. Import Behavior:",
            "   • If importing into a published version, a new draft version will be created automatically.",
            "   • Existing items/sections with matching order+name will be updated, others will be created.",
            "   • The import will fail if required columns are missing or invalid.",
            "",
            "8. Do NOT modify the Instructions sheet or add unrecognized sheets if you plan to re-import the file.",
        ]

        for note in notes:
            cell = sheet.cell(row=row, column=1, value=note)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if note and note[0].isdigit():
                cell.font = Font(bold=True, size=10)
            row += 1

        # Auto-size columns
        sheet.column_dimensions['A'].width = 100
        sheet.column_dimensions['B'].width = 5
        sheet.column_dimensions['C'].width = 5
        sheet.column_dimensions['D'].width = 5

        # Add borders to section headers for better visual separation
        for section_row in section_rows:
            for col in ['A', 'B', 'C', 'D']:
                cell = sheet[f'{col}{section_row}']
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='medium'),
                    bottom=Side(style='thin')
                )

        # Freeze first row
        sheet.freeze_panes = 'A2'

        # Protect the instructions sheet (read-only, but allow formatting)
        sheet.protection.sheet = True
        sheet.protection.formatCells = False
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False

    @classmethod
    def _create_excel_table(cls, sheet, table_name: str, num_columns: int, num_rows: int):
        """Create an Excel table from the data range.

        Args:
            sheet: The worksheet to add the table to
            table_name: Name for the table
            num_columns: Number of columns
            num_rows: Number of rows (including header)
            Note: Excel requires at least 2 rows (header + 1 data row), so if num_rows=1, we create with 2 rows
        """
        if num_rows < 1:  # Need at least header row
            return

        # Excel requires at least 2 rows for a table (header + at least one data row)
        # If we only have a header, create table with header + 1 empty data row
        if num_rows == 1:
            num_rows = 2  # Add one empty data row

        # Calculate range (A1 to last column, last row)
        start_col = get_column_letter(1)
        end_col = get_column_letter(num_columns)
        table_range = f"{start_col}1:{end_col}{num_rows}"

        # Create table
        table = Table(displayName=table_name, ref=table_range)

        # Style the table
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style

        # Add table to sheet
        sheet.add_table(table)

    @classmethod
    def _add_dropdown_validation(cls, sheet, column: int, options: List[str],
                                 start_row: int = 2, end_row: int = 1000):
        """Add data validation dropdown to a column.

        Args:
            sheet: The worksheet
            column: Column number (1-indexed)
            options: List of options for the dropdown
            start_row: Starting row (default 2, after header)
            end_row: Ending row
        """
        if not options:
            return

        # Create comma-separated list of options
        # Excel list validation formula format: "option1,option2,option3"
        options_str = ','.join(options)
        formula = f'"{options_str}"'

        # Create data validation
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = 'Invalid value'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = 'Please select from the dropdown'
        dv.promptTitle = 'Select Value'

        # Apply to column range
        col_letter = get_column_letter(column)
        dv_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
        dv.add(dv_range)

        # Add validation to sheet
        sheet.add_data_validation(dv)

    @classmethod
    def _add_sheet_reference_dropdown(cls, sheet, column: int, source_sheet,
                                      source_column_name: str,
                                      start_row: int = 2, end_row: int = 1000):
        """Add data validation dropdown that references values from another sheet using table column reference.

        Uses structured table references (e.g., PagesTable[id]) which automatically expand when new rows are added.

        Args:
            sheet: The worksheet to add validation to
            column: Column number (1-indexed) in the target sheet
            source_sheet: The source worksheet to reference
            source_column_name: Name of the column in source sheet to reference
            start_row: Starting row (default 2, after header)
            end_row: Ending row
        """
        try:
            # Find the source column index in the source sheet
            source_headers = [cell.value for cell in source_sheet[1]]
            if source_column_name not in source_headers:
                current_app.logger.warning(f"Source column '{source_column_name}' not found in source sheet")
                return

            source_col_idx = source_headers.index(source_column_name) + 1
            source_col_letter = get_column_letter(source_col_idx)

            # Prepare sheet reference
            source_sheet_name = source_sheet.title
            # Escape sheet name if it contains spaces or special characters
            if ' ' in source_sheet_name or '-' in source_sheet_name:
                source_sheet_ref = f"'{source_sheet_name}'"
            else:
                source_sheet_ref = source_sheet_name

            # Use OFFSET with COUNTA to create a dynamic range that only includes non-empty cells
            # OFFSET(start_cell, rows, cols, height, width)
            # COUNTA counts all non-empty cells in the column, subtract 1 to exclude header row
            # This creates a range that automatically expands/contracts based on actual data
            # Format: OFFSET(Pages!$A$2,0,0,COUNTA(Pages!$A:$A)-1,1)
            source_range = f"OFFSET({source_sheet_ref}!${source_col_letter}$2,0,0,COUNTA({source_sheet_ref}!${source_col_letter}:${source_col_letter})-1,1)"

            current_app.logger.debug(f"Creating dropdown with OFFSET formula: {source_range}")

            # Create data validation with reference formula
            dv = DataValidation(type="list", formula1=source_range, allow_blank=True)
            # Generic error messages since this is used for both page_id and section_id
            dv.error = f'Invalid {source_column_name}. Please select from the dropdown.'
            dv.errorTitle = 'Invalid Entry'
            dv.prompt = f'Please select a {source_column_name} from the {source_sheet.title} sheet'
            dv.promptTitle = f'Select {source_column_name.replace("_", " ").title()}'

            # Apply to column range
            col_letter = get_column_letter(column)
            dv_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
            dv.add(dv_range)

            # Add validation to sheet
            sheet.add_data_validation(dv)

            current_app.logger.info(f"Successfully added sheet reference dropdown: {source_range} for column {col_letter}, range {dv_range}")

        except Exception as e:
            current_app.logger.error(f"Could not add sheet reference dropdown: {e}", exc_info=True)

    @classmethod
    def _get_item_type_options(cls) -> List[str]:
        """Get all supported item types including plugin types.

        Returns:
            List of all item types: standard types (indicator, question, document_field, matrix)
            plus all active plugin field types (prefixed with 'plugin_')

        Note: These match the hardcoded values in form_builder.html dropdown:
        - indicator (Indicator from Bank)
        - question (New Question)
        - document_field (Document Field - stored as document_field in DB, shown as 'document' in UI)
        - matrix (Matrix Table)
        - plugin_* (dynamically added from plugin system)
        """
        item_types = []

        # Add standard item types - these are the hardcoded fallback values currently in the system
        # Matching the form_builder.html dropdown options (lines 1624-1627)
        # Note: UI shows 'document' but DB stores 'document_field', so we use 'document_field' here
        standard_types = ['indicator', 'question', 'document_field', 'matrix']
        item_types.extend(standard_types)

        # Add plugin field types if plugin manager is available
        try:
            if hasattr(current_app, 'plugin_manager'):
                plugin_manager = current_app.plugin_manager
                active_field_types = plugin_manager.list_active_field_types()

                # Prefix each plugin field type with 'plugin_'
                for field_type in active_field_types:
                    plugin_item_type = f'plugin_{field_type}'
                    if plugin_item_type not in item_types:
                        item_types.append(plugin_item_type)
        except Exception as e:
            current_app.logger.warning(f"Could not fetch plugin field types: {e}")

        # Sort for consistent ordering
        return sorted(item_types)

    @classmethod
    def _get_type_options_from_database(cls) -> List[str]:
        """Get distinct type values from FormItem table dynamically.

        Returns:
            List of unique type values found in the database, sorted alphabetically
        """
        try:
            # Query distinct non-null type values from FormItem
            distinct_types = db.session.query(FormItem.type).filter(
                FormItem.type.isnot(None),
                FormItem.type != ''
            ).distinct().all()

            # Extract values and sort
            type_options = sorted([t[0] for t in distinct_types if t[0]])

            # If no types found in database, return common defaults
            if not type_options:
                type_options = ['Number', 'Percentage', 'Text', 'Boolean', 'Date', 'Choice']

            return type_options

        except Exception as e:
            current_app.logger.warning(f"Could not fetch type options from database: {e}")
            # Return common defaults as fallback
            return ['Number', 'Percentage', 'Text', 'Boolean', 'Date', 'Choice']

    @classmethod
    def _add_duplicate_highlighting(cls, sheet, column_name: str, headers: List[str],
                                     num_rows: int):
        """Add conditional formatting to highlight duplicate values in a column.

        Args:
            sheet: The worksheet
            column_name: Name of the column to check for duplicates
            headers: List of header names (to find column index)
            num_rows: Total number of rows (including header)
        """
        if num_rows < 3:  # Need at least header + 2 data rows to have duplicates
            return

        try:
            # Find column index
            if column_name not in headers:
                return

            col_idx = headers.index(column_name) + 1  # +1 because Excel is 1-indexed
            col_letter = get_column_letter(col_idx)

            # Create range for data rows (skip header row)
            data_range = f"{col_letter}2:{col_letter}{num_rows}"

            # Create duplicate values rule with yellow background
            # Use COUNTIF formula to detect duplicates: COUNTIF($A$2:$A$100, A2)>1
            # This formula checks if the current cell value appears more than once in the range
            duplicate_fill = PatternFill(start_color=cls.IFRC_COLORS['YELLOW'],
                                        end_color=cls.IFRC_COLORS['YELLOW'],
                                        fill_type='solid')

            # Formula: COUNTIF(absolute_range, relative_cell) > 1
            # $col_letter$2:$col_letter$num_rows is the absolute range
            # col_letter2 is the relative cell reference (will adjust per row)
            formula = f'COUNTIF(${col_letter}$2:${col_letter}${num_rows},{col_letter}2)>1'

            duplicate_rule = FormulaRule(formula=[formula], fill=duplicate_fill)

            # Add conditional formatting rule
            sheet.conditional_formatting.add(data_range, duplicate_rule)

        except Exception as e:
            # Log but don't fail if conditional formatting can't be added
            current_app.logger.warning(f"Could not add duplicate highlighting for column {column_name}: {e}")

    @classmethod
    def _style_header_cell(cls, cell, is_required=True):
        """Apply IFRC styling to header cell.

        Args:
            cell: The cell to style
            is_required: If True, style as required (red), if False, style as optional (blue)
        """
        cell.font = Font(bold=True, color=cls.IFRC_COLORS['WHITE'])
        # Use red for required columns, blue for optional columns
        fill_color = cls.IFRC_COLORS['DARK_RED'] if is_required else cls.IFRC_COLORS['DARK_BLUE']
        cell.fill = PatternFill(start_color=fill_color,
                               end_color=fill_color,
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    @classmethod
    def _auto_size_columns(cls, sheet, num_columns):
        """Auto-size columns for better readability."""
        for col_idx in range(1, num_columns + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in sheet[column_letter]:
                with suppress(Exception):
                    if row.value:
                        lines = str(row.value).splitlines() or [str(row.value)]
                        max_length = max(max_length, max(len(line) for line in lines))
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            sheet.column_dimensions[column_letter].width = adjusted_width

    @classmethod
    def _count_nonempty_data_rows(cls, sheet) -> int:
        """Count non-empty data rows (row 2 onward) in a worksheet."""
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and not all(cell is None for cell in row):
                count += 1
        return count

    @classmethod
    def _validate_template_sheet_headers(cls, headers: List[Any]) -> List[str]:
        """Validate Template sheet headers (row layout or legacy column layout)."""
        if cls._template_sheet_uses_row_layout(headers):
            return []
        required_headers = cls.REQUIRED_COLUMNS.get('Template', ['name'])
        header_set = {h for h in headers if isinstance(h, str) and h}
        required_set = set(required_headers)
        if not required_set.issubset(header_set):
            missing = required_set - header_set
            return [
                f"Template sheet headers missing required columns. "
                f"Required: {required_headers}, Missing: {list(missing)}, Got: {headers}"
            ]
        return []

    @classmethod
    def _validate_import_sheet_headers(cls, sheet_name: str, headers: List[Any]) -> List[str]:
        _, _, error = cls._resolve_sheet_headers(sheet_name, headers)
        return [error] if error else []

    @classmethod
    def validate_import_file(cls, excel_file) -> Dict[str, Any]:
        """
        Validate a Humanitarian Databank Excel export before import.

        Returns dict with keys: valid, message, errors, preview
        (preview: name, pages, sections, items).
        """
        errors: List[str] = []
        preview = {'name': None, 'pages': 0, 'sections': 0, 'items': 0}

        try:
            workbook = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
        except Exception as e:
            current_app.logger.error(f"Failed to load Excel file for validation: {e}", exc_info=True)
            cls._matrix_import_log(f"validate: failed to load workbook: {e}", level='error')
            return {
                'valid': False,
                'message': 'Invalid Excel file. Check the file format and try again.',
                'errors': ['Failed to load Excel file.'],
                'preview': preview,
            }

        required_sheets = ['Template', 'Pages', 'Sections', 'Items']
        missing_sheets = [s for s in required_sheets if s not in workbook.sheetnames]
        if missing_sheets:
            msg = f"Missing required sheets: {', '.join(missing_sheets)}"
            return {
                'valid': False,
                'message': msg,
                'errors': [msg],
                'preview': preview,
            }

        template_sheet = workbook['Template']
        template_headers = [cell.value for cell in template_sheet[1]]
        errors.extend(cls._validate_template_sheet_headers(template_headers))
        row_data, legacy_column_layout, template_parse_errors = cls._parse_template_sheet(template_sheet)
        errors.extend(template_parse_errors)
        if not errors:
            if not row_data:
                errors.append('Template sheet has no data row.')
            else:
                template_name = row_data.get('name')
                if template_name is None or str(template_name).strip() == '':
                    errors.append('Template sheet data row is missing a template name.')
                else:
                    preview['name'] = str(template_name).strip()

        pages_sheet = workbook['Pages']
        pages_headers = [cell.value for cell in pages_sheet[1]]
        errors.extend(cls._validate_import_sheet_headers('Pages', pages_headers))
        if not any(e.startswith('Pages sheet') for e in errors):
            preview['pages'] = cls._count_nonempty_data_rows(pages_sheet)

        sections_sheet = workbook['Sections']
        sections_headers = [cell.value for cell in sections_sheet[1]]
        errors.extend(cls._validate_import_sheet_headers('Sections', sections_headers))
        if not any(e.startswith('Sections sheet') for e in errors):
            preview['sections'] = cls._count_nonempty_data_rows(sections_sheet)

        items_sheet = workbook['Items']
        items_headers = [cell.value for cell in items_sheet[1]]
        errors.extend(cls._validate_import_sheet_headers('Items', items_headers))
        if not any(e.startswith('Items sheet') for e in errors):
            preview['items'] = cls._count_nonempty_data_rows(items_sheet)

        cls._scan_workbook_matrix_items(workbook, stage='validate')

        valid = len(errors) == 0
        if valid:
            message = (
                f"Valid export: {preview['pages']} pages, "
                f"{preview['sections']} sections, {preview['items']} items."
            )
        else:
            if len(errors) == 1:
                message = errors[0]
            else:
                message = f"Found {len(errors)} validation errors."

        return {
            'valid': valid,
            'message': message,
            'errors': errors,
            'preview': preview,
        }

    @classmethod
    def _parse_import_version_mode(cls, import_version_mode: Optional[str], target_version) -> bool:
        """Return True when the import should land in a draft (new or existing)."""
        if import_version_mode == 'create_draft':
            return True
        if import_version_mode == 'current_version':
            return False
        return target_version.status == 'published'

    @classmethod
    def _resolve_target_version(cls, template, version_id: Optional[int]):
        """Resolve the base version row for an import request."""
        target_version = None
        if version_id:
            target_version = FormTemplateVersion.query.filter_by(
                id=version_id, template_id=template.id
            ).first()

        if not target_version:
            if template.published_version_id:
                target_version = FormTemplateVersion.query.get(template.published_version_id)
            else:
                target_version = FormTemplateVersion.query.filter_by(
                    template_id=template.id
                ).order_by(FormTemplateVersion.created_at.desc()).first()

        return target_version

    @classmethod
    def resolve_import_target_version_id(
        cls,
        template_id: int,
        version_id: Optional[int],
        import_version_mode: Optional[str] = None,
    ) -> Optional[int]:
        """Return the version_id that will receive the import (without creating rows)."""
        template = FormTemplate.query.get(template_id)
        if not template:
            return version_id

        target_version = cls._resolve_target_version(template, version_id)
        if not target_version:
            return version_id

        if cls._parse_import_version_mode(import_version_mode, target_version):
            if target_version.status == 'published':
                existing_draft = FormTemplateVersion.query.filter_by(
                    template_id=template.id, status='draft'
                ).first()
                return existing_draft.id if existing_draft else None
            return target_version.id

        return target_version.id

    @classmethod
    def _get_or_create_draft_for_import(cls, template, source_version) -> FormTemplateVersion:
        """Return an existing draft or create one cloned from source_version."""
        existing_draft = FormTemplateVersion.query.filter_by(
            template_id=template.id, status='draft'
        ).first()
        if existing_draft:
            current_app.logger.info(
                f"Draft version already exists (ID={existing_draft.id}); using it for Excel import"
            )
            return existing_draft

        from sqlalchemy import func
        max_version = db.session.query(func.max(FormTemplateVersion.version_number)).filter_by(
            template_id=template.id
        ).scalar()
        next_version_number = (max_version + 1) if max_version else 1

        new_draft = FormTemplateVersion(
            template_id=template.id,
            version_number=next_version_number,
            status='draft',
            created_by=current_user.id,
            updated_by=current_user.id,
            based_on_version_id=source_version.id,
            comment='Created automatically for Excel import',
            name=source_version.name,
            name_translations=source_version.name_translations.copy() if source_version.name_translations else None,
            description=source_version.description,
            add_to_self_report=source_version.add_to_self_report,
            display_order_visible=source_version.display_order_visible,
            is_paginated=source_version.is_paginated,
            enable_export_pdf=source_version.enable_export_pdf,
            enable_export_excel=source_version.enable_export_excel,
            enable_import_excel=source_version.enable_import_excel,
            enable_ai_validation=getattr(source_version, 'enable_ai_validation', False),
            enable_data_quality=getattr(source_version, 'enable_data_quality', False),
            data_quality_methodology=getattr(source_version, 'data_quality_methodology', None),
            validation_rule_pack=getattr(source_version, 'validation_rule_pack', None),
            variables=source_version.variables.copy() if source_version.variables else None
        )
        db.session.add(new_draft)
        db.session.flush()
        current_app.logger.info(
            f"Created new draft version: ID={new_draft.id}, Version #={new_draft.version_number}"
        )
        cls._matrix_import_log(
            f"Published import: created draft {new_draft.id} from version {source_version.id} "
            f"(structure will be rebuilt from Excel)"
        )
        return new_draft

    @classmethod
    @memory_tracker("Template Excel Import", log_top_allocations=True)
    def import_template(
        cls,
        template_id: int,
        excel_file,
        version_id: Optional[int] = None,
        import_version_mode: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Import template structure from Excel.

        Args:
            template_id: Template ID to import into
            excel_file: File-like object containing Excel file
            version_id: Optional version ID to import into (defaults to active version or creates draft)
            import_version_mode: 'create_draft' | 'current_version' — UI choice for import target

        Returns:
            Dict with 'success', 'message', 'errors', 'created_count' keys
        """
        cls.matrix_import_entry_log(
            f"import_template START template_id={template_id} version_id={version_id}"
        )
        current_app.logger.info(f"=== TEMPLATE EXCEL IMPORT START ===")
        current_app.logger.info(f"Template ID: {template_id}, Version ID: {version_id}")

        template = FormTemplate.query.get_or_404(template_id)
        current_app.logger.info(f"Template found: '{template.name}'")

        errors = []
        warnings: List[str] = []
        created_counts = {'pages': 0, 'sections': 0, 'items': 0}

        try:
            # Load workbook
            current_app.logger.info("Loading Excel workbook...")
            workbook = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
            current_app.logger.info(f"Workbook loaded. Sheets found: {workbook.sheetnames}")
            cls._scan_workbook_matrix_items(workbook, stage='import-pre-scan')

            # Validate required sheets exist (ignore Instructions, _Metadata and any other unrecognized sheets)
            required_sheets = ['Template', 'Pages', 'Sections', 'Items']
            recognized_sheets = set(required_sheets + ['Instructions', '_Metadata'])  # Instructions and _Metadata are recognized but optional
            missing_sheets = [s for s in required_sheets if s not in workbook.sheetnames]

            # Log any unrecognized sheets (they will be ignored)
            unrecognized_sheets = [s for s in workbook.sheetnames if s not in recognized_sheets]
            if unrecognized_sheets:
                current_app.logger.info(f"Ignoring unrecognized sheets during import: {unrecognized_sheets}")

            if missing_sheets:
                current_app.logger.error(f"Missing required sheets: {missing_sheets}")
                return {
                    'success': False,
                    'message': f"Missing required sheets: {', '.join(missing_sheets)}",
                    'errors': errors,
                    'created_count': created_counts
                }

            # Determine target version
            current_app.logger.info("Determining target version...")
            target_version = cls._resolve_target_version(template, version_id)
            if target_version and version_id:
                current_app.logger.info(
                    f"Using specified version ID {version_id} "
                    f"(Status: {target_version.status}, Version #: {target_version.version_number})"
                )
            elif target_version:
                current_app.logger.info(
                    f"Using fallback version ID {target_version.id} "
                    f"(Status: {target_version.status})"
                )

            if not target_version:
                current_app.logger.error("No version found for template")
                return {
                    'success': False,
                    'message': "No version found for this template. Please create a version first.",
                    'errors': errors,
                    'created_count': created_counts
                }

            create_new_draft = cls._parse_import_version_mode(import_version_mode, target_version)

            # When importing into a draft, create/use a draft version from published.
            if create_new_draft and target_version.status == 'published':
                current_app.logger.info("Import target is a draft; preparing draft version...")
                target_version = cls._get_or_create_draft_for_import(template, target_version)
                version_id = target_version.id

            current_app.logger.info(
                f"Target version: ID={target_version.id}, Status={target_version.status}, "
                f"Version #={target_version.version_number}"
            )

            if not create_new_draft and target_version.status == 'published':
                deletion_impact = cls._count_deletion_impact(template.id, target_version.id)
                if deletion_impact.get('has_data'):
                    return {
                        'success': False,
                        'message': (
                            'This template has submitted data. Import into a draft version '
                            'and deploy it instead.'
                        ),
                        'errors': [
                            'Excel import into the published version is blocked when submission data exists.'
                        ],
                        'created_count': created_counts,
                    }

            # Import Template metadata first (including version-specific name)
            current_app.logger.info("=== IMPORTING TEMPLATE METADATA ===")
            template_sheet = workbook['Template']
            template_errors = cls._import_template_metadata(template_sheet, template, target_version)
            errors.extend(template_errors)
            if template_errors:
                current_app.logger.warning(f"Template metadata import errors: {template_errors}")
            else:
                current_app.logger.info("Template metadata imported successfully")

            # Excel import is a full structure replace: drop existing pages/sections/items first
            # so renamed sections, changed types, and removed subsections don't linger.
            published_stable_key_context = cls._build_published_stable_key_context(template)
            cls._clear_version_structure(template.id, target_version.id)

            # ID mapping dictionaries (export ID -> new database ID)
            # Export IDs are sequential (1, 2, 3...) and map to new database IDs
            page_id_map = {}  # export_id -> new_db_id
            section_id_map = {}  # export_id -> new_db_id

            # Import Pages
            current_app.logger.info("=== IMPORTING PAGES ===")
            pages_sheet = workbook['Pages']
            page_errors = cls._import_pages(pages_sheet, template, target_version, page_id_map)
            errors.extend(page_errors)
            created_counts['pages'] = len(page_id_map)
            current_app.logger.info(f"Pages imported: {created_counts['pages']} pages created")
            current_app.logger.info(f"Page ID mapping: {page_id_map}")
            if page_errors:
                current_app.logger.warning(f"Page import errors: {page_errors}")

            # Import Sections
            current_app.logger.info("=== IMPORTING SECTIONS ===")
            sections_sheet = workbook['Sections']
            section_errors = cls._import_sections(
                sections_sheet,
                template,
                target_version,
                page_id_map,
                section_id_map,
                published_stable_key_context=published_stable_key_context,
            )
            errors.extend(section_errors)
            created_counts['sections'] = len(section_id_map)
            current_app.logger.info(f"Sections imported: {created_counts['sections']} sections created")
            current_app.logger.info(f"Section ID mapping: {section_id_map}")
            if section_errors:
                current_app.logger.warning(f"Section import errors: {section_errors}")

            # Import Items
            current_app.logger.info("=== IMPORTING ITEMS ===")
            items_sheet = workbook['Items']
            item_errors = cls._import_items(
                items_sheet,
                template,
                target_version,
                section_id_map,
                warnings=warnings,
                published_stable_key_context=published_stable_key_context,
            )
            errors.extend(item_errors)
            created_counts['items'] = db.session.query(FormItem).filter_by(
                template_id=template.id, version_id=target_version.id
            ).count()
            current_app.logger.info(f"Items imported: {created_counts['items']} items created")
            if item_errors:
                current_app.logger.warning(f"Item import errors: {item_errors}")

            if errors:
                current_app.logger.error(f"Import completed with {len(errors)} errors. Rolling back...")
                db.session.rollback()
                return {
                    'success': False,
                    'message': f"Import completed with {len(errors)} errors",
                    'errors': errors,
                    'created_count': created_counts
                }

            # Update version timestamp
            current_app.logger.info("Updating version timestamp...")
            target_version.updated_at = utcnow()
            target_version.updated_by = current_user.id

            current_app.logger.info("Committing database changes...")
            db.session.commit()
            cls._log_matrix_items_in_version(template.id, target_version.id, stage='after-commit')
            current_app.logger.info("=== TEMPLATE EXCEL IMPORT SUCCESS ===")
            current_app.logger.info(f"Final counts - Pages: {created_counts['pages']}, Sections: {created_counts['sections']}, Items: {created_counts['items']}")

            return {
                'success': True,
                'message': f"Successfully imported {created_counts['pages']} pages, "
                          f"{created_counts['sections']} sections, "
                          f"{created_counts['items']} items",
                'errors': [],
                'warnings': warnings,
                'created_count': created_counts,
                'version_id': version_id  # Return the version ID (may be new draft if published was selected)
            }

        except Exception as e:
            current_app.logger.error(f"=== TEMPLATE EXCEL IMPORT FAILED ===")
            current_app.logger.error(f"Error importing template from Excel: {e}", exc_info=True)
            db.session.rollback()
            return {
                'success': False,
                'message': "Error importing template. Check the file format and try again.",
                'errors': ['Import failed. See logs for details.'],
                'created_count': created_counts
            }

    @classmethod
    def _import_template_metadata(cls, sheet, template: FormTemplate, version: FormTemplateVersion) -> List[str]:
        """Import template metadata from Template sheet, including version-specific name."""
        current_app.logger.info("Starting template metadata import...")
        errors = []

        row_data_all, legacy_column_layout, parse_errors = cls._parse_template_sheet(sheet)
        errors.extend(parse_errors)
        if parse_errors:
            return errors

        if not row_data_all:
            current_app.logger.warning("Template sheet has no data row")
            return errors

        headers = [cell.value for cell in sheet[1]]
        current_app.logger.info(f"Template sheet headers: {headers}")

        if legacy_column_layout:
            _, legacy_format, _ = cls._resolve_sheet_headers('Template', headers)
            expected_headers = (
                cls.TEMPLATE_LEGACY_COLUMNS if legacy_format else cls.get_template_columns()
            )
        else:
            legacy_format = False
            expected_headers = cls.get_template_columns()

        header_set = set(row_data_all.keys())
        required_headers = cls.REQUIRED_COLUMNS.get('Template', ['name'])
        expected_set = set(expected_headers)
        extra_columns = header_set - expected_set
        missing_optional = expected_set - header_set - set(required_headers)

        if extra_columns:
            current_app.logger.info(
                f"Template sheet contains extra/legacy fields (will be ignored): {list(extra_columns)}"
            )
        if missing_optional:
            current_app.logger.info(
                f"Template sheet missing optional fields (will use defaults): {list(missing_optional)}"
            )

        row_data = {h: row_data_all.get(h) for h in expected_headers}

        try:
            current_app.logger.info(f"Updating template metadata for template ID {template.id}")

            # Update version-specific fields (name is now only stored in versions)
            if 'name' in row_data and row_data['name']:
                version.name = row_data['name']
                current_app.logger.info(f"Updated version name: '{version.name}'")

            if 'description' in row_data:
                version.description = row_data['description']
                current_app.logger.info(f"Updated version/template description")

            if 'add_to_self_report' in row_data:
                version.add_to_self_report = bool(row_data['add_to_self_report'])
                current_app.logger.info(f"Updated version/template add_to_self_report: {version.add_to_self_report}")

            if 'display_order_visible' in row_data:
                version.display_order_visible = bool(row_data['display_order_visible'])
                current_app.logger.info(f"Updated version/template display_order_visible: {version.display_order_visible}")

            if 'is_paginated' in row_data:
                version.is_paginated = bool(row_data['is_paginated'])
                current_app.logger.info(f"Updated version/template is_paginated: {version.is_paginated}")

            if 'enable_export_pdf' in row_data:
                version.enable_export_pdf = bool(row_data['enable_export_pdf'])
                current_app.logger.info(f"Updated version/template enable_export_pdf: {version.enable_export_pdf}")

            if 'enable_export_excel' in row_data:
                version.enable_export_excel = bool(row_data['enable_export_excel'])
                current_app.logger.info(f"Updated version/template enable_export_excel: {version.enable_export_excel}")

            if 'enable_import_excel' in row_data:
                version.enable_import_excel = bool(row_data['enable_import_excel'])
                current_app.logger.info(f"Updated version/template enable_import_excel: {version.enable_import_excel}")

            if 'enable_ai_validation' in row_data:
                version.enable_ai_validation = bool(row_data['enable_ai_validation'])
                current_app.logger.info(f"Updated version/template enable_ai_validation: {version.enable_ai_validation}")

            name_translations = cls._collect_translations_from_row(
                row_data_all, 'name', legacy=legacy_format
            )
            if name_translations:
                version.name_translations = name_translations
                current_app.logger.info("Updated version name_translations")

            if 'variables' in row_data:
                raw_variables = row_data.get('variables')
                current_app.logger.info(
                    f"Variables cell raw value type={type(raw_variables).__name__} "
                    f"len={len(str(raw_variables)) if raw_variables is not None else 0}: "
                    f"{str(raw_variables)[:200]}"
                )
                variables = cls._parse_json(raw_variables)
                if variables is not None:  # Allow empty dict {} to clear variables
                    # Save to version (version-specific template variables)
                    version.variables = variables if variables else {}
                    from sqlalchemy.orm.attributes import flag_modified
                    flag_modified(version, 'variables')
                    current_app.logger.info(f"Updated version variables: {len(variables) if variables else 0} variable(s)")
                elif raw_variables is not None:
                    current_app.logger.warning(
                        f"Could not parse variables JSON from Excel (value truncated or malformed): "
                        f"{str(raw_variables)[:300]}"
                    )
                else:
                    current_app.logger.info("variables cell is empty in Excel — skipping (no overwrite)")

            db.session.add(template)
            db.session.add(version)
            current_app.logger.info("Template metadata import complete")

        except Exception as e:
            current_app.logger.error("Template metadata row: %s", e, exc_info=True)
            errors.append("Template metadata row: Validation error.")

        return errors

    @classmethod
    def _import_pages(cls, sheet, template: FormTemplate, version: FormTemplateVersion,
                     page_id_map: Dict[int, int]) -> List[str]:
        """Import pages from sheet."""
        current_app.logger.info("Starting pages import...")
        errors = []

        # Read headers
        headers = [cell.value for cell in sheet[1]]
        current_app.logger.info(f"Pages sheet headers: {headers}")

        # Validate headers match expected columns
        _, legacy_format, header_error = cls._resolve_sheet_headers('Pages', headers)
        if header_error:
            current_app.logger.error(header_error)
            errors.append(header_error)
            return errors

        current_app.logger.info("Headers validated. Processing page rows...")
        row_count = 0

        # Read data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue  # Skip empty rows

            row_count += 1
            try:
                row_data = cls._row_dict_from_sheet(headers, row)

                # Get export ID (sequential ID from Excel, e.g., 1, 2, 3...)
                export_id = int(row_data['id']) if row_data.get('id') else None

                if export_id is None:
                    error_msg = f"Pages row {row_idx}: Missing export ID"
                    current_app.logger.warning(error_msg)
                    errors.append(error_msg)
                    continue

                page_name = row_data['name'] or ''
                page_order = int(row_data['order']) if row_data['order'] is not None else 1
                current_app.logger.info(f"Processing page row {row_idx}: export_id={export_id}, name='{page_name}', order={page_order}")

                # Create new page
                new_page = FormPage(
                    template_id=template.id,
                    version_id=version.id,
                    name=page_name,
                    order=page_order,
                    name_translations=cls._collect_translations_from_row(
                        row_data, 'name', legacy=legacy_format
                    ),
                )

                db.session.add(new_page)
                db.session.flush()

                # Map export ID to new database ID
                page_id_map[export_id] = new_page.id
                current_app.logger.info(f"Page created: export_id={export_id} -> db_id={new_page.id}, name='{page_name}'")

            except Exception as e:
                current_app.logger.error("Pages row %s: %s", row_idx, e, exc_info=True)
                errors.append(f"Pages row {row_idx}: Validation error.")

        current_app.logger.info(f"Pages import complete: {row_count} rows processed, {len(page_id_map)} pages created, {len(errors)} errors")
        return errors

    @classmethod
    def _build_published_stable_key_context(cls, template: FormTemplate) -> Dict[str, Any]:
        """Map published structure to stable_key values for silent reuse on Excel import."""
        empty: Dict[str, Any] = {
            'sections': {},
            'items_by_indicator': {},
            'items_by_position': {},
        }
        if not template.published_version_id:
            return empty

        sections = FormSection.query.filter_by(
            template_id=template.id,
            version_id=template.published_version_id,
        ).all()
        if not sections:
            return empty

        sections_by_id = {section.id: section for section in sections}
        section_keys: Dict[Tuple[float, str], str] = {}
        for section in sections:
            if section.stable_key:
                section_keys[(float(section.order), section.name or '')] = section.stable_key

        items_by_indicator: Dict[Tuple[float, str, int, float], str] = {}
        items_by_position: Dict[Tuple[float, str, float, str, str], str] = {}
        items = FormItem.query.filter_by(
            template_id=template.id,
            version_id=template.published_version_id,
        ).all()
        for item in items:
            if not item.stable_key:
                continue
            section = sections_by_id.get(item.section_id)
            if not section:
                continue
            section_key = (float(section.order), section.name or '')
            item_order = float(item.order) if item.order is not None else 0.0
            item_type = (item.item_type or '').strip().lower()
            if item.indicator_bank_id:
                indicator_key = (
                    section_key[0],
                    section_key[1],
                    int(item.indicator_bank_id),
                    item_order,
                )
                items_by_indicator[indicator_key] = item.stable_key
            else:
                position_key = (
                    section_key[0],
                    section_key[1],
                    item_order,
                    item_type,
                    str(item.label or '').strip(),
                )
                items_by_position[position_key] = item.stable_key

        return {
            'sections': section_keys,
            'items_by_indicator': items_by_indicator,
            'items_by_position': items_by_position,
        }

    @classmethod
    def _published_item_stable_key_fallback(
        cls,
        *,
        published_stable_key_context: Dict[str, Any],
        section_order: float,
        section_name: str,
        item_type: str,
        item_label: str,
        item_order: float,
        indicator_bank_id: Optional[int],
    ) -> Optional[str]:
        section_key = (section_order, section_name or '')
        if indicator_bank_id is not None:
            indicator_key = (section_key[0], section_key[1], int(indicator_bank_id), item_order)
            return published_stable_key_context.get('items_by_indicator', {}).get(indicator_key)
        position_key = (
            section_key[0],
            section_key[1],
            item_order,
            (item_type or '').strip().lower(),
            str(item_label or '').strip(),
        )
        return published_stable_key_context.get('items_by_position', {}).get(position_key)

    @classmethod
    def _resolve_import_stable_key(
        cls,
        row_data: Dict[str, Any],
        row_idx: int,
        sheet_name: str,
        errors: List[str],
        published_fallback: Optional[str] = None,
    ) -> Optional[str]:
        raw = row_data.get('stable_key')
        if raw is not None and str(raw).strip() != '':
            key = normalize_stable_key(raw)
            if not key:
                errors.append(f"{sheet_name} row {row_idx}: Invalid stable_key '{raw}'")
                return None
            return key
        if published_fallback:
            return published_fallback
        return generate_stable_key()

    @classmethod
    def _validate_stable_key_duplicates_in_sheet(
        cls,
        rows: List[Tuple[int, Dict[str, Any]]],
        sheet_name: str,
        errors: List[str],
    ) -> None:
        seen: Dict[str, int] = {}
        for row_idx, row_data in rows:
            raw = row_data.get('stable_key')
            if raw is None or str(raw).strip() == '':
                continue
            key = normalize_stable_key(raw)
            if not key:
                continue
            if key in seen:
                errors.append(
                    f"{sheet_name} row {row_idx}: Duplicate stable_key '{key}' "
                    f"(also on row {seen[key]})"
                )
            else:
                seen[key] = row_idx

    @classmethod
    def _published_items_by_stable_key(cls, template: FormTemplate) -> Dict[str, FormItem]:
        if not template.published_version_id:
            return {}
        items = FormItem.query.filter_by(
            template_id=template.id,
            version_id=template.published_version_id,
        ).filter(FormItem.stable_key.isnot(None)).all()
        return {item.stable_key: item for item in items if item.stable_key}

    @classmethod
    def _check_stable_key_identity_mismatch(
        cls,
        *,
        template: FormTemplate,
        stable_key: str,
        item_type: str,
        indicator_bank_id: Optional[int],
        row_idx: int,
        warnings: List[str],
    ) -> None:
        published_by_key = cls._published_items_by_stable_key(template)
        published_item = published_by_key.get(stable_key)
        if not published_item:
            return
        pub_type = (published_item.item_type or '').strip().lower()
        cur_type = (item_type or '').strip().lower()
        pub_bank = published_item.indicator_bank_id
        if pub_type != cur_type or pub_bank != indicator_bank_id:
            warnings.append(
                f"Items row {row_idx}: stable_key {stable_key} matches published field "
                f"'{published_item.label}' (type={pub_type}, indicator_bank_id={pub_bank}) "
                f"but this row is type={cur_type}, indicator_bank_id={indicator_bank_id}. "
                f"Identity mismatch — verify this is intentional."
            )

    @classmethod
    def _import_sections(cls, sheet, template: FormTemplate, version: FormTemplateVersion,
                        page_id_map: Dict[int, int], section_id_map: Dict[int, int],
                        published_stable_key_context: Optional[Dict[str, Any]] = None) -> List[str]:
        """Import sections from sheet."""
        if published_stable_key_context is None:
            published_stable_key_context = cls._build_published_stable_key_context(template)
        current_app.logger.info("Starting sections import...")
        current_app.logger.info(f"Page ID mapping available: {len(page_id_map)} pages")
        errors = []

        # Read headers
        headers = [cell.value for cell in sheet[1]]
        current_app.logger.info(f"Sections sheet headers: {headers}")

        _, legacy_format, header_error = cls._resolve_sheet_headers('Sections', headers)
        if header_error:
            current_app.logger.error(header_error)
            errors.append(header_error)
            return errors

        # First pass: collect all section data
        current_app.logger.info("First pass: Collecting section data...")
        sections_data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            try:
                row_data = cls._row_dict_from_sheet(headers, row)
                sections_data.append((row_idx, row_data))
            except Exception as e:
                current_app.logger.error("Sections row %s: %s", row_idx, e)
                errors.append(f"Sections row {row_idx}: Validation error.")

        current_app.logger.info(f"Collected {len(sections_data)} section rows to process")

        cls._validate_stable_key_duplicates_in_sheet(sections_data, 'Sections', errors)
        if errors:
            return errors

        # Get existing sections for this version (for matching)
        existing_sections = FormSection.query.filter_by(
            template_id=template.id, version_id=version.id
        ).all()
        # Create lookup: (order, name) -> existing section (use order+name as match key)
        existing_sections_by_key = {(s.order, s.name): s for s in existing_sections}

        # Second pass: create or update sections and build ID map
        current_app.logger.info("Second pass: Creating/updating sections...")
        sections_created = 0
        sections_updated = 0

        for row_idx, row_data in sections_data:
            try:
                # Get export ID (sequential ID from Excel)
                export_id = int(row_data['id']) if row_data.get('id') else None

                if export_id is None:
                    error_msg = f"Sections row {row_idx}: Missing export ID"
                    current_app.logger.warning(error_msg)
                    errors.append(error_msg)
                    continue

                # Resolve page_id using export ID mapping
                page_id = None
                page_export_id = None
                if row_data.get('page_id'):
                    page_export_id = int(row_data['page_id']) if row_data.get('page_id') else None
                    if page_export_id and page_export_id in page_id_map:
                        page_id = page_id_map[page_export_id]
                        current_app.logger.debug(f"Section row {row_idx}: Mapped page export_id={page_export_id} -> db_id={page_id}")
                    elif page_export_id:
                        current_app.logger.warning(f"Section row {row_idx}: Page export_id={page_export_id} not found in page_id_map")

                section_name = row_data['name'] or ''
                section_order = float(row_data['order']) if row_data['order'] is not None else 0.0
                section_type = row_data.get('section_type') or 'standard'
                section_stable_key = cls._resolve_import_stable_key(
                    row_data,
                    row_idx,
                    'Sections',
                    errors,
                    published_fallback=published_stable_key_context.get('sections', {}).get(
                        (section_order, section_name)
                    ),
                )
                if section_stable_key is None:
                    continue
                current_app.logger.info(f"Processing section row {row_idx}: export_id={export_id}, name='{section_name}', order={section_order}, type={section_type}, page_export_id={page_export_id}")

                # Check if section already exists (match by order + name)
                section_key = (section_order, section_name)
                existing_section = existing_sections_by_key.get(section_key)

                if existing_section:
                    # Update existing section
                    current_app.logger.info(f"Updating existing section: db_id={existing_section.id}, order={section_order}, name='{section_name}'")
                    existing_section.page_id = page_id
                    existing_section.section_type = section_type
                    existing_section.max_dynamic_indicators = int(row_data['max_dynamic_indicators']) if row_data.get('max_dynamic_indicators') else None
                    existing_section.allowed_sectors = cls._parse_json(row_data.get('allowed_sectors'))
                    existing_section.indicator_filters = cls._parse_json(row_data.get('indicator_filters'))
                    existing_section.allow_data_not_available = bool(row_data.get('allow_data_not_available', False))
                    existing_section.allow_not_applicable = bool(row_data.get('allow_not_applicable', False))
                    existing_section.allowed_disaggregation_options = cls._parse_json(row_data.get('allowed_disaggregation_options'))
                    existing_section.data_entry_display_filters = cls._parse_json(row_data.get('data_entry_display_filters'))
                    existing_section.add_indicator_note = row_data.get('add_indicator_note')
                    existing_section.name_translations = cls._collect_translations_from_row(
                        row_data, 'name', legacy=legacy_format
                    )
                    existing_section.relevance_condition = row_data.get('relevance_condition')
                    existing_section.archived = bool(row_data.get('archived', False))
                    existing_section.stable_key = section_stable_key
                    # Note: parent_section_id will be updated in third pass
                    sections_updated += 1
                    section_id_map[export_id] = existing_section.id
                    current_app.logger.info(f"Section updated: export_id={export_id} -> db_id={existing_section.id}, name='{section_name}'")
                else:
                    # Create new section (parent_section_id will be resolved in third pass)
                    new_section = FormSection(
                        template_id=template.id,
                        version_id=version.id,
                        name=section_name,
                        order=section_order,
                        page_id=page_id,
                        parent_section_id=None,  # Will be set in third pass
                        section_type=section_type,
                        max_dynamic_indicators=int(row_data['max_dynamic_indicators']) if row_data.get('max_dynamic_indicators') else None,
                        allowed_sectors=cls._parse_json(row_data.get('allowed_sectors')),
                        indicator_filters=cls._parse_json(row_data.get('indicator_filters')),
                        allow_data_not_available=bool(row_data.get('allow_data_not_available', False)),
                        allow_not_applicable=bool(row_data.get('allow_not_applicable', False)),
                        allowed_disaggregation_options=cls._parse_json(row_data.get('allowed_disaggregation_options')),
                        data_entry_display_filters=cls._parse_json(row_data.get('data_entry_display_filters')),
                        add_indicator_note=row_data.get('add_indicator_note'),
                        name_translations=cls._collect_translations_from_row(
                            row_data, 'name', legacy=legacy_format
                        ),
                        relevance_condition=row_data.get('relevance_condition'),
                        archived=bool(row_data.get('archived', False)),
                        stable_key=section_stable_key,
                    )

                    db.session.add(new_section)
                    db.session.flush()

                    # Map export ID to new database ID
                    section_id_map[export_id] = new_section.id
                    sections_created += 1
                    current_app.logger.info(f"Section created: export_id={export_id} -> db_id={new_section.id}, name='{section_name}'")

            except Exception as e:
                current_app.logger.error("Sections row %s: %s", row_idx, e, exc_info=True)
                errors.append(f"Sections row {row_idx}: Validation error.")

        current_app.logger.info(f"Sections created/updated: {sections_created} created, {sections_updated} updated")

        current_app.logger.info(f"Sections created: {len(section_id_map)} sections")

        # Third pass: update parent_section_id references using export ID mapping
        current_app.logger.info("Third pass: Updating parent_section_id references...")
        parent_updates = 0
        for row_idx, row_data in sections_data:
            if row_data.get('parent_section_id'):
                try:
                    parent_export_id = int(row_data['parent_section_id']) if row_data.get('parent_section_id') else None
                    section_export_id = int(row_data['id']) if row_data.get('id') else None

                    if parent_export_id and section_export_id:
                        # Find new database IDs using export IDs
                        new_section_id = section_id_map.get(section_export_id)
                        new_parent_id = section_id_map.get(parent_export_id)

                        if new_section_id and new_parent_id:
                            # Update parent reference
                            section = FormSection.query.get(new_section_id)
                            if section:
                                section.parent_section_id = new_parent_id
                                parent_updates += 1
                                current_app.logger.debug(f"Updated parent: section export_id={section_export_id} (db_id={new_section_id}) -> parent export_id={parent_export_id} (db_id={new_parent_id})")
                        else:
                            current_app.logger.warning(f"Sections row {row_idx}: Could not resolve parent reference (section_export_id={section_export_id}, parent_export_id={parent_export_id})")
                except Exception as e:
                    current_app.logger.error("Sections row %s (parent update): %s", row_idx, e)
                    errors.append(f"Sections row {row_idx} (parent update): Validation error.")

        current_app.logger.info(f"Sections import complete: {len(sections_data)} rows processed, {len(section_id_map)} sections created, {parent_updates} parent relationships updated, {len(errors)} errors")
        return errors

    @classmethod
    def _import_items(cls, sheet, template: FormTemplate, version: FormTemplateVersion,
                     section_id_map: Dict[int, int], warnings: Optional[List[str]] = None,
                     published_stable_key_context: Optional[Dict[str, Any]] = None) -> List[str]:
        """Import items from sheet."""
        if warnings is None:
            warnings = []
        if published_stable_key_context is None:
            published_stable_key_context = cls._build_published_stable_key_context(template)
        current_app.logger.info("Starting items import...")
        current_app.logger.info(f"Section ID mapping available: {len(section_id_map)} sections")
        errors = []

        def _parse_int_like(value) -> Optional[int]:
            """Parse ints from Excel/JSON-ish values (e.g., 233, 233.0, '233')."""
            if value is None:
                return None
            try:
                # Excel frequently provides whole numbers as floats (e.g. 233.0)
                if isinstance(value, bool):
                    return None
                if isinstance(value, int):
                    return int(value)
                if isinstance(value, float):
                    if value.is_integer():
                        return int(value)
                    return None
                if isinstance(value, str):
                    s = value.strip()
                    if not s:
                        return None
                    if s.isdigit():
                        return int(s)
                    # Handle numeric strings like "233.0"
                    with suppress(Exception):
                        f = float(s)
                        if f.is_integer():
                            return int(f)
                return None
            except Exception as e:
                current_app.logger.debug("_parse_config_int failed: %s", e)
                return None

        def _ensure_import_issue(config: Any, *, code: str, message: str, meta: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
            """Attach an import issue to config['_import_issues'] and return the config dict."""
            cfg = config if isinstance(config, dict) else {}
            issues = cfg.get('_import_issues')
            if not isinstance(issues, list):
                issues = []
            issues.append({
                'code': code,
                'message': message,
                'meta': meta or {},
            })
            cfg['_import_issues'] = issues
            return cfg

        def _clear_import_issue_codes(config: Any, codes: List[str]) -> Any:
            """Remove matching import issue codes from config if present."""
            if not isinstance(config, dict):
                return config
            issues = config.get('_import_issues')
            if not isinstance(issues, list) or not issues:
                return config
            filtered = [i for i in issues if not (isinstance(i, dict) and i.get('code') in codes)]
            if filtered:
                config['_import_issues'] = filtered
            else:
                config.pop('_import_issues', None)
            return config

        # Read headers
        headers = [cell.value for cell in sheet[1]]
        current_app.logger.info(f"Items sheet headers: {headers}")

        _, legacy_format, header_error = cls._resolve_sheet_headers('Items', headers)
        if header_error:
            current_app.logger.error(header_error)
            errors.append(header_error)
            return errors

        current_app.logger.info("Headers validated. Processing item rows...")

        # Get existing items for this version (for matching)
        existing_items_lookup = cls._get_existing_items_lookup(template.id, version.id)
        ordered_existing_items = cls._get_items_for_version(template, version)
        export_id_to_existing = {idx + 1: item for idx, item in enumerate(ordered_existing_items)}
        cls._matrix_import_log(
            f"Starting items import for version_id={version.id}; "
            f"existing_items={len(existing_items_lookup)}, export_order_items={len(export_id_to_existing)}"
        )
        cls._log_matrix_items_in_version(template.id, version.id, stage='before-import')

        # Pre-scan rows for indicator_bank_id values so we can validate in one DB query.
        candidate_indicator_bank_ids: set[int] = set()
        rows_buffer: List[Tuple[int, Dict[str, Any]]] = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue
            row_data = cls._row_dict_from_sheet(headers, row)
            rows_buffer.append((row_idx, row_data))
            item_type = (row_data.get('item_type') or 'indicator')
            if item_type == 'indicator':
                raw_ib = row_data.get('indicator_bank_id')
                parsed_ib = _parse_int_like(raw_ib)
                if parsed_ib is not None:
                    candidate_indicator_bank_ids.add(parsed_ib)

        cls._validate_stable_key_duplicates_in_sheet(rows_buffer, 'Items', errors)
        if errors:
            return errors

        existing_indicator_bank_ids: set[int] = set()
        if candidate_indicator_bank_ids:
            existing_indicator_bank_ids = {
                int(x[0]) for x in db.session.query(IndicatorBank.id)
                .filter(IndicatorBank.id.in_(candidate_indicator_bank_ids))
                .all()
            }

        row_count = 0
        items_created = 0
        items_updated = 0

        # Map Items sheet export IDs -> actual FormItem objects (existing or newly created)
        item_export_to_obj: Dict[int, FormItem] = {}

        # Process buffered rows
        for row_idx, row_data in rows_buffer:
            row_count += 1
            try:
                # Get export ID
                export_id = int(row_data['id']) if row_data.get('id') else None

                # Resolve section_id using export ID mapping
                section_id = None
                section_export_id = None
                if row_data.get('section_id'):
                    section_export_id = int(row_data['section_id']) if row_data.get('section_id') else None
                    if section_export_id and section_export_id in section_id_map:
                        section_id = section_id_map[section_export_id]
                        current_app.logger.debug(f"Item row {row_idx}: Mapped section export_id={section_export_id} -> db_id={section_id}")
                    elif section_export_id:
                        current_app.logger.warning(f"Item row {row_idx}: Section export_id={section_export_id} not found in section_id_map")

                if not section_id:
                    error_msg = f"Items row {row_idx}: Could not resolve section_id (export_id: {section_export_id})"
                    current_app.logger.warning(error_msg)
                    errors.append(error_msg)
                    continue

                item_type = str(row_data.get('item_type') or 'indicator').strip().lower()
                item_label = str(row_data.get('label') or '').strip()
                item_order = float(row_data['order']) if row_data['order'] is not None else 0.0
                parsed_ib_for_mismatch = _parse_int_like(row_data.get('indicator_bank_id'))
                section = FormSection.query.get(section_id)
                section_order = float(section.order) if section and section.order is not None else 0.0
                section_name = (section.name or '') if section else ''
                item_stable_key = cls._resolve_import_stable_key(
                    row_data,
                    row_idx,
                    'Items',
                    errors,
                    published_fallback=cls._published_item_stable_key_fallback(
                        published_stable_key_context=published_stable_key_context,
                        section_order=section_order,
                        section_name=section_name,
                        item_type=item_type,
                        item_label=item_label,
                        item_order=item_order,
                        indicator_bank_id=parsed_ib_for_mismatch,
                    ),
                )
                if item_stable_key is None:
                    continue
                cls._check_stable_key_identity_mismatch(
                    template=template,
                    stable_key=item_stable_key,
                    item_type=item_type,
                    indicator_bank_id=parsed_ib_for_mismatch,
                    row_idx=row_idx,
                    warnings=warnings,
                )
                current_app.logger.info(f"Processing item row {row_idx}: export_id={export_id}, type={item_type}, label='{item_label[:50]}...', order={item_order}, section_export_id={section_export_id}")

                existing_item, match_method = cls._find_existing_item_for_import(
                    section_id=section_id,
                    item_type=item_type,
                    item_label=item_label,
                    item_order=item_order,
                    export_id=export_id,
                    export_id_to_existing=export_id_to_existing,
                    existing_items_lookup=existing_items_lookup,
                )

                if item_type == 'matrix':
                    target_db_id = existing_item.id if existing_item else None
                    parsed_preview = cls._parse_json(row_data.get('config'))
                    normalized_preview = cls._normalize_matrix_item_config(item_type, parsed_preview)
                    cls._matrix_import_log(
                        f"import row {row_idx} export_id={export_id} label={item_label!r} "
                        f"match={match_method or 'CREATE NEW'} target_db_id={target_db_id}; "
                        f"raw={cls._summarize_raw_config_cell(row_data.get('config'))}; "
                        f"normalized={cls._summarize_matrix_config_for_log(normalized_preview if isinstance(normalized_preview, dict) else {})}"
                    )

                if existing_item:
                    # Update existing item
                    current_app.logger.info(f"Updating existing item: db_id={existing_item.id}, section_id={section_id}, order={item_order}, type={item_type}")
                    existing_item.label = item_label
                    existing_item.order = item_order
                    existing_item.item_type = item_type
                    existing_item.relevance_condition = row_data.get('relevance_condition')
                    existing_item.archived = bool(row_data.get('archived', False))
                    cls._apply_item_config_from_import(existing_item, item_type, row_data.get('config'))

                    # Validate indicator_bank_id (only meaningful for indicator items)
                    if item_type == 'indicator':
                        raw_ib = row_data.get('indicator_bank_id')
                        parsed_ib = _parse_int_like(raw_ib)
                        if parsed_ib is None:
                            # Missing/invalid ID: import as NULL + flag issue in config
                            existing_item.indicator_bank_id = None
                            existing_item.config = _ensure_import_issue(
                                existing_item.config,
                                code='missing_or_invalid_indicator_bank_id',
                                message='Indicator is missing a valid Indicator Bank ID. Please select a valid indicator before deploying.',
                                meta={'raw_indicator_bank_id': raw_ib},
                            )
                        elif parsed_ib not in existing_indicator_bank_ids:
                            # References a non-existent IndicatorBank row: import as NULL + flag issue in config
                            existing_item.indicator_bank_id = None
                            existing_item.config = _ensure_import_issue(
                                existing_item.config,
                                code='missing_indicator_bank_reference',
                                message=f'Indicator references missing Indicator Bank ID {parsed_ib}. Please select a valid indicator before deploying.',
                                meta={'raw_indicator_bank_id': raw_ib, 'parsed_indicator_bank_id': parsed_ib},
                            )
                        else:
                            # Valid reference: clear any prior import issue flags and set the FK
                            existing_item.indicator_bank_id = parsed_ib
                            existing_item.config = _clear_import_issue_codes(
                                existing_item.config,
                                ['missing_or_invalid_indicator_bank_id', 'missing_indicator_bank_reference'],
                            )
                    else:
                        # Non-indicator: clear any stale indicator import issues
                        existing_item.config = _clear_import_issue_codes(
                            existing_item.config,
                            ['missing_or_invalid_indicator_bank_id', 'missing_indicator_bank_reference'],
                        )
                        existing_item.indicator_bank_id = _parse_int_like(row_data.get('indicator_bank_id'))
                    existing_item.type = row_data.get('type')
                    existing_item.unit = row_data.get('unit')
                    existing_item.validation_condition = row_data.get('validation_condition')
                    existing_item.validation_message = row_data.get('validation_message')
                    existing_item.definition = row_data.get('definition')
                    existing_item.options_json = cls._parse_json(row_data.get('options_json'))
                    existing_item.lookup_list_id = str(row_data['lookup_list_id']) if row_data.get('lookup_list_id') else None
                    existing_item.list_display_column = row_data.get('list_display_column')
                    existing_item.list_filters_json = cls._parse_json(row_data.get('list_filters_json'))
                    cls._apply_item_translations_from_row(existing_item, row_data, legacy=legacy_format)
                    existing_item.options_translations = cls._parse_json(row_data.get('options_translations'))
                    existing_item.description = row_data.get('description')
                    existing_item.stable_key = item_stable_key
                    if item_type == 'matrix':
                        cls._sync_matrix_item_fields_from_config(existing_item)
                        from sqlalchemy.orm.attributes import flag_modified
                        flag_modified(existing_item, 'config')
                        cls._matrix_import_log(
                            f"import updated matrix db_id={existing_item.id} label={item_label!r} "
                            f"via {match_method}: "
                            f"{cls._summarize_matrix_config_for_log(existing_item.config)}"
                        )
                    items_updated += 1
                    current_app.logger.info(f"Item updated: db_id={existing_item.id}, label='{item_label[:50]}...'")
                    if export_id is not None:
                        item_export_to_obj[export_id] = existing_item
                else:
                    # Parse and validate indicator bank reference up-front (so we never violate FK constraints)
                    parsed_indicator_bank_id: Optional[int] = None
                    parsed_config = cls._parse_json(row_data.get('config'))
                    config_payload = cls._normalize_matrix_item_config(item_type, parsed_config)
                    if item_type == 'matrix' and isinstance(config_payload, dict):
                        config_payload = cls._clean_imported_matrix_config(config_payload)
                    if item_type == 'matrix':
                        cls._matrix_import_log(
                            f"row {row_idx} CREATE NEW label={item_label!r}: "
                            f"parsed_config={cls._summarize_matrix_config_for_log(config_payload if isinstance(config_payload, dict) else {})}"
                        )
                    if item_type == 'indicator':
                        raw_ib = row_data.get('indicator_bank_id')
                        parsed_ib = _parse_int_like(raw_ib)
                        if parsed_ib is None:
                            parsed_indicator_bank_id = None
                            config_payload = _ensure_import_issue(
                                config_payload,
                                code='missing_or_invalid_indicator_bank_id',
                                message='Indicator is missing a valid Indicator Bank ID. Please select a valid indicator before deploying.',
                                meta={'raw_indicator_bank_id': raw_ib},
                            )
                        elif parsed_ib not in existing_indicator_bank_ids:
                            parsed_indicator_bank_id = None
                            config_payload = _ensure_import_issue(
                                config_payload,
                                code='missing_indicator_bank_reference',
                                message=f'Indicator references missing Indicator Bank ID {parsed_ib}. Please select a valid indicator before deploying.',
                                meta={'raw_indicator_bank_id': raw_ib, 'parsed_indicator_bank_id': parsed_ib},
                            )
                        else:
                            parsed_indicator_bank_id = parsed_ib
                            config_payload = _clear_import_issue_codes(
                                config_payload,
                                ['missing_or_invalid_indicator_bank_id', 'missing_indicator_bank_reference'],
                            )
                    else:
                        parsed_indicator_bank_id = _parse_int_like(row_data.get('indicator_bank_id'))

                    # Create new item
                    new_item = FormItem(
                        section_id=section_id,
                        template_id=template.id,
                        version_id=version.id,
                        item_type=item_type,
                        stable_key=item_stable_key,
                        label=item_label,
                        order=item_order,
                        relevance_condition=row_data.get('relevance_condition'),
                        archived=bool(row_data.get('archived', False)),
                        config=config_payload,
                        indicator_bank_id=parsed_indicator_bank_id,
                        type=row_data.get('type'),
                        unit=row_data.get('unit'),
                        validation_condition=row_data.get('validation_condition'),
                        validation_message=row_data.get('validation_message'),
                        definition=row_data.get('definition'),
                        options_json=cls._parse_json(row_data.get('options_json')),
                        lookup_list_id=str(row_data['lookup_list_id']) if row_data.get('lookup_list_id') else None,
                        list_display_column=row_data.get('list_display_column'),
                        list_filters_json=cls._parse_json(row_data.get('list_filters_json')),
                        options_translations=cls._parse_json(row_data.get('options_translations')),
                        description=row_data.get('description')
                    )
                    cls._apply_item_translations_from_row(new_item, row_data, legacy=legacy_format)

                    if item_type == 'matrix':
                        cls._sync_matrix_item_fields_from_config(new_item)

                    db.session.add(new_item)
                    items_created += 1
                    if export_id is not None:
                        item_export_to_obj[export_id] = new_item

                    if items_created % 10 == 0:
                        current_app.logger.debug(f"Items progress: {items_created} items created so far...")

            except Exception as e:
                current_app.logger.error("Items row %s: %s", row_idx, e, exc_info=True)
                errors.append(f"Items row {row_idx}: Validation error.")

        current_app.logger.info(f"Items import complete: {row_count} rows processed, {items_created} created, {items_updated} updated, {len(errors)} errors")
        cls._log_matrix_items_in_version(template.id, version.id, stage='after-import-pre-commit')

        # Rewrite rule JSON item IDs from export IDs -> new DB IDs so relevance/validation rules survive imports.
        try:
            db.session.flush()  # ensure new items have IDs
            item_id_map: Dict[int, int] = {}
            for exp_id, obj in item_export_to_obj.items():
                if obj is not None and getattr(obj, 'id', None) is not None:
                    item_id_map[int(exp_id)] = int(obj.id)

            if item_id_map:
                rewritten_items = 0
                for obj in item_export_to_obj.values():
                    if not obj:
                        continue
                    before_rel = obj.relevance_condition
                    before_val = getattr(obj, 'validation_condition', None)
                    obj.relevance_condition = cls._rewrite_rule_json_item_ids(before_rel, item_id_map)
                    if hasattr(obj, 'validation_condition'):
                        obj.validation_condition = cls._rewrite_rule_json_item_ids(before_val, item_id_map)
                    if before_rel != obj.relevance_condition or before_val != getattr(obj, 'validation_condition', None):
                        rewritten_items += 1

                rewritten_sections = 0
                sections = FormSection.query.filter_by(template_id=template.id, version_id=version.id).all()
                for s in sections:
                    before = s.relevance_condition
                    s.relevance_condition = cls._rewrite_rule_json_item_ids(before, item_id_map)
                    if before != s.relevance_condition:
                        rewritten_sections += 1

                current_app.logger.info(
                    f"Rewrote rule item IDs using item_id_map (size={len(item_id_map)}): "
                    f"items_updated={rewritten_items}, sections_updated={rewritten_sections}"
                )
        except Exception as e:
            # Don't fail the import for rule rewrite issues; log and continue.
            current_app.logger.warning(f"Could not rewrite rule item IDs after item import: {e}", exc_info=True)

        return errors

    @classmethod
    def _count_deletion_impact(cls, template_id: int, version_id: Optional[int]) -> Dict[str, Any]:
        """Return counts of submission-data rows that would be removed by a structure clear.

        Used by the preflight endpoint so the UI can warn the admin before the import
        actually destroys any data.  No data is modified.
        """
        if not version_id:
            return {'has_data': False, 'counts': {}}

        section_ids: List[int] = [
            row[0]
            for row in db.session.query(FormSection.id).filter_by(
                template_id=template_id, version_id=version_id
            ).all()
        ]

        # Join form_item so we count only data for this version without a large IN subquery.
        form_data: int = db.session.query(
            db.func.count(FormData.id)
        ).join(
            FormItem, FormData.form_item_id == FormItem.id
        ).filter(
            FormItem.template_id == template_id,
            FormItem.version_id == version_id,
        ).scalar() or 0

        repeat_instances: int = 0
        repeat_data: int = 0
        dynamic_indicators: int = 0
        dynamic_contexts: int = 0

        if section_ids:
            repeat_instances = db.session.query(
                db.func.count(RepeatGroupInstance.id)
            ).filter(RepeatGroupInstance.section_id.in_(section_ids)).scalar() or 0

            if repeat_instances:
                repeat_data = db.session.query(
                    db.func.count(RepeatGroupData.id)
                ).join(
                    RepeatGroupInstance,
                    RepeatGroupData.repeat_instance_id == RepeatGroupInstance.id,
                ).filter(
                    RepeatGroupInstance.section_id.in_(section_ids)
                ).scalar() or 0

            dynamic_indicators = db.session.query(
                db.func.count(DynamicIndicatorData.id)
            ).filter(DynamicIndicatorData.section_id.in_(section_ids)).scalar() or 0

            dynamic_contexts = db.session.query(
                db.func.count(DynamicSectionContext.id)
            ).filter(DynamicSectionContext.section_id.in_(section_ids)).scalar() or 0

        submitted_documents: int = db.session.query(
            db.func.count(SubmittedDocument.id)
        ).join(
            FormItem, SubmittedDocument.form_item_id == FormItem.id
        ).filter(
            FormItem.template_id == template_id,
            FormItem.version_id == version_id,
        ).scalar() or 0

        total = form_data + repeat_instances + repeat_data + dynamic_indicators + dynamic_contexts + submitted_documents
        return {
            'has_data': total > 0,
            'counts': {
                'form_data': form_data,
                'repeat_instances': repeat_instances,
                'repeat_data': repeat_data,
                'dynamic_indicators': dynamic_indicators,
                'dynamic_contexts': dynamic_contexts,
                'submitted_documents': submitted_documents,
            },
        }

    @classmethod
    def _clear_version_structure(cls, template_id: int, version_id: int) -> Dict[str, int]:
        """Delete all pages, sections, and items for a version (keeps the version row itself).

        Bulk .delete(synchronize_session=False) bypasses SQLAlchemy ORM cascade rules, so
        child records that have FK constraints pointing to form_section / form_item must be
        removed explicitly before the structural rows are deleted.
        """
        version = FormTemplateVersion.query.filter_by(id=version_id, template_id=template_id).first()
        if version and version.status == 'published':
            impact = cls._count_deletion_impact(template_id, version_id)
            if impact.get('has_data'):
                raise ValueError(
                    'Cannot clear structure on a published version that has submission data. '
                    'Import into a draft version and deploy it instead.'
                )

        current_app.logger.info(
            f"Clearing existing structure for template_id={template_id}, version_id={version_id} before Excel import"
        )

        # Collect the IDs of sections and items that are about to be removed so we can
        # clean up submission-data rows that hold FK references to them.
        section_ids: List[int] = [
            row[0]
            for row in db.session.query(FormSection.id).filter_by(
                template_id=template_id, version_id=version_id
            ).all()
        ]
        item_ids: List[int] = [
            row[0]
            for row in db.session.query(FormItem.id).filter_by(
                template_id=template_id, version_id=version_id
            ).all()
        ]

        if section_ids:
            # repeat_group_data.repeat_instance_id → repeat_group_instance.id has no DB CASCADE
            instance_ids_sq = db.session.query(RepeatGroupInstance.id).filter(
                RepeatGroupInstance.section_id.in_(section_ids)
            ).subquery()
            rgd_deleted = RepeatGroupData.query.filter(
                RepeatGroupData.repeat_instance_id.in_(instance_ids_sq)
            ).delete(synchronize_session=False)

            # repeat_group_instance.section_id → form_section.id has no DB CASCADE
            rgi_deleted = RepeatGroupInstance.query.filter(
                RepeatGroupInstance.section_id.in_(section_ids)
            ).delete(synchronize_session=False)

            # dynamic_indicator_data.section_id → form_section.id has no DB CASCADE
            did_deleted = DynamicIndicatorData.query.filter(
                DynamicIndicatorData.section_id.in_(section_ids)
            ).delete(synchronize_session=False)

            # dynamic_section_context.section_id → form_section.id has no DB CASCADE
            dsc_deleted = DynamicSectionContext.query.filter(
                DynamicSectionContext.section_id.in_(section_ids)
            ).delete(synchronize_session=False)

            if any([rgd_deleted, rgi_deleted, did_deleted, dsc_deleted]):
                current_app.logger.info(
                    f"Cleared submission data tied to version structure: "
                    f"{rgi_deleted} repeat_group_instance(s), {rgd_deleted} repeat_group_data row(s), "
                    f"{did_deleted} dynamic_indicator_data row(s), {dsc_deleted} dynamic_section_context row(s)"
                )

        if item_ids:
            # form_data.form_item_id → form_item.id has no DB CASCADE
            fd_deleted = FormData.query.filter(
                FormData.form_item_id.in_(item_ids)
            ).delete(synchronize_session=False)

            # Catch any repeat_group_data rows referencing these items that were not already
            # removed via the section/instance path above (edge-case guard).
            rgd_by_item = RepeatGroupData.query.filter(
                RepeatGroupData.form_item_id.in_(item_ids)
            ).delete(synchronize_session=False)

            if fd_deleted or rgd_by_item:
                current_app.logger.info(
                    f"Cleared {fd_deleted} form_data row(s) and "
                    f"{rgd_by_item} additional repeat_group_data row(s) tied to version items"
                )

        items_deleted = FormItem.query.filter_by(
            template_id=template_id, version_id=version_id
        ).delete(synchronize_session=False)
        sections_deleted = FormSection.query.filter_by(
            template_id=template_id, version_id=version_id
        ).delete(synchronize_session=False)
        pages_deleted = FormPage.query.filter_by(
            template_id=template_id, version_id=version_id
        ).delete(synchronize_session=False)
        db.session.flush()
        current_app.logger.info(
            f"Cleared version structure: {pages_deleted} pages, {sections_deleted} sections, {items_deleted} items"
        )
        return {'pages': pages_deleted, 'sections': sections_deleted, 'items': items_deleted}

    @classmethod
    def _clone_template_structure(cls, template_id: int, source_version_id: int, target_version_id: int) -> None:
        """Clone pages, sections, and items from source_version_id to target_version_id preserving order."""
        current_app.logger.info(f"Cloning template structure from version {source_version_id} to {target_version_id}")
        # Maps for old->new IDs
        page_id_map = {}
        section_id_map = {}

        # Clone pages
        src_pages = FormPage.query.filter_by(template_id=template_id, version_id=source_version_id).order_by(FormPage.order).all()
        current_app.logger.info(f"Cloning {len(src_pages)} pages")
        for p in src_pages:
            new_p = FormPage(
                template_id=template_id,
                version_id=target_version_id,
                name=p.name,
                order=p.order,
                name_translations=p.name_translations
            )
            db.session.add(new_p)
            db.session.flush()
            page_id_map[p.id] = new_p.id

        # Clone sections (first pass: main sections)
        src_sections = FormSection.query.filter_by(template_id=template_id, version_id=source_version_id).order_by(FormSection.order).all()
        current_app.logger.info(f"Cloning {len(src_sections)} sections")
        # Create all sections without parent refs first
        for s in src_sections:
            new_s = FormSection(
                template_id=template_id,
                version_id=target_version_id,
                name=s.name,
                order=s.order,
                parent_section_id=None,  # set later
                page_id=page_id_map.get(s.page_id) if s.page_id else None,
                section_type=s.section_type,
                max_dynamic_indicators=s.max_dynamic_indicators,
                allowed_sectors=s.allowed_sectors,
                indicator_filters=s.indicator_filters,
                allow_data_not_available=s.allow_data_not_available,
                allow_not_applicable=s.allow_not_applicable,
                allowed_disaggregation_options=s.allowed_disaggregation_options,
                data_entry_display_filters=s.data_entry_display_filters,
                add_indicator_note=s.add_indicator_note,
                name_translations=s.name_translations,
                relevance_condition=s.relevance_condition,
                archived=s.archived
            )
            db.session.add(new_s)
            db.session.flush()
            section_id_map[s.id] = new_s.id

        # Second pass: set parent_section_id now that all new IDs exist
        parent_updates = 0
        for s in src_sections:
            if s.parent_section_id:
                new_id = section_id_map[s.id]
                new_parent_id = section_id_map.get(s.parent_section_id)
                if new_parent_id:
                    FormSection.query.filter_by(id=new_id).update({'parent_section_id': new_parent_id})
                    parent_updates += 1
        current_app.logger.info(f"Cloned {len(section_id_map)} sections, updated {parent_updates} parent relationships")

        # Clone items
        src_items = FormItem.query.join(FormSection, FormItem.section_id == FormSection.id).\
            filter(FormItem.template_id == template_id, FormItem.version_id == source_version_id).\
            order_by(FormItem.order).all()
        current_app.logger.info(f"Cloning {len(src_items)} items")
        items_cloned = 0
        for it in src_items:
            # Deep copy config to avoid cross-version mutations
            _new_config = None
            try:
                _new_config = json.loads(json.dumps(it.config)) if it.config is not None else None
            except Exception as e:
                current_app.logger.debug("JSON roundtrip for item config failed: %s", e)
                try:
                    import copy as _copy
                    _new_config = _copy.deepcopy(it.config) if it.config is not None else None
                except Exception as e2:
                    current_app.logger.debug("deepcopy item config failed: %s", e2)
                    _new_config = it.config.copy() if isinstance(it.config, dict) else it.config

            new_it = FormItem(
                template_id=template_id,
                version_id=target_version_id,
                section_id=section_id_map.get(it.section_id),
                item_type=it.item_type,
                label=it.label,
                order=it.order,
                relevance_condition=it.relevance_condition,
                archived=it.archived,
                config=_new_config,
                indicator_bank_id=it.indicator_bank_id,
                type=it.type,
                unit=it.unit,
                validation_condition=it.validation_condition,
                validation_message=it.validation_message,
                definition=it.definition,
                options_json=it.options_json,
                lookup_list_id=getattr(it, 'lookup_list_id', None),
                list_display_column=getattr(it, 'list_display_column', None),
                list_filters_json=getattr(it, 'list_filters_json', None),
                label_translations=getattr(it, 'label_translations', None),
                definition_translations=getattr(it, 'definition_translations', None),
                options_translations=getattr(it, 'options_translations', None),
                description_translations=getattr(it, 'description_translations', None),
                description=getattr(it, 'description', None)
            )
            db.session.add(new_it)
            items_cloned += 1

        current_app.logger.info(f"Successfully cloned structure: {len(page_id_map)} pages, {len(section_id_map)} sections, {items_cloned} items")

    @classmethod
    def _get_existing_items_lookup(cls, template_id: int, version_id: int) -> Dict[Tuple, FormItem]:
        """
        Create a lookup dictionary for existing items.
        Key: (section_id, order, item_type, label) - combination to uniquely identify an item
        Value: FormItem object
        """
        existing_items = FormItem.query.join(FormSection).filter(
            FormItem.template_id == template_id,
            FormSection.version_id == version_id
        ).all()

        # Create lookup: (section_id, order, item_type, label) -> item
        # Using label as part of key since it's more stable than just order
        items_lookup = {}
        for item in existing_items:
            key = (
                item.section_id,
                float(item.order) if item.order is not None else 0.0,
                str(item.item_type or 'indicator').strip().lower(),
                str(item.label or '').strip(),
            )
            items_lookup[key] = item

        return items_lookup

    @classmethod
    def _parse_json(cls, value) -> Optional[Any]:
        """Parse JSON string to object; return None only when empty or invalid."""
        if value is None or value == '' or value == 'None':
            return None
        if isinstance(value, (dict, list)):
            return value
        try:
            text = str(value).strip()
            if text.startswith('\ufeff'):
                text = text[1:]
            parsed = json.loads(text)
            if isinstance(parsed, (dict, list)):
                return parsed
            return None
        except (json.JSONDecodeError, TypeError):
            return None
