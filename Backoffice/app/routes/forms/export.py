"""PDF / Excel export and import routes for forms.

.. deprecated::
   The Excel export/import endpoints in this module
   (``export_excel`` / ``import_excel``) are superseded by the dedicated
   ``excel_bp`` blueprint in ``app/routes/excel.py``
   (``/excel/assignment/<aes_id>/export`` and ``/excel/assignment/<aes_id>/import``).
   New callers should use the ``excel`` blueprint routes.  These legacy endpoints
   are kept for backward compatibility only and will be removed in a future release.
"""
from __future__ import annotations

from contextlib import suppress
import io
import json
import os
import re

from flask import current_app, flash, redirect, render_template, request, send_file, url_for
from flask_babel import _
from flask_login import current_user, login_required
from sqlalchemy.orm import joinedload

from app import get_locale
from app.models import (
    db, AssignedForm, AssignmentEntityStatus, Country, DynamicIndicatorData,
    FormData, FormItem, FormPage, FormSection, QuestionType,
    SubmittedDocument,
)
from app.services.imports.excel_service import ExcelService
from app.services.forms.processing_service import slugify_age_group
from app.utils.datetime_helpers import utcnow
from app.utils.form_localization import (
    get_localized_country_name,
    get_localized_indicator_name,
    get_localized_page_name,
    get_localized_section_name,
    get_localized_template_name,
    get_translation_key,
)
from app.utils.transactions import request_transaction_rollback
from config import Config

from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def _first_localized_translation(translations_map, translation_key):
    """Return the first non-empty translation for preferred language keys."""
    if not isinstance(translations_map, dict):
        return None
    for key in (translation_key, 'en', 'EN'):
        if not key:
            continue
        for candidate_key in (key, str(key).lower(), str(key).upper()):
            value = translations_map.get(candidate_key)
            if isinstance(value, str) and value.strip():
                return value.strip()
    return None


def _normalize_note_html_for_pdf(html):
    """Tighten note/blank HTML for PDF so spacing matches the entry form."""
    if not html or not isinstance(html, str):
        return html or ''
    cleaned = re.sub(
        r'<p>\s*(?:<br\s*/?>|&nbsp;|\u00a0|\s)*\s*</p>',
        '',
        html,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r'(\s*<br\s*/?>\s*){2,}', '<br>', cleaned, flags=re.IGNORECASE)
    return cleaned.strip()


def _resolve_note_body(form_item, translation_key, resolved_variables, variable_configs):
    """Resolve localized note/blank body HTML (definition), with variable substitution."""
    from app.services.forms.variable_resolution_service import VariableResolutionService

    body = _first_localized_translation(
        getattr(form_item, 'definition_translations', None),
        translation_key,
    )
    if not body:
        raw_definition = getattr(form_item, 'definition', None)
        body = raw_definition.strip() if isinstance(raw_definition, str) else ''
    if body and resolved_variables:
        with suppress(Exception):
            body = VariableResolutionService.replace_variables_in_text(
                body,
                resolved_variables,
                variable_configs,
            )
    return _normalize_note_html_for_pdf(body or '')


def _wrap_export_value_with_flags(value, *, data_not_available=False, not_applicable=False):
    """Attach availability flags to exported field values when set on FormData."""
    if not data_not_available and not not_applicable:
        return value
    if isinstance(value, dict):
        wrapped = dict(value)
        wrapped['data_not_available'] = bool(data_not_available)
        wrapped['not_applicable'] = bool(not_applicable)
        return wrapped
    return {
        'value': value,
        'data_not_available': bool(data_not_available),
        'not_applicable': bool(not_applicable),
    }


_FIELD_VALUE_KEY_RE = re.compile(r'^field_value\[(\d+)\](.*)$')
_DYNAMIC_VALUE_KEY_RE = re.compile(r'^field_value\[dynamic_(\d+)\](.*)$')
_AVAILABILITY_KEY_RES = (
    (re.compile(r'^matrix_(\d+)_data_not_available$'), 'data_not_available'),
    (re.compile(r'^indicator_(\d+)_data_not_available$'), 'data_not_available'),
    (re.compile(r'^question_(\d+)_data_not_available$'), 'data_not_available'),
    (re.compile(r'^matrix_(\d+)_not_applicable$'), 'not_applicable'),
    (re.compile(r'^indicator_(\d+)_not_applicable$'), 'not_applicable'),
    (re.compile(r'^question_(\d+)_not_applicable$'), 'not_applicable'),
    (re.compile(r'^dynamic_(\d+)_data_not_available$'), 'data_not_available'),
    (re.compile(r'^dynamic_(\d+)_not_applicable$'), 'not_applicable'),
)


def _normalize_matrix_cell_value(value):
    """Normalize a matrix cell value for prefilled/carry-forward comparison."""
    if value is None or value == '':
        return ''
    if isinstance(value, dict):
        if value.get('modified') not in (None, ''):
            return _normalize_matrix_cell_value(value.get('modified'))
        if value.get('original') not in (None, ''):
            return _normalize_matrix_cell_value(value.get('original'))
        return ''
    if isinstance(value, bool):
        return '1' if value else '0'
    text = str(value).strip()
    if text == 'true':
        return '1'
    if text == 'false':
        return '0'
    return text


def matrix_pdf_layout_strategy(col_count):
    """Choose PDF layout for wide matrices: portrait squeeze vs landscape page.

    Returns an empty string for normal (≤7 columns) matrices.
    """
    try:
        count = int(col_count or 0)
    except (TypeError, ValueError):
        count = 0
    if count <= 7:
        return ''
    if count <= 13:
        return 'portrait-compact'
    if count <= 16:
        return 'landscape'
    return 'landscape-scale'


def matrix_portrait_column_widths_mm(
    col_count,
    columns,
    *,
    show_row_totals=True,
    table_width_mm=190,
):
    """Return ``(name, width_mm)`` pairs for a portrait-compact matrix ``colgroup``."""
    try:
        count = int(col_count or 0)
    except (TypeError, ValueError):
        count = 0

    row_label_mm = 26
    if count >= 12:
        row_label_mm = 20
    elif count >= 10:
        row_label_mm = 22

    tick_mm = 7.0
    min_number_mm = 8.5

    specs = []
    for col in columns or []:
        if isinstance(col, dict):
            col_type = (col.get('type') or 'number').strip().lower()
            name = col.get('name') or ''
        else:
            col_type = 'number'
            name = str(col)
        specs.append({'name': name, 'is_tick': col_type == 'tick'})

    if show_row_totals:
        specs.append({'name': '__total__', 'is_tick': False})

    tick_total_mm = tick_mm * sum(1 for spec in specs if spec['is_tick'])
    number_count = sum(1 for spec in specs if not spec['is_tick'])
    remaining_mm = max(table_width_mm - row_label_mm - tick_total_mm, min_number_mm * max(number_count, 1))
    number_mm = remaining_mm / number_count if number_count else min_number_mm
    number_mm = max(min_number_mm, number_mm)

    widths = [('__row__', row_label_mm)]
    for spec in specs:
        if spec['is_tick']:
            widths.append((spec['name'], tick_mm))
        else:
            widths.append((spec['name'], round(number_mm, 1)))
    return widths


def matrix_cell_is_prefilled_highlight(
    cell_key,
    display_cell,
    *,
    is_prefilled=False,
    is_imputed=False,
    carry_forward_ref=None,
    is_variable_readonly=False,
):
    """Return True when a matrix cell should be highlighted as prefilled in PDF export."""
    if is_imputed or is_variable_readonly:
        return False
    if not is_prefilled and not (isinstance(carry_forward_ref, dict) and carry_forward_ref):
        return False

    normalized = _normalize_matrix_cell_value(display_cell)
    if not normalized:
        return False

    if isinstance(carry_forward_ref, dict) and carry_forward_ref:
        if cell_key not in carry_forward_ref:
            return False
        return normalized == _normalize_matrix_cell_value(carry_forward_ref.get(cell_key))

    return True


def _apply_availability_flag(export_data, export_key, flag_name):
    """Merge data-not-available / not-applicable flags into an export field payload."""
    current = export_data.get(export_key)
    if current is None:
        export_data[export_key] = {flag_name: True}
    elif isinstance(current, dict):
        merged = dict(current)
        merged[flag_name] = True
        export_data[export_key] = merged
    else:
        export_data[export_key] = {
            'value': current,
            flag_name: True,
        }


def _convert_entry_data_to_export_format(entry_data):
    """Map entry-form existing_data keys to PDF export template keys."""
    export_data = {}

    for key, value in (entry_data or {}).items():
        dynamic_match = _DYNAMIC_VALUE_KEY_RE.match(key)
        if dynamic_match:
            export_data[f"form_item_dynamic_{dynamic_match.group(1)}{dynamic_match.group(2)}"] = value
            continue

        field_match = _FIELD_VALUE_KEY_RE.match(key)
        if field_match:
            export_data[f"form_item_{field_match.group(1)}{field_match.group(2)}"] = value
            continue

        matched = False
        for pattern, flag_name in _AVAILABILITY_KEY_RES:
            avail_match = pattern.match(key)
            if not avail_match:
                continue
            item_id = avail_match.group(1)
            export_key = (
                f"form_item_dynamic_{item_id}"
                if key.startswith('dynamic_')
                else f"form_item_{item_id}"
            )
            _apply_availability_flag(export_data, export_key, flag_name)
            matched = True
            break
        if matched:
            continue

    return export_data


def _load_existing_data_for_pdf_export(assignment_entity_status, form_template):
    """Load assignment data with prefilled/imputed/carry-forward metadata for PDF export."""
    from app.routes.forms.helpers import _load_existing_data_for_assignment

    entry_data = _load_existing_data_for_assignment(assignment_entity_status, form_template)

    cf_items = []
    for section_model in form_template.sections.order_by(FormSection.order).all():
        for form_item in FormItem.query.filter_by(section_id=section_model.id, archived=False).all():
            config = form_item.config if isinstance(form_item.config, dict) else {}
            if config.get('carry_forward'):
                cf_items.append(form_item)

    if cf_items:
        from app.services.forms.carry_forward_service import CarryForwardService

        try:
            cf_results = CarryForwardService.resolve_for_aes(assignment_entity_status, cf_items)
            for item_id, result in cf_results.items():
                field_key = f'field_value[{item_id}]'
                if field_key in entry_data:
                    continue
                entry_data[field_key] = (
                    result['disagg_data'] if result.get('is_matrix') else result.get('value')
                )
                entry_data[f'{field_key}_is_prefilled'] = True
                entry_data[f'{field_key}_is_carry_forward'] = True

            cf_refs = CarryForwardService.resolve_references_for_aes(
                assignment_entity_status, cf_items
            )
            for item_id, result in cf_refs.items():
                ref_data = (
                    result['disagg_data'] if result.get('is_matrix') else result.get('value')
                )
                if ref_data is None:
                    continue
                entry_data[f'field_value[{item_id}]_carry_forward_ref'] = ref_data
                entry_data[f'field_value[{item_id}]_is_carry_forward'] = True
        except Exception as e:
            current_app.logger.warning(
                "Carry-forward resolution failed for PDF export (AES %s): %s",
                assignment_entity_status.id,
                e,
            )

    return _convert_entry_data_to_export_format(entry_data)


def _make_assignment_pdf_download_name(entity_name, assignment_name):
    """Build `{entity} - {assignment}.pdf`, keeping names intact aside from unsafe path chars."""
    entity = (entity_name or '').strip()
    assignment = (assignment_name or '').strip() or 'Assignment'
    base = f"{entity} - {assignment}" if entity else assignment
    cleaned = re.sub(r'[\\/:*?"<>|\x00-\x1f]', '', base)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip().rstrip('.')
    cleaned = cleaned[:200] or 'Assignment'
    return f"{cleaned}.pdf"


def _matrix_row_entity_ids(field_dict):
    """Collect numeric row entity IDs from a matrix field export dict."""
    row_ids = []
    seen = set()
    for r in field_dict.get('matrix_rows') or []:
        rid = r.get('text') if isinstance(r, dict) else r
        if rid is None:
            continue
        try:
            rid_int = int(rid)
        except (ValueError, TypeError):
            continue
        if rid_int not in seen:
            seen.add(rid_int)
            row_ids.append(rid_int)
    if not row_ids:
        for rid in (field_dict.get('matrix_row_labels') or {}).keys():
            try:
                rid_int = int(rid)
            except (ValueError, TypeError):
                continue
            if rid_int not in seen:
                seen.add(rid_int)
                row_ids.append(rid_int)
    return row_ids


def _matrix_has_variable_columns(columns):
    for col in columns or []:
        if isinstance(col, dict) and (col.get('is_variable') or col.get('type') == 'variable'):
            return True
    return False


def _merge_matrix_variable_values(field_dict, matrix_data, template_version, assignment_entity_status):
    """Fill variable matrix cells from template variable resolution (not stored in disagg_data)."""
    from app.services.forms.variable_resolution_service import VariableResolutionService

    columns = field_dict.get('matrix_columns') or []
    if not _matrix_has_variable_columns(columns) or not template_version:
        return matrix_data

    row_ids = _matrix_row_entity_ids(field_dict)
    if not row_ids:
        return matrix_data

    try:
        batch = VariableResolutionService.resolve_variables_batch(
            template_version,
            assignment_entity_status,
            row_ids,
        ) or {}
    except Exception as e:
        current_app.logger.warning(
            "Failed to resolve matrix variable columns for PDF export (field %s): %s",
            field_dict.get('id'),
            e,
            exc_info=True,
        )
        return matrix_data

    enriched = dict(matrix_data) if isinstance(matrix_data, dict) else {}
    for row_id, var_map in batch.items():
        if not isinstance(var_map, dict):
            continue
        row_key = str(row_id)
        for col in columns:
            if not isinstance(col, dict):
                continue
            if not (col.get('is_variable') or col.get('type') == 'variable'):
                continue
            col_name = col.get('name')
            var_name = col.get('variable') or col.get('variable_name')
            if not col_name or not var_name:
                continue
            cell_key = f"{row_key}_{col_name}"
            existing = enriched.get(cell_key)
            if existing is not None and existing != '':
                if isinstance(existing, dict):
                    if existing.get('modified') not in (None, ''):
                        continue
                    if existing.get('original') not in (None, ''):
                        continue
                else:
                    continue
            val = var_map.get(var_name)
            if val is not None and val != '':
                enriched[cell_key] = val
    return enriched


def _enrich_matrix_export_data(section_node, existing_data, template_version, assignment_entity_status):
    """Walk section tree and merge resolved variable values into matrix export data."""
    if not isinstance(section_node, dict):
        return
    for field_dict in section_node.get('fields_ordered') or []:
        if not isinstance(field_dict, dict) or field_dict.get('kind') != 'matrix':
            continue
        item_key = f"form_item_{field_dict.get('id')}"
        raw = existing_data.get(item_key)
        if raw is None:
            raw = {}
        elif not isinstance(raw, dict):
            raw = {'value': raw}
        existing_data[item_key] = _merge_matrix_variable_values(
            field_dict,
            raw,
            template_version,
            assignment_entity_status,
        )
    for child in section_node.get('subsections') or []:
        _enrich_matrix_export_data(child, existing_data, template_version, assignment_entity_status)


def register_export_routes(bp):
    """Register export/import routes onto the forms blueprint."""

    @bp.route("/assignment_status/<int:aes_id>/export_pdf", methods=["GET"])
    @login_required
    def export_assignment_pdf(aes_id):
        """Generate a high-quality PDF for an assignment using a print-optimized HTML template."""
        return _export_pdf_impl(aes_id)

    @bp.route("/assignment_status/<int:aes_id>/export_excel", methods=["GET"])
    @login_required
    def export_focal_data_excel(aes_id):
        return _export_excel_impl(aes_id)

    @bp.route("/assignment_status/<int:aes_id>/import_excel", methods=["POST"])
    @login_required
    def handle_excel_import(aes_id):
        """Legacy Excel import endpoint (delegates to shared ExcelService)."""
        return _import_excel_impl(aes_id)


def _export_pdf_impl(aes_id):
    """Generate a high-quality PDF for an assignment using a print-optimized HTML template.

    Uses WeasyPrint (HTML to PDF) for faithful rendering with proper pagination and styles.
    """
    try:
        assignment_entity_status = AssignmentEntityStatus.query.options(
            db.joinedload(AssignmentEntityStatus.assigned_form).joinedload(AssignedForm.template)
        ).get_or_404(aes_id)

        from app.services.organization.authorization_service import AuthorizationService
        if not AuthorizationService.can_access_assignment(assignment_entity_status, current_user):
            flash("You are not authorized to export data for this assignment and country.", "warning")
            return redirect(url_for("main.dashboard"))

        assignment = assignment_entity_status.assigned_form
        from app.utils.api_serialization import _country_for_aes
        country = _country_for_aes(assignment_entity_status)
        form_template_for_export = assignment.template

        from app.services.forms.variable_resolution_service import VariableResolutionService
        from app.models import FormTemplateVersion

        template_version = None
        resolved_variables = {}
        variable_configs = {}
        if form_template_for_export.published_version_id:
            template_version = FormTemplateVersion.query.get(form_template_for_export.published_version_id)
            if template_version:
                variable_configs = template_version.variables or {}
                resolved_variables = VariableResolutionService.resolve_variables(
                    template_version,
                    assignment_entity_status
                )

        if country:
            try:
                # Binding-aware resolution: keeps EO1/EO2/EO3 in the export aligned with the appeal
                # codes the data was actually entered against (see emergency_section_binding).
                from app.services.forms.emergency_section_binding import resolve_eo_variables
                eo_vars = resolve_eo_variables(assignment_entity_status)
                for key, value in eo_vars.items():
                    resolved_variables[key] = value or ''
            except Exception as e:
                current_app.logger.debug(
                    f"Could not resolve EO1/EO2/EO3 for PDF export (plugin or API): {e}"
                )

        translation_key = get_translation_key()

        assignment_display_name = None
        with suppress(Exception):
            assignment_display_name = get_localized_template_name(
                form_template_for_export,
                locale=translation_key,
                version=template_version,
            )
        country_display_name = None
        with suppress(Exception):
            country_display_name = get_localized_country_name(country) if country else None

        from app.services.organization.entity_service import EntityService
        entity_display_name = None
        with suppress(Exception):
            entity_display_name = EntityService.get_localized_entity_name(
                assignment_entity_status.entity_type,
                assignment_entity_status.entity_id,
            )

        sections_by_page = {}
        default_page_id = 0

        section_nodes_by_id = {}
        ordered_section_ids = []

        for section_model in form_template_for_export.sections.order_by(FormSection.order).all():
            section_display_name = None
            with suppress(Exception):
                section_display_name = get_localized_section_name(section_model)
            if not section_display_name:
                section_display_name = getattr(section_model, 'display_name', None) or section_model.name
            if resolved_variables and section_display_name:
                try:
                    section_display_name = VariableResolutionService.replace_variables_in_text(
                        section_display_name,
                        resolved_variables,
                        variable_configs
                    )
                except Exception as e:
                    current_app.logger.warning(
                        f"Error resolving variables in section name for section {section_model.id}: {e}",
                        exc_info=True
                    )

            section_data_for_export = {
                'name': section_model.name,
                'display_name': section_display_name,
                'id': section_model.id,
                'order': section_model.order,
                'page_id': section_model.page_id,
                'parent_section_id': section_model.parent_section_id,
                'relevance_condition': getattr(section_model, 'relevance_condition', None),
                'subsections': [],
                'fields_ordered': []
            }

            temp_fields = []
            form_items = FormItem.query.filter_by(section_id=section_model.id, archived=False).order_by(FormItem.order).all()
            if form_items:
                for form_item in form_items:
                    display_label = None
                    with suppress(Exception):
                        lt = getattr(form_item, 'label_translations', None)
                        if isinstance(lt, dict) and lt:
                            candidate = lt.get(translation_key) or lt.get('en')
                            if isinstance(candidate, str) and candidate.strip():
                                display_label = candidate.strip()
                    if not display_label:
                        display_label = getattr(form_item, 'display_label', None) or form_item.label
                    if resolved_variables and display_label:
                        try:
                            display_label = VariableResolutionService.replace_variables_in_text(
                                display_label,
                                resolved_variables,
                                variable_configs
                            )
                        except Exception as e:
                            current_app.logger.warning(
                                f"Error resolving variables in display_label for form_item {form_item.id}: {e}",
                                exc_info=True
                            )

                    base = {
                        'id': form_item.id,
                        'order': form_item.order,
                        'label': form_item.label,
                        'display_label': display_label,
                        'unit': getattr(form_item, 'unit', None),
                        'type': getattr(form_item, 'type', None),
                        'conditions': getattr(form_item, 'conditions', None),
                    }
                    if form_item.is_indicator:
                        base.update({'kind': 'indicator', 'model': form_item})
                        temp_fields.append(base)
                    elif form_item.is_question:
                        is_blank_note = (
                            getattr(form_item, 'type', None) == 'blank'
                            or (getattr(form_item, 'question_type', None) and getattr(form_item.question_type, 'value', None) == 'blank')
                        )
                        if is_blank_note:
                            note_body = _resolve_note_body(
                                form_item,
                                translation_key,
                                resolved_variables,
                                variable_configs,
                            )
                            base.update({
                                'kind': 'note',
                                'model': form_item,
                                'note_label': display_label or '',
                                'note_body': note_body,
                            })
                        else:
                            base.update({'kind': 'question', 'model': form_item})
                        temp_fields.append(base)
                    elif getattr(form_item, 'item_type', None) == 'matrix' or getattr(form_item, 'is_matrix', False):
                        matrix_config = {}
                        try:
                            if isinstance(getattr(form_item, 'config', None), dict):
                                matrix_config = form_item.config.get('matrix_config') or form_item.config or {}
                        except Exception as e:
                            current_app.logger.debug("matrix_config parse failed: %s", e)
                            matrix_config = {}

                        matrix_rows = getattr(form_item, '_display_matrix_rows', None)
                        if not matrix_rows and isinstance(matrix_config, dict):
                            matrix_rows = matrix_config.get('rows', []) or []

                        try:
                            if isinstance(matrix_config, dict):
                                row_mode = matrix_config.get('row_mode', 'manual')
                                if row_mode == 'manual' or not row_mode:
                                    if resolved_variables and matrix_rows and isinstance(matrix_rows, list):
                                        resolved_rows = []
                                        for r in matrix_rows:
                                            if isinstance(r, str):
                                                resolved_rows.append(
                                                    VariableResolutionService.replace_variables_in_text(
                                                        r, resolved_variables, variable_configs
                                                    )
                                                )
                                            elif isinstance(r, dict):
                                                row_text = r.get('text', '')
                                                resolved_text = VariableResolutionService.replace_variables_in_text(
                                                    row_text, resolved_variables, variable_configs
                                                ) if row_text else row_text
                                                resolved_rows.append({**r, 'text': resolved_text})
                                            else:
                                                resolved_rows.append(r)
                                        matrix_rows = resolved_rows
                        except Exception as e:
                            current_app.logger.warning(
                                f"Error resolving variables in matrix row labels for form_item {form_item.id}: {e}",
                                exc_info=True
                            )
                        matrix_columns = matrix_config.get('columns', []) if isinstance(matrix_config, dict) else []
                        matrix_column_groups = (
                            matrix_config.get('column_groups', {})
                            if isinstance(matrix_config, dict) else {}
                        )
                        try:
                            resolved_columns, resolved_groups = VariableResolutionService.resolve_matrix_display_headers(
                                matrix_config,
                                resolved_variables,
                                variable_configs,
                                replace_fn=lambda text: VariableResolutionService.replace_variables_in_text(
                                    text,
                                    resolved_variables,
                                    variable_configs,
                                ),
                            )
                            if resolved_columns:
                                matrix_columns = resolved_columns
                            if resolved_groups:
                                matrix_column_groups = resolved_groups
                        except Exception as e:
                            current_app.logger.debug(
                                "resolve_matrix_display_headers failed for form_item %s: %s",
                                form_item.id,
                                e,
                            )

                        base.update({
                            'kind': 'matrix',
                            'model': form_item,
                            'matrix_config': matrix_config,
                            'matrix_rows': matrix_rows,
                            'matrix_columns': matrix_columns,
                            'matrix_column_groups': matrix_column_groups,
                        })
                        temp_fields.append(base)
                    elif form_item.is_document_field:
                        base.update({'kind': 'document', 'model': form_item})
                        temp_fields.append(base)

            section_type = getattr(section_model, 'section_type', None) or 'standard'
            if section_type == 'dynamic_indicators':
                dynamic_assignments = DynamicIndicatorData.query.filter_by(
                    assignment_entity_status_id=assignment_entity_status.id,
                    section_id=section_model.id,
                ).order_by(DynamicIndicatorData.order).all()
                for dyn in dynamic_assignments:
                    display_label = dyn.custom_label
                    if not (display_label and str(display_label).strip()):
                        with suppress(Exception):
                            display_label = get_localized_indicator_name(dyn.indicator_bank)
                    if not display_label:
                        display_label = getattr(dyn.indicator_bank, 'name', '') or ''
                    if resolved_variables and display_label:
                        try:
                            display_label = VariableResolutionService.replace_variables_in_text(
                                display_label, resolved_variables, variable_configs
                            )
                        except Exception as e:
                            current_app.logger.debug("replace_variables for display_label failed: %s", e)
                    temp_fields.append({
                        'id': f'dynamic_{dyn.id}',
                        'order': dyn.order,
                        'label': display_label,
                        'display_label': display_label,
                        'unit': getattr(dyn.indicator_bank, 'unit', None),
                        'type': getattr(dyn.indicator_bank, 'type', None),
                        'conditions': None,
                        'kind': 'indicator',
                        'model': None,
                    })

            temp_fields.sort(key=lambda x: (x.get('order') is None, x.get('order')))
            section_data_for_export['fields_ordered'] = temp_fields

            section_nodes_by_id[section_model.id] = section_data_for_export
            ordered_section_ids.append(section_model.id)

        for section_id in ordered_section_ids:
            node = section_nodes_by_id.get(section_id)
            if not node:
                continue
            parent_id = node.get('parent_section_id')
            if parent_id and parent_id in section_nodes_by_id:
                section_nodes_by_id[parent_id]['subsections'].append(node)

        for section_id in ordered_section_ids:
            node = section_nodes_by_id.get(section_id)
            if not node or node.get('parent_section_id') is not None:
                continue
            page_id = node.get('page_id') if node.get('page_id') is not None else default_page_id
            if page_id not in sections_by_page:
                sections_by_page[page_id] = []
            sections_by_page[page_id].append(node)

        existing_data_processed_for_export = _load_existing_data_for_pdf_export(
            assignment_entity_status,
            form_template_for_export,
        )

        def _parse_hidden_ids_arg(arg_name):
            raw = (request.args.get(arg_name) or '').strip()
            if not raw:
                return set()
            out = set()
            for part in raw.split(','):
                part = (part or '').strip()
                if not part:
                    continue
                if part.isdigit():
                    try:
                        out.add(int(part))
                    except (ValueError, TypeError):
                        continue
            return out

        hidden_section_ids_from_client = _parse_hidden_ids_arg('hidden_sections')
        hidden_field_ids_from_client = _parse_hidden_ids_arg('hidden_fields')

        def _filter_section_node(section_node):
            if not isinstance(section_node, dict):
                return None

            try:
                if section_node.get('id') in hidden_section_ids_from_client:
                    return None
            except Exception as e:
                current_app.logger.debug("hidden section filter failed: %s", e)

            kept_fields = []
            for f in (section_node.get('fields_ordered') or []):
                if not isinstance(f, dict):
                    continue
                try:
                    if f.get('id') in hidden_field_ids_from_client:
                        continue
                except Exception as e:
                    current_app.logger.debug("hidden field filter failed: %s", e)
                kept_fields.append(f)
            section_node['fields_ordered'] = kept_fields

            kept_children = []
            for child in (section_node.get('subsections') or []):
                kept = _filter_section_node(child)
                if kept is not None:
                    kept_children.append(kept)
            section_node['subsections'] = kept_children
            return section_node

        filtered_sections_by_page = {}
        for page_id, root_sections in (sections_by_page or {}).items():
            kept_roots = []
            for sec in (root_sections or []):
                kept = _filter_section_node(sec)
                if kept is not None:
                    kept_roots.append(kept)
            filtered_sections_by_page[page_id] = kept_roots
        sections_by_page = filtered_sections_by_page

        def _infer_list_library_rows_and_labels(field_dict, matrix_data):
            try:
                if not isinstance(field_dict, dict):
                    return
                if field_dict.get('kind') != 'matrix':
                    return
                if not isinstance(matrix_data, dict) or not matrix_data:
                    return

                matrix_config = field_dict.get('matrix_config') if isinstance(field_dict.get('matrix_config'), dict) else {}
                row_mode = (matrix_config.get('row_mode') or '').strip().lower()
                if row_mode != 'list_library':
                    return

                cols = field_dict.get('matrix_columns') or []
                col_names = []
                for c in cols:
                    if isinstance(c, dict):
                        col_names.append(str(c.get('name') if c.get('name') else c))
                    else:
                        col_names.append(str(c))
                col_names = [c for c in col_names if c and c != 'None']
                if not col_names:
                    return

                col_names_sorted = sorted(col_names, key=len, reverse=True)
                row_ids = []
                seen = set()
                for k in matrix_data.keys():
                    if not isinstance(k, str):
                        continue
                    if k.startswith('_'):
                        continue
                    matched_row_id = None
                    for cn in col_names_sorted:
                        suffix = "_" + cn
                        if k.endswith(suffix):
                            matched_row_id = k[: -len(suffix)]
                            break
                    if matched_row_id and matched_row_id not in seen:
                        seen.add(matched_row_id)
                        row_ids.append(matched_row_id)

                if not row_ids:
                    return

                from flask import session
                lookup_list_id = (matrix_config.get('lookup_list_id') or '').strip()

                display_column = (matrix_config.get('display_column') or matrix_config.get('list_display_column') or 'name').strip() or 'name'

                row_labels = {}
                if lookup_list_id and str(lookup_list_id).isdigit():
                    from app.models import LookupListRow
                    for rid in row_ids:
                        try:
                            rid_int = int(rid)
                        except (ValueError, TypeError):
                            row_labels[rid] = rid
                            continue
                        row_obj = LookupListRow.query.get(rid_int)
                        if row_obj and isinstance(row_obj.data, dict):
                            row_labels[rid] = str(row_obj.data.get(display_column) or row_obj.data.get('name') or rid)
                        else:
                            row_labels[rid] = rid
                elif lookup_list_id in ('country_map', 'national_society', 'indicator_bank'):
                    current_locale = session.get('language', 'en') or (str(get_locale()) if get_locale() else 'en')
                    if isinstance(current_locale, str) and '_' in current_locale:
                        current_locale = current_locale.split('_', 1)[0]

                    if lookup_list_id == 'country_map':
                        for rid in row_ids:
                            try:
                                rid_int = int(rid)
                            except (ValueError, TypeError):
                                row_labels[rid] = rid
                                continue
                            obj = Country.query.get(rid_int)
                            row_labels[rid] = get_localized_country_name(obj) if obj else rid
                    elif lookup_list_id == 'national_society':
                        from app.models.organization import NationalSociety
                        for rid in row_ids:
                            try:
                                rid_int = int(rid)
                            except (ValueError, TypeError):
                                row_labels[rid] = rid
                                continue
                            obj = NationalSociety.query.get(rid_int)
                            if obj:
                                localized_name = obj.get_name_translation(current_locale) if hasattr(obj, 'get_name_translation') else None
                                row_labels[rid] = (localized_name.strip() if isinstance(localized_name, str) and localized_name.strip() else obj.name)
                            else:
                                row_labels[rid] = rid
                    else:
                        from app.models.indicator_bank import IndicatorBank
                        for rid in row_ids:
                            try:
                                rid_int = int(rid)
                            except (ValueError, TypeError):
                                row_labels[rid] = rid
                                continue
                            obj = IndicatorBank.query.get(rid_int)
                            row_labels[rid] = obj.name if obj else rid
                else:
                    row_labels = {rid: rid for rid in row_ids}

                field_dict['matrix_rows'] = row_ids
                field_dict['matrix_row_labels'] = row_labels
            except Exception as e:
                current_app.logger.warning(f"Failed to infer list-library matrix rows for PDF: {e}", exc_info=True)

        def _walk_sections_for_export(section_node):
            if not isinstance(section_node, dict):
                return
            fields = section_node.get('fields_ordered') or []
            if isinstance(fields, list):
                for f in fields:
                    if isinstance(f, dict) and f.get('kind') == 'matrix':
                        item_key = f"form_item_{f.get('id')}"
                        _infer_list_library_rows_and_labels(f, existing_data_processed_for_export.get(item_key))
            for child in section_node.get('subsections', []) or []:
                _walk_sections_for_export(child)

        for page_id, root_sections in (sections_by_page or {}).items():
            for sec in root_sections or []:
                _walk_sections_for_export(sec)

        if template_version:
            for page_id, root_sections in (sections_by_page or {}).items():
                for sec in root_sections or []:
                    _enrich_matrix_export_data(
                        sec,
                        existing_data_processed_for_export,
                        template_version,
                        assignment_entity_status,
                    )

        pages = list(form_template_for_export.pages) if form_template_for_export.is_paginated else [None]

        html_content = render_template(
            'forms/entry_form/export_pdf.html',
            assignment=assignment,
            assignment_display_name=assignment_display_name,
            country=country,
            country_display_name=country_display_name,
            entity_display_name=entity_display_name,
            aes=assignment_entity_status,
            form_template=form_template_for_export,
            sections_by_page=sections_by_page,
            pages=pages,
            existing_data=existing_data_processed_for_export,
            generated_at=utcnow(),
            get_localized_page_name=get_localized_page_name,
            matrix_cell_is_prefilled_highlight=matrix_cell_is_prefilled_highlight,
            matrix_pdf_layout_strategy=matrix_pdf_layout_strategy,
            matrix_portrait_column_widths_mm=matrix_portrait_column_widths_mm,
        )

        try:
            from weasyprint import HTML, CSS  # type: ignore
        except Exception as e:
            current_app.logger.error(f"WeasyPrint not available: {e}", exc_info=True)
            return current_app.response_class(
                response="PDF generation is not available on this deployment.",
                status=503,
                mimetype='text/plain'
            )

        static_dir = os.path.join(current_app.root_path, 'static')

        pdf_css_string = '''
            @page {
                size: A4;
                margin: 20mm 10mm 20mm 10mm;
                @bottom-right { content: "Page " counter(page); font-size: 10pt; color: #6b7280; }
            }
            body {
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Arial, sans-serif;
                color: #111827;
                line-height: 1.5;
            }
            h1, h2, h3, h4 { color: #111827; margin: 0 0 8px 0; }
            h1 { font-size: 20pt; }
            h2 { font-size: 14pt; border-bottom: 2px solid #cc0000; padding-bottom: 4px; margin-top: 16px; margin-bottom: 12px; }
            .form-page-title { page-break-after: avoid; }
            h3 { font-size: 12pt; margin-top: 10px; margin-bottom: 8px; color: #374151; }
            h4 { font-size: 11pt; margin-top: 8px; margin-bottom: 6px; color: #374151; }
            .meta {
                color: #374151;
                font-size: 10pt;
                margin-bottom: 16px;
                padding: 8px;
                background: #f9fafb;
                border-left: 3px solid #cc0000;
            }
            .meta div { margin: 4px 0; }
            .field-flag {
                display: inline-block;
                margin-left: 8px;
                padding: 1px 6px;
                border-radius: 3px;
                font-size: 8pt;
                font-weight: 600;
                vertical-align: middle;
            }
            .field-flag-prefilled {
                background: #fef9c3;
                color: #854d0e;
                border: 1px solid #fde047;
            }
            .field-flag-imputed {
                background: #dbeafe;
                color: #1e40af;
                border: 1px solid #93c5fd;
            }
            .matrix-legend {
                margin: 0 0 8px 0;
                padding: 6px 8px;
                background: #f9fafb;
                border: 1px solid #e5e7eb;
                border-radius: 4px;
                font-size: 9pt;
                color: #374151;
                page-break-inside: avoid;
            }
            .matrix-field-flag {
                text-align: right;
                margin: 0 0 6px 0;
                page-break-inside: avoid;
            }
            .matrix-legend-item {
                display: flex;
                align-items: center;
                gap: 8px;
            }
            .matrix-legend-swatch {
                width: 14px;
                height: 14px;
                border: 1px solid #fde047;
                background: #fef9c3;
                border-radius: 2px;
                flex-shrink: 0;
            }
            .field-box-imputed {
                background: #eff6ff;
                border-color: #93c5fd;
            }
            .cell-prefilled {
                background: #fef9c3 !important;
            }
            .cell-imputed {
                background: #eff6ff !important;
            }
            .field-value-prefilled {
                display: inline-block;
                background: #fef9c3;
                border: 1px solid #fde047;
                border-radius: 3px;
                padding: 2px 6px;
            }
            .field-value-imputed {
                display: inline-block;
                background: #eff6ff;
                border: 1px solid #93c5fd;
                border-radius: 3px;
                padding: 2px 6px;
            }
            .section {
                margin-bottom: 16px;
            }
            .section h2, .section h3, .section h4 {
                page-break-after: avoid;
            }
            .subsection {
                margin-left: 12px;
                padding-left: 10px;
                border-left: 2px solid #e5e7eb;
            }
            .section-empty-note {
                color: #6b7280;
                font-size: 10pt;
                font-style: italic;
                margin: 8px 0 0 0;
            }
            .form-note {
                margin: 8px 0;
                padding: 8px 12px;
                background: #f9fafb;
                border-left: 3px solid #9ca3af;
                border-radius: 0 4px 4px 0;
            }
            .form-note-heading {
                color: #111827;
                font-size: 10pt;
                font-weight: 600;
                margin: 0 0 4px 0;
                white-space: pre-wrap;
            }
            .form-note-body {
                color: #374151;
                font-size: 10pt;
                line-height: 1.25;
            }
            .form-note-body p {
                margin: 0;
            }
            .form-note-body ul, .form-note-body ol { margin: 4px 0 6px 1.2em; padding: 0; }
            .form-note-body li { margin: 2px 0; }
            .form-note-body a { color: #2563eb; text-decoration: underline; }
            .form-note-body strong, .form-note-body b { font-weight: 600; }

            .field-box {
                border: 1.5px solid #e5e7eb;
                border-radius: 4px;
                margin: 8px 0;
                page-break-inside: avoid;
                background: #ffffff;
            }
            .field-unlabeled {
                margin: 8px 0;
                border: none;
                background: transparent;
                page-break-inside: auto;
            }
            .field-content-unlabeled {
                padding: 0;
                min-height: 0;
            }
            .field-box-matrix {
                page-break-inside: auto;
            }
            /* Wide matrices (8–13 cols): squeeze into portrait via fixed layout + narrow columns */
            .field-box-matrix-wide.wide-matrix-portrait-compact,
            .field-unlabeled.field-box-matrix-wide.wide-matrix-portrait-compact {
                page-break-inside: auto;
            }
            .wide-matrix-portrait-compact .wide-matrix-compact .matrix-table {
                table-layout: fixed;
                width: 100%;
                font-size: 7pt;
            }
            .wide-matrix-portrait-compact .wide-matrix-compact .matrix-table th,
            .wide-matrix-portrait-compact .wide-matrix-compact .matrix-table td {
                padding: 2px 3px;
                font-size: 7pt;
                line-height: 1.25;
                overflow-wrap: anywhere;
                word-break: break-word;
                hyphens: auto;
                vertical-align: top;
            }
            .wide-matrix-portrait-compact .wide-matrix-compact .matrix-table th:first-child,
            .wide-matrix-portrait-compact .wide-matrix-compact .matrix-table td:first-child {
                min-width: 0;
                max-width: none;
            }
            .wide-matrix-portrait-compact .wide-matrix-compact .matrix-group-header {
                font-size: 6.5pt;
                padding: 2px 2px;
                text-align: center;
            }
            /* Wide matrices (14+ cols): landscape page */
            .field-box-matrix-wide.wide-matrix-landscape,
            .field-box-matrix-wide.wide-matrix-landscape-scale {
                page: wide;
                page-break-before: always;
                page-break-inside: auto;
            }
            .field-box-matrix-wide.wide-matrix-landscape-scale {
                overflow: hidden;
            }
            .field-box-matrix-wide.wide-matrix-landscape-scale .matrix-table {
                transform: scale(0.82);
                transform-origin: top left;
            }
            .field-box-matrix .matrix-table {
                page-break-inside: auto;
            }
            .field-box-matrix .matrix-table thead {
                display: table-header-group;
            }
            .field-box-matrix .matrix-table thead tr {
                page-break-inside: avoid;
                page-break-after: avoid;
            }
            .field-box-matrix .matrix-table tbody tr {
                page-break-inside: avoid;
            }
            .field-box-matrix .matrix-table td:first-child,
            .field-box-matrix .matrix-table th:first-child {
                min-width: 22mm;
                max-width: 34mm;
                vertical-align: top;
            }
            .matrix-group-header {
                background: #eef2ff;
                color: #3730a3;
                font-weight: 600;
                text-align: center;
            }

            .field-filled {
                border-left: 4px solid #10b981;
            }

            .field-empty {
                border-left: 4px solid #d1d5db;
                background: #f9fafb;
            }

            .field-empty-required {
                border-left: 4px solid #ef4444;
                background: #fef2f2;
            }

            .field-empty-optional {
                border-left: 4px solid #d1d5db;
                background: #f9fafb;
            }

            .field-header {
                background: #f9fafb;
                padding: 8px 12px;
                border-bottom: 1px solid #e5e7eb;
                font-weight: 600;
            }

            .field-label {
                color: #111827;
                font-size: 11pt;
                display: block;
            }

            .field-unit {
                color: #6b7280;
                font-size: 9pt;
                font-weight: normal;
                font-style: italic;
            }

            .field-content {
                padding: 10px 12px;
                min-height: 20px;
            }

            .field-value {
                color: #111827;
                font-size: 10pt;
                word-wrap: break-word;
                display: block;
            }
            .field-value-multiline {
                line-height: 1.4;
            }

            .disaggregation-caption {
                font-size: 9pt;
                font-weight: 600;
                color: #374151;
                margin: 6px 0 4px 0;
                display: block;
            }

            .not-reported {
                color: #dc2626;
                font-size: 10pt;
                font-weight: 600;
                font-style: italic;
                display: block;
            }

            .not-reported-optional {
                color: #6b7280;
                font-size: 10pt;
                font-weight: 600;
                font-style: italic;
                display: block;
            }

            html[dir="rtl"] body {
                direction: rtl;
                font-family: "Tajawal", -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Arial, sans-serif;
            }
            html[dir="rtl"] h1,
            html[dir="rtl"] h2,
            html[dir="rtl"] h3,
            html[dir="rtl"] h4,
            html[dir="rtl"] .meta,
            html[dir="rtl"] .field-header,
            html[dir="rtl"] .field-content,
            html[dir="rtl"] .field-value,
            html[dir="rtl"] .disaggregation-caption,
            html[dir="rtl"] .not-reported,
            html[dir="rtl"] .not-reported-optional {
                text-align: right;
                font-family: "Tajawal", -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Arial, sans-serif;
            }
            html[dir="rtl"] .meta { border-left: none; border-right: 3px solid #cc0000; }
            html[dir="rtl"] .form-note { border-left: none; border-right: 3px solid #9ca3af; border-radius: 4px 0 0 4px; }
            html[dir="rtl"] .form-note-heading,
            html[dir="rtl"] .form-note-body { text-align: right; }
            html[dir="rtl"] .subsection { margin-left: 0; margin-right: 12px; padding-left: 0; padding-right: 10px; border-left: none; border-right: 2px solid #e5e7eb; }
            html[dir="rtl"] table th, html[dir="rtl"] table td {
                text-align: right;
                font-family: "Tajawal", -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Arial, sans-serif;
            }

            .document-note {
                color: #6b7280;
                font-style: italic;
            }

            .table {
                width: 100%;
                border-collapse: collapse;
                margin: 8px 0;
                font-size: 9pt;
            }
            .table th, .table td {
                border: 1px solid #d1d5db;
                padding: 6px 8px;
            }
            .table th {
                background: #f3f4f6;
                text-align: left;
                font-weight: 600;
                color: #374151;
            }
            .table td {
                background: #ffffff;
            }
            .table tbody tr:nth-child(even) td {
                background: #f9fafb;
            }
            .table td.cell-tick {
                text-align: center;
            }
            /* Wide matrices (8+ columns): tighter cells */
            .wide-matrix-compact table.table,
            .wide-matrix-compact table.matrix-table {
                font-size: 7.5pt;
            }
            .wide-matrix-compact table.table th,
            .wide-matrix-compact table.table td,
            .wide-matrix-compact table.matrix-table th,
            .wide-matrix-compact table.matrix-table td {
                padding: 3px 4px;
                word-wrap: break-word;
                overflow-wrap: anywhere;
                hyphens: auto;
            }
            .wide-matrix-compact table.table th:first-child,
            .wide-matrix-compact table.table td:first-child,
            .wide-matrix-compact table.matrix-table th:first-child,
            .wide-matrix-compact table.matrix-table td:first-child {
                max-width: 28mm;
            }
            .wide-matrix-scale {
                width: 100%;
                overflow: hidden;
            }
            .wide-matrix-scale table.table {
                transform: scale(0.72);
                transform-origin: top left;
            }
            /* Matrix alignment — explicit classes; keep after generic table/RTL rules */
            html[dir="ltr"] .matrix-table thead th.matrix-row-header,
            html[dir="ltr"] .matrix-table thead th.matrix-col-header {
                text-align: left;
                hyphens: none;
            }
            html[dir="ltr"] .matrix-table thead th.matrix-group-header {
                text-align: center;
            }
            html[dir="ltr"] .matrix-table thead th.cell-tick,
            html[dir="ltr"] .matrix-table tbody td.cell-tick {
                text-align: center;
            }
            html[dir="ltr"] .matrix-table tbody td.matrix-col-data,
            html[dir="ltr"] .matrix-table tbody td.matrix-column-total-cell,
            html[dir="ltr"] .matrix-table tbody td.matrix-row-total-cell,
            html[dir="ltr"] .matrix-table tbody td.matrix-grand-total-cell {
                text-align: right;
            }
            html[dir="ltr"] .matrix-table tbody td.matrix-row-label {
                text-align: left;
            }
            html[dir="rtl"] .matrix-table thead th.matrix-row-header,
            html[dir="rtl"] .matrix-table thead th.matrix-col-header {
                text-align: right;
            }
            html[dir="rtl"] .matrix-table thead th.matrix-group-header {
                text-align: center;
            }
            html[dir="rtl"] .matrix-table tbody td.matrix-col-data,
            html[dir="rtl"] .matrix-table tbody td.matrix-column-total-cell,
            html[dir="rtl"] .matrix-table tbody td.matrix-row-total-cell,
            html[dir="rtl"] .matrix-table tbody td.matrix-grand-total-cell {
                text-align: left;
            }
            @page wide {
                size: A4 landscape;
                margin: 15mm 8mm 15mm 8mm;
                @bottom-right { content: "Page " counter(page); font-size: 10pt; color: #6b7280; }
            }
            .page-break { page-break-before: always; }
        '''
        with suppress(Exception):
            pdf_css_string = pdf_css_string.replace('content: "Page "', f'content: "{_("Page")} "')
        pdf_css = CSS(string=pdf_css_string)

        pdf_buffer = io.BytesIO()
        HTML(string=html_content, base_url=static_dir).write_pdf(
            pdf_buffer,
            stylesheets=[pdf_css],
            optimize_images=True,
        )

        pdf_buffer.seek(0)
        filename = _make_assignment_pdf_download_name(
            entity_display_name or country_display_name,
            assignment_display_name or (assignment.template.name if assignment and assignment.template else None),
        )
        return send_file(
            pdf_buffer,
            download_name=filename,
            as_attachment=True,
            mimetype='application/pdf'
        )
    except Exception as e:
        current_app.logger.error(f"Error generating PDF for ACS {aes_id}: {e}", exc_info=True)
        flash("Failed to generate PDF.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))


def _export_excel_impl(aes_id):
    assignment_entity_status = AssignmentEntityStatus.query.options(
        db.joinedload(AssignmentEntityStatus.assigned_form).joinedload(AssignedForm.template)
    ).get_or_404(aes_id)

    from app.services.organization.authorization_service import AuthorizationService
    if not AuthorizationService.can_access_assignment(assignment_entity_status, current_user):
         flash("You are not authorized to export data for this assignment and country.", "warning")
         return redirect(url_for("main.dashboard"))

    assignment = assignment_entity_status.assigned_form
    from app.utils.api_serialization import _country_for_aes
    country = _country_for_aes(assignment_entity_status)
    form_template_for_export = assignment.template

    from app.services.forms.variable_resolution_service import VariableResolutionService
    from app.models import FormTemplateVersion

    template_version = None
    resolved_variables = {}
    variable_configs = {}
    if form_template_for_export.published_version_id:
        template_version = FormTemplateVersion.query.get(form_template_for_export.published_version_id)
        if template_version:
            variable_configs = template_version.variables or {}
            resolved_variables = VariableResolutionService.resolve_variables(
                template_version,
                assignment_entity_status
            )

    sections_by_page = {}
    default_page_id = 0

    for section_model in form_template_for_export.sections.order_by(FormSection.order).all():
        page_id = section_model.page_id if section_model.page_id is not None else default_page_id
        if page_id not in sections_by_page:
            sections_by_page[page_id] = []

        section_display_name = getattr(section_model, 'display_name', None) or section_model.name
        if resolved_variables and section_display_name:
            try:
                section_display_name = VariableResolutionService.replace_variables_in_text(
                    section_display_name,
                    resolved_variables,
                    variable_configs
                )
            except Exception as e:
                current_app.logger.warning(
                    f"Error resolving variables in section name for section {section_model.id}: {e}",
                    exc_info=True
                )

        section_data_for_export = {'name': section_display_name, 'id': section_model.id, 'fields_ordered': []}
        temp_fields = []

        form_items = FormItem.query.filter_by(section_id=section_model.id, archived=False).order_by(FormItem.order).all()
        if form_items:
            for form_item in form_items:
                display_label = getattr(form_item, 'display_label', None) or form_item.label
                if resolved_variables and display_label:
                    try:
                        display_label = VariableResolutionService.replace_variables_in_text(
                            display_label,
                            resolved_variables,
                            variable_configs
                        )
                    except Exception as e:
                        current_app.logger.warning(
                            f"Error resolving variables in display_label for form_item {form_item.id}: {e}",
                            exc_info=True
                        )

                if form_item.is_indicator:
                    temp_fields.append({
                        'id': form_item.id, 'legacy_id': None, 'label': display_label,
                        'type': form_item.type, 'unit': form_item.unit, 'order': form_item.order,
                        'is_indicator': True, 'is_form_item': True, 'item_model': form_item
                    })
                elif form_item.is_question:
                    temp_fields.append({
                        'id': form_item.id, 'legacy_id': None, 'label': display_label,
                        'type': form_item.type, 'order': form_item.order,
                        'is_question': True, 'is_form_item': True, 'item_model': form_item
                    })
                elif form_item.is_document_field:
                    temp_fields.append({
                        'id': form_item.id, 'legacy_id': None, 'label': display_label,
                        'type': 'DOCUMENT', 'order': form_item.order, 'is_required': form_item.is_required,
                        'description': form_item.description, 'is_document': True, 'is_form_item': True, 'item_model': form_item
                    })

        temp_fields.sort(key=lambda x: x['order'])
        section_data_for_export['fields_ordered'] = temp_fields
        sections_by_page[page_id].append(section_data_for_export)

    existing_data_entries_for_export = FormData.query.filter_by(
        assignment_entity_status_id=assignment_entity_status.id
    ).all()
    existing_data_processed_for_export = {}
    for entry in existing_data_entries_for_export:
        if entry.form_item_id:
            item_key_suffix = f"form_item_{entry.form_item_id}"
            existing_data_processed_for_export[item_key_suffix] = entry.value

    workbook = openpyxl.Workbook()

    IFRC_RED = "FFED1B2E"
    IFRC_DARK_RED = "FFAF0E1B"
    IFRC_LIGHT_GRAY = "FFF5F5F5"
    IFRC_MEDIUM_GRAY = "FFE0E0E0"
    IFRC_DARK_GRAY = "FF666666"
    IFRC_WHITE = "FFFFFFFF"
    IFRC_YELLOW = "FFFFF9E6"

    title_font = Font(name='Arial', size=16, bold=True, color=IFRC_DARK_RED)
    section_title_font = Font(name='Arial', size=14, bold=True, color=IFRC_DARK_RED)
    item_label_font = Font(name='Arial', size=12, bold=True, color=IFRC_DARK_GRAY)
    header_font = Font(name='Arial', size=11, bold=True, color=IFRC_WHITE)
    normal_font = Font(name='Arial', size=11, color=IFRC_DARK_GRAY)

    header_fill = PatternFill(start_color=IFRC_DARK_GRAY, end_color=IFRC_DARK_GRAY, fill_type='solid')
    section_fill = PatternFill(start_color=IFRC_LIGHT_GRAY, end_color=IFRC_LIGHT_GRAY, fill_type='solid')
    data_entry_fill = PatternFill(start_color=IFRC_YELLOW, end_color=IFRC_YELLOW, fill_type='solid')
    alternate_row_fill = PatternFill(start_color=IFRC_LIGHT_GRAY, end_color=IFRC_LIGHT_GRAY, fill_type='solid')

    no_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )
    disagg_border = Border(
        left=Side(style='thin', color=IFRC_DARK_GRAY),
        right=Side(style='thin', color=IFRC_DARK_GRAY),
        top=Side(style='thin', color=IFRC_DARK_GRAY),
        bottom=Side(style='thin', color=IFRC_DARK_GRAY)
    )

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    pages = list(form_template_for_export.pages) if form_template_for_export.is_paginated else [None]
    first_sheet = True

    for page in pages:
        page_id = page.id if page else default_page_id
        sheet_name_raw = get_localized_page_name(page)

        invalid_chars = ['/', '\\', '?', '*', '[', ']']
        sheet_name = sheet_name_raw
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '-')

        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        if first_sheet:
            data_sheet = workbook.active
            data_sheet.title = sheet_name
            first_sheet = False
        else:
            data_sheet = workbook.create_sheet(sheet_name)

            for row in range(1, 1000):
                data_sheet.row_dimensions[row].height = 17

        current_row = 1
        data_sheet.cell(row=current_row, column=1).value = f"Assignment: {form_template_for_export.name} - {assignment.period_name} for {country.name}"
        data_sheet.cell(row=current_row, column=1).font = title_font
        current_row += 2

        for section_data in sections_by_page.get(page_id, []):
            data_sheet.cell(row=current_row, column=1).value = section_data['name']
            data_sheet.cell(row=current_row, column=1).font = section_title_font
            current_row += 1

            for field_data in section_data['fields_ordered']:
                field_model = field_data['item_model']
                data_sheet.cell(row=current_row, column=2).value = f"{field_data['order']}. {field_data['label']}"
                data_sheet.cell(row=current_row, column=2).font = item_label_font
                current_row += 1
                col_offset = 3

                if field_data.get('is_indicator'):
                    indicator = field_model
                    item_key = f"form_item_{field_data['id']}"
                    entry_data = existing_data_processed_for_export.get(item_key, {})

                    allowed_modes = indicator.allowed_disaggregation_options if indicator.supports_disaggregation else ['total']

                    if indicator.type == 'Number':
                        for mode in allowed_modes:
                            mode_display = Config.DISAGGREGATION_MODES.get(mode, mode.title())
                            mode_cell = data_sheet.cell(row=current_row, column=col_offset, value=mode_display)
                            mode_cell.font = header_font
                            mode_cell.fill = header_fill
                            mode_cell.border = disagg_border
                            mode_cell.alignment = center_align
                            current_row += 1

                            current_values = entry_data.get('values', {}) if isinstance(entry_data, dict) else {'value': entry_data if entry_data is not None else ''}

                            if mode == 'total':
                                val_cell = data_sheet.cell(row=current_row, column=col_offset, value=current_values.get('total', ''))
                                val_cell.fill = data_entry_fill
                                val_cell.border = disagg_border
                                val_cell.number_format = '0'
                                current_row += 1

                            elif mode == 'sex':
                                headers = indicator.effective_sex_categories
                                for col, header in enumerate(headers, col_offset):
                                    data_sheet.cell(row=current_row, column=col).value = header
                                    data_sheet.cell(row=current_row, column=col).font = header_font
                                current_row += 1

                                for col, category in enumerate(headers, col_offset):
                                    val_key = category.lower().replace(' ', '_')
                                    val_cell = data_sheet.cell(row=current_row, column=col, value=current_values.get(val_key, ''))
                                    val_cell.fill = data_entry_fill
                                    val_cell.border = disagg_border
                                    val_cell.number_format = '0'
                                    data_sheet.row_dimensions[current_row].height = 17
                                current_row += 1

                            elif mode == 'age':
                                headers = indicator.effective_age_groups
                                for col, header in enumerate(headers, col_offset):
                                    data_sheet.cell(row=current_row, column=col).value = header
                                    data_sheet.cell(row=current_row, column=col).font = header_font
                                current_row += 1

                                for col, category in enumerate(headers, col_offset):
                                    val_key = category.lower().replace(' ', '_').replace('+', 'plus')
                                    val_cell = data_sheet.cell(row=current_row, column=col, value=current_values.get(val_key, ''))
                                    val_cell.fill = data_entry_fill
                                    val_cell.border = disagg_border
                                    val_cell.number_format = '0'
                                current_row += 1

                            elif mode == 'sex_age':
                                age_groups = indicator.effective_age_groups
                                sex_categories = indicator.effective_sex_categories

                                category_cell = data_sheet.cell(row=current_row, column=col_offset, value="Category")
                                category_cell.font = header_font
                                category_cell.fill = header_fill
                                category_cell.border = disagg_border
                                category_cell.alignment = center_align

                                for col, age in enumerate(age_groups, col_offset + 1):
                                    header_cell = data_sheet.cell(row=current_row, column=col, value=age)
                                    header_cell.font = header_font
                                    header_cell.fill = header_fill
                                    header_cell.border = disagg_border
                                    header_cell.alignment = center_align
                                current_row += 1

                                for sex in sex_categories:
                                    sex_cell = data_sheet.cell(row=current_row, column=col_offset, value=sex)
                                    sex_cell.font = header_font
                                    sex_cell.fill = header_fill
                                    sex_cell.border = disagg_border
                                    sex_cell.alignment = center_align

                                    for col, age in enumerate(age_groups, col_offset + 1):
                                        val_key = f"{sex.lower().replace(' ', '_')}_{age.lower().replace(' ', '_').replace('+', 'plus')}"
                                        val_cell = data_sheet.cell(row=current_row, column=col, value=current_values.get(val_key, ''))
                                        val_cell.fill = data_entry_fill
                                        val_cell.border = disagg_border
                                        val_cell.number_format = '0'
                                    current_row += 1

                            current_row += 1
                    else:
                                                    # Non-numeric indicators just get a single value cell
                                val_cell = data_sheet.cell(row=current_row, column=col_offset, value=current_values.get('value', ''))
                                val_cell.fill = data_entry_fill
                                val_cell.border = no_border
                                current_row += 2

                elif field_data.get('is_question'):
                    question = field_model
                    item_key = f"form_item_{question.id}"
                    q_value = existing_data_processed_for_export.get(item_key, '')
                    val_cell = data_sheet.cell(row=current_row, column=col_offset, value=q_value)
                    val_cell.fill = data_entry_fill
                    val_cell.border = no_border

                    if question.type == QuestionType.number:
                        val_cell.number_format = '0'
                    elif question.type == QuestionType.date and q_value:
                        with suppress(Exception):
                            val_cell.value = datetime.strptime(q_value, '%Y-%m-%d').date()
                            val_cell.number_format = 'yyyy-mm-dd'

                    if question.type in [QuestionType.single_choice, QuestionType.multiple_choice] and question.options:
                        options_list = [str(opt.get('value', opt) if isinstance(opt, dict) else opt) for opt in question.options]
                        options_str = ','.join([f'"{opt}"' for opt in options_list])
                        dv = DataValidation(type="list", formula1=f"={options_str}", allow_blank=True)
                        data_sheet.add_data_validation(dv)
                        dv.add(val_cell)

                    current_row += 2

                elif field_data.get('is_document'):
                    data_sheet.cell(row=current_row, column=col_offset, value="(Manage in Web Form)")
                    current_row += 2

                current_row += 1

        data_sheet.column_dimensions['C'].width = 20
        for col_idx in range(1, data_sheet.max_column + 1):
            if get_column_letter(col_idx) != 'C':
                data_sheet.column_dimensions[get_column_letter(col_idx)].autosize = True

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"data_entry_{country.iso3}_{str(assignment.period_name).replace(' ', '_')}.xlsx"
    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    download_name=filename,
                    as_attachment=True)


def _import_excel_impl(aes_id):
    """Legacy Excel import endpoint (delegates to shared ExcelService).

    NOTE: This endpoint is maintained for backward compatibility and testing.
    New code should use excel.import_assignment_excel instead.
    """
    MAX_EXCEL_FILE_SIZE = 10 * 1024 * 1024

    assignment_entity_status = AssignmentEntityStatus.query.get_or_404(aes_id)

    from app.services.organization.authorization_service import AuthorizationService
    if not AuthorizationService.can_edit_assignment(assignment_entity_status, current_user):
        flash("You are not authorized to import data for this assignment or it's not in an editable state.", "warning")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    excel_file = request.files.get('excel_file')
    if not excel_file or excel_file.filename == '':
        flash("No Excel file selected for import.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    if not excel_file.filename.lower().endswith('.xlsx'):
        flash("Invalid file type. Please upload a .xlsx file.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    file_size = excel_file.content_length
    if file_size is None:
        excel_file.seek(0, 2)
        file_size = excel_file.tell()
        excel_file.seek(0)

    if file_size > MAX_EXCEL_FILE_SIZE:
        flash(f"File size ({file_size / (1024*1024):.2f}MB) exceeds the maximum allowed size of 10MB.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    try:
        from app.utils.advanced_validation import AdvancedValidator
        file_ext = os.path.splitext(excel_file.filename)[1].lower()
        is_valid_mime, detected_mime = AdvancedValidator.validate_mime_type(excel_file, [file_ext])
        if not is_valid_mime:
            current_app.logger.warning(f"Excel import MIME mismatch: claimed {file_ext}, detected {detected_mime}")
            flash("File content does not match its extension. Please upload a valid Excel file.", "danger")
            return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))
    except Exception as e:
        current_app.logger.warning(f"MIME validation error for Excel import: {e}", exc_info=True)
        flash("Unable to validate file type. Please try again.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    try:
        workbook = ExcelService.load_workbook(excel_file)
    except ValueError as exc:
        flash("An error occurred. Please try again.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    result = ExcelService.import_assignment_data(assignment_entity_status, workbook)
    if result['success']:
        if result.get('errors'):
            error_msg = f"Excel import completed with {result['updated_count']} values saved. Errors: {', '.join(result['errors'][:5])}"
            if len(result['errors']) > 5:
                error_msg += f" (and {len(result['errors']) - 5} more)"
            flash(error_msg, "warning")
        else:
            flash(f"Excel import completed: {result['updated_count']} values saved.", "success")
    else:
        error_msg = f"Excel import failed: {', '.join(result.get('errors', [])[:5])}"
        if len(result.get('errors', [])) > 5:
            error_msg += f" (and {len(result.get('errors', [])) - 5} more)"
        flash(error_msg, "danger")

    return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))
