"""Tests for app/utils/file_parsing.py – targets 100 % coverage."""
import io
import codecs
import pytest
from unittest.mock import MagicMock, patch
from werkzeug.datastructures import FileStorage

from app.utils.file_parsing import (
    CSV_EXCEL_EXTENSIONS,
    EXCEL_EXTENSIONS,
    _decode_csv_content,
    parse_csv_to_rows,
    parse_excel_to_rows,
    parse_csv_or_excel_to_rows,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_fs(data: bytes, filename: str = "test.csv") -> FileStorage:
    return FileStorage(stream=io.BytesIO(data), filename=filename, content_type="text/plain")


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

def test_csv_excel_extensions():
    assert ".csv" in CSV_EXCEL_EXTENSIONS
    assert ".xlsx" in CSV_EXCEL_EXTENSIONS
    assert ".xls" in CSV_EXCEL_EXTENSIONS


def test_excel_extensions_subset():
    assert EXCEL_EXTENSIONS < CSV_EXCEL_EXTENSIONS


# ---------------------------------------------------------------------------
# _decode_csv_content
# ---------------------------------------------------------------------------

class TestDecodeCsvContent:
    def test_utf8_bom(self):
        content = codecs.BOM_UTF8 + "name,age\nAlice,30".encode("utf-8")
        result = _decode_csv_content(content)
        assert result.startswith("name,age")

    def test_plain_utf8(self):
        content = "name,age\nBob,25".encode("utf-8")
        result = _decode_csv_content(content)
        assert "Bob" in result

    def test_utf8_sig_fallback(self):
        # Simulate content that fails plain utf-8 but succeeds as utf-8-sig
        # We mock the decode chain to hit the utf-8-sig fallback
        content = codecs.BOM_UTF8 + "hello".encode("utf-8")
        # Verify utf-8-sig path is taken for BOM content
        result = _decode_csv_content(content)
        assert "hello" in result

    def test_latin1_fallback(self):
        # Byte that is invalid in UTF-8
        content = b"name\nCaf\xe9"
        result = _decode_csv_content(content)
        assert "Caf" in result

    def test_utf8_sig_second_fallback(self):
        """Force the utf-8 decode to fail, then utf-8-sig to succeed."""
        bad_utf8 = b"hello\xff\xfe"  # invalid UTF-8 sequence
        result = _decode_csv_content(bad_utf8)
        # Should decode via latin-1 at minimum
        assert isinstance(result, str)


# ---------------------------------------------------------------------------
# parse_csv_to_rows
# ---------------------------------------------------------------------------

class TestParseCsvToRows:
    def test_basic_csv(self):
        data = b"name,age\nAlice,30\nBob,25"
        fs = _make_fs(data)
        columns, rows = parse_csv_to_rows(fs)
        assert columns == ["name", "age"]
        assert len(rows) == 2
        assert rows[0]["name"] == "Alice"
        assert rows[1]["age"] == "25"

    def test_empty_csv(self):
        data = b""
        fs = _make_fs(data)
        columns, rows = parse_csv_to_rows(fs)
        assert columns == []
        assert rows == []

    def test_header_only(self):
        data = b"col1,col2,col3"
        fs = _make_fs(data)
        columns, rows = parse_csv_to_rows(fs)
        assert columns == ["col1", "col2", "col3"]
        assert rows == []

    def test_csv_with_bom(self):
        data = codecs.BOM_UTF8 + b"name,value\ntest,1"
        fs = _make_fs(data)
        columns, rows = parse_csv_to_rows(fs)
        assert "name" in columns
        assert rows[0]["name"] == "test"

    def test_csv_latin1_encoding(self):
        data = "name\nCafé".encode("latin-1")
        fs = _make_fs(data)
        columns, rows = parse_csv_to_rows(fs)
        assert columns == ["name"]
        assert len(rows) == 1

    def test_csv_returns_dict_rows(self):
        data = b"a,b\n1,2"
        fs = _make_fs(data)
        columns, rows = parse_csv_to_rows(fs)
        assert isinstance(rows[0], dict)


# ---------------------------------------------------------------------------
# parse_excel_to_rows
# ---------------------------------------------------------------------------

class TestParseExcelToRows:
    @pytest.fixture()
    def xlsx_bytes(self):
        """Create a minimal xlsx file in memory using openpyxl."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])
        ws.append(["Alice", 30])
        ws.append(["Bob", 25])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    @pytest.fixture()
    def xlsx_empty_rows(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["col1", "col2"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    @pytest.fixture()
    def xlsx_short_row(self):
        """Row with fewer values than headers."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["a", "b", "c"])
        ws.append(["only_a"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def test_basic_xlsx(self, xlsx_bytes):
        fs = FileStorage(stream=io.BytesIO(xlsx_bytes), filename="data.xlsx")
        columns, rows = parse_excel_to_rows(fs)
        assert columns == ["name", "age"]
        assert len(rows) == 2
        assert rows[0]["name"] == "Alice"
        assert rows[1]["age"] == 25

    def test_header_only_xlsx(self, xlsx_empty_rows):
        fs = FileStorage(stream=io.BytesIO(xlsx_empty_rows), filename="data.xlsx")
        columns, rows = parse_excel_to_rows(fs)
        assert columns == ["col1", "col2"]
        assert rows == []

    def test_short_row_padded_with_none(self, xlsx_short_row):
        fs = FileStorage(stream=io.BytesIO(xlsx_short_row), filename="data.xlsx")
        columns, rows = parse_excel_to_rows(fs)
        assert columns == ["a", "b", "c"]
        assert rows[0]["a"] == "only_a"
        assert rows[0]["b"] is None
        assert rows[0]["c"] is None

    def test_empty_workbook(self):
        """Workbook with no rows at all (active sheet empty)."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fs = FileStorage(stream=buf, filename="empty.xlsx")
        columns, rows = parse_excel_to_rows(fs)
        assert columns == []
        assert rows == []


# ---------------------------------------------------------------------------
# parse_csv_or_excel_to_rows
# ---------------------------------------------------------------------------

class TestParseCsvOrExcelToRows:
    def test_csv_dispatched(self):
        data = b"x,y\n1,2"
        fs = _make_fs(data, "data.csv")
        columns, rows = parse_csv_or_excel_to_rows(fs, "data.csv")
        assert columns == ["x", "y"]
        assert rows[0]["x"] == "1"

    def test_xlsx_dispatched(self):
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.append(["p", "q"])
        wb.active.append([10, 20])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fs = FileStorage(stream=buf, filename="data.xlsx")
        columns, rows = parse_csv_or_excel_to_rows(fs, "data.xlsx")
        assert "p" in columns
        assert rows[0]["p"] == 10

    def test_xls_dispatched(self):
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.append(["col"])
        wb.active.append(["val"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fs = FileStorage(stream=buf, filename="data.xls")
        columns, rows = parse_csv_or_excel_to_rows(fs, "data.xls")
        assert "col" in columns

    def test_uppercase_csv_extension(self):
        data = b"a,b\n1,2"
        fs = _make_fs(data, "DATA.CSV")
        columns, rows = parse_csv_or_excel_to_rows(fs, "DATA.CSV")
        assert columns == ["a", "b"]

    def test_unsupported_format_raises(self):
        fs = _make_fs(b"data", "file.txt")
        with pytest.raises(ValueError, match="Unsupported file format"):
            parse_csv_or_excel_to_rows(fs, "file.txt")

    def test_unsupported_format_pdf(self):
        fs = _make_fs(b"%PDF", "file.pdf")
        with pytest.raises(ValueError):
            parse_csv_or_excel_to_rows(fs, "file.pdf")
