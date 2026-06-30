from contextlib import suppress
import json
import logging
import os
import threading
import uuid as _uuid_mod

from flask import current_app, flash, redirect, render_template, request, url_for
from flask_babel import _
from flask_login import current_user
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from app import db
from app.models import (
    CommonWord,
    IndicatorBank,
    IndicatorBankHistory,
    IndicatorBankType,
    IndicatorBankUnit,
    IndicatorSuggestion,
    Sector,
    SubSector,
)
from app.models.enums import IndicatorSuggestionStatusValue
from app.routes.admin.shared import permission_required
from app.routes.admin.system_admin.helpers import indicator_bank_history_snapshot
from app.routes.admin.utilities import bp
from app.utils.advanced_validation import validate_upload_extension_and_mime
from app.utils.api_helpers import GENERIC_ERROR_MESSAGE
from app.utils.api_responses import json_bad_request, json_ok
from app.utils.datetime_helpers import utcnow
from app.utils.error_handling import handle_json_view_exception
from app.utils.file_parsing import EXCEL_EXTENSIONS
from app.utils.transactions import request_transaction_rollback
from app.services.indicator_measurement_sync import backfill_fk_from_strings_bank

logger = logging.getLogger(__name__)


def _get_import_temp_path(token: str) -> str:
    """Return the expected temp file path for a given import token."""
    from app.utils.file_paths import get_temp_upload_path
    temp_dir = get_temp_upload_path()
    os.makedirs(temp_dir, exist_ok=True)
    return os.path.join(temp_dir, f"import_{token}.xlsx")

# === Import/Export Routes ===

@bp.route("/indicator_bank/import", methods=["GET", "POST"])
@permission_required('admin.indicator_bank.edit')
def import_indicators():
    """Legacy entry-point – redirects GET, delegates POST to the preview step."""
    if request.method == 'POST':
        # Forward to preview for backward compatibility
        return preview_indicator_import()
    return redirect(url_for("system_admin.manage_indicator_bank"))


@bp.route("/indicator_bank/import/preview", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def preview_indicator_import():
    """Step 1 of two-step import: parse the file and return a change summary.

    Saves the upload to a temp file keyed by a UUID token so the user can
    confirm and then call /apply without re-uploading.
    """
    temp_path = None
    try:
        if 'file' not in request.files:
            return json_bad_request('No file selected.')

        file = request.files['file']
        if not file or file.filename == '':
            return json_bad_request('No file selected.')

        valid, error_msg, ext = validate_upload_extension_and_mime(file, EXCEL_EXTENSIONS)
        if not valid:
            return json_bad_request(error_msg or 'Please upload an Excel file (.xlsx or .xls).')

        token = str(_uuid_mod.uuid4())
        temp_path = _get_import_temp_path(token)
        file.save(temp_path)

        current_app.logger.info("Import preview: saved to %s", temp_path)
        summary = _preview_indicator_import(temp_path)

        return json_ok(import_token=token, summary=summary)

    except Exception as e:
        if temp_path and os.path.exists(temp_path):
            with suppress(Exception):
                os.remove(temp_path)
        return handle_json_view_exception(e, 'Error processing import file.', status_code=500)


@bp.route("/indicator_bank/import/apply", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def apply_indicator_import():
    """Step 2 of two-step import: apply the changes using the token from preview."""
    temp_path = None
    try:
        data = request.get_json(silent=True) or {}
        token = (data.get('import_token') or '').strip()
        if not token:
            return json_bad_request('No import token provided.')

        # Validate token format (security: prevent path traversal)
        try:
            _uuid_mod.UUID(token)
        except ValueError:
            return json_bad_request('Invalid import token.')

        temp_path = _get_import_temp_path(token)
        if not os.path.exists(temp_path):
            return json_bad_request('Import session expired or not found. Please upload the file again.')

        current_app.logger.info("Import apply: processing %s", temp_path)
        result = _process_indicator_import(temp_path)
        current_app.logger.info("Import apply result: %s", result)

        if not result['success']:
            return json_bad_request(f"Import failed: {result.get('message', 'Unknown error')}")

        parts = []
        if result['imported']:
            parts.append(f"{result['imported']} indicator(s) created")
        if result['updated']:
            parts.append(f"{result['updated']} indicator(s) updated")
        for key, label in [
            ("measurement_types_imported", "type(s) created"),
            ("measurement_types_updated", "type(s) updated"),
            ("measurement_units_imported", "unit(s) created"),
            ("measurement_units_updated", "unit(s) updated"),
            ("sectors_imported", "sector(s) created"),
            ("sectors_updated", "sector(s) updated"),
            ("subsectors_imported", "sub-sector(s) created"),
            ("subsectors_updated", "sub-sector(s) updated"),
            ("common_words_imported", "common word(s) created"),
            ("common_words_updated", "common word(s) updated"),
        ]:
            v = result.get(key) or 0
            if v:
                parts.append(f"{v} {label}")

        message = "Import completed. " + (", ".join(parts) if parts else "No changes applied.")
        if result.get('errors'):
            message += f" ({len(result['errors'])} row error(s) skipped.)"

        return json_ok(message=message)

    except Exception as e:
        return handle_json_view_exception(e, 'Error applying import.', status_code=500)
    finally:
        if temp_path and os.path.exists(temp_path):
            def _cleanup(p):
                import time
                time.sleep(2)
                with suppress(Exception):
                    os.remove(p)
            threading.Thread(target=_cleanup, args=(temp_path,), daemon=True).start()

@bp.route("/indicator_bank/change_history", methods=["GET"])
@permission_required('admin.audit.view')
def indicator_change_history():
    """View indicator change history"""
    # Get all change history records (no pagination - AG Grid handles pagination client-side)
    changes = IndicatorBankHistory.query.order_by(
        IndicatorBankHistory.created_at.desc()
    ).all()

    return render_template("admin/indicator_bank/change_history.html",
                         changes=changes,
                         title="Indicator Change History")

# === Indicator Suggestions Management ===
@bp.route("/indicator_suggestions", methods=["GET"])
@permission_required('admin.indicator_bank.suggestions.review')
def manage_indicator_suggestions():
    """Manage indicator suggestions from users"""
    from app.utils.api_pagination import validate_pagination_params
    page, per_page = validate_pagination_params(request.args, default_per_page=20, max_per_page=100)
    status_filter = request.args.get('status', '')
    suggestion_type_filter = request.args.get('suggestion_type', '')

    query = IndicatorSuggestion.query

    if status_filter:
        query = query.filter(IndicatorSuggestion.status == status_filter)

    pagination = query.order_by(
        IndicatorSuggestion.submitted_at.desc()
    ).paginate(page=page, per_page=per_page, error_out=False)

    return render_template(
        "admin/indicator_suggestions.html",
        suggestions=pagination.items,
        pagination=pagination,
        status_filter=status_filter,
        suggestion_type_filter=suggestion_type_filter,
        title="Manage Indicator Suggestions",
    )

@bp.route("/indicator_suggestions/view/<int:suggestion_id>", methods=["GET"])
@permission_required('admin.indicator_bank.suggestions.review')
def view_indicator_suggestion(suggestion_id):
    """View individual indicator suggestion"""
    suggestion = IndicatorSuggestion.query.get_or_404(suggestion_id)
    return render_template("admin/view_indicator_suggestion.html",
                         suggestion=suggestion,
                         title=f"View Suggestion: {suggestion.indicator_name}")

@bp.route("/indicator_suggestions/update_status/<int:suggestion_id>", methods=["POST"])
@permission_required('admin.indicator_bank.suggestions.review')
def update_indicator_suggestion_status(suggestion_id):
    """Update indicator suggestion status"""
    suggestion = IndicatorSuggestion.query.get_or_404(suggestion_id)
    new_status = request.form.get('status')
    admin_notes = request.form.get('admin_notes', '').strip()

    try:
        if new_status in ['pending', 'approved', 'rejected', 'implemented', 'pending', 'reviewed']:
            suggestion.status = IndicatorSuggestionStatusValue.normalize(new_status)
            suggestion.admin_notes = admin_notes
            suggestion.reviewed_by = current_user
            suggestion.reviewed_at = utcnow()

            # If approved, optionally create the indicator automatically
            if new_status == 'approved':
                _create_indicator_from_suggestion(suggestion)

            db.session.flush()
            flash(_("Suggestion status updated to %(status)s.", status=new_status), "success")
        else:
            flash(_("Invalid status provided."), "danger")

    except Exception as e:
        request_transaction_rollback()
        current_app.logger.error(f"Error updating suggestion status: {e}", exc_info=True)
        flash(_("Error updating suggestion status."), "danger")

    return redirect(url_for("utilities.view_indicator_suggestion", suggestion_id=suggestion_id))

@bp.route("/indicator_suggestions/delete/<int:suggestion_id>", methods=["POST"])
@permission_required('admin.indicator_bank.suggestions.review')
def delete_indicator_suggestion(suggestion_id):
    """Delete indicator suggestion"""
    suggestion = IndicatorSuggestion.query.get_or_404(suggestion_id)

    try:
        db.session.delete(suggestion)
        db.session.flush()
        flash(_("Suggestion deleted successfully."), "success")

    except Exception as e:
        request_transaction_rollback()
        current_app.logger.error(f"Error deleting suggestion: {e}", exc_info=True)
        flash(_("Error deleting suggestion."), "danger")

    return redirect(url_for("utilities.manage_indicator_suggestions"))

# === Helper Functions ===

# ── Shared parsing utilities ──────────────────────────────────────────────────

def _norm_header(v):
    return str(v).strip().lower() if v is not None else ""


def _to_int(v):
    try:
        if v is None or v == "":
            return None
        if isinstance(v, bool):
            return int(v)
        if isinstance(v, (int, float)):
            return int(v)
        s = str(v).strip()
        return int(float(s)) if s else None
    except Exception as e:
        logger.debug("int parse failed for %r: %s", v, e)
        return None


def _to_bool(v):
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(int(v))
    s = str(v).strip().lower()
    if s in ("true", "t", "yes", "y", "1"):
        return True
    if s in ("false", "f", "no", "n", "0"):
        return False
    return None


def _to_json_dict(v):
    if v is None or v == "":
        return None
    if isinstance(v, dict):
        return v
    try:
        parsed = json.loads(str(v))
        return parsed if isinstance(parsed, dict) else None
    except Exception as e:
        logger.debug("json parse failed for %r: %s", v, e)
        return None


def _sheet_rows_as_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    raw_headers = list(rows[0])
    headers = [_norm_header(h) for h in raw_headers]
    out = []
    for r in rows[1:]:
        if not r:
            continue
        if all((c is None or str(c).strip() == "") for c in r):
            continue
        d = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            if idx < len(r):
                d[key] = r[idx]
        out.append(d)
    return out


def _build_levels_json(primary_id, secondary_id, tertiary_id):
    data = {}
    if primary_id is not None:
        data["primary"] = primary_id
    if secondary_id is not None:
        data["secondary"] = secondary_id
    if tertiary_id is not None:
        data["tertiary"] = tertiary_id
    return data or None


def _mq_str_to_list(raw):
    """Parse semicolon-separated monitoring questions string to list."""
    if not raw or not str(raw).strip():
        return None
    return [q.strip() for q in str(raw).split(";") if q.strip()] or None


def _tags_str_to_list(raw):
    """Parse comma-separated tags string to list."""
    if not raw or not str(raw).strip():
        return None
    return [t.strip() for t in str(raw).split(",") if t.strip()] or None


# ── Preview (dry-run analysis) ────────────────────────────────────────────────

def _preview_indicator_import(file_path):
    """Parse the import file and return a change-summary dict without touching the DB."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheetnames = set(wb.sheetnames or [])

    summary = {
        "indicators":    {"to_create": 0, "to_update": 0, "sample": []},
        "types":         {"to_create": 0, "to_update": 0},
        "units":         {"to_create": 0, "to_update": 0},
        "sectors":       {"to_create": 0, "to_update": 0},
        "subsectors":    {"to_create": 0, "to_update": 0},
        "common_words":  {"to_create": 0, "to_update": 0},
    }

    # Types ──────────────────────────────────────────────────────────────────
    for sname in ("Types", "DB_MeasurementTypes"):
        if sname not in sheetnames:
            continue
        for r in _sheet_rows_as_dicts(wb[sname]):
            rid  = _to_int(r.get("id"))
            code = (r.get("code") or "").strip().lower()
            if not rid and not code:
                continue
            obj = IndicatorBankType.query.get(rid) if rid else None
            if obj is None and code:
                obj = IndicatorBankType.query.filter(
                    db.func.lower(IndicatorBankType.code) == code
                ).first()
            if obj:
                summary["types"]["to_update"] += 1
            else:
                summary["types"]["to_create"] += 1
        break

    # Units ──────────────────────────────────────────────────────────────────
    for sname in ("Units", "DB_MeasurementUnits"):
        if sname not in sheetnames:
            continue
        for r in _sheet_rows_as_dicts(wb[sname]):
            rid  = _to_int(r.get("id"))
            code = (r.get("code") or "").strip().lower()
            if not rid and not code:
                continue
            obj = IndicatorBankUnit.query.get(rid) if rid else None
            if obj is None and code:
                obj = IndicatorBankUnit.query.filter(
                    db.func.lower(IndicatorBankUnit.code) == code
                ).first()
            if obj:
                summary["units"]["to_update"] += 1
            else:
                summary["units"]["to_create"] += 1
        break

    # Sectors ────────────────────────────────────────────────────────────────
    if "Sectors" in sheetnames:
        for r in _sheet_rows_as_dicts(wb["Sectors"]):
            rid  = _to_int(r.get("id"))
            name = (r.get("name") or "").strip()
            if not rid and not name:
                continue
            obj = Sector.query.get(rid) if rid else Sector.query.filter_by(name=name).first()
            if obj:
                summary["sectors"]["to_update"] += 1
            else:
                summary["sectors"]["to_create"] += 1
    elif "DB_Sectors_SubSectors" in sheetnames:
        for r in _sheet_rows_as_dicts(wb["DB_Sectors_SubSectors"]):
            rid  = _to_int(r.get("id"))
            name = (r.get("name") or "").strip()
            rt   = (r.get("record_type") or "").strip().lower()
            if rt == "sector":
                obj = Sector.query.get(rid) if rid else Sector.query.filter_by(name=name).first()
                if obj:
                    summary["sectors"]["to_update"] += 1
                else:
                    summary["sectors"]["to_create"] += 1
            elif rt == "subsector":
                obj = SubSector.query.get(rid) if rid else SubSector.query.filter_by(name=name).first()
                if obj:
                    summary["subsectors"]["to_update"] += 1
                else:
                    summary["subsectors"]["to_create"] += 1

    # Sub-Sectors ────────────────────────────────────────────────────────────
    if "Sub-Sectors" in sheetnames:
        for r in _sheet_rows_as_dicts(wb["Sub-Sectors"]):
            rid  = _to_int(r.get("id"))
            name = (r.get("name") or "").strip()
            if not rid and not name:
                continue
            obj = SubSector.query.get(rid) if rid else SubSector.query.filter_by(name=name).first()
            if obj:
                summary["subsectors"]["to_update"] += 1
            else:
                summary["subsectors"]["to_create"] += 1

    # Common Words ───────────────────────────────────────────────────────────
    for sname in ("Common Words", "DB_CommonWords"):
        if sname not in sheetnames:
            continue
        for r in _sheet_rows_as_dicts(wb[sname]):
            rid  = _to_int(r.get("id"))
            term = (r.get("term") or "").strip()
            if not rid and not term:
                continue
            obj = CommonWord.query.get(rid) if rid else CommonWord.query.filter_by(term=term).first()
            if obj:
                summary["common_words"]["to_update"] += 1
            else:
                summary["common_words"]["to_create"] += 1
        break

    # Indicators ─────────────────────────────────────────────────────────────
    SAMPLE_LIMIT = 25
    if "DB_Indicators" in sheetnames:
        for r in _sheet_rows_as_dicts(wb["DB_Indicators"]):
            rid  = _to_int(r.get("id"))
            name = (r.get("name") or "").strip()
            if not name and not rid:
                continue
            obj    = IndicatorBank.query.get(rid) if rid else IndicatorBank.query.filter_by(name=name).first()
            action = "update" if obj else "create"
            summary["indicators"]["to_" + action] += 1
            if len(summary["indicators"]["sample"]) < SAMPLE_LIMIT:
                summary["indicators"]["sample"].append({"name": name or f"ID:{rid}", "action": action})
    else:
        ws_ind = wb.active
        rows_raw = list(ws_ind.iter_rows(values_only=True))
        if rows_raw:
            headers = [_norm_header(h) for h in rows_raw[0]]
            hmap    = {h: i for i, h in enumerate(headers) if h}
            name_idx = hmap.get("name", 1)
            for r in rows_raw[1:]:
                if not r or all((c is None or str(c).strip() == "") for c in r):
                    continue
                if name_idx >= len(r):
                    continue
                name = str(r[name_idx] or "").strip()
                if not name:
                    continue
                obj    = IndicatorBank.query.filter_by(name=name).first()
                action = "update" if obj else "create"
                summary["indicators"]["to_" + action] += 1
                if len(summary["indicators"]["sample"]) < SAMPLE_LIMIT:
                    summary["indicators"]["sample"].append({"name": name, "action": action})

    return summary


# ── Full apply ────────────────────────────────────────────────────────────────

def _process_indicator_import(file_path):
    """Process Excel file for indicator import – writes to the DB."""
    try:
        current_app.logger.info("Loading workbook from: %s", file_path)
        wb = load_workbook(file_path, read_only=True, data_only=True)

        result = {
            'success': True,
            'imported': 0,
            'updated': 0,
            'sectors_imported': 0,
            'sectors_updated': 0,
            'subsectors_imported': 0,
            'subsectors_updated': 0,
            'common_words_imported': 0,
            'common_words_updated': 0,
            'measurement_types_imported': 0,
            'measurement_types_updated': 0,
            'measurement_units_imported': 0,
            'measurement_units_updated': 0,
            'errors': [],
            'message': '',
        }

        sheetnames = set(wb.sheetnames or [])

        # ── Branch A: legacy hidden DB_* sheets ──────────────────────────────
        if (
            "DB_Indicators" in sheetnames
            or "DB_Sectors_SubSectors" in sheetnames
            or "DB_CommonWords" in sheetnames
            or "DB_MeasurementTypes" in sheetnames
            or "DB_MeasurementUnits" in sheetnames
        ):
            _import_db_sectors_subsectors(wb, sheetnames, result)
            _import_db_common_words(wb, sheetnames, result)
            _import_db_measurement_types(wb, sheetnames, result)
            _import_db_measurement_units(wb, sheetnames, result)
            _import_db_indicators(wb, sheetnames, result)
            db.session.flush()
            return result

        # ── Branch B: new human-readable lookup sheets ────────────────────────
        # Process lookup sheets first (order matters: sectors before sub-sectors,
        # types/units before indicators so FK backfill can resolve them).
        _import_types_sheet(wb, sheetnames, result)
        _import_units_sheet(wb, sheetnames, result)
        _import_sectors_sheet(wb, sheetnames, result)
        _import_subsectors_sheet(wb, sheetnames, result)
        _import_common_words_sheet(wb, sheetnames, result)
        # Flush so that sector/unit/type PKs are available for indicator backfill.
        db.session.flush()

        # ── Branch B/C: main Indicators sheet (same logic for both new & plain)
        _import_indicators_main_sheet(wb, result)

        db.session.flush()

    except Exception as e:
        result['success'] = False
        result['message'] = GENERIC_ERROR_MESSAGE
        request_transaction_rollback()

    return result


# ── Sub-routines for each sheet type ─────────────────────────────────────────

def _import_db_sectors_subsectors(wb, sheetnames, result):
    if "DB_Sectors_SubSectors" not in sheetnames:
        return
    from sqlalchemy.orm.attributes import flag_modified as _flag
    for idx, r in enumerate(_sheet_rows_as_dicts(wb["DB_Sectors_SubSectors"]), 2):
        try:
            rt           = (r.get("record_type") or "").strip().lower()
            rid          = _to_int(r.get("id"))
            name         = (r.get("name") or "").strip()
            description  = r.get("description") or None
            sector_id    = _to_int(r.get("sector_id"))
            display_order= _to_int(r.get("display_order"))
            is_active    = _to_bool(r.get("is_active"))
            icon_class   = r.get("icon_class") or None
            logo_filename= r.get("logo_filename") or None
            name_translations = _to_json_dict(r.get("name_translations_json")) or {}

            if rt == "sector":
                if not rid and not name:
                    continue
                obj = Sector.query.get(rid) if rid else Sector.query.filter_by(name=name).first()
                is_new = obj is None
                if is_new:
                    obj = Sector()
                    if rid:
                        obj.id = rid
                    db.session.add(obj)
                if name:
                    obj.name = name
                obj.description = description
                if display_order is not None:
                    obj.display_order = display_order
                if is_active is not None:
                    obj.is_active = is_active
                obj.icon_class = icon_class
                obj.logo_filename = logo_filename
                obj.name_translations = name_translations
                result["sectors_imported" if is_new else "sectors_updated"] += 1

            elif rt == "subsector":
                if not rid and not name:
                    continue
                obj = SubSector.query.get(rid) if rid else SubSector.query.filter_by(name=name).first()
                is_new = obj is None
                if is_new:
                    obj = SubSector()
                    if rid:
                        obj.id = rid
                    db.session.add(obj)
                if name:
                    obj.name = name
                obj.description = description
                if sector_id is not None:
                    obj.sector_id = sector_id
                if display_order is not None:
                    obj.display_order = display_order
                if is_active is not None:
                    obj.is_active = is_active
                obj.icon_class = icon_class
                obj.name_translations = name_translations
                result["subsectors_imported" if is_new else "subsectors_updated"] += 1
        except Exception as e:
            current_app.logger.error("Error processing DB_Sectors_SubSectors row %d: %s", idx, e, exc_info=True)
            result["errors"].append(f"DB_Sectors_SubSectors row {idx}: error.")


def _import_db_common_words(wb, sheetnames, result):
    if "DB_CommonWords" not in sheetnames:
        return
    for idx, r in enumerate(_sheet_rows_as_dicts(wb["DB_CommonWords"]), 2):
        try:
            rid     = _to_int(r.get("id"))
            term    = (r.get("term") or "").strip()
            meaning = (r.get("meaning") or "").strip()
            if not term and not rid:
                continue
            is_active = _to_bool(r.get("is_active"))
            meaning_translations = _to_json_dict(r.get("meaning_translations_json")) or {}

            obj = CommonWord.query.get(rid) if rid else CommonWord.query.filter_by(term=term).first()
            is_new = obj is None
            if is_new:
                obj = CommonWord()
                if rid:
                    obj.id = rid
                db.session.add(obj)
            if term:
                obj.term = term
            if meaning:
                obj.meaning = meaning
            if is_active is not None:
                obj.is_active = is_active
            obj.meaning_translations = meaning_translations
            result["common_words_imported" if is_new else "common_words_updated"] += 1
        except Exception as e:
            current_app.logger.error("Error processing DB_CommonWords row %d: %s", idx, e, exc_info=True)
            result["errors"].append(f"DB_CommonWords row {idx}: error.")


def _import_db_measurement_types(wb, sheetnames, result):
    if "DB_MeasurementTypes" not in sheetnames:
        return
    from sqlalchemy.orm.attributes import flag_modified as _flag
    for idx, r in enumerate(_sheet_rows_as_dicts(wb["DB_MeasurementTypes"]), 2):
        try:
            rid  = _to_int(r.get("id"))
            code = (r.get("code") or "").strip().lower()
            name = (r.get("name") or "").strip()
            name_translations = _to_json_dict(r.get("name_translations_json")) or {}
            sort_order = _to_int(r.get("sort_order"))
            is_active  = _to_bool(r.get("is_active"))
            if not code and not rid:
                continue
            obj = IndicatorBankType.query.get(rid) if rid else None
            if obj is None and code:
                obj = IndicatorBankType.query.filter(
                    db.func.lower(IndicatorBankType.code) == code
                ).first()
            is_new = obj is None
            if is_new:
                obj = IndicatorBankType()
                if rid:
                    obj.id = rid
                db.session.add(obj)
            if code:
                obj.code = code
            if name:
                obj.name = name
            obj.name_translations = name_translations
            _flag(obj, "name_translations")
            if sort_order is not None:
                obj.sort_order = sort_order
            if is_active is not None:
                obj.is_active = is_active
            result["measurement_types_imported" if is_new else "measurement_types_updated"] += 1
        except Exception as e:
            current_app.logger.error("Error processing DB_MeasurementTypes row %d: %s", idx, e, exc_info=True)
            result["errors"].append(f"DB_MeasurementTypes row {idx}: error.")


def _import_db_measurement_units(wb, sheetnames, result):
    if "DB_MeasurementUnits" not in sheetnames:
        return
    from sqlalchemy.orm.attributes import flag_modified as _flag
    for idx, r in enumerate(_sheet_rows_as_dicts(wb["DB_MeasurementUnits"]), 2):
        try:
            rid  = _to_int(r.get("id"))
            code = (r.get("code") or "").strip().lower()
            name = (r.get("name") or "").strip()
            name_translations = _to_json_dict(r.get("name_translations_json")) or {}
            sort_order            = _to_int(r.get("sort_order"))
            is_active             = _to_bool(r.get("is_active"))
            allows_disaggregation = _to_bool(r.get("allows_disaggregation"))
            if not code and not rid:
                continue
            obj = IndicatorBankUnit.query.get(rid) if rid else None
            if obj is None and code:
                obj = IndicatorBankUnit.query.filter(
                    db.func.lower(IndicatorBankUnit.code) == code
                ).first()
            is_new = obj is None
            if is_new:
                obj = IndicatorBankUnit()
                if rid:
                    obj.id = rid
                db.session.add(obj)
            if code:
                obj.code = code
            if name:
                obj.name = name
            obj.name_translations = name_translations
            _flag(obj, "name_translations")
            if sort_order is not None:
                obj.sort_order = sort_order
            if is_active is not None:
                obj.is_active = is_active
            if allows_disaggregation is not None:
                obj.allows_disaggregation = allows_disaggregation
            result["measurement_units_imported" if is_new else "measurement_units_updated"] += 1
        except Exception as e:
            current_app.logger.error("Error processing DB_MeasurementUnits row %d: %s", idx, e, exc_info=True)
            result["errors"].append(f"DB_MeasurementUnits row {idx}: error.")


def _import_db_indicators(wb, sheetnames, result):
    if "DB_Indicators" not in sheetnames:
        return
    for idx, r in enumerate(_sheet_rows_as_dicts(wb["DB_Indicators"]), 2):
        try:
            rid  = _to_int(r.get("id"))
            name = (r.get("name") or "").strip()
            if not name and not rid:
                continue

            definition             = (r.get("definition") or "").strip()
            aggregated_label       = (r.get("aggregated_label") or "").strip() or None
            area                   = (r.get("area") or "").strip() or None
            data_source            = (r.get("data_source") or "").strip() or None
            disaggregation_guidance= (r.get("disaggregation_guidance") or "").strip() or None
            monitoring_questions   = (
                _to_json_dict(r.get("monitoring_questions_json"))
                or _to_json_dict(r.get("monitoring_questions"))
            )
            tags                   = (
                _to_json_dict(r.get("tags_json"))
                or _to_json_dict(r.get("tags"))
            )
            indicator_type  = (r.get("type") or "").strip() or "numeric"
            unit            = (r.get("unit") or "").strip()
            fdrs_kpi_code   = (r.get("fdrs_kpi_code") or "").strip() or None
            emergency       = _to_bool(r.get("emergency"))
            archived        = _to_bool(r.get("archived"))
            comments        = r.get("comments") or None
            programs        = str(r.get("related_programs") or "").strip()

            s_primary   = _to_int(r.get("sector_primary_id"))
            s_secondary = _to_int(r.get("sector_secondary_id"))
            s_tertiary  = _to_int(r.get("sector_tertiary_id"))
            ss_primary  = _to_int(r.get("subsector_primary_id"))
            ss_secondary= _to_int(r.get("subsector_secondary_id"))
            ss_tertiary = _to_int(r.get("subsector_tertiary_id"))

            name_translations             = _to_json_dict(r.get("name_translations_json")) or {}
            definition_translations       = _to_json_dict(r.get("definition_translations_json")) or {}
            aggregated_label_translations = _to_json_dict(r.get("aggregated_label_translations_json")) or {}
            itid = _to_int(r.get("indicator_type_id"))
            iuid = _to_int(r.get("indicator_unit_id"))

            existing = IndicatorBank.query.get(rid) if rid else IndicatorBank.query.filter_by(name=name).first()
            is_new   = existing is None

            if is_new:
                existing = IndicatorBank(
                    name=name, definition=definition, type=indicator_type, unit=unit,
                    fdrs_kpi_code=fdrs_kpi_code,
                    emergency=bool(emergency) if emergency is not None else False,
                    related_programs=programs, aggregated_label=aggregated_label,
                    area=area, data_source=data_source,
                    disaggregation_guidance=disaggregation_guidance,
                    monitoring_questions=monitoring_questions if isinstance(monitoring_questions, list) else None,
                    tags=tags if isinstance(tags, list) else None,
                )
                if rid:
                    existing.id = rid
                db.session.add(existing)
                db.session.flush()
            else:
                if name:
                    existing.name = name
                existing.definition = definition
                existing.type       = indicator_type
                existing.unit       = unit
                existing.fdrs_kpi_code = fdrs_kpi_code
                if emergency is not None:
                    existing.emergency = emergency
                if archived is not None:
                    existing.archived = archived
                existing.comments   = comments
                existing.related_programs = programs
                existing.aggregated_label = aggregated_label
                existing.area       = area
                existing.data_source = data_source
                existing.disaggregation_guidance = disaggregation_guidance
                existing.monitoring_questions = (
                    monitoring_questions if isinstance(monitoring_questions, list) else None
                )
                existing.tags = tags if isinstance(tags, list) else None

            existing.sector                       = _build_levels_json(s_primary, s_secondary, s_tertiary)
            existing.sub_sector                   = _build_levels_json(ss_primary, ss_secondary, ss_tertiary)
            existing.name_translations            = name_translations or {}
            existing.definition_translations      = definition_translations or {}
            existing.aggregated_label_translations= aggregated_label_translations or {}

            backfill_fk_from_strings_bank(existing)
            if itid is not None:
                existing.indicator_type_id = itid
            if iuid is not None:
                existing.indicator_unit_id = iuid
            if hasattr(existing, "sync_type_unit_string_columns"):
                existing.sync_type_unit_string_columns()

            history = IndicatorBankHistory(
                indicator_bank_id=existing.id,
                user_id=current_user.id,
                change_type='CREATED' if is_new else 'UPDATED',
                change_description=(
                    f'Indicator "{existing.name}" created via import by {current_user.name or current_user.email}'
                    if is_new else
                    f'Indicator "{existing.name}" updated via import by {current_user.name or current_user.email}'
                ),
                **indicator_bank_history_snapshot(existing),
            )
            db.session.add(history)
            result["imported" if is_new else "updated"] += 1
        except Exception as e:
            current_app.logger.error("Error processing DB_Indicators row %d: %s", idx, e, exc_info=True)
            result["errors"].append(f"DB_Indicators row {idx}: error.")


# ── New human-readable sheet importers ───────────────────────────────────────

def _import_types_sheet(wb, sheetnames, result):
    """Import from the 'Types' sheet (new human-readable format)."""
    if "Types" not in sheetnames:
        return
    from sqlalchemy.orm.attributes import flag_modified as _flag
    for idx, r in enumerate(_sheet_rows_as_dicts(wb["Types"]), 2):
        try:
            rid       = _to_int(r.get("id"))
            code      = (r.get("code") or "").strip().lower()
            name      = (r.get("name") or "").strip()
            is_active = _to_bool(r.get("active"))
            sort_order= _to_int(r.get("sort order") or r.get("sort_order"))
            if not code and not rid:
                continue

            # Collect translations: columns like "name (fr)"
            name_translations = {}
            for key, val in r.items():
                if key.startswith("name (") and key.endswith(")"):
                    lang = key[6:-1].strip().lower()
                    if lang and lang != "en" and val and str(val).strip():
                        name_translations[lang] = str(val).strip()

            obj = IndicatorBankType.query.get(rid) if rid else None
            if obj is None and code:
                obj = IndicatorBankType.query.filter(
                    db.func.lower(IndicatorBankType.code) == code
                ).first()
            is_new = obj is None
            if is_new:
                obj = IndicatorBankType()
                if rid:
                    obj.id = rid
                db.session.add(obj)
            if code:
                obj.code = code
            if name:
                obj.name = name
            if name_translations:
                obj.name_translations = {**(obj.name_translations or {}), **name_translations}
                _flag(obj, "name_translations")
            if sort_order is not None:
                obj.sort_order = sort_order
            if is_active is not None:
                obj.is_active = is_active
            result["measurement_types_imported" if is_new else "measurement_types_updated"] += 1
        except Exception as e:
            current_app.logger.error("Error processing Types row %d: %s", idx, e, exc_info=True)
            result["errors"].append(f"Types row {idx}: error.")


def _import_units_sheet(wb, sheetnames, result):
    """Import from the 'Units' sheet (new human-readable format)."""
    if "Units" not in sheetnames:
        return
    from sqlalchemy.orm.attributes import flag_modified as _flag
    for idx, r in enumerate(_sheet_rows_as_dicts(wb["Units"]), 2):
        try:
            rid                   = _to_int(r.get("id"))
            code                  = (r.get("code") or "").strip().lower()
            name                  = (r.get("name") or "").strip()
            is_active             = _to_bool(r.get("active"))
            sort_order            = _to_int(r.get("sort order") or r.get("sort_order"))
            allows_disaggregation = _to_bool(r.get("allows disaggregation") or r.get("allows_disaggregation"))
            if not code and not rid:
                continue

            name_translations = {}
            for key, val in r.items():
                if key.startswith("name (") and key.endswith(")"):
                    lang = key[6:-1].strip().lower()
                    if lang and lang != "en" and val and str(val).strip():
                        name_translations[lang] = str(val).strip()

            obj = IndicatorBankUnit.query.get(rid) if rid else None
            if obj is None and code:
                obj = IndicatorBankUnit.query.filter(
                    db.func.lower(IndicatorBankUnit.code) == code
                ).first()
            is_new = obj is None
            if is_new:
                obj = IndicatorBankUnit()
                if rid:
                    obj.id = rid
                db.session.add(obj)
            if code:
                obj.code = code
            if name:
                obj.name = name
            if name_translations:
                obj.name_translations = {**(obj.name_translations or {}), **name_translations}
                _flag(obj, "name_translations")
            if sort_order is not None:
                obj.sort_order = sort_order
            if is_active is not None:
                obj.is_active = is_active
            if allows_disaggregation is not None:
                obj.allows_disaggregation = allows_disaggregation
            result["measurement_units_imported" if is_new else "measurement_units_updated"] += 1
        except Exception as e:
            current_app.logger.error("Error processing Units row %d: %s", idx, e, exc_info=True)
            result["errors"].append(f"Units row {idx}: error.")


def _import_sectors_sheet(wb, sheetnames, result):
    """Import from the 'Sectors' sheet (new human-readable format)."""
    if "Sectors" not in sheetnames:
        return
    for idx, r in enumerate(_sheet_rows_as_dicts(wb["Sectors"]), 2):
        try:
            rid          = _to_int(r.get("id"))
            name         = (r.get("name") or "").strip()
            description  = r.get("description") or None
            is_active    = _to_bool(r.get("active"))
            display_order= _to_int(r.get("display order") or r.get("display_order"))
            icon_class   = r.get("icon class") or r.get("icon_class") or None
            if not rid and not name:
                continue

            name_translations = {}
            for key, val in r.items():
                if key.startswith("name (") and key.endswith(")"):
                    lang = key[6:-1].strip().lower()
                    if lang and lang != "en" and val and str(val).strip():
                        name_translations[lang] = str(val).strip()

            obj = Sector.query.get(rid) if rid else Sector.query.filter_by(name=name).first()
            is_new = obj is None
            if is_new:
                obj = Sector()
                if rid:
                    obj.id = rid
                db.session.add(obj)
            if name:
                obj.name = name
            if description is not None:
                obj.description = description
            if display_order is not None:
                obj.display_order = display_order
            if is_active is not None:
                obj.is_active = is_active
            if icon_class is not None:
                obj.icon_class = icon_class
            if name_translations:
                obj.name_translations = {**(obj.name_translations or {}), **name_translations}
            result["sectors_imported" if is_new else "sectors_updated"] += 1
        except Exception as e:
            current_app.logger.error("Error processing Sectors row %d: %s", idx, e, exc_info=True)
            result["errors"].append(f"Sectors row {idx}: error.")


def _import_subsectors_sheet(wb, sheetnames, result):
    """Import from the 'Sub-Sectors' sheet (new human-readable format)."""
    if "Sub-Sectors" not in sheetnames:
        return
    sector_name_to_id = {s.name: s.id for s in Sector.query.all()}
    for idx, r in enumerate(_sheet_rows_as_dicts(wb["Sub-Sectors"]), 2):
        try:
            rid          = _to_int(r.get("id"))
            name         = (r.get("name") or "").strip()
            sector_name  = (r.get("sector") or "").strip()
            description  = r.get("description") or None
            is_active    = _to_bool(r.get("active"))
            display_order= _to_int(r.get("display order") or r.get("display_order"))
            icon_class   = r.get("icon class") or r.get("icon_class") or None
            if not rid and not name:
                continue

            name_translations = {}
            for key, val in r.items():
                if key.startswith("name (") and key.endswith(")"):
                    lang = key[6:-1].strip().lower()
                    if lang and lang != "en" and val and str(val).strip():
                        name_translations[lang] = str(val).strip()

            obj = SubSector.query.get(rid) if rid else SubSector.query.filter_by(name=name).first()
            is_new = obj is None
            if is_new:
                obj = SubSector()
                if rid:
                    obj.id = rid
                db.session.add(obj)
            if name:
                obj.name = name
            if sector_name:
                sid = sector_name_to_id.get(sector_name)
                if sid:
                    obj.sector_id = sid
            if description is not None:
                obj.description = description
            if display_order is not None:
                obj.display_order = display_order
            if is_active is not None:
                obj.is_active = is_active
            if icon_class is not None:
                obj.icon_class = icon_class
            if name_translations:
                obj.name_translations = {**(obj.name_translations or {}), **name_translations}
            result["subsectors_imported" if is_new else "subsectors_updated"] += 1
        except Exception as e:
            current_app.logger.error("Error processing Sub-Sectors row %d: %s", idx, e, exc_info=True)
            result["errors"].append(f"Sub-Sectors row {idx}: error.")


def _import_common_words_sheet(wb, sheetnames, result):
    """Import from the 'Common Words' sheet (new human-readable format)."""
    if "Common Words" not in sheetnames:
        return
    for idx, r in enumerate(_sheet_rows_as_dicts(wb["Common Words"]), 2):
        try:
            rid     = _to_int(r.get("id"))
            term    = (r.get("term") or "").strip()
            meaning = (r.get("meaning") or "").strip()
            is_active = _to_bool(r.get("active"))
            if not rid and not term:
                continue

            meaning_translations = {}
            for key, val in r.items():
                if key.startswith("meaning (") and key.endswith(")"):
                    lang = key[9:-1].strip().lower()
                    if lang and lang != "en" and val and str(val).strip():
                        meaning_translations[lang] = str(val).strip()

            obj = CommonWord.query.get(rid) if rid else CommonWord.query.filter_by(term=term).first()
            is_new = obj is None
            if is_new:
                obj = CommonWord()
                if rid:
                    obj.id = rid
                db.session.add(obj)
            if term:
                obj.term = term
            if meaning:
                obj.meaning = meaning
            if is_active is not None:
                obj.is_active = is_active
            if meaning_translations:
                obj.meaning_translations = {**(obj.meaning_translations or {}), **meaning_translations}
            result["common_words_imported" if is_new else "common_words_updated"] += 1
        except Exception as e:
            current_app.logger.error("Error processing Common Words row %d: %s", idx, e, exc_info=True)
            result["errors"].append(f"Common Words row {idx}: error.")


def _import_indicators_main_sheet(wb, result):
    """Import indicators from the first/active sheet (Indicators)."""
    ws = wb.active
    headers_raw = [cell.value for cell in ws[1]]
    headers = [str(h).strip() if h is not None else "" for h in headers_raw]
    header_to_col = {}
    for idx, h in enumerate(headers):
        key = _norm_header(h) if h else ""
        if key and key not in header_to_col:
            header_to_col[key] = idx
    current_app.logger.info("Indicators sheet headers: %s", headers)

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    def _val(row, idx):
        if row is None or idx is None or idx < 0 or idx >= len(row):
            return None
        return row[idx]

    def _resolve_sector_id(name_val):
        if name_val is None or str(name_val).strip() == "":
            return None
        s = Sector.query.filter_by(name=str(name_val).strip()).first()
        return s.id if s else None

    def _resolve_subsector_id(name_val):
        if name_val is None or str(name_val).strip() == "":
            return None
        ss = SubSector.query.filter_by(name=str(name_val).strip()).first()
        return ss.id if ss else None

    for row_num, row in enumerate(rows, 2):
        try:
            if row is None or all((c is None or str(c).strip() == "") for c in row):
                continue

            name = (str(_val(row, header_to_col.get("name", 1)) or "")).strip()
            if not name:
                continue

            definition     = str(_val(row, header_to_col.get("definition", 2)) or "").strip()
            indicator_type = str(_val(row, header_to_col.get("type", 5)) or "").strip() or "numeric"
            unit           = str(_val(row, header_to_col.get("unit", 6)) or "").strip()
            fdrs_kpi_code  = (
                str(_val(row, header_to_col["fdrs kpi code"]) or "").strip() or None
            ) if "fdrs kpi code" in header_to_col else None
            aggregated_label = str(_val(row, header_to_col.get("aggregated label")) or "").strip() or None
            area           = str(_val(row, header_to_col.get("area")) or "").strip() or None
            data_source    = str(_val(row, header_to_col.get("data source")) or "").strip() or None
            disaggregation_guidance = str(_val(row, header_to_col.get("disaggregation guidance")) or "").strip() or None
            comments       = str(_val(row, header_to_col.get("comments")) or "").strip() or None

            # Monitoring questions: semicolon-separated string → list
            mq_raw         = _val(row, header_to_col.get("monitoring questions"))
            monitoring_questions = _mq_str_to_list(mq_raw)

            # Tags: comma-separated
            tags_raw       = _val(row, header_to_col.get("tags"))
            tags           = _tags_str_to_list(tags_raw)

            _emergency     = _to_bool(_val(row, header_to_col.get("emergency")))
            emergency      = _emergency if _emergency is not None else False
            _archived      = _to_bool(_val(row, header_to_col.get("archived")))
            programs       = str(_val(row, header_to_col.get("related programs")) or "").strip()

            # Localized names / definitions from "Name (xx)" columns
            name_translations        = {}
            definition_translations  = {}
            agg_label_translations   = {}
            for key, col_idx in header_to_col.items():
                val = _val(row, col_idx)
                if not val or not str(val).strip():
                    continue
                if key.startswith("name (") and key.endswith(")"):
                    lang = key[6:-1].strip().lower()
                    if lang and len(lang) <= 6:
                        name_translations[lang] = str(val).strip()
                elif key.startswith("definition (") and key.endswith(")"):
                    lang = key[12:-1].strip().lower()
                    if lang and len(lang) <= 6:
                        definition_translations[lang] = str(val).strip()
                elif key.startswith("aggregated label (") and key.endswith(")"):
                    lang = key[18:-1].strip().lower()
                    if lang and len(lang) <= 6:
                        agg_label_translations[lang] = str(val).strip()

            # Sector / subsector (resolved by name)
            s_primary   = _resolve_sector_id(_val(row, header_to_col.get("sector primary")))
            s_secondary = _resolve_sector_id(_val(row, header_to_col.get("sector secondary")))
            s_tertiary  = _resolve_sector_id(_val(row, header_to_col.get("sector tertiary")))
            ss_primary  = _resolve_subsector_id(_val(row, header_to_col.get("subsector primary")))
            ss_secondary= _resolve_subsector_id(_val(row, header_to_col.get("subsector secondary")))
            ss_tertiary = _resolve_subsector_id(_val(row, header_to_col.get("subsector tertiary")))
            sector_json     = _build_levels_json(s_primary, s_secondary, s_tertiary)
            sub_sector_json = _build_levels_json(ss_primary, ss_secondary, ss_tertiary)

            existing = IndicatorBank.query.filter_by(name=name).first()

            if existing:
                changes = []
                if existing.definition != definition:
                    od, nd = existing.definition or '', definition or ''
                    changes.append(f"Definition changed from '{od[:50]}...' to '{nd[:50]}...'")
                if existing.type != indicator_type:
                    changes.append(f"Type: {existing.type!r} → {indicator_type!r}")
                if existing.unit != unit:
                    changes.append(f"Unit: {existing.unit!r} → {unit!r}")
                if getattr(existing, 'fdrs_kpi_code', None) != fdrs_kpi_code:
                    changes.append(f"FDRS KPI Code: {existing.fdrs_kpi_code!r} → {fdrs_kpi_code!r}")
                if existing.emergency != emergency:
                    changes.append(f"Emergency: {existing.emergency} → {emergency}")

                existing.definition  = definition
                existing.type        = indicator_type
                existing.unit        = unit
                existing.fdrs_kpi_code = fdrs_kpi_code
                existing.emergency   = emergency
                existing.related_programs = programs
                if aggregated_label is not None:
                    existing.aggregated_label = aggregated_label
                if area is not None:
                    existing.area = area
                if data_source is not None:
                    existing.data_source = data_source
                if disaggregation_guidance is not None:
                    existing.disaggregation_guidance = disaggregation_guidance
                if monitoring_questions is not None:
                    existing.monitoring_questions = monitoring_questions
                if tags is not None:
                    existing.tags = tags
                if comments is not None:
                    existing.comments = comments
                if _archived is not None:
                    existing.archived = _archived
                if name_translations:
                    existing.name_translations = {**(existing.name_translations or {}), **name_translations}
                if definition_translations:
                    existing.definition_translations = {**(existing.definition_translations or {}), **definition_translations}
                if agg_label_translations:
                    existing.aggregated_label_translations = {**(existing.aggregated_label_translations or {}), **agg_label_translations}
                if sector_json is not None:
                    existing.sector = sector_json
                if sub_sector_json is not None:
                    existing.sub_sector = sub_sector_json

                backfill_fk_from_strings_bank(existing)

                change_description = "; ".join(changes) if changes else \
                    f"Indicator updated via import by {current_user.name or current_user.email}"
                history = IndicatorBankHistory(
                    indicator_bank_id=existing.id,
                    user_id=current_user.id,
                    change_type='UPDATED',
                    change_description=change_description,
                    **indicator_bank_history_snapshot(existing),
                )
                db.session.add(history)
                result['updated'] += 1

            else:
                new_indicator = IndicatorBank(
                    name=name, definition=definition, type=indicator_type, unit=unit,
                    fdrs_kpi_code=fdrs_kpi_code, emergency=emergency,
                    related_programs=programs, aggregated_label=aggregated_label,
                    area=area, data_source=data_source,
                    disaggregation_guidance=disaggregation_guidance,
                    monitoring_questions=monitoring_questions,
                    tags=tags, comments=comments,
                )
                if name_translations:
                    new_indicator.name_translations = name_translations
                if definition_translations:
                    new_indicator.definition_translations = definition_translations
                if agg_label_translations:
                    new_indicator.aggregated_label_translations = agg_label_translations
                if sector_json is not None:
                    new_indicator.sector = sector_json
                if sub_sector_json is not None:
                    new_indicator.sub_sector = sub_sector_json
                db.session.add(new_indicator)
                db.session.flush()
                backfill_fk_from_strings_bank(new_indicator)

                history = IndicatorBankHistory(
                    indicator_bank_id=new_indicator.id,
                    user_id=current_user.id,
                    change_type='CREATED',
                    change_description=f'Indicator "{new_indicator.name}" created via import by {current_user.name or current_user.email}',
                    **indicator_bank_history_snapshot(new_indicator),
                )
                db.session.add(history)
                result['imported'] += 1

        except Exception as e:
            current_app.logger.error("Error processing Indicators row %d: %s", row_num, e, exc_info=True)
            result['errors'].append(f"Row {row_num}: error.")
            continue

def _create_indicator_from_suggestion(suggestion):
    """Create an indicator from an approved suggestion"""
    try:
        new_indicator = IndicatorBank(
            name=suggestion.suggested_name,
            definition=suggestion.suggested_definition or '',
            type=suggestion.suggested_type or 'numeric',
            unit=suggestion.suggested_unit or '',
            fdrs_kpi_code=None,
            emergency=suggestion.suggested_emergency or False,
            related_programs=suggestion.suggested_programs or ''
        )

        db.session.add(new_indicator)
        db.session.flush()  # Get the ID
        backfill_fk_from_strings_bank(new_indicator)

        # Create history record
        history = IndicatorBankHistory(
            indicator_bank_id=new_indicator.id,
            user_id=current_user.id,
            change_type='CREATED',
            change_description=f'Indicator "{new_indicator.name}" created from suggestion by {current_user.name or current_user.email}',
            **indicator_bank_history_snapshot(new_indicator),
        )
        db.session.add(history)

        # Update suggestion status
        suggestion.status = IndicatorSuggestionStatusValue.implemented
        suggestion.indicator_id = new_indicator.id

        return new_indicator

    except Exception as e:
        current_app.logger.error(f"Error creating indicator from suggestion: {e}", exc_info=True)
        return None
