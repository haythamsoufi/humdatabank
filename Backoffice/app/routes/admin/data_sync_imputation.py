from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, send_file, after_this_request
from flask_login import current_user
import tempfile
from app.models import db, FormTemplate, FormItem, FormSection, FormPage, AssignedForm, FormData, Country, TemplateShare
from app.models.assignments import AssignmentEntityStatus
from app.routes.admin.shared import admin_permission_required, check_template_access
from app.services.forms.imputation_service import ImputationService
from app.services.forms.reporting_period_service import sort_period_names
import io
import os
import sys
import threading
import time
import uuid
import logging
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment
from contextlib import suppress
from datetime import datetime
from typing import Any, Dict, List, Optional

from app.utils.transactions import request_transaction_rollback
from app.utils.api_helpers import GENERIC_ERROR_MESSAGE, get_json_safe
from app.utils.request_utils import get_json_or_form, is_json_request
from app.utils.error_handling import handle_json_view_exception
from app.utils.api_responses import json_accepted, json_bad_request, json_error, json_forbidden, json_not_found, json_ok, json_server_error
from app.services.imports.async_import_job_store import (
    FDRS_DATA_SYNC_JOB_TYPE,
    create_import_job,
    cleanup_expired_import_jobs,
    clear_import_job_logging_state,
    get_import_job,
    get_import_job_logging_state,
    is_import_job_cancel_requested,
    request_import_job_cancel,
    update_import_job,
)
bp = Blueprint("data_sync_imputation", __name__, url_prefix="/admin/templates/data-sync")


def _country_by_aes_id_for_assignments(assignments):
    """Batch-resolve Country for country-type AES rows (avoids ``aes.country`` N+1)."""
    from app.utils.api_serialization import batch_countries_for_aes_list, _country_for_aes

    all_aes = []
    for af in assignments or []:
        all_aes.extend(af.country_statuses.all())
    if not all_aes:
        return {}
    batch = batch_countries_for_aes_list(all_aes)
    return {aes.id: _country_for_aes(aes, batch) for aes in all_aes}


# ----------------------------------
# Data sync cancel signals (per-worker thread events)
# ----------------------------------
_DATA_SYNC_LOCK = threading.Lock()
_DATA_SYNC_CANCEL_EVENTS: Dict[str, threading.Event] = {}
_DATA_SYNC_STALE_SECONDS = 15 * 60


def _reconcile_stale_data_sync_job(job_id: str) -> None:
    """Mark long-idle running jobs as failed/cancelled (worker recycle orphan recovery)."""
    job = get_import_job(job_id)
    if not job:
        return
    status = job.get("status")
    if status not in ("running", "cancel_requested"):
        return
    updated_ts = job.get("updated_ts")
    if updated_ts is None:
        return
    if time.time() - float(updated_ts) <= _DATA_SYNC_STALE_SECONDS:
        return
    if status == "cancel_requested":
        update_import_job(
            job_id,
            force=True,
            status="cancelled",
            stage="cancelled",
            message="Cancelled",
            error="Sync cancelled by user.",
        )
        return
    update_import_job(
        job_id,
        force=True,
        status="failed",
        stage="failed",
        message="Stopped",
        error="The sync worker stopped responding (likely after an app restart). Re-run the sync.",
    )


def _cleanup_data_sync_jobs_locked(now_ts: Optional[float] = None) -> None:
    """Remove old finished import jobs from PostgreSQL (safe across workers)."""
    cleanup_expired_import_jobs(now_ts)


_SYNC_ALLOWED_STATES = frozenset({0, 100, 200, 300, 400, 500})


def _parse_reported_import_states(data: Dict[str, Any]) -> Optional[List[int]]:
    """
    Parse JSON fdrs_reported_import_states (list of ints or comma-separated string).
    If the key is absent, return None (importer uses env / default allowlist).
    """
    if "fdrs_reported_import_states" not in data:
        return None
    raw = data.get("fdrs_reported_import_states")
    if raw is None:
        return None
    out: List[int] = []
    if isinstance(raw, str):
        parts = [p.strip() for p in raw.split(",") if p.strip()]
        for p in parts:
            try:
                out.append(int(p))
            except ValueError:
                raise ValueError("Each data status must be a whole number (IFRC State).")
    elif isinstance(raw, (list, tuple)):
        for p in raw:
            try:
                out.append(int(p))
            except (TypeError, ValueError):
                raise ValueError("Each data status must be a whole number (IFRC State).")
    else:
        raise ValueError("Data statuses must be sent as a list or comma-separated numbers.")
    if not out:
        raise ValueError("Select at least one data status to include.")
    bad = [x for x in out if x not in _SYNC_ALLOWED_STATES]
    if bad:
        raise ValueError(
            "Unknown data status value(s): %s. Use only the statuses shown in the sync dialog."
            % (", ".join(str(x) for x in sorted(set(bad))),)
        )
    return out


def _get_data_sync_cancel_event(job_id: str) -> threading.Event:
    with _DATA_SYNC_LOCK:
        ev = _DATA_SYNC_CANCEL_EVENTS.get(job_id)
        if ev is None:
            ev = threading.Event()
            _DATA_SYNC_CANCEL_EVENTS[job_id] = ev
        return ev


def _clear_data_sync_cancel_event(job_id: str) -> None:
    with _DATA_SYNC_LOCK:
        _DATA_SYNC_CANCEL_EVENTS.pop(job_id, None)


def _published_sections_query(template: FormTemplate):
    """Sections for the published template version only (avoids duplicate rows across versions)."""
    query = FormSection.query.filter(
        FormSection.template_id == template.id,
        FormSection.archived == False,
    )
    if template.published_version_id:
        query = query.filter(FormSection.version_id == template.published_version_id)
    return query.order_by(FormSection.order)


def _published_items_query(template: FormTemplate):
    """Form items for the published template version only.

    Aligns with stable_key deploy migration: form_data FKs and sync/imputation
    must target the live published structure, not draft or archived versions.
    """
    query = FormItem.query.filter(
        FormItem.template_id == template.id,
        FormItem.archived == False,
    )
    if template.published_version_id:
        query = query.filter(FormItem.version_id == template.published_version_id)
    return query.order_by(FormItem.order)


def _sort_sections_by_order(sections: List[FormSection]) -> List[FormSection]:
    return sorted(sections, key=lambda s: (float(s.order or 0), s.id))


def _build_ordered_sections_with_items(
    template: FormTemplate,
    all_sections: List[FormSection],
    items_by_section_id: Dict[int, List[FormItem]],
) -> List[Dict[str, Any]]:
    """Sections-with-items in form-builder order: pages, main sections, then subsections."""
    sections_with_items: List[Dict[str, Any]] = []
    handled_section_ids: set[int] = set()

    def _append_if_has_items(section: FormSection) -> None:
        section_items = items_by_section_id.get(section.id, [])
        if not section_items:
            return
        sections_with_items.append({
            'section': section,
            'section_items': list(section_items),
        })
        handled_section_ids.add(section.id)

    main_sections: List[FormSection] = []
    sub_sections_by_parent: Dict[int, List[FormSection]] = {}
    for section in all_sections:
        if section.parent_section_id is None:
            main_sections.append(section)
        else:
            sub_sections_by_parent.setdefault(section.parent_section_id, []).append(section)

    def _emit_main_section_block(section: FormSection) -> None:
        _append_if_has_items(section)
        for subsection in _sort_sections_by_order(sub_sections_by_parent.get(section.id, [])):
            _append_if_has_items(subsection)

    is_paginated = template.is_paginated
    if is_paginated and template.published_version_id:
        pages = (
            FormPage.query
            .filter_by(template_id=template.id, version_id=template.published_version_id)
            .order_by(FormPage.order)
            .all()
        )
        sections_by_page: Dict[Any, List[FormSection]] = {}
        for section in main_sections:
            page_key = section.page_id if section.page_id else 'default'
            sections_by_page.setdefault(page_key, []).append(section)

        for page in pages:
            for section in _sort_sections_by_order(sections_by_page.get(page.id, [])):
                _emit_main_section_block(section)

        for section in _sort_sections_by_order(sections_by_page.get('default', [])):
            _emit_main_section_block(section)
    else:
        for section in _sort_sections_by_order(main_sections):
            _emit_main_section_block(section)

    # Subsections whose parent is not in the published set (edge case)
    orphan_subsections = [
        section for section in all_sections
        if section.parent_section_id and section.id not in handled_section_ids
        and items_by_section_id.get(section.id)
    ]
    for section in _sort_sections_by_order(orphan_subsections):
        _append_if_has_items(section)

    return sections_with_items


def _imputable_items_for_template(
    template: FormTemplate,
    *,
    item_filter: Optional[str] = None,
    type_filter: Optional[str] = None,
) -> List[FormItem]:
    """Published-version items that participate in imputation, in template section order."""
    all_sections = _published_sections_query(template).all()
    items_by_section_id: Dict[int, List[FormItem]] = {}
    for item in _published_items_query(template).all():
        imputation_method = item.config.get('imputation_method', 'no_imputation') if item.config else 'no_imputation'
        if imputation_method == 'no_imputation':
            continue
        if item_filter and item.label != item_filter:
            continue
        if type_filter and item.type != type_filter:
            continue
        items_by_section_id.setdefault(item.section_id, []).append(item)

    items: List[FormItem] = []
    for section_block in _build_ordered_sections_with_items(template, all_sections, items_by_section_id):
        items.extend(section_block['section_items'])
    return items


# Templates that support external data sync (FDRS pipeline).
# Other templates get imputation only until their sync source is configured.
_TEMPLATES_WITH_DATA_SYNC: frozenset = frozenset({21})
_UPR_EXCEL_TEMPLATE_IDS: frozenset = frozenset({22, 23, 24, 33})


def _accessible_templates_for_user(user) -> List[Dict[str, Any]]:
    """Templates the user may open in the data sync & imputation tool."""
    from app.services.organization.authorization_service import AuthorizationService
    from app.services.security.api_authentication import get_user_allowed_template_ids

    if AuthorizationService.is_system_manager(user):
        templates = FormTemplate.query.all()
    else:
        allowed_ids = get_user_allowed_template_ids(user.id)
        if not allowed_ids:
            return []
        templates = FormTemplate.query.filter(FormTemplate.id.in_(allowed_ids)).all()

    templates.sort(key=lambda t: (t.name or "").lower())
    return [{"id": t.id, "name": t.name} for t in templates]


def _sections_with_items_for_template(template: FormTemplate) -> List[Dict[str, Any]]:
    template_sections = _published_sections_query(template).all()
    published_items = _published_items_query(template).all()
    items_by_section_id: Dict[int, List[FormItem]] = {}
    for item in published_items:
        items_by_section_id.setdefault(item.section_id, []).append(item)
    return _build_ordered_sections_with_items(
        template,
        template_sections,
        items_by_section_id,
    )


def _fdrs_imports_dir() -> str:
    return os.path.normpath(os.path.join(current_app.root_path, "..", "scripts", "imports"))


def _fdrs_default_years_bounds() -> tuple[int, int]:
    imports_dir = _fdrs_imports_dir()
    if imports_dir not in sys.path:
        sys.path.insert(0, imports_dir)
    from fdrs_data_fetcher import DEFAULT_FDRS_YEARS_END, DEFAULT_FDRS_YEARS_START

    return DEFAULT_FDRS_YEARS_START, DEFAULT_FDRS_YEARS_END


def _template_has_data_sync(template_id: int) -> bool:
    imports_dir = _fdrs_imports_dir()
    sync_script_available = os.path.isfile(os.path.join(imports_dir, "import_fdrs_form_data.py"))
    return (template_id in _TEMPLATES_WITH_DATA_SYNC) and sync_script_available


def render_data_sync_imputation_page(template_id: int):
    """Render data sync & imputation UI for the given template."""
    template = FormTemplate.query.get_or_404(template_id)

    if not check_template_access(template_id, current_user.id):
        flash("Access denied. You don't have permission to access this template.", "warning")
        return redirect(url_for("form_builder.manage_templates"))

    sections_with_items = _sections_with_items_for_template(template)
    has_data_sync = _template_has_data_sync(template_id)
    accessible_templates = _accessible_templates_for_user(current_user)
    fdrs_years_start, fdrs_years_end = _fdrs_default_years_bounds()

    return render_template(
        "admin/templates/data_sync_imputation.html",
        template=template,
        sections_with_items=sections_with_items,
        title=f"{template.name} — Data Sync & Imputation",
        has_data_sync=has_data_sync,
        has_upr_excel=template_id in _UPR_EXCEL_TEMPLATE_IDS,
        accessible_templates=accessible_templates,
        fdrs_years_start=fdrs_years_start,
        fdrs_years_end=fdrs_years_end,
    )


@bp.route("/<int:template_id>", methods=["GET"])
@admin_permission_required('admin.templates.view')
def data_sync_view(template_id: int):
    return render_data_sync_imputation_page(template_id)


@bp.route("/<int:template_id>/context", methods=["GET"])
@admin_permission_required('admin.templates.view')
def get_template_context(template_id: int):
    """JSON payload for switching templates without a full page reload."""
    template = FormTemplate.query.get_or_404(template_id)
    if not check_template_access(template_id, current_user.id):
        return json_forbidden("Access denied. You don't have permission to access this template.")

    sections_with_items = _sections_with_items_for_template(template)
    methods_html = render_template(
        "admin/templates/partials/imputation_methods_rows.html",
        sections_with_items=sections_with_items,
    )
    return json_ok({
        "template_id": template_id,
        "template_name": template.name,
        "page_title": f"{template.name} — Data Sync & Imputation",
        "has_data_sync": _template_has_data_sync(template_id),
        "has_upr_excel": template_id in _UPR_EXCEL_TEMPLATE_IDS,
        "methods_html": methods_html,
    })


@bp.route("/impute/template2", methods=["POST"])
@admin_permission_required('admin.templates.edit')
def impute_template2():
    # Accept from form or JSON; trim whitespace
    target_period = (request.form.get('target_period') or '').strip()
    if not target_period and is_json_request():
        try:
            payload = get_json_safe()
            target_period = str(payload.get('target_period') or '').strip()
        except Exception as e:
            current_app.logger.debug("target_period extraction failed: %s", e)
            target_period = ''
    if not target_period:
        flash("Target period is required (e.g., 2025)", "warning")
        return redirect(url_for('data_sync_imputation.data_sync_view', template_id=1))

    try:
        result = ImputationService.impute_template_2(target_period)
        if result.get("success"):
            flash(
                f"Imputation completed for {result['target_period']} from {result['source_period']}. "
                f"Countries: {result['countries_processed']}, Items: {result['items_imputed']}, "
                f"Rows created: {result['rows_created']}, updated: {result['rows_updated']}",
                "success"
            )
        else:
            flash(result.get("error") or "Imputation failed", "danger")
    except Exception as e:
        current_app.logger.error(f"Imputation error: {e}", exc_info=True)
        flash("An error occurred during imputation.", "danger")

    return redirect(url_for('data_sync_imputation.data_sync_view', template_id=1))


@bp.route("/update-imputation-methods-batch", methods=["POST"])
@admin_permission_required('admin.templates.edit')
def update_imputation_methods_batch():
    """Update imputation methods for multiple form items in a single request."""
    try:
        data = get_json_safe()
        updates = data.get('updates', [])

        if not updates:
            return json_bad_request('No updates provided')

        results = []
        for update in updates:
            item_id = update.get('item_id')
            method = update.get('method')

            if item_id is None or method is None:
                results.append({'item_id': item_id, 'success': False, 'error': 'Missing item_id or method'})
                continue

            # Coerce item_id to int
            try:
                item_id_int = int(item_id)
            except Exception as e:
                current_app.logger.debug("item_id int parse failed for %r: %s", item_id, e)
                results.append({'item_id': item_id, 'success': False, 'error': 'Invalid item_id'})
                continue

            if method not in ['last_year', 'three_year_avg', 'no_imputation']:
                results.append({'item_id': item_id_int, 'success': False, 'error': 'Invalid method'})
                continue

            # Get the form item
            form_item = FormItem.query.get(item_id_int)
            if not form_item:
                results.append({'item_id': item_id_int, 'success': False, 'error': 'Form item not found'})
                continue

            # Update the config - force a new dict to trigger SQLAlchemy change detection
            if form_item.config is None:
                new_config = {'imputation_method': method}
            else:
                new_config = form_item.config.copy()
                new_config['imputation_method'] = method

            # Force assignment to trigger change detection
            form_item.config = new_config
            db.session.add(form_item)
            results.append({'item_id': item_id_int, 'success': True})

        # Commit all changes at once
        db.session.flush()

        return json_ok(
            success=True,
            results=results,
            total_updated=len([r for r in results if r.get('success')])
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/preview-data", methods=["GET"])
@admin_permission_required('admin.templates.view')
def preview_data(template_id: int):
    """Get current data for preview table with progress tracking."""
    try:
        year = request.args.get('year')
        if not year:
            return json_bad_request('Year parameter required')

        template = FormTemplate.query.get_or_404(template_id)

        # Get all assignments for this template and year
        assignments = AssignedForm.query.filter_by(
            template_id=template_id,
            period_name=year
        ).all()

        form_items = _imputable_items_for_template(template)

        preview_data = []
        total_operations = 0
        completed_operations = 0
        country_by_aes = _country_by_aes_id_for_assignments(assignments)

        # Calculate total operations for progress tracking
        for assignment in assignments:
            total_operations += len(assignment.country_statuses.all()) * len(form_items)

        for assignment in assignments:
            for aes in assignment.country_statuses.all():
                country = country_by_aes.get(aes.id)
                if not country:
                    completed_operations += len(form_items)
                    continue

                for item in form_items:
                    # Get current data
                    form_data = FormData.query.filter_by(
                        assignment_entity_status_id=aes.id,
                        form_item_id=item.id
                    ).first()

                    current_value = None
                    if form_data:
                        try:
                            current_value = form_data.total_value
                        except AttributeError:
                            current_value = form_data.value

                    preview_data.append({
                        'country': country.name,
                        'item_label': item.label,
                        'item_unit': item.unit,
                        'current_value': current_value,
                        'imputed_value': None,  # Will be filled by preview imputation
                        'method': None,
                        'source_periods': None
                    })

                    completed_operations += 1

        return json_ok(
            success=True,
            data=preview_data,
            progress={
                'total': total_operations,
                'completed': completed_operations,
                'percentage': 100
            }
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/preview-imputation", methods=["POST"])
@admin_permission_required('admin.templates.edit')
def preview_imputation(template_id: int):
    """Preview imputation results without saving to database."""
    try:
        data = get_json_safe()
        year = data.get('year')
        if not year:
            return json_bad_request('Year parameter required')

        template = FormTemplate.query.get_or_404(template_id)

        # Calculate previous year
        try:
            prev_year = str(int(year) - 1)
        except (ValueError, TypeError):
            return json_bad_request('Invalid year format')

        # Get all assignments for target and previous year
        target_assignments = AssignedForm.query.filter_by(
            template_id=template_id,
            period_name=year
        ).all()

        prev_assignments = AssignedForm.query.filter_by(template_id=template_id, period_name=prev_year).all()
        country_by_aes = _country_by_aes_id_for_assignments(target_assignments + prev_assignments)

        prev_assignments_by_country = {}
        for af in prev_assignments:
            for aes in af.country_statuses.all():
                if aes.entity_id not in prev_assignments_by_country:
                    prev_assignments_by_country[aes.entity_id] = af

        form_items = _imputable_items_for_template(template)

        preview_data = []

        for assignment in target_assignments:
            for aes in assignment.country_statuses.all():
                country = country_by_aes.get(aes.id)
                if not country:
                    continue

                prev_af = prev_assignments_by_country.get(country.id)
                prev_aes = None
                if prev_af:
                    prev_aes = prev_af.country_statuses.filter_by(
                        entity_id=country.id, entity_type='country',
                    ).first()

                for item in form_items:
                    # Get current data
                    current_fd = FormData.query.filter_by(
                        assignment_entity_status_id=aes.id,
                        form_item_id=item.id
                    ).first()

                    current_value = None
                    if current_fd:
                        try:
                            current_value = current_fd.total_value
                        except AttributeError:
                            current_value = current_fd.value

                    # Calculate imputed value
                    imputed_value = None
                    method = None
                    source_periods = []

                    if prev_aes:
                        imputation_method = item.config.get('imputation_method', 'no_imputation') if item.config else 'no_imputation'

                        if imputation_method == 'three_year_avg':
                            # Get 3 years of data
                            values = []
                            for year_offset in range(1, 4):
                                source_year = str(int(year) - year_offset)
                                source_af = AssignedForm.query.filter_by(
                                    template_id=template_id,
                                    period_name=source_year
                                ).first()
                                if source_af:
                                    source_aes = source_af.country_statuses.filter_by(
                                        entity_id=country.id, entity_type='country',
                                    ).first()
                                    if source_aes:
                                        source_fd = FormData.query.filter_by(
                                            assignment_entity_status_id=source_aes.id,
                                            form_item_id=item.id
                                        ).first()
                                        if source_fd:
                                            with suppress(Exception):
                                                val = source_fd.total_value
                                                if val is not None:
                                                    values.append(float(val))
                                                    source_periods.append(source_year)

                            if values:
                                imputed_value = sum(values) / len(values)
                                method = 'Three Year Average'

                        else:  # last_year
                            prev_fd = FormData.query.filter_by(
                                assignment_entity_status_id=prev_aes.id,
                                form_item_id=item.id
                            ).first()
                            if prev_fd:
                                with suppress(Exception):
                                    # For single_choice items, prioritize value field over total_value
                                    # since text values like "Male" are stored in the value field
                                    val = prev_fd.value
                                    if val is not None and val != '':
                                        # Try to convert to float for numeric values, keep as string for text
                                        try:
                                            imputed_value = float(val)
                                        except (ValueError, TypeError):
                                            imputed_value = val  # Keep as string for text values
                                        method = 'Last Year\'s Value'
                                        source_periods = [prev_year]
                                    else:
                                        # Fallback to total_value for numeric items
                                        val = prev_fd.total_value
                                        if val is not None:
                                            imputed_value = float(val)
                                            method = 'Last Year\'s Value'
                                            source_periods = [prev_year]

                    preview_data.append({
                        'country': country.name,
                        'item_label': item.label,
                        'item_unit': item.unit,
                        'current_value': current_value,
                        'imputed_value': round(imputed_value, 2) if imputed_value is not None and isinstance(imputed_value, (int, float)) else imputed_value,
                        'method': method,
                        'source_periods': source_periods
                    })

        return json_ok(success=True, data=preview_data)

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/preview-data-chunked", methods=["POST"])
@admin_permission_required('admin.templates.view')
def preview_data_chunked(template_id: int):
    """Get current data for preview table with real progress tracking using chunked processing.

    Optimized to avoid building the full (ACS x items) combinations list on each request.
    """
    try:
        data = get_json_safe()
        year = data.get('year')
        chunk_size = data.get('chunk_size', 2000)
        offset = data.get('offset', 0)
        country_filter = data.get('country_filter')
        item_filter = data.get('item_filter')
        type_filter = data.get('type_filter')
        imputation_mode = data.get('imputation_mode', 'missing_only')  # Unused here but kept for parity

        if not year:
            return json_bad_request('Year parameter required')

        template = FormTemplate.query.get_or_404(template_id)

        # Gather filtered ACS (with country) for the target year
        assignments = AssignedForm.query.filter_by(template_id=template_id, period_name=year).all()
        country_by_aes = _country_by_aes_id_for_assignments(assignments)
        filtered_aess = []  # List[Tuple[AssignmentEntityStatus, Country]]
        for assignment in assignments:
            for aes in assignment.country_statuses.all():
                country = country_by_aes.get(aes.id)
                if not country:
                    continue
                if country_filter and country.name != country_filter:
                    continue
                filtered_aess.append((aes, country))

        filtered_items = _imputable_items_for_template(
            template,
            item_filter=item_filter,
            type_filter=type_filter,
        )

        num_aess = len(filtered_aess)
        num_items = len(filtered_items)
        total = num_aess * num_items

        if total == 0:
            return json_ok(
                success=True,
                data=[],
                progress={
                    'total': 0,
                    'completed': 0,
                    'percentage': 100,
                    'is_complete': True,
                    'next_offset': None
                }
            )

        end_offset = min(offset + chunk_size, total)

        # Build needed sets for preloading FormData for this chunk only
        aes_ids_needed = set()
        item_ids_needed = set()
        for i in range(offset, end_offset):
            aes_idx = i // num_items
            item_idx = i % num_items
            aes, _country = filtered_aess[aes_idx]
            item = filtered_items[item_idx]
            aes_ids_needed.add(aes.id)
            item_ids_needed.add(item.id)

        formdata_map = {}
        if aes_ids_needed and item_ids_needed:
            fds = FormData.query.filter(
                FormData.assignment_entity_status_id.in_(aes_ids_needed),
                FormData.form_item_id.in_(item_ids_needed)
            ).all()
            for fd in fds:
                formdata_map[(fd.assignment_entity_status_id, fd.form_item_id)] = fd

        # Build chunk data using index math
        chunk_data = []
        for i in range(offset, end_offset):
            aes_idx = i // num_items
            item_idx = i % num_items
            aes, country = filtered_aess[aes_idx]
            item = filtered_items[item_idx]

            fd = formdata_map.get((aes.id, item.id))
            current_value = None
            if fd:
                try:
                    current_value = fd.total_value
                except Exception as e:
                    current_app.logger.debug("fd.total_value failed: %s", e)
                    current_value = fd.value

            chunk_data.append({
                'country': country.name,
                'item_label': item.label,
                'item_unit': item.unit,
                'current_value': current_value,
                'imputed_value': None,
                'method': None,
                'source_periods': None
            })

        is_complete = end_offset >= total
        progress_percentage = (end_offset / total) * 100 if total > 0 else 100

        return json_ok(
            success=True,
            data=chunk_data,
            progress={
                'total': total,
                'completed': end_offset,
                'percentage': round(progress_percentage, 1),
                'is_complete': is_complete,
                'next_offset': end_offset if not is_complete else None
            }
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/preview-imputation-chunked", methods=["POST"])
@admin_permission_required('admin.templates.edit')
def preview_imputation_chunked(template_id: int):
    """Preview imputation results with real progress tracking using chunked processing.

    Optimized to avoid building the full (ACS x items) combinations list on each request,
    and preload only the FormData entries needed for the current chunk.
    """
    try:
        data = get_json_safe()
        year = data.get('year')
        chunk_size = data.get('chunk_size', 2000)
        offset = data.get('offset', 0)
        country_filter = data.get('country_filter')
        item_filter = data.get('item_filter')
        type_filter = data.get('type_filter')
        imputation_mode = data.get('imputation_mode', 'missing_only')

        if not year:
            return json_bad_request('Year parameter required')

        source_period = data.get('source_period')

        template = FormTemplate.query.get_or_404(template_id)
        if source_period:
            prev_year = source_period
        else:
            try:
                prev_year = str(int(year) - 1)
            except Exception as e:
                current_app.logger.debug("prev_year parse failed, source_period required: %s", e)
                return json_bad_request(
                    'The target period is not a plain year number. '
                    'Please select a Source Period to carry values forward from.'
                )

        # Target AES list with countries (filtered)
        target_assignments = AssignedForm.query.filter_by(template_id=template_id, period_name=year).all()
        country_by_aes = _country_by_aes_id_for_assignments(target_assignments)
        filtered_aess = []  # List[Tuple[AssignmentEntityStatus, Country]]
        for assignment in target_assignments:
            for aes in assignment.country_statuses.all():
                country = country_by_aes.get(aes.id)
                if not country:
                    continue
                if country_filter and country.name != country_filter:
                    continue
                filtered_aess.append((aes, country))

        filtered_items = _imputable_items_for_template(
            template,
            item_filter=item_filter,
            type_filter=type_filter,
        )

        num_aess = len(filtered_aess)
        num_items = len(filtered_items)
        total = num_aess * num_items

        if total == 0:
            return json_ok(
                success=True,
                data=[],
                progress={
                    'total': 0,
                    'completed': 0,
                    'percentage': 100,
                    'is_complete': True,
                    'next_offset': None
                }
            )

        end_offset = min(offset + chunk_size, total)

        # Determine source years for three-year average
        source_years = []
        try:
            base_year = int(year)
            source_years = [str(base_year - k) for k in (1, 2, 3)]
        except Exception as e:
            current_app.logger.debug(
                "source_years parse failed (non-numeric period), falling back to single source_period: %s", e
            )
            if prev_year:
                source_years = [prev_year]

        # Preload mappings of (year -> country_id -> aes_id) for needed years, limited to countries present
        country_ids_in_scope = {country.id for (_aes, country) in filtered_aess}
        country_year_to_aes_id = {}
        if source_years:
            prev_afs = AssignedForm.query.filter(
                AssignedForm.template_id == template_id,
                AssignedForm.period_name.in_(source_years)
            ).all()
            for paf in prev_afs:
                for p_aes in paf.country_statuses.all():
                    if p_aes.entity_id in country_ids_in_scope:
                        country_year_to_aes_id[(paf.period_name, p_aes.entity_id)] = p_aes.id

        # Collect needed FormData pairs for this chunk only
        current_pairs = set()
        source_pairs = set()
        for i in range(offset, end_offset):
            aes_idx = i // num_items
            item_idx = i % num_items
            aes, country = filtered_aess[aes_idx]
            item = filtered_items[item_idx]

            current_pairs.add((aes.id, item.id))

            imputation_method = item.config.get('imputation_method', 'no_imputation') if item.config else 'no_imputation'
            if imputation_method == 'three_year_avg':
                for sy in source_years:
                    src_aes_id = country_year_to_aes_id.get((sy, country.id))
                    if src_aes_id:
                        source_pairs.add((src_aes_id, item.id))
            elif imputation_method == 'last_year':
                src_aes_id = country_year_to_aes_id.get((prev_year, country.id))
                if src_aes_id:
                    source_pairs.add((src_aes_id, item.id))

        # Preload FormData for all needed pairs
        needed_pairs = current_pairs | source_pairs
        all_aes_ids_needed = {aes_id for (aes_id, _iid) in needed_pairs}
        all_item_ids_needed = {iid for (_aes_id, iid) in needed_pairs}

        formdata_map = {}
        if all_aes_ids_needed and all_item_ids_needed:
            fds = FormData.query.filter(
                FormData.assignment_entity_status_id.in_(all_aes_ids_needed),
                FormData.form_item_id.in_(all_item_ids_needed)
            ).all()
            for fd in fds:
                formdata_map[(fd.assignment_entity_status_id, fd.form_item_id)] = fd

        # Build chunk data
        chunk_data = []
        for i in range(offset, end_offset):
            aes_idx = i // num_items
            item_idx = i % num_items
            aes, country = filtered_aess[aes_idx]
            item = filtered_items[item_idx]

            fd_current = formdata_map.get((aes.id, item.id))
            current_value = None
            if fd_current:
                try:
                    current_value = fd_current.total_value
                except Exception as e:
                    current_app.logger.debug("fd_current.total_value failed: %s", e)
                    current_value = fd_current.value

            imputed_value = None
            method = None
            source_periods = []

            imputation_method = item.config.get('imputation_method', 'no_imputation') if item.config else 'no_imputation'
            if imputation_method != 'no_imputation':
                if imputation_method == 'three_year_avg' and source_years:
                    values = []
                    used_years = []
                    for sy in source_years:
                        src_aes_id = country_year_to_aes_id.get((sy, country.id))
                        if not src_aes_id:
                            continue
                        fd_src = formdata_map.get((src_aes_id, item.id))
                        if not fd_src:
                            continue
                        with suppress(Exception):
                            val = fd_src.total_value
                            if val is None:
                                val = fd_src.value
                            if val is not None:
                                values.append(float(val))
                                used_years.append(sy)
                    if values:
                        imputed_value = sum(values) / len(values)
                        method = 'Three Year Average'
                        source_periods = used_years
                elif imputation_method == 'last_year':
                    src_aes_id = country_year_to_aes_id.get((prev_year, country.id))
                    if src_aes_id:
                        fd_prev = formdata_map.get((src_aes_id, item.id))
                        if fd_prev:
                            with suppress(Exception):
                                # For single_choice items, prioritize value field over total_value
                                # since text values like "Male" are stored in the value field
                                val = fd_prev.value
                                if val is not None and val != '':
                                    # Try to convert to float for numeric values, keep as string for text
                                    try:
                                        imputed_value = float(val)
                                    except (ValueError, TypeError):
                                        imputed_value = val  # Keep as string for text values
                                    method = 'Last Year\'s Value'
                                    source_periods = [prev_year]
                                else:
                                    # Fallback to total_value for numeric items
                                    val = fd_prev.total_value
                                    if val is not None:
                                        imputed_value = float(val)
                                        method = 'Last Year\'s Value'
                                        source_periods = [prev_year]

            # Respect imputation mode
            if imputation_mode == 'missing_only' and imputed_value is not None:
                has_existing_data = (
                    (current_value is not None and current_value != '') or
                    (fd_current and fd_current.imputed_value is not None and fd_current.imputed_value != '')
                )
                if has_existing_data:
                    imputed_value = None
                    method = None
                    source_periods = []

            chunk_data.append({
                'country': country.name,
                'item_label': item.label,
                'item_unit': item.unit,
                'current_value': current_value,
                'imputed_value': round(imputed_value, 2) if imputed_value is not None and isinstance(imputed_value, (int, float)) else imputed_value,
                'method': method,
                'source_periods': source_periods
            })

        is_complete = end_offset >= total
        progress_percentage = (end_offset / total) * 100 if total > 0 else 100

        # Log summary for this chunk
        imputed_count = len([row for row in chunk_data if row['imputed_value'] is not None])
        current_app.logger.info(f"Chunk {offset}-{end_offset}: {imputed_count}/{len(chunk_data)} items have imputed values")

        return json_ok(
            success=True,
            data=chunk_data,
            progress={
                'total': total,
                'completed': end_offset,
                'percentage': round(progress_percentage, 1),
                'is_complete': is_complete,
                'next_offset': end_offset if not is_complete else None
            }
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/filter-options", methods=["GET"])
@admin_permission_required('admin.templates.view')
def get_filter_options(template_id: int):
    """Get available countries and items for filters."""
    try:
        year = request.args.get('year')
        if not year:
            return json_bad_request('Year parameter required')

        template = FormTemplate.query.get_or_404(template_id)

        # Get all assignments for this template and year
        assignments = AssignedForm.query.filter_by(
            template_id=template_id,
            period_name=year
        ).all()

        # Get unique countries from assignments
        country_by_aes = _country_by_aes_id_for_assignments(assignments)
        countries = []
        seen_names = set()
        for assignment in assignments:
            for aes in assignment.country_statuses.all():
                country = country_by_aes.get(aes.id)
                if country and country.name not in seen_names:
                    seen_names.add(country.name)
                    countries.append({
                        'id': country.id,
                        'name': country.name
                    })

        # Sort countries by name
        countries.sort(key=lambda x: x['name'])

        items = [
            {
                'id': item.id,
                'label': item.label,
                'unit': item.unit,
            }
            for item in _imputable_items_for_template(template)
        ]

        return json_ok(success=True, countries=countries, items=items)

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/available-periods", methods=["GET"])
@admin_permission_required('admin.templates.view')
def get_available_periods(template_id: int):
    """Return all distinct period_name values for this template, sorted most-recent-first."""
    try:
        FormTemplate.query.get_or_404(template_id)

        rows = (
            AssignedForm.query
            .filter_by(template_id=template_id)
            .with_entities(AssignedForm.period_name)
            .distinct()
            .all()
        )
        period_names = sort_period_names([row.period_name for row in rows if row.period_name])

        return json_ok(
            success=True,
            periods=period_names,
            latest=period_names[0] if period_names else None,
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/run-imputation-filtered", methods=["POST"])
@admin_permission_required('admin.templates.edit')
def run_imputation_filtered(template_id: int):
    """Run imputation for a specific year with optional country and item filters."""
    try:
        data = get_json_safe()
        year = data.get('year')
        source_period = data.get('source_period')
        country_filter = data.get('country_filter')
        item_filter = data.get('item_filter')
        type_filter = data.get('type_filter')
        imputation_mode = data.get('imputation_mode', 'missing_only')

        if not year:
            return json_bad_request('Year parameter required')

        result = ImputationService.impute_template_filtered(
            template_id=template_id,
            target_year=year,
            source_period=source_period,
            country_filter=country_filter,
            item_filter=item_filter,
            type_filter=type_filter,
            imputation_mode=imputation_mode
        )

        if result.get("success"):
            return json_ok(
                success=True,
                target_period=result['target_period'],
                source_period=result['source_period'],
                countries_processed=result['countries_processed'],
                items_imputed=result['items_imputed'],
                rows_created=result['rows_created'],
                rows_updated=result['rows_updated']
            )
        else:
            return json_bad_request(result.get('error', 'Imputation failed'))

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/export-excel", methods=["POST"])
@admin_permission_required('admin.templates.export_excel')
def export_preview_excel(template_id: int):
    """Export preview data to Excel."""
    try:
        data = get_json_safe()
        year = data.get('year')
        preview_data = data.get('preview_data', [])

        if not year or not preview_data:
            return json_bad_request('Year and preview data required')

        # Create workbook with formatting
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Preview Data"

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        imputed_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        # Headers
        headers = ['Country', 'Item', 'Unit', 'Current Value', 'Imputed Value', 'Method', 'Source Periods']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, row_data in enumerate(preview_data, 2):
            sheet.cell(row=row_idx, column=1, value=row_data.get('country', ''))
            sheet.cell(row=row_idx, column=2, value=row_data.get('item_label', ''))
            sheet.cell(row=row_idx, column=3, value=row_data.get('item_unit', ''))

            # Current value
            current_val = row_data.get('current_value')
            sheet.cell(row=row_idx, column=4, value=current_val if current_val is not None else 'No data')

            # Imputed value with highlighting
            imputed_val = row_data.get('imputed_value')
            imputed_cell = sheet.cell(row=row_idx, column=5, value=imputed_val if imputed_val is not None else '-')
            if imputed_val is not None:
                imputed_cell.fill = imputed_fill

            sheet.cell(row=row_idx, column=6, value=row_data.get('method', '-'))
            sheet.cell(row=row_idx, column=7, value=', '.join(row_data.get('source_periods', [])) or '-')

        # Auto-size columns
        for column in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            column_letter = column[0].column_letter
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 30)

        # Freeze header row
        sheet.freeze_panes = "A2"

        # Save and return
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name=f"preview_data_{year}.xlsx",
            as_attachment=True
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/run-data-sync", methods=["POST"])
@admin_permission_required("admin.templates.edit")
def run_data_sync(template_id: int):
    """
    Trigger data sync for a template (import_fdrs_form_data.py pipeline).
    Expects JSON: dry_run (bool), batch_size (int), fdrs_years (str, comma-separated), test (bool),
    imputed_use_cache (bool), sync_documents (bool), async (bool), and optional fdrs_reported_import_states (list of IFRC State ints).
    If fdrs_reported_import_states is omitted, the importer uses FDRS_REPORTED_IMPORT_STATES env or default all except Not filled (0).
    """
    try:
        template = FormTemplate.query.get_or_404(template_id)
        if not check_template_access(template_id, current_user.id):
            return json_forbidden("Access denied")

        data = get_json_safe()
        dry_run = bool(data.get("dry_run", False))
        batch_size_raw = data.get("batch_size", None)
        if batch_size_raw in (None, ""):
            batch_size = 1000
        else:
            try:
                batch_size = int(batch_size_raw)
            except Exception as e:
                current_app.logger.debug("batch_size parse failed: %s", e)
                return json_bad_request("Invalid batch_size: must be an integer (or omit it)")
        if batch_size < 100:
            return json_bad_request("Invalid batch_size: must be >= 100")
        fdrs_years_raw = (data.get("fdrs_years") or "").strip()
        test_mode = bool(data.get("test", False))
        async_mode = bool(data.get("async", False))
        imputed_use_cache = bool(data.get("imputed_use_cache", True))
        sync_documents = bool(data.get("sync_documents", True))
        try:
            fdrs_reported_import_states = _parse_reported_import_states(data)
        except ValueError as e:
            return json_bad_request(str(e))

        fdrs_years = None
        test_limit = None
        if test_mode:
            fdrs_years = [2024]
            test_limit = 1000
        elif fdrs_years_raw:
            try:
                fdrs_years = [int(y.strip()) for y in fdrs_years_raw.split(",") if y.strip()]
            except ValueError:
                return json_bad_request("Invalid fdrs_years: use comma-separated integers")

        imports_dir = _fdrs_imports_dir()
        if imports_dir not in sys.path:
            sys.path.insert(0, imports_dir)
        from import_fdrs_form_data import run_import

        preview_path = None
        if dry_run:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp.close()
            preview_path = tmp.name

        # Async mode: run in background and expose progress via polling endpoint.
        if async_mode:
            job_id = uuid.uuid4().hex
            sync_user_id = int(getattr(current_user, "id", 0) or 0) or None
            with _DATA_SYNC_LOCK:
                _cleanup_data_sync_jobs_locked(time.time())
            create_import_job(
                job_id=job_id,
                job_type=FDRS_DATA_SYNC_JOB_TYPE,
                user_id=int(getattr(current_user, "id", 0) or 0),
                initial={
                    "template_id": template_id,
                    "status": "queued",
                    "stage": "queued",
                    "message": "Queued",
                    "current": 0,
                    "total": None,
                    "percent": 0.0,
                    "stats": None,
                    "error": None,
                    "preview_path": preview_path,
                    "download_ready": False,
                    "last_logged_pct": None,
                },
            )

            worker_app = current_app._get_current_object()

            def _run_job(app=worker_app) -> None:
                log = logging.getLogger(__name__)
                cancel_ev = _get_data_sync_cancel_event(job_id)
                last_cancel_db_check = 0.0

                def _progress_cb(payload: Dict[str, Any]) -> None:
                    stage = payload.get("stage") or ""
                    pct = payload.get("percent")
                    msg = payload.get("message") or ""
                    existing = get_import_job(job_id) or {}
                    if stage.startswith(("documents", "assignment_status")):
                        current = payload.get("current")
                        total = payload.get("total")
                    else:
                        current = payload.get("current") if payload.get("current") is not None else existing.get("current")
                        total = payload.get("total") if payload.get("total") is not None else existing.get("total")
                    update_import_job(
                        job_id,
                        status="running",
                        stage=payload.get("stage") or existing.get("stage"),
                        message=payload.get("message") or existing.get("message"),
                        current=current,
                        total=total,
                        percent=float(payload.get("percent") or existing.get("percent") or 0.0),
                        stats=payload.get("stats") if payload.get("stats") is not None else existing.get("stats"),
                    )

                    try:
                        pct_f = float(pct) if pct is not None else None
                    except Exception:
                        pct_f = None
                    log_state = get_import_job_logging_state(job_id)
                    last_logged = log_state.get("last_logged_pct")
                    should_log = (
                        stage
                        in (
                            "documents_plan",
                            "documents_done",
                            "assignment_status_plan",
                            "assignment_status_done",
                            "complete",
                            "failed",
                            "cancelled",
                        )
                        or (stage and stage != "upsert" and not stage.endswith("_upsert"))
                        or (
                            pct_f is not None
                            and (last_logged is None or abs(pct_f - float(last_logged)) >= 5.0)
                        )
                    )
                    if should_log and pct_f is not None:
                        log_state["last_logged_pct"] = pct_f
                        update_import_job(job_id, last_logged_pct=pct_f)
                    if should_log:
                        app.logger.info(
                            "Data sync %s: %s %s%% %s",
                            job_id,
                            stage or "-",
                            f"{pct_f:.1f}" if pct_f is not None else "-",
                            msg,
                        )

                def _cancel_check() -> bool:
                    nonlocal last_cancel_db_check
                    if cancel_ev.is_set():
                        return True
                    now = time.time()
                    if now - last_cancel_db_check >= 1.0:
                        last_cancel_db_check = now
                        if is_import_job_cancel_requested(job_id):
                            cancel_ev.set()
                            return True
                    return False

                with app.app_context():
                    update_import_job(
                        job_id,
                        force=True,
                        status="running",
                        stage="starting",
                        message="Starting...",
                        worker_pid=os.getpid(),
                    )
                    app.logger.info(
                        "Data sync %s: starting (template_id=%s, dry_run=%s, test=%s, sync_documents=%s)",
                        job_id,
                        template_id,
                        dry_run,
                        test_mode,
                        sync_documents,
                    )

                    try:
                        from import_fdrs_form_data import FdrsSyncCancelled

                        stats = run_import(
                            input_path=None,
                            fdrs_api_url=None,
                            fdrs_from_data_api=True,
                            fdrs_data_api_base=None,
                            fdrs_data_api_key=None,
                            fdrs_imputed_url=None,
                            fdrs_imputed_from_api=False,
                            fdrs_imputed_kpi_codes_path=None,
                            fdrs_imputed_use_cache=imputed_use_cache,
                            fdrs_years=fdrs_years,
                            fdrs_reported_import_states=fdrs_reported_import_states,
                            indicator_mapping_path=None,
                            indicator_bank_api_base=None,
                            indicator_bank_api_key=None,
                            databank_base_url=None,
                            databank_api_key=None,
                            preview_excel_path=preview_path if dry_run else None,
                            test_limit=test_limit,
                            dry_run=dry_run,
                            batch_size=batch_size,
                            template_id=template_id,
                            progress_cb=_progress_cb,
                            cancel_check=_cancel_check,
                            sync_user_id=sync_user_id,
                            sync_documents=sync_documents,
                        )
                        update_import_job(
                            job_id,
                            force=True,
                            status="completed",
                            stage="complete",
                            message="Completed",
                            percent=100.0,
                            stats=dict(stats or {}),
                            download_ready=bool(dry_run and preview_path and os.path.isfile(preview_path)),
                        )
                        app.logger.info(
                            "Data sync %s: completed loaded=%s skipped=%s inserted=%s updated=%s errors=%s",
                            job_id,
                            stats.get("loaded"),
                            stats.get("skipped"),
                            stats.get("inserted"),
                            stats.get("updated"),
                            stats.get("errors"),
                        )
                    except FdrsSyncCancelled:
                        update_import_job(
                            job_id,
                            force=True,
                            status="cancelled",
                            stage="cancelled",
                            message="Cancelled",
                            error="Sync cancelled by user.",
                        )
                        app.logger.info("Data sync %s: cancelled", job_id)
                    except Exception as e:
                        log.exception("Async data sync job failed: %s", e)
                        err_msg = str(e).strip() or type(e).__name__
                        if len(err_msg) > 2000:
                            err_msg = err_msg[:1997] + "..."
                        update_import_job(
                            job_id,
                            force=True,
                            status="failed",
                            stage="failed",
                            message="Failed",
                            error=err_msg,
                        )
                        app.logger.error("Data sync %s: failed: %s", job_id, e, exc_info=True)
                    finally:
                        _clear_data_sync_cancel_event(job_id)
                        clear_import_job_logging_state(job_id)
                        db.session.remove()

            if current_app.config.get("TESTING"):
                _run_job(worker_app)
            else:
                threading.Thread(target=_run_job, args=(worker_app,), daemon=True).start()
            return json_accepted(job_id=job_id)

        stats = run_import(
            input_path=None,
            fdrs_api_url=None,
            fdrs_from_data_api=True,
            fdrs_data_api_base=None,
            fdrs_data_api_key=None,
            fdrs_imputed_url=None,
            fdrs_imputed_from_api=False,
            fdrs_imputed_kpi_codes_path=None,
            fdrs_imputed_use_cache=imputed_use_cache,
            fdrs_years=fdrs_years,
            fdrs_reported_import_states=fdrs_reported_import_states,
            indicator_mapping_path=None,
            indicator_bank_api_base=None,
            indicator_bank_api_key=None,
            databank_base_url=None,
            databank_api_key=None,
            preview_excel_path=preview_path if dry_run else None,
            test_limit=test_limit,
            dry_run=dry_run,
            batch_size=batch_size,
            template_id=template_id,
            sync_user_id=int(getattr(current_user, "id", 0) or 0) or None,
            sync_documents=sync_documents,
        )

        if dry_run and preview_path and os.path.isfile(preview_path):
            @after_this_request
            def _remove_preview(resp):
                try:
                    os.unlink(preview_path)
                except OSError:
                    pass
                return resp
            return send_file(
                preview_path,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"data_sync_preview_{template_id}.xlsx",
            )

        return json_ok(
            success=True,
            dry_run=dry_run,
            stats=stats,
            message=(
                f"Loaded: {stats['loaded']}, Skipped: {stats['skipped']}, "
                f"Inserted: {stats['inserted']}, Updated: {stats['updated']}, Errors: {stats['errors']}"
                + (
                    f"; Documents: +{stats.get('documents_inserted', 0)} "
                    f"~{stats.get('documents_updated', 0)} "
                    f"err={stats.get('documents_errors', 0)}"
                    if stats.get("documents_inserted") is not None or stats.get("documents_updated")
                    else ""
                )
            ),
        )
    except (ValueError, RuntimeError) as e:
        current_app.logger.error(f"Data sync error: {e}", exc_info=True)
        msg = str(e).strip() or "Sync failed."
        return json_bad_request(msg[:2000] if len(msg) > 2000 else msg)
    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/data-sync-status/<job_id>", methods=["GET"])
@admin_permission_required("admin.templates.edit")
def data_sync_status(template_id: int, job_id: str):
    """Poll data sync job status (for live UI progress)."""
    with _DATA_SYNC_LOCK:
        _cleanup_data_sync_jobs_locked(time.time())
    job = get_import_job(job_id)
    if not job or int(job.get("template_id") or 0) != int(template_id):
        return json_not_found("Job not found")
    if int(job.get("user_id") or 0) != int(getattr(current_user, "id", 0) or 0):
        return json_forbidden("Access denied")

    _reconcile_stale_data_sync_job(job_id)
    job = get_import_job(job_id) or job

    resp = {
        "success": True,
        "job": {
            "job_id": job_id,
            "status": job.get("status"),
            "stage": job.get("stage"),
            "message": job.get("message"),
            "current": job.get("current"),
            "total": job.get("total"),
            "percent": job.get("percent"),
            "stats": job.get("stats"),
            "error": job.get("error"),
            "started_at": job.get("started_at"),
            "updated_at": job.get("updated_at"),
            "download_ready": bool(job.get("download_ready")),
        },
    }
    if resp["job"]["download_ready"]:
        resp["job"]["download_url"] = url_for("data_sync_imputation.data_sync_download", template_id=template_id, job_id=job_id)
    return json_ok(**resp) if isinstance(resp, dict) else json_ok(data=resp)


@bp.route("/<int:template_id>/data-sync-cancel/<job_id>", methods=["POST"])
@admin_permission_required("admin.templates.edit")
def data_sync_cancel(template_id: int, job_id: str):
    """Request cancellation for a running data sync job (best-effort)."""
    job = get_import_job(job_id)
    if not job or int(job.get("template_id") or 0) != int(template_id):
        return json_not_found("Job not found")
    if int(job.get("user_id") or 0) != int(getattr(current_user, "id", 0) or 0):
        return json_forbidden("Access denied")

    status = job.get("status")
    if status in ("completed", "failed", "cancelled"):
        return json_ok(status=status)

    request_import_job_cancel(job_id)
    _get_data_sync_cancel_event(job_id).set()
    return json_ok(status="cancel_requested")


@bp.route("/<int:template_id>/data-sync-download/<job_id>", methods=["GET"])
@admin_permission_required("admin.templates.edit")
def data_sync_download(template_id: int, job_id: str):
    """Download preview Excel generated by an async dry-run sync."""
    job = get_import_job(job_id)
    if not job or int(job.get("template_id") or 0) != int(template_id):
        return json_not_found("Job not found")
    if int(job.get("user_id") or 0) != int(getattr(current_user, "id", 0) or 0):
        return json_forbidden("Access denied")
    path = job.get("preview_path")
    if not path or not os.path.isfile(path):
        return json_not_found("Preview file not available")

    @after_this_request
    def _remove_preview(resp):
        with suppress(Exception):
            os.unlink(path)
        update_import_job(job_id, force=True, download_ready=False, preview_path=None)
        return resp

    return send_file(
        path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"data_sync_preview_{template_id}.xlsx",
    )
