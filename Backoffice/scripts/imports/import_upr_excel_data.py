#!/usr/bin/env python3
"""
Import Unified Planning & Reporting (UPR) Excel data into form submissions.

Reads the ``UPR Data`` sheet from UPR Master.xlsx and maps rows into form_data.

Planning templates (rounds P*):
    24  Unified Country Plan — funding uses hybrid matrices (items 967/968/974)
    22  Annual Planning – International Bilateral Support

Reporting templates:
    33  Reporting – Country  (rounds AR*, MYR*)
    23  Reporting – PNS      (rounds AR*)

Usage:
    python scripts/imports/import_upr_excel_data.py --input path/to/UPR\\ Master.xlsx
    python scripts/imports/import_upr_excel_data.py --input path/to/file.xlsx --rounds P25,P26 --dry-run
    python scripts/imports/import_upr_excel_data.py --input path/to/file.xlsx --rounds AR25 --templates 33,23
    python scripts/imports/import_upr_excel_data.py --input path/to/file.xlsx --rounds MYR26 --templates 33
    python scripts/imports/import_upr_excel_data.py --input path/to/file.xlsx --templates 24,22
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import pickle
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any, Callable, Dict, List, Optional, Set, Tuple

logger = logging.getLogger(__name__)

script_dir = os.path.dirname(os.path.abspath(__file__))
backoffice_dir = os.path.dirname(os.path.dirname(script_dir))
if backoffice_dir not in sys.path:
    sys.path.insert(0, backoffice_dir)
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

if "FLASK_CONFIG" not in os.environ:
    os.environ["FLASK_CONFIG"] = "development"

from import_fdrs_form_data import (  # noqa: E402
    COL_ASSIGNMENT,
    COL_DISAGG,
    COL_ITEM,
    COL_NA,
    COL_DATA_NA,
    COL_IMPUTED,
    COL_PREFILLED,
    COL_PUBLIC,
    COL_SUBMITTED,
    COL_VALUE,
    FdrsSyncCancelled,
    upsert_form_data_rows,
    write_rows_to_excel,
)
from upr_import_versioning import (  # noqa: E402
    find_item_by_label,
)
from upr_import_warnings import summarize_warnings  # noqa: E402

UPR_DATA_SHEET = "UPR Data"
HEADER_ROW_INDEX = 2  # 0-based row 3 in Excel
ROWS_CACHE_VERSION = 1
TRANSFORM_CACHE_VERSION = 4

UPR_TEMPLATE_PROFILES: Dict[int, Dict[str, Any]] = {
    # ── Planning ──────────────────────────────────────────────────────────────
    24: {
        "name": "Unified Country Plan",
        "round_prefixes": ("P",),
        "sections": frozenset({"Reach", "NS Data", "Funding", "Comments", "Support"}),
    },
    22: {
        "name": "Annual Planning - International Bilateral Support",
        "round_prefixes": ("P",),
        # Staff rows are filtered here; PNS Funding is written from the Funding section
        # when template 22 is included alongside template 24.
        "sections": frozenset({"Staff"}),
    },
    # ── Reporting ─────────────────────────────────────────────────────────────
    33: {
        "name": "Reporting - Country",
        "round_prefixes": ("AR", "MYR"),
        "sections": frozenset({
            "NS Data",
            "Funding",
            "Core indicators",
            "Other indicators",
            "Support",
            "Emergency 1",
            "Emergency 2",
            "Emergency 3",
        }),
    },
    23: {
        "name": "Reporting - PNS",
        "round_prefixes": ("AR",),
        # Only AR rounds have T23 assignments. Funding covers PNS-reported totals.
        "sections": frozenset({"Funding"}),
    },
}

STAFF_INDICATOR_COLUMNS: Dict[str, str] = {
    "# of international delegates integrated with the HNS": "intl_delegates_hns",
    "# of international delegates integrated with the IFRC": "intl_delegates_ifrc",
    "# of national staff hired through the HNS (PNS operating under HNS legal umbrella)": "national_staff_hns_hns",
    "# of national staff hired through the HNS (PNS operating under IFRC legal umbrella)": "national_staff_hns_ifrc",
    "# of national staff hired through the IFRC (PNS operating under IFRC legal umbrella)": "national_staff_ifrc_ifrc",
}

STAFF_MATRIX_LABEL = "PNS staff contributions"

# Template 24 planning funding — one hybrid matrix per year offset (static HNS/IFRC rows + PNS list rows).
FUNDING_MATRIX_BY_YEAR_OFFSET: Dict[int, int] = {
    0: 967,
    1: 968,
    2: 974,
}

ITEM_LONGER_TERM_PROGRAMMES = 954
ITEM_EMERGENCY_APPEALS = 960
ITEM_BILATERAL_SUPPORT = 955
ITEM_COMMENTS = 956  # Legacy textarea item; comments now import into discussion panel
ITEM_FUNDING_REQUIREMENTS_T22 = 1303  # Template 22 – Funding Requirements (rows=country_map)
T22_ROW_TOTAL_COLUMN = "Total"  # matrix row-total cell suffix (row_total_manual_enabled)
# Item 1303 variable columns (Excel Area names) — overridden when PNS reports totals only.
T22_BREAKDOWN_AREAS: Tuple[str, ...] = ("SP1", "SP2", "SP3", "SP4", "SP5", "EFs")
# Legacy UPR Master ``Area`` codes on Funding rows (pre-SP1/EFs workbook naming).
# E1/E2/E3/EO remain skipped; EA1–EA3 map to selectable-header columns on the hybrid
# funding matrix (items 967/968/974) — see ``PLANNING_EA_FUNDING_AREAS``.
SKIPPED_LEGACY_FUNDING_AREAS = frozenset({"E1", "E2", "E3", "EO"})
PLANNING_EA_FUNDING_AREAS = frozenset({"EA1", "EA2", "EA3"})
EMERGENCY_APPEALS_COLUMN = "Total People to be reached"

# ── Reporting country template (T33) ───────────────────────────────────────────
REPORTING_COUNTRY_TEMPLATE_ID = 33

# Label needles for version-aware item resolution (substring match, case-insensitive).
REPORTING_SPECIAL_ITEM_LABELS: Dict[str, Tuple[str, ...]] = {
    "funding": ("ns total funding", "ns 2025 total funding", "ns 2026 total funding"),
    "expenditure": ("ns total expenditure", "ns 2025 total expenditure", "ns 2026 total expenditure"),
    "sp_breakdown": ("optional breakdown by sp/ef",),
    "support": ("received support",),
}
T22_STAFF_MATRIX_LABELS: Tuple[str, ...] = ("pns staff contributions",)
T23_PNS_FUNDING_LABELS: Tuple[str, ...] = ("pns funding", "funding matrix")

# Fallback item ids when label lookup fails on the published version.
ITEM_REPORTING_COUNTRY_FUNDING = 1403
ITEM_REPORTING_COUNTRY_EXPENDITURE = 1404
ITEM_REPORTING_COUNTRY_SP_BREAKDOWN = 1405
ITEM_REPORTING_COUNTRY_SUPPORT = 1407
ITEM_REPORTING_PNS_FUNDING = 952

# Row names in country-reporting Total Funding matrix.
REPORTING_FUNDING_ROW_IFRC = "IFRC Secretariat"
REPORTING_FUNDING_ROW_PNS = "PNSs"
REPORTING_FUNDING_ROW_OTHER = "HNS other sources"
# Fallback matrix column ``name`` for item 1403 when published config cannot be read.
# Cell keys are ``{row}_{column_name}`` — must match FormItem matrix_config.columns[0].name.
REPORTING_FUNDING_MATRIX_COLUMN = "ns_fun"

# SP/EF breakdown matrix: Excel Area → manual matrix row label
REPORTING_SP_BREAKDOWN_AREA_TO_ROW: Dict[str, str] = {
    "SP1": "Resilience - Climate and environment",
    "SP2": "Response - Disasters and crises",
    "SP3": "Resilience - Health and wellbeing",
    "SP4": "Resilience - Migration and displacement",
    "SP5": "Respect - Values, power and inclusion",
    "EFs": "Enabling functions",
}

# SP/EF breakdown matrix: indicator bank id → matrix column name
REPORTING_SP_BREAKDOWN_COLUMNS: Dict[int, str] = {
    733: "Funding (CHF)",
    734: "Expenditure (CHF)",
}

REPORTING_EMERGENCY_EXCEL_SECTION_TO_SLOT: Dict[str, int] = {
    "Emergency 1": 1,
    "Emergency 2": 2,
    "Emergency 3": 3,
}
NS_EMERGENCY_NAME_PREFIX = "data_eo"
NS_EMERGENCY_CODE_PREFIX = "data_mdr"

# Core/Other indicators: Excel Area → form section name (when the same bank id
# appears on multiple section-scoped items, e.g. 619 on Cross Cutting + SP2).
REPORTING_EXCEL_AREA_TO_SECTION: Dict[str, str] = {
    "Cross-cutting": "Cross Cutting",
    "SP1": "Resilience - Climate and environment",
    "SP2": "Response - Disasters and crises",
    "SP3": "Resilience - Health and wellbeing",
    "SP4": "Resilience - Migration and displacement",
    "SP5": "Respect - Values, power and inclusion",
    "EF1": "Strategic and operational coordination",
    "EF2": "National Society development",
    "EF3": "Humanitarian diplomacy and communication",
    "EF4": "Accountability and agility",
}

# Normalized indicatorId → T23 item 952 column name for PNS-reported funding
T23_PNS_FUNDING_COLUMNS: Dict[int, str] = {
    733: "Total Funding",
    734: "Total Expenditure",
    5: "Total Transferred to HNS",   # indicatorId '00005' normalises to 5
    # 2 = Funding Requirement ('00002') — variable/readonly, populated from planning, skip
}

# Excel Comments_* indicator codes → labels used in discussion panel import rows.
COMMENT_INDICATOR_LABELS: Dict[str, str] = {
    "comments_fundingrequirements": "Funding requirements",
    "comments_keyfigures": "Key figures",
    "comments_reach": "People to be reached",
    "comments_support": "Bilateral support",
}

AGGREGATE_ATTRIBUTE = frozenset({"Total", "SubTotal"})
AGGREGATE_AREA = frozenset({"Total", "SubTotal", "EAs"})


class UprImportCancelled(FdrsSyncCancelled):
    """Raised when a UPR Excel import is cancelled."""


@dataclass
class UprImportContext:
    """Resolved template items and assignment lookups."""

    template_ids: List[int]
    assignment_by_period_iso: Dict[Tuple[str, str], int] = field(default_factory=dict)
    # Per-template assignment maps (avoids ISO3 key collisions between tpl 24 and tpl 22).
    assignment_by_template: Dict[int, Dict[Tuple[str, str], int]] = field(default_factory=dict)
    items_by_bank_id: Dict[int, Dict[int, int]] = field(default_factory=dict)  # template_id -> bank_id -> item_id
    # template_id -> bank_id -> section_name -> item_id (for section-scoped duplicate bank ids)
    items_by_bank_section: Dict[int, Dict[int, Dict[str, int]]] = field(default_factory=dict)
    item_ids_by_label: Dict[int, Dict[str, int]] = field(default_factory=dict)
    published_version_ids: Dict[int, int] = field(default_factory=dict)
    # template_id -> {funding, expenditure, sp_breakdown, support, funding_col}
    reporting_special_items: Dict[int, Dict[str, Any]] = field(default_factory=dict)
    other_indicators_section_id: Optional[int] = None
    ea_repeat_section_id: Optional[int] = None
    ea_dynamic_section_id: Optional[int] = None
    emergency_choice_item_id: Optional[int] = None
    emergency_slot_meta: Dict[Tuple[int, int], Dict[str, str]] = field(default_factory=dict)
    dynamic_indicator_entries: List[Dict[str, Any]] = field(default_factory=list)
    discussion_comment_entries: List[Dict[str, Any]] = field(default_factory=list)
    staff_matrix_item_id: int = 1314  # fallback when label lookup fails (prod T22)
    pns_funding_item_id: int = ITEM_REPORTING_PNS_FUNDING
    ns_name_to_id: Dict[str, int] = field(default_factory=dict)
    # NS name (lower) → home country ISO3 (for PNS funding → template 22 lookup)
    ns_home_country_iso3: Dict[str, str] = field(default_factory=dict)
    # ISO3 → Country.id (used as row key in country_map list_library matrices)
    country_id_by_iso3: Dict[str, int] = field(default_factory=dict)
    # ISO3 → NationalSociety.id for the host country's primary NS.
    # Used as the matrix row key for T22 Staff and T23 Funding (list_library national_society).
    iso3_to_hns_id: Dict[str, int] = field(default_factory=dict)
    emergency_ops_by_iso: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    emergency_ops_ordered_by_iso: Dict[str, List[Dict[str, Any]]] = field(default_factory=dict)
    emergency_matrix_plugin_config: Dict[str, Any] = field(default_factory=dict)
    emergency_go_plugin_config: Dict[str, Any] = field(default_factory=dict)
    yes_no_bank_ids: Set[int] = field(default_factory=set)
    percentage_bank_ids: Set[int] = field(default_factory=set)
    percentage_allow_over_100_bank_ids: Set[int] = field(default_factory=set)
    indicator_bank_ids: Set[int] = field(default_factory=set)
    core_yes_no_item_ids: List[int] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    # PNS assignment AES ids with at least one Excel Funding row where PNS reported = Yes.
    pns_t22_reported_aes: Set[int] = field(default_factory=set)
    pns_t23_reported_aes: Set[int] = field(default_factory=set)
    # (aes_id, item_id) / (aes_id, section_id, bank_id, repeat_instance_number) pairs that
    # already carry a real sex/age/sex_age breakdown in the DB — see
    # _load_existing_disaggregated_keys. UPR Master's flat ValueNum has no breakdown columns,
    # so writing it here unguarded would silently erase that existing breakdown.
    existing_disagg_static_keys: Set[Tuple[int, int]] = field(default_factory=set)
    existing_disagg_dynamic_keys: Set[Tuple[int, int, int, Optional[int]]] = field(default_factory=set)


def round_to_period(round_code: str) -> Optional[str]:
    """Map UPR round code to assignment period name (e.g. P26 -> 2026)."""
    r = (round_code or "").strip().upper()
    if not r:
        return None
    if r.startswith("P") and len(r) >= 3 and r[1:].isdigit():
        return str(2000 + int(r[1:]))
    if r.startswith("AR") and len(r) >= 4 and r[2:].isdigit():
        return str(2000 + int(r[2:]))
    if r.startswith("MYR") and len(r) >= 5 and r[3:].isdigit():
        return f"Jan-Jun {2000 + int(r[3:])}"
    return None


FUNDING_REQUIREMENT_BANK_ID = 2


def normalize_indicator_id(raw: Any) -> Optional[int]:
    if raw is None or raw == "":
        return None
    try:
        return int(str(raw).strip())
    except (ValueError, TypeError):
        return None


def is_skipped_legacy_funding_area(area: Any) -> bool:
    """True when UPR Master uses a legacy Funding ``Area`` code we do not map yet."""
    return str(area or "").strip() in SKIPPED_LEGACY_FUNDING_AREAS


def is_planning_funding_requirement_row(row: Dict[str, Any]) -> bool:
    """True for planning ``Funding Requirement`` rows (bank id 2) we import to templates 24/22.

    ``Confirmed Funding`` is a separate UPR Master indicator with no backoffice form field yet —
    those rows are ignored by the import.
    """
    indicator = str(row.get("Indicator") or "").strip().lower()
    if indicator == "confirmed funding":
        return False
    indicator_id = normalize_indicator_id(row.get("indicatorId"))
    return indicator_id == FUNDING_REQUIREMENT_BANK_ID or indicator == "funding requirement"


def parse_value_num(raw: Any) -> Optional[float]:
    if raw is None or raw == "":
        return None
    try:
        return float(raw)
    except (ValueError, TypeError):
        return None


COL_PNS_REPORTED = "PNS reported"


def parse_pns_reported_yes(row: Dict[str, Any]) -> bool:
    """True when the Excel ``PNS reported`` column is Yes for this Funding row."""
    raw = row.get(COL_PNS_REPORTED)
    if raw is None or raw == "":
        return False
    return str(raw).strip().lower() in ("yes", "y", "1", "true")


def _build_pns_reported_yes_sets(
    rows: List[Dict[str, Any]],
    ctx: UprImportContext,
    template_ids: List[int],
    *,
    rounds: Optional[Set[str]] = None,
) -> Tuple[Set[Tuple[int, str]], Set[Tuple[int, str]]]:
    """Collect ``(pns_aes_id, host_iso3)`` pairs where Excel ``PNS reported`` is Yes."""
    t22_yes: Set[Tuple[int, str]] = set()
    t23_yes: Set[Tuple[int, str]] = set()
    if 22 not in template_ids and 23 not in template_ids:
        return t22_yes, t23_yes

    for row in rows:
        rnd = str(row.get("Round") or "").strip().upper()
        if rounds and rnd not in rounds:
            continue
        if not parse_pns_reported_yes(row):
            continue
        if str(row.get("Section") or "").strip() != "Funding":
            continue
        if str(row.get("Entity") or "").strip().upper() != "PNS":
            continue
        if str(row.get("Round") or "").strip().upper().startswith("P") and not is_planning_funding_requirement_row(
            row
        ):
            continue
        iso3 = str(row.get("ISO3") or "").strip().upper()
        rnd = str(row.get("Round") or "").strip().upper()
        period = round_to_period(rnd)
        ns_name = str(row.get("NS") or "").strip()
        if not iso3 or not period or not ns_name or ns_name.lower() == "country":
            continue
        pns_iso3 = ctx.ns_home_country_iso3.get(ns_name.lower())
        if not pns_iso3:
            continue
        if rnd.startswith("P") and 22 in template_ids:
            pns_aes = ctx.assignment_by_template.get(22, {}).get((period, pns_iso3))
            if pns_aes:
                t22_yes.add((pns_aes, iso3))
        elif (rnd.startswith("AR") or rnd.startswith("MYR")) and 23 in template_ids:
            pns_aes = ctx.assignment_by_template.get(23, {}).get((period, pns_iso3))
            if pns_aes:
                t23_yes.add((pns_aes, iso3))
    return t22_yes, t23_yes


def _periods_for_import_rounds(rounds: Optional[Set[str]]) -> Optional[Set[str]]:
    """Map selected UPR round codes to assignment period names (None = all periods)."""
    if not rounds:
        return None
    periods: Set[str] = set()
    for rnd in rounds:
        period = round_to_period(rnd)
        if period:
            periods.add(period)
    return periods or None


def plan_non_reported_pns_aes_by_template(
    ctx: UprImportContext,
    template_ids: List[int],
    *,
    periods: Optional[Set[str]] = None,
) -> Dict[int, Set[int]]:
    """Map template id → AES ids with no ``PNS reported = Yes`` in the import scope."""
    out: Dict[int, Set[int]] = {22: set(), 23: set()}
    template_reported = (
        (22, ctx.pns_t22_reported_aes),
        (23, ctx.pns_t23_reported_aes),
    )
    for tpl_id, reported_aes in template_reported:
        if tpl_id not in template_ids:
            continue
        for (period, _iso3), aes_id in ctx.assignment_by_template.get(tpl_id, {}).items():
            if periods and period not in periods:
                continue
            if int(aes_id) in reported_aes:
                continue
            out[tpl_id].add(int(aes_id))
    return out


def upr_pns_import_item_ids(ctx: UprImportContext, template_id: int) -> Set[int]:
    """Form item ids written by the UPR Excel import for a PNS template."""
    if template_id == 22:
        return {ITEM_FUNDING_REQUIREMENTS_T22, int(ctx.staff_matrix_item_id)}
    if template_id == 23:
        return {int(ctx.pns_funding_item_id)}
    return set()


def plan_pns_assignment_status_updates(
    ctx: UprImportContext,
    template_ids: List[int],
    *,
    periods: Optional[Set[str]] = None,
) -> List[Dict[str, Any]]:
    """Plan ``pending`` status for PNS templates with no ``PNS reported = Yes`` in Excel."""
    by_template = plan_non_reported_pns_aes_by_template(ctx, template_ids, periods=periods)
    plan: List[Dict[str, Any]] = []
    for aes_ids in by_template.values():
        for aes_id in aes_ids:
            plan.append(
                {
                    "assignment_entity_status_id": int(aes_id),
                    "status": "pending",
                }
            )
    return plan


def clear_non_reported_pns_form_data(
    ctx: UprImportContext,
    template_ids: List[int],
    *,
    rounds: Optional[Set[str]] = None,
    dry_run: bool = False,
) -> Tuple[Dict[str, int], Set[Tuple[int, int]]]:
    """
    Delete UPR-imported form_data on PNS assignments with no ``PNS reported = Yes``.

    Returns stats and ``(aes_id, form_item_id)`` source pairs for variable refresh.
    """
    from app.extensions import db
    from app.models.forms import FormData

    stats = {"form_data_deleted": 0, "form_data_delete_skipped": 0, "form_data_delete_errors": 0}
    updated_sources: Set[Tuple[int, int]] = set()
    by_template = plan_non_reported_pns_aes_by_template(
        ctx,
        template_ids,
        periods=_periods_for_import_rounds(rounds),
    )

    for tpl_id, aes_ids in by_template.items():
        item_ids = upr_pns_import_item_ids(ctx, tpl_id)
        if not aes_ids or not item_ids:
            continue
        for aes_id in aes_ids:
            try:
                query = FormData.query.filter(
                    FormData.assignment_entity_status_id == int(aes_id),
                    FormData.form_item_id.in_(sorted(item_ids)),
                )
                rows = query.all()
                if not rows:
                    stats["form_data_delete_skipped"] += 1
                    continue
                for row in rows:
                    updated_sources.add((int(aes_id), int(row.form_item_id)))
                if dry_run:
                    stats["form_data_deleted"] += len(rows)
                    continue
                deleted = query.delete(synchronize_session=False)
                stats["form_data_deleted"] += int(deleted or 0)
            except Exception as exc:
                stats["form_data_delete_errors"] += 1
                logger.error(
                    "PNS form_data clear error (aes_id=%s, template=%s): %s",
                    aes_id,
                    tpl_id,
                    exc,
                )

    if not dry_run and stats["form_data_deleted"] > 0:
        db.session.commit()
    return stats, updated_sources


def sync_non_reported_pns_assignments(
    ctx: UprImportContext,
    template_ids: List[int],
    *,
    rounds: Optional[Set[str]] = None,
    dry_run: bool = False,
) -> Dict[str, int]:
    """Clear stale form_data and reset workflow status for non-reported PNS assignments."""
    clear_stats, updated_sources = clear_non_reported_pns_form_data(
        ctx,
        template_ids,
        rounds=rounds,
        dry_run=dry_run,
    )
    status_plan = plan_pns_assignment_status_updates(
        ctx,
        template_ids,
        periods=_periods_for_import_rounds(rounds),
    )
    status_stats = apply_pns_pending_status_resets(status_plan, dry_run=dry_run)

    if not dry_run and updated_sources:
        from app.services.forms.variable_resolution_service import VariableResolutionService

        refresh_stats = VariableResolutionService.refresh_stale_variable_matrix_cells(
            updated_sources
        )
        clear_stats.update(
            {
                "variable_consumer_rows_checked": refresh_stats.get("consumer_rows_checked", 0),
                "variable_cells_cleared": refresh_stats.get("cells_cleared", 0),
                "variable_consumer_rows_updated": refresh_stats.get("rows_updated", 0),
            }
        )

    return {
        **clear_stats,
        "assignment_status_updated": status_stats.get("updated", 0),
        "assignment_status_skipped": status_stats.get("skipped", 0),
        "assignment_status_errors": status_stats.get("errors", 0),
        "assignment_status_planned": len(status_plan),
    }


def _pns_pending_reset_needs_update(aes: Any, *, assigned_at: Any) -> bool:
    """True when AES should be reset to pending with assignment-date metadata."""
    from app.models.enums import AssignmentEntityStatusValue

    current = aes.status.value if hasattr(aes.status, "value") else str(aes.status)
    if current != AssignmentEntityStatusValue.pending.value:
        return True
    if aes.status_timestamp != assigned_at:
        return True
    if aes.submitted_at is not None:
        return True
    if aes.submitted_by_user_id is not None:
        return True
    if aes.approved_by_user_id is not None:
        return True
    if aes.sent_for_review_by_user_id is not None:
        return True
    if aes.sent_for_review_at is not None:
        return True
    return False


def _apply_pns_pending_reset_fields(aes: Any, *, assigned_at: Any) -> None:
    """Reset workflow fields for a non-reported PNS assignment."""
    from app.models.enums import AssignmentEntityStatusValue

    aes.status = AssignmentEntityStatusValue.pending
    aes.status_timestamp = assigned_at
    aes.submitted_at = None
    aes.submitted_by_user_id = None
    aes.approved_by_user_id = None
    aes.sent_for_review_by_user_id = None
    aes.sent_for_review_at = None


def apply_pns_pending_status_resets(
    plan: List[Dict[str, Any]],
    *,
    dry_run: bool = False,
) -> Dict[str, int]:
    """Apply pending resets using parent assignment ``assigned_at`` as status date."""
    from app.extensions import db
    from app.models.assignments import AssignedForm, AssignmentEntityStatus
    from app.utils.datetime_helpers import utcnow

    stats = {"updated": 0, "skipped": 0, "errors": 0}
    if not plan:
        return stats

    aes_ids = [int(row["assignment_entity_status_id"]) for row in plan]
    aes_rows = AssignmentEntityStatus.query.filter(AssignmentEntityStatus.id.in_(aes_ids)).all()
    aes_by_id = {int(row.id): row for row in aes_rows}
    assigned_form_ids = {int(aes.assigned_form_id) for aes in aes_rows if aes.assigned_form_id}
    assigned_at_by_form = {
        int(form.id): form.assigned_at
        for form in AssignedForm.query.filter(AssignedForm.id.in_(assigned_form_ids)).all()
    }

    for row in plan:
        try:
            aes_id = int(row["assignment_entity_status_id"])
            aes = aes_by_id.get(aes_id)
            if aes is None:
                stats["skipped"] += 1
                continue

            assigned_at = assigned_at_by_form.get(int(aes.assigned_form_id)) or utcnow()
            if not _pns_pending_reset_needs_update(aes, assigned_at=assigned_at):
                stats["skipped"] += 1
                continue

            if dry_run:
                stats["updated"] += 1
                continue

            _apply_pns_pending_reset_fields(aes, assigned_at=assigned_at)
            db.session.add(aes)
            stats["updated"] += 1
        except Exception as exc:
            stats["errors"] += 1
            logger.error(
                "PNS pending status reset error (aes_id=%s): %s",
                row.get("assignment_entity_status_id"),
                exc,
            )

    if not dry_run and stats["updated"] > 0:
        db.session.commit()
    return stats


def _is_yes_no_indicator_type(type_value: Any) -> bool:
    normalized = str(type_value or "").strip().lower().replace("-", "").replace("_", "")
    return normalized in ("yesno", "boolean", "bool")


def _load_yes_no_bank_ids() -> Set[int]:
    """Indicator bank ids whose type is Yes/No (UPR Master stores 1/0 in ValueNum)."""
    from app.models.indicator_bank import IndicatorBank

    out: Set[int] = set()
    for row in IndicatorBank.query.filter_by(archived=False).all():
        if _is_yes_no_indicator_type(row.type):
            out.add(int(row.id))
    return out


def _load_indicator_bank_ids() -> Set[int]:
    """All indicator bank primary keys (used to validate dynamic import FK targets)."""
    from app.models.indicator_bank import IndicatorBank

    return {int(row.id) for row in IndicatorBank.query.with_entities(IndicatorBank.id).all()}


def _is_percentage_indicator_type(type_value: Any) -> bool:
    normalized = str(type_value or "").strip().lower().replace("/", "").replace("-", "").replace(" ", "")
    return normalized in ("percentage", "percent", "pct")


def _load_percentage_bank_ids() -> Set[int]:
    """Indicator bank ids whose type is Percentage (UPR Master ValueNum expected 0-100)."""
    from app.models.indicator_bank import IndicatorBank

    out: Set[int] = set()
    for row in IndicatorBank.query.filter_by(archived=False).all():
        if _is_percentage_indicator_type(row.type):
            out.add(int(row.id))
    return out


def _load_percentage_allow_over_100_bank_ids(bank_ids: Set[int]) -> Set[int]:
    """Percentage bank ids where at least one live form item explicitly allows values
    over 100% (cumulative/ratio-based indicators configured with ``allow_over_100``)."""
    if not bank_ids:
        return set()
    from app.models.form_items import FormItem

    out: Set[int] = set()
    rows = FormItem.query.filter(
        FormItem.indicator_bank_id.in_(bank_ids),
        FormItem.archived == False,  # noqa: E712
    ).all()
    for item in rows:
        config = item.config or {}
        if config.get("allow_over_100") in (True, "true", "True", 1, "1"):
            out.add(int(item.indicator_bank_id))
    return out


def _percentage_scalar_range_warning(
    ctx: "UprImportContext",
    indicator_bank_id: Optional[int],
    value_num: Optional[float],
    *,
    iso3: str = "",
    rnd: str = "",
    label: str = "",
) -> Optional[str]:
    """Flag a percentage-type indicator whose Master ``ValueNum`` is outside the
    plausible 0-100% range — the classic 'entered 500 instead of 50' mistake.
    Non-blocking: the value still imports as entered, but the mistake becomes visible
    instead of silently landing as a nonsensical number.
    """
    if not indicator_bank_id or value_num is None:
        return None
    if indicator_bank_id not in ctx.percentage_bank_ids:
        return None
    upper_bound = None if indicator_bank_id in ctx.percentage_allow_over_100_bank_ids else 100.0
    if value_num >= 0 and (upper_bound is None or value_num <= upper_bound):
        return None
    name = label or f"bank {indicator_bank_id}"
    where = " ".join(part for part in (iso3, rnd) if part)
    suffix = f" ({where})" if where else ""
    return (
        f"Percentage indicator {name!r} = {value_num:g} is outside the valid 0-100% "
        f"range{suffix} — please check for a data-entry mistake (e.g. 500 instead of 50)."
    )


def _master_yes_no_value(value_num: Optional[float]) -> str:
    """Map UPR Master ValueNum to entry-form yes/no storage."""
    if value_num is not None and float(value_num) == 1:
        return "yes"
    return "no"


def _reporting_indicator_has_import_value(
    ctx: UprImportContext,
    indicator_bank_id: int,
    value_num: Optional[float],
    *,
    is_dna: bool,
) -> bool:
    if is_dna:
        return True
    if indicator_bank_id in ctx.yes_no_bank_ids:
        return True
    return value_num is not None


def _reporting_indicator_import_value(
    ctx: UprImportContext,
    indicator_bank_id: int,
    value_num: Optional[float],
    *,
    iso3: str = "",
    rnd: str = "",
    label: str = "",
) -> Any:
    if indicator_bank_id in ctx.yes_no_bank_ids:
        return _master_yes_no_value(value_num)
    warning = _percentage_scalar_range_warning(
        ctx, indicator_bank_id, value_num, iso3=iso3, rnd=rnd, label=label
    )
    if warning:
        ctx.warnings.append(warning)
    return value_num


def _has_real_disagg_breakdown(disagg: Any) -> bool:
    """True when a disagg_data payload is an actual sex/age/sex_age breakdown
    with at least one non-null sub-value (not just ``{"mode": "total", ...}``).
    """
    if not isinstance(disagg, dict):
        return False
    if disagg.get("mode") not in ("sex", "age", "sex_age"):
        return False
    values = disagg.get("values")
    if not isinstance(values, dict):
        return False
    direct = values.get("direct", values)
    return isinstance(direct, dict) and any(v is not None for v in direct.values())


def _load_existing_disaggregated_keys(
    aes_ids: Set[int],
) -> Tuple[Set[Tuple[int, int]], Set[Tuple[int, int, int, Optional[int]]]]:
    """Find static (aes_id, item_id) and dynamic (aes_id, section_id, bank_id,
    repeat_instance_number) keys that already carry a real sex/age/sex_age breakdown.

    UPR Master's "UPR Data" sheet is flat — one ``ValueNum`` per country/indicator/round,
    with no Male/Female/Age columns — so a scalar value built from it must never be written
    over an indicator that already has a real breakdown (e.g. entered via the T33
    per-country Excel import or directly in the web form): both
    ``FormData.set_simple_value()`` and ``DynamicIndicatorData.set_simple_value()``
    (``DataEntryMixin``, app/models/forms.py) unconditionally clear ``disagg_data`` when a
    plain value is set, which would silently destroy the breakdown.
    """
    static_keys: Set[Tuple[int, int]] = set()
    dynamic_keys: Set[Tuple[int, int, int, Optional[int]]] = set()
    if not aes_ids:
        return static_keys, dynamic_keys

    from app.models.forms import DynamicIndicatorData, FormData

    aes_list = list(aes_ids)
    chunk_size = 1000
    for i in range(0, len(aes_list), chunk_size):
        chunk = aes_list[i : i + chunk_size]
        static_rows = (
            FormData.query.filter(
                FormData.assignment_entity_status_id.in_(chunk),
                FormData.disagg_data.isnot(None),
            )
            .with_entities(
                FormData.assignment_entity_status_id,
                FormData.form_item_id,
                FormData.disagg_data,
            )
            .all()
        )
        for aes_id, item_id, disagg in static_rows:
            if _has_real_disagg_breakdown(disagg):
                static_keys.add((int(aes_id), int(item_id)))

        dynamic_rows = (
            DynamicIndicatorData.query.filter(
                DynamicIndicatorData.assignment_entity_status_id.in_(chunk),
                DynamicIndicatorData.disagg_data.isnot(None),
            )
            .with_entities(
                DynamicIndicatorData.assignment_entity_status_id,
                DynamicIndicatorData.section_id,
                DynamicIndicatorData.indicator_bank_id,
                DynamicIndicatorData.repeat_instance_number,
                DynamicIndicatorData.disagg_data,
            )
            .all()
        )
        for aes_id, section_id, bank_id, repeat_num, disagg in dynamic_rows:
            if _has_real_disagg_breakdown(disagg):
                dynamic_keys.add((int(aes_id), int(section_id), int(bank_id), repeat_num))

    return static_keys, dynamic_keys


def _disaggregation_overwrite_warning(
    ctx: UprImportContext,
    *,
    value_num: Optional[float],
    is_dna: bool,
    static_key: Optional[Tuple[int, int]] = None,
    dynamic_key: Optional[Tuple[int, int, int, Optional[int]]] = None,
    iso3: str = "",
    rnd: str = "",
    label: str = "",
) -> Optional[str]:
    """Guard against UPR Master's flat ValueNum silently erasing an existing sex/age
    breakdown for the same indicator/assignment (see _load_existing_disaggregated_keys).
    Returns a warning message when the write should be skipped, else None.
    """
    if is_dna or value_num is None:
        return None
    conflict = bool(
        (static_key and static_key in ctx.existing_disagg_static_keys)
        or (dynamic_key and dynamic_key in ctx.existing_disagg_dynamic_keys)
    )
    if not conflict:
        return None
    name = label or "indicator"
    where = " ".join(part for part in (iso3, rnd) if part)
    suffix = f" ({where})" if where else ""
    return (
        f"Skipped UPR Master value for {name!r}{suffix} — this indicator already has a "
        f"sex/age breakdown saved from a richer source. Importing the flat total "
        f"({value_num:g}) would have erased that breakdown, so the existing data was left "
        f"unchanged; update the breakdown directly if the total has actually changed."
    )


def _load_core_yes_no_item_ids(
    template_id: int,
    version_id: int,
    *,
    other_section_id: Optional[int],
    yes_no_bank_ids: Set[int],
) -> List[int]:
    """Published-version static form items for Yes/No core indicators (excludes Other dynamic)."""
    from app.models.form_items import FormItem

    skip_sections = {int(other_section_id)} if other_section_id else set()
    out: List[int] = []
    for item in FormItem.query.filter_by(
        template_id=int(template_id),
        version_id=int(version_id),
        archived=False,
    ):
        if item.section_id in skip_sections:
            continue
        bank_id = item.indicator_bank_id
        is_yes_no = bool(
            (bank_id and int(bank_id) in yes_no_bank_ids)
            or _is_yes_no_indicator_type(getattr(item, "type", None))
        )
        if is_yes_no:
            out.append(int(item.id))
    return out


def _reporting_aes_ids_from_excel(
    rows: List[Dict[str, Any]],
    ctx: UprImportContext,
    *,
    template_ids: List[int],
) -> Set[int]:
    """T33 assignment ids for reporting round+country pairs present in the Excel rows.

    Yes/No defaults must not touch assignments for rounds that are absent from UPR
    Master (e.g. MYR26 when the workbook only contains MYR25), even when those
    assignments already exist in the database.
    """
    if REPORTING_COUNTRY_TEMPLATE_ID not in template_ids:
        return set()
    t33_sections = UPR_TEMPLATE_PROFILES[REPORTING_COUNTRY_TEMPLATE_ID]["sections"]
    tpl_map = ctx.assignment_by_template.get(REPORTING_COUNTRY_TEMPLATE_ID, {})
    if not tpl_map:
        return set()
    out: Set[int] = set()
    seen_period_iso: Set[Tuple[str, str]] = set()
    for row in rows:
        rnd = str(row.get("Round") or "").strip().upper()
        if not (rnd.startswith("AR") or rnd.startswith("MYR")):
            continue
        sec = str(row.get("Section") or "").strip()
        if sec not in t33_sections:
            continue
        iso3 = str(row.get("ISO3") or "").strip().upper()
        if not iso3:
            continue
        period = round_to_period(rnd)
        if not period:
            continue
        key = (period, iso3)
        if key in seen_period_iso:
            continue
        seen_period_iso.add(key)
        aes_id = tpl_map.get(key)
        if aes_id is not None:
            out.add(int(aes_id))
    return out


def _fill_missing_core_yes_no_defaults(
    *,
    ctx: UprImportContext,
    import_rows: List[Dict[str, str]],
    filled_core_yes_no: Set[Tuple[int, int]],
    target_aes_ids: Set[int],
    aes_meta: Dict[int, Tuple[str, str]],
) -> None:
    """Default missing Yes/No core indicators to ``no`` for Excel-present reporting rounds."""
    if not ctx.core_yes_no_item_ids or not target_aes_ids:
        return
    core_ids = set(ctx.core_yes_no_item_ids)
    for aes_id in target_aes_ids:
        iso3, period = aes_meta.get(aes_id, ("", ""))
        for item_id in core_ids:
            if (aes_id, item_id) in filled_core_yes_no:
                continue
            built = _scalar_row(
                aes_id=aes_id,
                item_id=item_id,
                value="no",
                iso3=iso3,
                period=period,
                debug_kpi=f"core_yesno_default_{item_id}",
            )
            if built:
                import_rows.append(built)
                filled_core_yes_no.add((aes_id, item_id))


def is_aggregate_row(row: Dict[str, Any]) -> bool:
    """Skip rollup rows (EAs, section totals).

    NS Data and Comments use Area=Total for real data rows, not rollups.
    Funding Area=Total rows carry PNS aggregate values for T22 row-total cells.
    """
    sec = (row.get("Section") or "").strip()
    area = (row.get("Area") or "").strip()
    if area == "EAs":
        return True
    if area in AGGREGATE_AREA and sec not in ("NS Data", "Comments"):
        if sec == "Funding" and area == "Total":
            return False
        return True
    return False


def humanize_comment_label(indicator: str) -> str:
    """Map Excel Comments_* codes to readable section labels."""
    key = (indicator or "").strip().lower()
    if key in COMMENT_INDICATOR_LABELS:
        return COMMENT_INDICATOR_LABELS[key]
    if key.startswith("comments_"):
        slug = key[len("comments_") :].replace("_", " ").strip()
        if slug:
            return slug[0].upper() + slug[1:]
    label = (indicator or "").strip()
    return label or "Comment"


def parse_comment_value(row: Dict[str, Any]) -> Optional[str]:
    """Comment text lives in Value (not ValueNum). UPR Value is the same column duplicated in some exports."""
    for key in ("Value", "UPR Value"):
        raw = row.get(key)
        if raw is None:
            continue
        text = str(raw).strip()
        if text:
            return text
    return None


def _file_fingerprint(path: str) -> Tuple[int, int]:
    st = os.stat(path)
    return (st.st_mtime_ns, st.st_size)


def rows_cache_path(path: str) -> str:
    return f"{path}.rows.v{ROWS_CACHE_VERSION}.pkl"


def clear_upr_import_caches(path: Optional[str]) -> None:
    """Remove on-disk row/transform caches for an uploaded workbook path."""
    if not path:
        return
    for suffix in (rows_cache_path(path),):
        if os.path.isfile(suffix):
            try:
                os.remove(suffix)
            except OSError:
                pass
    prefix = f"{path}.transform."
    parent = os.path.dirname(path) or "."
    base = os.path.basename(path)
    try:
        for name in os.listdir(parent):
            if name.startswith(f"{base}.transform.") and name.endswith(".pkl"):
                try:
                    os.remove(os.path.join(parent, name))
                except OSError:
                    pass
    except OSError:
        pass


def load_upr_data_sheet(path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Excel support requires openpyxl: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if UPR_DATA_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet {UPR_DATA_SHEET!r} not found in workbook")
    ws = wb[UPR_DATA_SHEET]
    headers: List[str] = []
    rows: List[Dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == HEADER_ROW_INDEX:
            headers = [str(h).strip() if h is not None else "" for h in row]
            continue
        if i < HEADER_ROW_INDEX:
            continue
        if not any(c is not None and str(c).strip() for c in row):
            continue
        record = {}
        for j, h in enumerate(headers):
            if not h:
                continue
            val = row[j] if j < len(row) else None
            record[h] = val
        rows.append(record)
    wb.close()
    return headers, rows


def load_upr_data_sheet_cached(path: str, *, use_cache: bool = True) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Load UPR Data rows, reusing a pickle cache keyed by file mtime/size."""
    cache_path = rows_cache_path(path)
    fingerprint = _file_fingerprint(path)
    if use_cache and os.path.isfile(cache_path):
        try:
            with open(cache_path, "rb") as fh:
                cached_fp, headers, rows = pickle.load(fh)
            if cached_fp == fingerprint:
                return headers, rows
        except Exception as exc:
            logger.debug("UPR rows cache read failed: %s", exc)

    headers, rows = load_upr_data_sheet(path)
    try:
        with open(cache_path, "wb") as fh:
            pickle.dump((fingerprint, headers, rows), fh, protocol=pickle.HIGHEST_PROTOCOL)
    except Exception as exc:
        logger.debug("UPR rows cache write failed: %s", exc)
    return headers, rows


def summarize_workbook_from_rows(headers: List[str], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Summarize already-loaded UPR Data rows for the import wizard."""
    rounds: Set[str] = set()
    sections: Set[str] = set()
    years: Set[str] = set()
    iso3s: Set[str] = set()
    by_template_section: Dict[str, int] = defaultdict(int)

    for row in rows:
        rnd = str(row.get("Round") or "").strip()
        sec = str(row.get("Section") or "").strip()
        yr = row.get("Year")
        iso3 = str(row.get("ISO3") or "").strip()
        if rnd:
            rounds.add(rnd)
        if sec:
            sections.add(sec)
        if yr not in (None, ""):
            years.add(str(yr).strip())
        if iso3:
            iso3s.add(iso3)
        for tid, profile in UPR_TEMPLATE_PROFILES.items():
            prefixes = profile.get("round_prefixes") or ()
            if sec in profile.get("sections", frozenset()) and any(rnd.startswith(p) for p in prefixes):
                by_template_section[f"{tid}:{sec}"] += 1

    planning_rounds = sorted(r for r in rounds if r.startswith("P"))
    ar_rounds = sorted(r for r in rounds if r.startswith("AR"))
    myr_rounds = sorted(r for r in rounds if r.startswith("MYR"))
    return {
        "success": True,
        "sheet": UPR_DATA_SHEET,
        "headers": headers,
        "total_rows": len(rows),
        "rounds": sorted(rounds),
        "planning_rounds": planning_rounds,
        "ar_rounds": ar_rounds,
        "myr_rounds": myr_rounds,
        "sections": sorted(sections),
        "years": sorted(years),
        "countries": len(iso3s),
        "templates": [
            {"id": tid, "name": prof["name"], "sections": sorted(prof["sections"])}
            for tid, prof in UPR_TEMPLATE_PROFILES.items()
        ],
        "row_counts_by_template_section": dict(by_template_section),
    }


def analyze_workbook(path: str, *, use_cache: bool = True) -> Dict[str, Any]:
    """Summarize workbook for the import wizard."""
    headers, rows = load_upr_data_sheet_cached(path, use_cache=use_cache)
    return summarize_workbook_from_rows(headers, rows)


def _normalize_round_set(rounds: Optional[List[str]]) -> Optional[Set[str]]:
    if not rounds:
        return None
    out = {r.strip().upper() for r in rounds if r and str(r).strip()}
    return out or None


def _transform_cache_key(path: str, template_ids: List[int], rounds: Optional[Set[str]]) -> str:
    fingerprint = _file_fingerprint(path)
    tids = ",".join(str(t) for t in sorted(template_ids))
    rnd = ",".join(sorted(rounds)) if rounds else "*"
    digest = hashlib.sha256(
        f"{TRANSFORM_CACHE_VERSION}|{fingerprint}|{tids}|{rnd}".encode()
    ).hexdigest()
    return digest[:20]


def transform_cache_path(path: str, cache_key: str) -> str:
    return f"{path}.transform.{cache_key}.pkl"


def _write_transform_cache(
    path: str,
    template_ids: List[int],
    rounds: Optional[Set[str]],
    import_rows: List[Dict[str, str]],
    ctx: "UprImportContext",
) -> None:
    cache_key = _transform_cache_key(path, template_ids, rounds)
    cache_path = transform_cache_path(path, cache_key)
    try:
        with open(cache_path, "wb") as fh:
            pickle.dump(
                (
                    TRANSFORM_CACHE_VERSION,
                    _file_fingerprint(path),
                    sorted(template_ids),
                    rounds,
                    import_rows,
                    ctx,
                ),
                fh,
                protocol=pickle.HIGHEST_PROTOCOL,
            )
    except Exception as exc:
        logger.debug("UPR transform cache write failed: %s", exc)


def load_transform_cache(
    path: str,
    template_ids: List[int],
    rounds: Optional[Set[str]],
) -> Optional[Tuple[List[Dict[str, str]], "UprImportContext"]]:
    cache_key = _transform_cache_key(path, template_ids, rounds)
    cache_path = transform_cache_path(path, cache_key)
    if not os.path.isfile(cache_path):
        return None
    try:
        with open(cache_path, "rb") as fh:
            (
                version,
                fingerprint,
                cached_tids,
                cached_rounds,
                import_rows,
                ctx,
            ) = pickle.load(fh)
        if version != TRANSFORM_CACHE_VERSION:
            return None
        if fingerprint != _file_fingerprint(path):
            return None
        if sorted(template_ids) != list(cached_tids):
            return None
        if cached_rounds != rounds:
            return None
        return import_rows, ctx
    except Exception as exc:
        logger.debug("UPR transform cache read failed: %s", exc)
        return None


def prepare_upr_transform(
    input_path: str,
    template_ids: List[int],
    *,
    rounds: Optional[List[str]] = None,
    use_row_cache: bool = True,
    use_transform_cache: bool = False,
    save_transform_cache: bool = False,
) -> Tuple[List[Dict[str, str]], "UprImportContext", bool]:
    """Load workbook rows, build context, and transform to import rows."""
    tids = [int(t) for t in template_ids]
    round_set = _normalize_round_set(rounds)

    if use_transform_cache:
        cached = load_transform_cache(input_path, tids, round_set)
        if cached:
            return cached[0], cached[1], True

    _, rows = load_upr_data_sheet_cached(input_path, use_cache=use_row_cache)
    ctx = build_import_context(tids)
    import_rows = transform_to_import_rows(rows, ctx, template_ids=tids, rounds=round_set)

    if save_transform_cache:
        _write_transform_cache(input_path, tids, round_set, import_rows, ctx)

    return import_rows, ctx, False


def _build_ns_name_index() -> Dict[str, int]:
    from app.models.organization import NationalSociety

    out: Dict[str, int] = {}
    for ns in NationalSociety.query.filter_by(is_active=True).all():
        name = (ns.name or "").strip()
        if name:
            out[name.lower()] = int(ns.id)
    return out


def _build_ns_home_country_index() -> Tuple[Dict[str, str], Dict[str, int], Dict[str, int]]:
    """Build three indexes from the NationalSociety / Country tables in one pass.

    Returns:
        ns_home_iso3  – NS name (lower) → home country ISO3
        country_id    – ISO3 (upper) → Country.id  (for country_map matrix row keys)
        iso3_to_ns_id – ISO3 (upper) → NationalSociety.id for the country's primary active NS
                        (row key for list_library national_society matrices in T22 Staff / T23 Funding)
    """
    from app.models.organization import NationalSociety
    from app.models.core import Country
    from app.extensions import db

    rows = (
        db.session.query(NationalSociety.id, NationalSociety.name, Country.iso3, Country.id)
        .join(Country, Country.id == NationalSociety.country_id)
        .filter(NationalSociety.is_active == True, Country.iso3.isnot(None))
        .all()
    )
    ns_home: Dict[str, str] = {}
    cid: Dict[str, int] = {}
    iso3_ns: Dict[str, int] = {}
    for ns_id, ns_name, iso3, c_id in rows:
        name_key = (ns_name or "").strip().lower()
        iso3_up = iso3.strip().upper()
        if name_key:
            ns_home[name_key] = iso3_up
        if iso3_up:
            cid[iso3_up] = int(c_id)
            # First active NS per ISO3 is the host NS (HNS).
            if iso3_up not in iso3_ns:
                iso3_ns[iso3_up] = int(ns_id)
    # Also fill country_id for countries that may not have an active NS
    for c in Country.query.filter(Country.iso3.isnot(None)).all():
        iso3_up = c.iso3.strip().upper()
        if iso3_up not in cid:
            cid[iso3_up] = int(c.id)
    return ns_home, cid, iso3_ns


def _load_emergency_choice_plugin_config(choice_item_id: int) -> Dict[str, Any]:
    """GO filter config from the T33 emergency_operations single-choice field."""
    from app.models.form_items import FormItem

    item = FormItem.query.get(int(choice_item_id))
    if not item or not isinstance(item.config, dict):
        return {}
    cfg = item.config
    if item.lookup_list_id == "emergency_operations":
        qpc = cfg.get("question_plugin_config")
        if isinstance(qpc, dict):
            return qpc
    plugin_cfg = cfg.get("plugin_config")
    return plugin_cfg if isinstance(plugin_cfg, dict) else {}


def _active_emergency_go_config(ctx: UprImportContext) -> Dict[str, Any]:
    return ctx.emergency_go_plugin_config or ctx.emergency_matrix_plugin_config


def _load_emergency_matrix_plugin_config(template_id: int = 24) -> Dict[str, Any]:
    """Plugin filter config from the Emergency Appeals matrix (item 960)."""
    from app.models.form_items import FormItem

    item = FormItem.query.get(ITEM_EMERGENCY_APPEALS)
    if not item or int(item.template_id or 0) != template_id:
        return {}
    matrix_cfg = (item.config or {}).get("matrix_config") or {}
    plugin_cfg = matrix_cfg.get("plugin_config") or {}
    return plugin_cfg if isinstance(plugin_cfg, dict) else {}


def _emergency_go_config(plugin_cfg: Dict[str, Any]) -> Dict[str, Any]:
    """Map matrix plugin_config keys to get_emergency_operations_data config."""
    operation_types = plugin_cfg.get("emops_operation_types") or plugin_cfg.get("operation_types") or ["All"]
    show_closed = plugin_cfg.get("emops_show_closed_operations")
    if show_closed is None:
        show_closed = plugin_cfg.get("show_closed_operations", True)
    cfg: Dict[str, Any] = {
        "operation_types": operation_types,
        "show_closed_operations": bool(show_closed),
    }
    end_date_gt = plugin_cfg.get("emops_end_date_gt") or plugin_cfg.get("end_date_gt")
    if end_date_gt:
        cfg["end_date_gt"] = str(end_date_gt)
    start_date = plugin_cfg.get("emops_start_date") or plugin_cfg.get("start_date")
    if start_date:
        cfg["start_date"] = str(start_date)
    return cfg


def _emergency_op_row_id(op: Dict[str, Any]) -> str:
    """Matrix row id for emergency_operations list_library (name_with_code)."""
    name = (op.get("name") or "").strip()
    code = (op.get("code") or "").strip()
    return f"{name} ({code})" if code else name


def _fetch_emergency_ops_for_country(iso3: str, plugin_cfg: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    """Fetch GO emergency operations for a country; return ordered list and code index."""
    from app.services.forms.emergency_section_binding import _fetch_ordered_operations

    go_cfg = _emergency_go_config(plugin_cfg)
    ops = _fetch_ordered_operations(iso3.upper(), go_cfg)
    by_code = {(op.get("code") or "").strip().upper(): op for op in ops if (op.get("code") or "").strip()}
    return ops, by_code


def _ensure_emergency_ops(ctx: UprImportContext, iso3: str) -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    key = iso3.upper()
    if key not in ctx.emergency_ops_ordered_by_iso:
        plugin_cfg = _active_emergency_go_config(ctx)
        ordered, by_code = _fetch_emergency_ops_for_country(key, plugin_cfg)
        ctx.emergency_ops_ordered_by_iso[key] = ordered
        ctx.emergency_ops_by_iso[key] = by_code
    return ctx.emergency_ops_ordered_by_iso[key], ctx.emergency_ops_by_iso[key]


def _aes_id_to_iso3(ctx: UprImportContext, aes_id: int) -> str:
    tpl_map = ctx.assignment_by_template.get(REPORTING_COUNTRY_TEMPLATE_ID, {})
    for (_period, iso), aid in tpl_map.items():
        if int(aid) == int(aes_id):
            return str(iso).upper()
    return ""


def _format_emergency_operation_display(name: str, code: str) -> str:
    name = (name or "").strip()
    code = (code or "").strip()
    if name and code:
        return f"{name} ({code})"
    return name or code


def _resolve_emergency_operation_labels(
    ctx: UprImportContext,
    *,
    iso3: str,
    excel_name: str,
    excel_code: str,
) -> Tuple[str, str, str]:
    """Return (name, code, display_value). Prefer GO API labels when MDR code matches."""
    name = (excel_name or "").strip()
    code = (excel_code or "").strip()
    code_upper = code.upper()
    if code_upper and iso3:
        _, by_code = _ensure_emergency_ops(ctx, iso3.upper())
        op = by_code.get(code_upper)
        if op:
            api_name = (op.get("name") or "").strip()
            api_code = (op.get("code") or "").strip()
            return api_name or name, api_code or code, _emergency_op_row_id(op)
        ctx.warnings.append(
            f"Emergency code {code!r} not found in GO API for {iso3} — using Excel name/code"
        )
    return name, code, _format_emergency_operation_display(name, code)


def _parse_ea_slot(area: str) -> Optional[int]:
    area = (area or "").strip().upper()
    if area in PLANNING_EA_FUNDING_AREAS:
        return int(area[-1])
    return None


def _build_reach_ea_code_index(
    rows: List[Dict[str, Any]],
    *,
    rounds: Optional[Set[str]] = None,
) -> Dict[Tuple[str, str, str], str]:
    """Map (iso3, round, EA slot) → EA Code from Reach rows (Funding fallback)."""
    index: Dict[Tuple[str, str, str], str] = {}
    for row in rows:
        rnd = str(row.get("Round") or "").strip().upper()
        if rounds and rnd not in rounds:
            continue
        if str(row.get("Section") or "").strip() != "Reach":
            continue
        area = str(row.get("Area") or "").strip()
        if area not in PLANNING_EA_FUNDING_AREAS:
            continue
        code = row.get("EA Code")
        if code in (None, ""):
            continue
        iso3 = str(row.get("ISO3") or "").strip().upper()
        if iso3 and rnd:
            index[(iso3, rnd, area)] = str(code).strip().upper()
    return index


def _resolve_ea_operation(
    ctx: UprImportContext,
    *,
    iso3: str,
    area: str,
    ea_code: Any,
    context: str = "Reach",
) -> Optional[Dict[str, Any]]:
    """Resolve a GO emergency operation for Excel EA1/EA2/EA3 slot or EA Code."""
    ordered, by_code = _ensure_emergency_ops(ctx, iso3)
    code = (str(ea_code).strip().upper() if ea_code not in (None, "") else "")
    if code:
        op = by_code.get(code)
        if op is None:
            ctx.warnings.append(f"Emergency appeal code {code!r} not found in GO API for {iso3}")
            return None
        return op

    slot = _parse_ea_slot(area)
    if slot is None:
        return None
    if slot < 1 or slot > len(ordered):
        ctx.warnings.append(
            f"No EA Code for {area} and only {len(ordered)} appeal(s) in GO API for {iso3} — skipped"
        )
        return None
    op = ordered[slot - 1]
    ctx.warnings.append(
        f"{context} {area} for {iso3} missing EA Code — using GO slot {slot}: {_emergency_op_row_id(op)!r}"
    )
    return op


def _resolve_ea_operation_display(
    ctx: UprImportContext,
    *,
    iso3: str,
    area: str,
    ea_code: Any,
    context: str = "Reach",
) -> Optional[str]:
    """Display label for a selectable emergency-appeal column header (name_with_code)."""
    op = _resolve_ea_operation(ctx, iso3=iso3, area=area, ea_code=ea_code, context=context)
    if not op:
        return None
    return _emergency_op_row_id(op) or None


def _ensure_funding_ea_col_header(
    matrix_cells: Dict[Tuple[int, int], Dict[str, Any]],
    ctx: UprImportContext,
    *,
    aes_id: int,
    funding_item_id: int,
    iso3: str,
    rnd: str,
    area: str,
    ea_code_raw: Any,
    reach_ea_codes: Dict[Tuple[str, str, str], str],
) -> bool:
    """Set ``col_header|EA*`` once per matrix item when importing planning funding."""
    cells = matrix_cells[(aes_id, funding_item_id)]
    header_key = f"col_header|{area}"
    if header_key in cells:
        return True
    ea_code = ea_code_raw
    if ea_code in (None, ""):
        ea_code = reach_ea_codes.get((iso3, rnd, area))
    display = _resolve_ea_operation_display(
        ctx, iso3=iso3, area=area, ea_code=ea_code, context="Funding"
    )
    if not display:
        return False
    cells[header_key] = display
    return True


def _resolve_emergency_row_key(
    ctx: UprImportContext,
    *,
    iso3: str,
    area: str,
    ea_code: Any,
) -> Optional[str]:
    """Resolve Emergency Appeals matrix row key from Excel EA slot and/or EA Code."""
    op = _resolve_ea_operation(ctx, iso3=iso3, area=area, ea_code=ea_code, context="Reach")
    if not op:
        return None
    row_id = _emergency_op_row_id(op)
    if not row_id:
        return None
    return f"{row_id}_{EMERGENCY_APPEALS_COLUMN}"


def _load_assignment_map(template_ids: List[int]) -> Dict[int, Dict[Tuple[str, str], int]]:
    from app.extensions import db
    from app.models.assignments import AssignedForm, AssignmentEntityStatus
    from app.models.core import Country

    by_template: Dict[int, Dict[Tuple[str, str], int]] = {tid: {} for tid in template_ids}
    for template_id, period_name, aes_id, iso3_raw in (
        db.session.query(
            AssignedForm.template_id,
            AssignedForm.period_name,
            AssignmentEntityStatus.id,
            Country.iso3,
        )
        .join(AssignmentEntityStatus, AssignmentEntityStatus.assigned_form_id == AssignedForm.id)
        .join(Country, Country.id == AssignmentEntityStatus.entity_id)
        .filter(
            AssignedForm.template_id.in_(template_ids),
            AssignmentEntityStatus.entity_type == "country",
        )
        .all()
    ):
        tid = int(template_id)
        pn = (period_name or "").strip()
        iso3 = (iso3_raw or "").strip().upper()
        if pn and iso3 and aes_id is not None and tid in by_template:
            by_template[tid][(pn, iso3)] = int(aes_id)
    return by_template


def _matrix_column_name_from_form_item(item) -> Optional[str]:
    """Return the first matrix column ``name`` from a FormItem config, if defined."""
    if not item or not isinstance(getattr(item, "config", None), dict):
        return None
    matrix_cfg = item.config.get("matrix_config") or item.config
    if not isinstance(matrix_cfg, dict):
        return None
    columns = matrix_cfg.get("columns") or []
    if not columns:
        return None
    first = columns[0]
    name = first.get("name") if isinstance(first, dict) else first
    return str(name) if name else None


def _published_funding_matrix_item(template_id: int):
    """Resolve the published-version funding matrix item for a reporting template."""
    from app.models.form_items import FormItem
    from app.models.forms import FormTemplate

    template = FormTemplate.query.get(int(template_id))
    pub_vid = getattr(template, "published_version_id", None) if template else None
    if not pub_vid:
        return None
    labels = REPORTING_SPECIAL_ITEM_LABELS.get("funding") or ()
    query = FormItem.query.filter(
        FormItem.template_id == int(template_id),
        FormItem.version_id == int(pub_vid),
        FormItem.item_type == "matrix",
        FormItem.archived == False,
    )
    for needle in labels:
        item = query.filter(FormItem.label.ilike(f"%{needle}%")).first()
        if item:
            return item
    return query.first()


def _matrix_column_name_from_item_id(item_id: int) -> str:
    from app.models.form_items import FormItem

    item = FormItem.query.get(item_id)
    name = _matrix_column_name_from_form_item(item)
    if name:
        return name

    template_id = getattr(item, "template_id", None) if item else None
    if template_id:
        pub_item = _published_funding_matrix_item(int(template_id))
        name = _matrix_column_name_from_form_item(pub_item)
        if name:
            return name

    return REPORTING_FUNDING_MATRIX_COLUMN


def _load_published_version_ids(template_ids: List[int]) -> Dict[int, int]:
    from app.models.forms import FormTemplate

    out: Dict[int, int] = {}
    for tid in template_ids:
        template = FormTemplate.query.get(int(tid))
        if template and template.published_version_id:
            out[int(tid)] = int(template.published_version_id)
    return out


def _load_published_item_indexes(
    template_ids: List[int],
    published_version_ids: Dict[int, int],
) -> Tuple[Dict[int, Dict[int, int]], Dict[int, Dict[int, Dict[str, int]]], Dict[int, Dict[str, int]]]:
    from app.models.form_items import FormItem
    from app.models.forms import FormSection

    version_ids = set(published_version_ids.values())
    by_bank: Dict[int, Dict[int, int]] = {tid: {} for tid in template_ids}
    by_bank_section: Dict[int, Dict[int, Dict[str, int]]] = {tid: {} for tid in template_ids}
    by_label: Dict[int, Dict[str, int]] = {tid: {} for tid in template_ids}
    if not version_ids:
        return by_bank, by_bank_section, by_label

    items = (
        FormItem.query.outerjoin(FormSection, FormItem.section_id == FormSection.id)
        .filter(
            FormItem.template_id.in_(template_ids),
            FormItem.version_id.in_(version_ids),
            FormItem.archived == False,
        )
        .all()
    )
    for item in items:
        if not item.template_id or not item.version_id:
            continue
        tid = int(item.template_id)
        expected_vid = published_version_ids.get(tid)
        if expected_vid is None or int(item.version_id) != int(expected_vid):
            continue
        item_id = int(item.id)
        label = (item.label or "").strip().lower()
        if label:
            by_label.setdefault(tid, {})[label] = item_id
        if item.indicator_bank_id:
            bank_id = int(item.indicator_bank_id)
            by_bank.setdefault(tid, {})[bank_id] = item_id
            section_name = (item.form_section.name if item.form_section else "").strip()
            if section_name:
                by_bank_section.setdefault(tid, {}).setdefault(bank_id, {})[section_name] = item_id
    return by_bank, by_bank_section, by_label


def _build_reporting_special_items(labels: Dict[str, int]) -> Dict[str, Any]:
    fallbacks = {
        "funding": ITEM_REPORTING_COUNTRY_FUNDING,
        "expenditure": ITEM_REPORTING_COUNTRY_EXPENDITURE,
        "sp_breakdown": ITEM_REPORTING_COUNTRY_SP_BREAKDOWN,
        "support": ITEM_REPORTING_COUNTRY_SUPPORT,
    }
    special: Dict[str, Any] = {}
    for key, needles in REPORTING_SPECIAL_ITEM_LABELS.items():
        item_id = find_item_by_label(labels, *needles) or fallbacks.get(key)
        if item_id:
            special[key] = int(item_id)
    funding_id = special.get("funding")
    if funding_id:
        special["funding_col"] = _matrix_column_name_from_item_id(int(funding_id))
    else:
        special["funding_col"] = REPORTING_FUNDING_MATRIX_COLUMN
    return special


def _load_other_indicators_section_id() -> Optional[int]:
    from app.models.forms import FormSection, FormTemplateVersion

    sections = (
        FormSection.query.join(FormTemplateVersion, FormSection.version_id == FormTemplateVersion.id)
        .filter(
            FormSection.template_id == REPORTING_COUNTRY_TEMPLATE_ID,
            FormTemplateVersion.status == "published",
        )
        .all()
    )
    for section in sections:
        name = (section.name or "").strip().lower()
        if name == "other indicators" and section.section_type == "dynamic_indicators":
            return int(section.id)
    return None


def _load_reporting_emergency_structure(template_id: int, version_id: int) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    """Return (ea_repeat_section_id, ea_dynamic_section_id, emergency_choice_item_id)."""
    from app.models.form_items import FormItem
    from app.models.forms import FormSection, FormTemplateVersion

    ea_repeat_id: Optional[int] = None
    ea_dynamic_id: Optional[int] = None
    sections = (
        FormSection.query.join(FormTemplateVersion, FormSection.version_id == FormTemplateVersion.id)
        .filter(
            FormSection.template_id == int(template_id),
            FormTemplateVersion.id == int(version_id),
        )
        .all()
    )
    for section in sections:
        name = (section.name or "").strip().lower()
        stype = (section.section_type or "").strip().lower()
        if name == "emergency appeals indicators" and stype == "repeat":
            ea_repeat_id = int(section.id)
        elif name == "emergency appeal indicators" and stype == "dynamic_indicators":
            ea_dynamic_id = int(section.id)
    if ea_repeat_id:
        for section in sections:
            if (
                section.parent_section_id == ea_repeat_id
                and (section.section_type or "").strip().lower() == "dynamic_indicators"
            ):
                ea_dynamic_id = int(section.id)
                break
    choice_item_id: Optional[int] = None
    if ea_repeat_id:
        item = (
            FormItem.query.filter_by(
                template_id=int(template_id),
                version_id=int(version_id),
                section_id=ea_repeat_id,
                archived=False,
            )
            .filter(FormItem.lookup_list_id == "emergency_operations")
            .first()
        )
        if item:
            choice_item_id = int(item.id)
    return ea_repeat_id, ea_dynamic_id, choice_item_id


def _parse_row_text_value(row: Dict[str, Any]) -> Optional[str]:
    for key in ("Value", "UPR Value"):
        raw = row.get(key)
        if raw is None:
            continue
        text = str(raw).strip()
        if text:
            return text
    return None


def _parse_ns_emergency_slot_field(indicator_name: str) -> Optional[Tuple[int, str]]:
    """Return (slot 1–3, 'name'|'code') for Data_EO* / Data_MDR* NS Data indicators."""
    key = (indicator_name or "").strip().lower().replace(" ", "")
    for prefix, field in ((NS_EMERGENCY_NAME_PREFIX, "name"), (NS_EMERGENCY_CODE_PREFIX, "code")):
        m = re.fullmatch(rf"{prefix}(\d)", key)
        if m:
            return int(m.group(1)), field
    return None


def _stage_emergency_slot_meta(
    ctx: UprImportContext,
    *,
    aes_id: int,
    slot: int,
    field: str,
    value: str,
) -> None:
    if slot not in (1, 2, 3) or field not in ("name", "code"):
        return
    meta = ctx.emergency_slot_meta.setdefault((int(aes_id), int(slot)), {"name": "", "code": ""})
    meta[field] = value.strip()


def _ensure_repeat_instance(
    aes_id: int,
    repeat_section_id: int,
    instance_number: int,
    label: Optional[str] = None,
    *,
    user_id: Optional[int] = None,
):
    from app.extensions import db
    from app.models.forms import RepeatGroupInstance

    with db.session.no_autoflush:
        inst = RepeatGroupInstance.query.filter_by(
            assignment_entity_status_id=aes_id,
            section_id=repeat_section_id,
            instance_number=instance_number,
        ).first()
    if not inst:
        effective_user_id = int(user_id) if user_id is not None else _default_user_id_for_import()
        inst = RepeatGroupInstance(
            assignment_entity_status_id=aes_id,
            section_id=repeat_section_id,
            instance_number=instance_number,
            instance_label=label or f"Emergency Appeal {instance_number}",
            created_by_user_id=effective_user_id,
        )
        db.session.add(inst)
        db.session.flush()
    elif label:
        inst.instance_label = label
    return inst


def _upsert_emergency_repeat_choice(
    *,
    repeat_instance,
    choice_item_id: int,
    appeal_name: str,
    mdr_code: str,
    display_value: Optional[str] = None,
) -> None:
    from app.extensions import db
    from app.models.forms import RepeatGroupData

    name = (appeal_name or "").strip()
    code = (mdr_code or "").strip()
    if not name and not code:
        return

    display = (display_value or "").strip() or _format_emergency_operation_display(name, code)
    with db.session.no_autoflush:
        entry = RepeatGroupData.query.filter_by(
            repeat_instance_id=repeat_instance.id,
            form_item_id=choice_item_id,
        ).first()
    if not entry:
        entry = RepeatGroupData(
            repeat_instance_id=repeat_instance.id,
            form_item_id=choice_item_id,
        )
        db.session.add(entry)

    entry.value = display
    entry.disagg_data = {"name": name, "code": code}
    entry.disagg_type = "emergency_operation"
    entry.data_not_available = False
    entry.not_applicable = False


def upsert_emergency_repeat_slots(
    ctx: UprImportContext,
    *,
    dry_run: bool = False,
) -> Dict[str, int]:
    """Create repeat-group instances and emergency operation choices from staged NS Data."""
    from app.extensions import db

    stats = {"emergency_slots_upserted": 0}
    repeat_section_id = ctx.ea_repeat_section_id
    choice_item_id = ctx.emergency_choice_item_id
    if dry_run or not repeat_section_id or not choice_item_id:
        return stats

    slots_needed: Set[Tuple[int, int]] = set(ctx.emergency_slot_meta.keys())
    ea_dynamic = ctx.ea_dynamic_section_id
    if ea_dynamic:
        for entry in ctx.dynamic_indicator_entries:
            repeat_num = entry.get("repeat_instance_number")
            if repeat_num is not None and int(entry.get("section_id") or 0) == int(ea_dynamic):
                slots_needed.add((int(entry["aes_id"]), int(repeat_num)))

    user_id = _default_user_id_for_import()
    for aes_id, slot_num in sorted(slots_needed):
        meta = ctx.emergency_slot_meta.get((aes_id, slot_num), {"name": "", "code": ""})
        excel_name = (meta.get("name") or "").strip()
        excel_code = (meta.get("code") or "").strip()
        if not excel_name and not excel_code:
            continue
        iso3 = _aes_id_to_iso3(ctx, aes_id)
        name, code, display = _resolve_emergency_operation_labels(
            ctx,
            iso3=iso3,
            excel_name=excel_name,
            excel_code=excel_code,
        )
        label = name or code or display
        inst = _ensure_repeat_instance(
            aes_id,
            int(repeat_section_id),
            int(slot_num),
            label,
            user_id=user_id,
        )
        if inst:
            _upsert_emergency_repeat_choice(
                repeat_instance=inst,
                choice_item_id=int(choice_item_id),
                appeal_name=name,
                mdr_code=code,
                display_value=display,
            )
            stats["emergency_slots_upserted"] += 1

    if stats["emergency_slots_upserted"]:
        db.session.commit()
    return stats


def reporting_special_item(
    ctx: UprImportContext,
    key: str,
    *,
    template_id: int = REPORTING_COUNTRY_TEMPLATE_ID,
) -> Optional[int]:
    special = ctx.reporting_special_items.get(int(template_id), {})
    item_id = special.get(key)
    return int(item_id) if item_id else None


def reporting_funding_matrix_column(
    ctx: UprImportContext,
    *,
    template_id: int = REPORTING_COUNTRY_TEMPLATE_ID,
) -> str:
    special = ctx.reporting_special_items.get(int(template_id), {})
    return str(special.get("funding_col") or REPORTING_FUNDING_MATRIX_COLUMN)


def _resolve_item_by_bank_and_area(
    ctx: UprImportContext,
    template_id: int,
    bank_id: int,
    area: str,
) -> Optional[int]:
    """Resolve a published-version form item when bank ids repeat across sections."""
    section_map = ctx.items_by_bank_section.get(template_id, {}).get(bank_id)
    if section_map:
        section_name = REPORTING_EXCEL_AREA_TO_SECTION.get(area)
        if section_name and section_name in section_map:
            return section_map[section_name]
        if len(section_map) == 1:
            return next(iter(section_map.values()))
    return ctx.items_by_bank_id.get(template_id, {}).get(bank_id)


def _queue_dynamic_indicator_entry(
    ctx: UprImportContext,
    *,
    section_id: Optional[int],
    aes_id: int,
    indicator_bank_id: int,
    value: Optional[Any],
    data_not_available: bool,
    order_counters: Dict[Tuple[Any, ...], float],
    order_key: Tuple[Any, ...],
    repeat_instance_number: Optional[int] = None,
    disagg_data: Optional[Dict[str, Any]] = None,
    missing_section_warning: str = "",
) -> None:
    if not section_id:
        if missing_section_warning:
            ctx.warnings.append(missing_section_warning)
        return
    if ctx.indicator_bank_ids and indicator_bank_id not in ctx.indicator_bank_ids:
        ctx.warnings.append(
            f"Indicator bank id {indicator_bank_id} not found; skipping dynamic import"
        )
        return
    order_counters[order_key] = order_counters.get(order_key, 0.0) + 1.0
    entry = {
        "aes_id": aes_id,
        "section_id": section_id,
        "indicator_bank_id": indicator_bank_id,
        "repeat_instance_number": repeat_instance_number,
        "value": value,
        "data_not_available": data_not_available,
        "order": order_counters[order_key],
        "disagg_data": disagg_data,
    }
    for idx, existing in enumerate(ctx.dynamic_indicator_entries):
        if (
            existing.get("aes_id") == aes_id
            and existing.get("indicator_bank_id") == indicator_bank_id
            and existing.get("section_id") == section_id
            and existing.get("repeat_instance_number") == repeat_instance_number
        ):
            ctx.dynamic_indicator_entries[idx] = entry
            return
    ctx.dynamic_indicator_entries.append(entry)


def _queue_other_dynamic_indicator(
    ctx: UprImportContext,
    *,
    aes_id: int,
    indicator_bank_id: int,
    value: Optional[Any],
    data_not_available: bool,
    order_counters: Dict[int, float],
    disagg_data: Optional[Dict[str, Any]] = None,
) -> None:
    _queue_dynamic_indicator_entry(
        ctx,
        section_id=ctx.other_indicators_section_id,
        aes_id=aes_id,
        indicator_bank_id=indicator_bank_id,
        value=value,
        data_not_available=data_not_available,
        order_counters=order_counters,
        order_key=(aes_id, "other"),
        repeat_instance_number=None,
        disagg_data=disagg_data,
        missing_section_warning=(
            f"Other indicators dynamic section missing; cannot import bank {indicator_bank_id} (aes {aes_id})"
        ),
    )


def _queue_emergency_dynamic_indicator(
    ctx: UprImportContext,
    *,
    aes_id: int,
    slot: int,
    indicator_bank_id: int,
    value: Optional[Any],
    data_not_available: bool,
    order_counters: Dict[Tuple[int, int], float],
    disagg_data: Optional[Dict[str, Any]] = None,
) -> None:
    _queue_dynamic_indicator_entry(
        ctx,
        section_id=ctx.ea_dynamic_section_id,
        aes_id=aes_id,
        indicator_bank_id=indicator_bank_id,
        value=value,
        data_not_available=data_not_available,
        order_counters=order_counters,
        order_key=(aes_id, slot),
        repeat_instance_number=int(slot),
        disagg_data=disagg_data,
        missing_section_warning=(
            f"Emergency appeal dynamic section missing; cannot import bank {indicator_bank_id} "
            f"(aes {aes_id}, slot {slot})"
        ),
    )


def _default_user_id_for_import() -> int:
    try:
        from flask_login import current_user

        if getattr(current_user, "is_authenticated", False) and getattr(current_user, "id", None):
            return int(current_user.id)
    except Exception:
        pass
    from app.models.core import User

    user = User.query.order_by(User.id.asc()).first()
    return int(user.id) if user else 1


def upsert_dynamic_indicator_entries(
    entries: List[Dict[str, Any]],
    *,
    dry_run: bool = False,
) -> Dict[str, int]:
    from app.extensions import db
    from app.models.forms import DynamicIndicatorData
    from sqlalchemy import tuple_

    stats = {"dynamic_inserted": 0, "dynamic_updated": 0, "dynamic_skipped": 0}
    if dry_run or not entries:
        return stats

    from app.models.indicator_bank import IndicatorBank

    valid_bank_ids = {int(row.id) for row in IndicatorBank.query.with_entities(IndicatorBank.id).all()}
    user_id = _default_user_id_for_import()

    pending: List[Dict[str, Any]] = []
    for entry in entries:
        bank_id = int(entry["indicator_bank_id"])
        if bank_id not in valid_bank_ids:
            stats["dynamic_skipped"] += 1
            continue
        pending.append(entry)

    keys: List[Tuple[int, int, int, Any]] = [
        (
            int(entry["aes_id"]),
            int(entry["section_id"]),
            int(entry["indicator_bank_id"]),
            entry.get("repeat_instance_number"),
        )
        for entry in pending
    ]
    existing_by_key: Dict[Tuple[int, int, int, Any], DynamicIndicatorData] = {}
    chunk_size = 500
    for i in range(0, len(keys), chunk_size):
        chunk = keys[i : i + chunk_size]
        if not chunk:
            continue
        for row in DynamicIndicatorData.query.filter(
            tuple_(
                DynamicIndicatorData.assignment_entity_status_id,
                DynamicIndicatorData.section_id,
                DynamicIndicatorData.indicator_bank_id,
                DynamicIndicatorData.repeat_instance_number,
            ).in_(chunk)
        ).all():
            existing_by_key[
                (
                    int(row.assignment_entity_status_id),
                    int(row.section_id),
                    int(row.indicator_bank_id),
                    row.repeat_instance_number,
                )
            ] = row

    for entry in pending:
        aes_id = int(entry["aes_id"])
        section_id = int(entry["section_id"])
        bank_id = int(entry["indicator_bank_id"])
        repeat_num = entry.get("repeat_instance_number")
        key = (aes_id, section_id, bank_id, repeat_num)
        row = existing_by_key.get(key)
        action = "dynamic_updated"
        if not row:
            row = DynamicIndicatorData(
                assignment_entity_status_id=aes_id,
                section_id=section_id,
                indicator_bank_id=bank_id,
                repeat_instance_number=repeat_num,
                added_by_user_id=user_id,
                order=float(entry.get("order") or 0),
            )
            db.session.add(row)
            existing_by_key[key] = row
            action = "dynamic_inserted"
        elif entry.get("order") is not None:
            row.order = float(entry["order"])

        if entry.get("data_not_available"):
            row.set_data_availability(data_not_available=True)
        elif entry.get("disagg_data"):
            payload = entry["disagg_data"]
            row.set_disaggregated_data(payload["mode"], payload["values"])
        else:
            value = entry.get("value")
            if value is not None:
                if isinstance(value, float) and value.is_integer():
                    row.set_simple_value(str(int(value)))
                else:
                    row.set_simple_value(str(value))

        stats[action] += 1

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise
    return stats


def build_import_context(template_ids: List[int]) -> UprImportContext:
    ids = [int(t) for t in template_ids]
    ctx = UprImportContext(template_ids=ids)
    ctx.indicator_bank_ids = _load_indicator_bank_ids()
    ctx.assignment_by_template = _load_assignment_map(ids)
    ctx.assignment_by_period_iso = ctx.assignment_by_template.get(24, {})
    ctx.published_version_ids = _load_published_version_ids(ids)
    ctx.items_by_bank_id, ctx.items_by_bank_section, ctx.item_ids_by_label = _load_published_item_indexes(
        ids, ctx.published_version_ids
    )
    if REPORTING_COUNTRY_TEMPLATE_ID in ids:
        ctx.reporting_special_items[REPORTING_COUNTRY_TEMPLATE_ID] = _build_reporting_special_items(
            ctx.item_ids_by_label.get(REPORTING_COUNTRY_TEMPLATE_ID, {})
        )
        ctx.other_indicators_section_id = _load_other_indicators_section_id()
        ctx.yes_no_bank_ids = _load_yes_no_bank_ids()
        ctx.percentage_bank_ids = _load_percentage_bank_ids()
        ctx.percentage_allow_over_100_bank_ids = _load_percentage_allow_over_100_bank_ids(
            ctx.percentage_bank_ids
        )
        reporting_aes_ids = set(
            ctx.assignment_by_template.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).values()
        )
        ctx.existing_disagg_static_keys, ctx.existing_disagg_dynamic_keys = (
            _load_existing_disaggregated_keys(reporting_aes_ids)
        )
        pub_vid = ctx.published_version_ids.get(REPORTING_COUNTRY_TEMPLATE_ID)
        if pub_vid:
            ctx.core_yes_no_item_ids = _load_core_yes_no_item_ids(
                REPORTING_COUNTRY_TEMPLATE_ID,
                pub_vid,
                other_section_id=ctx.other_indicators_section_id,
                yes_no_bank_ids=ctx.yes_no_bank_ids,
            )
            (
                ctx.ea_repeat_section_id,
                ctx.ea_dynamic_section_id,
                ctx.emergency_choice_item_id,
            ) = _load_reporting_emergency_structure(REPORTING_COUNTRY_TEMPLATE_ID, pub_vid)
            if ctx.emergency_choice_item_id:
                ctx.emergency_go_plugin_config = _load_emergency_choice_plugin_config(
                    ctx.emergency_choice_item_id
                )
    if 22 in ids:
        staff_id = find_item_by_label(ctx.item_ids_by_label.get(22, {}), *T22_STAFF_MATRIX_LABELS)
        if staff_id:
            ctx.staff_matrix_item_id = int(staff_id)
    if 23 in ids:
        pns_id = find_item_by_label(ctx.item_ids_by_label.get(23, {}), *T23_PNS_FUNDING_LABELS)
        if pns_id:
            ctx.pns_funding_item_id = int(pns_id)
    ctx.ns_name_to_id = _build_ns_name_index()
    ctx.ns_home_country_iso3, ctx.country_id_by_iso3, ctx.iso3_to_hns_id = _build_ns_home_country_index()
    if 24 in ids:
        ctx.emergency_matrix_plugin_config = _load_emergency_matrix_plugin_config(24)
    return ctx


def _resolve_aes(ctx: UprImportContext, period: str, iso3: str) -> Optional[int]:
    return ctx.assignment_by_period_iso.get((period, iso3.upper()))


def _resolve_ns_row_id(ctx: UprImportContext, ns_name: str) -> Optional[int]:
    key = (ns_name or "").strip().lower()
    if not key:
        return None
    ns_id = ctx.ns_name_to_id.get(key)
    if ns_id is None:
        ctx.warnings.append(f"National Society not found: {ns_name!r}")
    return ns_id


def _matrix_row(
    *,
    aes_id: int,
    item_id: int,
    cells: Dict[str, Any],
    iso3: str,
    period: str,
    debug_kpi: str,
) -> Dict[str, str]:
    return {
        "_debug_iso3": iso3,
        "_debug_year": period,
        "_debug_kpi_code": debug_kpi,
        COL_ASSIGNMENT: str(aes_id),
        COL_PUBLIC: "",
        COL_ITEM: str(item_id),
        COL_VALUE: "",
        COL_DISAGG: json.dumps(cells),
        COL_DATA_NA: "",
        COL_NA: "",
        COL_PREFILLED: "",
        COL_IMPUTED: "",
        COL_SUBMITTED: "",
    }


def _format_scalar_value(value: Any) -> str:
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    if num.is_integer():
        return str(int(num))
    return str(value)


def _scalar_row(
    *,
    aes_id: int,
    item_id: int,
    value: Any,
    iso3: str,
    period: str,
    debug_kpi: str,
) -> Dict[str, str]:
    if value is None:
        return {}
    return {
        "_debug_iso3": iso3,
        "_debug_year": period,
        "_debug_kpi_code": debug_kpi,
        COL_ASSIGNMENT: str(aes_id),
        COL_PUBLIC: "",
        COL_ITEM: str(item_id),
        COL_VALUE: _format_scalar_value(value),
        COL_DISAGG: "",
        COL_DATA_NA: "",
        COL_NA: "",
        COL_PREFILLED: "",
        COL_IMPUTED: "",
        COL_SUBMITTED: "",
    }


def _data_na_row(
    *,
    aes_id: int,
    item_id: int,
    iso3: str,
    period: str,
    debug_kpi: str,
) -> Dict[str, str]:
    """Build a form_data row that sets is_data_not_available = True (no value written)."""
    return {
        "_debug_iso3": iso3,
        "_debug_year": period,
        "_debug_kpi_code": debug_kpi,
        COL_ASSIGNMENT: str(aes_id),
        COL_PUBLIC: "",
        COL_ITEM: str(item_id),
        COL_VALUE: "",
        COL_DISAGG: "",
        COL_DATA_NA: "1",
        COL_NA: "",
        COL_PREFILLED: "",
        COL_IMPUTED: "",
        COL_SUBMITTED: "",
    }


def _year_offset(base_period: str, year_val: Any) -> Optional[int]:
    try:
        base = int(str(base_period).strip())
        year = int(float(str(year_val).strip()))  # float-safe: handles "2026.0" from Excel
    except (ValueError, TypeError):
        return None
    offset = year - base
    if offset not in (0, 1, 2):
        return None
    return offset


def _filter_rows(
    rows: List[Dict[str, Any]],
    *,
    template_ids: List[int],
    rounds: Optional[Set[str]] = None,
) -> List[Dict[str, Any]]:
    allowed_sections: Set[str] = set()
    allowed_prefixes: Set[str] = set()
    for tid in template_ids:
        prof = UPR_TEMPLATE_PROFILES.get(tid)
        if not prof:
            continue
        allowed_sections.update(prof.get("sections", frozenset()))
        allowed_prefixes.update(prof.get("round_prefixes", ()))

    out = []
    for row in rows:
        sec = str(row.get("Section") or "").strip()
        rnd = str(row.get("Round") or "").strip().upper()
        if sec not in allowed_sections:
            continue
        if allowed_prefixes and not any(rnd.startswith(p) for p in allowed_prefixes):
            continue
        if rounds and rnd not in rounds:
            continue
        if is_aggregate_row(row):
            continue
        out.append(row)
    return out


def _t22_pns_import_cell_value(
    country_val: Optional[float],
    pns_val: Optional[float],
) -> Optional[Any]:
    """Persist T22 funding when ``PNS reported = Yes``.

    - ``pns_val`` present → use PNS self-report (scalar)
    - ``pns_val`` blank but ``country_val`` present → PNS cleared the lookup
    - both blank → no cell
    """
    if pns_val is not None:
        return pns_val
    if country_val is not None:
        return {
            "original": country_val,
            "modified": "",
            "isModified": True,
        }
    return None


def _t22_total_only_breakdown_cell(
    pns_t22_staging: Dict[Tuple[int, int, str], Tuple[Optional[float], Optional[float]]],
    pns_aes: int,
    host_cid: int,
    area: str,
) -> Optional[Any]:
    """PNS reported a row total only — blank breakdown where country reported a value."""
    cv, pv = pns_t22_staging.get((pns_aes, host_cid, area), (None, None))
    if cv is None and pv is None:
        return None
    return _t22_pns_import_cell_value(cv, pv)


def transform_to_import_rows(
    rows: List[Dict[str, Any]],
    ctx: UprImportContext,
    *,
    template_ids: Optional[List[int]] = None,
    rounds: Optional[Set[str]] = None,
) -> List[Dict[str, str]]:
    """Transform UPR Excel rows into ready-to-import form_data rows."""
    tids = template_ids or ctx.template_ids
    filtered = _filter_rows(rows, template_ids=tids, rounds=rounds)
    reach_ea_codes = _build_reach_ea_code_index(rows, rounds=rounds)

    pns_t22_reported_yes, pns_t23_reported_yes = _build_pns_reported_yes_sets(
        rows, ctx, tids, rounds=rounds
    )
    iso3_by_host_cid = {cid: iso for iso, cid in ctx.country_id_by_iso3.items()}

    matrix_cells: Dict[Tuple[int, int], Dict[str, Any]] = defaultdict(dict)
    import_rows: List[Dict[str, str]] = []
    dynamic_order: Dict[int, float] = {}
    emergency_dynamic_order: Dict[Tuple[int, int], float] = {}
    filled_core_yes_no: Set[Tuple[int, int]] = set()
    core_yes_no_ids = set(ctx.core_yes_no_item_ids)

    # ── Planning T22 PNS funding staging ──────────────────────────────────────
    # Collected across all rows then converted to scalar matrix cells.
    # Keyed by (pns_aes_id, host_country_id, area) → (country_val, pns_val).
    pns_t22_staging: Dict[Tuple[int, int, str], Tuple[Optional[float], Optional[float]]] = {}
    # Area=Total rows (PNS reported aggregate only) → row-total column on item 1303.
    pns_t22_total_staging: Dict[Tuple[int, int], Tuple[Optional[float], Optional[float]]] = {}

    # ── Reporting country funding staging ─────────────────────────────────────
    # Entity=IFRC/PNS/Other rows (Attribute=Funding Source only, Indicator=Funding) accumulated
    # per aes_id into item 1403 rows.  Keyed by (aes_id, row_name) → accumulated total.
    reporting_funding_staging: Dict[Tuple[int, int, str], float] = defaultdict(float)

    for row in filtered:
        iso3 = str(row.get("ISO3") or "").strip().upper()
        rnd = str(row.get("Round") or "").strip().upper()
        sec = str(row.get("Section") or "").strip()
        period = round_to_period(rnd)
        if not iso3 or not period:
            continue

        # Classify round type so planning and reporting handlers don't cross-fire when
        # both planning (T24/T22) and reporting (T33/T23) templates are imported together.
        rnd_is_planning = rnd.startswith("P")
        rnd_is_reporting = rnd.startswith("AR") or rnd.startswith("MYR")

        value_num = parse_value_num(row.get("ValueNum"))
        indicator = str(row.get("Indicator") or "").strip()
        indicator_id = normalize_indicator_id(row.get("indicatorId"))
        entity = str(row.get("Entity") or "").strip()
        ns_name = str(row.get("NS") or "").strip()
        area = str(row.get("Area") or "").strip()
        year_val = row.get("Year")

        # ════════════════════════════════════════════════════════════════════════
        # PLANNING HANDLERS  (rounds P*)
        # ════════════════════════════════════════════════════════════════════════

        # --- Template 24: NS Data ---
        if 24 in tids and sec == "NS Data" and rnd_is_planning:
            if not indicator_id:
                continue
            aes_id = _resolve_aes(ctx, period, iso3)
            item_id = ctx.items_by_bank_id.get(24, {}).get(indicator_id)
            if not aes_id or not item_id or value_num is None:
                continue
            built = _scalar_row(
                aes_id=aes_id,
                item_id=item_id,
                value=value_num,
                iso3=iso3,
                period=period,
                debug_kpi=indicator or f"bank_{indicator_id}",
            )
            if built:
                import_rows.append(built)
            continue

        # --- Template 24: Comments ---
        if 24 in tids and sec == "Comments" and rnd_is_planning:
            aes_id = _resolve_aes(ctx, period, iso3)
            if not aes_id:
                continue
            text_val = parse_comment_value(row)
            if not text_val:
                continue
            ctx.discussion_comment_entries.append(
                {
                    "aes_id": aes_id,
                    "body": f"{humanize_comment_label(indicator)}: {text_val}",
                    "source": "upr_excel_import",
                }
            )
            continue

        # --- Template 24: Support (bilateral ticks) ---
        if 24 in tids and sec == "Support" and rnd_is_planning:
            if indicator_id != 3 and indicator.lower() != "bilateral support":
                continue
            if value_num != 1:
                continue
            if not area or area in AGGREGATE_AREA:
                continue
            aes_id = _resolve_aes(ctx, period, iso3)
            ns_id = _resolve_ns_row_id(ctx, ns_name)
            if not aes_id or ns_id is None:
                continue
            matrix_cells[(aes_id, ITEM_BILATERAL_SUPPORT)][f"{ns_id}_{area}"] = 1
            continue

        # --- Funding (T24: HNS/IFRC/PNS Country-Value; T22: PNS-reported) ---
        if sec == "Funding" and rnd_is_planning:
            if not is_planning_funding_requirement_row(row):
                continue
            if year_val in (None, ""):
                ctx.warnings.append(f"Funding row missing Year for {iso3} {rnd}")
                continue
            offset = _year_offset(period, year_val)
            if offset is None:
                continue

            ent_upper = entity.upper()
            country_val = parse_value_num(row.get("Country Value"))
            pns_val = parse_value_num(row.get("PNS Value"))

            # Aggregate Excel areas (Total) → T22 row-total column when PNS has no SP/EF breakdown.
            if (not area or area in AGGREGATE_AREA):
                if (
                    area == "Total"
                    and offset == 0
                    and 22 in tids
                    and ent_upper == "PNS"
                    and parse_pns_reported_yes(row)
                    and (country_val or pns_val)
                ):
                    pns_iso3 = ctx.ns_home_country_iso3.get(ns_name.lower())
                    if not pns_iso3:
                        if pns_val:
                            ctx.warnings.append(f"Cannot resolve home country for NS: {ns_name!r}")
                    else:
                        pns_aes = ctx.assignment_by_template.get(22, {}).get((period, pns_iso3))
                        if not pns_aes:
                            if pns_val:
                                ctx.warnings.append(
                                    f"No template 22 assignment for {pns_iso3} {period} (NS: {ns_name!r})"
                                )
                        else:
                            host_cid = ctx.country_id_by_iso3.get(iso3)
                            if not host_cid:
                                if pns_val:
                                    ctx.warnings.append(f"Cannot resolve Country.id for ISO3: {iso3!r}")
                            else:
                                prev_cv, prev_pv = pns_t22_total_staging.get((pns_aes, host_cid), (None, None))
                                pns_t22_total_staging[(pns_aes, host_cid)] = (
                                    country_val if country_val is not None else prev_cv,
                                    pns_val if pns_val is not None else prev_pv,
                                )
                continue

            if not area:
                continue

            if is_skipped_legacy_funding_area(area):
                continue

            # ── HNS / IFRC Secretariat → country-reported → template 24 ──
            if 24 in tids and ent_upper in ("HNS", "IFRC SECRETARIAT"):
                hns_country_val = country_val
                if hns_country_val is None:
                    hns_country_val = value_num  # older export fallback
                if not hns_country_val:
                    continue
                aes_id = ctx.assignment_by_template.get(24, {}).get((period, iso3))
                if not aes_id:
                    continue
                funding_item_id = FUNDING_MATRIX_BY_YEAR_OFFSET.get(offset)
                if not funding_item_id:
                    continue
                if area in PLANNING_EA_FUNDING_AREAS:
                    if not _ensure_funding_ea_col_header(
                        matrix_cells,
                        ctx,
                        aes_id=aes_id,
                        funding_item_id=funding_item_id,
                        iso3=iso3,
                        rnd=rnd,
                        area=area,
                        ea_code_raw=row.get("EA Code"),
                        reach_ea_codes=reach_ea_codes,
                    ):
                        continue
                row_key = "HNS" if ent_upper == "HNS" else "IFRC Secretariat"
                matrix_cells[(aes_id, funding_item_id)][f"{row_key}_{area}"] = hns_country_val
                continue

            # ── PNS — Country Value and PNS Value processed independently ──
            if ent_upper != "PNS":
                continue

            # Country Value → template 24 hybrid funding matrix (same item as HNS/IFRC above)
            if country_val and 24 in tids:
                t24_aes = ctx.assignment_by_template.get(24, {}).get((period, iso3))
                funding_item_id = FUNDING_MATRIX_BY_YEAR_OFFSET.get(offset)
                if t24_aes and funding_item_id:
                    ns_id = _resolve_ns_row_id(ctx, ns_name)
                    if ns_id is not None:
                        if area in PLANNING_EA_FUNDING_AREAS:
                            if not _ensure_funding_ea_col_header(
                                matrix_cells,
                                ctx,
                                aes_id=t24_aes,
                                funding_item_id=funding_item_id,
                                iso3=iso3,
                                rnd=rnd,
                                area=area,
                                ea_code_raw=row.get("EA Code"),
                                reach_ea_codes=reach_ea_codes,
                            ):
                                continue
                        matrix_cells[(t24_aes, funding_item_id)][f"{ns_id}_{area}"] = country_val

            # T22 item 1303 only has SP1–SP5/EFs columns (no EA* selectable headers).
            if area in PLANNING_EA_FUNDING_AREAS:
                continue

            # T22 item 1303 is the current planning year only (offset 0).
            if 22 in tids and offset == 0 and parse_pns_reported_yes(row) and (country_val or pns_val):
                pns_iso3 = ctx.ns_home_country_iso3.get(ns_name.lower())
                if not pns_iso3:
                    if pns_val:
                        ctx.warnings.append(f"Cannot resolve home country for NS: {ns_name!r}")
                else:
                    pns_aes = ctx.assignment_by_template.get(22, {}).get((period, pns_iso3))
                    if not pns_aes:
                        if pns_val:
                            ctx.warnings.append(
                                f"No template 22 assignment for {pns_iso3} {period} (NS: {ns_name!r})"
                            )
                    else:
                        host_cid = ctx.country_id_by_iso3.get(iso3)
                        if not host_cid:
                            if pns_val:
                                ctx.warnings.append(f"Cannot resolve Country.id for ISO3: {iso3!r}")
                        else:
                            key = (pns_aes, host_cid, area)
                            prev_cv, prev_pv = pns_t22_staging.get(key, (None, None))
                            pns_t22_staging[key] = (
                                country_val if country_val is not None else prev_cv,
                                pns_val if pns_val is not None else prev_pv,
                            )
            continue

        # --- Template 24: Reach ---
        if 24 in tids and sec == "Reach" and rnd_is_planning:
            if indicator.lower() != "people to be reached":
                continue
            if not area or area in AGGREGATE_AREA:
                continue
            aes_id = _resolve_aes(ctx, period, iso3)
            if not aes_id or not value_num:
                continue
            if area.startswith("EA") and area != "EAs":
                cell_key = _resolve_emergency_row_key(
                    ctx,
                    iso3=iso3,
                    area=area,
                    ea_code=row.get("EA Code"),
                )
                if not cell_key:
                    continue
                matrix_cells[(aes_id, ITEM_EMERGENCY_APPEALS)][cell_key] = value_num
            elif area.startswith("SP"):
                if year_val in (None, ""):
                    ctx.warnings.append(f"Reach row missing Year for {iso3} {rnd} {area}")
                    continue
                row_year = str(int(float(str(year_val)))) if str(year_val).strip() else ""
                matrix_cells[(aes_id, ITEM_LONGER_TERM_PROGRAMMES)][f"{row_year}_{area}"] = value_num
            continue

        # --- Template 22: Staff ---
        # AES is keyed by the PNS home country (not host ISO3); row key is the host NS id.
        if 22 in tids and sec == "Staff" and rnd_is_planning:
            col_name = STAFF_INDICATOR_COLUMNS.get(indicator)
            if not col_name or entity.upper() != "PNS" or not value_num:
                continue
            pns_iso3 = ctx.ns_home_country_iso3.get(ns_name.lower())
            if not pns_iso3:
                ctx.warnings.append(f"Cannot resolve home country for NS: {ns_name!r}")
                continue
            pns_aes = ctx.assignment_by_template.get(22, {}).get((period, pns_iso3))
            if not pns_aes:
                ctx.warnings.append(
                    f"No template 22 assignment for {pns_iso3} {period} (NS: {ns_name!r})"
                )
                continue
            if (pns_aes, iso3) not in pns_t22_reported_yes:
                continue
            host_ns_id = ctx.iso3_to_hns_id.get(iso3)
            if not host_ns_id:
                ctx.warnings.append(f"No active NS found for host country: {iso3!r}")
                continue
            matrix_cells[(pns_aes, ctx.staff_matrix_item_id)][f"{host_ns_id}_{col_name}"] = value_num
            continue

        # ════════════════════════════════════════════════════════════════════════
        # REPORTING HANDLERS  (rounds AR*, MYR*)
        # ════════════════════════════════════════════════════════════════════════

        # --- Template 33: NS Data ---
        # KPI scalars (723/724/727/1117) plus emergency slot metadata (Data_EO*/Data_MDR*).
        if REPORTING_COUNTRY_TEMPLATE_ID in tids and sec == "NS Data" and rnd_is_reporting:
            aes_id = ctx.assignment_by_template.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get((period, iso3))
            if not aes_id:
                continue
            slot_field = _parse_ns_emergency_slot_field(indicator)
            if slot_field:
                slot, field = slot_field
                text = _parse_row_text_value(row)
                if text:
                    _stage_emergency_slot_meta(ctx, aes_id=aes_id, slot=slot, field=field, value=text)
                continue
            if not indicator_id or value_num is None:
                continue
            item_id = ctx.items_by_bank_id.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get(indicator_id)
            if not item_id:
                continue
            built = _scalar_row(
                aes_id=aes_id,
                item_id=item_id,
                value=value_num,
                iso3=iso3,
                period=period,
                debug_kpi=indicator or f"bank_{indicator_id}",
            )
            if built:
                import_rows.append(built)
            continue

        # --- Template 33: Emergency 1/2/3 (repeat-group dynamic indicators) ---
        if (
            REPORTING_COUNTRY_TEMPLATE_ID in tids
            and sec in REPORTING_EMERGENCY_EXCEL_SECTION_TO_SLOT
            and rnd_is_reporting
        ):
            if not indicator_id:
                continue
            if not area or area in AGGREGATE_AREA:
                continue
            slot = REPORTING_EMERGENCY_EXCEL_SECTION_TO_SLOT[sec]
            aes_id = ctx.assignment_by_template.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get((period, iso3))
            if not aes_id:
                continue
            section_b = str(row.get("SectionB") or "").strip()
            if section_b:
                code = section_b.upper()
                meta = ctx.emergency_slot_meta.setdefault((aes_id, slot), {"name": "", "code": ""})
                if not meta.get("code"):
                    meta["code"] = code
                elif meta["code"].upper() != code:
                    ctx.warnings.append(
                        f"Emergency slot {slot} SectionB {code!r} differs from NS Data MDR {meta['code']!r} ({iso3} {rnd})"
                    )
            applicable_raw = str(row.get("Applicable/Data not available") or "").strip().lower()
            is_dna = "data not available" in applicable_raw
            has_value = _reporting_indicator_has_import_value(
                ctx, indicator_id, value_num, is_dna=is_dna
            )
            overwrite_warning = _disaggregation_overwrite_warning(
                ctx,
                value_num=value_num,
                is_dna=is_dna,
                dynamic_key=(aes_id, ctx.ea_dynamic_section_id, indicator_id, slot),
                iso3=iso3, rnd=rnd, label=indicator,
            )
            if overwrite_warning:
                ctx.warnings.append(overwrite_warning)
            elif has_value:
                _queue_emergency_dynamic_indicator(
                    ctx,
                    aes_id=aes_id,
                    slot=slot,
                    indicator_bank_id=indicator_id,
                    value=None if is_dna else _reporting_indicator_import_value(
                        ctx, indicator_id, value_num, iso3=iso3, rnd=rnd, label=indicator
                    ),
                    data_not_available=is_dna,
                    order_counters=emergency_dynamic_order,
                )
            continue

        # --- Template 33: Core indicators + Other indicators ---
        # Core: static form items on the published template when the indicator still exists
        # in that section; otherwise sync to the Other indicators dynamic section.
        # Other indicators Excel rows always go to the dynamic section.
        if REPORTING_COUNTRY_TEMPLATE_ID in tids and sec in ("Core indicators", "Other indicators") and rnd_is_reporting:
            if not indicator_id:
                continue
            if not area or area in AGGREGATE_AREA:
                continue
            aes_id = ctx.assignment_by_template.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get((period, iso3))
            if not aes_id:
                continue
            applicable_raw = str(row.get("Applicable/Data not available") or "").strip().lower()
            is_dna = "data not available" in applicable_raw
            has_value = _reporting_indicator_has_import_value(
                ctx, indicator_id, value_num, is_dna=is_dna
            )

            if sec == "Other indicators":
                other_overwrite_warning = _disaggregation_overwrite_warning(
                    ctx,
                    value_num=value_num,
                    is_dna=is_dna,
                    dynamic_key=(aes_id, ctx.other_indicators_section_id, indicator_id, None),
                    iso3=iso3, rnd=rnd, label=indicator,
                )
                if other_overwrite_warning:
                    ctx.warnings.append(other_overwrite_warning)
                elif has_value:
                    _queue_other_dynamic_indicator(
                        ctx,
                        aes_id=aes_id,
                        indicator_bank_id=indicator_id,
                        value=None if is_dna else _reporting_indicator_import_value(
                            ctx, indicator_id, value_num, iso3=iso3, rnd=rnd, label=indicator
                        ),
                        data_not_available=is_dna,
                        order_counters=dynamic_order,
                    )
                continue

            item_id = _resolve_item_by_bank_and_area(ctx, REPORTING_COUNTRY_TEMPLATE_ID, indicator_id, area)
            core_overwrite_warning = _disaggregation_overwrite_warning(
                ctx,
                value_num=value_num,
                is_dna=is_dna,
                static_key=(aes_id, item_id) if item_id else None,
                dynamic_key=None if item_id else (aes_id, ctx.other_indicators_section_id, indicator_id, None),
                iso3=iso3, rnd=rnd, label=indicator,
            )
            if core_overwrite_warning:
                ctx.warnings.append(core_overwrite_warning)
            elif item_id:
                if is_dna:
                    import_rows.append(
                        _data_na_row(
                            aes_id=aes_id,
                            item_id=item_id,
                            iso3=iso3,
                            period=period,
                            debug_kpi=f"bank_{indicator_id}",
                        )
                    )
                    if item_id in core_yes_no_ids:
                        filled_core_yes_no.add((aes_id, item_id))
                elif has_value:
                    built = _scalar_row(
                        aes_id=aes_id,
                        item_id=item_id,
                        value=_reporting_indicator_import_value(
                            ctx, indicator_id, value_num, iso3=iso3, rnd=rnd, label=indicator
                        ),
                        iso3=iso3,
                        period=period,
                        debug_kpi=indicator or f"bank_{indicator_id}",
                    )
                    if built:
                        import_rows.append(built)
                        if item_id in core_yes_no_ids:
                            filled_core_yes_no.add((aes_id, item_id))
            elif has_value:
                _queue_other_dynamic_indicator(
                    ctx,
                    aes_id=aes_id,
                    indicator_bank_id=indicator_id,
                    value=None if is_dna else _reporting_indicator_import_value(
                        ctx, indicator_id, value_num, iso3=iso3, rnd=rnd, label=indicator
                    ),
                    data_not_available=is_dna,
                    order_counters=dynamic_order,
                )
            continue

        # --- Reporting Funding (T33 items 1403 + 1404 + 1405; T23 item 952) ---
        if sec == "Funding" and rnd_is_reporting:
            if value_num is None:
                continue
            ent_upper = entity.upper()
            attr = str(row.get("Attribute") or "").strip()
            attr_lower = attr.lower()

            # ── T33: SP Breakdown (Attribute=SP Breakdown, Entity=HNS) → item 1405 ──
            # Area (SP1–SP5, EFs) → matrix row; Funding/Expenditure → column.
            if (
                REPORTING_COUNTRY_TEMPLATE_ID in tids
                and ent_upper == "HNS"
                and attr_lower == "sp breakdown"
                and area
                and area not in AGGREGATE_AREA
            ):
                row_name = REPORTING_SP_BREAKDOWN_AREA_TO_ROW.get(area)
                col_name = REPORTING_SP_BREAKDOWN_COLUMNS.get(indicator_id)
                if not col_name and indicator.lower() == "funding":
                    col_name = REPORTING_SP_BREAKDOWN_COLUMNS[733]
                elif not col_name and indicator.lower() == "expenditure":
                    col_name = REPORTING_SP_BREAKDOWN_COLUMNS[734]
                if not row_name:
                    ctx.warnings.append(
                        f"Unknown SP/EF area for reporting SP breakdown: {area!r} ({iso3} {rnd})"
                    )
                elif col_name and value_num:
                    aes_id = ctx.assignment_by_template.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get((period, iso3))
                    sp_item = reporting_special_item(ctx, "sp_breakdown")
                    if aes_id and sp_item:
                        cell_key = f"{row_name}_{col_name}"
                        matrix_cells[(aes_id, sp_item)][cell_key] = value_num
                continue

            # ── T33: HNS Expenditure total (Attribute=Total, bank=734) → scalar item 1404 ──
            if (
                REPORTING_COUNTRY_TEMPLATE_ID in tids
                and ent_upper == "HNS"
                and attr_lower == "total"
                and indicator_id == 734
            ):
                aes_id = ctx.assignment_by_template.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get((period, iso3))
                item_id = ctx.items_by_bank_id.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get(734) or reporting_special_item(
                    ctx, "expenditure"
                )
                if aes_id and item_id:
                    built = _scalar_row(
                        aes_id=aes_id,
                        item_id=item_id,
                        value=value_num,
                        iso3=iso3,
                        period=period,
                        debug_kpi="Reporting_Expenditure",
                    )
                    if built:
                        import_rows.append(built)
                continue

            # ── T33: Funding by source (Attribute=Funding Source only, Indicator=Funding) → item 1403 ──
            is_funding_source_row = (
                attr_lower == "funding source"
                and indicator_id == 733
            )

            if REPORTING_COUNTRY_TEMPLATE_ID in tids and is_funding_source_row:
                aes_id = ctx.assignment_by_template.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get((period, iso3))
                if aes_id:
                    funding_item = reporting_special_item(ctx, "funding")
                    if funding_item:
                        if ent_upper == "IFRC SECRETARIAT":
                            reporting_funding_staging[(aes_id, funding_item, REPORTING_FUNDING_ROW_IFRC)] += value_num
                        elif ent_upper == "PNS" and ns_name and ns_name.lower() not in ("country", ""):
                            reporting_funding_staging[(aes_id, funding_item, REPORTING_FUNDING_ROW_PNS)] += value_num
                        elif ent_upper == "OTHER SOURCES":
                            reporting_funding_staging[(aes_id, funding_item, REPORTING_FUNDING_ROW_OTHER)] += value_num

            # ── T23: PNS-reported Funding / Expenditure / Transferred → item 952 ──
            # AES is keyed by PNS home country ISO3; row is the host country's NS id.
            if 23 in tids and ent_upper == "PNS":
                col_name = T23_PNS_FUNDING_COLUMNS.get(indicator_id)
                if col_name and ns_name and ns_name.lower() not in ("country", ""):
                    pns_iso3 = ctx.ns_home_country_iso3.get(ns_name.lower())
                    if not pns_iso3:
                        ctx.warnings.append(f"Cannot resolve home country for NS: {ns_name!r}")
                    else:
                        pns_aes = ctx.assignment_by_template.get(23, {}).get((period, pns_iso3))
                        if pns_aes:
                            if (pns_aes, iso3) not in pns_t23_reported_yes:
                                continue
                            host_ns_id = ctx.iso3_to_hns_id.get(iso3)
                            if host_ns_id:
                                cell_key = f"{host_ns_id}_{col_name}"
                                matrix_cells[(pns_aes, ctx.pns_funding_item_id)][cell_key] = value_num
                            else:
                                ctx.warnings.append(f"No active NS found for host country: {iso3!r}")
            continue

        # --- Template 33: Support (Received Support) ---
        # Records which PNSs provided support per SP/EF area in the reporting year.
        # Writes to the "{area} Supported" column of item 1407 (list_library national_society).
        if REPORTING_COUNTRY_TEMPLATE_ID in tids and sec == "Support" and rnd_is_reporting:
            if entity.upper() != "PNS" or not area or area in AGGREGATE_AREA:
                continue
            if value_num != 1:
                continue
            aes_id = ctx.assignment_by_template.get(REPORTING_COUNTRY_TEMPLATE_ID, {}).get((period, iso3))
            ns_id = _resolve_ns_row_id(ctx, ns_name)
            if not aes_id or ns_id is None:
                continue
            support_item = reporting_special_item(ctx, "support")
            if not support_item:
                continue
            cell_key = f"{ns_id}_{area} Supported"
            matrix_cells[(aes_id, support_item)][cell_key] = 1
            continue

    # ── Post-loop: T22 PNS funding → submitted cells (PNS Value only when reported) ──
    for (pns_aes, host_cid, area), (cv, pv) in pns_t22_staging.items():
        host_iso3 = iso3_by_host_cid.get(host_cid)
        if not host_iso3 or (pns_aes, host_iso3) not in pns_t22_reported_yes:
            continue
        cell = _t22_pns_import_cell_value(cv, pv)
        if cell is not None:
            matrix_cells[(pns_aes, ITEM_FUNDING_REQUIREMENTS_T22)][f"{host_cid}_{area}"] = cell

    # ── Post-loop: T22 PNS Total (Excel Area=Total) → row-total column ──
    for (pns_aes, host_cid), (cv, pv) in pns_t22_total_staging.items():
        host_iso3 = iso3_by_host_cid.get(host_cid)
        pns_reported = bool(host_iso3 and (pns_aes, host_iso3) in pns_t22_reported_yes)
        if not pns_reported:
            continue
        cell = _t22_pns_import_cell_value(cv, pv)
        if cell is not None:
            item_cells = matrix_cells[(pns_aes, ITEM_FUNDING_REQUIREMENTS_T22)]
            item_cells[f"{host_cid}_{T22_ROW_TOTAL_COLUMN}"] = cell
        # PNS reported row totals only (no SP/EF breakdown) — blank breakdown cells
        # where country reported amounts but PNS did not provide SP/EF values.
        has_pns_breakdown = any(
            k[0] == pns_aes
            and k[1] == host_cid
            and pns_t22_staging[k][1] is not None
            for k in pns_t22_staging
        )
        if not has_pns_breakdown:
            item_cells = matrix_cells[(pns_aes, ITEM_FUNDING_REQUIREMENTS_T22)]
            for area in T22_BREAKDOWN_AREAS:
                breakdown_cell = _t22_total_only_breakdown_cell(
                    pns_t22_staging, pns_aes, host_cid, area
                )
                if breakdown_cell is not None:
                    item_cells[f"{host_cid}_{area}"] = breakdown_cell

    # ── Post-loop: reporting country funding staging → item 1403 matrix cells ──
    reporting_funding_col = reporting_funding_matrix_column(ctx)
    for (aes_id, funding_item_id, row_name), total in reporting_funding_staging.items():
        if total:
            cell_key = f"{row_name}_{reporting_funding_col}"
            matrix_cells[(aes_id, funding_item_id)][cell_key] = total

    # Build reverse map: aes_id → (iso3, period) across ALL templates.
    aes_meta: Dict[int, Tuple[str, str]] = {}
    for tpl_map in ctx.assignment_by_template.values():
        for (pn, iso), aid in tpl_map.items():
            aes_meta.setdefault(aid, (iso, pn))

    for (aes_id, item_id), cells in matrix_cells.items():
        if not cells:
            continue
        iso3, period = aes_meta.get(aes_id, ("", ""))
        import_rows.append(
            _matrix_row(
                aes_id=aes_id,
                item_id=item_id,
                cells=cells,
                iso3=iso3,
                period=period,
                debug_kpi=f"matrix_{item_id}",
            )
        )

    if REPORTING_COUNTRY_TEMPLATE_ID in tids:
        _fill_missing_core_yes_no_defaults(
            ctx=ctx,
            import_rows=import_rows,
            filled_core_yes_no=filled_core_yes_no,
            target_aes_ids=_reporting_aes_ids_from_excel(filtered, ctx, template_ids=tids),
            aes_meta=aes_meta,
        )

    ctx.pns_t22_reported_aes = {int(aes) for aes, _ in pns_t22_reported_yes}
    ctx.pns_t23_reported_aes = {int(aes) for aes, _ in pns_t23_reported_yes}

    return import_rows


def upsert_upr_discussion_comments(
    entries: List[Dict[str, Any]],
    *,
    dry_run: bool = False,
) -> Dict[str, int]:
    """Replace UPR-import discussion comments for affected assignments."""
    from app.extensions import db
    from app.models import SubmissionDiscussionComment
    from app.utils.datetime_helpers import utcnow
    from app.utils.discussion_comments import DISCUSSION_SOURCE_UPR_EXCEL

    stats = {
        "discussion_inserted": 0,
        "discussion_deleted": 0,
        "discussion_skipped": 0,
    }
    if not entries:
        return stats

    aes_ids = sorted({int(entry["aes_id"]) for entry in entries if entry.get("aes_id")})
    if dry_run:
        stats["discussion_inserted"] = sum(
            1 for entry in entries if (entry.get("body") or "").strip()
        )
        return stats

    if aes_ids:
        deleted = (
            SubmissionDiscussionComment.query.filter(
                SubmissionDiscussionComment.assignment_entity_status_id.in_(aes_ids),
                SubmissionDiscussionComment.source == DISCUSSION_SOURCE_UPR_EXCEL,
            ).delete(synchronize_session=False)
        )
        stats["discussion_deleted"] = int(deleted or 0)

    now = utcnow()
    for entry in entries:
        body = (entry.get("body") or "").strip()
        if not body:
            stats["discussion_skipped"] += 1
            continue
        db.session.add(
            SubmissionDiscussionComment(
                assignment_entity_status_id=int(entry["aes_id"]),
                body=body,
                created_by_user_id=None,
                created_at=now,
                source=entry.get("source") or DISCUSSION_SOURCE_UPR_EXCEL,
            )
        )
        stats["discussion_inserted"] += 1

    db.session.commit()
    return stats


def run_upr_import(
    input_path: str,
    *,
    template_ids: Optional[List[int]] = None,
    rounds: Optional[List[str]] = None,
    dry_run: bool = False,
    batch_size: int = 1000,
    preview_excel_path: Optional[str] = None,
    progress_cb: Optional[Callable[[Dict[str, Any]], None]] = None,
    cancel_check: Optional[Callable[[], bool]] = None,
    ensure_staff_matrix: bool = True,  # kept for API backward compat, no longer used
    use_row_cache: bool = True,
    use_transform_cache: bool = False,
) -> Dict[str, Any]:
    """Load UPR Excel, transform, and upsert into form_data."""
    from app.extensions import db
    from contextlib import nullcontext

    tids = template_ids or list(UPR_TEMPLATE_PROFILES.keys())
    for tid in tids:
        if tid not in UPR_TEMPLATE_PROFILES:
            raise ValueError(f"Template {tid} is not configured for UPR Excel import")

    def _progress(stage: str, message: str, percent: float, **extra: Any) -> None:
        if not progress_cb:
            return
        payload = {"stage": stage, "message": message, "percent": percent, **extra}
        progress_cb(payload)

    # Reuse an existing Flask app context (e.g. from a web request or the async
    # worker thread) instead of creating a nested one. Only push a fresh context
    # when running from the CLI or any other context-free environment.
    try:
        from flask import current_app as _cur
        _cur._get_current_object()
        _ctx = nullcontext()
    except RuntimeError:
        from app import create_app
        _ctx = create_app().app_context()

    stats: Dict[str, Any] = {"loaded": 0, "skipped": 0, "inserted": 0, "updated": 0, "errors": 0, "warnings": []}

    with _ctx:
        _progress("read", "Reading UPR Data sheet...", 5.0)
        import_rows, ctx, from_transform_cache = prepare_upr_transform(
            input_path,
            tids,
            rounds=rounds,
            use_row_cache=use_row_cache,
            use_transform_cache=use_transform_cache,
            save_transform_cache=False,
        )
        if from_transform_cache:
            _progress("transform", "Using cached transform from preview...", 15.0)
        else:
            _progress("transform", "Mapping Excel rows to form items...", 15.0)
        stats.update(summarize_warnings(ctx.warnings))
        stats["transformed"] = len(import_rows)
        stats["dynamic_transformed"] = len(ctx.dynamic_indicator_entries)
        stats["emergency_slots_staged"] = len(ctx.emergency_slot_meta)

        if preview_excel_path and import_rows:
            write_rows_to_excel(import_rows, preview_excel_path)

        valid_item_ids: Optional[Set[int]] = None
        if tids:
            from app.utils.stable_key import published_form_item_id_set_for_templates

            valid_item_ids = published_form_item_id_set_for_templates(tids)

        valid_aes_ids: Set[int] = set()
        for tpl_map in ctx.assignment_by_template.values():
            valid_aes_ids.update(int(v) for v in tpl_map.values())

        def _emit(payload: Dict[str, Any]) -> None:
            if cancel_check and cancel_check():
                db.session.rollback()
                raise UprImportCancelled()
            if progress_cb:
                progress_cb(payload)

        upsert_stats = upsert_form_data_rows(
            import_rows,
            dry_run=dry_run,
            batch_size=batch_size,
            valid_form_item_ids=valid_item_ids,
            valid_assignment_entity_status_ids=valid_aes_ids,
            progress_cb=_emit,
            cancel_check=cancel_check,
            progress_start_pct=25.0,
            progress_end_pct=85.0,
            stats=stats,
        )
        discussion_stats = upsert_upr_discussion_comments(
            ctx.discussion_comment_entries,
            dry_run=dry_run,
        )
        upsert_stats.update(discussion_stats)
        if not dry_run and import_rows:
            from app.services.forms.variable_resolution_service import VariableResolutionService

            updated_sources = set()
            for row in import_rows:
                try:
                    aes_raw = row.get(COL_ASSIGNMENT) or row.get("assignment_entity_status_id")
                    item_raw = row.get(COL_ITEM) or row.get("form_item_id")
                    aes_id = int(aes_raw) if aes_raw else 0
                    form_item_id = int(item_raw) if item_raw else 0
                except (TypeError, ValueError):
                    continue
                if aes_id and form_item_id:
                    updated_sources.add((aes_id, form_item_id))
            if updated_sources:
                refresh_stats = VariableResolutionService.refresh_stale_variable_matrix_cells(
                    updated_sources
                )
                upsert_stats.update(
                    {
                        "variable_consumer_rows_checked": refresh_stats.get(
                            "consumer_rows_checked", 0
                        ),
                        "variable_cells_cleared": refresh_stats.get("cells_cleared", 0),
                        "variable_consumer_rows_updated": refresh_stats.get("rows_updated", 0),
                    }
                )
        emergency_stats = upsert_emergency_repeat_slots(ctx, dry_run=dry_run)
        upsert_stats.update(emergency_stats)
        dyn_stats = upsert_dynamic_indicator_entries(
            ctx.dynamic_indicator_entries,
            dry_run=dry_run,
        )
        upsert_stats.update(dyn_stats)
        if 22 in tids or 23 in tids:
            _progress("status", "Clearing non-reported PNS data and syncing status...", 92.0)
            pns_sync_stats = sync_non_reported_pns_assignments(
                ctx,
                tids,
                rounds=_normalize_round_set(rounds),
                dry_run=dry_run,
            )
            upsert_stats.update(pns_sync_stats)
        upsert_stats.update(summarize_warnings(ctx.warnings))
        upsert_stats["transformed"] = len(import_rows)
        upsert_stats["dynamic_transformed"] = len(ctx.dynamic_indicator_entries)
        if preview_excel_path:
            upsert_stats["preview_path"] = preview_excel_path
        _progress("complete", "UPR import completed.", 100.0, stats=upsert_stats)
        return upsert_stats


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    parser = argparse.ArgumentParser(description="Import UPR Master Excel into form data")
    parser.add_argument("--input", required=True, help="Path to UPR Master.xlsx")
    parser.add_argument(
        "--templates",
        default=",".join(str(t) for t in UPR_TEMPLATE_PROFILES),
        help="Comma-separated template IDs (default: 24,22)",
    )
    parser.add_argument("--rounds", default="", help="Comma-separated round codes (e.g. P26, AR25, MYR26). Default: all rounds matching the selected templates.")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, no DB writes")
    parser.add_argument("--batch-size", type=int, default=1000)
    parser.add_argument("--preview-excel", default="", help="Write ready-to-import preview Excel")
    parser.add_argument("--analyze-only", action="store_true", help="Print workbook summary and exit")
    args = parser.parse_args()

    if not os.path.isfile(args.input):
        logger.error("File not found: %s", args.input)
        return 1

    if args.analyze_only:
        summary = analyze_workbook(args.input)
        print(json.dumps(summary, indent=2))
        return 0

    template_ids = [int(x.strip()) for x in args.templates.split(",") if x.strip()]
    rounds = [x.strip() for x in args.rounds.split(",") if x.strip()] or None
    preview = args.preview_excel or None

    try:
        stats = run_upr_import(
            args.input,
            template_ids=template_ids,
            rounds=rounds,
            dry_run=args.dry_run,
            batch_size=args.batch_size,
            preview_excel_path=preview,
        )
    except UprImportCancelled:
        logger.warning("Import cancelled")
        return 2

    print(json.dumps(stats, indent=2))
    return 0 if stats.get("errors", 0) == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
