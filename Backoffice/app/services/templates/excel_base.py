"""Shared constants and helpers for template Excel import/export."""

# ========== Template Excel Import/Export Service ==========
from app.utils.datetime_helpers import utcnow
"""
Service for exporting and importing form templates to/from Excel.
"""
from flask import current_app
from flask_login import current_user
from app import db
from app.models import (
    FormTemplate, FormPage, FormSection, FormItem, FormTemplateVersion,
    FormData,
    RepeatGroupInstance, RepeatGroupData, DynamicIndicatorData, DynamicSectionContext,
)
from app.models.documents import SubmittedDocument
from app.models.indicator_bank import IndicatorBank
from contextlib import suppress
from app.services.monitoring.memory import memory_tracker
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import io
import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from app.utils.stable_key import generate_stable_key, is_valid_stable_key, normalize_stable_key



class TemplateExcelBase(object):
    """Mixin for TemplateExcelService."""

    IFRC_COLORS = {
        'RED': "FFED1B2E",
        'DARK_RED': "FFAF0E1B",
        'LIGHT_GRAY': "FFF5F5F5",
        'MEDIUM_GRAY': "FFE0E0E0",
        'DARK_GRAY': "FF666666",
        'WHITE': "FFFFFFFF",
        'YELLOW': "FFFFF9E6",
        'BLUE': "FF0066CC",
        'DARK_BLUE': "FF004499",
    }

    REQUIRED_COLUMNS = {
        'Template': ['name'],
        'Pages': ['id', 'name', 'order'],
        'Sections': ['id', 'name', 'order'],
        'Items': ['id', 'section_id', 'item_type', 'label', 'order']
    }

    OPTIONAL_HEADER_COLUMNS = {
        'Sections': {'stable_key'},
        'Items': {'stable_key'},
    }

    EXCEL_EXPORT_VERSION = 'V2'

    FALLBACK_TRANSLATABLE_LANGUAGES = ['fr', 'es', 'ar', 'ru', 'zh']

    TRANSLATION_DB_FIELDS = {
        'name': 'name_translations',
        'label': 'label_translations',
        'definition': 'definition_translations',
        'description': 'description_translations',
    }

    TEMPLATE_SHEET_ROW_HEADERS = ['field', 'value']

    TEMPLATE_TRANSLATABLE_FIELDS = ['name']

    PAGE_TRANSLATABLE_FIELDS = ['name']

    SECTION_TRANSLATABLE_FIELDS = ['name']

    ITEM_TRANSLATABLE_FIELDS = ['label', 'definition', 'description']

    TEMPLATE_BASE_COLUMNS = [
        'name',
        'description',
        'is_paginated',
        'add_to_self_report',
        'display_order_visible',
        'enable_export_pdf',
        'enable_export_excel',
        'enable_import_excel',
        'enable_ai_validation',
        'variables',
    ]

    PAGE_BASE_COLUMNS = ['id', 'order', 'name']

    SECTION_BASE_COLUMNS = [
        'id',
        'order',
        'name',
        'page_id',
        'parent_section_id',
        'section_type',
        'archived',
        'max_dynamic_indicators',
        'allowed_sectors',
        'indicator_filters',
        'allow_data_not_available',
        'allow_not_applicable',
        'allowed_disaggregation_options',
        'data_entry_display_filters',
        'add_indicator_note',
        'relevance_condition',
    ]

    ITEM_BASE_COLUMNS = [
        'id',
        'section_id',
        'item_type',
        'order',
        'archived',
        'label',
        'definition',
        'description',
        'options_json',
        'options_translations',
        'lookup_list_id',
        'list_display_column',
        'list_filters_json',
        'indicator_bank_id',
        'type',
        'unit',
        'relevance_condition',
        'validation_condition',
        'validation_message',
        'config',
    ]

    TEMPLATE_LEGACY_COLUMNS = [
        'name', 'description',
        'add_to_self_report', 'display_order_visible',
        'is_paginated', 'enable_export_pdf', 'enable_export_excel',
        'enable_import_excel', 'enable_ai_validation', 'name_translations', 'variables',
    ]

    PAGE_LEGACY_COLUMNS = ['id', 'name', 'order', 'name_translations']

    SECTION_LEGACY_COLUMNS = [
        'id', 'name', 'order', 'parent_section_id', 'page_id',
        'section_type', 'max_dynamic_indicators', 'allowed_sectors',
        'indicator_filters', 'allow_data_not_available', 'allow_not_applicable',
        'allowed_disaggregation_options', 'data_entry_display_filters',
        'add_indicator_note', 'name_translations', 'relevance_condition', 'archived',
    ]

    ITEM_LEGACY_COLUMNS = [
        'id', 'section_id', 'item_type', 'label', 'order',
        'relevance_condition', 'archived', 'config', 'indicator_bank_id',
        'type', 'unit', 'validation_condition', 'validation_message',
        'definition', 'options_json', 'lookup_list_id', 'list_display_column',
        'list_filters_json', 'label_translations', 'definition_translations',
        'options_translations', 'description_translations', 'description',
    ]

    JSON_EXPORT_COLUMNS = frozenset({
        'variables',
        'allowed_sectors',
        'indicator_filters',
        'allowed_disaggregation_options',
        'data_entry_display_filters',
        'relevance_condition',
        'validation_condition',
        'config',
        'options_json',
        'list_filters_json',
        'options_translations',
        # Legacy JSON translation columns (import-only)
        'name_translations',
        'label_translations',
        'definition_translations',
        'description_translations',
    })

    @classmethod
    def _get_translatable_languages(cls) -> List[str]:
        try:
            langs = current_app.config.get('TRANSLATABLE_LANGUAGES')
            if langs:
                return list(langs)
        except RuntimeError:
            pass
        return list(cls.FALLBACK_TRANSLATABLE_LANGUAGES)

    @classmethod
    def _expand_columns_with_translations(
        cls, base_columns: List[str], translatable_fields: List[str]
    ) -> List[str]:
        """Insert {field}_{lang} columns immediately after each translatable base field."""
        translatable = set(translatable_fields)
        expanded: List[str] = []
        for col in base_columns:
            expanded.append(col)
            if col in translatable:
                for lang in cls._get_translatable_languages():
                    expanded.append(f"{col}_{lang}")
        return expanded

    @classmethod
    def get_template_columns(cls) -> List[str]:
        """Ordered Template sheet field names (one row per field in export)."""
        return cls._expand_columns_with_translations(
            cls.TEMPLATE_BASE_COLUMNS, cls.TEMPLATE_TRANSLATABLE_FIELDS
        )

    @classmethod
    def _template_sheet_uses_row_layout(cls, headers: List[Any]) -> bool:
        if len(headers) < 2:
            return False
        first = str(headers[0]).strip().lower() if headers[0] is not None else ''
        second = str(headers[1]).strip().lower() if headers[1] is not None else ''
        return first == 'field' and second == 'value'

    @classmethod
    def _build_template_field_values(
        cls, template: FormTemplate, version: FormTemplateVersion
    ) -> Dict[str, Any]:
        version_name = version.name if version.name else template.name
        version_name_translations = (
            version.name_translations if version.name_translations else template.name_translations
        )
        version_description = version.description
        if version_description is None:
            version_description = getattr(template, 'description', None)

        row_values = cls._build_export_row(
            cls.TEMPLATE_BASE_COLUMNS,
            cls.TEMPLATE_TRANSLATABLE_FIELDS,
            {
                'name': version_name,
                'description': version_description,
                'add_to_self_report': (
                    version.add_to_self_report
                    if version.add_to_self_report is not None
                    else getattr(template, 'add_to_self_report', False)
                ),
                'display_order_visible': (
                    version.display_order_visible
                    if version.display_order_visible is not None
                    else getattr(template, 'display_order_visible', False)
                ),
                'is_paginated': (
                    version.is_paginated
                    if version.is_paginated is not None
                    else getattr(template, 'is_paginated', False)
                ),
                'enable_export_pdf': (
                    version.enable_export_pdf
                    if version.enable_export_pdf is not None
                    else getattr(template, 'enable_export_pdf', False)
                ),
                'enable_export_excel': (
                    version.enable_export_excel
                    if version.enable_export_excel is not None
                    else getattr(template, 'enable_export_excel', False)
                ),
                'enable_import_excel': (
                    version.enable_import_excel
                    if version.enable_import_excel is not None
                    else getattr(template, 'enable_import_excel', False)
                ),
                'enable_ai_validation': (
                    version.enable_ai_validation
                    if getattr(version, 'enable_ai_validation', None) is not None
                    else getattr(template, 'enable_ai_validation', False)
                ),
                'variables': cls._format_json_for_excel(
                    version.variables if version.variables else None
                ),
            },
            {'name': version_name_translations},
        )
        return dict(zip(cls.get_template_columns(), row_values))

    @classmethod
    def _parse_template_sheet(
        cls, sheet,
    ) -> Tuple[Dict[str, Any], bool, List[str]]:
        """Parse Template sheet. Returns (field_values, legacy_column_layout, errors)."""
        headers = [cell.value for cell in sheet[1]]
        errors: List[str] = []

        if cls._template_sheet_uses_row_layout(headers):
            row_data: Dict[str, Any] = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None for cell in row):
                    continue
                field_name = row[0]
                if field_name is None or str(field_name).strip() == '':
                    continue
                value = row[1] if len(row) > 1 else None
                row_data[str(field_name).strip()] = value
            if 'name' not in row_data or str(row_data.get('name') or '').strip() == '':
                errors.append('Template sheet is missing required field: name')
            return row_data, False, errors

        # Legacy column layout (headers in row 1, values in row 2)
        header_list = [h for h in headers if h is not None]
        required_headers = cls.REQUIRED_COLUMNS.get('Template', ['name'])
        header_set = {h for h in header_list if isinstance(h, str) and h}
        missing = set(required_headers) - header_set
        if missing:
            errors.append(
                f"Template sheet headers missing required columns. "
                f"Required: {required_headers}, Missing: {list(missing)}, Got: {headers}"
            )
            return {}, True, errors

        row = next(sheet.iter_rows(min_row=2, values_only=True), None)
        if not row or all(cell is None for cell in row):
            return {}, True, []

        row_data = cls._row_dict_from_sheet(headers, row)
        return row_data, True, errors

    @classmethod
    def get_page_columns(cls) -> List[str]:
        return cls._expand_columns_with_translations(
            cls.PAGE_BASE_COLUMNS, cls.PAGE_TRANSLATABLE_FIELDS
        )

    @classmethod
    def get_section_columns(cls) -> List[str]:
        return cls._expand_columns_with_translations(
            cls.SECTION_BASE_COLUMNS, cls.SECTION_TRANSLATABLE_FIELDS
        )

    @classmethod
    def get_item_columns(cls) -> List[str]:
        return cls._expand_columns_with_translations(
            cls.ITEM_BASE_COLUMNS, cls.ITEM_TRANSLATABLE_FIELDS
        )

    @classmethod
    def _format_sheet_header_error(
        cls,
        sheet_name: str,
        *,
        summary: str,
        got: Optional[List[str]] = None,
        missing: Optional[List[str]] = None,
        unrecognized: Optional[List[str]] = None,
    ) -> str:
        parts = [f"{sheet_name} sheet: {summary}"]
        if missing:
            parts.append(f"Missing column(s): {', '.join(missing)}")
        if unrecognized:
            parts.append(f"Unrecognized column(s): {', '.join(unrecognized)}")
        if got:
            parts.append(f"Found ({len(got)}): {', '.join(got)}")
        return ' '.join(parts)

    @classmethod
    def _sheet_uses_legacy_translation_columns(cls, sheet_name: str, header_set: set) -> bool:
        legacy_json_cols = {
            'Template': ['name_translations'],
            'Pages': ['name_translations'],
            'Sections': ['name_translations'],
            'Items': ['label_translations', 'definition_translations', 'description_translations'],
        }.get(sheet_name, [])
        return any(col in header_set for col in legacy_json_cols)

    @classmethod
    def _resolve_sheet_headers(cls, sheet_name: str, headers: List[Any]) -> Tuple[List[str], bool, Optional[str]]:
        """Return (expected_headers, legacy_format, error_message).

        Column order is ignored; only required columns, recognized names, and format matter.
        """
        normalized = [str(h).strip() for h in headers if h is not None and str(h).strip()]
        if not normalized:
            return [], False, cls._format_sheet_header_error(
                sheet_name, summary='no column headers found'
            )

        header_set = set(normalized)
        layouts = {
            'Template': (cls.get_template_columns(), cls.TEMPLATE_LEGACY_COLUMNS),
            'Pages': (cls.get_page_columns(), cls.PAGE_LEGACY_COLUMNS),
            'Sections': (cls.get_section_columns(), cls.SECTION_LEGACY_COLUMNS),
            'Items': (cls.get_item_columns(), cls.ITEM_LEGACY_COLUMNS),
        }
        current, legacy = layouts[sheet_name]
        current_set = set(current)
        legacy_set = set(legacy)
        optional_import = cls.OPTIONAL_HEADER_COLUMNS.get(sheet_name, set())
        required = set(cls.REQUIRED_COLUMNS.get(sheet_name, []))

        missing_required = sorted(required - header_set)
        if missing_required:
            return current, False, cls._format_sheet_header_error(
                sheet_name,
                summary='missing required column(s)',
                missing=missing_required,
                got=normalized,
            )

        known_set = current_set | legacy_set | optional_import
        unrecognized = sorted(header_set - known_set)
        if unrecognized:
            return current, False, cls._format_sheet_header_error(
                sheet_name,
                summary='unrecognized column(s)',
                unrecognized=unrecognized,
                got=normalized,
            )

        legacy_format = cls._sheet_uses_legacy_translation_columns(sheet_name, header_set)
        if legacy_format:
            if header_set <= legacy_set:
                return legacy, True, None
            missing_legacy = sorted(legacy_set - header_set)
            return legacy, True, cls._format_sheet_header_error(
                sheet_name,
                summary='legacy export format but missing expected column(s)',
                missing=missing_legacy,
                got=normalized,
            )

        if header_set <= current_set:
            return current, False, None

        missing_from_current = current_set - header_set
        if missing_from_current <= optional_import:
            return current, False, None

        return current, False, cls._format_sheet_header_error(
            sheet_name,
            summary='missing expected column(s) for the current export format',
            missing=sorted(missing_from_current - optional_import),
            got=normalized,
        )

    @classmethod
    def _translation_export_values(cls, translations: Any) -> List[Any]:
        data = cls._normalize_json_for_export(translations)
        if not isinstance(data, dict):
            data = {}
        values: List[Any] = []
        for lang in cls._get_translatable_languages():
            val = data.get(lang)
            if val is None or str(val).strip() == '':
                values.append(None)
            elif isinstance(val, str):
                values.append(cls._decode_unicode_escapes_in_str(val))
            else:
                values.append(val)
        return values

    @classmethod
    def _build_export_row(
        cls,
        base_columns: List[str],
        translatable_fields: List[str],
        scalar_values: Dict[str, Any],
        translation_values: Optional[Dict[str, Any]] = None,
    ) -> List[Any]:
        translatable = set(translatable_fields)
        translation_values = translation_values or {}
        row: List[Any] = []
        for col in base_columns:
            row.append(scalar_values.get(col))
            if col in translatable:
                row.extend(cls._translation_export_values(translation_values.get(col)))
        return row

    @classmethod
    def _collect_translations_from_row(
        cls,
        row_data: Dict[str, Any],
        base_field: str,
        *,
        legacy: bool = False,
    ) -> Optional[Dict[str, str]]:
        translations: Dict[str, str] = {}
        if not legacy:
            for lang in cls._get_translatable_languages():
                col = f"{base_field}_{lang}"
                val = row_data.get(col)
                if val is not None and str(val).strip() != '':
                    translations[lang] = cls._decode_unicode_escapes_in_str(str(val).strip())
        legacy_col = cls.TRANSLATION_DB_FIELDS.get(base_field)
        if legacy_col:
            parsed = cls._parse_json(row_data.get(legacy_col))
            if isinstance(parsed, dict):
                for lang, val in parsed.items():
                    if val is not None and str(val).strip() and lang not in translations:
                        translations[str(lang)] = cls._decode_unicode_escapes_in_str(str(val).strip())
        return translations or None

    @classmethod
    def _row_dict_from_sheet(cls, headers: List[Any], row: Tuple[Any, ...]) -> Dict[str, Any]:
        row_data: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if header is None:
                continue
            row_data[str(header)] = row[idx] if idx < len(row) else None
        return row_data

    @classmethod
    def _apply_item_translations_from_row(
        cls,
        item: FormItem,
        row_data: Dict[str, Any],
        *,
        legacy: bool = False,
    ) -> None:
        for base_field in cls.ITEM_TRANSLATABLE_FIELDS:
            db_field = cls.TRANSLATION_DB_FIELDS[base_field]
            setattr(
                item,
                db_field,
                cls._collect_translations_from_row(row_data, base_field, legacy=legacy),
            )

    DROPDOWN_OPTIONS = {
        'section_type': ['standard', 'repeat', 'dynamic_indicators'],
        # Note: item_type is now dynamic - see _get_item_type_options()
        'archived': ['TRUE', 'FALSE'],
        'add_to_self_report': ['TRUE', 'FALSE'],
        'display_order_visible': ['TRUE', 'FALSE'],
        'is_paginated': ['TRUE', 'FALSE'],
        'enable_export_pdf': ['TRUE', 'FALSE'],
        'enable_export_excel': ['TRUE', 'FALSE'],
        'enable_import_excel': ['TRUE', 'FALSE'],
        'enable_ai_validation': ['TRUE', 'FALSE'],
        'allow_data_not_available': ['TRUE', 'FALSE'],
        'allow_not_applicable': ['TRUE', 'FALSE'],
    }

    _UNICODE_ESCAPE_RE = re.compile(r'\\u([0-9a-fA-F]{4})')

    @classmethod
    def _decode_unicode_escapes_in_str(cls, value: str) -> str:
        """Convert literal \\uXXXX sequences to real Unicode characters."""
        if not isinstance(value, str) or '\\u' not in value:
            return value

        def _repl(match: re.Match[str]) -> str:
            try:
                return chr(int(match.group(1), 16))
            except ValueError:
                return match.group(0)

        return cls._UNICODE_ESCAPE_RE.sub(_repl, value)

    @classmethod
    def _normalize_json_for_export(cls, value: Any) -> Any:
        """Normalize JSON payloads so exports show readable Unicode text."""
        if value is None or value == '' or value == 'None':
            return None
        if isinstance(value, str):
            parsed = cls._parse_json(value)
            if isinstance(parsed, (dict, list)):
                return cls._normalize_json_for_export(parsed)
            if isinstance(parsed, str) and parsed != value:
                return cls._normalize_json_for_export(parsed)
            return cls._decode_unicode_escapes_in_str(value)
        if isinstance(value, dict):
            return {key: cls._normalize_json_for_export(val) for key, val in value.items()}
        if isinstance(value, list):
            return [cls._normalize_json_for_export(item) for item in value]
        return value

    @classmethod
    def _format_json_for_excel(cls, value: Any) -> Optional[str]:
        """Serialize JSON values with line breaks for readable Excel cells."""
        normalized = cls._normalize_json_for_export(value)
        if normalized is None:
            return None
        if isinstance(normalized, str):
            return normalized
        try:
            return json.dumps(normalized, ensure_ascii=False, indent=2)
        except (TypeError, ValueError):
            try:
                return json.dumps(normalized, ensure_ascii=False, default=str, indent=2)
            except (TypeError, ValueError):
                return str(normalized)

    @classmethod
    def _write_data_row(cls, sheet, row_idx: int, headers: List[str], row_data: List[Any]) -> None:
        """Write a data row."""
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    @classmethod
    def _get_items_for_version(cls, template: FormTemplate, version: FormTemplateVersion) -> List[FormItem]:
        """Get items for a template version in a deterministic order.

        Rows are grouped by section display order, then item order within each section.
        Stable secondary sorts on section/item IDs keep export IDs deterministic.
        """
        return FormItem.query.join(FormSection).filter(
            FormItem.template_id == template.id,
            FormSection.version_id == version.id
        ).order_by(
            FormSection.order,
            FormSection.id,
            FormItem.order,
            FormItem.id,
        ).all()

    @classmethod
    def _build_item_db_to_export_map(cls, template: FormTemplate, version: FormTemplateVersion) -> Dict[int, int]:
        """Build mapping from DB item IDs -> sequential export IDs (1, 2, 3...)."""
        items = cls._get_items_for_version(template, version)
        return {item.id: idx + 1 for idx, item in enumerate(items)}

    @classmethod
    def _rewrite_rule_json_item_ids(cls, rule_json: Any, id_map: Dict[int, int]) -> Any:
        """Rewrite item references inside a relevance/validation rule JSON.

        The rule builder stores references under the key 'item_id'. Values can be:
        - numeric strings (e.g., "66") for regular items
        - prefixed strings (e.g., "plugin_123" or "plugin_123_measure_id")
        - legacy prefixed strings (e.g., "question_66")

        This method is used in two directions depending on the provided id_map:
        - Export: db_id -> export_id
        - Import: export_id -> new_db_id
        """
        if rule_json is None:
            return None
        if isinstance(rule_json, str) and rule_json.strip() == '':
            return rule_json

        # Parse (handle occasional double-encoded JSON)
        parsed = None
        if isinstance(rule_json, (dict, list)):
            parsed = rule_json
        else:
            raw = str(rule_json)
            with suppress(Exception):
                parsed = json.loads(raw)
                if isinstance(parsed, str):
                    parsed2 = json.loads(parsed)
                    parsed = parsed2
        if parsed is None:
            return rule_json

        def _rewrite_item_id_value(val: Any) -> Any:
            if val is None:
                return val
            # Numeric id stored as int/float/string
            if isinstance(val, (int, float)) or (isinstance(val, str) and val.isdigit()):
                old_id = int(val)
                new_id = id_map.get(old_id)
                return str(new_id) if new_id is not None else val

            if isinstance(val, str):
                # plugin_123 or plugin_123_measure_id
                if val.startswith('plugin_'):
                    parts = val.split('_')
                    if len(parts) >= 2 and parts[1].isdigit():
                        old_id = int(parts[1])
                        new_id = id_map.get(old_id)
                        if new_id is not None:
                            parts[1] = str(new_id)
                            return '_'.join(parts)
                    return val

                # Legacy prefixed IDs (question_66, indicator_12, document_field_7, form_item_99)
                m = re.match(r'^(question|indicator|document_field|form_item)_(\d+)$', val)
                if m:
                    old_id = int(m.group(2))
                    new_id = id_map.get(old_id)
                    if new_id is not None:
                        return f"{m.group(1)}_{new_id}"
                    return val

            return val

        def _walk(obj: Any) -> Any:
            if isinstance(obj, list):
                for i in range(len(obj)):
                    obj[i] = _walk(obj[i])
                return obj
            if isinstance(obj, dict):
                for k, v in list(obj.items()):
                    if k == 'item_id':
                        obj[k] = _rewrite_item_id_value(v)
                    else:
                        obj[k] = _walk(v)
                return obj
            return obj

        parsed = _walk(parsed)
        try:
            return cls._format_json_for_excel(parsed)
        except Exception as e:
            current_app.logger.debug("Rule JSON serialization failed, using original: %s", e)
            # Fall back to the original if serialization fails
            return rule_json

    @classmethod
    def _parse_json(cls, value) -> Optional[Any]:
        """Parse JSON string to object; return None only when empty or invalid."""
        if value is None or value == '' or value == 'None':
            return None
        if isinstance(value, (dict, list)):
            return value
        try:
            text = str(value).strip()
            if text.startswith('\ufeff'):
                text = text[1:]
            parsed = json.loads(text)
            if isinstance(parsed, (dict, list)):
                return parsed
            return None
        except (json.JSONDecodeError, TypeError):
            return None

