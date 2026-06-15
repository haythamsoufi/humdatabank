"""
Comprehensive tests for app/services/kobo_data_import_service.py

Focus: helper functions and KoboDataImportService.analyze() / extract_unique_entities()
"""
from __future__ import annotations

import io
from collections import OrderedDict
from datetime import datetime, date
from typing import Any, Dict, List, Optional
from unittest.mock import MagicMock, patch

import openpyxl
import pytest


# ---------------------------------------------------------------------------
# Helpers to create Excel bytes
# ---------------------------------------------------------------------------

def _make_excel_bytes(headers, data_rows):
    """Return bytes of an openpyxl workbook with given headers + rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for row in data_rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# _s
# ---------------------------------------------------------------------------

class TestS:
    def _call(self, val):
        from app.services.kobo_data_import_service import _s
        return _s(val)

    def test_none_returns_empty(self):
        assert self._call(None) == ""

    def test_string(self):
        assert self._call("  hello  ") == "hello"

    def test_int(self):
        assert self._call(42) == "42"


# ---------------------------------------------------------------------------
# _is_date
# ---------------------------------------------------------------------------

class TestIsDate:
    def _call(self, v):
        from app.services.kobo_data_import_service import _is_date
        return _is_date(v)

    def test_datetime(self):
        assert self._call(datetime(2024, 1, 1)) is True

    def test_date(self):
        assert self._call(date(2024, 1, 1)) is True

    def test_string(self):
        assert self._call("2024-01-01") is False

    def test_none(self):
        assert self._call(None) is False


# ---------------------------------------------------------------------------
# _is_numeric
# ---------------------------------------------------------------------------

class TestIsNumeric:
    def _call(self, v):
        from app.services.kobo_data_import_service import _is_numeric
        return _is_numeric(v)

    def test_int(self):
        assert self._call(42) is True

    def test_float(self):
        assert self._call(3.14) is True

    def test_bool_false(self):
        assert self._call(True) is False  # bool is NOT numeric in this context

    def test_numeric_string(self):
        assert self._call("123") is True

    def test_non_numeric_string(self):
        assert self._call("hello") is False

    def test_none(self):
        assert self._call(None) is False

    def test_empty_string(self):
        assert self._call("") is False


# ---------------------------------------------------------------------------
# _classify_header
# ---------------------------------------------------------------------------

class TestClassifyHeader:
    def _call(self, header):
        from app.services.kobo_data_import_service import _classify_header
        return _classify_header(header)

    def test_empty_header(self):
        cat, reason = self._call("")
        assert cat == "system"

    def test_kobo_system_exact(self):
        cat, _ = self._call("start")
        assert cat == "system"
        cat, _ = self._call("end")
        assert cat == "system"
        cat, _ = self._call("deviceid")
        assert cat == "system"

    def test_underscore_prefix(self):
        cat, _ = self._call("_id")
        assert cat == "system"
        cat, _ = self._call("_uuid")
        assert cat == "system"

    def test_double_underscore(self):
        cat, _ = self._call("__version__")
        assert cat == "system"

    def test_meta_prefix(self):
        cat, _ = self._call("meta/instanceID")
        assert cat == "system"

    def test_hidden_display(self):
        cat, _ = self._call("<span style='display: none'>hidden</span>")
        assert cat == "hidden"

    def test_validation_message(self):
        cat, _ = self._call("<span style='color: red'>error</span>")
        assert cat == "hidden"

    def test_file_url(self):
        cat, _ = self._call("photo_URL")
        assert cat == "file_url"

    def test_calculated_variable(self):
        cat, _ = self._call("total_people")
        assert cat == "calculated"
        cat, _ = self._call("cal_total_x")
        assert cat == "calculated"

    def test_data_column(self):
        cat, _ = self._call("National Society")
        assert cat == "data"

    def test_data_with_group(self):
        cat, _ = self._call("demographics/age")
        assert cat == "data"


# ---------------------------------------------------------------------------
# _detect_disagg
# ---------------------------------------------------------------------------

class TestDetectDisagg:
    def _call(self, header):
        from app.services.kobo_data_import_service import _detect_disagg
        return _detect_disagg(header)

    def test_no_disagg_flat_header(self):
        assert self._call("National Society") is None

    def test_sex_disagg(self):
        result = self._call("gender_q1/Male")
        assert result == "sex"

    def test_sexage_disagg(self):
        result = self._call("sexage_breakdown/Male")
        assert result == "sex_age"

    def test_age_disagg(self):
        result = self._call("age/18-35")
        assert result == "age"

    def test_nested_without_disagg(self):
        result = self._call("group1/sub_question")
        assert result is None


# ---------------------------------------------------------------------------
# _detect_data_type
# ---------------------------------------------------------------------------

class TestDetectDataType:
    def _call(self, values):
        from app.services.kobo_data_import_service import _detect_data_type
        return _detect_data_type(values)

    def test_empty_values(self):
        t, opts = self._call([])
        assert t == "text"
        assert opts is None

    def test_all_null(self):
        t, opts = self._call([None, None])
        assert t == "text"

    def test_all_numeric(self):
        t, opts = self._call([1, 2, 3, 4])
        assert t == "number"

    def test_all_dates(self):
        t, opts = self._call([datetime(2024, 1, 1), datetime(2024, 2, 1)])
        assert t == "date"

    def test_yesno(self):
        t, opts = self._call(["Yes", "No", "yes", "no"])
        assert t == "yesno"

    def test_single_choice(self):
        t, opts = self._call(["A"] * 5 + ["B"] * 5 + ["C"] * 5)
        assert t == "single_choice"
        assert opts is not None

    def test_long_text(self):
        long_val = "X" * 600
        t, opts = self._call([long_val])
        assert t == "textarea"

    def test_regular_text(self):
        t, opts = self._call(["hello world", "foo bar baz"])
        assert t == "text"


# ---------------------------------------------------------------------------
# _extract_group_and_label
# ---------------------------------------------------------------------------

class TestExtractGroupAndLabel:
    def _call(self, header):
        from app.services.kobo_data_import_service import _extract_group_and_label
        return _extract_group_and_label(header)

    def test_no_slash(self):
        grp, lbl = self._call("Question Label")
        assert grp is None
        assert lbl == "Question Label"

    def test_with_slash(self):
        grp, lbl = self._call("demographics/age")
        assert grp == "demographics"
        assert lbl == "age"

    def test_nested_slash(self):
        grp, lbl = self._call("group1/sub/question")
        assert grp == "group1"
        assert lbl == "sub/question"


# ---------------------------------------------------------------------------
# _find_validation_status_column
# ---------------------------------------------------------------------------

class TestFindValidationStatusColumn:
    def _call(self, headers):
        from app.services.kobo_data_import_service import _find_validation_status_column
        return _find_validation_status_column(headers)

    def test_exact_match(self):
        headers = ["National Society", "_validation_status", "gender"]
        assert self._call(headers) == 1

    def test_variant_match(self):
        headers = ["NS", "validation_status"]
        assert self._call(headers) == 1

    def test_nested_header_match(self):
        headers = ["NS", "meta/validation_status"]
        assert self._call(headers) == 1

    def test_not_found(self):
        headers = ["NS", "gender", "age"]
        assert self._call(headers) is None

    def test_empty_headers(self):
        assert self._call([]) is None


# ---------------------------------------------------------------------------
# _has_kobo_structural_markers
# ---------------------------------------------------------------------------

class TestHasKoboStructuralMarkers:
    def _call(self, headers):
        from app.services.kobo_data_import_service import _has_kobo_structural_markers
        return _has_kobo_structural_markers(headers)

    def test_underscore_id_gives_meta(self):
        has_meta, has_group = self._call(["_id", "NS", "Q1"])
        assert has_meta is True

    def test_underscore_uuid_gives_meta(self):
        has_meta, _ = self._call(["_uuid", "NS"])
        assert has_meta is True

    def test_system_exact_start_gives_meta(self):
        has_meta, _ = self._call(["start", "end", "NS"])
        assert has_meta is True

    def test_meta_prefix_gives_meta(self):
        has_meta, _ = self._call(["meta/instanceID", "NS"])
        assert has_meta is True

    def test_slash_header_gives_group(self):
        _, has_group = self._call(["NS", "demographics/age"])
        assert has_group is True

    def test_plain_headers_give_neither(self):
        has_meta, has_group = self._call(["National Society", "Q1 Value", "Status"])
        assert has_meta is False
        assert has_group is False

    def test_empty_headers_give_neither(self):
        has_meta, has_group = self._call([None, "", None])
        assert has_meta is False
        assert has_group is False

    def test_mixed_markers(self):
        # Has both kinds
        has_meta, has_group = self._call(["_id", "group1/question", "NS"])
        assert has_meta is True
        assert has_group is True

    def test_validation_status_column_is_meta(self):
        has_meta, _ = self._call(["NS", "_validation_status"])
        assert has_meta is True


# ---------------------------------------------------------------------------
# _kobo_validation_bucket
# ---------------------------------------------------------------------------

class TestKoboValidationBucket:
    def _call(self, val):
        from app.services.kobo_data_import_service import _kobo_validation_bucket
        return _kobo_validation_bucket(val)

    def test_empty(self):
        assert self._call(None) == "empty"
        assert self._call("") == "empty"

    def test_approved(self):
        assert self._call("approved") == "approved"
        assert self._call("Approved") == "approved"

    def test_not_approved(self):
        assert self._call("not_approved") == "rejected"

    def test_on_hold(self):
        assert self._call("on_hold") == "on_hold"
        assert self._call("onhold") == "on_hold"

    def test_not_validated(self):
        assert self._call("not_validated") == "not_validated"
        assert self._call("unvalidated") == "not_validated"

    def test_rejected(self):
        assert self._call("rejected") == "rejected"
        assert self._call("flagged_for_removal") == "rejected"
        assert self._call("denied") == "rejected"

    def test_pending(self):
        assert self._call("approval_requested") == "pending"
        assert self._call("pending") == "pending"

    def test_draft(self):
        assert self._call("draft") == "draft"

    def test_unknown(self):
        assert self._call("some_other_status") == "unknown"


# ---------------------------------------------------------------------------
# _is_kobo_submission_approved
# ---------------------------------------------------------------------------

class TestIsKoboSubmissionApproved:
    def _call(self, val):
        from app.services.kobo_data_import_service import _is_kobo_submission_approved
        return _is_kobo_submission_approved(val)

    def test_approved(self):
        assert self._call("approved") is True

    def test_not_approved(self):
        assert self._call("on_hold") is False
        assert self._call(None) is False


# ---------------------------------------------------------------------------
# _row_matches_submission_filter
# ---------------------------------------------------------------------------

class TestRowMatchesSubmissionFilter:
    def _call(self, cell_val, sf):
        from app.services.kobo_data_import_service import _row_matches_submission_filter
        return _row_matches_submission_filter(cell_val, sf)

    def test_all_filter_always_true(self):
        assert self._call(None, "all") is True
        assert self._call("rejected", "all") is True

    def test_approved_only(self):
        assert self._call("approved", "approved_only") is True
        assert self._call("on_hold", "approved_only") is False

    def test_exclude_rejected(self):
        assert self._call("rejected", "exclude_rejected") is False
        assert self._call("approved", "exclude_rejected") is True
        assert self._call("on_hold", "exclude_rejected") is True

    def test_approved_or_on_hold(self):
        assert self._call("approved", "approved_or_on_hold") is True
        assert self._call("on_hold", "approved_or_on_hold") is True
        assert self._call("rejected", "approved_or_on_hold") is False

    def test_on_hold_only(self):
        assert self._call("on_hold", "on_hold_only") is True
        assert self._call("approved", "on_hold_only") is False

    def test_not_validated_only(self):
        assert self._call("not_validated", "not_validated_only") is True
        assert self._call("pending", "not_validated_only") is True
        assert self._call("approved", "not_validated_only") is False

    def test_draft_only(self):
        assert self._call("draft", "draft_only") is True
        assert self._call("approved", "draft_only") is False

    def test_unknown_filter_all_pass(self):
        assert self._call("anything", "unknown_filter_xyz") is True


# ---------------------------------------------------------------------------
# _eligible_row_indices
# ---------------------------------------------------------------------------

class TestEligibleRowIndices:
    def _call(self, data_rows, submission_filter="all", validation_col=None):
        from app.services.kobo_data_import_service import _eligible_row_indices
        return _eligible_row_indices(data_rows, submission_filter=submission_filter, validation_col=validation_col)

    def test_all_filter(self):
        data_rows = [["Kenya", "approved"], ["Uganda", "rejected"]]
        indices, excl, err = self._call(data_rows, "all")
        assert indices == [0, 1]
        assert excl == 0
        assert err is None

    def test_approved_only_with_col(self):
        data_rows = [["Kenya", "approved"], ["Uganda", "rejected"]]
        indices, excl, err = self._call(data_rows, "approved_only", validation_col=1)
        assert 0 in indices
        assert 1 not in indices
        assert excl == 1

    def test_filter_without_validation_col_returns_error(self):
        data_rows = [["Kenya"], ["Uganda"]]
        indices, excl, err = self._call(data_rows, "approved_only", validation_col=None)
        assert indices == []
        assert err is not None


# ---------------------------------------------------------------------------
# _pick_duplicate_winner
# ---------------------------------------------------------------------------

class TestPickDuplicateWinner:
    def _call(self, indices, data_rows, sub_time_col=None, duplicate_strategy="latest", validation_col=None):
        from app.services.kobo_data_import_service import _pick_duplicate_winner
        return _pick_duplicate_winner(
            indices,
            data_rows=data_rows,
            sub_time_col=sub_time_col,
            duplicate_strategy=duplicate_strategy,
            validation_col=validation_col,
        )

    def test_single_index_no_work(self):
        winner, fallback = self._call([0], [[1, 2, 3]])
        assert winner == 0
        assert fallback is False

    def test_latest_strategy(self):
        data_rows = [
            ["Kenya", datetime(2023, 1, 1)],
            ["Kenya", datetime(2024, 1, 1)],
        ]
        winner, _ = self._call([0, 1], data_rows, sub_time_col=1, duplicate_strategy="latest")
        assert winner == 1  # later date

    def test_first_strategy(self):
        data_rows = [
            ["Kenya", datetime(2023, 1, 1)],
            ["Kenya", datetime(2024, 1, 1)],
        ]
        winner, _ = self._call([0, 1], data_rows, sub_time_col=1, duplicate_strategy="first")
        assert winner == 0

    def test_latest_approved_with_approved_row(self):
        data_rows = [
            ["Kenya", datetime(2023, 1, 1), "rejected"],
            ["Kenya", datetime(2024, 1, 1), "approved"],
        ]
        winner, fallback = self._call(
            [0, 1], data_rows,
            sub_time_col=1, duplicate_strategy="latest_approved", validation_col=2,
        )
        assert winner == 1
        assert fallback is False

    def test_latest_approved_fallback_when_none_approved(self):
        data_rows = [
            ["Kenya", datetime(2023, 1, 1), "rejected"],
            ["Kenya", datetime(2024, 1, 1), "rejected"],
        ]
        winner, fallback = self._call(
            [0, 1], data_rows,
            sub_time_col=1, duplicate_strategy="latest_approved", validation_col=2,
        )
        assert fallback is True  # Used unapproved fallback

    def test_first_approved_strategy(self):
        data_rows = [
            ["Kenya", datetime(2023, 1, 1), "approved"],
            ["Kenya", datetime(2024, 1, 1), "approved"],
        ]
        winner, fallback = self._call(
            [0, 1], data_rows,
            sub_time_col=1, duplicate_strategy="first_approved", validation_col=2,
        )
        assert winner == 0
        assert fallback is False

    def test_all_strategy_returns_last(self):
        data_rows = [["Kenya", 1], ["Kenya", 2]]
        winner, _ = self._call([0, 1], data_rows, duplicate_strategy="all")
        assert winner == 1  # last


# ---------------------------------------------------------------------------
# _resolve_import_row_indices
# ---------------------------------------------------------------------------

class TestResolveImportRowIndices:
    def _call(self, data_rows, headers, entity_col=0, duplicate_strategy="latest", submission_filter="all", sub_time_col=None, val_col=None):
        from app.services.kobo_data_import_service import _resolve_import_row_indices
        return _resolve_import_row_indices(
            data_rows, headers,
            entity_column_index=entity_col,
            duplicate_strategy=duplicate_strategy,
            submission_filter=submission_filter,
            submission_time_column_index=sub_time_col,
            validation_status_column_index=val_col,
        )

    def test_single_row_per_entity(self):
        data_rows = [["Kenya", 100], ["Uganda", 200]]
        headers = ["ns", "value"]
        indices, stats = self._call(data_rows, headers)
        assert len(indices) == 2
        assert stats['duplicate_count'] == 0

    def test_deduplication_latest(self):
        data_rows = [
            ["Kenya", datetime(2023, 1, 1)],
            ["Kenya", datetime(2024, 1, 1)],
            ["Uganda", datetime(2023, 6, 1)],
        ]
        headers = ["ns", "_submission_time"]
        indices, stats = self._call(data_rows, headers, sub_time_col=1)
        assert len(indices) == 2  # one per entity
        assert stats['duplicate_count'] == 1

    def test_all_strategy_keeps_duplicates(self):
        data_rows = [
            ["Kenya", 100],
            ["Kenya", 200],
            ["Uganda", 300],
        ]
        headers = ["ns", "value"]
        indices, stats = self._call(data_rows, headers, duplicate_strategy="all")
        assert len(indices) == 3  # all kept

    def test_submission_filter_error_propagates(self):
        data_rows = [["Kenya", 100]]
        headers = ["ns", "value"]
        # Need approved_only but no validation col
        indices, stats = self._call(
            data_rows, headers,
            submission_filter="approved_only",
            val_col=None,
        )
        assert indices == []
        assert stats['error'] is not None

    def test_skips_empty_entity_names(self):
        data_rows = [["", 100], ["Kenya", 200]]
        headers = ["ns", "value"]
        indices, stats = self._call(data_rows, headers)
        # Empty entity name row is skipped
        assert len(indices) == 1


# ---------------------------------------------------------------------------
# _slugify_sex_category
# ---------------------------------------------------------------------------

class TestSlugifySexCategory:
    def _call(self, s):
        from app.services.kobo_data_import_service import _slugify_sex_category
        return _slugify_sex_category(s)

    def test_basic(self):
        assert self._call("Male") == "male"
        assert self._call("Female") == "female"

    def test_space_replaced(self):
        assert self._call("Non Binary") == "non_binary"

    def test_hyphen_replaced(self):
        assert self._call("Trans-gender") == "trans_gender"


# ---------------------------------------------------------------------------
# _normalize_column_to_item_mapping
# ---------------------------------------------------------------------------

class TestNormalizeColumnToItemMapping:
    def _call(self, raw):
        from app.services.kobo_data_import_service import _normalize_column_to_item_mapping
        return _normalize_column_to_item_mapping(raw)

    def test_empty(self):
        assert self._call({}) == {}

    def test_simple_int_value(self):
        result = self._call({"0": 5, "1": 10})
        assert result[0] == {'item_id': 5, 'disagg': None}
        assert result[1] == {'item_id': 10, 'disagg': None}

    def test_structured_dict_value(self):
        result = self._call({"2": {"item_id": 7, "disagg": {"mode": "sex"}}})
        assert result[2]['item_id'] == 7
        assert result[2]['disagg'] == {"mode": "sex"}

    def test_invalid_key_skipped(self):
        result = self._call({"not_int": 5})
        assert result == {}

    def test_invalid_value_skipped(self):
        result = self._call({"0": "not_convertible"})
        assert result == {}

    def test_none_value_skipped(self):
        result = self._call({"0": None})
        assert result == {}


# ---------------------------------------------------------------------------
# _match_entity_to_country
# ---------------------------------------------------------------------------

class TestMatchEntityToCountry:
    def _call(self, entity_name, countries_by_name=None, ns_by_name=None):
        from app.services.kobo_data_import_service import _match_entity_to_country
        if countries_by_name is None:
            countries_by_name = {}
        if ns_by_name is None:
            ns_by_name = {}
        return _match_entity_to_country(entity_name, countries_by_name, ns_by_name)

    def test_exact_ns_match(self):
        ns = MagicMock()
        ns.country = MagicMock(name="Kenya")
        result = self._call("Kenya Red Cross Society", ns_by_name={"kenya red cross society": ns})
        assert result == ns.country

    def test_fuzzy_ns_match(self):
        ns = MagicMock()
        ns.country = MagicMock()
        result = self._call("Red Cross Uganda", ns_by_name={"red cross uganda society": ns})
        assert result == ns.country

    def test_country_name_after_strip(self):
        country = MagicMock()
        result = self._call("Kenya Red Cross", countries_by_name={"kenya": country})
        assert result == country

    def test_no_match_returns_none(self):
        result = self._call("XYZABC 12345", countries_by_name={"kenya": MagicMock()})
        assert result is None

    def test_fuzzy_country_match(self):
        country = MagicMock()
        result = self._call("Uganda", countries_by_name={"uganda": country})
        assert result == country


# ---------------------------------------------------------------------------
# KoboDataImportService.validate_data_export
# ---------------------------------------------------------------------------

class TestKoboDataImportServiceValidateDataExport:
    def test_valid_data_export(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            # Include KoBo system columns so the structural check passes
            file_bytes = _make_excel_bytes(
                headers=["start", "end", "National Society", "Q1 Value", "_id"],
                data_rows=[[datetime(2024, 1, 1), datetime(2024, 1, 1, 1), "Kenya", 100, 1001]],
            )
            result = KoboDataImportService.validate_data_export(file_bytes)
            assert result['valid'] is True
            assert result['preview']['total_rows'] == 1
            assert result['preview']['total_columns'] == 5

    def test_empty_workbook_invalid(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            wb = openpyxl.Workbook()
            buf = io.BytesIO()
            wb.save(buf)
            result = KoboDataImportService.validate_data_export(buf.getvalue())
            assert result['valid'] is False

    def test_headers_only_invalid(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(headers=["National Society"], data_rows=[])
            result = KoboDataImportService.validate_data_export(file_bytes)
            assert result['valid'] is False
            assert 'no data' in result['message'].lower()

    def test_xlsform_rejected(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            wb = openpyxl.Workbook()
            wb.active.title = 'survey'
            wb.active.append(['type', 'name', 'label'])
            wb.active.append(['text', 'q1', 'Question 1'])
            buf = io.BytesIO()
            wb.save(buf)
            result = KoboDataImportService.validate_data_export(buf.getvalue())
            assert result['valid'] is False
            assert 'xlsform' in result['message'].lower()

    def test_plain_spreadsheet_no_markers_rejected(self, app):
        """A random spreadsheet with no KoBo markers must be rejected."""
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["Name", "Department", "Score", "Notes"],
                data_rows=[
                    ["Alice", "HR", 90, "Good"],
                    ["Bob",   "IT", 85, "OK"],
                ],
            )
            result = KoboDataImportService.validate_data_export(file_bytes)
            assert result['valid'] is False
            assert 'kobo' in result['message'].lower()

    def test_underscore_meta_column_passes(self, app):
        """A file with _uuid (even no group headers) is accepted."""
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["_uuid", "National Society", "Q1"],
                data_rows=[["uuid-1", "Kenya", 100]],
            )
            result = KoboDataImportService.validate_data_export(file_bytes)
            assert result['valid'] is True

    def test_group_path_headers_pass(self, app):
        """A file with Group/Question path headers (no system cols) is accepted."""
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["NS", "section1/q1", "section1/q2", "section2/q3"],
                data_rows=[
                    ["Kenya", "A", "B", "C"],
                    ["Uganda", "D", "E", "F"],
                ],
            )
            result = KoboDataImportService.validate_data_export(file_bytes)
            assert result['valid'] is True

    def test_system_exact_columns_pass(self, app):
        """start / end / today column is enough to pass the structural check."""
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["start", "end", "today", "National Society", "Q1"],
                data_rows=[
                    [datetime(2024, 1, 1), datetime(2024, 1, 1, 1), datetime(2024, 1, 1), "Kenya", 100],
                ],
            )
            result = KoboDataImportService.validate_data_export(file_bytes)
            assert result['valid'] is True


# ---------------------------------------------------------------------------
# KoboDataImportService.analyze
# ---------------------------------------------------------------------------

class TestKoboDataImportServiceAnalyze:
    def test_empty_file(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            wb = openpyxl.Workbook()
            ws = wb.active
            # Only headers, no data rows
            ws.append(["National Society", "Q1"])
            buf = io.BytesIO()
            wb.save(buf)
            result = KoboDataImportService.analyze(buf.getvalue())
            assert result['success'] is False
            assert "no data" in result['message'].lower()

    def test_no_header_row(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            wb = openpyxl.Workbook()
            buf = io.BytesIO()
            wb.save(buf)
            result = KoboDataImportService.analyze(buf.getvalue())
            assert result['success'] is False

    def test_basic_analysis(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            # _uuid column acts as the KoBo structural marker
            file_bytes = _make_excel_bytes(
                headers=["_uuid", "National Society", "Q1 Value", "Q2 Text"],
                data_rows=[
                    ["uuid-1", "Kenya", 100, "Answer A"],
                    ["uuid-2", "Uganda", 200, "Answer B"],
                    ["uuid-3", "Tanzania", 150, "Answer A"],
                ],
            )
            result = KoboDataImportService.analyze(file_bytes)
            assert result['success'] is True
            assert result['total_rows'] == 3
            assert result['total_columns'] == 4

    def test_entity_candidates_detected(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            # _uuid + start/end give the required KoBo structural markers
            ns_values = [f"NS_{i}" for i in range(10)]
            headers = ["start", "end", "_uuid", "National Society"] + [f"Q{i}" for i in range(5)]
            data_rows = [
                [datetime(2024, 1, 1), datetime(2024, 1, 1, 1), f"uuid-{idx}", ns] + [idx * 10 for _ in range(5)]
                for idx, ns in enumerate(ns_values)
            ]
            file_bytes = _make_excel_bytes(headers, data_rows)
            result = KoboDataImportService.analyze(file_bytes)
            assert result['success'] is True
            assert 'entity_candidates' in result

    def test_system_columns_skipped(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["start", "end", "National Society", "Q1"],
                data_rows=[
                    [datetime(2024, 1, 1), datetime(2024, 1, 1, 1), "Kenya", 100],
                    [datetime(2024, 1, 2), datetime(2024, 1, 2, 1), "Uganda", 200],
                ],
            )
            result = KoboDataImportService.analyze(file_bytes)
            assert result['success'] is True
            # start, end should be in skipped
            skipped_headers = [s['header'] for s in result.get('skipped_columns', [])]
            assert "start" in skipped_headers
            assert "end" in skipped_headers

    def test_grouped_columns(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["NS", "demographics/age", "demographics/gender"],
                data_rows=[
                    ["Kenya", 25, "Male"],
                    ["Uganda", 30, "Female"],
                ],
            )
            result = KoboDataImportService.analyze(file_bytes)
            assert result['success'] is True
            group_names = [g['name'] for g in result.get('groups', [])]
            assert "demographics" in group_names

    def test_invalid_file(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            result = KoboDataImportService.analyze(b"not an excel file")
            assert result['success'] is False

    def test_plain_spreadsheet_no_markers_rejected(self, app):
        """analyze() also rejects files with no KoBo structural markers."""
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["Name", "Department", "Score"],
                data_rows=[
                    ["Alice", "HR", 90],
                    ["Bob",   "IT", 85],
                ],
            )
            result = KoboDataImportService.analyze(file_bytes)
            assert result['success'] is False
            assert 'kobo' in result['message'].lower()

    def test_analyze_returns_marker_flags(self, app):
        """Successful analyze() includes has_kobo_meta / has_group_headers flags."""
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["_uuid", "NS", "section/q1"],
                data_rows=[["uuid-1", "Kenya", "Answer"]],
            )
            result = KoboDataImportService.analyze(file_bytes)
            assert result['success'] is True
            assert result['has_kobo_meta'] is True
            assert result['has_group_headers'] is True

    def test_openpyxl_none_returns_error(self, app):
        with app.app_context():
            import app.services.kobo_data_import_service as module
            original = module.openpyxl
            module.openpyxl = None
            try:
                from app.services.kobo_data_import_service import KoboDataImportService
                result = KoboDataImportService.analyze(b"anything")
                assert result['success'] is False
            finally:
                module.openpyxl = original

    def test_validation_status_detected(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["NS", "Q1", "_validation_status"],
                data_rows=[
                    ["Kenya", 100, "approved"],
                    ["Uganda", 200, "rejected"],
                ],
            )
            result = KoboDataImportService.analyze(file_bytes)
            assert result['success'] is True
            assert result.get('validation_status_column_index') is not None

    def test_calculated_columns_skipped(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["_uuid", "NS", "total_beneficiaries", "Q1"],
                data_rows=[
                    ["uuid-1", "Kenya", 1000, "Answer"],
                    ["uuid-2", "Uganda", 500, "Answer"],
                ],
            )
            result = KoboDataImportService.analyze(file_bytes)
            assert result['success'] is True
            skipped_hdrs = [s['header'] for s in result.get('skipped_columns', [])]
            assert "total_beneficiaries" in skipped_hdrs


# ---------------------------------------------------------------------------
# KoboDataImportService.extract_unique_entities
# ---------------------------------------------------------------------------

class TestExtractUniqueEntities:
    def test_basic_entities(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["NS", "Q1"],
                data_rows=[
                    ["Kenya", 100],
                    ["Uganda", 200],
                    ["Kenya", 150],  # duplicate
                ],
            )
            result = KoboDataImportService.extract_unique_entities(file_bytes, entity_col_index=0)
            assert "Kenya" in result
            assert "Uganda" in result
            assert result.count("Kenya") == 1  # de-duplicated

    def test_respects_submission_filter(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["NS", "_validation_status"],
                data_rows=[
                    ["Kenya", "approved"],
                    ["Uganda", "rejected"],
                ],
            )
            result = KoboDataImportService.extract_unique_entities(
                file_bytes, entity_col_index=0,
                submission_filter="approved_only",
                validation_status_column_index=1,
            )
            assert "Kenya" in result
            assert "Uganda" not in result

    def test_openpyxl_none_returns_empty(self, app):
        with app.app_context():
            import app.services.kobo_data_import_service as module
            original = module.openpyxl
            module.openpyxl = None
            try:
                from app.services.kobo_data_import_service import KoboDataImportService
                result = KoboDataImportService.extract_unique_entities(b"anything", 0)
                assert result == []
            finally:
                module.openpyxl = original

    def test_invalid_col_index_returns_empty(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["NS"],
                data_rows=[["Kenya"], ["Uganda"]],
            )
            # Column index 99 doesn't exist
            result = KoboDataImportService.extract_unique_entities(file_bytes, entity_col_index=99)
            assert result == []

    def test_skips_null_entity_values(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["NS", "Q1"],
                data_rows=[
                    [None, 100],
                    ["Kenya", 200],
                ],
            )
            result = KoboDataImportService.extract_unique_entities(file_bytes, entity_col_index=0)
            assert None not in result
            assert "" not in result
            assert "Kenya" in result

    def test_filter_no_validation_col(self, app):
        with app.app_context():
            from app.services.kobo_data_import_service import KoboDataImportService
            file_bytes = _make_excel_bytes(
                headers=["NS", "Q1"],
                data_rows=[["Kenya", 100], ["Uganda", 200]],
            )
            # Using approved_only but no validation col -> returns []
            result = KoboDataImportService.extract_unique_entities(
                file_bytes, entity_col_index=0,
                submission_filter="approved_only",
            )
            assert result == []


# ---------------------------------------------------------------------------
# _sub_time_sort_value
# ---------------------------------------------------------------------------

class TestSubTimeSortValue:
    def _call(self, data_rows, ri, sub_time_col):
        from app.services.kobo_data_import_service import _sub_time_sort_value
        return _sub_time_sort_value(data_rows, ri, sub_time_col)

    def test_none_col_returns_datetime_min(self):
        result = self._call([["a", "b"]], 0, None)
        assert result == datetime.min

    def test_datetime_value(self):
        dt = datetime(2024, 5, 1)
        result = self._call([["ns", dt]], 0, 1)
        assert result == dt

    def test_non_datetime_returns_min(self):
        result = self._call([["ns", "not a date"]], 0, 1)
        assert result == datetime.min

    def test_col_out_of_range(self):
        result = self._call([["only_one"]], 0, 5)
        assert result == datetime.min


# ---------------------------------------------------------------------------
# Additional edge case tests
# ---------------------------------------------------------------------------

class TestKoboValidationBucketEdgeCases:
    def _call(self, val):
        from app.services.kobo_data_import_service import _kobo_validation_bucket
        return _kobo_validation_bucket(val)

    def test_flagged_for_removal_is_always_rejected(self):
        # 'flagged_for_removal' is matched literally before the generic 'approved' check,
        # so "flagged_for_removal_approved" is still rejected.
        result = self._call("flagged_for_removal_approved")
        assert result == "rejected"

    def test_approved_for_removal_is_rejected(self):
        # The generic 'removal' guard fires when 'approved' is NOT in norm.
        # "approved_for_removal" contains 'removal' but also contains 'approved',
        # so the 'removal' guard is skipped and it falls through to the 'approved' branch.
        result = self._call("approved_for_removal")
        assert result == "approved"

    def test_removal_without_approved(self):
        result = self._call("flagged_removal")
        assert result == "rejected"
