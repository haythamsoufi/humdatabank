"""
Comprehensive tests for app/services/excel_service.py
"""
from __future__ import annotations

import io
import json
from typing import Any, Dict
from unittest.mock import MagicMock, patch, PropertyMock

import openpyxl
import pytest


def _make_workbook_with_data(rows):
    """Create an openpyxl workbook with given rows in sheet 1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    return wb


def _workbook_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# _calculate_label_rows_needed
# ---------------------------------------------------------------------------

class TestCalculateLabelRowsNeeded:
    def _call(self, label_text, max_chars_per_row=40, min_rows=1, max_rows=5):
        from app.services.imports.excel_service import ExcelService
        return ExcelService._calculate_label_rows_needed(
            label_text, max_chars_per_row=max_chars_per_row, min_rows=min_rows, max_rows=max_rows
        )

    def test_empty_label_returns_min_rows(self):
        assert self._call("") == 1
        assert self._call(None) == 1

    def test_short_label(self):
        result = self._call("Short label")
        assert result == 1

    def test_long_label_needs_more_rows(self):
        long_label = "A" * 100
        result = self._call(long_label)
        assert result > 1

    def test_max_rows_capped(self):
        very_long = "A" * 1000
        result = self._call(very_long, max_rows=3)
        assert result <= 3

    def test_min_rows_respected(self):
        result = self._call("Tiny", min_rows=2)
        assert result >= 2

    def test_near_threshold_gets_two_rows(self):
        # 36 chars > 35 threshold triggers bump to 2
        label = "A" * 36
        result = self._call(label)
        assert result >= 2


# ---------------------------------------------------------------------------
# extract_field_values
# ---------------------------------------------------------------------------

class TestExtractFieldValues:
    def _call(self, workbook):
        from app.services.imports.excel_service import ExcelService
        return ExcelService.extract_field_values(workbook)

    def test_new_form_format_basic(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Entry"
        # item_id in col A, label in B+C, value in D
        ws.append([1, "Label B", "Label C", "42"])
        ws.append([2, "Label B2", "Label C2", "Hello"])
        result = self._call(wb)
        assert 1 in result
        assert result[1]['value'] == "42"
        assert 2 in result
        assert result[2]['value'] == "Hello"

    def test_old_table_format(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Entry"
        # Old format: first cell is "item_id"
        ws.append(["item_id", "Section", "Label", "Value"])
        ws.append([1, "Section A", "Question 1", "100"])
        ws.append([2, "Section A", "Question 2", "200"])
        result = self._call(wb)
        assert 1 in result
        assert result[1]['value'] == "100"

    def test_disaggregation_data_parsed(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Entry"
        values_json = json.dumps({"male": 50, "female": 60})
        ws.append([5, "Label B", "Label C", None, "Mode: sex", values_json])
        result = self._call(wb)
        assert 5 in result
        assert result[5]['disagg_data'] is not None
        assert result[5]['disagg_data']['mode'] == 'sex'

    def test_none_item_id_skipped(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Entry"
        ws.append([None, "Label", "C", "value"])  # no item_id
        ws.append([1, "Label", "C", "50"])
        result = self._call(wb)
        assert None not in result
        assert 1 in result

    def test_invalid_item_id_skipped(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Entry"
        ws.append(["not_an_int", "Label", "C", "value"])
        ws.append([1, "Label", "C", "50"])
        result = self._call(wb)
        assert len([k for k in result if not isinstance(k, int)]) == 0

    def test_empty_value_becomes_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Entry"
        ws.append([3, "Label B", "Label C", ""])
        result = self._call(wb)
        assert 3 in result
        assert result[3]['value'] is None

    def test_null_string_becomes_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Entry"
        ws.append([4, "Label", "C", "null"])
        result = self._call(wb)
        assert result[4]['value'] is None

    def test_mode_in_value_cell_skipped(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Entry"
        ws.append([7, "Label", "C", "Mode: sex"])  # misplaced data
        result = self._call(wb)
        assert 7 not in result

    def test_empty_sheet_raises(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Don't add any rows
        with pytest.raises(ValueError, match="valid data sheet"):
            self._call(wb)

    def test_no_valid_item_id_rows_raises(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, "A", "B", "C"])  # no valid item_id
        with pytest.raises(ValueError, match="no valid data rows"):
            self._call(wb)

    def test_multiple_sheets(self):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Page 1"
        ws1.append([1, "Label", "C", "10"])
        ws2 = wb.create_sheet("Page 2")
        ws2.append([2, "Label", "C", "20"])
        result = self._call(wb)
        assert 1 in result
        assert 2 in result

    def test_row_with_fewer_than_3_cols_skipped(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Entry"
        ws.append([1, "Only two"])  # too few columns
        ws.append([2, "Label", "C", "50"])
        result = self._call(wb)
        assert 1 not in result
        assert 2 in result

    def test_disagg_json_parse_failure_skips_disagg(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Entry"
        ws.append([8, "Label", "C", None, "Mode: sex", "INVALID JSON {{{"])
        result = self._call(wb)
        assert 8 in result
        assert result[8]['disagg_data'] is None


# ---------------------------------------------------------------------------
# load_workbook
# ---------------------------------------------------------------------------

class TestLoadWorkbook:
    def test_load_valid_workbook(self):
        from app.services.imports.excel_service import ExcelService
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([1, "test"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        mock_file = MagicMock()
        mock_file.read.return_value = buf.read()
        mock_file.stream.seek = MagicMock()

        result = ExcelService.load_workbook(mock_file)
        assert result is not None
        assert len(result.sheetnames) >= 1


# ---------------------------------------------------------------------------
# _bulk_save_fields_with_disagg (mocked DB)
# ---------------------------------------------------------------------------

class TestBulkSaveFieldsWithDisagg:
    def _make_aes(self):
        aes = MagicMock()
        aes.id = 1
        return aes

    def test_simple_value_saved(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            aes = self._make_aes()
            field_data = {
                1: {'value': '42', 'disagg_data': None}
            }
            mock_entry = MagicMock()
            mock_entry.set_simple_value = MagicMock()

            with patch("app.services.imports.excel_service.FormData") as MockFormData:
                with patch("app.services.imports.excel_service.PublicSubmission") as MockPS:
                    MockPS.__instancecheck__ = lambda cls, inst: False
                    MockFormData.query.filter_by.return_value.first.return_value = mock_entry
                    with patch("app.services.imports.excel_service.db") as mock_db:
                        result = ExcelService._bulk_save_fields_with_disagg(aes, field_data)
                        assert result['success'] is True
                        assert result['updated_count'] >= 1

    def test_disagg_data_saved(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            aes = self._make_aes()
            field_data = {
                1: {
                    'value': None,
                    'disagg_data': {'mode': 'sex', 'values': {'male': 50, 'female': 60}}
                }
            }
            mock_entry = MagicMock()
            mock_entry.set_disaggregated_data = MagicMock()

            with patch("app.services.imports.excel_service.FormData") as MockFormData:
                with patch("app.services.imports.excel_service.PublicSubmission") as MockPS:
                    MockPS.__instancecheck__ = lambda cls, inst: False
                    MockFormData.query.filter_by.return_value.first.return_value = mock_entry
                    with patch("app.services.imports.excel_service.db") as mock_db:
                        result = ExcelService._bulk_save_fields_with_disagg(aes, field_data)
                        assert result['success'] is True

    def test_new_entry_created_when_missing(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            aes = self._make_aes()
            field_data = {2: {'value': '100', 'disagg_data': None}}
            mock_entry = MagicMock()

            with patch("app.services.imports.excel_service.FormData") as MockFormData:
                with patch("app.services.imports.excel_service.PublicSubmission") as MockPS:
                    MockPS.__instancecheck__ = lambda cls, inst: False
                    MockFormData.query.filter_by.return_value.first.return_value = None
                    MockFormData.return_value = mock_entry
                    with patch("app.services.imports.excel_service.db") as mock_db:
                        result = ExcelService._bulk_save_fields_with_disagg(aes, field_data)
                        mock_db.session.add.assert_called()

    def test_rollback_on_exception(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            aes = self._make_aes()
            field_data = {1: {'value': '42', 'disagg_data': None}}

            with patch("app.services.imports.excel_service.FormData") as MockFormData:
                with patch("app.services.imports.excel_service.PublicSubmission") as MockPS:
                    MockPS.__instancecheck__ = lambda cls, inst: False
                    MockFormData.query.filter_by.return_value.first.side_effect = RuntimeError("DB error")
                    with patch("app.services.imports.excel_service.db") as mock_db:
                        mock_db.session.commit.side_effect = RuntimeError("commit failed")
                        result = ExcelService._bulk_save_fields_with_disagg(aes, field_data)
                        assert result['success'] is False

    def test_invalid_disagg_data_structure(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            aes = self._make_aes()
            # disagg_data without mode or values
            field_data = {3: {'value': None, 'disagg_data': {'mode': None, 'values': {}}}}
            mock_entry = MagicMock()

            with patch("app.services.imports.excel_service.FormData") as MockFormData:
                with patch("app.services.imports.excel_service.PublicSubmission") as MockPS:
                    MockPS.__instancecheck__ = lambda cls, inst: False
                    MockFormData.query.filter_by.return_value.first.return_value = mock_entry
                    with patch("app.services.imports.excel_service.db") as mock_db:
                        result = ExcelService._bulk_save_fields_with_disagg(aes, field_data)
                        assert len(result['errors']) > 0

    def test_empty_value_clears_entry(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            aes = self._make_aes()
            field_data = {5: {'value': None, 'disagg_data': None}}
            mock_entry = MagicMock()

            with patch("app.services.imports.excel_service.FormData") as MockFormData:
                with patch("app.services.imports.excel_service.PublicSubmission") as MockPS:
                    MockPS.__instancecheck__ = lambda cls, inst: False
                    MockFormData.query.filter_by.return_value.first.return_value = mock_entry
                    with patch("app.services.imports.excel_service.db") as mock_db:
                        result = ExcelService._bulk_save_fields_with_disagg(aes, field_data)
                        assert result['success'] is True
                        mock_entry.set_simple_value.assert_called_with(None)

    def test_public_submission_uses_different_filter(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            from app.models.assignments import PublicSubmission

            # Create a real PublicSubmission mock that passes isinstance check
            ps = MagicMock(spec=PublicSubmission)
            ps.id = 99
            field_data = {1: {'value': '55', 'disagg_data': None}}
            mock_entry = MagicMock()

            with patch("app.services.imports.excel_service.FormData") as MockFormData:
                with patch("app.services.imports.excel_service.db") as mock_db:
                    MockFormData.query.filter_by.return_value.first.return_value = mock_entry
                    result = ExcelService._bulk_save_fields_with_disagg(ps, field_data)
                    assert result['success'] is True


# ---------------------------------------------------------------------------
# import_assignment_data
# ---------------------------------------------------------------------------

class TestImportAssignmentData:
    def test_calls_extract_and_bulk_save(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            aes = MagicMock()
            aes.id = 1

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([1, "Label", "C", "42"])

            with patch.object(ExcelService, 'extract_field_values', return_value={1: {'value': '42', 'disagg_data': None}}) as mock_extract:
                with patch.object(ExcelService, '_bulk_save_fields_with_disagg', return_value={'success': True, 'updated_count': 1, 'errors': []}) as mock_save:
                    result = ExcelService.import_assignment_data(aes, wb)
                    assert result['success'] is True
                    mock_extract.assert_called_once_with(wb)
                    mock_save.assert_called_once()


# ---------------------------------------------------------------------------
# _write_section (basic smoke test)
# ---------------------------------------------------------------------------

class TestWriteSection:
    def test_basic_section_written(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active

            # Mocks
            section = MagicMock()
            section.display_name = "Test Section"
            section.name = "Test Section"
            section.id = 1

            # Create mock form item
            item = MagicMock()
            item.is_document_field = False
            item.order = 1
            item.label = "Question 1"
            item.display_label = None
            item.unit = None
            item.item_type = "question"
            item.id = 1

            section.form_items.order_by.return_value.all.return_value = [item]

            entries_map = {}

            with patch("app.services.imports.excel_service.get_unified_form_item_id", return_value=1):
                section_fill = PatternFill(start_color="FFCC0000", end_color="FFCC0000", fill_type="solid")
                section_font = Font(name="Arial", size=14, bold=True, color="FFFFFFFF")
                label_font = Font(name="Arial", size=11, bold=True)
                value_fill = PatternFill(start_color="FFFFF9E6", end_color="FFFFF9E6", fill_type="solid")
                value_font = Font(name="Arial", size=11)
                disagg_label_font = Font(name="Arial", size=10, bold=True)
                disagg_value_fill = PatternFill(start_color="FFE6F3FF", end_color="FFE6F3FF", fill_type="solid")
                left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

                end_row = ExcelService._write_section(
                    ws, section, entries_map, value_fill, section_fill, section_font,
                    label_font, value_font, disagg_label_font, disagg_value_fill,
                    left_align, start_row=1,
                )
                assert end_row > 1

    def test_section_with_disagg_data(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active

            section = MagicMock()
            section.display_name = "Section"
            section.name = "Section"
            section.id = 1

            item = MagicMock()
            item.is_document_field = False
            item.order = 1
            item.label = "Sex Disagg"
            item.display_label = None
            item.unit = None
            item.item_type = "question"
            item.id = 2

            section.form_items.order_by.return_value.all.return_value = [item]

            # Entry with disagg data
            mock_entry = MagicMock()
            mock_entry.disagg_data = {'mode': 'sex', 'values': {'male': 50, 'female': 60}}
            mock_entry.value = None
            entries_map = {2: mock_entry}

            with patch("app.services.imports.excel_service.get_unified_form_item_id", return_value=2):
                section_fill = PatternFill(start_color="FFCC0000", end_color="FFCC0000", fill_type="solid")
                section_font = Font(name="Arial", size=14, bold=True, color="FFFFFFFF")
                label_font = Font(name="Arial", size=11, bold=True)
                value_fill = PatternFill(start_color="FFFFF9E6", end_color="FFFFF9E6", fill_type="solid")
                value_font = Font(name="Arial", size=11)
                disagg_label_font = Font(name="Arial", size=10, bold=True)
                disagg_value_fill = PatternFill(start_color="FFE6F3FF", end_color="FFE6F3FF", fill_type="solid")
                left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

                end_row = ExcelService._write_section(
                    ws, section, entries_map, value_fill, section_fill, section_font,
                    label_font, value_font, disagg_label_font, disagg_value_fill,
                    left_align, start_row=1,
                )
                assert end_row > 1

    def test_section_with_matrix_data(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active

            section = MagicMock()
            section.display_name = "Section"
            section.name = "Section"
            section.id = 1

            item = MagicMock()
            item.is_document_field = False
            item.order = 1
            item.label = "Matrix Q"
            item.display_label = None
            item.unit = None
            item.item_type = "matrix"
            item.id = 3

            section.form_items.order_by.return_value.all.return_value = [item]

            mock_entry = MagicMock()
            mock_entry.disagg_data = {'row1': {'col1': 10, 'col2': 20}}
            mock_entry.value = None
            entries_map = {3: mock_entry}

            with patch("app.services.imports.excel_service.get_unified_form_item_id", return_value=3):
                section_fill = PatternFill(start_color="FFCC0000", end_color="FFCC0000", fill_type="solid")
                section_font = Font(name="Arial", size=14, bold=True, color="FFFFFFFF")
                label_font = Font(name="Arial", size=11, bold=True)
                value_fill = PatternFill(start_color="FFFFF9E6", end_color="FFFFF9E6", fill_type="solid")
                value_font = Font(name="Arial", size=11)
                disagg_label_font = Font(name="Arial", size=10, bold=True)
                disagg_value_fill = PatternFill(start_color="FFE6F3FF", end_color="FFE6F3FF", fill_type="solid")
                left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

                end_row = ExcelService._write_section(
                    ws, section, entries_map, value_fill, section_fill, section_font,
                    label_font, value_font, disagg_label_font, disagg_value_fill,
                    left_align, start_row=1,
                )
                assert end_row > 1

    def test_document_field_skipped(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active

            section = MagicMock()
            section.display_name = "Section"
            section.name = "Section"
            section.id = 1

            item = MagicMock()
            item.is_document_field = True  # Should be skipped
            item.order = 1
            item.label = "Upload"
            item.id = 4

            section.form_items.order_by.return_value.all.return_value = [item]
            entries_map = {}

            section_fill = PatternFill(start_color="FFCC0000", end_color="FFCC0000", fill_type="solid")
            section_font = Font(name="Arial", size=14, bold=True, color="FFFFFFFF")
            label_font = Font(name="Arial", size=11, bold=True)
            value_fill = PatternFill(start_color="FFFFF9E6", end_color="FFFFF9E6", fill_type="solid")
            value_font = Font(name="Arial", size=11)
            disagg_label_font = Font(name="Arial", size=10, bold=True)
            disagg_value_fill = PatternFill(start_color="FFE6F3FF", end_color="FFE6F3FF", fill_type="solid")
            left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

            with patch("app.services.imports.excel_service.get_unified_form_item_id", return_value=None):
                end_row = ExcelService._write_section(
                    ws, section, entries_map, value_fill, section_fill, section_font,
                    label_font, value_font, disagg_label_font, disagg_value_fill,
                    left_align, start_row=1,
                )
                assert end_row > 0

    def test_section_with_variable_resolution(self, app):
        with app.app_context():
            from app.services.imports.excel_service import ExcelService
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active

            section = MagicMock()
            section.display_name = "{{country}} Section"
            section.name = "Country Section"
            section.id = 1
            section.form_items.order_by.return_value.all.return_value = []

            section_fill = PatternFill(start_color="FFCC0000", end_color="FFCC0000", fill_type="solid")
            section_font = Font(name="Arial", size=14, bold=True, color="FFFFFFFF")
            label_font = Font(name="Arial", size=11, bold=True)
            value_fill = PatternFill(start_color="FFFFF9E6", end_color="FFFFF9E6", fill_type="solid")
            value_font = Font(name="Arial", size=11)
            disagg_label_font = Font(name="Arial", size=10, bold=True)
            disagg_value_fill = PatternFill(start_color="FFE6F3FF", end_color="FFE6F3FF", fill_type="solid")
            left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

            with patch("app.services.imports.excel_service.VariableResolutionService") as mock_vrs:
                mock_vrs.replace_variables_in_text.return_value = "Kenya Section"
                end_row = ExcelService._write_section(
                    ws, section, {}, value_fill, section_fill, section_font,
                    label_font, value_font, disagg_label_font, disagg_value_fill,
                    left_align, start_row=1,
                    resolved_variables={'country': 'Kenya'},
                    variable_configs={},
                )
                assert end_row > 0
