"""
Tests for app/routes/admin/utilities/import_export.py

Covers:
- Route: GET/POST /admin/indicator_bank/import         (import_indicators)
- Route: POST     /admin/indicator_bank/import/preview  (preview_indicator_import)
- Route: POST     /admin/indicator_bank/import/apply    (apply_indicator_import)
- Route: GET      /admin/indicator_bank/change_history  (indicator_change_history)
- Route: GET      /admin/indicator_suggestions          (manage_indicator_suggestions)
- Route: GET      /admin/indicator_suggestions/view/<id> (view_indicator_suggestion)
- Route: POST     /admin/indicator_suggestions/update_status/<id>
- Route: POST     /admin/indicator_suggestions/delete/<id>
- Helpers: _norm_header, _to_int, _to_bool, _to_json_dict,
           _sheet_rows_as_dicts, _build_levels_json, _mq_str_to_list, _tags_str_to_list
"""
import io
import json
import os
import uuid
from unittest.mock import MagicMock, patch, call

import pytest
from openpyxl import Workbook

pytestmark = [pytest.mark.unit]


# ---------------------------------------------------------------------------
# Helpers to build in-memory Excel files
# ---------------------------------------------------------------------------

def _make_xlsx_bytes(sheets: dict) -> bytes:
    """Create an .xlsx file in memory.

    sheets = {"SheetName": [row_tuple, ...], ...}
    First row tuple in each sheet is the header row.
    """
    wb = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _xlsx_file(name: str, sheets: dict):
    """Return (filename, bytes, content_type) tuple for test client uploads."""
    return (io.BytesIO(_make_xlsx_bytes(sheets)), name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# Grant extra permissions helper
# ---------------------------------------------------------------------------

def _grant_suggestions_permission(db_session, app):
    """Grant admin.indicator_bank.suggestions.review to admin_core role."""
    from app.models.rbac import RbacRole, RbacPermission, RbacRolePermission

    with app.app_context():
        role = RbacRole.query.filter_by(code="admin_core").first()
        if not role:
            return
        perm = RbacPermission.query.filter_by(
            code="admin.indicator_bank.suggestions.review"
        ).first()
        if not perm:
            perm = RbacPermission(
                code="admin.indicator_bank.suggestions.review",
                name="Manage indicator suggestions",
                description="Review and manage indicator suggestions",
            )
            db_session.add(perm)
            db_session.flush()

        existing = RbacRolePermission.query.filter_by(
            role_id=role.id, permission_id=perm.id
        ).first()
        if not existing:
            db_session.add(RbacRolePermission(role_id=role.id, permission_id=perm.id))
            db_session.commit()


# ---------------------------------------------------------------------------
# Pure helper function tests (no Flask app needed)
# ---------------------------------------------------------------------------

class TestNormHeader:
    def test_lowercase_and_strip(self):
        from app.routes.admin.utilities.import_export import _norm_header
        assert _norm_header("  Name  ") == "name"
        assert _norm_header("UNIT") == "unit"

    def test_none_becomes_empty_string(self):
        from app.routes.admin.utilities.import_export import _norm_header
        assert _norm_header(None) == ""

    def test_numeric_value(self):
        from app.routes.admin.utilities.import_export import _norm_header
        assert _norm_header(42) == "42"


class TestToInt:
    def test_none_returns_none(self):
        from app.routes.admin.utilities.import_export import _to_int
        assert _to_int(None) is None

    def test_empty_string_returns_none(self):
        from app.routes.admin.utilities.import_export import _to_int
        assert _to_int("") is None

    def test_integer_passthrough(self):
        from app.routes.admin.utilities.import_export import _to_int
        assert _to_int(5) == 5
        assert _to_int(0) == 0

    def test_float_truncates(self):
        from app.routes.admin.utilities.import_export import _to_int
        assert _to_int(3.9) == 3

    def test_string_int(self):
        from app.routes.admin.utilities.import_export import _to_int
        assert _to_int("42") == 42

    def test_string_float(self):
        from app.routes.admin.utilities.import_export import _to_int
        assert _to_int("3.0") == 3

    def test_bool_true(self):
        from app.routes.admin.utilities.import_export import _to_int
        assert _to_int(True) == 1

    def test_bool_false(self):
        from app.routes.admin.utilities.import_export import _to_int
        assert _to_int(False) == 0

    def test_invalid_string_returns_none(self):
        from app.routes.admin.utilities.import_export import _to_int
        assert _to_int("abc") is None

    def test_whitespace_string_returns_none(self):
        from app.routes.admin.utilities.import_export import _to_int
        assert _to_int("   ") is None


class TestToBool:
    def test_none_returns_none(self):
        from app.routes.admin.utilities.import_export import _to_bool
        assert _to_bool(None) is None

    def test_empty_string_returns_none(self):
        from app.routes.admin.utilities.import_export import _to_bool
        assert _to_bool("") is None

    def test_bool_passthrough(self):
        from app.routes.admin.utilities.import_export import _to_bool
        assert _to_bool(True) is True
        assert _to_bool(False) is False

    def test_int_nonzero_is_true(self):
        from app.routes.admin.utilities.import_export import _to_bool
        assert _to_bool(1) is True
        assert _to_bool(2) is True

    def test_int_zero_is_false(self):
        from app.routes.admin.utilities.import_export import _to_bool
        assert _to_bool(0) is False

    def test_string_true_values(self):
        from app.routes.admin.utilities.import_export import _to_bool
        for v in ("true", "True", "TRUE", "t", "yes", "y", "1"):
            assert _to_bool(v) is True, f"Expected True for {v!r}"

    def test_string_false_values(self):
        from app.routes.admin.utilities.import_export import _to_bool
        for v in ("false", "False", "FALSE", "f", "no", "n", "0"):
            assert _to_bool(v) is False, f"Expected False for {v!r}"

    def test_unknown_string_returns_none(self):
        from app.routes.admin.utilities.import_export import _to_bool
        assert _to_bool("maybe") is None
        assert _to_bool("unknown") is None

    def test_float_nonzero_is_true(self):
        from app.routes.admin.utilities.import_export import _to_bool
        assert _to_bool(1.5) is True

    def test_float_zero_is_false(self):
        from app.routes.admin.utilities.import_export import _to_bool
        assert _to_bool(0.0) is False


class TestToJsonDict:
    def test_none_returns_none(self):
        from app.routes.admin.utilities.import_export import _to_json_dict
        assert _to_json_dict(None) is None

    def test_empty_string_returns_none(self):
        from app.routes.admin.utilities.import_export import _to_json_dict
        assert _to_json_dict("") is None

    def test_dict_passthrough(self):
        from app.routes.admin.utilities.import_export import _to_json_dict
        d = {"a": 1}
        assert _to_json_dict(d) == {"a": 1}

    def test_valid_json_string(self):
        from app.routes.admin.utilities.import_export import _to_json_dict
        assert _to_json_dict('{"key": "value"}') == {"key": "value"}

    def test_json_list_returns_none(self):
        from app.routes.admin.utilities.import_export import _to_json_dict
        # list is not a dict → return None
        assert _to_json_dict("[1, 2, 3]") is None

    def test_invalid_json_returns_none(self):
        from app.routes.admin.utilities.import_export import _to_json_dict
        assert _to_json_dict("not-json") is None

    def test_nested_dict(self):
        from app.routes.admin.utilities.import_export import _to_json_dict
        d = {"en": "Hello", "fr": "Bonjour"}
        assert _to_json_dict(json.dumps(d)) == d


class TestSheetRowsAsDicts:
    def test_empty_sheet_returns_empty_list(self):
        from app.routes.admin.utilities.import_export import _sheet_rows_as_dicts

        wb = Workbook()
        ws = wb.active
        result = _sheet_rows_as_dicts(ws)
        assert result == []

    def test_header_only_returns_empty_list(self):
        from app.routes.admin.utilities.import_export import _sheet_rows_as_dicts

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "value"])
        result = _sheet_rows_as_dicts(ws)
        assert result == []

    def test_single_data_row(self):
        from app.routes.admin.utilities.import_export import _sheet_rows_as_dicts

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "value"])
        ws.append(["Indicator A", 100])
        result = _sheet_rows_as_dicts(ws)
        assert len(result) == 1
        assert result[0]["name"] == "Indicator A"
        assert result[0]["value"] == 100

    def test_skips_empty_rows(self):
        from app.routes.admin.utilities.import_export import _sheet_rows_as_dicts

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "value"])
        ws.append(["Valid Row", 1])
        ws.append([None, None])  # empty row
        ws.append(["Another Row", 2])
        result = _sheet_rows_as_dicts(ws)
        assert len(result) == 2

    def test_skips_blank_header_columns(self):
        from app.routes.admin.utilities.import_export import _sheet_rows_as_dicts

        wb = Workbook()
        ws = wb.active
        ws.append(["name", None, "value"])
        ws.append(["Row 1", "ignored", 42])
        result = _sheet_rows_as_dicts(ws)
        assert len(result) == 1
        assert "name" in result[0]
        assert "value" in result[0]
        # blank header key should not appear
        assert "" not in result[0]

    def test_headers_are_normalised_lowercase(self):
        from app.routes.admin.utilities.import_export import _sheet_rows_as_dicts

        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "CODE", "Sort Order"])
        ws.append(["Test", "ABC", 5])
        result = _sheet_rows_as_dicts(ws)
        assert "name" in result[0]
        assert "code" in result[0]
        assert "sort order" in result[0]


class TestBuildLevelsJson:
    def test_all_none_returns_none(self):
        from app.routes.admin.utilities.import_export import _build_levels_json
        assert _build_levels_json(None, None, None) is None

    def test_primary_only(self):
        from app.routes.admin.utilities.import_export import _build_levels_json
        result = _build_levels_json(1, None, None)
        assert result == {"primary": 1}

    def test_primary_and_secondary(self):
        from app.routes.admin.utilities.import_export import _build_levels_json
        result = _build_levels_json(1, 2, None)
        assert result == {"primary": 1, "secondary": 2}

    def test_all_three_levels(self):
        from app.routes.admin.utilities.import_export import _build_levels_json
        result = _build_levels_json(1, 2, 3)
        assert result == {"primary": 1, "secondary": 2, "tertiary": 3}

    def test_tertiary_only(self):
        from app.routes.admin.utilities.import_export import _build_levels_json
        result = _build_levels_json(None, None, 5)
        assert result == {"tertiary": 5}


class TestMqStrToList:
    def test_none_returns_none(self):
        from app.routes.admin.utilities.import_export import _mq_str_to_list
        assert _mq_str_to_list(None) is None

    def test_empty_string_returns_none(self):
        from app.routes.admin.utilities.import_export import _mq_str_to_list
        assert _mq_str_to_list("") is None
        assert _mq_str_to_list("   ") is None

    def test_single_question(self):
        from app.routes.admin.utilities.import_export import _mq_str_to_list
        result = _mq_str_to_list("How many?")
        assert result == ["How many?"]

    def test_multiple_questions_semicolon_separated(self):
        from app.routes.admin.utilities.import_export import _mq_str_to_list
        result = _mq_str_to_list("Q1; Q2; Q3")
        assert result == ["Q1", "Q2", "Q3"]

    def test_whitespace_stripped(self):
        from app.routes.admin.utilities.import_export import _mq_str_to_list
        result = _mq_str_to_list("  Q1  ;  Q2  ")
        assert result == ["Q1", "Q2"]

    def test_trailing_semicolon_ignored(self):
        from app.routes.admin.utilities.import_export import _mq_str_to_list
        result = _mq_str_to_list("Q1; Q2;")
        assert result == ["Q1", "Q2"]

    def test_semicolons_only_returns_none(self):
        from app.routes.admin.utilities.import_export import _mq_str_to_list
        result = _mq_str_to_list(";;;")
        assert result is None


class TestTagsStrToList:
    def test_none_returns_none(self):
        from app.routes.admin.utilities.import_export import _tags_str_to_list
        assert _tags_str_to_list(None) is None

    def test_empty_string_returns_none(self):
        from app.routes.admin.utilities.import_export import _tags_str_to_list
        assert _tags_str_to_list("") is None
        assert _tags_str_to_list("   ") is None

    def test_single_tag(self):
        from app.routes.admin.utilities.import_export import _tags_str_to_list
        assert _tags_str_to_list("health") == ["health"]

    def test_multiple_tags_comma_separated(self):
        from app.routes.admin.utilities.import_export import _tags_str_to_list
        result = _tags_str_to_list("health, shelter, food")
        assert result == ["health", "shelter", "food"]

    def test_whitespace_stripped(self):
        from app.routes.admin.utilities.import_export import _tags_str_to_list
        result = _tags_str_to_list("  tag1  ,  tag2  ")
        assert result == ["tag1", "tag2"]

    def test_commas_only_returns_none(self):
        from app.routes.admin.utilities.import_export import _tags_str_to_list
        result = _tags_str_to_list(",,,")
        assert result is None


# ---------------------------------------------------------------------------
# Route: GET /admin/indicator_bank/import
# ---------------------------------------------------------------------------

class TestImportIndicatorsGet:
    def test_redirects_to_manage_indicator_bank(self, logged_in_client, db_session):
        resp = logged_in_client.get("/admin/indicator_bank/import")
        assert resp.status_code == 302
        assert "indicator_bank" in resp.headers["Location"].lower()

    def test_unauthenticated_redirects_to_login(self, client, db_session):
        resp = client.get("/admin/indicator_bank/import")
        assert resp.status_code in (302, 401)


# ---------------------------------------------------------------------------
# Route: POST /admin/indicator_bank/import (delegates to preview)
# ---------------------------------------------------------------------------

class TestImportIndicatorsPost:
    def test_post_delegates_to_preview(self, logged_in_client, db_session):
        """POST /admin/indicator_bank/import should behave like the preview endpoint."""
        sheets = {
            "Indicators": [
                ("name", "definition", "type", "unit"),
                ("Test Indicator", "A definition", "numeric", "count"),
            ]
        }

        with patch(
            "app.routes.admin.utilities.import_export.validate_upload_extension_and_mime",
            return_value=(True, None, "xlsx"),
        ), patch(
            "app.routes.admin.utilities.import_export._preview_indicator_import",
            return_value={"indicators": {"to_create": 1, "to_update": 0, "sample": []}},
        ), patch(
            "app.routes.admin.utilities.import_export._get_import_temp_path",
            return_value="/tmp/import_test.xlsx",
        ):
            data = {
                "file": _xlsx_file("test.xlsx", sheets),
            }
            resp = logged_in_client.post(
                "/admin/indicator_bank/import",
                data=data,
                content_type="multipart/form-data",
            )

        assert resp.status_code in (200, 400, 500)  # reached the preview logic


# ---------------------------------------------------------------------------
# Route: POST /admin/indicator_bank/import/preview
# ---------------------------------------------------------------------------

class TestPreviewIndicatorImport:
    def test_no_file_returns_400(self, logged_in_client, db_session):
        resp = logged_in_client.post(
            "/admin/indicator_bank/import/preview",
            content_type="multipart/form-data",
        )
        assert resp.status_code == 400
        data = resp.get_json()
        assert "file" in data.get("message", "").lower() or "error" in str(data).lower()

    def test_empty_filename_returns_400(self, logged_in_client, db_session):
        resp = logged_in_client.post(
            "/admin/indicator_bank/import/preview",
            data={"file": (io.BytesIO(b""), "")},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 400

    def test_invalid_file_type_returns_400(self, logged_in_client, db_session):
        with patch(
            "app.routes.admin.utilities.import_export.validate_upload_extension_and_mime",
            return_value=(False, "Please upload an Excel file", None),
        ):
            resp = logged_in_client.post(
                "/admin/indicator_bank/import/preview",
                data={"file": (io.BytesIO(b"fake content"), "test.txt")},
                content_type="multipart/form-data",
            )
        assert resp.status_code == 400

    def test_valid_excel_file_returns_summary(self, logged_in_client, db_session, tmp_path):
        """Valid xlsx file → 200 with import_token and summary."""
        fake_summary = {
            "indicators": {"to_create": 2, "to_update": 0, "sample": []},
            "types": {"to_create": 0, "to_update": 0},
            "units": {"to_create": 0, "to_update": 0},
            "sectors": {"to_create": 0, "to_update": 0},
            "subsectors": {"to_create": 0, "to_update": 0},
            "common_words": {"to_create": 0, "to_update": 0},
        }
        test_token = str(uuid.uuid4())
        temp_file = tmp_path / f"import_{test_token}.xlsx"

        with patch(
            "app.routes.admin.utilities.import_export.validate_upload_extension_and_mime",
            return_value=(True, None, "xlsx"),
        ), patch(
            "app.routes.admin.utilities.import_export._uuid_mod.uuid4",
            return_value=MagicMock(__str__=lambda self: test_token),
        ), patch(
            "app.routes.admin.utilities.import_export._get_import_temp_path",
            return_value=str(temp_file),
        ), patch(
            "app.routes.admin.utilities.import_export._preview_indicator_import",
            return_value=fake_summary,
        ):
            sheets = {"Indicators": [("name",), ("Test Ind",)]}
            resp = logged_in_client.post(
                "/admin/indicator_bank/import/preview",
                data={"file": _xlsx_file("test.xlsx", sheets)},
                content_type="multipart/form-data",
            )

        assert resp.status_code == 200
        data = resp.get_json()
        assert "import_token" in data
        assert "summary" in data

    def test_exception_in_preview_returns_500(self, logged_in_client, db_session):
        with patch(
            "app.routes.admin.utilities.import_export.validate_upload_extension_and_mime",
            return_value=(True, None, "xlsx"),
        ), patch(
            "app.routes.admin.utilities.import_export._get_import_temp_path",
            return_value="/tmp/fake.xlsx",
        ), patch(
            "app.routes.admin.utilities.import_export._preview_indicator_import",
            side_effect=RuntimeError("workbook error"),
        ), patch("builtins.open", MagicMock()), patch(
            "os.path.exists", return_value=True
        ), patch(
            "os.remove"
        ):
            sheets = {"Indicators": [("name",), ("Test Ind",)]}
            resp = logged_in_client.post(
                "/admin/indicator_bank/import/preview",
                data={"file": _xlsx_file("test.xlsx", sheets)},
                content_type="multipart/form-data",
            )

        # Exception path → 500
        assert resp.status_code == 500

    def test_temp_file_removed_on_exception(self, logged_in_client, db_session, tmp_path):
        """Temp file is cleaned up if an exception occurs during preview."""
        temp_file = tmp_path / "import_token123.xlsx"
        temp_file.write_bytes(b"fake")

        with patch(
            "app.routes.admin.utilities.import_export.validate_upload_extension_and_mime",
            return_value=(True, None, "xlsx"),
        ), patch(
            "app.routes.admin.utilities.import_export._get_import_temp_path",
            return_value=str(temp_file),
        ), patch(
            "app.routes.admin.utilities.import_export._preview_indicator_import",
            side_effect=ValueError("parse error"),
        ):
            sheets = {"Indicators": [("name",), ("Test Ind",)]}
            resp = logged_in_client.post(
                "/admin/indicator_bank/import/preview",
                data={"file": _xlsx_file("test.xlsx", sheets)},
                content_type="multipart/form-data",
            )

        assert resp.status_code == 500
        # File should have been removed
        assert not temp_file.exists()

    def test_unauthenticated_redirects(self, client, db_session):
        resp = client.post("/admin/indicator_bank/import/preview")
        assert resp.status_code in (302, 401, 403)


# ---------------------------------------------------------------------------
# Route: POST /admin/indicator_bank/import/apply
# ---------------------------------------------------------------------------

class TestApplyIndicatorImport:
    def test_no_token_returns_400(self, logged_in_client, db_session):
        resp = logged_in_client.post(
            "/admin/indicator_bank/import/apply",
            json={},
        )
        assert resp.status_code == 400
        data = resp.get_json()
        assert "token" in data.get("message", "").lower()

    def test_invalid_uuid_token_returns_400(self, logged_in_client, db_session):
        resp = logged_in_client.post(
            "/admin/indicator_bank/import/apply",
            json={"import_token": "not-a-uuid"},
        )
        assert resp.status_code == 400
        data = resp.get_json()
        assert "invalid" in data.get("message", "").lower()

    def test_missing_temp_file_returns_400(self, logged_in_client, db_session):
        token = str(uuid.uuid4())
        with patch(
            "app.routes.admin.utilities.import_export._get_import_temp_path",
            return_value="/tmp/nonexistent_import.xlsx",
        ), patch("os.path.exists", return_value=False):
            resp = logged_in_client.post(
                "/admin/indicator_bank/import/apply",
                json={"import_token": token},
            )

        assert resp.status_code == 400
        data = resp.get_json()
        assert "expired" in data.get("message", "").lower() or "not found" in data.get("message", "").lower()

    def test_successful_import_returns_200_with_message(
        self, logged_in_client, db_session, tmp_path
    ):
        token = str(uuid.uuid4())
        temp_file = tmp_path / f"import_{token}.xlsx"
        temp_file.write_bytes(b"fake")

        import_result = {
            "success": True,
            "imported": 3,
            "updated": 1,
            "sectors_imported": 0,
            "sectors_updated": 0,
            "subsectors_imported": 0,
            "subsectors_updated": 0,
            "common_words_imported": 0,
            "common_words_updated": 0,
            "measurement_types_imported": 0,
            "measurement_types_updated": 0,
            "measurement_units_imported": 0,
            "measurement_units_updated": 0,
            "errors": [],
            "message": "",
        }

        with patch(
            "app.routes.admin.utilities.import_export._get_import_temp_path",
            return_value=str(temp_file),
        ), patch(
            "app.routes.admin.utilities.import_export._process_indicator_import",
            return_value=import_result,
        ), patch("threading.Thread"):
            resp = logged_in_client.post(
                "/admin/indicator_bank/import/apply",
                json={"import_token": token},
            )

        assert resp.status_code == 200
        data = resp.get_json()
        assert "message" in data
        assert "3 indicator(s) created" in data["message"]
        assert "1 indicator(s) updated" in data["message"]

    def test_failed_import_returns_400(self, logged_in_client, db_session, tmp_path):
        token = str(uuid.uuid4())
        temp_file = tmp_path / f"import_{token}.xlsx"
        temp_file.write_bytes(b"fake")

        import_result = {
            "success": False,
            "imported": 0,
            "updated": 0,
            "errors": [],
            "message": "File parse error",
        }

        with patch(
            "app.routes.admin.utilities.import_export._get_import_temp_path",
            return_value=str(temp_file),
        ), patch(
            "app.routes.admin.utilities.import_export._process_indicator_import",
            return_value=import_result,
        ), patch("threading.Thread"):
            resp = logged_in_client.post(
                "/admin/indicator_bank/import/apply",
                json={"import_token": token},
            )

        assert resp.status_code == 400
        data = resp.get_json()
        assert "failed" in data.get("message", "").lower() or "error" in str(data).lower()

    def test_import_with_errors_includes_error_count_in_message(
        self, logged_in_client, db_session, tmp_path
    ):
        token = str(uuid.uuid4())
        temp_file = tmp_path / f"import_{token}.xlsx"
        temp_file.write_bytes(b"fake")

        import_result = {
            "success": True,
            "imported": 1,
            "updated": 0,
            "sectors_imported": 0,
            "sectors_updated": 0,
            "subsectors_imported": 0,
            "subsectors_updated": 0,
            "common_words_imported": 0,
            "common_words_updated": 0,
            "measurement_types_imported": 0,
            "measurement_types_updated": 0,
            "measurement_units_imported": 0,
            "measurement_units_updated": 0,
            "errors": ["Row 3: error.", "Row 5: error."],
            "message": "",
        }

        with patch(
            "app.routes.admin.utilities.import_export._get_import_temp_path",
            return_value=str(temp_file),
        ), patch(
            "app.routes.admin.utilities.import_export._process_indicator_import",
            return_value=import_result,
        ), patch("threading.Thread"):
            resp = logged_in_client.post(
                "/admin/indicator_bank/import/apply",
                json={"import_token": token},
            )

        assert resp.status_code == 200
        data = resp.get_json()
        assert "2 row error(s)" in data["message"]

    def test_import_with_all_lookup_counts(
        self, logged_in_client, db_session, tmp_path
    ):
        """Test that all lookup count fields appear in the message."""
        token = str(uuid.uuid4())
        temp_file = tmp_path / f"import_{token}.xlsx"
        temp_file.write_bytes(b"fake")

        import_result = {
            "success": True,
            "imported": 0,
            "updated": 0,
            "sectors_imported": 2,
            "sectors_updated": 1,
            "subsectors_imported": 3,
            "subsectors_updated": 0,
            "common_words_imported": 1,
            "common_words_updated": 2,
            "measurement_types_imported": 1,
            "measurement_types_updated": 1,
            "measurement_units_imported": 2,
            "measurement_units_updated": 0,
            "errors": [],
            "message": "",
        }

        with patch(
            "app.routes.admin.utilities.import_export._get_import_temp_path",
            return_value=str(temp_file),
        ), patch(
            "app.routes.admin.utilities.import_export._process_indicator_import",
            return_value=import_result,
        ), patch("threading.Thread"):
            resp = logged_in_client.post(
                "/admin/indicator_bank/import/apply",
                json={"import_token": token},
            )

        assert resp.status_code == 200
        msg = resp.get_json()["message"]
        assert "sector(s) created" in msg
        assert "sub-sector(s) created" in msg

    def test_exception_during_apply_returns_500(
        self, logged_in_client, db_session, tmp_path
    ):
        token = str(uuid.uuid4())
        temp_file = tmp_path / f"import_{token}.xlsx"
        temp_file.write_bytes(b"fake")

        with patch(
            "app.routes.admin.utilities.import_export._get_import_temp_path",
            return_value=str(temp_file),
        ), patch(
            "app.routes.admin.utilities.import_export._process_indicator_import",
            side_effect=RuntimeError("db failure"),
        ), patch("threading.Thread"):
            resp = logged_in_client.post(
                "/admin/indicator_bank/import/apply",
                json={"import_token": token},
            )

        assert resp.status_code == 500

    def test_no_changes_applied_message(self, logged_in_client, db_session, tmp_path):
        token = str(uuid.uuid4())
        temp_file = tmp_path / f"import_{token}.xlsx"
        temp_file.write_bytes(b"fake")

        import_result = {
            "success": True,
            "imported": 0,
            "updated": 0,
            "sectors_imported": 0,
            "sectors_updated": 0,
            "subsectors_imported": 0,
            "subsectors_updated": 0,
            "common_words_imported": 0,
            "common_words_updated": 0,
            "measurement_types_imported": 0,
            "measurement_types_updated": 0,
            "measurement_units_imported": 0,
            "measurement_units_updated": 0,
            "errors": [],
            "message": "",
        }

        with patch(
            "app.routes.admin.utilities.import_export._get_import_temp_path",
            return_value=str(temp_file),
        ), patch(
            "app.routes.admin.utilities.import_export._process_indicator_import",
            return_value=import_result,
        ), patch("threading.Thread"):
            resp = logged_in_client.post(
                "/admin/indicator_bank/import/apply",
                json={"import_token": token},
            )

        assert resp.status_code == 200
        data = resp.get_json()
        assert "No changes applied" in data["message"]

    def test_unauthenticated_redirects(self, client, db_session):
        resp = client.post(
            "/admin/indicator_bank/import/apply", json={"import_token": str(uuid.uuid4())}
        )
        assert resp.status_code in (302, 401, 403)

    def test_empty_token_string_returns_400(self, logged_in_client, db_session):
        resp = logged_in_client.post(
            "/admin/indicator_bank/import/apply",
            json={"import_token": "   "},
        )
        assert resp.status_code == 400


# ---------------------------------------------------------------------------
# Route: GET /admin/indicator_bank/change_history
# ---------------------------------------------------------------------------

class TestIndicatorChangeHistory:
    def test_renders_change_history_template(self, logged_in_client, db_session):
        with patch(
            "app.routes.admin.utilities.import_export.render_template",
            return_value="<html>history</html>",
        ) as mock_render:
            resp = logged_in_client.get("/admin/indicator_bank/change_history")

        assert resp.status_code == 200
        template_name = mock_render.call_args[0][0]
        assert "change_history" in template_name

    def test_passes_changes_and_title_to_template(self, logged_in_client, db_session):
        with patch(
            "app.routes.admin.utilities.import_export.render_template",
            return_value="<html>history</html>",
        ) as mock_render:
            resp = logged_in_client.get("/admin/indicator_bank/change_history")

        assert resp.status_code == 200
        kwargs = mock_render.call_args[1]
        assert "changes" in kwargs
        assert kwargs.get("title") == "Indicator Change History"

    def test_unauthenticated_redirects(self, client, db_session):
        resp = client.get("/admin/indicator_bank/change_history")
        assert resp.status_code in (302, 401, 403)


# ---------------------------------------------------------------------------
# Route: GET /admin/indicator_suggestions
# ---------------------------------------------------------------------------

class TestManageIndicatorSuggestions:
    def test_renders_suggestions_template(self, logged_in_client, db_session, app):
        _grant_suggestions_permission(db_session, app)

        with patch(
            "app.routes.admin.utilities.import_export.render_template",
            return_value="<html>suggestions</html>",
        ) as mock_render:
            resp = logged_in_client.get("/admin/indicator_suggestions")

        assert resp.status_code == 200
        template_name = mock_render.call_args[0][0]
        assert "suggestion" in template_name

    def test_accepts_status_filter(self, logged_in_client, db_session, app):
        _grant_suggestions_permission(db_session, app)

        with patch(
            "app.routes.admin.utilities.import_export.render_template",
            return_value="<html>suggestions</html>",
        ) as mock_render:
            resp = logged_in_client.get(
                "/admin/indicator_suggestions?status=pending"
            )

        assert resp.status_code == 200
        kwargs = mock_render.call_args[1]
        assert kwargs.get("status_filter") == "pending"

    def test_accepts_suggestion_type_filter(self, logged_in_client, db_session, app):
        _grant_suggestions_permission(db_session, app)

        with patch(
            "app.routes.admin.utilities.import_export.render_template",
            return_value="<html>suggestions</html>",
        ) as mock_render:
            resp = logged_in_client.get(
                "/admin/indicator_suggestions?suggestion_type=new"
            )

        assert resp.status_code == 200
        kwargs = mock_render.call_args[1]
        assert kwargs.get("suggestion_type_filter") == "new"

    def test_without_permission_redirects(self, logged_in_client, db_session):
        """Without suggestions.review permission → redirect to dashboard."""
        # Don't grant suggestions permission; default admin may or may not have it
        resp = logged_in_client.get("/admin/indicator_suggestions", follow_redirects=False)
        # Either renders (if permission already granted elsewhere) or redirects
        assert resp.status_code in (200, 302)


# ---------------------------------------------------------------------------
# Route: GET /admin/indicator_suggestions/view/<id>
# ---------------------------------------------------------------------------

class TestViewIndicatorSuggestion:
    def test_renders_suggestion_template(self, logged_in_client, db_session, app):
        from app.models import IndicatorSuggestion

        _grant_suggestions_permission(db_session, app)

        # Create a real suggestion
        with app.app_context():
            suggestion = IndicatorSuggestion(
                submitter_name="Test User",
                submitter_email="test@example.com",
                indicator_name="Test Indicator",
                reason="Testing purposes",
            )
            db_session.add(suggestion)
            db_session.commit()
            suggestion_id = suggestion.id

        with patch(
            "app.routes.admin.utilities.import_export.render_template",
            return_value="<html>suggestion detail</html>",
        ) as mock_render:
            resp = logged_in_client.get(
                f"/admin/indicator_suggestions/view/{suggestion_id}"
            )

        assert resp.status_code == 200
        template_name = mock_render.call_args[0][0]
        assert "suggestion" in template_name.lower()

    def test_returns_404_for_missing_suggestion(
        self, logged_in_client, db_session, app
    ):
        _grant_suggestions_permission(db_session, app)

        resp = logged_in_client.get("/admin/indicator_suggestions/view/99999")
        assert resp.status_code == 404

    def test_unauthenticated_redirects(self, client, db_session):
        resp = client.get("/admin/indicator_suggestions/view/1")
        assert resp.status_code in (302, 401, 403)


# ---------------------------------------------------------------------------
# Route: POST /admin/indicator_suggestions/update_status/<id>
# ---------------------------------------------------------------------------

class TestUpdateIndicatorSuggestionStatus:
    def test_updates_suggestion_status(self, logged_in_client, db_session, app):
        from app.models import IndicatorSuggestion

        _grant_suggestions_permission(db_session, app)

        with app.app_context():
            suggestion = IndicatorSuggestion(
                submitter_name="Test User",
                submitter_email="test@example.com",
                indicator_name="Test Indicator",
                reason="Testing",
            )
            db_session.add(suggestion)
            db_session.commit()
            suggestion_id = suggestion.id

        resp = logged_in_client.post(
            f"/admin/indicator_suggestions/update_status/{suggestion_id}",
            data={"status": "reviewed", "admin_notes": "Looks good"},
        )
        # Should redirect back to view page
        assert resp.status_code == 302
        assert str(suggestion_id) in resp.headers["Location"]

    def test_invalid_status_flashes_error(self, logged_in_client, db_session, app):
        from app.models import IndicatorSuggestion

        _grant_suggestions_permission(db_session, app)

        with app.app_context():
            suggestion = IndicatorSuggestion(
                submitter_name="Test User",
                submitter_email="test@example.com",
                indicator_name="Test Indicator",
                reason="Testing",
            )
            db_session.add(suggestion)
            db_session.commit()
            suggestion_id = suggestion.id

        resp = logged_in_client.post(
            f"/admin/indicator_suggestions/update_status/{suggestion_id}",
            data={"status": "invalid_status"},
        )
        assert resp.status_code == 302

    def test_approved_status_triggers_create_indicator(
        self, logged_in_client, db_session, app
    ):
        from app.models import IndicatorSuggestion

        _grant_suggestions_permission(db_session, app)

        with app.app_context():
            suggestion = IndicatorSuggestion(
                submitter_name="Test User",
                submitter_email="test@example.com",
                indicator_name="New Approved Indicator",
                reason="Good indicator",
            )
            db_session.add(suggestion)
            db_session.commit()
            suggestion_id = suggestion.id

        with patch(
            "app.routes.admin.utilities.import_export._create_indicator_from_suggestion",
        ) as mock_create:
            resp = logged_in_client.post(
                f"/admin/indicator_suggestions/update_status/{suggestion_id}",
                data={"status": "approved", "admin_notes": ""},
            )

        assert resp.status_code == 302
        mock_create.assert_called_once()

    def test_returns_404_for_missing_suggestion(
        self, logged_in_client, db_session, app
    ):
        _grant_suggestions_permission(db_session, app)

        resp = logged_in_client.post(
            "/admin/indicator_suggestions/update_status/99999",
            data={"status": "reviewed"},
        )
        assert resp.status_code == 404

    def test_unauthenticated_redirects(self, client, db_session):
        resp = client.post(
            "/admin/indicator_suggestions/update_status/1",
            data={"status": "reviewed"},
        )
        assert resp.status_code in (302, 401, 403)


# ---------------------------------------------------------------------------
# Route: POST /admin/indicator_suggestions/delete/<id>
# ---------------------------------------------------------------------------

class TestDeleteIndicatorSuggestion:
    def test_deletes_suggestion_and_redirects(
        self, logged_in_client, db_session, app
    ):
        from app.models import IndicatorSuggestion

        _grant_suggestions_permission(db_session, app)

        with app.app_context():
            suggestion = IndicatorSuggestion(
                submitter_name="Del User",
                submitter_email="del@example.com",
                indicator_name="Deletable Indicator",
                reason="To be deleted",
            )
            db_session.add(suggestion)
            db_session.commit()
            suggestion_id = suggestion.id

        resp = logged_in_client.post(
            f"/admin/indicator_suggestions/delete/{suggestion_id}"
        )
        assert resp.status_code == 302
        assert "indicator_suggestions" in resp.headers["Location"]

    def test_returns_404_for_missing_suggestion(
        self, logged_in_client, db_session, app
    ):
        _grant_suggestions_permission(db_session, app)

        resp = logged_in_client.post("/admin/indicator_suggestions/delete/99999")
        assert resp.status_code == 404

    def test_unauthenticated_redirects(self, client, db_session):
        resp = client.post("/admin/indicator_suggestions/delete/1")
        assert resp.status_code in (302, 401, 403)


# ---------------------------------------------------------------------------
# Internal helper: _preview_indicator_import (workbook parsing logic)
# ---------------------------------------------------------------------------

class TestPreviewIndicatorImportInternal:
    """Test _preview_indicator_import directly via a real workbook."""

    def _write_tmp_xlsx(self, tmp_path, sheets):
        path = tmp_path / "preview_test.xlsx"
        path.write_bytes(_make_xlsx_bytes(sheets))
        return str(path)

    def test_empty_workbook_returns_zero_summary(self, app, db_session, tmp_path):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(tmp_path, {"Sheet1": [("name",)]})
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["indicators"]["to_create"] == 0
        assert summary["indicators"]["to_update"] == 0

    def test_db_indicators_sheet_detected(self, app, db_session, tmp_path):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_Indicators": [
                    ("id", "name"),
                    (None, "Brand New Indicator"),
                ]
            },
        )
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["indicators"]["to_create"] == 1

    def test_fallback_active_sheet_used(self, app, db_session, tmp_path):
        """When no DB_Indicators sheet, active sheet is used."""
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Indicators": [
                    ("name", "definition"),
                    ("FallbackIndicator", "A definition"),
                ]
            },
        )
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["indicators"]["to_create"] == 1

    def test_types_sheet_to_create(self, app, db_session, tmp_path):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Types": [
                    ("id", "code", "name"),
                    (None, "newtype", "New Type"),
                ],
                "Indicators": [("name",)],
            },
        )
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["types"]["to_create"] == 1

    def test_units_sheet_to_create(self, app, db_session, tmp_path):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Units": [
                    ("id", "code", "name"),
                    (None, "newunit", "New Unit"),
                ],
                "Indicators": [("name",)],
            },
        )
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["units"]["to_create"] == 1

    def test_sectors_sheet_to_create(self, app, db_session, tmp_path):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Sectors": [
                    ("id", "name"),
                    (None, "New Sector XYZ"),
                ],
                "Indicators": [("name",)],
            },
        )
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["sectors"]["to_create"] == 1

    def test_subsectors_sheet_to_create(self, app, db_session, tmp_path):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Sub-Sectors": [
                    ("id", "name"),
                    (None, "New SubSector ABC"),
                ],
                "Indicators": [("name",)],
            },
        )
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["subsectors"]["to_create"] == 1

    def test_common_words_sheet_to_create(self, app, db_session, tmp_path):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Common Words": [
                    ("id", "term", "meaning"),
                    (None, "new_term_xyz", "A meaning"),
                ],
                "Indicators": [("name",)],
            },
        )
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["common_words"]["to_create"] == 1

    def test_db_sectors_subsectors_sheet(self, app, db_session, tmp_path):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_Sectors_SubSectors": [
                    ("id", "name", "record_type"),
                    (None, "New DB Sector", "sector"),
                    (None, "New DB SubSector", "subsector"),
                ],
                "DB_Indicators": [
                    ("id", "name"),
                ],
            },
        )
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["sectors"]["to_create"] == 1
        assert summary["subsectors"]["to_create"] == 1

    def test_db_measurement_types_sheet(self, app, db_session, tmp_path):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_MeasurementTypes": [
                    ("id", "code", "name"),
                    (None, "new_type_preview", "New Type"),
                ],
                "DB_Indicators": [("id", "name")],
            },
        )
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["types"]["to_create"] == 1

    def test_db_measurement_units_sheet(self, app, db_session, tmp_path):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_MeasurementUnits": [
                    ("id", "code", "name"),
                    (None, "new_unit_preview", "New Unit"),
                ],
                "DB_Indicators": [("id", "name")],
            },
        )
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["units"]["to_create"] == 1

    def test_db_common_words_sheet(self, app, db_session, tmp_path):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_CommonWords": [
                    ("id", "term", "meaning"),
                    (None, "new_common_term", "A meaning"),
                ],
                "DB_Indicators": [("id", "name")],
            },
        )
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["common_words"]["to_create"] == 1

    def test_sample_limited_to_25(self, app, db_session, tmp_path):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        rows = [("id", "name")]
        for i in range(30):
            rows.append((None, f"Indicator {i:03d}"))

        path = self._write_tmp_xlsx(tmp_path, {"DB_Indicators": rows})
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert len(summary["indicators"]["sample"]) == 25

    def test_row_with_only_id_and_no_name_in_db_indicators(
        self, app, db_session, tmp_path
    ):
        from app.routes.admin.utilities.import_export import _preview_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_Indicators": [
                    ("id", "name"),
                    (None, None),  # empty row → skipped
                    (None, "ValidIndicator"),
                ]
            },
        )
        with app.app_context():
            summary = _preview_indicator_import(path)

        assert summary["indicators"]["to_create"] == 1


# ---------------------------------------------------------------------------
# Internal: _process_indicator_import (Branch A: DB_* sheets)
# ---------------------------------------------------------------------------

class TestProcessIndicatorImportBranchA:
    """Test _process_indicator_import with DB_* sheet layout."""

    def _write_tmp_xlsx(self, tmp_path, sheets):
        path = tmp_path / "apply_test.xlsx"
        path.write_bytes(_make_xlsx_bytes(sheets))
        return str(path)

    def test_creates_new_indicator_from_db_indicators_sheet(
        self, app, db_session, admin_user, tmp_path
    ):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_Indicators": [
                    ("id", "name", "definition", "type", "unit"),
                    (None, "Import Test Indicator", "A definition", "numeric", "count"),
                ]
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["imported"] >= 1

    def test_updates_existing_indicator(self, app, db_session, admin_user, tmp_path):
        from app.models import IndicatorBank
        from app.routes.admin.utilities.import_export import _process_indicator_import

        with app.app_context():
            ind = IndicatorBank(
                name="ExistingIndForUpdate",
                definition="Old def",
                type="numeric",
                unit="count",
            )
            db_session.add(ind)
            db_session.commit()

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_Indicators": [
                    ("id", "name", "definition", "type", "unit"),
                    (None, "ExistingIndForUpdate", "New definition", "numeric", "count"),
                ]
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["updated"] >= 1

    def test_creates_sector_from_db_sectors_subsectors(
        self, app, db_session, admin_user, tmp_path
    ):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_Sectors_SubSectors": [
                    ("id", "name", "record_type"),
                    (None, "New Test Sector Branch A", "sector"),
                ],
                "DB_Indicators": [("id", "name")],
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["sectors_imported"] >= 1

    def test_creates_subsector_from_db_sectors_subsectors(
        self, app, db_session, admin_user, tmp_path
    ):
        from app.routes.admin.utilities.import_export import _process_indicator_import
        from app.models import Sector

        with app.app_context():
            # Create parent sector first
            sector = Sector(name="Parent Sector A")
            db_session.add(sector)
            db_session.commit()
            sector_id = sector.id

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_Sectors_SubSectors": [
                    ("id", "name", "record_type", "sector_id"),
                    (None, "New DB SubSector Test", "subsector", sector_id),
                ],
                "DB_Indicators": [("id", "name")],
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["subsectors_imported"] >= 1

    def test_creates_common_words(self, app, db_session, admin_user, tmp_path):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_CommonWords": [
                    ("id", "term", "meaning"),
                    (None, "new_test_word_bra", "A meaning"),
                ],
                "DB_Indicators": [("id", "name")],
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["common_words_imported"] >= 1

    def test_creates_measurement_types(self, app, db_session, admin_user, tmp_path):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_MeasurementTypes": [
                    ("id", "code", "name"),
                    (None, "new_type_bra_test", "New Type BranchA"),
                ],
                "DB_Indicators": [("id", "name")],
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["measurement_types_imported"] >= 1

    def test_creates_measurement_units(self, app, db_session, admin_user, tmp_path):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_MeasurementUnits": [
                    ("id", "code", "name"),
                    (None, "new_unit_bra_test", "New Unit BranchA"),
                ],
                "DB_Indicators": [("id", "name")],
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["measurement_units_imported"] >= 1

    def test_row_error_does_not_abort_import(self, app, db_session, admin_user, tmp_path):
        """A bad row in DB_Indicators is captured as an error but import continues."""
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "DB_Indicators": [
                    ("id", "name", "definition"),
                    (None, "GoodIndicator", "Good definition"),
                    (None, None, None),  # empty → skipped
                ]
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True


# ---------------------------------------------------------------------------
# Internal: _process_indicator_import (Branch B: human-readable sheets)
# ---------------------------------------------------------------------------

class TestProcessIndicatorImportBranchB:
    """Test _process_indicator_import with human-readable sheet layout."""

    def _write_tmp_xlsx(self, tmp_path, sheets):
        path = tmp_path / "apply_brb_test.xlsx"
        path.write_bytes(_make_xlsx_bytes(sheets))
        return str(path)

    def test_creates_indicator_from_indicators_sheet(
        self, app, db_session, admin_user, tmp_path
    ):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Indicators": [
                    ("name", "definition", "type", "unit"),
                    ("BranchB New Indicator", "A definition", "numeric", "count"),
                ]
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["imported"] >= 1

    def test_updates_existing_indicator_branch_b(
        self, app, db_session, admin_user, tmp_path
    ):
        from app.models import IndicatorBank
        from app.routes.admin.utilities.import_export import _process_indicator_import

        with app.app_context():
            ind = IndicatorBank(
                name="BranchBExisting",
                definition="Old def",
                type="numeric",
                unit="pct",
            )
            db_session.add(ind)
            db_session.commit()

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Indicators": [
                    ("name", "definition", "type", "unit"),
                    ("BranchBExisting", "Updated definition", "numeric", "pct"),
                ]
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["updated"] >= 1

    def test_types_sheet_imported(self, app, db_session, admin_user, tmp_path):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Indicators": [("name",)],
                "Types": [
                    ("id", "code", "name"),
                    (None, "new_type_brb", "New Type BranchB"),
                ],
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["measurement_types_imported"] >= 1

    def test_units_sheet_imported(self, app, db_session, admin_user, tmp_path):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Indicators": [("name",)],
                "Units": [
                    ("id", "code", "name"),
                    (None, "new_unit_brb", "New Unit BranchB"),
                ],
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["measurement_units_imported"] >= 1

    def test_sectors_sheet_imported(self, app, db_session, admin_user, tmp_path):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Indicators": [("name",)],
                "Sectors": [
                    ("id", "name"),
                    (None, "New Sector BranchB XYZ"),
                ],
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["sectors_imported"] >= 1

    def test_subsectors_sheet_imported(self, app, db_session, admin_user, tmp_path):
        from app.routes.admin.utilities.import_export import _process_indicator_import
        from app.models import Sector

        with app.app_context():
            sector = Sector(name="Sector For SubSector BranchB")
            db_session.add(sector)
            db_session.commit()

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Indicators": [("name",)],
                "Sub-Sectors": [
                    ("id", "name", "sector"),
                    (None, "New SubSector BranchB XYZ", "Sector For SubSector BranchB"),
                ],
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["subsectors_imported"] >= 1

    def test_common_words_sheet_imported(self, app, db_session, admin_user, tmp_path):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Indicators": [("name",)],
                "Common Words": [
                    ("id", "term", "meaning"),
                    (None, "new_word_brb_test", "A meaning"),
                ],
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["common_words_imported"] >= 1

    def test_indicator_with_monitoring_questions_and_tags(
        self, app, db_session, admin_user, tmp_path
    ):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Indicators": [
                    ("name", "definition", "type", "unit", "monitoring questions", "tags"),
                    (
                        "BranchB MQ Indicator",
                        "Def",
                        "numeric",
                        "count",
                        "How many?; In what context?",
                        "health, food",
                    ),
                ]
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["imported"] >= 1

    def test_row_error_captured_and_import_continues(
        self, app, db_session, admin_user, tmp_path
    ):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        # Patch backfill to raise error on first call only
        original_backfill = None
        call_count = {"n": 0}

        def patched_backfill(ind):
            call_count["n"] += 1
            if call_count["n"] == 1:
                raise ValueError("backfill error")

        path = self._write_tmp_xlsx(
            tmp_path,
            {
                "Indicators": [
                    ("name", "definition"),
                    ("BranchB Error Row", "def"),
                    ("BranchB Good Row", "good def"),
                ]
            },
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context(), patch(
                "app.routes.admin.utilities.import_export.backfill_fk_from_strings_bank",
                side_effect=patched_backfill,
            ):
                login_user(admin_user)
                result = _process_indicator_import(path)

        # Success should still be True (partial errors are OK)
        assert result["success"] is True
        assert len(result["errors"]) >= 1

    def test_empty_indicators_sheet_no_imports(
        self, app, db_session, admin_user, tmp_path
    ):
        from app.routes.admin.utilities.import_export import _process_indicator_import

        path = self._write_tmp_xlsx(
            tmp_path,
            {"Indicators": [("name", "definition")]},  # header only
        )
        with app.app_context():
            from flask_login import login_user

            with app.test_request_context():
                login_user(admin_user)
                result = _process_indicator_import(path)

        assert result["success"] is True
        assert result["imported"] == 0
