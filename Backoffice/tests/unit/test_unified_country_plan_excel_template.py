"""Unit tests for Unified Country Plan Excel template helpers."""

from __future__ import annotations

import os
import sys

import pytest

BACKOFFICE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
IMPORTS_DIR = os.path.join(BACKOFFICE_DIR, "scripts", "imports")
if IMPORTS_DIR not in sys.path:
    sys.path.insert(0, IMPORTS_DIR)

from import_upr_excel_data import UprImportContext  # noqa: E402
from unified_country_plan_excel_template import (  # noqa: E402
    START_REGION_CELL,
    START_SHEET,
    FUNDING_SHEET,
    FUNDING_TABLE,
    SUPPORT_SHEET,
    SUPPORT_TABLE,
    COMMENT_NAMED_CELL,
    _cell_is_tick,
    _export_support_to_workbook,
    _import_funding_matrices,
    _import_support_matrix,
    _parse_funding_row_entity,
    _parse_emergency_row_id,
    _parse_planning_support_ticks,
    _quiet_openpyxl_io,
    _refresh_funding_pns_array_formula,
    restore_workbook_dynamic_array_metadata,
    _workbook_region_for_ns_name,
    _workbook_region_for_country,
    _workbook_region_label,
    _write_start_sheet_selection,
    funding_column_header,
    parse_funding_column_header,
    parse_comment,
    parse_version,
    period_to_workbook_version,
    planning_year_triplet,
    read_named_table,
    read_table_cell,
    rewrite_planning_year_headers,
    validate_unified_country_plan_import_file,
    write_table_cell,
)
from upr_country_reporting_excel_template import _bilateral_ns_name_for_row, write_named_cell, _write_bilateral_ns_source_cell  # noqa: E402
TEMPLATE_PATH = os.path.join(
    BACKOFFICE_DIR,
    "app",
    "static",
    "templates",
    "unified_country_plan.xlsx",
)


@pytest.fixture(scope="module")
def ucp_workbook():
    if not os.path.isfile(TEMPLATE_PATH):
        pytest.skip("Unified Country Plan template file not present")
    import openpyxl

    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    yield wb
    wb.close()


def test_period_to_workbook_version():
    assert period_to_workbook_version("2027") == "P27.V1.0"
    assert period_to_workbook_version("2030") == "P30.V1.0"


def test_planning_year_triplet():
    assert planning_year_triplet("2027") == (2027, 2028, 2029)


def test_funding_header_round_trip():
    assert funding_column_header("SP1", 2028) == "SP1_2028"
    assert parse_funding_column_header("EA2_2027") == ("EA2", 2027)
    assert parse_funding_column_header("NS") is None


def test_parse_funding_row_entity():
    assert _parse_funding_row_entity("IFRC Secretariat") == ("ifrc", "IFRC Secretariat")
    assert _parse_funding_row_entity("Host National Society\nCameroon Red Cross") == ("hns", "HNS")
    assert _parse_funding_row_entity("Netherlands Red Cross") == ("pns", "Netherlands Red Cross")


def test_parse_emergency_row_id():
    assert _parse_emergency_row_id("Flood response (MDR123)") == ("Flood response", "MDR123")
    assert _parse_emergency_row_id("Appeal only") == ("Appeal only", "")


def test_cell_is_tick():
    assert _cell_is_tick(1)
    assert _cell_is_tick("1")
    assert _cell_is_tick("x")
    assert not _cell_is_tick(0)
    assert not _cell_is_tick(None)


def test_validate_template_structure(ucp_workbook):
    result = validate_unified_country_plan_import_file(
        ucp_workbook,
        expected_country="",
        expected_period="",
    )
    assert result["valid"] is True
    assert result["preview"]["round_code"] == "P27"


def test_rewrite_planning_year_headers(ucp_workbook):
    import openpyxl
    import re
    import tempfile
    import zipfile

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    try:
        with _quiet_openpyxl_io():
            wb = openpyxl.load_workbook(TEMPLATE_PATH)
        rewrite_planning_year_headers(wb, "2029")
        with _quiet_openpyxl_io():
            wb.save(path)
        wb.close()

        wb2 = openpyxl.load_workbook(path, data_only=True)
        _, people_rows = read_named_table(wb2, "People to be reached", "Data_People")
        years = [row.get("Year") for row in people_rows]
        assert years[:3] == [2029, 2030, 2031]
        headers, _ = read_named_table(wb2, "Funding requirements", "Data_FR")
        funding_years = sorted(
            {parse_funding_column_header(h)[1] for h in headers if parse_funding_column_header(h)}
        )
        assert funding_years == [2029, 2030, 2031]
        with zipfile.ZipFile(path) as z:
            xml = z.read("xl/tables/table11.xml").decode("utf-8")
        xml_cols = re.findall(r'<tableColumn[^>]*name="([^"]+)"', xml)
        assert headers == xml_cols
        wb2.close()
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


def test_rewrite_planning_year_headers_same_period_keeps_table_metadata(ucp_workbook):
    """Rewriting headers for the template's default period must not corrupt Data_FR."""
    import openpyxl
    import re
    import tempfile
    import zipfile

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    try:
        with _quiet_openpyxl_io():
            wb = openpyxl.load_workbook(TEMPLATE_PATH)
        rewrite_planning_year_headers(wb, "2027")
        with _quiet_openpyxl_io():
            wb.save(path)
        wb.close()

        wb2 = openpyxl.load_workbook(path, data_only=True)
        headers, _ = read_named_table(wb2, "Funding requirements", "Data_FR")
        with zipfile.ZipFile(path) as z:
            xml = z.read("xl/tables/table11.xml").decode("utf-8")
        xml_cols = re.findall(r'<tableColumn[^>]*name="([^"]+)"', xml)
        assert headers == xml_cols
        assert headers[15:18] == ["SP2_2029", "SP1_2029", "SP3_2029"]
        wb2.close()
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


def test_parse_version(ucp_workbook):
    round_code, period = parse_version(ucp_workbook)
    assert round_code == "P27"
    assert period == "2027"


def test_workbook_region_for_country_from_template_table(ucp_workbook):
    assert _workbook_region_for_country(ucp_workbook, "Afghanistan") == "Asia Pacific"
    assert _workbook_region_for_country(ucp_workbook, "afghanistan") == "Asia Pacific"


def test_workbook_region_label_maps_mena():
    assert _workbook_region_label("MENA") == "Middle East and North Africa"
    assert _workbook_region_label("Asia Pacific") == "Asia Pacific"


def test_write_start_sheet_selection(ucp_workbook):
    import openpyxl

    with _quiet_openpyxl_io():
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    _write_start_sheet_selection(wb, "Afghanistan", "Asia Pacific")
    assert wb[START_SHEET][START_REGION_CELL].value == "Asia Pacific"
    assert wb[START_SHEET]["K12"].value == "Afghanistan"
    wb.close()


def test_parse_planning_support_ticks():
    cells = {"166_SP1": 1, "166_SP2": 0, "42_EFs": 1, "foo": 1}
    assert _parse_planning_support_ticks(cells) == {166: {"SP1": True}, 42: {"EFs": True}}


def test_workbook_region_for_ns_name_from_table9(ucp_workbook):
    assert _workbook_region_for_ns_name(ucp_workbook, "Swiss Red Cross") == "Europe and Central Asia"


def test_bilateral_support_row_reads_column_c():
    import openpyxl

    with _quiet_openpyxl_io():
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    _write_bilateral_ns_source_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, 0, "Swiss Red Cross")
    assert _bilateral_ns_name_for_row(wb, SUPPORT_SHEET, SUPPORT_TABLE, 0) == "Swiss Red Cross"
    wb.close()


def test_export_support_writes_ns_and_ticks():
    import openpyxl
    from unittest.mock import patch

    with _quiet_openpyxl_io():
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ctx = type("Ctx", (), {"ns_name_to_id": {"swiss red cross": 166}})()
    entry = type("Entry", (), {"disagg_data": {"166_SP1": 1, "166_EFs": 1}})()

    with patch("app.models.organization.NationalSociety") as mock_ns:
        mock_ns.query.filter.return_value.all.return_value = [
            type("NS", (), {"id": 166, "name": "Swiss Red Cross"})(),
        ]
        _export_support_to_workbook(wb, entry, ctx)

    assert _bilateral_ns_name_for_row(wb, SUPPORT_SHEET, SUPPORT_TABLE, 0) == "Swiss Red Cross"
    assert wb[SUPPORT_SHEET]["B5"].value == "Europe and Central Asia"
    assert read_table_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, 0, "SP1") == "X"
    assert read_table_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, 0, "EFs") == "X"
    wb.close()


def test_import_support_matrix_reads_column_c():
    import openpyxl

    with _quiet_openpyxl_io():
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    _write_bilateral_ns_source_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, 0, "Swiss Red Cross")
    write_table_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, 0, "SP2", 1)
    write_table_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, 0, "EFs", 1)
    ctx = type("Ctx", (), {"ns_name_to_id": {"swiss red cross": 166}})()
    cells = _import_support_matrix(wb, ctx, aes_id=1, warnings=[])
    assert cells == {"166_SP2": 1, "166_EFs": 1}
    wb.close()


def test_import_funding_adds_pns_rows_from_bilateral_support_without_values():
    import openpyxl
    import tempfile

    with _quiet_openpyxl_io():
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    _write_bilateral_ns_source_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, 0, "Swiss Red Cross")
    _write_bilateral_ns_source_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, 1, "Norwegian Red Cross")
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb2 = None
    try:
        wb.save(tmp)
        wb.close()
        wb = None
        wb2 = openpyxl.load_workbook(tmp, data_only=True)
        ctx = UprImportContext(template_ids=[24])
        ctx.ns_name_to_id = {"swiss red cross": 166, "norwegian red cross": 200}
        matrices = _import_funding_matrices(
            wb2,
            ctx,
            aes_id=1,
            iso3="AFG",
            period="2027",
            rnd="P27",
            warnings=[],
        )
        for item_id in (967, 968, 974):
            cells = matrices[item_id]
            assert cells["166_SP2"] == ""
            assert cells["200_SP2"] == ""
    finally:
        if wb2 is not None:
            wb2.close()
        if wb is not None:
            wb.close()
        try:
            os.unlink(tmp)
        except OSError:
            pass


def test_import_funding_reads_pns_values_by_bilateral_row_index():
    import openpyxl
    import tempfile

    with _quiet_openpyxl_io():
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    _write_bilateral_ns_source_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, 0, "Swiss Red Cross")
    _write_bilateral_ns_source_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, 1, "Norwegian Red Cross")
    headers, rows = read_named_table(wb, FUNDING_SHEET, FUNDING_TABLE)
    row_offsets = {row["_row"]: idx for idx, row in enumerate(rows)}
    write_table_cell(wb, FUNDING_SHEET, FUNDING_TABLE, row_offsets[11], "SP1_2027", 750)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb2 = None
    try:
        wb.save(tmp)
        wb.close()
        wb = None
        wb2 = openpyxl.load_workbook(tmp, data_only=True)
        ctx = UprImportContext(template_ids=[24])
        ctx.ns_name_to_id = {"swiss red cross": 166, "norwegian red cross": 200}
        matrices = _import_funding_matrices(
            wb2,
            ctx,
            aes_id=1,
            iso3="AFG",
            period="2027",
            rnd="P27",
            warnings=[],
        )
        assert matrices[967]["200_SP1"] == 750.0
        assert matrices[967]["166_SP2"] == ""
    finally:
        if wb2 is not None:
            wb2.close()
        if wb is not None:
            wb.close()
        try:
            os.unlink(tmp)
        except OSError:
            pass


def test_refresh_funding_pns_array_formula_keeps_dynamic_formula():
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula

    with _quiet_openpyxl_io():
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    names = ["Swiss Red Cross", "Norwegian Red Cross"]
    for idx, name in enumerate(names):
        _write_bilateral_ns_source_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, idx, name)
    _refresh_funding_pns_array_formula(wb)
    val = wb[FUNDING_SHEET]["B10"].value
    assert isinstance(val, ArrayFormula)
    assert "FILTER" in val.text
    assert wb[FUNDING_SHEET]["B11"].value is None
    wb.close()


def test_parse_and_write_comment_single_cell(ucp_workbook):
    write_named_cell(ucp_workbook, COMMENT_NAMED_CELL, "Planning notes from the NS.")
    assert parse_comment(ucp_workbook) == "Planning notes from the NS."


def test_restore_workbook_dynamic_array_metadata_preserves_cm_flag():
    import openpyxl
    import tempfile
    import zipfile
    from openpyxl.worksheet.formula import ArrayFormula

    with _quiet_openpyxl_io():
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    _write_bilateral_ns_source_cell(wb, SUPPORT_SHEET, SUPPORT_TABLE, 0, "Swiss Red Cross")
    _refresh_funding_pns_array_formula(wb)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    try:
        with _quiet_openpyxl_io():
            wb.save(path)
        wb.close()
        restore_workbook_dynamic_array_metadata(TEMPLATE_PATH, path)
        with zipfile.ZipFile(path) as z:
            assert "xl/metadata.xml" in z.namelist()
            funding_xml = z.read("xl/worksheets/sheet5.xml").decode("utf-8")
            assert 'r="B10"' in funding_xml and 'cm="1"' in funding_xml
            assert isinstance(
                openpyxl.load_workbook(path)[FUNDING_SHEET]["B10"].value,
                ArrayFormula,
            )
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass
