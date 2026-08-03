"""Matrix-specific template Excel import logic."""

# ========== Template Excel Import/Export Service ==========
from app.utils.datetime_helpers import utcnow
"""
Service for exporting and importing form templates to/from Excel.
"""
from flask import current_app
from flask_login import current_user
from app import db
from app.models import (
    FormTemplate, FormPage, FormSection, FormItem, FormTemplateVersion,
    FormData,
    RepeatGroupInstance, RepeatGroupData, DynamicIndicatorData, DynamicSectionContext,
)
from app.models.documents import SubmittedDocument
from app.models.indicator_bank import IndicatorBank
from contextlib import suppress
from app.services.monitoring.memory import memory_tracker
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import io
import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from app.utils.stable_key import generate_stable_key, is_valid_stable_key, normalize_stable_key

from app.services.templates.excel_base import TemplateExcelBase


class TemplateExcelMatrixMixin(TemplateExcelBase):
    """Mixin for TemplateExcelService."""

    _matrix_import_logger: Optional[logging.Logger] = None

    _MATRIX_IMPORT_LOG_PREFIX = '[excel-import:matrix]'

    _MATRIX_CONFIG_KEYS = frozenset({
        'type', 'columns', 'column_groups', 'rows', 'row_mode', 'show_row_totals',
        'show_column_totals', 'auto_load_entities', 'highlight_manual_rows',
        'legend_text', 'legend_text_translations', 'legend_hide', 'lookup_list_id',
        'list_display_column', 'list_filters', 'group_by_column', 'group_dropdown_enabled',
        'group_table_enabled', 'search_placeholder', 'search_placeholder_translations',
        'plugin_config',
    })

    @classmethod
    def _get_matrix_import_logger(cls) -> logging.Logger:
        """Dedicated logger: always writes to stdout and instance/logs/excel_matrix_import.log."""
        if cls._matrix_import_logger is not None:
            return cls._matrix_import_logger

        logger = logging.getLogger('excel_matrix_import')
        logger.setLevel(logging.DEBUG)
        logger.propagate = False

        if not logger.handlers:
            formatter = logging.Formatter(
                '[%(asctime)s] %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S',
            )
            stdout_handler = logging.StreamHandler(sys.stdout)
            stdout_handler.setFormatter(formatter)
            logger.addHandler(stdout_handler)

            try:
                log_dir = Path(__file__).resolve().parents[2] / 'instance' / 'logs'
                log_dir.mkdir(parents=True, exist_ok=True)
                file_handler = logging.FileHandler(
                    log_dir / 'excel_matrix_import.log',
                    encoding='utf-8',
                )
                file_handler.setFormatter(formatter)
                logger.addHandler(file_handler)
            except OSError as exc:
                logger.warning(
                    '%s Could not open excel_matrix_import.log: %s',
                    cls._MATRIX_IMPORT_LOG_PREFIX,
                    exc,
                )

        cls._matrix_import_logger = logger
        return logger

    @classmethod
    def matrix_import_entry_log(cls, message: str) -> None:
        """High-visibility entry point log (routes, import start)."""
        cls._matrix_import_log(message)

    @classmethod
    def _matrix_import_log(cls, message: str, *, level: str = 'info') -> None:
        """Log matrix Excel import diagnostics to dedicated logger, app logger, and log file."""
        text = f"{cls._MATRIX_IMPORT_LOG_PREFIX} {message}"
        logger = cls._get_matrix_import_logger()
        log_fn = getattr(logger, level, logger.info)
        log_fn(text)
        try:
            app_log_fn = getattr(current_app.logger, level, current_app.logger.info)
            app_log_fn(text)
        except RuntimeError:
            pass

    @classmethod
    def _scan_workbook_matrix_items(cls, workbook, *, stage: str) -> int:
        """Scan Items sheet for matrix rows and log config/group hints (validate or import)."""
        if 'Items' not in workbook.sheetnames:
            cls._matrix_import_log(f"{stage}: Items sheet missing in workbook", level='warning')
            return 0

        sheet = workbook['Items']
        headers = [cell.value for cell in sheet[1]]
        if 'config' not in headers or 'item_type' not in headers:
            cls._matrix_import_log(
                f"{stage}: Items sheet missing config or item_type column (headers={headers})",
                level='warning',
            )
            return 0

        matrix_count = 0
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue
            row_data = cls._row_dict_from_sheet(headers, row)
            item_type = str(row_data.get('item_type') or '').strip().lower()
            if item_type != 'matrix':
                continue
            matrix_count += 1
            parsed = cls._parse_json(row_data.get('config'))
            normalized = cls._normalize_matrix_item_config('matrix', parsed) if parsed else None
            cls._matrix_import_log(
                f"{stage} Items row {row_idx} export_id={row_data.get('id')!r} "
                f"label={row_data.get('label')!r}: "
                f"raw={cls._summarize_raw_config_cell(row_data.get('config'))}; "
                f"normalized={cls._summarize_matrix_config_for_log(normalized if isinstance(normalized, dict) else {})}"
            )

        cls._matrix_import_log(f"{stage}: found {matrix_count} matrix row(s) in workbook")
        return matrix_count

    @classmethod
    def _summarize_matrix_config_for_log(cls, config: Any) -> str:
        if not isinstance(config, dict):
            return f"type={type(config).__name__}, value={config!r}"

        mc = config.get('matrix_config')
        if not isinstance(mc, dict):
            if config.get('type') == 'matrix' or isinstance(config.get('columns'), list):
                mc = config
            else:
                return "no matrix_config key"

        columns = mc.get('columns') if isinstance(mc.get('columns'), list) else []
        column_groups = mc.get('column_groups') if isinstance(mc.get('column_groups'), dict) else {}
        grouped = [
            f"{col.get('name')}→{col.get('group')}"
            for col in columns
            if isinstance(col, dict) and col.get('group')
        ]
        ungrouped = [
            str(col.get('name', col))
            for col in columns
            if isinstance(col, dict) and not col.get('group')
        ]
        grouped_preview = grouped[:10]
        if len(grouped) > 10:
            grouped_preview.append(f"...+{len(grouped) - 10} more")
        return (
            f"columns={len(columns)}, "
            f"column_group_keys={list(column_groups.keys())}, "
            f"grouped={grouped_preview or 'none'}, "
            f"ungrouped={ungrouped[:5]}{'...' if len(ungrouped) > 5 else ''}"
        )

    @classmethod
    def _summarize_raw_config_cell(cls, raw: Any) -> str:
        if raw is None or raw == '' or raw == 'None':
            return 'empty'
        if isinstance(raw, dict):
            return cls._summarize_matrix_config_for_log(raw)
        text = str(raw).strip()
        has_column_groups = 'column_groups' in text
        has_group_key = '"group"' in text or "'group'" in text
        preview = text[:160].replace('\n', '\\n')
        if len(text) > 160:
            preview += '...'
        return (
            f"text_len={len(text)}, has_column_groups={has_column_groups}, "
            f"has_group_key={has_group_key}, preview={preview!r}"
        )

    @classmethod
    def _log_matrix_items_in_version(cls, template_id: int, version_id: int, stage: str) -> None:
        """Log all matrix items currently in the target version."""
        matrix_items = (
            FormItem.query.join(FormSection, FormItem.section_id == FormSection.id)
            .filter(
                FormItem.template_id == template_id,
                FormSection.version_id == version_id,
                FormItem.item_type == 'matrix',
            )
            .order_by(FormItem.order, FormItem.id)
            .all()
        )
        cls._matrix_import_log(
            f"{stage}: found {len(matrix_items)} matrix item(s) in version_id={version_id}"
        )
        for item in matrix_items:
            cls._matrix_import_log(
                f"{stage}: item db_id={item.id} label={item.label!r} "
                f"section_id={item.section_id} order={item.order} "
                f"config={cls._summarize_matrix_config_for_log(item.config)}"
            )

    @classmethod
    def _deep_copy_json(cls, value: Any) -> Any:
        try:
            return json.loads(json.dumps(value))
        except (TypeError, ValueError):
            import copy
            return copy.deepcopy(value)

    @classmethod
    def _clean_imported_matrix_config(cls, config: Dict[str, Any]) -> Dict[str, Any]:
        """Remove legacy matrix keys duplicated at config root; keep matrix_config canonical."""
        out = cls._deep_copy_json(config)
        if not isinstance(out, dict):
            return out
        if isinstance(out.get('matrix_config'), dict):
            for key in cls._MATRIX_CONFIG_KEYS:
                out.pop(key, None)
            out.pop('column_groups', None)
        return out

    @classmethod
    def _find_existing_item_for_import(
        cls,
        *,
        section_id: int,
        item_type: str,
        item_label: str,
        item_order: float,
        export_id: Optional[int],
        export_id_to_existing: Dict[int, FormItem],
        existing_items_lookup: Dict[Tuple, FormItem],
    ) -> Tuple[Optional[FormItem], Optional[str]]:
        """Resolve an existing FormItem for an imported row (matrix: label match wins over export id)."""
        existing_item: Optional[FormItem] = None
        match_method: Optional[str] = None

        if item_label:
            label_key = (section_id, item_type, item_label)
            existing_item = existing_items_lookup.get(label_key)
            if existing_item:
                return existing_item, f'label_key={label_key!r}'

            for key, candidate in existing_items_lookup.items():
                if key[0] == section_id and key[2] == item_type and key[3] == item_label:
                    return candidate, f'label={item_label!r} section_id={section_id}'

        if export_id and export_id in export_id_to_existing:
            candidate = export_id_to_existing[export_id]
            if candidate.section_id == section_id:
                candidate_label = str(candidate.label or '').strip()
                if not item_label or not candidate_label or candidate_label == item_label:
                    return candidate, f'export_id={export_id}'
                cls._matrix_import_log(
                    f"export_id={export_id} label mismatch: excel={item_label!r} "
                    f"db_id={candidate.id} db_label={candidate_label!r}; skipping export_id match",
                    level='warning',
                )

        item_key = (section_id, item_order, item_type, item_label)
        existing_item = existing_items_lookup.get(item_key)
        if existing_item:
            return existing_item, f'order_key={item_key!r}'

        return None, None

    @classmethod
    def _normalize_matrix_item_config(cls, item_type: str, config: Any) -> Any:
        """Ensure matrix items store grouped columns under config.matrix_config."""
        if item_type != 'matrix' or config is None:
            return config
        if isinstance(config, str):
            config = cls._parse_json(config)
        if not isinstance(config, dict):
            cls._matrix_import_log(
                f"normalize skipped: config is {type(config).__name__}, not dict",
                level='warning',
            )
            return config

        cls._matrix_import_log(
            f"normalize input: {cls._summarize_matrix_config_for_log(config)}"
        )

        out = dict(config)
        matrix_config = out.get('matrix_config')
        if isinstance(matrix_config, str):
            matrix_config = cls._parse_json(matrix_config)

        root_column_groups = out.pop('column_groups', None)

        if isinstance(matrix_config, dict):
            mc = dict(matrix_config)
        elif out.get('type') == 'matrix' or isinstance(out.get('columns'), list):
            mc = {k: out.pop(k) for k in list(out.keys()) if k in cls._MATRIX_CONFIG_KEYS}
        else:
            cls._matrix_import_log(
                "normalize: no matrix_config/columns found in config; leaving unchanged",
                level='warning',
            )
            return out

        if root_column_groups is not None and 'column_groups' not in mc:
            cls._matrix_import_log(
                f"normalize: hoisted root column_groups keys={list(root_column_groups.keys()) if isinstance(root_column_groups, dict) else root_column_groups!r}"
            )
            mc['column_groups'] = root_column_groups

        columns = mc.get('columns')
        if isinstance(columns, list):
            normalized_columns = []
            for col in columns:
                if isinstance(col, dict):
                    normalized_columns.append(dict(col))
                elif isinstance(col, str) and col.strip():
                    normalized_columns.append({'name': col.strip(), 'type': 'number'})
            mc['columns'] = normalized_columns

        mc = cls._repair_matrix_column_groups(mc)
        out['matrix_config'] = mc
        cls._matrix_import_log(
            f"normalize output: {cls._summarize_matrix_config_for_log(out)}"
        )
        return out

    @classmethod
    def _repair_matrix_column_groups(cls, matrix_config: Any) -> Dict[str, Any]:
        """Keep column_groups translations and per-column group keys in sync."""
        if not isinstance(matrix_config, dict):
            return matrix_config

        before = cls._summarize_matrix_config_for_log({'matrix_config': matrix_config})

        mc = dict(matrix_config)
        columns = mc.get('columns')
        if not isinstance(columns, list):
            return mc

        columns = [dict(col) if isinstance(col, dict) else col for col in columns]
        raw_groups = mc.get('column_groups')
        column_groups: Dict[str, Any] = raw_groups if isinstance(raw_groups, dict) else {}

        repaired_groups: Dict[str, Any] = {}
        for group_label, group_value in column_groups.items():
            group_key = str(group_label).strip()
            if not group_key:
                continue
            if isinstance(group_value, list):
                for col_name in group_value:
                    target_name = str(col_name).strip()
                    for col in columns:
                        if isinstance(col, dict) and str(col.get('name', '')).strip() == target_name:
                            col['group'] = group_key
                repaired_groups[group_key] = {}
            elif isinstance(group_value, dict):
                repaired_groups[group_key] = group_value
            else:
                repaired_groups[group_key] = {}

        for col in columns:
            if not isinstance(col, dict):
                continue
            if 'group' not in col and col.get('group_label'):
                col['group'] = str(col.pop('group_label')).strip()
            elif 'group' not in col and col.get('groupName'):
                col['group'] = str(col.pop('groupName')).strip()
            group_name = col.get('group')
            if group_name and str(group_name).strip() and str(group_name).strip() not in repaired_groups:
                repaired_groups[str(group_name).strip()] = {}

        # Infer groups from column_groups keys when columns share a name prefix (e.g. "SP1 Planned" -> "SP1")
        for group_key in list(repaired_groups.keys()):
            prefix = f"{group_key} "
            for col in columns:
                if not isinstance(col, dict) or col.get('group'):
                    continue
                name = str(col.get('name', '')).strip()
                if name == group_key or name.startswith(prefix):
                    col['group'] = group_key

        mc['columns'] = columns
        mc['column_groups'] = repaired_groups
        after = cls._summarize_matrix_config_for_log({'matrix_config': mc})
        if before != after:
            cls._matrix_import_log(f"repair changed config: before={before} after={after}")
        else:
            cls._matrix_import_log(f"repair unchanged: {after}")
        return mc

    @classmethod
    def _apply_item_config_from_import(
        cls,
        item: FormItem,
        item_type: str,
        raw_config: Any,
        *,
        keep_existing_on_failure: bool = True,
    ) -> None:
        """Parse, normalize, and persist imported item config without wiping on parse failure."""
        from sqlalchemy.orm.attributes import flag_modified

        item_id = getattr(item, 'id', None)
        item_label = getattr(item, 'label', None)
        log_matrix = item_type == 'matrix'

        if raw_config is None or raw_config == '' or raw_config == 'None':
            if log_matrix:
                cls._matrix_import_log(
                    f"item id={item_id} label={item_label!r}: config cell empty; keeping existing config "
                    f"({cls._summarize_matrix_config_for_log(item.config if isinstance(item.config, dict) else {})})"
                )
            return

        if log_matrix:
            cls._matrix_import_log(
                f"item id={item_id} label={item_label!r}: raw config cell "
                f"{cls._summarize_raw_config_cell(raw_config)}"
            )

        parsed = cls._parse_json(raw_config)
        if parsed is None:
            if keep_existing_on_failure and log_matrix:
                cls._matrix_import_log(
                    f"item id={item_id} label={item_label!r}: JSON parse FAILED; keeping existing config "
                    f"({cls._summarize_matrix_config_for_log(item.config if isinstance(item.config, dict) else {})})",
                    level='warning',
                )
            return

        before_summary = cls._summarize_matrix_config_for_log(
            item.config if isinstance(item.config, dict) else {}
        ) if log_matrix else ''

        normalized = cls._normalize_matrix_item_config(item_type, parsed)
        if normalized is None:
            if log_matrix:
                cls._matrix_import_log(
                    f"item id={item_id} label={item_label!r}: normalize returned None; config not updated",
                    level='warning',
                )
            return

        if item_type == 'matrix' and isinstance(normalized, dict):
            item.config = cls._clean_imported_matrix_config(normalized)
            cls._sync_matrix_item_fields_from_config(item)
        else:
            item.config = cls._deep_copy_json(normalized) if isinstance(normalized, dict) else normalized

        flag_modified(item, 'config')
        if log_matrix:
            cls._matrix_import_log(
                f"item id={item_id} label={item_label!r}: config applied "
                f"before={before_summary} after={cls._summarize_matrix_config_for_log(item.config)}"
            )

    @classmethod
    def _sync_matrix_item_fields_from_config(cls, item: FormItem) -> None:
        """Mirror matrix_config list-library fields onto FormItem columns."""
        if item.item_type != 'matrix' or not isinstance(item.config, dict):
            return
        matrix_config = item.config.get('matrix_config')
        if not isinstance(matrix_config, dict):
            return
        if matrix_config.get('row_mode') == 'list_library':
            if 'lookup_list_id' in matrix_config:
                item.lookup_list_id = matrix_config['lookup_list_id']
            if matrix_config.get('list_display_column'):
                item.list_display_column = matrix_config['list_display_column']
            if matrix_config.get('list_filters') is not None:
                item.list_filters_json = matrix_config['list_filters']

