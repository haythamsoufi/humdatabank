import openpyxl
from openpyxl.utils import range_boundaries

path = r"app/static/templates/unified_country_plan.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["Planned Bilateral Support"]
print("tables:", {k: (v.ref if hasattr(v, "ref") else v) for k, v in ws.tables.items()})

tbl = ws.tables.get("Data_Support")
if tbl:
    ref = tbl.ref if hasattr(tbl, "ref") else tbl
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    headers = [ws.cell(min_row, c).value for c in range(min_col, max_col + 1)]
    print("headers:", headers)
    print("first 5 data rows:")
    for row in range(min_row + 1, min(min_row + 6, max_row + 1)):
        print([ws.cell(row, c).value for c in range(min_col, max_col + 1)])

td = wb["TemplateData"]
if "Table9" in td.tables:
    tbl = td.tables["Table9"]
    ref = tbl.ref if hasattr(tbl, "ref") else tbl
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    headers = [td.cell(min_row, c).value for c in range(min_col, max_col + 1)]
    print("Table9 headers:", headers)
    print("Table9 sample:", [td.cell(min_row + 1, c).value for c in range(min_col, max_col + 1)])

wb.close()
