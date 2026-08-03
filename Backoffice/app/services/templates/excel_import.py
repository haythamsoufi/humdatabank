"""Template Excel import paths."""

# ========== Template Excel Import/Export Service ==========
from app.utils.datetime_helpers import utcnow
"""
Service for exporting and importing form templates to/from Excel.
"""
from flask import current_app
from flask_login import current_user
from app import db
from app.models import (
    FormTemplate, FormPage, FormSection, FormItem, FormTemplateVersion,
    FormData,
    RepeatGroupInstance, RepeatGroupData, DynamicIndicatorData, DynamicSectionContext,
)
from app.models.documents import SubmittedDocument
from app.models.indicator_bank import IndicatorBank
from contextlib import suppress
from app.services.monitoring.memory import memory_tracker
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import io
import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from app.utils.stable_key import generate_stable_key, is_valid_stable_key, normalize_stable_key

from app.services.templates.matrix_import import TemplateExcelMatrixMixin


class TemplateExcelImportMixin(TemplateExcelMatrixMixin):
    """Mixin for TemplateExcelService."""

    @classmethod
    def _count_nonempty_data_rows(cls, sheet) -> int:
        """Count non-empty data rows (row 2 onward) in a worksheet."""
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and not all(cell is None for cell in row):
                count += 1
        return count

    @classmethod
    def _validate_template_sheet_headers(cls, headers: List[Any]) -> List[str]:
        """Validate Template sheet headers (row layout or legacy column layout)."""
        if cls._template_sheet_uses_row_layout(headers):
            return []
        required_headers = cls.REQUIRED_COLUMNS.get('Template', ['name'])
        header_set = {h for h in headers if isinstance(h, str) and h}
        required_set = set(required_headers)
        if not required_set.issubset(header_set):
            missing = required_set - header_set
            return [
                f"Template sheet headers missing required columns. "
                f"Required: {required_headers}, Missing: {list(missing)}, Got: {headers}"
            ]
        return []

    @classmethod
    def _validate_import_sheet_headers(cls, sheet_name: str, headers: List[Any]) -> List[str]:
        _, _, error = cls._resolve_sheet_headers(sheet_name, headers)
        return [error] if error else []

    @classmethod
    def validate_import_file(cls, excel_file) -> Dict[str, Any]:
        """
        Validate a Humanitarian Databank Excel export before import.

        Returns dict with keys: valid, message, errors, preview
        (preview: name, pages, sections, items).
        """
        errors: List[str] = []
        preview = {'name': None, 'pages': 0, 'sections': 0, 'items': 0}

        try:
            workbook = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
        except Exception as e:
            current_app.logger.error(f"Failed to load Excel file for validation: {e}", exc_info=True)
            cls._matrix_import_log(f"validate: failed to load workbook: {e}", level='error')
            return {
                'valid': False,
                'message': 'Invalid Excel file. Check the file format and try again.',
                'errors': ['Failed to load Excel file.'],
                'preview': preview,
            }

        required_sheets = ['Template', 'Pages', 'Sections', 'Items']
        missing_sheets = [s for s in required_sheets if s not in workbook.sheetnames]
        if missing_sheets:
            msg = f"Missing required sheets: {', '.join(missing_sheets)}"
            return {
                'valid': False,
                'message': msg,
                'errors': [msg],
                'preview': preview,
            }

        template_sheet = workbook['Template']
        template_headers = [cell.value for cell in template_sheet[1]]
        errors.extend(cls._validate_template_sheet_headers(template_headers))
        row_data, legacy_column_layout, template_parse_errors = cls._parse_template_sheet(template_sheet)
        errors.extend(template_parse_errors)
        if not errors:
            if not row_data:
                errors.append('Template sheet has no data row.')
            else:
                template_name = row_data.get('name')
                if template_name is None or str(template_name).strip() == '':
                    errors.append('Template sheet data row is missing a template name.')
                else:
                    preview['name'] = str(template_name).strip()

        pages_sheet = workbook['Pages']
        pages_headers = [cell.value for cell in pages_sheet[1]]
        errors.extend(cls._validate_import_sheet_headers('Pages', pages_headers))
        if not any(e.startswith('Pages sheet') for e in errors):
            preview['pages'] = cls._count_nonempty_data_rows(pages_sheet)

        sections_sheet = workbook['Sections']
        sections_headers = [cell.value for cell in sections_sheet[1]]
        errors.extend(cls._validate_import_sheet_headers('Sections', sections_headers))
        if not any(e.startswith('Sections sheet') for e in errors):
            preview['sections'] = cls._count_nonempty_data_rows(sections_sheet)

        items_sheet = workbook['Items']
        items_headers = [cell.value for cell in items_sheet[1]]
        errors.extend(cls._validate_import_sheet_headers('Items', items_headers))
        if not any(e.startswith('Items sheet') for e in errors):
            preview['items'] = cls._count_nonempty_data_rows(items_sheet)

        cls._scan_workbook_matrix_items(workbook, stage='validate')

        valid = len(errors) == 0
        if valid:
            message = (
                f"Valid export: {preview['pages']} pages, "
                f"{preview['sections']} sections, {preview['items']} items."
            )
        else:
            if len(errors) == 1:
                message = errors[0]
            else:
                message = f"Found {len(errors)} validation errors."

        return {
            'valid': valid,
            'message': message,
            'errors': errors,
            'preview': preview,
        }

    @classmethod
    def _parse_import_version_mode(cls, import_version_mode: Optional[str], target_version) -> bool:
        """Return True when the import should land in a draft (new or existing)."""
        if import_version_mode == 'create_draft':
            return True
        if import_version_mode == 'current_version':
            return False
        return target_version.status == 'published'

    @classmethod
    def _resolve_target_version(cls, template, version_id: Optional[int]):
        """Resolve the base version row for an import request."""
        target_version = None
        if version_id:
            target_version = FormTemplateVersion.query.filter_by(
                id=version_id, template_id=template.id
            ).first()

        if not target_version:
            if template.published_version_id:
                target_version = FormTemplateVersion.query.get(template.published_version_id)
            else:
                target_version = FormTemplateVersion.query.filter_by(
                    template_id=template.id
                ).order_by(FormTemplateVersion.created_at.desc()).first()

        return target_version

    @classmethod
    def resolve_import_target_version_id(
        cls,
        template_id: int,
        version_id: Optional[int],
        import_version_mode: Optional[str] = None,
    ) -> Optional[int]:
        """Return the version_id that will receive the import (without creating rows)."""
        template = FormTemplate.query.get(template_id)
        if not template:
            return version_id

        target_version = cls._resolve_target_version(template, version_id)
        if not target_version:
            return version_id

        if cls._parse_import_version_mode(import_version_mode, target_version):
            if target_version.status == 'published':
                existing_draft = FormTemplateVersion.query.filter_by(
                    template_id=template.id, status='draft'
                ).first()
                return existing_draft.id if existing_draft else None
            return target_version.id

        return target_version.id

    @classmethod
    def _get_or_create_draft_for_import(cls, template, source_version) -> FormTemplateVersion:
        """Return an existing draft or create one cloned from source_version."""
        existing_draft = FormTemplateVersion.query.filter_by(
            template_id=template.id, status='draft'
        ).first()
        if existing_draft:
            current_app.logger.info(
                f"Draft version already exists (ID={existing_draft.id}); using it for Excel import"
            )
            return existing_draft

        from sqlalchemy import func
        max_version = db.session.query(func.max(FormTemplateVersion.version_number)).filter_by(
            template_id=template.id
        ).scalar()
        next_version_number = (max_version + 1) if max_version else 1

        new_draft = FormTemplateVersion(
            template_id=template.id,
            version_number=next_version_number,
            status='draft',
            created_by=current_user.id,
            updated_by=current_user.id,
            based_on_version_id=source_version.id,
            comment='Created automatically for Excel import',
            name=source_version.name,
            name_translations=source_version.name_translations.copy() if source_version.name_translations else None,
            description=source_version.description,
            add_to_self_report=source_version.add_to_self_report,
            display_order_visible=source_version.display_order_visible,
            is_paginated=source_version.is_paginated,
            enable_export_pdf=source_version.enable_export_pdf,
            enable_export_excel=source_version.enable_export_excel,
            enable_import_excel=source_version.enable_import_excel,
            enable_ai_validation=getattr(source_version, 'enable_ai_validation', False),
            enable_data_quality=getattr(source_version, 'enable_data_quality', False),
            enable_discussion=getattr(source_version, 'enable_discussion', False),
            discussion_config=source_version.discussion_config.copy() if getattr(source_version, 'discussion_config', None) else None,
            data_quality_methodology=getattr(source_version, 'data_quality_methodology', None),
            validation_rule_pack=getattr(source_version, 'validation_rule_pack', None),
            variables=source_version.variables.copy() if source_version.variables else None
        )
        db.session.add(new_draft)
        db.session.flush()
        current_app.logger.info(
            f"Created new draft version: ID={new_draft.id}, Version #={new_draft.version_number}"
        )
        cls._matrix_import_log(
            f"Published import: created draft {new_draft.id} from version {source_version.id} "
            f"(structure will be rebuilt from Excel)"
        )
        return new_draft

    @classmethod
    @memory_tracker("Template Excel Import", log_top_allocations=True)
    def import_template(
        cls,
        template_id: int,
        excel_file,
        version_id: Optional[int] = None,
        import_version_mode: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Import template structure from Excel.

        Args:
            template_id: Template ID to import into
            excel_file: File-like object containing Excel file
            version_id: Optional version ID to import into (defaults to active version or creates draft)
            import_version_mode: 'create_draft' | 'current_version' — UI choice for import target

        Returns:
            Dict with 'success', 'message', 'errors', 'created_count' keys
        """
        cls.matrix_import_entry_log(
            f"import_template START template_id={template_id} version_id={version_id}"
        )
        current_app.logger.info(f"=== TEMPLATE EXCEL IMPORT START ===")
        current_app.logger.info(f"Template ID: {template_id}, Version ID: {version_id}")

        template = FormTemplate.query.get_or_404(template_id)
        current_app.logger.info(f"Template found: '{template.name}'")

        errors = []
        warnings: List[str] = []
        created_counts = {'pages': 0, 'sections': 0, 'items': 0}

        try:
            # Load workbook
            current_app.logger.info("Loading Excel workbook...")
            workbook = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
            current_app.logger.info(f"Workbook loaded. Sheets found: {workbook.sheetnames}")
            cls._scan_workbook_matrix_items(workbook, stage='import-pre-scan')

            # Validate required sheets exist (ignore Instructions, _Metadata and any other unrecognized sheets)
            required_sheets = ['Template', 'Pages', 'Sections', 'Items']
            recognized_sheets = set(required_sheets + ['Instructions', '_Metadata'])  # Instructions and _Metadata are recognized but optional
            missing_sheets = [s for s in required_sheets if s not in workbook.sheetnames]

            # Log any unrecognized sheets (they will be ignored)
            unrecognized_sheets = [s for s in workbook.sheetnames if s not in recognized_sheets]
            if unrecognized_sheets:
                current_app.logger.info(f"Ignoring unrecognized sheets during import: {unrecognized_sheets}")

            if missing_sheets:
                current_app.logger.error(f"Missing required sheets: {missing_sheets}")
                return {
                    'success': False,
                    'message': f"Missing required sheets: {', '.join(missing_sheets)}",
                    'errors': errors,
                    'created_count': created_counts
                }

            # Determine target version
            current_app.logger.info("Determining target version...")
            target_version = cls._resolve_target_version(template, version_id)
            if target_version and version_id:
                current_app.logger.info(
                    f"Using specified version ID {version_id} "
                    f"(Status: {target_version.status}, Version #: {target_version.version_number})"
                )
            elif target_version:
                current_app.logger.info(
                    f"Using fallback version ID {target_version.id} "
                    f"(Status: {target_version.status})"
                )

            if not target_version:
                current_app.logger.error("No version found for template")
                return {
                    'success': False,
                    'message': "No version found for this template. Please create a version first.",
                    'errors': errors,
                    'created_count': created_counts
                }

            create_new_draft = cls._parse_import_version_mode(import_version_mode, target_version)

            # When importing into a draft, create/use a draft version from published.
            if create_new_draft and target_version.status == 'published':
                current_app.logger.info("Import target is a draft; preparing draft version...")
                target_version = cls._get_or_create_draft_for_import(template, target_version)
                version_id = target_version.id

            current_app.logger.info(
                f"Target version: ID={target_version.id}, Status={target_version.status}, "
                f"Version #={target_version.version_number}"
            )

            if not create_new_draft and target_version.status == 'published':
                deletion_impact = cls._count_deletion_impact(template.id, target_version.id)
                if deletion_impact.get('has_data'):
                    return {
                        'success': False,
                        'message': (
                            'This template has submitted data. Import into a draft version '
                            'and deploy it instead.'
                        ),
                        'errors': [
                            'Excel import into the published version is blocked when submission data exists.'
                        ],
                        'created_count': created_counts,
                    }

            # Import Template metadata first (including version-specific name)
            current_app.logger.info("=== IMPORTING TEMPLATE METADATA ===")
            template_sheet = workbook['Template']
            template_errors = cls._import_template_metadata(template_sheet, template, target_version)
            errors.extend(template_errors)
            if template_errors:
                current_app.logger.warning(f"Template metadata import errors: {template_errors}")
            else:
                current_app.logger.info("Template metadata imported successfully")

            # Excel import is a full structure replace: drop existing pages/sections/items first
            # so renamed sections, changed types, and removed subsections don't linger.
            published_stable_key_context = cls._build_published_stable_key_context(template)
            cls._clear_version_structure(template.id, target_version.id)

            # ID mapping dictionaries (export ID -> new database ID)
            # Export IDs are sequential (1, 2, 3...) and map to new database IDs
            page_id_map = {}  # export_id -> new_db_id
            section_id_map = {}  # export_id -> new_db_id

            # Import Pages
            current_app.logger.info("=== IMPORTING PAGES ===")
            pages_sheet = workbook['Pages']
            page_errors = cls._import_pages(pages_sheet, template, target_version, page_id_map)
            errors.extend(page_errors)
            created_counts['pages'] = len(page_id_map)
            current_app.logger.info(f"Pages imported: {created_counts['pages']} pages created")
            current_app.logger.info(f"Page ID mapping: {page_id_map}")
            if page_errors:
                current_app.logger.warning(f"Page import errors: {page_errors}")

            # Import Sections
            current_app.logger.info("=== IMPORTING SECTIONS ===")
            sections_sheet = workbook['Sections']
            section_errors = cls._import_sections(
                sections_sheet,
                template,
                target_version,
                page_id_map,
                section_id_map,
                published_stable_key_context=published_stable_key_context,
            )
            errors.extend(section_errors)
            created_counts['sections'] = len(section_id_map)
            current_app.logger.info(f"Sections imported: {created_counts['sections']} sections created")
            current_app.logger.info(f"Section ID mapping: {section_id_map}")
            if section_errors:
                current_app.logger.warning(f"Section import errors: {section_errors}")

            # Import Items
            current_app.logger.info("=== IMPORTING ITEMS ===")
            items_sheet = workbook['Items']
            item_errors = cls._import_items(
                items_sheet,
                template,
                target_version,
                section_id_map,
                warnings=warnings,
                published_stable_key_context=published_stable_key_context,
            )
            errors.extend(item_errors)
            created_counts['items'] = db.session.query(FormItem).filter_by(
                template_id=template.id, version_id=target_version.id
            ).count()
            current_app.logger.info(f"Items imported: {created_counts['items']} items created")
            if item_errors:
                current_app.logger.warning(f"Item import errors: {item_errors}")

            if errors:
                current_app.logger.error(f"Import completed with {len(errors)} errors. Rolling back...")
                db.session.rollback()
                return {
                    'success': False,
                    'message': f"Import completed with {len(errors)} errors",
                    'errors': errors,
                    'created_count': created_counts
                }

            # Update version timestamp
            current_app.logger.info("Updating version timestamp...")
            target_version.updated_at = utcnow()
            target_version.updated_by = current_user.id

            current_app.logger.info("Committing database changes...")
            db.session.commit()
            cls._log_matrix_items_in_version(template.id, target_version.id, stage='after-commit')
            current_app.logger.info("=== TEMPLATE EXCEL IMPORT SUCCESS ===")
            current_app.logger.info(f"Final counts - Pages: {created_counts['pages']}, Sections: {created_counts['sections']}, Items: {created_counts['items']}")

            return {
                'success': True,
                'message': f"Successfully imported {created_counts['pages']} pages, "
                          f"{created_counts['sections']} sections, "
                          f"{created_counts['items']} items",
                'errors': [],
                'warnings': warnings,
                'created_count': created_counts,
                'version_id': version_id  # Return the version ID (may be new draft if published was selected)
            }

        except Exception as e:
            current_app.logger.error(f"=== TEMPLATE EXCEL IMPORT FAILED ===")
            current_app.logger.error(f"Error importing template from Excel: {e}", exc_info=True)
            db.session.rollback()
            return {
                'success': False,
                'message': "Error importing template. Check the file format and try again.",
                'errors': ['Import failed. See logs for details.'],
                'created_count': created_counts
            }

    @classmethod
    def _import_template_metadata(cls, sheet, template: FormTemplate, version: FormTemplateVersion) -> List[str]:
        """Import template metadata from Template sheet, including version-specific name."""
        current_app.logger.info("Starting template metadata import...")
        errors = []

        row_data_all, legacy_column_layout, parse_errors = cls._parse_template_sheet(sheet)
        errors.extend(parse_errors)
        if parse_errors:
            return errors

        if not row_data_all:
            current_app.logger.warning("Template sheet has no data row")
            return errors

        headers = [cell.value for cell in sheet[1]]
        current_app.logger.info(f"Template sheet headers: {headers}")

        if legacy_column_layout:
            _, legacy_format, _ = cls._resolve_sheet_headers('Template', headers)
            expected_headers = (
                cls.TEMPLATE_LEGACY_COLUMNS if legacy_format else cls.get_template_columns()
            )
        else:
            legacy_format = False
            expected_headers = cls.get_template_columns()

        header_set = set(row_data_all.keys())
        required_headers = cls.REQUIRED_COLUMNS.get('Template', ['name'])
        expected_set = set(expected_headers)
        extra_columns = header_set - expected_set
        missing_optional = expected_set - header_set - set(required_headers)

        if extra_columns:
            current_app.logger.info(
                f"Template sheet contains extra/legacy fields (will be ignored): {list(extra_columns)}"
            )
        if missing_optional:
            current_app.logger.info(
                f"Template sheet missing optional fields (will use defaults): {list(missing_optional)}"
            )

        row_data = {h: row_data_all.get(h) for h in expected_headers}

        try:
            current_app.logger.info(f"Updating template metadata for template ID {template.id}")

            # Update version-specific fields (name is now only stored in versions)
            if 'name' in row_data and row_data['name']:
                version.name = row_data['name']
                current_app.logger.info(f"Updated version name: '{version.name}'")

            if 'description' in row_data:
                version.description = row_data['description']
                current_app.logger.info(f"Updated version/template description")

            if 'add_to_self_report' in row_data:
                version.add_to_self_report = bool(row_data['add_to_self_report'])
                current_app.logger.info(f"Updated version/template add_to_self_report: {version.add_to_self_report}")

            if 'display_order_visible' in row_data:
                version.display_order_visible = bool(row_data['display_order_visible'])
                current_app.logger.info(f"Updated version/template display_order_visible: {version.display_order_visible}")

            if 'is_paginated' in row_data:
                version.is_paginated = bool(row_data['is_paginated'])
                current_app.logger.info(f"Updated version/template is_paginated: {version.is_paginated}")

            if 'enable_export_pdf' in row_data:
                version.enable_export_pdf = bool(row_data['enable_export_pdf'])
                current_app.logger.info(f"Updated version/template enable_export_pdf: {version.enable_export_pdf}")

            if 'enable_export_excel' in row_data:
                version.enable_export_excel = bool(row_data['enable_export_excel'])
                current_app.logger.info(f"Updated version/template enable_export_excel: {version.enable_export_excel}")

            if 'enable_import_excel' in row_data:
                version.enable_import_excel = bool(row_data['enable_import_excel'])
                current_app.logger.info(f"Updated version/template enable_import_excel: {version.enable_import_excel}")

            if 'enable_ai_validation' in row_data:
                version.enable_ai_validation = bool(row_data['enable_ai_validation'])
                current_app.logger.info(f"Updated version/template enable_ai_validation: {version.enable_ai_validation}")

            name_translations = cls._collect_translations_from_row(
                row_data_all, 'name', legacy=legacy_format
            )
            if name_translations:
                version.name_translations = name_translations
                current_app.logger.info("Updated version name_translations")

            if 'variables' in row_data:
                raw_variables = row_data.get('variables')
                current_app.logger.info(
                    f"Variables cell raw value type={type(raw_variables).__name__} "
                    f"len={len(str(raw_variables)) if raw_variables is not None else 0}: "
                    f"{str(raw_variables)[:200]}"
                )
                variables = cls._parse_json(raw_variables)
                if variables is not None:  # Allow empty dict {} to clear variables
                    # Save to version (version-specific template variables)
                    version.variables = variables if variables else {}
                    from sqlalchemy.orm.attributes import flag_modified
                    flag_modified(version, 'variables')
                    current_app.logger.info(f"Updated version variables: {len(variables) if variables else 0} variable(s)")
                elif raw_variables is not None:
                    current_app.logger.warning(
                        f"Could not parse variables JSON from Excel (value truncated or malformed): "
                        f"{str(raw_variables)[:300]}"
                    )
                else:
                    current_app.logger.info("variables cell is empty in Excel — skipping (no overwrite)")

            db.session.add(template)
            db.session.add(version)
            current_app.logger.info("Template metadata import complete")

        except Exception as e:
            current_app.logger.error("Template metadata row: %s", e, exc_info=True)
            errors.append("Template metadata row: Validation error.")

        return errors

    @classmethod
    def _import_pages(cls, sheet, template: FormTemplate, version: FormTemplateVersion,
                     page_id_map: Dict[int, int]) -> List[str]:
        """Import pages from sheet."""
        current_app.logger.info("Starting pages import...")
        errors = []

        # Read headers
        headers = [cell.value for cell in sheet[1]]
        current_app.logger.info(f"Pages sheet headers: {headers}")

        # Validate headers match expected columns
        _, legacy_format, header_error = cls._resolve_sheet_headers('Pages', headers)
        if header_error:
            current_app.logger.error(header_error)
            errors.append(header_error)
            return errors

        current_app.logger.info("Headers validated. Processing page rows...")
        row_count = 0

        # Read data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue  # Skip empty rows

            row_count += 1
            try:
                row_data = cls._row_dict_from_sheet(headers, row)

                # Get export ID (sequential ID from Excel, e.g., 1, 2, 3...)
                export_id = int(row_data['id']) if row_data.get('id') else None

                if export_id is None:
                    error_msg = f"Pages row {row_idx}: Missing export ID"
                    current_app.logger.warning(error_msg)
                    errors.append(error_msg)
                    continue

                page_name = row_data['name'] or ''
                page_order = int(row_data['order']) if row_data['order'] is not None else 1
                current_app.logger.info(f"Processing page row {row_idx}: export_id={export_id}, name='{page_name}', order={page_order}")

                # Create new page
                new_page = FormPage(
                    template_id=template.id,
                    version_id=version.id,
                    name=page_name,
                    order=page_order,
                    name_translations=cls._collect_translations_from_row(
                        row_data, 'name', legacy=legacy_format
                    ),
                )

                db.session.add(new_page)
                db.session.flush()

                # Map export ID to new database ID
                page_id_map[export_id] = new_page.id
                current_app.logger.info(f"Page created: export_id={export_id} -> db_id={new_page.id}, name='{page_name}'")

            except Exception as e:
                current_app.logger.error("Pages row %s: %s", row_idx, e, exc_info=True)
                errors.append(f"Pages row {row_idx}: Validation error.")

        current_app.logger.info(f"Pages import complete: {row_count} rows processed, {len(page_id_map)} pages created, {len(errors)} errors")
        return errors

    @classmethod
    def _build_published_stable_key_context(cls, template: FormTemplate) -> Dict[str, Any]:
        """Map published structure to stable_key values for silent reuse on Excel import."""
        empty: Dict[str, Any] = {
            'sections': {},
            'items_by_indicator': {},
            'items_by_position': {},
        }
        if not template.published_version_id:
            return empty

        sections = FormSection.query.filter_by(
            template_id=template.id,
            version_id=template.published_version_id,
        ).all()
        if not sections:
            return empty

        sections_by_id = {section.id: section for section in sections}
        section_keys: Dict[Tuple[float, str], str] = {}
        for section in sections:
            if section.stable_key:
                section_keys[(float(section.order), section.name or '')] = section.stable_key

        items_by_indicator: Dict[Tuple[float, str, int, float], str] = {}
        items_by_position: Dict[Tuple[float, str, float, str, str], str] = {}
        items = FormItem.query.filter_by(
            template_id=template.id,
            version_id=template.published_version_id,
        ).all()
        for item in items:
            if not item.stable_key:
                continue
            section = sections_by_id.get(item.section_id)
            if not section:
                continue
            section_key = (float(section.order), section.name or '')
            item_order = float(item.order) if item.order is not None else 0.0
            item_type = (item.item_type or '').strip().lower()
            if item.indicator_bank_id:
                indicator_key = (
                    section_key[0],
                    section_key[1],
                    int(item.indicator_bank_id),
                    item_order,
                )
                items_by_indicator[indicator_key] = item.stable_key
            else:
                position_key = (
                    section_key[0],
                    section_key[1],
                    item_order,
                    item_type,
                    str(item.label or '').strip(),
                )
                items_by_position[position_key] = item.stable_key

        return {
            'sections': section_keys,
            'items_by_indicator': items_by_indicator,
            'items_by_position': items_by_position,
        }

    @classmethod
    def _published_item_stable_key_fallback(
        cls,
        *,
        published_stable_key_context: Dict[str, Any],
        section_order: float,
        section_name: str,
        item_type: str,
        item_label: str,
        item_order: float,
        indicator_bank_id: Optional[int],
    ) -> Optional[str]:
        section_key = (section_order, section_name or '')
        if indicator_bank_id is not None:
            indicator_key = (section_key[0], section_key[1], int(indicator_bank_id), item_order)
            return published_stable_key_context.get('items_by_indicator', {}).get(indicator_key)
        position_key = (
            section_key[0],
            section_key[1],
            item_order,
            (item_type or '').strip().lower(),
            str(item_label or '').strip(),
        )
        return published_stable_key_context.get('items_by_position', {}).get(position_key)

    @classmethod
    def _resolve_import_stable_key(
        cls,
        row_data: Dict[str, Any],
        row_idx: int,
        sheet_name: str,
        errors: List[str],
        published_fallback: Optional[str] = None,
    ) -> Optional[str]:
        raw = row_data.get('stable_key')
        if raw is not None and str(raw).strip() != '':
            key = normalize_stable_key(raw)
            if not key:
                errors.append(f"{sheet_name} row {row_idx}: Invalid stable_key '{raw}'")
                return None
            return key
        if published_fallback:
            return published_fallback
        return generate_stable_key()

    @classmethod
    def _validate_stable_key_duplicates_in_sheet(
        cls,
        rows: List[Tuple[int, Dict[str, Any]]],
        sheet_name: str,
        errors: List[str],
    ) -> None:
        seen: Dict[str, int] = {}
        for row_idx, row_data in rows:
            raw = row_data.get('stable_key')
            if raw is None or str(raw).strip() == '':
                continue
            key = normalize_stable_key(raw)
            if not key:
                continue
            if key in seen:
                errors.append(
                    f"{sheet_name} row {row_idx}: Duplicate stable_key '{key}' "
                    f"(also on row {seen[key]})"
                )
            else:
                seen[key] = row_idx

    @classmethod
    def _published_items_by_stable_key(cls, template: FormTemplate) -> Dict[str, FormItem]:
        if not template.published_version_id:
            return {}
        items = FormItem.query.filter_by(
            template_id=template.id,
            version_id=template.published_version_id,
        ).filter(FormItem.stable_key.isnot(None)).all()
        return {item.stable_key: item for item in items if item.stable_key}

    @classmethod
    def _check_stable_key_identity_mismatch(
        cls,
        *,
        template: FormTemplate,
        stable_key: str,
        item_type: str,
        indicator_bank_id: Optional[int],
        row_idx: int,
        warnings: List[str],
    ) -> None:
        published_by_key = cls._published_items_by_stable_key(template)
        published_item = published_by_key.get(stable_key)
        if not published_item:
            return
        pub_type = (published_item.item_type or '').strip().lower()
        cur_type = (item_type or '').strip().lower()
        pub_bank = published_item.indicator_bank_id
        if pub_type != cur_type or pub_bank != indicator_bank_id:
            warnings.append(
                f"Items row {row_idx}: stable_key {stable_key} matches published field "
                f"'{published_item.label}' (type={pub_type}, indicator_bank_id={pub_bank}) "
                f"but this row is type={cur_type}, indicator_bank_id={indicator_bank_id}. "
                f"Identity mismatch — verify this is intentional."
            )

    @classmethod
    def _import_sections(cls, sheet, template: FormTemplate, version: FormTemplateVersion,
                        page_id_map: Dict[int, int], section_id_map: Dict[int, int],
                        published_stable_key_context: Optional[Dict[str, Any]] = None) -> List[str]:
        """Import sections from sheet."""
        if published_stable_key_context is None:
            published_stable_key_context = cls._build_published_stable_key_context(template)
        current_app.logger.info("Starting sections import...")
        current_app.logger.info(f"Page ID mapping available: {len(page_id_map)} pages")
        errors = []

        # Read headers
        headers = [cell.value for cell in sheet[1]]
        current_app.logger.info(f"Sections sheet headers: {headers}")

        _, legacy_format, header_error = cls._resolve_sheet_headers('Sections', headers)
        if header_error:
            current_app.logger.error(header_error)
            errors.append(header_error)
            return errors

        # First pass: collect all section data
        current_app.logger.info("First pass: Collecting section data...")
        sections_data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            try:
                row_data = cls._row_dict_from_sheet(headers, row)
                sections_data.append((row_idx, row_data))
            except Exception as e:
                current_app.logger.error("Sections row %s: %s", row_idx, e)
                errors.append(f"Sections row {row_idx}: Validation error.")

        current_app.logger.info(f"Collected {len(sections_data)} section rows to process")

        cls._validate_stable_key_duplicates_in_sheet(sections_data, 'Sections', errors)
        if errors:
            return errors

        # Get existing sections for this version (for matching)
        existing_sections = FormSection.query.filter_by(
            template_id=template.id, version_id=version.id
        ).all()
        # Create lookup: (order, name) -> existing section (use order+name as match key)
        existing_sections_by_key = {(s.order, s.name): s for s in existing_sections}

        # Second pass: create or update sections and build ID map
        current_app.logger.info("Second pass: Creating/updating sections...")
        sections_created = 0
        sections_updated = 0

        for row_idx, row_data in sections_data:
            try:
                # Get export ID (sequential ID from Excel)
                export_id = int(row_data['id']) if row_data.get('id') else None

                if export_id is None:
                    error_msg = f"Sections row {row_idx}: Missing export ID"
                    current_app.logger.warning(error_msg)
                    errors.append(error_msg)
                    continue

                # Resolve page_id using export ID mapping
                page_id = None
                page_export_id = None
                if row_data.get('page_id'):
                    page_export_id = int(row_data['page_id']) if row_data.get('page_id') else None
                    if page_export_id and page_export_id in page_id_map:
                        page_id = page_id_map[page_export_id]
                        current_app.logger.debug(f"Section row {row_idx}: Mapped page export_id={page_export_id} -> db_id={page_id}")
                    elif page_export_id:
                        current_app.logger.warning(f"Section row {row_idx}: Page export_id={page_export_id} not found in page_id_map")

                section_name = row_data['name'] or ''
                section_order = float(row_data['order']) if row_data['order'] is not None else 0.0
                section_type = row_data.get('section_type') or 'standard'
                section_stable_key = cls._resolve_import_stable_key(
                    row_data,
                    row_idx,
                    'Sections',
                    errors,
                    published_fallback=published_stable_key_context.get('sections', {}).get(
                        (section_order, section_name)
                    ),
                )
                if section_stable_key is None:
                    continue
                current_app.logger.info(f"Processing section row {row_idx}: export_id={export_id}, name='{section_name}', order={section_order}, type={section_type}, page_export_id={page_export_id}")

                # Check if section already exists (match by order + name)
                section_key = (section_order, section_name)
                existing_section = existing_sections_by_key.get(section_key)

                if existing_section:
                    # Update existing section
                    current_app.logger.info(f"Updating existing section: db_id={existing_section.id}, order={section_order}, name='{section_name}'")
                    existing_section.page_id = page_id
                    existing_section.section_type = section_type
                    existing_section.max_dynamic_indicators = int(row_data['max_dynamic_indicators']) if row_data.get('max_dynamic_indicators') else None
                    existing_section.allowed_sectors = cls._parse_json(row_data.get('allowed_sectors'))
                    existing_section.indicator_filters = cls._parse_json(row_data.get('indicator_filters'))
                    existing_section.allow_data_not_available = bool(row_data.get('allow_data_not_available', False))
                    existing_section.allow_not_applicable = bool(row_data.get('allow_not_applicable', False))
                    existing_section.allowed_disaggregation_options = cls._parse_json(row_data.get('allowed_disaggregation_options'))
                    existing_section.data_entry_display_filters = cls._parse_json(row_data.get('data_entry_display_filters'))
                    existing_section.add_indicator_note = row_data.get('add_indicator_note')
                    existing_section.name_translations = cls._collect_translations_from_row(
                        row_data, 'name', legacy=legacy_format
                    )
                    existing_section.relevance_condition = row_data.get('relevance_condition')
                    existing_section.archived = bool(row_data.get('archived', False))
                    existing_section.stable_key = section_stable_key
                    # Note: parent_section_id will be updated in third pass
                    sections_updated += 1
                    section_id_map[export_id] = existing_section.id
                    current_app.logger.info(f"Section updated: export_id={export_id} -> db_id={existing_section.id}, name='{section_name}'")
                else:
                    # Create new section (parent_section_id will be resolved in third pass)
                    new_section = FormSection(
                        template_id=template.id,
                        version_id=version.id,
                        name=section_name,
                        order=section_order,
                        page_id=page_id,
                        parent_section_id=None,  # Will be set in third pass
                        section_type=section_type,
                        max_dynamic_indicators=int(row_data['max_dynamic_indicators']) if row_data.get('max_dynamic_indicators') else None,
                        allowed_sectors=cls._parse_json(row_data.get('allowed_sectors')),
                        indicator_filters=cls._parse_json(row_data.get('indicator_filters')),
                        allow_data_not_available=bool(row_data.get('allow_data_not_available', False)),
                        allow_not_applicable=bool(row_data.get('allow_not_applicable', False)),
                        allowed_disaggregation_options=cls._parse_json(row_data.get('allowed_disaggregation_options')),
                        data_entry_display_filters=cls._parse_json(row_data.get('data_entry_display_filters')),
                        add_indicator_note=row_data.get('add_indicator_note'),
                        name_translations=cls._collect_translations_from_row(
                            row_data, 'name', legacy=legacy_format
                        ),
                        relevance_condition=row_data.get('relevance_condition'),
                        archived=bool(row_data.get('archived', False)),
                        stable_key=section_stable_key,
                    )

                    db.session.add(new_section)
                    db.session.flush()

                    # Map export ID to new database ID
                    section_id_map[export_id] = new_section.id
                    sections_created += 1
                    current_app.logger.info(f"Section created: export_id={export_id} -> db_id={new_section.id}, name='{section_name}'")

            except Exception as e:
                current_app.logger.error("Sections row %s: %s", row_idx, e, exc_info=True)
                errors.append(f"Sections row {row_idx}: Validation error.")

        current_app.logger.info(f"Sections created/updated: {sections_created} created, {sections_updated} updated")

        current_app.logger.info(f"Sections created: {len(section_id_map)} sections")

        # Third pass: update parent_section_id references using export ID mapping
        current_app.logger.info("Third pass: Updating parent_section_id references...")
        parent_updates = 0
        for row_idx, row_data in sections_data:
            if row_data.get('parent_section_id'):
                try:
                    parent_export_id = int(row_data['parent_section_id']) if row_data.get('parent_section_id') else None
                    section_export_id = int(row_data['id']) if row_data.get('id') else None

                    if parent_export_id and section_export_id:
                        # Find new database IDs using export IDs
                        new_section_id = section_id_map.get(section_export_id)
                        new_parent_id = section_id_map.get(parent_export_id)

                        if new_section_id and new_parent_id:
                            # Update parent reference
                            section = FormSection.query.get(new_section_id)
                            if section:
                                section.parent_section_id = new_parent_id
                                parent_updates += 1
                                current_app.logger.debug(f"Updated parent: section export_id={section_export_id} (db_id={new_section_id}) -> parent export_id={parent_export_id} (db_id={new_parent_id})")
                        else:
                            current_app.logger.warning(f"Sections row {row_idx}: Could not resolve parent reference (section_export_id={section_export_id}, parent_export_id={parent_export_id})")
                except Exception as e:
                    current_app.logger.error("Sections row %s (parent update): %s", row_idx, e)
                    errors.append(f"Sections row {row_idx} (parent update): Validation error.")

        current_app.logger.info(f"Sections import complete: {len(sections_data)} rows processed, {len(section_id_map)} sections created, {parent_updates} parent relationships updated, {len(errors)} errors")
        return errors

    @classmethod
    def _import_items(cls, sheet, template: FormTemplate, version: FormTemplateVersion,
                     section_id_map: Dict[int, int], warnings: Optional[List[str]] = None,
                     published_stable_key_context: Optional[Dict[str, Any]] = None) -> List[str]:
        """Import items from sheet."""
        if warnings is None:
            warnings = []
        if published_stable_key_context is None:
            published_stable_key_context = cls._build_published_stable_key_context(template)
        current_app.logger.info("Starting items import...")
        current_app.logger.info(f"Section ID mapping available: {len(section_id_map)} sections")
        errors = []

        def _parse_int_like(value) -> Optional[int]:
            """Parse ints from Excel/JSON-ish values (e.g., 233, 233.0, '233')."""
            if value is None:
                return None
            try:
                # Excel frequently provides whole numbers as floats (e.g. 233.0)
                if isinstance(value, bool):
                    return None
                if isinstance(value, int):
                    return int(value)
                if isinstance(value, float):
                    if value.is_integer():
                        return int(value)
                    return None
                if isinstance(value, str):
                    s = value.strip()
                    if not s:
                        return None
                    if s.isdigit():
                        return int(s)
                    # Handle numeric strings like "233.0"
                    with suppress(Exception):
                        f = float(s)
                        if f.is_integer():
                            return int(f)
                return None
            except Exception as e:
                current_app.logger.debug("_parse_config_int failed: %s", e)
                return None

        def _ensure_import_issue(config: Any, *, code: str, message: str, meta: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
            """Attach an import issue to config['_import_issues'] and return the config dict."""
            cfg = config if isinstance(config, dict) else {}
            issues = cfg.get('_import_issues')
            if not isinstance(issues, list):
                issues = []
            issues.append({
                'code': code,
                'message': message,
                'meta': meta or {},
            })
            cfg['_import_issues'] = issues
            return cfg

        def _clear_import_issue_codes(config: Any, codes: List[str]) -> Any:
            """Remove matching import issue codes from config if present."""
            if not isinstance(config, dict):
                return config
            issues = config.get('_import_issues')
            if not isinstance(issues, list) or not issues:
                return config
            filtered = [i for i in issues if not (isinstance(i, dict) and i.get('code') in codes)]
            if filtered:
                config['_import_issues'] = filtered
            else:
                config.pop('_import_issues', None)
            return config

        # Read headers
        headers = [cell.value for cell in sheet[1]]
        current_app.logger.info(f"Items sheet headers: {headers}")

        _, legacy_format, header_error = cls._resolve_sheet_headers('Items', headers)
        if header_error:
            current_app.logger.error(header_error)
            errors.append(header_error)
            return errors

        current_app.logger.info("Headers validated. Processing item rows...")

        # Get existing items for this version (for matching)
        existing_items_lookup = cls._get_existing_items_lookup(template.id, version.id)
        ordered_existing_items = cls._get_items_for_version(template, version)
        export_id_to_existing = {idx + 1: item for idx, item in enumerate(ordered_existing_items)}
        cls._matrix_import_log(
            f"Starting items import for version_id={version.id}; "
            f"existing_items={len(existing_items_lookup)}, export_order_items={len(export_id_to_existing)}"
        )
        cls._log_matrix_items_in_version(template.id, version.id, stage='before-import')

        # Pre-scan rows for indicator_bank_id values so we can validate in one DB query.
        candidate_indicator_bank_ids: set[int] = set()
        rows_buffer: List[Tuple[int, Dict[str, Any]]] = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue
            row_data = cls._row_dict_from_sheet(headers, row)
            rows_buffer.append((row_idx, row_data))
            item_type = (row_data.get('item_type') or 'indicator')
            if item_type == 'indicator':
                raw_ib = row_data.get('indicator_bank_id')
                parsed_ib = _parse_int_like(raw_ib)
                if parsed_ib is not None:
                    candidate_indicator_bank_ids.add(parsed_ib)

        cls._validate_stable_key_duplicates_in_sheet(rows_buffer, 'Items', errors)
        if errors:
            return errors

        existing_indicator_bank_ids: set[int] = set()
        if candidate_indicator_bank_ids:
            existing_indicator_bank_ids = {
                int(x[0]) for x in db.session.query(IndicatorBank.id)
                .filter(IndicatorBank.id.in_(candidate_indicator_bank_ids))
                .all()
            }

        row_count = 0
        items_created = 0
        items_updated = 0

        # Map Items sheet export IDs -> actual FormItem objects (existing or newly created)
        item_export_to_obj: Dict[int, FormItem] = {}

        # Process buffered rows
        for row_idx, row_data in rows_buffer:
            row_count += 1
            try:
                # Get export ID
                export_id = int(row_data['id']) if row_data.get('id') else None

                # Resolve section_id using export ID mapping
                section_id = None
                section_export_id = None
                if row_data.get('section_id'):
                    section_export_id = int(row_data['section_id']) if row_data.get('section_id') else None
                    if section_export_id and section_export_id in section_id_map:
                        section_id = section_id_map[section_export_id]
                        current_app.logger.debug(f"Item row {row_idx}: Mapped section export_id={section_export_id} -> db_id={section_id}")
                    elif section_export_id:
                        current_app.logger.warning(f"Item row {row_idx}: Section export_id={section_export_id} not found in section_id_map")

                if not section_id:
                    error_msg = f"Items row {row_idx}: Could not resolve section_id (export_id: {section_export_id})"
                    current_app.logger.warning(error_msg)
                    errors.append(error_msg)
                    continue

                item_type = str(row_data.get('item_type') or 'indicator').strip().lower()
                item_label = str(row_data.get('label') or '').strip()
                item_order = float(row_data['order']) if row_data['order'] is not None else 0.0
                parsed_ib_for_mismatch = _parse_int_like(row_data.get('indicator_bank_id'))
                section = FormSection.query.get(section_id)
                section_order = float(section.order) if section and section.order is not None else 0.0
                section_name = (section.name or '') if section else ''
                item_stable_key = cls._resolve_import_stable_key(
                    row_data,
                    row_idx,
                    'Items',
                    errors,
                    published_fallback=cls._published_item_stable_key_fallback(
                        published_stable_key_context=published_stable_key_context,
                        section_order=section_order,
                        section_name=section_name,
                        item_type=item_type,
                        item_label=item_label,
                        item_order=item_order,
                        indicator_bank_id=parsed_ib_for_mismatch,
                    ),
                )
                if item_stable_key is None:
                    continue
                cls._check_stable_key_identity_mismatch(
                    template=template,
                    stable_key=item_stable_key,
                    item_type=item_type,
                    indicator_bank_id=parsed_ib_for_mismatch,
                    row_idx=row_idx,
                    warnings=warnings,
                )
                current_app.logger.info(f"Processing item row {row_idx}: export_id={export_id}, type={item_type}, label='{item_label[:50]}...', order={item_order}, section_export_id={section_export_id}")

                existing_item, match_method = cls._find_existing_item_for_import(
                    section_id=section_id,
                    item_type=item_type,
                    item_label=item_label,
                    item_order=item_order,
                    export_id=export_id,
                    export_id_to_existing=export_id_to_existing,
                    existing_items_lookup=existing_items_lookup,
                )

                if item_type == 'matrix':
                    target_db_id = existing_item.id if existing_item else None
                    parsed_preview = cls._parse_json(row_data.get('config'))
                    normalized_preview = cls._normalize_matrix_item_config(item_type, parsed_preview)
                    cls._matrix_import_log(
                        f"import row {row_idx} export_id={export_id} label={item_label!r} "
                        f"match={match_method or 'CREATE NEW'} target_db_id={target_db_id}; "
                        f"raw={cls._summarize_raw_config_cell(row_data.get('config'))}; "
                        f"normalized={cls._summarize_matrix_config_for_log(normalized_preview if isinstance(normalized_preview, dict) else {})}"
                    )

                if existing_item:
                    # Update existing item
                    current_app.logger.info(f"Updating existing item: db_id={existing_item.id}, section_id={section_id}, order={item_order}, type={item_type}")
                    existing_item.label = item_label
                    existing_item.order = item_order
                    existing_item.item_type = item_type
                    existing_item.relevance_condition = row_data.get('relevance_condition')
                    existing_item.archived = bool(row_data.get('archived', False))
                    cls._apply_item_config_from_import(existing_item, item_type, row_data.get('config'))

                    # Validate indicator_bank_id (only meaningful for indicator items)
                    if item_type == 'indicator':
                        raw_ib = row_data.get('indicator_bank_id')
                        parsed_ib = _parse_int_like(raw_ib)
                        if parsed_ib is None:
                            # Missing/invalid ID: import as NULL + flag issue in config
                            existing_item.indicator_bank_id = None
                            existing_item.config = _ensure_import_issue(
                                existing_item.config,
                                code='missing_or_invalid_indicator_bank_id',
                                message='Indicator is missing a valid Indicator Bank ID. Please select a valid indicator before deploying.',
                                meta={'raw_indicator_bank_id': raw_ib},
                            )
                        elif parsed_ib not in existing_indicator_bank_ids:
                            # References a non-existent IndicatorBank row: import as NULL + flag issue in config
                            existing_item.indicator_bank_id = None
                            existing_item.config = _ensure_import_issue(
                                existing_item.config,
                                code='missing_indicator_bank_reference',
                                message=f'Indicator references missing Indicator Bank ID {parsed_ib}. Please select a valid indicator before deploying.',
                                meta={'raw_indicator_bank_id': raw_ib, 'parsed_indicator_bank_id': parsed_ib},
                            )
                        else:
                            # Valid reference: clear any prior import issue flags and set the FK
                            existing_item.indicator_bank_id = parsed_ib
                            existing_item.config = _clear_import_issue_codes(
                                existing_item.config,
                                ['missing_or_invalid_indicator_bank_id', 'missing_indicator_bank_reference'],
                            )
                    else:
                        # Non-indicator: clear any stale indicator import issues
                        existing_item.config = _clear_import_issue_codes(
                            existing_item.config,
                            ['missing_or_invalid_indicator_bank_id', 'missing_indicator_bank_reference'],
                        )
                        existing_item.indicator_bank_id = _parse_int_like(row_data.get('indicator_bank_id'))
                    existing_item.type = row_data.get('type')
                    existing_item.unit = row_data.get('unit')
                    existing_item.validation_condition = row_data.get('validation_condition')
                    existing_item.validation_message = row_data.get('validation_message')
                    existing_item.definition = row_data.get('definition')
                    existing_item.options_json = cls._parse_json(row_data.get('options_json'))
                    existing_item.lookup_list_id = str(row_data['lookup_list_id']) if row_data.get('lookup_list_id') else None
                    existing_item.list_display_column = row_data.get('list_display_column')
                    existing_item.list_filters_json = cls._parse_json(row_data.get('list_filters_json'))
                    cls._apply_item_translations_from_row(existing_item, row_data, legacy=legacy_format)
                    existing_item.options_translations = cls._parse_json(row_data.get('options_translations'))
                    existing_item.description = row_data.get('description')
                    existing_item.stable_key = item_stable_key
                    if item_type == 'matrix':
                        cls._sync_matrix_item_fields_from_config(existing_item)
                        from sqlalchemy.orm.attributes import flag_modified
                        flag_modified(existing_item, 'config')
                        cls._matrix_import_log(
                            f"import updated matrix db_id={existing_item.id} label={item_label!r} "
                            f"via {match_method}: "
                            f"{cls._summarize_matrix_config_for_log(existing_item.config)}"
                        )
                    items_updated += 1
                    current_app.logger.info(f"Item updated: db_id={existing_item.id}, label='{item_label[:50]}...'")
                    if export_id is not None:
                        item_export_to_obj[export_id] = existing_item
                else:
                    # Parse and validate indicator bank reference up-front (so we never violate FK constraints)
                    parsed_indicator_bank_id: Optional[int] = None
                    parsed_config = cls._parse_json(row_data.get('config'))
                    config_payload = cls._normalize_matrix_item_config(item_type, parsed_config)
                    if item_type == 'matrix' and isinstance(config_payload, dict):
                        config_payload = cls._clean_imported_matrix_config(config_payload)
                    if item_type == 'matrix':
                        cls._matrix_import_log(
                            f"row {row_idx} CREATE NEW label={item_label!r}: "
                            f"parsed_config={cls._summarize_matrix_config_for_log(config_payload if isinstance(config_payload, dict) else {})}"
                        )
                    if item_type == 'indicator':
                        raw_ib = row_data.get('indicator_bank_id')
                        parsed_ib = _parse_int_like(raw_ib)
                        if parsed_ib is None:
                            parsed_indicator_bank_id = None
                            config_payload = _ensure_import_issue(
                                config_payload,
                                code='missing_or_invalid_indicator_bank_id',
                                message='Indicator is missing a valid Indicator Bank ID. Please select a valid indicator before deploying.',
                                meta={'raw_indicator_bank_id': raw_ib},
                            )
                        elif parsed_ib not in existing_indicator_bank_ids:
                            parsed_indicator_bank_id = None
                            config_payload = _ensure_import_issue(
                                config_payload,
                                code='missing_indicator_bank_reference',
                                message=f'Indicator references missing Indicator Bank ID {parsed_ib}. Please select a valid indicator before deploying.',
                                meta={'raw_indicator_bank_id': raw_ib, 'parsed_indicator_bank_id': parsed_ib},
                            )
                        else:
                            parsed_indicator_bank_id = parsed_ib
                            config_payload = _clear_import_issue_codes(
                                config_payload,
                                ['missing_or_invalid_indicator_bank_id', 'missing_indicator_bank_reference'],
                            )
                    else:
                        parsed_indicator_bank_id = _parse_int_like(row_data.get('indicator_bank_id'))

                    # Create new item
                    new_item = FormItem(
                        section_id=section_id,
                        template_id=template.id,
                        version_id=version.id,
                        item_type=item_type,
                        stable_key=item_stable_key,
                        label=item_label,
                        order=item_order,
                        relevance_condition=row_data.get('relevance_condition'),
                        archived=bool(row_data.get('archived', False)),
                        config=config_payload,
                        indicator_bank_id=parsed_indicator_bank_id,
                        type=row_data.get('type'),
                        unit=row_data.get('unit'),
                        validation_condition=row_data.get('validation_condition'),
                        validation_message=row_data.get('validation_message'),
                        definition=row_data.get('definition'),
                        options_json=cls._parse_json(row_data.get('options_json')),
                        lookup_list_id=str(row_data['lookup_list_id']) if row_data.get('lookup_list_id') else None,
                        list_display_column=row_data.get('list_display_column'),
                        list_filters_json=cls._parse_json(row_data.get('list_filters_json')),
                        options_translations=cls._parse_json(row_data.get('options_translations')),
                        description=row_data.get('description')
                    )
                    cls._apply_item_translations_from_row(new_item, row_data, legacy=legacy_format)

                    if item_type == 'matrix':
                        cls._sync_matrix_item_fields_from_config(new_item)

                    db.session.add(new_item)
                    items_created += 1
                    if export_id is not None:
                        item_export_to_obj[export_id] = new_item

                    if items_created % 10 == 0:
                        current_app.logger.debug(f"Items progress: {items_created} items created so far...")

            except Exception as e:
                current_app.logger.error("Items row %s: %s", row_idx, e, exc_info=True)
                errors.append(f"Items row {row_idx}: Validation error.")

        current_app.logger.info(f"Items import complete: {row_count} rows processed, {items_created} created, {items_updated} updated, {len(errors)} errors")
        cls._log_matrix_items_in_version(template.id, version.id, stage='after-import-pre-commit')

        # Rewrite rule JSON item IDs from export IDs -> new DB IDs so relevance/validation rules survive imports.
        try:
            db.session.flush()  # ensure new items have IDs
            item_id_map: Dict[int, int] = {}
            for exp_id, obj in item_export_to_obj.items():
                if obj is not None and getattr(obj, 'id', None) is not None:
                    item_id_map[int(exp_id)] = int(obj.id)

            if item_id_map:
                rewritten_items = 0
                for obj in item_export_to_obj.values():
                    if not obj:
                        continue
                    before_rel = obj.relevance_condition
                    before_val = getattr(obj, 'validation_condition', None)
                    obj.relevance_condition = cls._rewrite_rule_json_item_ids(before_rel, item_id_map)
                    if hasattr(obj, 'validation_condition'):
                        obj.validation_condition = cls._rewrite_rule_json_item_ids(before_val, item_id_map)
                    if before_rel != obj.relevance_condition or before_val != getattr(obj, 'validation_condition', None):
                        rewritten_items += 1

                rewritten_sections = 0
                sections = FormSection.query.filter_by(template_id=template.id, version_id=version.id).all()
                for s in sections:
                    before = s.relevance_condition
                    s.relevance_condition = cls._rewrite_rule_json_item_ids(before, item_id_map)
                    if before != s.relevance_condition:
                        rewritten_sections += 1

                current_app.logger.info(
                    f"Rewrote rule item IDs using item_id_map (size={len(item_id_map)}): "
                    f"items_updated={rewritten_items}, sections_updated={rewritten_sections}"
                )
        except Exception as e:
            # Don't fail the import for rule rewrite issues; log and continue.
            current_app.logger.warning(f"Could not rewrite rule item IDs after item import: {e}", exc_info=True)

        return errors

    @classmethod
    def _count_deletion_impact(cls, template_id: int, version_id: Optional[int]) -> Dict[str, Any]:
        """Return counts of submission-data rows that would be removed by a structure clear.

        Used by the preflight endpoint so the UI can warn the admin before the import
        actually destroys any data.  No data is modified.
        """
        if not version_id:
            return {'has_data': False, 'counts': {}}

        section_ids: List[int] = [
            row[0]
            for row in db.session.query(FormSection.id).filter_by(
                template_id=template_id, version_id=version_id
            ).all()
        ]

        # Join form_item so we count only data for this version without a large IN subquery.
        form_data: int = db.session.query(
            db.func.count(FormData.id)
        ).join(
            FormItem, FormData.form_item_id == FormItem.id
        ).filter(
            FormItem.template_id == template_id,
            FormItem.version_id == version_id,
        ).scalar() or 0

        repeat_instances: int = 0
        repeat_data: int = 0
        dynamic_indicators: int = 0
        dynamic_contexts: int = 0

        if section_ids:
            repeat_instances = db.session.query(
                db.func.count(RepeatGroupInstance.id)
            ).filter(RepeatGroupInstance.section_id.in_(section_ids)).scalar() or 0

            if repeat_instances:
                repeat_data = db.session.query(
                    db.func.count(RepeatGroupData.id)
                ).join(
                    RepeatGroupInstance,
                    RepeatGroupData.repeat_instance_id == RepeatGroupInstance.id,
                ).filter(
                    RepeatGroupInstance.section_id.in_(section_ids)
                ).scalar() or 0

            dynamic_indicators = db.session.query(
                db.func.count(DynamicIndicatorData.id)
            ).filter(DynamicIndicatorData.section_id.in_(section_ids)).scalar() or 0

            dynamic_contexts = db.session.query(
                db.func.count(DynamicSectionContext.id)
            ).filter(DynamicSectionContext.section_id.in_(section_ids)).scalar() or 0

        submitted_documents: int = db.session.query(
            db.func.count(SubmittedDocument.id)
        ).join(
            FormItem, SubmittedDocument.form_item_id == FormItem.id
        ).filter(
            FormItem.template_id == template_id,
            FormItem.version_id == version_id,
        ).scalar() or 0

        total = form_data + repeat_instances + repeat_data + dynamic_indicators + dynamic_contexts + submitted_documents
        return {
            'has_data': total > 0,
            'counts': {
                'form_data': form_data,
                'repeat_instances': repeat_instances,
                'repeat_data': repeat_data,
                'dynamic_indicators': dynamic_indicators,
                'dynamic_contexts': dynamic_contexts,
                'submitted_documents': submitted_documents,
            },
        }

    @classmethod
    def _clear_version_structure(cls, template_id: int, version_id: int) -> Dict[str, int]:
        """Delete all pages, sections, and items for a version (keeps the version row itself).

        Bulk .delete(synchronize_session=False) bypasses SQLAlchemy ORM cascade rules, so
        child records that have FK constraints pointing to form_section / form_item must be
        removed explicitly before the structural rows are deleted.
        """
        version = FormTemplateVersion.query.filter_by(id=version_id, template_id=template_id).first()
        if version and version.status == 'published':
            impact = cls._count_deletion_impact(template_id, version_id)
            if impact.get('has_data'):
                raise ValueError(
                    'Cannot clear structure on a published version that has submission data. '
                    'Import into a draft version and deploy it instead.'
                )

        current_app.logger.info(
            f"Clearing existing structure for template_id={template_id}, version_id={version_id} before Excel import"
        )

        # Collect the IDs of sections and items that are about to be removed so we can
        # clean up submission-data rows that hold FK references to them.
        section_ids: List[int] = [
            row[0]
            for row in db.session.query(FormSection.id).filter_by(
                template_id=template_id, version_id=version_id
            ).all()
        ]
        item_ids: List[int] = [
            row[0]
            for row in db.session.query(FormItem.id).filter_by(
                template_id=template_id, version_id=version_id
            ).all()
        ]

        if section_ids:
            # repeat_group_data.repeat_instance_id → repeat_group_instance.id has no DB CASCADE
            instance_ids_sq = db.session.query(RepeatGroupInstance.id).filter(
                RepeatGroupInstance.section_id.in_(section_ids)
            ).subquery()
            rgd_deleted = RepeatGroupData.query.filter(
                RepeatGroupData.repeat_instance_id.in_(instance_ids_sq)
            ).delete(synchronize_session=False)

            # repeat_group_instance.section_id → form_section.id has no DB CASCADE
            rgi_deleted = RepeatGroupInstance.query.filter(
                RepeatGroupInstance.section_id.in_(section_ids)
            ).delete(synchronize_session=False)

            # dynamic_indicator_data.section_id → form_section.id has no DB CASCADE
            did_deleted = DynamicIndicatorData.query.filter(
                DynamicIndicatorData.section_id.in_(section_ids)
            ).delete(synchronize_session=False)

            # dynamic_section_context.section_id → form_section.id has no DB CASCADE
            dsc_deleted = DynamicSectionContext.query.filter(
                DynamicSectionContext.section_id.in_(section_ids)
            ).delete(synchronize_session=False)

            if any([rgd_deleted, rgi_deleted, did_deleted, dsc_deleted]):
                current_app.logger.info(
                    f"Cleared submission data tied to version structure: "
                    f"{rgi_deleted} repeat_group_instance(s), {rgd_deleted} repeat_group_data row(s), "
                    f"{did_deleted} dynamic_indicator_data row(s), {dsc_deleted} dynamic_section_context row(s)"
                )

        if item_ids:
            # form_data.form_item_id → form_item.id has no DB CASCADE
            fd_deleted = FormData.query.filter(
                FormData.form_item_id.in_(item_ids)
            ).delete(synchronize_session=False)

            # Catch any repeat_group_data rows referencing these items that were not already
            # removed via the section/instance path above (edge-case guard).
            rgd_by_item = RepeatGroupData.query.filter(
                RepeatGroupData.form_item_id.in_(item_ids)
            ).delete(synchronize_session=False)

            if fd_deleted or rgd_by_item:
                current_app.logger.info(
                    f"Cleared {fd_deleted} form_data row(s) and "
                    f"{rgd_by_item} additional repeat_group_data row(s) tied to version items"
                )

        items_deleted = FormItem.query.filter_by(
            template_id=template_id, version_id=version_id
        ).delete(synchronize_session=False)
        sections_deleted = FormSection.query.filter_by(
            template_id=template_id, version_id=version_id
        ).delete(synchronize_session=False)
        pages_deleted = FormPage.query.filter_by(
            template_id=template_id, version_id=version_id
        ).delete(synchronize_session=False)
        db.session.flush()
        current_app.logger.info(
            f"Cleared version structure: {pages_deleted} pages, {sections_deleted} sections, {items_deleted} items"
        )
        return {'pages': pages_deleted, 'sections': sections_deleted, 'items': items_deleted}

    @classmethod
    def _clone_template_structure(cls, template_id: int, source_version_id: int, target_version_id: int) -> None:
        """Clone pages, sections, and items from source_version_id to target_version_id preserving order."""
        current_app.logger.info(f"Cloning template structure from version {source_version_id} to {target_version_id}")
        # Maps for old->new IDs
        page_id_map = {}
        section_id_map = {}

        # Clone pages
        src_pages = FormPage.query.filter_by(template_id=template_id, version_id=source_version_id).order_by(FormPage.order).all()
        current_app.logger.info(f"Cloning {len(src_pages)} pages")
        for p in src_pages:
            new_p = FormPage(
                template_id=template_id,
                version_id=target_version_id,
                name=p.name,
                order=p.order,
                name_translations=p.name_translations
            )
            db.session.add(new_p)
            db.session.flush()
            page_id_map[p.id] = new_p.id

        # Clone sections (first pass: main sections)
        src_sections = FormSection.query.filter_by(template_id=template_id, version_id=source_version_id).order_by(FormSection.order).all()
        current_app.logger.info(f"Cloning {len(src_sections)} sections")
        # Create all sections without parent refs first
        for s in src_sections:
            new_s = FormSection(
                template_id=template_id,
                version_id=target_version_id,
                name=s.name,
                order=s.order,
                parent_section_id=None,  # set later
                page_id=page_id_map.get(s.page_id) if s.page_id else None,
                section_type=s.section_type,
                max_dynamic_indicators=s.max_dynamic_indicators,
                allowed_sectors=s.allowed_sectors,
                indicator_filters=s.indicator_filters,
                allow_data_not_available=s.allow_data_not_available,
                allow_not_applicable=s.allow_not_applicable,
                allowed_disaggregation_options=s.allowed_disaggregation_options,
                data_entry_display_filters=s.data_entry_display_filters,
                add_indicator_note=s.add_indicator_note,
                name_translations=s.name_translations,
                relevance_condition=s.relevance_condition,
                archived=s.archived
            )
            db.session.add(new_s)
            db.session.flush()
            section_id_map[s.id] = new_s.id

        # Second pass: set parent_section_id now that all new IDs exist
        parent_updates = 0
        for s in src_sections:
            if s.parent_section_id:
                new_id = section_id_map[s.id]
                new_parent_id = section_id_map.get(s.parent_section_id)
                if new_parent_id:
                    FormSection.query.filter_by(id=new_id).update({'parent_section_id': new_parent_id})
                    parent_updates += 1
        current_app.logger.info(f"Cloned {len(section_id_map)} sections, updated {parent_updates} parent relationships")

        # Clone items
        src_items = FormItem.query.join(FormSection, FormItem.section_id == FormSection.id).\
            filter(FormItem.template_id == template_id, FormItem.version_id == source_version_id).\
            order_by(FormItem.order).all()
        current_app.logger.info(f"Cloning {len(src_items)} items")
        items_cloned = 0
        for it in src_items:
            # Deep copy config to avoid cross-version mutations
            _new_config = None
            try:
                _new_config = json.loads(json.dumps(it.config)) if it.config is not None else None
            except Exception as e:
                current_app.logger.debug("JSON roundtrip for item config failed: %s", e)
                try:
                    import copy as _copy
                    _new_config = _copy.deepcopy(it.config) if it.config is not None else None
                except Exception as e2:
                    current_app.logger.debug("deepcopy item config failed: %s", e2)
                    _new_config = it.config.copy() if isinstance(it.config, dict) else it.config

            new_it = FormItem(
                template_id=template_id,
                version_id=target_version_id,
                section_id=section_id_map.get(it.section_id),
                item_type=it.item_type,
                label=it.label,
                order=it.order,
                relevance_condition=it.relevance_condition,
                archived=it.archived,
                config=_new_config,
                indicator_bank_id=it.indicator_bank_id,
                type=it.type,
                unit=it.unit,
                validation_condition=it.validation_condition,
                validation_message=it.validation_message,
                definition=it.definition,
                options_json=it.options_json,
                lookup_list_id=getattr(it, 'lookup_list_id', None),
                list_display_column=getattr(it, 'list_display_column', None),
                list_filters_json=getattr(it, 'list_filters_json', None),
                label_translations=getattr(it, 'label_translations', None),
                definition_translations=getattr(it, 'definition_translations', None),
                options_translations=getattr(it, 'options_translations', None),
                description_translations=getattr(it, 'description_translations', None),
                description=getattr(it, 'description', None)
            )
            db.session.add(new_it)
            items_cloned += 1

        current_app.logger.info(f"Successfully cloned structure: {len(page_id_map)} pages, {len(section_id_map)} sections, {items_cloned} items")

    @classmethod
    def _get_existing_items_lookup(cls, template_id: int, version_id: int) -> Dict[Tuple, FormItem]:
        """
        Create a lookup dictionary for existing items.
        Key: (section_id, order, item_type, label) - combination to uniquely identify an item
        Value: FormItem object
        """
        existing_items = FormItem.query.join(FormSection).filter(
            FormItem.template_id == template_id,
            FormSection.version_id == version_id
        ).all()

        # Create lookup: (section_id, order, item_type, label) -> item
        # Using label as part of key since it's more stable than just order
        items_lookup = {}
        for item in existing_items:
            key = (
                item.section_id,
                float(item.order) if item.order is not None else 0.0,
                str(item.item_type or 'indicator').strip().lower(),
                str(item.label or '').strip(),
            )
            items_lookup[key] = item

        return items_lookup

